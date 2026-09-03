"""
Views для LiftTeam v2.59.0.
CRUD операции, дашборд, отчёты, визуальная сетка кассетниц, печать этикеток,
импорт радиодеталей из Excel.
"""
import re
import json
from collections import Counter, defaultdict
from contextlib import contextmanager
from datetime import datetime, time, timedelta
from decimal import Decimal
import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import (
    Count, DecimalField, F, Max, OuterRef, Prefetch, Q, Subquery, Sum, Value,
)
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST
from django.db import transaction
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer

from .models import (
    Client, EquipmentModel, EquipmentType, EquipmentVersion, Equipment,
    PriceList, PriceListLine, TechCard, available_equipment_for_order,
    FaultType, FaultTypePart, RepairOrder, RepairOrderEquipment,
    SparePart, StorageCell, StockMovement, StockAllocation, OrderCost, RepairOrderDetail,
    OrderStatusHistory, Employee,
    Notification, Payment, Organization, BankOperation, Cabinet,
    InventorySession, InventorySessionLine, SettingChange,
    LastAdminError, Position, admin_access_exists, permissions_by_section,
    add_months, warranty_cutoff, warranty_months, plural_genitive,
)
from .forms import (
    LoginForm, ClientForm, ClientContactFormSet, EquipmentModelForm, EquipmentTypeForm,
    EquipmentVersionForm, EquipmentForm,
    RepairOrderForm, RepairOrderDetailForm, SparePartForm,
    StockMovementForm, StockOutgoingForm, EmployeeForm, StatusChangeForm,
    PriceListForm, PriceListLineFormSet, EquipmentMaterialFormSet,
    PositionForm,
    TechCardForm, TechCardStepFormSet, UnitDiagnosisForm, UnitRepairForm,
    UnitDiskFolderForm, OrderInfoForm,
    RepairOrderIntakeForm,
    RepairOrderEquipmentIntakeFormSet, PartImportForm, PaymentForm, OrganizationForm,
    InvoiceSendForm, QuoteForm, QuoteLineFormSet,
    CabinetForm, MyNotificationsForm, FaultTypeForm, FaultTypePartFormSet,
    make_fault_type_part_formset,
)
from .utils import (
    generate_qr_image,
    build_workbook, add_sheet, xlsx_response, excel_datetime,
)
from .decorators import permission_required
from . import (
    envfile, invoicing, messengers, notifications, restarter, scanning,
    selfcheck, tbank, updater, yadisk,
)


def _send_stock_update(part):
    """Отправка обновления остатка через WebSocket."""
    if part is None:
        return
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "stock_updates",
        {
            "type": "stock_update",
            "data": {
                "part_id": part.id,
                "part_number": part.part_number,
                "name": part.name,
                "current_stock": part.current_stock,
                "min_stock": part.min_stock,
                "is_below_min": part.is_below_min_stock(),
            }
        }
    )


# ==================== АУТЕНТИФИКАЦИЯ ====================

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = LoginForm(request=request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Неверный логин или пароль')
    else:
        form = LoginForm()
    return render(request, 'core/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def my_notifications(request):
    """Личный выбор канала внутренних оповещений — своя страница,
    без параметра в URL: всегда про текущего пользователя, доступна любой
    роли. Меняет только эти три поля — за роль и доступ отвечает
    администратор через карточку пользователя."""
    if request.method == 'POST':
        form = MyNotificationsForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Настройки оповещений сохранены')
            return redirect('my_notifications')
    else:
        form = MyNotificationsForm(instance=request.user)
    return render(request, 'core/my_notifications.html', {'form': form})


@login_required
def presence(request):
    """Кто из сотрудников сейчас на связи.

    Видно всем вошедшим, права не спрашиваем: это не надзор
    за подчинёнными, а способ не идти через лабораторию к пустому
    терминалу. Правами закрыто только то, что и раньше было закрыто
    ролями, — деньги, удаление и администрирование.

    Список рисуется сразу из базы, а не ждёт сокета: отметки хранятся
    в Employee.last_seen и переживают перезапуск сервера. Дальше страницу
    обновляет ws/presence/.
    """
    employees = Employee.objects.filter(is_active=True)
    return render(request, 'core/presence.html', {
        'employees': employees,
        'timeout_seconds': settings.PRESENCE_TIMEOUT_SECONDS,
    })


# ==================== ДАШБОРД ====================

@login_required
def dashboard(request):
    """Главная страница — статистика и алерты."""
    total_orders = RepairOrder.objects.count()
    # «Завершённых» статусов три: отгружен, признан неремонтопригодным
    # и частично отремонтирован (смешанный исход тоже финал) — все три
    # означают, что по заказу больше ничего не ждут
    active_orders = RepairOrder.objects.exclude(
        status__in=['shipped', 'unrepairable', 'partially_repaired']
    ).count()

    # Дефицит и «ровно на минимуме» — разные вещи: первое уже в плане закупок,
    # второе означает, что следующее списание уведёт деталь в минус
    low_stock_parts = SparePart.objects.below_minimum().order_by('part_number')
    low_stock_count = low_stock_parts.count()
    at_minimum_parts = SparePart.objects.at_minimum().order_by('part_number')
    at_minimum_count = at_minimum_parts.count()
    recent_orders = RepairOrder.objects.select_related('client').prefetch_related('order_equipments__equipment__model').order_by('-date_received')[:10]

    # Разбивка по статусам — с нулями для тех, которых сейчас нет: пустая
    # колонка «Готов к отгрузке» тоже сведение, а прыгающий набор ячеек
    # читать труднее, чем постоянный
    counts = {
        item['status']: item['count']
        for item in RepairOrder.objects.values('status').annotate(count=Count('id'))
    }
    status_stats = [
        {'code': code, 'label': label, 'count': counts.get(code, 0)}
        for code, label in RepairOrder.STATUS_CHOICES
    ]

    # Что доделано и ждёт отправки — первый вопрос дня. Раньше ответа
    # на него на дашборде не было вовсе: приходилось открывать список
    # заказов и в нём каждый заказ по очереди.
    #
    # Берём заказы, ещё не отгруженные, и оставляем те, по которым
    # заполнено всё, что нужно для документов и склада. Готовность
    # считается по единицам (READINESS_CHECKS), поэтому нужны и строки
    # заказа, и то, что спрашивает чек-лист.
    ready_to_ship = [
        order for order in (
            RepairOrder.objects
            .filter(status__in=('repair', 'ready_for_shipment'))
            .select_related('client')
            .prefetch_related('order_equipments__faults',
                              'order_equipments__details')
            .order_by('date_received')
        )
        if order.readiness()['all_ready']
    ]

    # Должники (не оплаченные заказы)
    debtors = RepairOrder.objects.with_debt().select_related('client')
    total_debt = _total_debt(debtors)

    # Неразнесённые поступления показываем только тем, кто их разносит:
    # мастеру эта цифра ничего не говорит, а место на дашборде занимает
    unapplied_operations = (
        BankOperation.objects.filter(status='new').count()
        if request.user.allows('bank_statement') else 0
    )

    context = {
        'total_orders': total_orders,
        'active_orders': active_orders,
        'low_stock_count': low_stock_count,
        'low_stock_parts': low_stock_parts[:20],
        'at_minimum_count': at_minimum_count,
        'at_minimum_parts': at_minimum_parts[:10],
        'recent_orders': recent_orders,
        'ready_to_ship': ready_to_ship[:10],
        'ready_to_ship_count': len(ready_to_ship),
        'status_stats': status_stats,
        'debtors': debtors[:10],
        'total_debt': total_debt,
        'unapplied_operations': unapplied_operations,
        'now': timezone.now(),
    }
    return render(request, 'core/dashboard.html', context)


# ==================== КЛИЕНТЫ ====================

CLIENT_LIST_SORT_FIELDS = {
    'name': 'name',
    'inn': 'inn',
    'contact_person': 'contact_person',
    'phone': 'phone',
}


@login_required
def client_list(request):
    search = request.GET.get('q', '')
    clients = Client.objects.all()
    if search:
        clients = clients.filter(Q(name__icontains=search) | Q(inn__icontains=search))
    clients = sorted_by_request(clients.order_by('name'), request, CLIENT_LIST_SORT_FIELDS)
    paginator = Paginator(clients, 25)
    page = request.GET.get('page')
    return render(request, 'core/clients/list.html', {
        'clients': paginator.get_page(page),
        'search': search
    })


@login_required
def client_export(request):
    """Список заказчиков в Excel — с числом заказов и суммой долга."""
    search = request.GET.get('q', '').strip()
    clients = Client.objects.all()
    if search:
        clients = clients.filter(Q(name__icontains=search) | Q(inn__icontains=search))
    clients = clients.order_by('name')

    orders = dict(
        RepairOrder.objects
        .filter(client__in=clients)
        .values_list('client_id')
        .annotate(count=Count('id'))
    )
    debts = dict(
        RepairOrderEquipment.objects
        .filter(
            repair_order__client__in=clients,
            repair_order__payment_status__in=['unpaid', 'partially_paid'],
        )
        .exclude(repair_order__status='unrepairable')
        .values_list('repair_order__client_id')
        .annotate(total=Sum('repair_cost'))
    )

    headers = [
        'Название', 'ИНН', 'КПП', 'Контактное лицо', 'Телефон', 'Email',
        'Заказов', 'Долг, ₽',
    ]
    rows = [
        [
            client.name, client.inn, client.kpp, client.contact_person,
            client.phone, client.email,
            orders.get(client.pk, 0), debts.get(client.pk) or 0,
        ]
        for client in clients
    ]

    wb = build_workbook('Заказчики', headers, rows)
    return xlsx_response(wb, f'Заказчики {timezone.localdate():%Y-%m-%d}.xlsx')


@login_required
def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Заказчик добавлен')
            return redirect('client_list')
    else:
        form = ClientForm()
    return render(request, 'core/clients/form.html', {'form': form, 'title': 'Новый заказчик'})


@login_required
def client_edit(request, pk):
    """Правка заказчика вместе с его контактами.

    Контакты только здесь, на странице создания их нет: пока заказчика
    не сохранили, вешать контакты не на что — тот же приём, что
    у материалов модели (`EquipmentMaterialFormSet`).
    """
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        formset = ClientContactFormSet(request.POST, instance=client)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Заказчик обновлён')
            return redirect('client_list')
        messages.error(request, 'Не сохранено: проверьте отмеченные поля')
    else:
        form = ClientForm(instance=client)
        formset = ClientContactFormSet(instance=client)
    return render(request, 'core/clients/form.html', {
        'form': form, 'formset': formset,
        'title': 'Редактирование заказчика', 'client': client,
    })


@permission_required('clients_delete')
def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        client.delete()
        messages.success(request, 'Заказчик удалён')
        return redirect('client_list')
    return render(request, 'core/clients/delete.html', {'client': client})


# ==================== ОБОРУДОВАНИЕ ====================

def _filter_equipment(request):
    """Отбор оборудования по параметрам GET-запроса."""
    search = request.GET.get('q', '').strip()
    warranty_filter = request.GET.get('warranty', '')

    equipment = Equipment.objects.select_related('model', 'version', 'current_client')

    if search:
        # Ищем и по нормализованному номеру: запрос «буад 1234» должен найти
        # «БУАД-1234», иначе сотрудник не увидит уже заведённую единицу
        # и заведёт её второй раз под другим написанием.
        normalized = Equipment.normalize_serial(search)
        condition = (
            Q(serial_number__icontains=search) |
            Q(model__name__icontains=search) |
            Q(current_client__name__icontains=search)
        )
        if normalized:
            condition |= Q(serial_normalized__contains=normalized)
        equipment = equipment.filter(condition)

    # Гарантия не хранится, поэтому отбираем по её признаку: заказ с этой
    # единицей завершён не раньше, чем срок гарантии назад
    cutoff = warranty_cutoff()
    if cutoff is not None and warranty_filter in ('active', 'expired'):
        covered = Q(repairorderequipment__repair_order__date_completed__gte=cutoff)
        if warranty_filter == 'active':
            equipment = equipment.filter(covered)
        else:
            equipment = equipment.exclude(covered)

    return equipment.distinct().order_by('serial_number'), {
        'search': search,
        'warranty_filter': warranty_filter,
    }


def remember_list_query(request, key):
    """Запомнить отбор списка, чтобы «Назад» с карточки вернул к нему.

    «Назад» в программе — это обычная ссылка, а не кнопка браузера: она
    ведёт на список без параметров, и отбор терялся. Искали «BAV», открыли
    деталь, вернулись — и снова весь каталог, ищи заново.

    Хранится в сессии, а не передаётся в адресе: на карточку попадают
    не только из списка — ещё сканером, из заказа, по ссылке из письма, —
    и во всех этих случаях «Назад» обязан вести туда же, куда и обычно.
    """
    request.session['list_query:%s' % key] = request.GET.urlencode()


def list_back_url(request, key, url_name):
    """Адрес списка вместе с последним отбором."""
    query = request.session.get('list_query:%s' % key, '')
    return f'{reverse(url_name)}?{query}' if query else reverse(url_name)


def sorted_by_request(queryset, request, fields):
    """Сортировка списка по клику на шапку таблицы — одно место на всю
    программу (с v2.106.0), под общий `sortable-table.js`.

    `fields` — словарь {имя в адресе: выражение для `order_by`}. Список
    разрешённых явный, тем же приёмом, что у `envfile.EDITABLE`
    и `READINESS_CHECKS`: подставленное в запрос имя мимо списка не сортирует
    ничего, а не роняет страницу и не сортирует по случайному полю модели.

    Без `?sort=` в адресе список остаётся в том порядке, что уже задал сам
    вызывающий (обычно самое естественное упорядочение — по номеру, по дате
    приёма), и трогать его, пока по шапке не кликнули, незачем.

    Сортирует **сервер**, а не браузер: списки почти всегда постраничные,
    и пересортировка в браузере отсортировала бы только одну открытую
    страницу. Разметке нужен только `<th data-sort="имя">` — ссылки
    и обработчик клика рисует сам скрипт.
    """
    key = request.GET.get('sort', '')
    if key not in fields:
        return queryset
    direction = 'desc' if request.GET.get('dir') == 'desc' else 'asc'
    order_expr = fields[key]
    if direction == 'desc':
        order_expr = '-' + order_expr
    return queryset.order_by(order_expr)


@login_required
def equipment_list(request):
    equipment, filter_context = _filter_equipment(request)

    paginator = Paginator(equipment, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Гарантия считается сразу по всей странице: поштучная проверка
    # означала бы запрос на каждую строку списка
    warranty_map = Equipment.warranty_map(list(page_obj))
    for item in page_obj:
        item.warranty = warranty_map.get(item.pk)

    remember_list_query(request, 'equipment')
    return render(request, 'core/equipment/list.html', {
        'equipment': page_obj,
        'found_count': paginator.count,
        'filters_active': any(filter_context.values()),
        **filter_context,
    })


@login_required
def equipment_export(request):
    """Список оборудования в Excel — с учётом поиска и фильтра гарантии."""
    equipment, _ = _filter_equipment(request)
    equipment = list(equipment)
    warranty_map = Equipment.warranty_map(equipment)

    headers = [
        'Модель', 'Серийный номер', 'Текущий заказчик',
        'Гарантия до', 'Заказ по гарантии', 'Всего ремонтов',
    ]
    # Число ремонтов по всем единицам одним запросом: свойство на каждую
    # строку означало бы запрос на строку
    repairs = dict(
        RepairOrderEquipment.objects
        .filter(equipment__in=equipment)
        .values_list('equipment_id')
        .annotate(count=Count('id'))
    )

    rows = []
    for item in equipment:
        warranty = warranty_map.get(item.pk)
        rows.append([
            item.model.name,
            item.serial_number,
            item.current_client.name if item.current_client else '',
            # Только дата: час окончания гарантии никому не нужен
            excel_datetime(warranty.warranty_until).date() if warranty else '',
            warranty.repair_order.order_number if warranty else '',
            repairs.get(item.pk, 0),
        ])

    wb = build_workbook('Оборудование', headers, rows)
    return xlsx_response(wb, f'Оборудование {timezone.localdate():%Y-%m-%d}.xlsx')


@login_required
def equipment_create(request):
    """Новое оборудование в справочнике.

    `?next=` возвращает туда, откуда пришли, — с карточки заказа,
    когда прибор забыли внести при приёме. Раньше такое оборудование
    заводили окном на странице правки заказа; страницы больше нет,
    а окно рисует Bootstrap, которого может не быть.

    Адрес возврата принимается только свой: чужой превратил бы кнопку
    в способ увести человека на постороннюю страницу.
    """
    next_url = request.POST.get('next') or request.GET.get('next') or ''
    if not url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        next_url = ''

    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            equipment = form.save()
            messages.success(request, f'{equipment} добавлено в справочник')
            return redirect(next_url or 'equipment_list')
    else:
        # Заказчик заказа, из которого пришли: у прибора он и станет
        # владельцем, а перевыбирать его руками незачем
        initial = {}
        client_id = request.GET.get('client', '')
        if str(client_id).isdigit():
            initial['current_client'] = client_id
        form = EquipmentForm(initial=initial)
    return render(request, 'core/equipment/form.html', {
        'form': form, 'title': 'Новое оборудование', 'next_url': next_url,
    })


@login_required
def equipment_edit(request, pk):
    eq = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=eq)
        if form.is_valid():
            form.save()
            messages.success(request, 'Оборудование обновлено')
            return redirect('equipment_list')
    else:
        form = EquipmentForm(instance=eq)
    return render(request, 'core/equipment/form.html', {
        'form': form, 'title': 'Редактирование оборудования', 'equipment': eq,
        'back_url': list_back_url(request, 'equipment', 'equipment_list'),
    })


@login_required
def equipment_history(request, pk):
    """История ремонтов одной физической единицы оборудования.

    Оборудование уже выделено в отдельную сущность с уникальным серийным
    номером, поэтому история собирается обычной связью, без поиска по строке.
    """
    equipment = get_object_or_404(
        Equipment.objects.select_related('model', 'current_client'), pk=pk
    )
    visits = list(equipment.repair_history())

    # Детали в проекте списываются на заказ целиком, а не на конкретную
    # единицу оборудования в нём. Когда в заказе одна единица — список
    # деталей относится именно к ней; когда несколько, точнее сказать
    # нельзя, и это помечается в шаблоне, а не выдаётся за точные данные.
    visit_rows = []
    for visit in visits:
        order = visit.repair_order
        units_in_order = order.order_equipments.count()
        visit_rows.append({
            'roe': visit,
            'order': order,
            'parts': order.details.select_related('part'),
            'parts_are_exact': units_in_order == 1,
            'units_in_order': units_in_order,
        })

    similar = Equipment.find_similar(equipment.serial_number, exclude_pk=equipment.pk)

    return render(request, 'core/equipment/history.html', {
        'equipment': equipment,
        'back_url': list_back_url(request, 'equipment', 'equipment_list'),
        'visit_rows': visit_rows,
        'visits_count': len(visit_rows),
        'similar': similar,
        'warranty': equipment.active_warranty(),
    })


@login_required
def equipment_history_export(request, pk):
    """История ремонтов одной единицы в Excel — приложить к акту или письму."""
    equipment = get_object_or_404(Equipment.objects.select_related('model'), pk=pk)
    visits = equipment.repair_history()

    headers = [
        '№ заказа', 'Дата приёма', 'Дата завершения', 'Заказчик',
        'Неисправность', 'Начальное состояние', 'Номера пломб',
        'Стоимость, ₽', 'Гарантия до', 'Детали',
    ]
    rows = []
    for visit in visits:
        order = visit.repair_order
        # Детали списываются на заказ целиком. Когда единица в заказе одна,
        # список относится к ней; иначе честно помечаем, что он общий
        units = order.order_equipments.count()
        parts = ', '.join(
            f'{detail.part.name} x{detail.quantity_used}'
            for detail in order.details.select_related('part')
        )
        if parts and units > 1:
            parts = f'{parts} (на весь заказ, единиц в нём: {units})'

        warranty_until = visit.warranty_until
        rows.append([
            order.order_number,
            excel_datetime(order.date_received),
            excel_datetime(order.date_completed),
            order.client.name,
            visit.fault_description,
            visit.initial_condition,
            visit.seal_numbers,
            visit.repair_cost if visit.repair_cost is not None else '',
            excel_datetime(warranty_until).date() if warranty_until else '',
            parts,
        ])

    wb = build_workbook('История ремонтов', headers, rows)
    return xlsx_response(
        wb, f'История {equipment.serial_number} {timezone.localdate():%Y-%m-%d}.xlsx'
    )


@permission_required('catalog_delete')
def equipment_delete(request, pk):
    eq = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        eq.delete()
        messages.success(request, 'Оборудование удалено')
        return redirect('equipment_list')
    return render(request, 'core/equipment/delete.html', {'equipment': eq})


# ==================== МОДЕЛИ ОБОРУДОВАНИЯ ====================

@login_required
def equipment_model_list(request):
    models = (
        EquipmentModel.objects.select_related('equipment_type')
        .prefetch_related('versions').order_by('name')
    )
    return render(request, 'core/equipment/model_list.html', {'models': models})


def _equipment_kinds():
    """Родовые названия, которые уже вводили — подсказка в форме модели."""
    return list(
        EquipmentModel.objects.exclude(kind='')
        .order_by('kind').values_list('kind', flat=True).distinct()
    )


@login_required
def equipment_model_create(request):
    if request.method == 'POST':
        form = EquipmentModelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Модель добавлена')
            return redirect('equipment_model_list')
    else:
        form = EquipmentModelForm()
    return render(request, 'core/equipment/model_form.html', {
        'form': form, 'title': 'Новая модель оборудования', 'kinds': _equipment_kinds(),
        'back_url': reverse('equipment_model_list'),
    })


@login_required
def equipment_model_edit(request, pk):
    """Правка модели вместе с её материалами — схемами и инструкциями.

    Материалы только здесь, на странице создания их нет: пока модель
    не сохранена, вешать ссылки не на что. Набор форм создаётся
    с `form_kwargs={'equipment_model': model}` — иначе в выборе исполнения
    оказались бы исполнения всех моделей разом.

    `?next=` возвращает туда, откуда пришли, — обычно со страницы
    единицы, когда материалов или типа у модели ещё нет и мастер зашёл
    их завести. Без него — список моделей, как и раньше.
    """
    model = get_object_or_404(EquipmentModel, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or ''
    if not url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        next_url = ''
    if request.method == 'POST':
        form = EquipmentModelForm(request.POST, instance=model)
        formset = EquipmentMaterialFormSet(
            request.POST, instance=model, form_kwargs={'equipment_model': model}
        )
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Модель обновлена')
            return redirect(next_url or 'equipment_model_list')
        messages.error(request, 'Модель не сохранена: проверьте отмеченные поля')
    else:
        form = EquipmentModelForm(instance=model)
        formset = EquipmentMaterialFormSet(
            instance=model, form_kwargs={'equipment_model': model}
        )
    return render(request, 'core/equipment/model_form.html', {
        'form': form, 'formset': formset,
        'title': 'Редактирование модели', 'model': model,
        'kinds': _equipment_kinds(),
        'back_url': next_url or reverse('equipment_model_list'),
        'next_url': next_url,
    })


@permission_required('catalog_delete')
def equipment_model_delete(request, pk):
    model = get_object_or_404(EquipmentModel, pk=pk)
    if request.method == 'POST':
        model.delete()
        messages.success(request, 'Модель удалена')
        return redirect('equipment_model_list')
    return render(request, 'core/equipment/delete.html', {'equipment': model, 'is_model': True})


# ==================== ТИПЫ ОБОРУДОВАНИЯ ====================
# Права — как у EquipmentModel, соседнего справочника: список, создание
# и правка открыты любому вошедшему, удаление — по праву `catalog_delete`
# (должность с полным доступом проходит везде, см. decorators.py).

@login_required
def equipment_type_list(request):
    types = EquipmentType.objects.annotate(model_count=Count('models')).order_by('name')
    return render(request, 'core/equipment/type_list.html', {'types': types})


@login_required
def equipment_type_create(request):
    if request.method == 'POST':
        form = EquipmentTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Тип оборудования добавлен')
            return redirect('equipment_type_list')
    else:
        form = EquipmentTypeForm()
    return render(request, 'core/equipment/type_form.html', {
        'form': form, 'title': 'Новый тип оборудования',
    })


@login_required
def equipment_type_edit(request, pk):
    equipment_type = get_object_or_404(EquipmentType, pk=pk)
    if request.method == 'POST':
        form = EquipmentTypeForm(request.POST, instance=equipment_type)
        if form.is_valid():
            form.save()
            messages.success(request, 'Тип оборудования обновлён')
            return redirect('equipment_type_list')
    else:
        form = EquipmentTypeForm(instance=equipment_type)
    return render(request, 'core/equipment/type_form.html', {
        'form': form, 'title': 'Редактирование типа оборудования',
        'equipment_type': equipment_type,
    })


@permission_required('catalog_delete')
def equipment_type_delete(request, pk):
    equipment_type = get_object_or_404(EquipmentType, pk=pk)
    if request.method == 'POST':
        equipment_type.delete()
        messages.success(request, 'Тип оборудования удалён')
        return redirect('equipment_type_list')
    return render(request, 'core/equipment/type_delete.html', {
        'equipment_type': equipment_type,
        'model_count': equipment_type.models.count(),
    })


# ==================== ВЕРСИИ МОДЕЛЕЙ ====================

@login_required
def equipment_version_list(request):
    versions = (
        EquipmentVersion.objects.select_related('equipment_model')
        .annotate(equipment_count=Count('equipments'))
        .order_by('equipment_model__name', 'name')
    )
    return render(request, 'core/equipment/version_list.html', {'versions': versions})


def _equipment_model_names_json():
    """Названия моделей для подсказки «что напечатается».

    Кладём их в разметку сервером: отдельный запрос ради подсказки —
    ещё одна вещь, которая не приедет, когда со связью плохо, а подсказка
    нужна ровно в тот момент, когда обозначение набирают.
    """
    return json.dumps(
        {str(pk): name for pk, name in EquipmentModel.objects.values_list('pk', 'name')},
        ensure_ascii=False,
    )


@login_required
def equipment_version_create(request):
    if request.method == 'POST':
        form = EquipmentVersionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Версия добавлена')
            return redirect('equipment_version_list')
    else:
        # Со страницы модели версию заводят для неё же — подставляем её сразу
        form = EquipmentVersionForm(initial={'equipment_model': request.GET.get('model') or None})
    return render(request, 'core/equipment/version_form.html', {
        'form': form, 'title': 'Новая версия модели',
        'model_names': _equipment_model_names_json(),
    })


@login_required
def equipment_version_edit(request, pk):
    version = get_object_or_404(EquipmentVersion, pk=pk)
    if request.method == 'POST':
        form = EquipmentVersionForm(request.POST, instance=version)
        if form.is_valid():
            form.save()
            messages.success(request, 'Версия обновлена')
            return redirect('equipment_version_list')
    else:
        form = EquipmentVersionForm(instance=version)
    return render(request, 'core/equipment/version_form.html', {
        'form': form, 'title': 'Редактирование версии', 'version': version,
        'model_names': _equipment_model_names_json(),
    })


@permission_required('catalog_delete')
def equipment_version_delete(request, pk):
    version = get_object_or_404(EquipmentVersion, pk=pk)
    if request.method == 'POST':
        version.delete()
        messages.success(request, 'Версия удалена')
        return redirect('equipment_version_list')
    return render(request, 'core/equipment/version_delete.html', {
        'version': version,
        'equipment_count': version.equipments.count(),
        'recipe_count': version.fault_type_parts.count(),
    })


# ==================== ТИПОВЫЕ НЕИСПРАВНОСТИ ====================
# Права — как у EquipmentModel, ближайшего по смыслу справочника:
# создание и редактирование открыты любому вошедшему, удаление —
# по праву `catalog_delete` (полный доступ проходит везде).

@login_required
def fault_type_list(request):
    fault_types = (
        FaultType.objects.select_related('equipment_model')
        .prefetch_related('parts__part', 'parts__version')
        .order_by('equipment_model__name', 'name')
    )
    return render(request, 'core/faults/list.html', {'fault_types': fault_types})


# ==================== ТЕХНОЛОГИЧЕСКИЕ КАРТЫ ====================
# Карта отвечает на вопрос «как это делают руками». Привязана к модели
# всегда, к неисправности — по желанию: «как разобрать корпус» не про
# поломку, а про прибор. Права — как у соседних справочников: смотреть
# и править может любой вошедший, удалять — склад и мастер.


@login_required
def tech_card_list(request):
    cards = (
        TechCard.objects.select_related('equipment_model', 'fault_type')
        .annotate(step_count=Count('steps'))
        .order_by('equipment_model__name', 'title')
    )
    model_id = request.GET.get('model', '')
    if model_id:
        cards = cards.filter(equipment_model_id=model_id)
    return render(request, 'core/tech_cards/list.html', {
        'cards': cards,
        'models': EquipmentModel.objects.order_by('name'),
        'model_id': model_id,
    })


@login_required
def tech_card_detail(request, pk):
    """Карта на экране и на бумаге.

    `?version=` сужает шаги до одного исполнения — так на неё приходят
    со страницы дефектации, где исполнение единицы уже известно. Без него
    показываются только общие шаги: шаг чужого исполнения хуже, чем его
    отсутствие.
    """
    card = get_object_or_404(
        TechCard.objects.select_related('equipment_model', 'fault_type'), pk=pk
    )
    version = None
    version_id = request.GET.get('version', '')
    if version_id:
        version = card.equipment_model.versions.filter(pk=version_id).first()
    return render(request, 'core/tech_cards/detail.html', {
        'card': card,
        'version': version,
        'steps': card.steps_for(version).select_related('version'),
        # Исполнение просили, а его у этой модели нет — молчать нельзя:
        # человек решит, что карта полная, а половины шагов не увидит
        'version_missing': bool(version_id) and version is None,
    })


def _tech_card_page(request, card, title):
    """Общая работа страниц создания и правки карты.

    Одна на обе: набор шагов создаётся с моделью карты, иначе в выборе
    исполнения оказались бы исполнения всех моделей разом.

    «Назад» ведёт не всегда в список: карту чаще заводят со страницы
    единицы («у модели карт пока нет — завести карту»), и это она даёт
    `?next=`. Без него — карточка карты при правке (туда же вернулась бы
    и отмена) или список при создании без известного места, откуда пришли.
    """
    next_url = request.POST.get('next') or request.GET.get('next') or ''
    if not url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        next_url = ''

    model = card.equipment_model if card.pk else None
    if request.method == 'POST':
        form = TechCardForm(request.POST, instance=card)
        # Модель могли поменять в этом же запросе — исполнения берём
        # из присланной, а не из сохранённой: иначе правка «перенести
        # карту на другую модель» отвергала бы собственные исполнения
        posted = EquipmentModel.objects.filter(
            pk=request.POST.get('equipment_model') or 0
        ).first()
        formset = TechCardStepFormSet(
            request.POST, request.FILES, instance=card,
            form_kwargs={'equipment_model': posted or model},
        )
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                saved = form.save()
                formset.instance = saved
                formset.save()
            messages.success(request, 'Технологическая карта сохранена')
            return redirect('tech_card_detail', pk=saved.pk)
        messages.error(request, 'Карта не сохранена: проверьте отмеченные поля')
    else:
        form = TechCardForm(instance=card)
        formset = TechCardStepFormSet(
            instance=card, form_kwargs={'equipment_model': model}
        )
    back_url = next_url or (
        reverse('tech_card_detail', args=[card.pk]) if card.pk else reverse('tech_card_list')
    )
    return render(request, 'core/tech_cards/form.html', {
        'form': form, 'formset': formset, 'title': title, 'card': card,
        'back_url': back_url, 'next_url': next_url,
    })


@login_required
def tech_card_create(request):
    return _tech_card_page(request, TechCard(), 'Новая технологическая карта')


@login_required
def tech_card_edit(request, pk):
    card = get_object_or_404(TechCard, pk=pk)
    return _tech_card_page(request, card, 'Правка технологической карты')


@permission_required('catalog_delete')
def tech_card_delete(request, pk):
    card = get_object_or_404(TechCard, pk=pk)
    if request.method == 'POST':
        card.delete()
        messages.success(request, 'Технологическая карта удалена')
        return redirect('tech_card_list')
    return render(request, 'core/tech_cards/delete.html', {'card': card})


def _fault_type_copy_initial(source):
    """Чем заполнить форму новой неисправности, копируемой с образца.

    Копия — это только заполненная форма: пока человек не нажал
    «Сохранить», в базе не появляется ничего. Название помечается словом
    «копия», чтобы два одинаковых имени в списке не выглядели ошибкой.
    """
    form_initial = {
        'equipment_model': source.equipment_model_id,
        'name': f'{source.name} (копия)'[:255],
        'description': source.description,
        'work_description': source.work_description,
        'complexity': source.complexity,
        'versions': list(source.versions.values_list('pk', flat=True)),
    }
    lines_initial = [
        {'part': line.part_id, 'quantity': line.quantity, 'version': line.version_id}
        for line in source.parts.select_related('part', 'version').order_by('id')
    ]
    return form_initial, lines_initial


@login_required
def fault_type_create(request):
    """Новая типовая неисправность, в том числе копия существующей.

    `?copy_from=<номер>` подставляет в форму образец вместе со всем его
    рецептом — включая уточнения по версиям. Записи при этом не создаётся:
    копия существует ровно до тех пор, пока её не сохранили.
    """
    source = None
    copy_from = request.GET.get('copy_from') or request.POST.get('copy_from') or ''
    if str(copy_from).isdigit():
        source = FaultType.objects.filter(pk=copy_from).first()

    if request.method == 'POST':
        form = FaultTypeForm(request.POST)
        formset = FaultTypePartFormSet(request.POST, prefix='parts')
        # Модель нужна формсету до его проверки: уточнение рецепта сверяется
        # с моделью неисправности, а у ещё не сохранённой она известна
        # только из этой же формы
        form_is_valid = form.is_valid()
        if form_is_valid:
            formset.instance = form.instance
        if form_is_valid and formset.is_valid():
            fault_type = form.save()
            formset.instance = fault_type
            formset.save()
            messages.success(request, 'Типовая неисправность добавлена')
            return redirect('fault_type_list')
        messages.error(request, 'Неисправность не сохранена: проверьте отмеченные поля')
    elif source is not None:
        form_initial, lines_initial = _fault_type_copy_initial(source)
        form = FaultTypeForm(initial=form_initial)
        formset_class = make_fault_type_part_formset(extra=len(lines_initial) + 1)
        formset = formset_class(prefix='parts', initial=lines_initial)
    else:
        form = FaultTypeForm()
        formset = FaultTypePartFormSet(prefix='parts')
    return render(request, 'core/faults/form.html', {
        'form': form, 'formset': formset,
        'title': 'Копия типовой неисправности' if source is not None else 'Новая типовая неисправность',
        'copy_source': source,
        'versions': _fault_versions(),
    })


def _fault_versions():
    """Версии всех моделей — для выбора в строке рецепта.

    Отдаются списком с номером модели: страница прячет чужие версии
    сама, когда в форме выбрана модель.
    """
    return (
        EquipmentVersion.objects.select_related('equipment_model')
        .order_by('equipment_model__name', 'name')
    )


@login_required
def fault_type_edit(request, pk):
    fault_type = get_object_or_404(FaultType, pk=pk)
    if request.method == 'POST':
        form = FaultTypeForm(request.POST, instance=fault_type)
        formset = FaultTypePartFormSet(request.POST, instance=fault_type, prefix='parts')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Типовая неисправность обновлена')
            return redirect('fault_type_list')
        messages.error(request, 'Неисправность не сохранена: проверьте отмеченные поля')
    else:
        form = FaultTypeForm(instance=fault_type)
        formset = FaultTypePartFormSet(instance=fault_type, prefix='parts')
    return render(request, 'core/faults/form.html', {
        'form': form, 'formset': formset, 'title': 'Редактирование типовой неисправности',
        'fault_type': fault_type,
        'versions': _fault_versions(),
    })


@permission_required('catalog_delete')
def fault_type_delete(request, pk):
    fault_type = get_object_or_404(FaultType, pk=pk)
    if request.method == 'POST':
        fault_type.delete()
        messages.success(request, 'Типовая неисправность удалена')
        return redirect('fault_type_list')
    return render(request, 'core/faults/delete.html', {'fault_type': fault_type})


# ==================== ЗАКАЗЫ НА РЕМОНТ ====================

def _filter_orders(request):
    """Отбор заказов по параметрам GET-запроса.

    Поиск идёт по всему, что сотрудник может помнить о заказе: номеру,
    заказчику, серийнику любой единицы в заказе, номеру счёта, трек-номеру
    и описанию неисправности — как в самом заказе, так и по единицам.
    Искать «где-то это было» по одному полю за раз бессмысленно: обычно
    помнят обрывок, но не помнят, какое это было поле.
    """
    search = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    payment_status = request.GET.get('payment_status', '')
    client_id = request.GET.get('client', '')
    model_id = request.GET.get('model', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    orders = RepairOrder.objects.select_related('client').prefetch_related(
        'order_equipments__equipment__model'
    )

    if search:
        condition = (
            Q(order_number__icontains=search) |
            Q(client__name__icontains=search) |
            Q(client__inn__icontains=search) |
            Q(invoice_number__icontains=search) |
            Q(tracking_number__icontains=search) |
            Q(order_equipments__fault_description__icontains=search) |
            Q(order_equipments__equipment__serial_number__icontains=search) |
            Q(order_equipments__equipment__model__name__icontains=search)
        )
        # Серийник ищем и в нормализованном виде — «буад 1234» должен найти
        # заказ с «БУАД-1234», так же как это уже работает в оборудовании
        normalized = Equipment.normalize_serial(search)
        if normalized:
            condition |= Q(order_equipments__equipment__serial_normalized__contains=normalized)
        orders = orders.filter(condition)

    if status in dict(RepairOrder.STATUS_CHOICES):
        orders = orders.filter(status=status)
    if payment_status in dict(RepairOrder.PAYMENT_STATUS_CHOICES):
        orders = orders.filter(payment_status=payment_status)
    if client_id.isdigit():
        orders = orders.filter(client_id=int(client_id))
    if model_id.isdigit():
        orders = orders.filter(order_equipments__equipment__model_id=int(model_id))
    # Даты приходят из адресной строки и правятся руками — неразобранное
    # значение не применяется, иначе страница падала бы с ошибкой базы
    if parse_date(date_from):
        orders = orders.filter(date_received__date__gte=date_from)
    if parse_date(date_to):
        orders = orders.filter(date_received__date__lte=date_to)

    # distinct нужен из-за связи с оборудованием: заказ с тремя единицами
    # иначе попал бы в результат трижды
    return orders.distinct().order_by('-date_received'), {
        'search': search,
        'status_filter': status,
        'payment_status_filter': payment_status,
        'client_filter': client_id,
        'model_filter': model_id,
        'date_from': date_from,
        'date_to': date_to,
    }


# «Стоимость» и «Готовность» тут нет: первая — агрегат по единицам
# (`total_repair_cost`), сортировка по нему потребовала бы Sum-аннотации
# поверх уже двух M2M-джойнов (единицы, поиск) и рисковала бы задвоить
# суммы вместе с distinct(); вторая — вовсе не поле, а разбор чек-листа.
ORDER_LIST_SORT_FIELDS = {
    'order_number': 'order_number',
    'client': 'client__name',
    'status': 'status',
    'payment_status': 'payment_status',
    'date_received': 'date_received',
}


@login_required
def repair_order_list(request):
    orders, filter_context = _filter_orders(request)
    orders = sorted_by_request(orders, request, ORDER_LIST_SORT_FIELDS)

    # Готовность считается по единицам, поэтому по каждому заказу нужны
    # и его строки, и то, что спрашивает чек-лист. Предзагрузка — на одну
    # страницу списка, а не на всю выборку: без неё двадцать пять заказов
    # уходили бы в сотню запросов
    paginator = Paginator(
        orders.prefetch_related(
            'order_equipments__faults', 'order_equipments__details'
        ),
        25,
    )
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    # Признак «фильтры выставлены» — чтобы показать, сколько всего нашлось,
    # и дать кнопку сброса: иначе пустой список выглядит как пустая база
    filters_active = any(
        value for key, value in filter_context.items() if key != 'search'
    ) or bool(filter_context['search'])

    return render(request, 'core/repair_orders/list.html', {
        'orders': page_obj,
        'status_choices': RepairOrder.STATUS_CHOICES,
        'payment_status_choices': RepairOrder.PAYMENT_STATUS_CHOICES,
        'clients': Client.objects.order_by('name'),
        'equipment_models': EquipmentModel.objects.order_by('name'),
        'filters_active': filters_active,
        'found_count': paginator.count,
        **filter_context,
    })


@login_required
def repair_order_export(request):
    """Список заказов в Excel — с теми же условиями, что выставлены на странице."""
    orders, _ = _filter_orders(request)

    # Стоимости одним запросом. Считать их свойством заказа нельзя: это запрос
    # на строку, а выгружают обычно как раз длинные списки. Отдельный запрос
    # вместо annotate — потому что фильтры уже соединяют таблицу с
    # оборудованием, и сумма по такому соединению задвоилась бы.
    costs = dict(
        RepairOrderEquipment.objects
        .filter(repair_order__in=orders)
        .values_list('repair_order_id')
        .annotate(total=Sum('repair_cost'))
    )

    headers = [
        '№ заказа', 'Дата приёма', 'Заказчик', 'ИНН', 'Оборудование',
        'Статус ремонта', 'Статус оплаты', '№ счёта', 'Дата счёта',
        'Дата завершения', 'Трек-номер', 'Сумма, ₽',
    ]
    rows = []
    total_sum = 0
    for order in orders:
        cost = costs.get(order.pk) or 0
        total_sum += cost
        rows.append([
            order.order_number,
            excel_datetime(order.date_received),
            order.client.name,
            order.client.inn,
            ', '.join(str(roe.equipment) for roe in order.order_equipments.all()),
            order.get_status_display(),
            order.get_payment_status_display(),
            order.invoice_number,
            order.invoice_date,
            excel_datetime(order.date_completed),
            order.tracking_number,
            cost,
        ])
    if rows:
        rows.append(['Итого'] + [''] * (len(headers) - 2) + [total_sum])

    wb = build_workbook('Заказы', headers, rows)
    return xlsx_response(wb, f'Заказы {timezone.localdate():%Y-%m-%d}.xlsx')


@login_required
def repair_order_create(request):
    if request.method == 'POST':
        form = RepairOrderIntakeForm(request.POST)
        formset = RepairOrderEquipmentIntakeFormSet(request.POST, prefix='equipments')
        if form.is_valid() and formset.is_valid():
            order = form.save()
            formset.instance = order
            formset.save()
            order.assign_equipment_owners()
            messages.success(request, f'Заказ {order.order_number} создан')
            return redirect('repair_order_detail', pk=order.pk)
        # Без этого сообщения неудачное сохранение выглядело как успешное:
        # страница просто перезагружалась, а ошибки полей, которые шаблон
        # не выводил, оставались невидимыми — заказ молча не создавался
        messages.error(request, 'Заказ не сохранён: проверьте отмеченные поля')
    else:
        form = RepairOrderIntakeForm()
        formset = RepairOrderEquipmentIntakeFormSet(prefix='equipments')
    return render(request, 'core/repair_orders/form.html', {
        'form': form,
        'formset': formset,
        'title': 'Новый заказ на ремонт',
        # Форма приёма: остальные поля не спрятаны, а не показаны — их
        # заполняют позже, каждое в свой момент, и шаблон о них не знает.
        'intake': True,
    })


@login_required
@require_POST
def repair_order_edit_info(request, pk):
    """Заказчик и общее описание заказа — прямо на карточке.

    Всё, что относится к прибору, правится на странице единицы; сюда
    осталось то, что относится к заказу целиком. Отдельной страницы
    правки заказа больше нет: она держала те же поля вторым местом,
    и статус оплаты в ней спорил с формой на карточке.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    form = OrderInfoForm(request.POST, instance=order)
    if form.is_valid():
        saved = form.save()
        # Заказчика могли поставить только сейчас — а владелец
        # оборудования проставляется от него. Непустого не трогаем
        saved.assign_equipment_owners()
        messages.success(request, 'Заказ обновлён')
    else:
        messages.error(request, 'Не сохранено: проверьте отмеченные поля')
    return redirect('repair_order_detail', pk=pk)


@login_required
@require_POST
def repair_order_add_unit(request, pk):
    """Принять в уже существующий заказ ещё одну единицу — и сразу
    напечатать на неё наклейку.

    Так устроен приём: коробку поставили на стол, описали, наклеили ярлык,
    взяли следующую. Пока заказ сохранялся целиком, ярлыки печатались
    пачкой в конце, и их приходилось раскладывать по коробкам, сверяя
    номер позиции. Теперь заказ растёт по одной единице, и каждая уходит
    с наклейкой сразу.

    Отвечаем переходом на страницу наклейки, а не открытием окна из
    скрипта: окно, открытое не по щелчку человека, браузер часто
    не показывает вовсе — и наклейка молча не печаталась бы.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    equipment_id = request.POST.get('equipment')

    equipment = Equipment.objects.filter(pk=equipment_id).first() if equipment_id else None
    if equipment is None:
        messages.error(request, 'Единица не принята: выберите оборудование')
        return redirect('repair_order_detail', pk=order.pk)

    # Молчать нельзя: человек решит, что не сработало, и нажмёт ещё раз
    existing = order.order_equipments.filter(equipment=equipment).first()
    if existing:
        position = list(order.order_equipments.order_by('id')).index(existing) + 1
        messages.error(
            request,
            f'{equipment} уже в этом заказе, позиция {position} — второй раз не добавлено'
        )
        return redirect('repair_order_detail', pk=order.pk)

    # Прибор не может одновременно лежать на двух верстаках. Называем заказ,
    # в котором он сейчас: без этого непонятно, где его искать.
    busy = (
        RepairOrderEquipment.objects
        .filter(equipment=equipment, repair_order__status__in=RepairOrder.OPEN_STATUSES)
        .exclude(repair_order=order)
        .select_related('repair_order')
        .first()
    )
    if busy:
        messages.error(
            request,
            f'{equipment} сейчас в заказе {busy.repair_order.order_number} '
            f'({busy.repair_order.get_status_display().lower()}) — принять второй раз нельзя'
        )
        return redirect('repair_order_detail', pk=order.pk)

    roe = RepairOrderEquipment.objects.create(
        repair_order=order,
        equipment=equipment,
        fault_description=(request.POST.get('fault_description') or '').strip(),
        initial_condition=(request.POST.get('initial_condition') or '').strip(),
    )
    # Заказчик заказа — владелец привезённого прибора, если владельца ещё нет
    order.assign_equipment_owners()

    messages.success(request, f'{equipment} принята — печатайте наклейку')
    return redirect('repair_order_equipment_label', order_pk=order.pk, roe_pk=roe.pk)


@login_required
def repair_order_detail(request, pk):
    order = get_object_or_404(
        RepairOrder.objects.prefetch_related('order_equipments__equipment__model', 'client'),
        pk=pk
    )
    details = order.details.select_related('part', 'order_equipment__equipment__model')
    history = order.status_history.select_related('changed_by').order_by('-changed_at')
    # faults и details — ради готовности единицы: она спрашивает и то,
    # и другое по каждой строке, и без предзагрузки карточка заказа
    # с пятью приборами делала бы десяток лишних запросов
    order_equipments = list(
        order.order_equipments
        .select_related('equipment__model')
        .prefetch_related('faults', 'details')
        .order_by('id')
    )

    # Гарантия по прошлым ремонтам: текущий заказ исключён, иначе сразу после
    # его завершения он же и попадал бы в «повторное обращение по гарантии»
    previous_warranty = Equipment.warranty_map(
        [oe.equipment for oe in order_equipments], exclude_order_id=order.pk
    )
    for oe in order_equipments:
        oe.previous_warranty = previous_warranty.get(oe.equipment_id)

    readiness_pending = [oe for oe in order_equipments if not oe.is_ready]

    detail_form = RepairOrderDetailForm()
    status_form = StatusChangeForm()
    info_form = OrderInfoForm(instance=order)
    # Для формы «принять ещё единицу»: то, чего в этом заказе ещё нет
    # и что не лежит в другом незакрытом заказе
    available_equipment = (
        available_equipment_for_order(order)
        .select_related('model', 'version')
        .exclude(pk__in=[oe.equipment_id for oe in order_equipments])
        .order_by('model__name', 'serial_number')
    )
    # Себестоимость деталей: детали без цены в сумму не входят, поэтому
    # складываем то, что известно, а не подставляем ноль
    details_cost = sum(
        detail.cost for detail in details if detail.cost is not None
    )
    return render(request, 'core/repair_orders/detail.html', {
        'order': order,
        'details': details,
        'history': history,
        'order_equipments': order_equipments,
        # Готовность считаем по уже загруженным строкам, а не заново:
        # order.readiness() сходил бы в базу второй раз за тем же самым
        'readiness_pending': readiness_pending,
        'readiness_ready': len(order_equipments) - len(readiness_pending),
        'details_cost': details_cost,
        'payments': order.payments.select_related('created_by'),
        'payment_form': PaymentForm(),
        'detail_form': detail_form,
        'info_form': info_form,
        'available_equipment': available_equipment,
        'status_form': status_form,
    })


def _cost_from_allocations(movement):
    """Себестоимость исходящего движения `movement` по фактически задействованным
    партиям (`StockAllocation`), уже созданным для него.

    None — сумма неизвестна целиком: либо часть количества осталась без
    партии (нехватка остатка на момент списания), либо хотя бы у одной
    задействованной партии не заполнена цена. Частичная сумма без учёта
    неизвестного остатка не возвращается — она вводила бы в заблуждение,
    будто себестоимость посчитана полностью.
    """
    allocations = list(movement.allocations.select_related('incoming'))
    covered = sum(a.quantity for a in allocations)
    if covered < movement.quantity:
        return None
    if any(a.incoming.unit_price is None for a in allocations):
        return None
    return sum((a.quantity * a.incoming.unit_price for a in allocations), Decimal('0'))


def _use_repair_order_part(order, part, quantity, employee, history_note,
                           order_equipment=None):
    """Списывает деталь со склада и создаёт запись использованной в заказе
    детали — общая часть добавления детали вручную и применения шаблона
    неисправности.

    Не форма и не отдельная транзакция: вызывающая сторона решает, сколько
    таких вызовов войдёт в одну атомарную операцию — единственную деталь
    (`repair_order_add_detail`) или пачку по шаблону
    (`apply_fault_templates`), где частичный сбой посреди применения не
    должен оставить заказ с половиной добавленных деталей.

    Списание распределяется по партиям прихода (FIFO — см.
    `StockAllocation.allocate`), и по нему заводится запись затраты
    (`OrderCost`, category='parts') — этим и отличается от ручного
    списания (`part_stock_outgoing`), не привязанного к заказу.

    `order_equipment` — в какую единицу ушла деталь. Не передана и прибор
    в заказе один — привязывается к нему: ответ очевиден, и спрашивать
    незачем. Приборов несколько — деталь остаётся общей по заказу, пока
    мастер не укажет; угадывать программа не станет.

    Возвращает True, если остатка не хватило и он ушёл в минус.
    """
    shortage = part.current_stock < quantity
    if order_equipment is None:
        order_equipment = order.sole_equipment
    detail = RepairOrderDetail.objects.create(
        repair_order=order, order_equipment=order_equipment,
        part=part, quantity_used=quantity
    )
    part.current_stock -= quantity
    part.save(update_fields=['current_stock'])
    movement = StockMovement.objects.create(
        part=part,
        quantity=quantity,
        movement_type='outgoing',
        repair_order=order,
        notes=f'Списано по заказу {order.order_number}',
        created_by=employee
    )
    StockAllocation.allocate(movement)
    # Связь нужна возврату: чтобы вернуть деталь в её партию, надо знать,
    # из каких партий её брали, а это записано в распределении расхода.
    detail.movement = movement
    detail.save(update_fields=['movement'])
    OrderCost.objects.create(
        repair_order=order, category='parts', amount=_cost_from_allocations(movement)
    )
    OrderStatusHistory.objects.create(
        order=order,
        status=order.status,
        changed_by=employee,
        notes=history_note
    )
    return shortage


def apply_fault_templates(order, fault_types, employee, version=None,
                          order_equipment=None):
    """Строит объединённый рецепт деталей по выбранным неисправностям и
    списывает его одной атомарной транзакцией на все позиции сразу.

    Одна и та же деталь в рецептах нескольких выбранных неисправностей даёт
    одну позицию с суммарным количеством — а не несколько строк с частями
    этого количества. С уже имеющимися в заказе деталями (введёнными вручную
    или добавленными прошлым применением шаблона) слияния нет: шаблон
    дополняет список, а не пересчитывает его целиком.

    `version` — версия исполнения той единицы, к которой применяют рецепт.
    Известна — берутся уточнения рецепта для неё (см.
    `FaultType.recipe_lines`), неизвестна — только общие строки.

    `order_equipment` — та же единица, но чтобы записать, в какую железку
    ушли детали: рецепт применяют из её строки, значит и детали её.

    Возвращает (added, shortages): added — список (SparePart, количество)
    добавленных позиций в порядке первого появления детали среди рецептов;
    shortages — те из них, на которые не хватило остатка на складе.
    """
    merged_qty = {}
    merged_part = {}
    order_of_appearance = []
    for fault in fault_types:
        for line in fault.recipe_lines(version):
            if line.part_id not in merged_qty:
                merged_qty[line.part_id] = 0
                merged_part[line.part_id] = line.part
                order_of_appearance.append(line.part_id)
            merged_qty[line.part_id] += line.quantity

    added = []
    shortages = []
    with transaction.atomic():
        for part_id in order_of_appearance:
            part = merged_part[part_id]
            quantity = merged_qty[part_id]
            shortage = _use_repair_order_part(
                order, part, quantity, employee,
                f'Деталь {part.name} x{quantity} добавлена по шаблону неисправности',
                order_equipment=order_equipment
            )
            added.append((part, quantity))
            if shortage:
                shortages.append(part)
    return added, shortages


def _back_after_detail(request, order_pk):
    """Куда вернуться после работы с деталями заказа.

    Детали списывают из двух мест: на карточке заказа (там же — «на заказ
    целиком») и на странице единицы, где не надо выбирать, в какую железку.
    Возвращаться надо туда, откуда пришли, иначе мастер, списавший три
    детали подряд, три раза уезжает в заказ.

    Адрес принимается только свой: чужой превратил бы форму в способ
    увести человека на постороннюю страницу.
    """
    target = request.POST.get('next') or ''
    if target and url_has_allowed_host_and_scheme(
        target, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        return redirect(target)
    return redirect('repair_order_detail', pk=order_pk)


@login_required
@require_POST
def repair_order_add_detail(request, pk):
    """Добавление детали в заказ со списанием со склада (явная транзакция)."""
    order = get_object_or_404(RepairOrder, pk=pk)
    form = RepairOrderDetailForm(request.POST)
    if form.is_valid():
        part = form.cleaned_data['part']
        quantity = form.cleaned_data['quantity_used']
        # В какую единицу ушла деталь. С v2.96.0 «на заказ целиком» больше
        # не выбирают (решение владельца): деталь ставят в конкретный
        # прибор, иначе по этой записи потом не посчитать, во сколько
        # обошёлся его ремонт. Прибор в заказе один — привязка
        # проставляется сама (см. _use_repair_order_part).
        unit_id = request.POST.get('order_equipment')
        unit = order.order_equipments.filter(pk=unit_id).first() if unit_id else None
        if unit is None and order.order_equipments.count() > 1:
            messages.error(
                request,
                'Не списано: укажите, в какой прибор ушла деталь. '
                'В заказе их несколько, и угадывать программа не станет.'
            )
            return _back_after_detail(request, pk)
        # «Запланировать» вместо «Добавить»: деталь нужна, но со склада
        # ещё не взята — остаток трогать рано.
        if request.POST.get('plan'):
            if unit is None:
                unit = order.sole_equipment
            RepairOrderDetail.objects.create(
                repair_order=order, order_equipment=unit,
                part=part, quantity_used=quantity, is_planned=True
            )
            messages.success(
                request, f'{part.name} x{quantity} запланирована — склад не тронут'
            )
            return _back_after_detail(request, pk)

        if part.current_stock < quantity:
            messages.warning(request,
                f'Внимание: недостаточно {part.name} на складе! Текущий остаток: {part.current_stock}. '
                f'Будет списано с отрицательным остатком.')

        with transaction.atomic():
            _use_repair_order_part(
                order, part, quantity, request.user,
                f'Добавлена деталь {part.name} x{quantity}',
                order_equipment=unit
            )

        messages.success(request, f'Деталь {part.name} добавлена в заказ')
    else:
        messages.error(request, 'Ошибка при добавлении детали')
    return _back_after_detail(request, pk)


def _write_off_planned(detail, employee):
    """Превратить запланированную строку в настоящее списание.

    Одно место на оба случая — одну строку и весь список разом: второй
    такой код однажды разошёлся бы с первым, а склад не то место, где
    это можно позволить. Списание идёт ровно тем же путём, что и обычное
    (`_use_repair_order_part`): распределение по партиям FIFO, затрата
    по заказу, запись в истории. Отличается только тем, что строка
    в заказе уже есть — её убираем, а не заводим вторую.

    Возвращает True, если списывать пришлось в минус.
    """
    order = detail.repair_order
    part = detail.part
    quantity = detail.quantity_used
    shortage = part.current_stock < quantity
    unit = detail.order_equipment

    detail.delete()
    _use_repair_order_part(
        order, part, quantity, employee,
        f'Списана запланированная деталь {part.name} x{quantity}',
        order_equipment=unit
    )
    return shortage


@login_required
@require_POST
def repair_order_write_off_detail(request, pk, detail_pk):
    """Списать запланированную деталь: теперь она действительно ушла
    в прибор.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    detail = get_object_or_404(
        RepairOrderDetail, pk=detail_pk, repair_order=order, is_planned=True
    )
    part = detail.part
    quantity = detail.quantity_used

    if part.current_stock < quantity:
        messages.warning(
            request,
            f'Внимание: недостаточно {part.name} на складе! Текущий остаток: '
            f'{part.current_stock}. Будет списано с отрицательным остатком.'
        )

    with transaction.atomic():
        _write_off_planned(detail, request.user)

    messages.success(request, f'{part.name} x{quantity} списана со склада')
    return _back_after_detail(request, order.pk)


@login_required
@require_POST
def repair_order_unit_write_off_planned(request, order_pk, roe_pk):
    """Списать со склада весь намеченный список деталей этой единицы
    (с v2.96.0).

    Так это и происходит у стола: мастер вскрыл прибор, набрал список
    того, чтоменять, и только потом идёт к стеллажу. До этого каждая
    строка списывалась в тот же миг, когда её вписали, — то есть склад
    менялся, пока мастер ещё думал.

    Списывается каждая строка **тем же** путём, что и по одной
    (`_write_off_planned`), и всё одной транзакцией: наполовину списанный
    список хуже не списанного вовсе — по нему уже не понять, что взято.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)
    planned = list(
        order_equipment.details.filter(is_planned=True).select_related('part')
    )

    if not planned:
        messages.info(request, 'Списывать нечего: намеченных деталей нет.')
        return redirect(
            reverse('repair_order_unit_detail', args=[order_pk, roe_pk]) + '#parts'
        )

    short = []
    written = []
    with transaction.atomic():
        for detail in planned:
            name = detail.part.name
            written.append(f'{detail.part.part_number} x{detail.quantity_used}')
            if _write_off_planned(detail, request.user):
                short.append(name)

    # Перечисляем, что именно ушло с полки, а не считаем строки: мастер
    # сверяет это с тем, что взял в руки. Длинный список подрезаем —
    # в полосе сообщения читают первые слова
    shown = ', '.join(written[:6])
    if len(written) > 6:
        shown += f' и ещё {len(written) - 6}'
    messages.success(request, f'Списано со склада: {shown}')
    # Списали в минус — сказать вслух: остаток на полке и в программе
    # разошлись, и знать об этом надо сейчас, а не при инвентаризации
    if short:
        messages.warning(
            request,
            'Списано с отрицательным остатком: %s.' % ', '.join(short),
        )
    return redirect(
        reverse('repair_order_unit_detail', args=[order_pk, roe_pk]) + '#parts'
    )


@login_required
@require_POST
def repair_order_cancel_planned_detail(request, pk, detail_pk):
    """Убрать деталь из плана. Со склада её не брали — возвращать нечего."""
    order = get_object_or_404(RepairOrder, pk=pk)
    detail = get_object_or_404(
        RepairOrderDetail, pk=detail_pk, repair_order=order, is_planned=True
    )
    name = detail.part.name
    detail.delete()
    messages.success(request, f'{name} убрана из плана')
    return _back_after_detail(request, order.pk)


@login_required
@require_POST
def repair_order_return_detail(request, pk, detail_pk):
    """Вернуть деталь из заказа на склад — в те партии, из которых брали.

    Мастер вскрыл прибор, деталь не понадобилась или взял с запасом.
    Возврат отменяет списание ровно наоборот тому, как оно делалось,
    и себестоимость заказа уменьшается на стоимость возвращённого.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    detail = get_object_or_404(
        RepairOrderDetail, pk=detail_pk, repair_order=order, is_planned=False
    )

    raw = request.POST.get('quantity', '')
    if not str(raw).isdigit():
        messages.error(request, 'Возврат не сделан: укажите количество')
        return redirect('repair_order_detail', pk=order.pk)

    try:
        detail.return_to_stock(int(raw), request.user)
    except ValidationError as error:
        messages.error(request, '; '.join(error.messages))
    else:
        messages.success(
            request, f'{detail.part.name}: возвращено {raw} шт. на склад'
        )
    return _back_after_detail(request, order.pk)


@login_required
@require_POST
def repair_order_apply_fault_template(request, pk):
    """Применение рецептов деталей выбранных типовых неисправностей.

    Дополняет список использованных в заказе деталей, ничего в нём не
    заменяя — та же логика списания, что и у ручного добавления детали
    (см. `_use_repair_order_part`), но одним вызовом на каждую позицию уже
    слитого по всем выбранным неисправностям словаря и одной транзакцией
    на всю пачку сразу.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    fault_ids = [v for v in request.POST.getlist('fault_ids') if str(v).isdigit()]
    fault_types = list(FaultType.objects.filter(pk__in=fault_ids).prefetch_related('parts__part'))

    # Версия исполнения берётся у единицы, из строки которой нажали кнопку:
    # уточнения рецепта расписаны именно по версиям. Единица не передана
    # или версии у неё нет — действует общая часть рецепта
    version = None
    target_unit = None
    equipment_id = request.POST.get('equipment_id', '')
    if str(equipment_id).isdigit():
        equipment = Equipment.objects.filter(pk=equipment_id).select_related('version').first()
        if equipment is not None:
            version = equipment.version
            # Рецепт применяют из строки конкретной единицы — значит
            # и детали уходят в неё, а не «в заказ вообще»
            target_unit = order.order_equipments.filter(equipment=equipment).first()

    if not fault_types:
        return JsonResponse({
            'success': False,
            'error': 'Выберите хотя бы одну неисправность из списка — «Другое» своего рецепта не имеет.'
        })

    added, shortages = apply_fault_templates(
        order, fault_types, request.user, version=version, order_equipment=target_unit
    )
    if not added:
        return JsonResponse({
            'success': False,
            'error': 'В рецепте выбранных неисправностей деталей не указано.'
        })

    lines = ', '.join(f'{part.name} ×{quantity}' for part, quantity in added)
    message = f'Шаблон применён, в заказ добавлено: {lines}.'
    if shortages:
        names = ', '.join(sorted({part.name for part in shortages}))
        message += f' Внимание: не хватило на складе — {names}, списано с отрицательным остатком.'

    return JsonResponse({'success': True, 'message': message, 'count': len(added)})


@login_required
@require_POST
def repair_order_change_status(request, pk):
    """Изменение статуса заказа с логированием."""
    order = get_object_or_404(RepairOrder, pk=pk)
    form = StatusChangeForm(request.POST)
    if form.is_valid():
        new_status = form.cleaned_data['new_status']
        old_status = order.status

        if old_status == new_status:
            messages.info(request, 'Статус не изменился')
            return redirect('repair_order_detail', pk=pk)

        order.status = new_status
        if new_status == 'shipped':
            order.shipping_date = timezone.now()
            order.date_completed = timezone.now()
        else:
            # Сбрасываем даты, если заказ вернули из отгруженного
            order.date_completed = None
            if old_status == 'shipped':
                order.shipping_date = None
        order.save()

        # Создаём новую запись истории (а не обновляем старую). Текст
        # подставляет программа — своего примечания у мастера здесь
        # больше не спрашивают (v2.94.0)
        OrderStatusHistory.objects.create(
            order=order,
            status=new_status,
            changed_by=request.user,
            notes=f'Статус изменён с "{dict(RepairOrder.STATUS_CHOICES).get(old_status)}"'
        )

        # Оповещение заказчику — в очередь, не отправкой на месте: SMTP через
        # домашний канал может думать секундами, а страница ждать не должна
        notifications.notify_order_status(order, changed_by=request.user)

        messages.success(request, f'Статус изменён на «{order.get_status_display()}»')
        _warn_about_unfinished_units(request, order, new_status)
    return redirect('repair_order_detail', pk=pk)


# Статусы, при которых недоделки по единицам стоит называть вслух: дальше
# прибор уезжает к заказчику, и незаполненный акт всплывёт уже у него.
READINESS_WARNING_STATUSES = ('ready_for_shipment', 'shipped', 'partially_repaired')


def _warn_about_unfinished_units(request, order, new_status):
    """Сказать, по каким единицам осталась работа.

    Именно сказать, а не запретить: мастер видит прибор, а программа нет.
    Бывает ремонт, где записывать нечего, и запрет здесь означал бы, что
    статус проставляют «как-нибудь», лишь бы программа пропустила.

    Перечисляем не больше трёх единиц: в сообщении на весь экран список
    из десяти всё равно не читают, а полный разбор стоит карточкой
    на самой странице заказа.
    """
    if new_status not in READINESS_WARNING_STATUSES:
        return
    pending = order.readiness()['pending']
    if not pending:
        return
    shown = ', '.join(unit.equipment.serial_number for unit in pending[:3])
    if len(pending) > 3:
        shown += f' и ещё {len(pending) - 3}'
    messages.warning(
        request,
        f'Осталась работа по оборудованию ({len(pending)} из '
        f'{order.order_equipments.count()}): {shown}. Что именно — '
        f'в чек-листе готовности на карточке заказа.'
    )


# Единственный переход, разрешённый массовой сменой статуса. Курьер обычно
# забирает сразу пачку готовых заказов — это одно реальное событие, и ошибка
# «применил не к тому заказу» здесь маловероятна. У остальных переходов
# («в ремонт», «готов к отгрузке» и т.д.) риск выше: часть выбранных заказов
# может незаметно не подойти, и решение о них лучше принимать по одному.
BULK_STATUS_FROM = 'ready_for_shipment'
BULK_STATUS_TO = 'shipped'


@login_required
def repair_order_bulk_status(request):
    """Массовая отгрузка отмеченных заказов.

    GET — только показ: что применится, что будет пропущено и почему,
    без единого изменения в базе (страница подтверждения). POST выполняет
    переход для тех заказов, что на этот момент действительно «Готовы
    к отгрузке» — остальные пропускаются с явной причиной, а не молча.
    """
    ids = request.POST.getlist('ids') if request.method == 'POST' else _selected_ids(request)
    orders = list(
        RepairOrder.objects
        .filter(pk__in=[value for value in ids if str(value).isdigit()])
        .select_related('client')
        .prefetch_related(
            'order_equipments__equipment__model',
            # Готовность спрашивает и то, и другое по каждой единице:
            # без предзагрузки страница подтверждения на десяток заказов
            # делала бы сотню запросов
            'order_equipments__faults', 'order_equipments__details',
        )
        .order_by('pk')
    )

    ineligible = [
        {'order': order, 'reason': f'Статус «{order.get_status_display()}», а не «Готов к отгрузке»'}
        for order in orders if order.status != BULK_STATUS_FROM
    ]
    eligible = [order for order in orders if order.status == BULK_STATUS_FROM]

    if request.method == 'POST':
        if not orders:
            messages.warning(request, 'Отгружать нечего: ни один заказ не отмечен')
            return redirect('repair_order_list')

        now = timezone.now()
        with transaction.atomic():
            for order in eligible:
                order.status = BULK_STATUS_TO
                order.shipping_date = now
                order.date_completed = now
                order.save()
                OrderStatusHistory.objects.create(
                    order=order,
                    status=BULK_STATUS_TO,
                    changed_by=request.user,
                    notes='Массовая отгрузка отмеченных заказов',
                )

        # Оповещения — после того, как переходы записаны: письмо не должно
        # мешать основной операции, и по той же причине, что и у одиночной
        # смены статуса, кладём их в очередь, а не отправляем на месте
        for order in eligible:
            notifications.notify_order_status(order, changed_by=request.user)

        if eligible:
            messages.success(request, f'Отгружено заказов: {len(eligible)}')
        if ineligible:
            messages.warning(
                request,
                f'Пропущено без изменений: {len(ineligible)} — '
                'не в статусе «Готов к отгрузке»'
            )

        return render(request, 'core/repair_orders/bulk_status.html', {
            'done': True,
            'applied': eligible,
            'skipped': ineligible,
        })

    return render(request, 'core/repair_orders/bulk_status.html', {
        'done': False,
        'orders': orders,
        'eligible': eligible,
        'ineligible': ineligible,
        'from_status': dict(RepairOrder.STATUS_CHOICES).get(BULK_STATUS_FROM),
        'to_status': dict(RepairOrder.STATUS_CHOICES).get(BULK_STATUS_TO),
    })


@permission_required('payments_manage')
@require_POST
def repair_order_add_payment(request, pk):
    """Внести поступившие деньги по заказу.

    Статус оплаты после этого пересчитывается сам (см. core/signals.py):
    вручную выставленный «частично оплачен» после последнего платежа
    оставлял бы заказ в должниках навсегда.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    form = PaymentForm(request.POST)
    if not form.is_valid():
        error = next(iter(form.errors.values()))[0]
        messages.error(request, f'Оплата не внесена: {error}')
        return redirect('repair_order_detail', pk=pk)

    payment = form.save(commit=False)
    payment.repair_order = order
    payment.created_by = request.user
    payment.save()

    order.refresh_from_db()
    OrderStatusHistory.objects.create(
        order=order,
        payment_status=order.payment_status,
        changed_by=request.user,
        notes=f'Внесена оплата {payment.amount} ₽'
              + (f' — {payment.note}' if payment.note else ''),
    )

    remaining = order.debt
    messages.success(
        request,
        f'Оплата {payment.amount} ₽ внесена. '
        + (f'Остаток: {remaining} ₽' if remaining else 'Заказ оплачен полностью')
    )
    return redirect('repair_order_detail', pk=pk)


@permission_required('payments_manage')
@require_POST
def repair_order_delete_payment(request, pk, payment_pk):
    """Убрать ошибочно внесённую оплату.

    Суммы вбивают руками, и опечатка в разряде — обычное дело; без этой
    кнопки чинить пришлось бы через административную часть Django.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    payment = get_object_or_404(Payment, pk=payment_pk, repair_order=order)
    amount = payment.amount
    payment.delete()

    order.refresh_from_db()
    OrderStatusHistory.objects.create(
        order=order,
        payment_status=order.payment_status,
        changed_by=request.user,
        notes=f'Удалена оплата {amount} ₽',
    )
    messages.success(request, f'Оплата {amount} ₽ удалена')
    return redirect('repair_order_detail', pk=pk)


@permission_required('payment_status_change')
@require_POST
def repair_order_change_payment_status(request, pk):
    """Изменение статуса оплаты заказа с логированием."""
    order = get_object_or_404(RepairOrder, pk=pk)
    new_payment_status = request.POST.get('payment_status')
    notes = request.POST.get('notes', '')

    if new_payment_status not in dict(RepairOrder.PAYMENT_STATUS_CHOICES):
        messages.error(request, 'Недопустимый статус оплаты')
        return redirect('repair_order_detail', pk=pk)

    old_payment_status = order.payment_status
    if old_payment_status == new_payment_status:
        messages.info(request, 'Статус оплаты не изменился')
        return redirect('repair_order_detail', pk=pk)

    order.payment_status = new_payment_status
    order.save()

    # Логируем изменение статуса оплаты
    OrderStatusHistory.objects.create(
        order=order,
        payment_status=new_payment_status,
        changed_by=request.user,
        notes=notes or f'Статус оплаты изменён с "{dict(RepairOrder.PAYMENT_STATUS_CHOICES).get(old_payment_status)}"'
    )

    messages.success(request, f'Статус оплаты изменён на «{order.get_payment_status_display()}»')
    return redirect('repair_order_detail', pk=pk)


@permission_required('orders_delete')
def repair_order_delete(request, pk):
    order = get_object_or_404(RepairOrder, pk=pk)
    if request.method == 'POST':
        order.delete()
        messages.success(request, 'Заказ удалён')
        return redirect('repair_order_list')
    return render(request, 'core/repair_orders/delete.html', {'order': order})


# ==================== ДЕТАЛИ / ЗАПЧАСТИ ====================

def _as_number(value):
    """Число из строки запроса или None.

    Значения приходят из адресной строки, и её правят руками. Раньше здесь
    стоял голый float(), и «5 В» вместо «5» роняло страницу целиком.
    """
    try:
        return float(str(value).replace(',', '.'))
    except (TypeError, ValueError):
        return None


def _filter_parts(params):
    """Общая фильтрация деталей по параметрам запроса.

    Одна на всех, кто отбирает детали: список, выгрузка и поиск для выбора
    детали из формы. Второй набор условий рядом с этим неизбежно разошёлся
    бы с ним — и «найдено 3» в списке не совпадало бы с тем, что предлагает
    выбор детали в заказе.

    `params` — словарь параметров запроса (`request.GET`), а не сам запрос:
    так же отбирает и то, что фильтров в адресной строке не имеет.
    """
    search = params.get('q', '')
    component_type = params.get('component_type', '')
    package = params.get('package', '')

    # Диапазоны характеристик: поле в модели -> префикс параметра
    ranges = {
        'voltage': 'voltage',
        'current': 'current',
        'resistance': 'resistance',
        'capacitance': 'capacitance',
        'power': 'power',
    }
    stock_state = params.get('stock_state', '')
    # Старая ссылка вида ?below_min=1 продолжает работать
    if not stock_state and params.get('below_min'):
        stock_state = 'below'

    stock_from = params.get('stock_from', '')
    stock_to = params.get('stock_to', '')
    # «Только то, что есть в наличии» — отдельно от условий про минимум:
    # выбирающему деталь важно, лежит она на полке или нет, а не насколько
    # запас близок к минимальному
    in_stock = params.get('in_stock', '')
    # Деталь без ячейки — та, что ещё некуда положить (или сняли с прежней
    # и не переложили). storage_cells — M2M, isnull=True не двоит строки:
    # непопавшую в join запись выбрать пусто, а не отсутствие.
    no_cell = params.get('no_cell', '')

    parts = SparePart.objects.all()
    if search:
        parts = parts.filter(
            Q(part_number__icontains=search) |
            Q(name__icontains=search) |
            Q(component_type__icontains=search) |
            Q(package__icontains=search) |
            Q(description__icontains=search)
        )
    if component_type:
        parts = parts.filter(component_type=component_type)
    if package:
        parts = parts.filter(package=package)
    if in_stock:
        parts = parts.filter(current_stock__gt=0)
    if no_cell:
        parts = parts.filter(storage_cells__isnull=True)

    context = {
        'search': search,
        'component_type': component_type,
        'package': package,
        'stock_from': stock_from,
        'stock_to': stock_to,
        'stock_state': stock_state,
        'in_stock': in_stock,
        'no_cell': no_cell,
    }

    for field, prefix in ranges.items():
        for bound, lookup in (('from', 'gte'), ('to', 'lte')):
            raw = params.get(f'{prefix}_{bound}', '')
            context[f'{prefix}_{bound}'] = raw
            value = _as_number(raw)
            if value is not None:
                parts = parts.filter(**{f'{field}__{lookup}': value})

    stock_low = _as_number(stock_from)
    stock_high = _as_number(stock_to)
    if stock_low is not None:
        parts = parts.filter(current_stock__gte=int(stock_low))
    if stock_high is not None:
        parts = parts.filter(current_stock__lte=int(stock_high))

    if stock_state == 'below':
        parts = parts.below_minimum()
    elif stock_state == 'at_minimum':
        parts = parts.at_minimum()
    elif stock_state == 'attention':
        # Дефицит и «ровно на минимуме» вместе — то, на что стоит смотреть
        parts = parts.filter(
            Q(current_stock__lt=F('min_stock')) |
            Q(min_stock__gt=0, current_stock=F('min_stock'))
        )

    return parts, context


# Сортируемые столбцы списка радиодеталей — «Ячейка» не в их числе:
# `current_cell` не поле, а python-свойство (первая ячейка из M2M),
# `order_by` до неё не дотянется.
PART_LIST_SORT_FIELDS = {
    'part_number': 'part_number',
    'name': 'name',
    'component_type': 'component_type',
    'current_stock': 'current_stock',
    'min_stock': 'min_stock',
}


@login_required
def part_list(request):
    parts, filter_context = _filter_parts(request.GET)
    parts = sorted_by_request(parts.order_by('part_number'), request, PART_LIST_SORT_FIELDS)

    paginator = Paginator(parts, 25)
    page = request.GET.get('page')
    remember_list_query(request, 'parts')
    return render(request, 'core/parts/list.html', {
        'parts': paginator.get_page(page),
        **_part_choices(),
        'found_count': paginator.count,
        'filters_active': any(filter_context.values()),
        **filter_context,
    })


# Сколько деталей отдавать в выбор детали за один запрос. Каталог — сотни
# записей, и вываливать их целиком незачем: список длиннее полусотни уже
# не просматривают, его уточняют. Что показано не всё, ответ говорит прямо —
# иначе человек решит, что остальных деталей в программе нет
PART_SEARCH_LIMIT = 50


def _part_search_row(part):
    """Одна строка выбора детали.

    Остаток и адрес ячейки — не украшение: по ним решают, идти за деталью
    к полке или заказывать её, и выбор без них просто отодвигает вопрос
    на шаг дальше.
    """
    # Ячейку берём из уже загруженного списка, а не через current_cell:
    # тот делает свой запрос на каждую деталь и сводит prefetch на нет
    cells = list(part.storage_cells.all())
    return {
        'id': part.pk,
        'part_number': part.part_number,
        'name': part.name,
        'label': f'{part.part_number} — {part.name}',
        'specs': part.specs_display,
        'component_type': part.component_type,
        'package': part.package,
        'stock': part.current_stock,
        'min_stock': part.min_stock,
        'stock_state': part.stock_state,
        'cell': cells[0].address if cells else '',
    }


@login_required
def part_search(request):
    """Поиск детали для выбора из формы (JSON).

    Отбор — тот же `_filter_parts`, что у списка радиодеталей: выбор детали
    в заказе должен находить ровно то же, что и склад.

    `exclude` — детали, которые в этом месте выбирать не из чего (например,
    уже лежащие в этой ячейке). `id` — обратный запрос: как называется
    деталь с таким номером; им подписывается уже сделанный выбор, когда
    страница вернулась с ошибкой формы.
    """
    parts, _ = _filter_parts(request.GET)

    part_id = request.GET.get('id', '')
    if part_id.isdigit():
        parts = parts.filter(pk=int(part_id))

    exclude = [value for value in request.GET.get('exclude', '').split(',') if value.isdigit()]
    if exclude:
        parts = parts.exclude(pk__in=[int(value) for value in exclude])

    parts = parts.prefetch_related('storage_cells__cabinet').order_by('part_number')
    total = parts.count()
    data = {
        'results': [_part_search_row(part) for part in parts[:PART_SEARCH_LIMIT]],
        'total': total,
        'limit': PART_SEARCH_LIMIT,
        'limited': total > PART_SEARCH_LIMIT,
    }
    # Список типов компонентов нужен выбору один раз — на первое открытие.
    # Отдавать его на каждое нажатие клавиши значило бы гонять лишний
    # запрос к базе на каждую букву
    if request.GET.get('with_types'):
        data['component_types'] = _part_choices()['component_types']
    return JsonResponse(data)


@login_required
def part_export(request):
    """Экспорт радиодеталей в Excel (с учётом текущих фильтров списка)."""
    parts, _ = _filter_parts(request.GET)

    # Заголовки — имена полей, а не русские подписи: этот же файл принимает
    # импорт, и переименование колонок разорвало бы выгрузку и загрузку
    headers = [
        'part_number', 'name', 'component_type', 'package',
        'resistance', 'resistance_unit',
        'power', 'power_unit',
        'voltage', 'voltage_unit',
        'current', 'current_unit',
        'capacitance', 'capacitance_unit',
        'min_stock', 'current_stock', 'lead_time_days',
        'price', 'preferred_supplier', 'application', 'description',
    ]
    rows = [
        [getattr(part, field) for field in headers]
        for part in parts.order_by('part_number')
    ]
    wb = build_workbook('Радиодетали', headers, rows)
    return xlsx_response(wb, 'spare_parts.xlsx')


@login_required
def part_detail(request, pk):
    part = get_object_or_404(SparePart.objects.prefetch_related('movements'), pk=pk)
    movements = part.movements.order_by('-movement_date')[:50]
    # Все ячейки, кроме уже назначенной этой детали — деталь можно добавить
    # и в занятую другой деталью ячейку (в одной ячейке может быть несколько деталей)
    available_cells = StorageCell.objects.prefetch_related('parts').exclude(
        parts=part
    ).order_by('cabinet__number', 'row_number', 'cell_row')
    stock_form = StockMovementForm()
    return render(request, 'core/parts/detail.html', {
        'part': part,
        'movements': movements,
        'available_cells': available_cells,
        'stock_form': stock_form,
        'back_url': list_back_url(request, 'parts', 'part_list'),
    })


def _part_choices():
    """Уже встречавшиеся типы компонентов и корпуса — для подсказок в форме
    и отбора в списке. Оба поля заполняются руками, и без готового списка
    одно и то же обозначение расходится на «TO-220» и «то220»."""
    return {
        'component_types': list(
            SparePart.objects.exclude(component_type='')
            .values_list('component_type', flat=True).distinct().order_by('component_type')
        ),
        'packages': list(
            SparePart.objects.exclude(package='')
            .values_list('package', flat=True).distinct().order_by('package')
        ),
        'applications': list(
            SparePart.objects.exclude(application='')
            .values_list('application', flat=True).distinct().order_by('application')
        ),
    }


def _measurement_pairs(form):
    """Пары «значение — единица измерения» для вывода в одной строке формы."""
    return [
        (form[value], form[unit])
        for value, unit in (
            ('resistance', 'resistance_unit'),
            ('power', 'power_unit'),
            ('voltage', 'voltage_unit'),
            ('current', 'current_unit'),
            ('capacitance', 'capacitance_unit'),
        )
    ]


@login_required
def part_create(request):
    """Новая деталь, в том числе копия существующей.

    `?copy_from=<номер>` подставляет в форму карточку-образец: у соседних
    номиналов из одной серии совпадает почти всё, кроме одного значения,
    и набирать их заново — работа впустую. Артикул при этом не
    подставляется: он у каждой детали свой и уникален, а два одинаковых
    форма всё равно не примет. Остаток, ячейки и история движений
    к копии не относятся — копируется карточка, а не деталь на полке.
    """
    source = None
    copy_from = request.GET.get('copy_from', '')
    if str(copy_from).isdigit():
        source = SparePart.objects.filter(pk=copy_from).first()

    if request.method == 'POST':
        form = SparePartForm(request.POST)
        if form.is_valid():
            part = form.save()
            messages.success(request, f'Деталь {part.part_number} добавлена')
            return redirect('part_detail', pk=part.pk)
        messages.error(request, 'Деталь не сохранена: проверьте отмеченные поля')
    elif source is not None:
        initial = {
            name: getattr(source, name)
            for name in SparePartForm.Meta.fields if name != 'part_number'
        }
        form = SparePartForm(initial=initial)
    else:
        form = SparePartForm()
    # Отмена копии возвращает к образцу — там и стояли, когда решили
    # копировать; отмена новой детали с нуля — в список, другого места нет
    back_url = reverse('part_detail', args=[source.pk]) if source is not None else reverse('part_list')
    return render(request, 'core/parts/form.html', {
        'form': form,
        'title': 'Копия детали' if source is not None else 'Новая деталь',
        'copy_source': source,
        'measurement_pairs': _measurement_pairs(form),
        'back_url': back_url,
        **_part_choices(),
    })


@login_required
def part_edit(request, pk):
    part = get_object_or_404(SparePart, pk=pk)
    if request.method == 'POST':
        form = SparePartForm(request.POST, instance=part)
        if form.is_valid():
            form.save()
            messages.success(request, 'Деталь обновлена')
            return redirect('part_detail', pk=part.pk)
        messages.error(request, 'Изменения не сохранены: проверьте отмеченные поля')
    else:
        form = SparePartForm(instance=part)
    return render(request, 'core/parts/form.html', {
        'form': form,
        'title': 'Редактирование детали',
        'part': part,
        'measurement_pairs': _measurement_pairs(form),
        'back_url': reverse('part_detail', args=[part.pk]),
        **_part_choices(),
    })


@permission_required('parts_delete')
def part_bulk_delete(request):
    """Удаление отмеченных деталей — списком, а не по одной.

    После загрузки каталога из файла лишние позиции удаляют десятками,
    и открывать на каждую отдельную страницу подтверждения бессмысленно.

    Страница подтверждения показывает, что уйдёт вместе с деталями:
    удаление уносит за собой историю движений и записи о том, что деталь
    ставили в заказ. Восстановить это можно только из резервной копии,
    поэтому сказано об этом прямо, а не мелким шрифтом.
    """
    ids = request.POST.getlist('ids') if request.method == 'POST' else _selected_ids(request)
    parts = (
        SparePart.objects
        .filter(pk__in=[value for value in ids if str(value).isdigit()])
        .annotate(
            movement_count=Count('movements', distinct=True),
            order_count=Count('repairorderdetail', distinct=True),
        )
        .prefetch_related('storage_cells')
        .order_by('part_number')
    )

    parts = list(parts)

    if request.method == 'POST':
        if not parts:
            messages.warning(request, 'Удалять нечего: ни одна деталь не отмечена')
            return redirect('part_list')
        numbers = [part.part_number for part in parts]
        SparePart.objects.filter(pk__in=[part.pk for part in parts]).delete()
        messages.success(request, (
            f'Удалено деталей: {len(numbers)}'
            if len(numbers) > 1 else f'Деталь {numbers[0]} удалена'
        ))
        return redirect('part_list')

    return render(request, 'core/parts/bulk_delete.html', {
        'parts': parts,
        'in_stock': [part for part in parts if part.current_stock],
        'in_orders': [part for part in parts if part.order_count],
    })


@permission_required('parts_delete')
def part_delete(request, pk):
    part = get_object_or_404(SparePart, pk=pk)
    if request.method == 'POST':
        part.delete()
        messages.success(request, 'Деталь удалена')
        return redirect('part_list')
    return render(request, 'core/parts/delete.html', {'part': part})


@login_required
@require_POST
def part_stock_incoming(request, pk):
    """Приход детали на склад."""
    part = get_object_or_404(SparePart, pk=pk)
    form = StockMovementForm(request.POST)
    if form.is_valid():
        with transaction.atomic():
            movement = form.save(commit=False)
            movement.part = part
            movement.movement_type = 'incoming'
            movement.created_by = request.user
            part.current_stock += movement.quantity
            updated = ['current_stock']
            # Цена детали идёт следом за последней поставкой: вводить её
            # дважды — в приходе и в карточке — никто не будет, а расходится
            # она молча
            if movement.unit_price is not None:
                part.price = movement.unit_price
                updated.append('price')
            part.save(update_fields=updated)
            movement.save()
        message = f'Приход +{movement.quantity} {part.name} оформлен'
        if movement.unit_price is not None:
            message += f'. Цена детали обновлена: {movement.unit_price} ₽'
        messages.success(request, message)
    else:
        messages.error(request, 'Ошибка при оформлении прихода')
    return redirect('part_detail', pk=pk)


@login_required
@require_POST
def part_stock_outgoing(request, pk):
    """Ручное списание деталей (фактический расход / инвентаризация)."""
    part = get_object_or_404(SparePart, pk=pk)
    form = StockOutgoingForm(request.POST)
    if form.is_valid():
        qty = form.cleaned_data['quantity']
        reason = form.cleaned_data['reason']
        notes = form.cleaned_data.get('notes', '')
        doc_num = form.cleaned_data.get('document_number', '')

        if part.current_stock < qty:
            messages.warning(request,
                f'Внимание: списываем {qty} {part.name}, но на складе только {part.current_stock}!')

        with transaction.atomic():
            part.current_stock -= qty
            part.save(update_fields=['current_stock'])
            movement = StockMovement.objects.create(
                part=part,
                quantity=qty,
                movement_type='outgoing',
                document_number=doc_num,
                notes=f'{notes} (причина: {dict(form.fields["reason"].choices).get(reason)})',
                created_by=request.user
            )
            # Не привязано к заказу — OrderCost для ручного списания не
            # заводится (нет заказа, на который отнести затрату), только
            # распределение по партиям для будущей себестоимости
            StockAllocation.allocate(movement)
        messages.success(request, f'Списано {qty} {part.name} ({reason})')
    else:
        messages.error(request, 'Ошибка при оформлении списания')
    return redirect('part_detail', pk=pk)


@login_required
@require_POST
def part_assign_cell(request, pk):
    """Назначение детали на ячейку хранения (ячейка может содержать несколько деталей)."""
    part = get_object_or_404(SparePart, pk=pk)
    cell_id = request.POST.get('cell_id')
    if cell_id:
        cell = get_object_or_404(StorageCell, pk=cell_id)
        with transaction.atomic():
            # У детали может быть только одна ячейка — снимаем с предыдущей, если была
            for old_cell in part.storage_cells.exclude(pk=cell.pk):
                old_cell.parts.remove(part)
            cell.parts.add(part)
        messages.success(request, f'Деталь назначена на ячейку {cell.address}')
    return redirect('part_detail', pk=pk)


# Пары «значение — единица измерения» характеристик детали
SPEC_FIELDS = (
    ('voltage', 'voltage_unit'),
    ('current', 'current_unit'),
    ('power', 'power_unit'),
    ('resistance', 'resistance_unit'),
    ('capacitance', 'capacitance_unit'),
)


@login_required
def part_import(request):
    """Импорт радиодеталей из Excel."""
    if request.method == 'POST':
        form = PartImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            update_existing = form.cleaned_data['update_existing']
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

                required = ['part_number', 'name', 'component_type']
                missing = [h for h in required if h not in headers]
                if missing:
                    messages.error(request, f'Отсутствуют обязательные колонки: {", ".join(missing)}')
                    return redirect('part_import')

                created_count = 0
                updated_count = 0
                error_count = 0
                error_details = []

                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    part_number = str(data.get('part_number', '')).strip()
                    if not part_number:
                        continue

                    # --- Вспомогательная функция для парсинга чисел ---
                    def _parse_decimal(val):
                        """Парсит значение в Decimal или None."""
                        if val is None or val == '':
                            return None
                        try:
                            # Если это уже число (int/float из Excel)
                            if isinstance(val, (int, float)):
                                return val
                            # Если строка — чистим и парсим
                            s = str(val).strip().replace(',', '.')
                            # Убираем единицы измерения из строки если они есть
                            s = re.sub(r'[^0-9.\-]', '', s)
                            if s == '' or s == '.':
                                return None
                            return float(s)
                        except (ValueError, TypeError):
                            return None

                    def _parse_int(val, default):
                        """Парсит значение в int или возвращает default."""
                        if val is None or val == '':
                            return default
                        try:
                            if isinstance(val, (int, float)):
                                return int(val)
                            return int(str(val).strip())
                        except (ValueError, TypeError):
                            return default

                    def _get_str(val):
                        """Возвращает строку или пустую строку."""
                        if val is None:
                            return ''
                        return str(val).strip()

                    # --- Формируем defaults ---
                    defaults = {
                        'name': _get_str(data.get('name')) or part_number,
                        'component_type': _get_str(data.get('component_type')),
                        'package': _get_str(data.get('package')),
                        'voltage': _parse_decimal(data.get('voltage')),
                        'voltage_unit': _get_str(data.get('voltage_unit')),
                        'current': _parse_decimal(data.get('current')),
                        'current_unit': _get_str(data.get('current_unit')),
                        'power': _parse_decimal(data.get('power')),
                        'power_unit': _get_str(data.get('power_unit')),
                        'resistance': _parse_decimal(data.get('resistance')),
                        'resistance_unit': _get_str(data.get('resistance_unit')),
                        'capacitance': _parse_decimal(data.get('capacitance')),
                        'capacitance_unit': _get_str(data.get('capacitance_unit')),
                        'min_stock': _parse_int(data.get('min_stock'), 5),
                        'current_stock': _parse_int(data.get('current_stock'), 0),
                        'lead_time_days': _parse_int(data.get('lead_time_days'), 14),
                        'price': _parse_decimal(data.get('price')),
                        'application': _get_str(data.get('application'))[:100],
                        'description': _get_str(data.get('description')),
                    }

                    # Единица без значения — мусор: «Ом» у диода не значит
                    # ничего, а в форме и в выгрузке выглядит как заполненное
                    # поле. В присланных файлах единицы обычно проставлены
                    # во всех строках подряд, независимо от значений.
                    # Пустую единицу при заполненном значении, наоборот,
                    # не переносим: она затёрла бы уже введённую руками.
                    for value_field, unit_field in SPEC_FIELDS:
                        if defaults[value_field] is None:
                            defaults[unit_field] = ''
                        elif not defaults[unit_field]:
                            del defaults[unit_field]

                    try:
                        if update_existing:
                            part, created = SparePart.objects.update_or_create(
                                part_number=part_number,
                                defaults=defaults
                            )
                        else:
                            part, created = SparePart.objects.get_or_create(
                                part_number=part_number,
                                defaults=defaults
                            )
                            if not created:
                                continue
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                    except Exception as e:
                        error_count += 1
                        error_details.append(f'{part_number}: {str(e)[:100]}')
                        if error_count <= 5:
                            import traceback
                            traceback.print_exc()

                msg = f'Импорт завершён: создано {created_count}, обновлено {updated_count}, ошибок {error_count}'
                if error_details and error_count <= 10:
                    msg += ' | ' + '; '.join(error_details[:5])
                messages.success(request, msg)
                return redirect('part_list')
            except Exception as e:
                messages.error(request, f'Ошибка импорта: {e}')
    else:
        form = PartImportForm()
    return render(request, 'core/parts/import.html', {'form': form})

@login_required
def storage_cell_grid(request):
    """Визуальная сетка кассетниц. Одна ячейка может содержать несколько деталей.

    Ряды рисуются по фактическим ячейкам, а не по восьмёркам: в ряду из
    четырёх ячеек каждая занимает четверть ширины и выглядит крупной —
    ровно как крупные ящики внизу настоящего органайзера.
    """
    cabinets = list(Cabinet.objects.all())
    if not cabinets:
        # Пустой набор ячеек нужен даже здесь: страница вставляет его
        # в скрипт как есть, и без него получалось `const CELLS_DATA = ;` —
        # синтаксическая ошибка, гасившая весь скрипт страницы разом
        return render(request, 'core/storage_cells/grid.html', {
            'cabinets': [], 'cabinet': None, 'rows': [],
            'cells_data_json': '{}',
        })

    requested = request.GET.get('cabinet', '')
    cabinet = next(
        (item for item in cabinets if str(item.number) == requested), cabinets[0]
    )

    selected_part_id = request.GET.get('selected_part', '')
    move_from = request.GET.get('move_from', '')
    selected_part = None
    if selected_part_id:
        selected_part = get_object_or_404(SparePart, pk=selected_part_id)

    cells = cabinet.cells.prefetch_related('parts').order_by('row_number', 'cell_row')
    rows_map = {}
    cells_data = {}
    for cell in cells:
        cells_data[cell.pk] = {
            'address': cell.address,
            'parts': [
                {
                    'id': p.pk,
                    'part_number': p.part_number,
                    'name': p.name,
                    'component_type': p.component_type,
                    'package': p.package,
                    'stock': p.current_stock,
                    'min_stock': p.min_stock,
                }
                for p in cell.parts.all()
            ],
        }

        status = cell.get_status()
        cell_parts = cells_data[cell.pk]['parts']
        if selected_part and any(p['id'] == selected_part.pk for p in cell_parts):
            status = 'selected'
        rows_map.setdefault(cell.row_number, []).append({
            'cell': cell,
            'status': status,
            'part_count': len(cell_parts),
        })

    # Ширина ячейки в процентах, чтобы ряд из четырёх выглядел вчетверо
    # крупнее ряда из шестнадцати. Считаем здесь, а не в шаблоне: там
    # деления нет, а подгонять классами Bootstrap под любое число нельзя.
    #
    # Строкой, а не числом: локаль ru-ru печатает дробь через запятую,
    # и «calc(12,5% - 4px)» — невалидный CSS. Ячейки тогда остаются
    # без ширины, а на экране это выглядит как развалившаяся вёрстка
    rows = [
        {
            'number': number,
            'cells': rows_map[number],
            'width': f'{100 / len(rows_map[number]):.4f}',
        }
        for number in sorted(rows_map)
    ]

    # Каталог деталей на страницу больше не выгружается: выбор детали
    # спрашивает сервер сам. Прежде здесь в каждую сетку уезжали сотни
    # записей — ровно тот список, который человек и не мог пролистать
    return render(request, 'core/storage_cells/grid.html', {
        'rows': rows,
        'cabinet': cabinet,
        'cabinets': cabinets,
        'selected_part': selected_part,
        'move_from': move_from,
        'cells_data_json': json.dumps(cells_data, ensure_ascii=False),
    })


# ==================== КАССЕТНИЦЫ ====================

@permission_required('cabinets_manage')
def cabinet_list(request):
    """Список кассетниц с их раскладкой."""
    cabinets = Cabinet.objects.prefetch_related('cells')
    rows = [
        {
            'cabinet': cabinet,
            'layout': cabinet.layout_text,
            'cell_count': cabinet.cell_count,
            'occupied': sum(1 for cell in cabinet.cells.all() if cell.parts.exists()),
        }
        for cabinet in cabinets
    ]
    return render(request, 'core/storage_cells/cabinets.html', {'rows': rows})


@permission_required('cabinets_manage')
def cabinet_create(request):
    """Новая кассетница: номер, название и раскладка по рядам."""
    if request.method == 'POST':
        form = CabinetForm(request.POST)
        if form.is_valid():
            cabinet = form.save()
            added, _ = cabinet.apply_layout(form.cleaned_data['layout'])
            messages.success(
                request, f'Кассетница {cabinet.number} создана, ячеек: {added}')
            return redirect('cabinet_list')
        messages.error(request, 'Кассетница не создана: проверьте отмеченные поля')
    else:
        form = CabinetForm()
    return render(request, 'core/storage_cells/cabinet_form.html', {
        'form': form, 'title': 'Новая кассетница', 'back_url': reverse('cabinet_list'),
    })


@permission_required('cabinets_manage')
def cabinet_edit(request, pk):
    """Правка кассетницы, в том числе раскладки.

    Ячейки, оказавшиеся за пределами новой раскладки, удаляются — но
    только пустые: занятые форма не пропускает.

    `?next=` возвращает туда, откуда пришли, — обычно с сетки ячеек,
    когда у кассетницы ещё нет раскладки («задайте раскладку»). Без него
    — список кассетниц, как и раньше.
    """
    cabinet = get_object_or_404(Cabinet, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or ''
    if not url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        next_url = ''
    if request.method == 'POST':
        form = CabinetForm(request.POST, instance=cabinet)
        if form.is_valid():
            cabinet = form.save()
            added, removed = cabinet.apply_layout(form.cleaned_data['layout'])
            messages.success(request, (
                f'Кассетница {cabinet.number} сохранена. '
                f'Добавлено ячеек: {added}, удалено: {removed}'
            ))
            return redirect(next_url or 'cabinet_list')
        messages.error(request, 'Изменения не сохранены: проверьте отмеченные поля')
    else:
        form = CabinetForm(instance=cabinet)
    return render(request, 'core/storage_cells/cabinet_form.html', {
        'form': form, 'cabinet': cabinet,
        'title': f'Кассетница {cabinet.number}',
        'back_url': next_url or reverse('cabinet_list'),
        'next_url': next_url,
    })


@permission_required('cabinets_manage')
def cabinet_delete(request, pk):
    """Удаление кассетницы вместе с её ячейками.

    Занятые ячейки удалять не даём: сведения о том, где лежит деталь,
    восстановить будет неоткуда.
    """
    cabinet = get_object_or_404(Cabinet, pk=pk)
    occupied = [cell for cell in cabinet.cells.prefetch_related('parts') if cell.parts.exists()]

    if request.method == 'POST':
        if occupied:
            messages.error(request, (
                f'Кассетница {cabinet.number} не удалена: в её ячейках лежат '
                f'детали ({len(occupied)} шт.). Сначала переложите их.'
            ))
            return redirect('cabinet_list')
        number = cabinet.number
        cabinet.delete()
        messages.success(request, f'Кассетница {number} удалена')
        return redirect('cabinet_list')

    return render(request, 'core/storage_cells/cabinet_delete.html', {
        'cabinet': cabinet,
        'occupied': occupied,
    })


@login_required
@require_POST
def storage_cell_move(request):
    """Перемещение конкретной детали между ячейками (AJAX)."""
    from_cell_id = request.POST.get('from_cell')
    to_cell_id = request.POST.get('to_cell')
    part_id = request.POST.get('part_id')

    if not part_id:
        return JsonResponse({'error': 'Не указана деталь'}, status=400)

    try:
        with transaction.atomic():
            part = SparePart.objects.get(pk=part_id)

            if from_cell_id:
                from_cell = StorageCell.objects.select_for_update().get(pk=from_cell_id)
                from_cell.parts.remove(part)

            to_cell = None
            if to_cell_id:
                to_cell = StorageCell.objects.select_for_update().get(pk=to_cell_id)
                to_cell.parts.add(part)

            _send_stock_update(part)

            return JsonResponse({'success': True, 'message': 'Перемещение выполнено', 'address': to_cell.address if to_cell else None})
    except (SparePart.DoesNotExist, StorageCell.DoesNotExist):
        return JsonResponse({'error': 'Деталь или ячейка не найдена'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_POST
def storage_cell_move_all(request):
    """Перемещение ячейки целиком — со всем содержимым — в другую (AJAX).

    Тот же приём, что и у перемещения одной детали (`storage_cell_move`,
    выше): адресат может быть не пустым, тогда детали просто объединяются.
    Сама ячейка-источник никуда не девается — адрес и место в кассетнице
    у неё не меняются, меняется только то, что в ней лежит.
    """
    from_cell_id = request.POST.get('from_cell')
    to_cell_id = request.POST.get('to_cell')

    if not from_cell_id or not to_cell_id:
        return JsonResponse({'error': 'Не указана ячейка отправления или назначения'}, status=400)
    if from_cell_id == to_cell_id:
        return JsonResponse({'error': 'Ячейка отправления и назначения совпадают'}, status=400)

    try:
        with transaction.atomic():
            from_cell = StorageCell.objects.select_for_update().get(pk=from_cell_id)
            to_cell = StorageCell.objects.select_for_update().get(pk=to_cell_id)

            parts = list(from_cell.parts.all())
            if not parts:
                return JsonResponse({'error': 'В ячейке нет деталей'}, status=400)

            for part in parts:
                from_cell.parts.remove(part)
                to_cell.parts.add(part)
                _send_stock_update(part)

            return JsonResponse({
                'success': True,
                'message': f'Перемещено деталей: {len(parts)}',
                'address': to_cell.address,
                'count': len(parts),
            })
    except StorageCell.DoesNotExist:
        return JsonResponse({'error': 'Ячейка не найдена'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_POST
def storage_cell_add_part(request, pk):
    """Добавление детали в ячейку (одна из нескольких деталей, которые может хранить ячейка)."""
    cell = get_object_or_404(StorageCell, pk=pk)
    part_id = request.POST.get('part_id')
    part = get_object_or_404(SparePart, pk=part_id)

    with transaction.atomic():
        # У детали может быть только одна ячейка — снимаем с предыдущей, если была
        for old_cell in part.storage_cells.exclude(pk=cell.pk):
            old_cell.parts.remove(part)
        cell.parts.add(part)

    return JsonResponse({
        'success': True,
        'message': f'Деталь {part.part_number} добавлена в ячейку {cell.address}',
        'part': {
            'id': part.pk,
            'part_number': part.part_number,
            'name': part.name,
            'component_type': part.component_type,
            'stock': part.current_stock,
            'min_stock': part.min_stock,
        },
    })


@login_required
@require_POST
def storage_cell_remove_part(request, pk):
    """Удаление детали из ячейки (без удаления самой детали)."""
    cell = get_object_or_404(StorageCell, pk=pk)
    part_id = request.POST.get('part_id')
    part = get_object_or_404(SparePart, pk=part_id)
    cell.parts.remove(part)
    return JsonResponse({'success': True, 'message': f'Деталь {part.part_number} удалена из ячейки {cell.address}'})


# ==================== ИНВЕНТАРИЗАЦИЯ ====================

def _apply_inventory_discrepancy(line, discrepancy, employee):
    """Применяет расхождение по одной строке инвентаризации.

    `discrepancy` — фактически посчитанное минус ЖИВОЙ остаток детали на
    момент применения; вызывающая сторона (`inventory_confirm`) уже
    пересчитала его против `line.part.current_stock`, здесь это не
    делается повторно, чтобы несколько строк в одной сессии применялись
    против согласованного набора чисел, а не каждая против своего момента.

    Расхождение создаёт обычный `StockMovement` — не отдельную сущность
    корректировки — и для недостачи распределяется по партиям прихода
    (FIFO) точно так же, как любое другое списание в проекте (см.
    `_use_repair_order_part`, `part_stock_outgoing`,
    `StockAllocation.allocate`): решение задачи сознательно не даёт
    вручную выбирать партию под недостачу — причина расхождения
    (пересортица, утеря, ошибка ввода) ни с одной конкретной партией
    не связана сильнее прочих.

    Избыток заводится приходом без цены (находка — не покупка, угадывать
    её нельзя) и поэтому не должен обновлять `SparePart.price` — тот же
    аккуратный флажок, что уже есть в `part_stock_incoming`.

    После применения `part.current_stock` равен ровно
    `line.counted_quantity`, что бы ни случилось с остатком раньше.
    """
    part = line.part
    session = line.session
    where = line.cell.address if line.cell else part.part_number

    if discrepancy > 0:
        movement = StockMovement.objects.create(
            part=part,
            quantity=discrepancy,
            movement_type='incoming',
            unit_price=None,
            notes=f'{line.comment} (причина: Инвентаризация) — сессия №{session.pk}, {where}',
            created_by=employee,
        )
        part.current_stock += discrepancy
        part.save(update_fields=['current_stock'])
        # unit_price не заполнен — цену детали не трогаем, см. part_stock_incoming
    else:
        qty = -discrepancy
        movement = StockMovement.objects.create(
            part=part,
            quantity=qty,
            movement_type='outgoing',
            notes=f'Недостача при инвентаризации (причина: Инвентаризация) — сессия №{session.pk}, {where}',
            created_by=employee,
        )
        part.current_stock -= qty
        part.save(update_fields=['current_stock'])
        StockAllocation.allocate(movement)

    line.movement = movement
    line.save(update_fields=['movement', 'comment'])
    return movement


@login_required
@require_POST
def inventory_start(request, pk):
    """Начинает инвентаризацию кассетницы `pk`.

    Если по ней уже идёт незавершённая сессия — не открываем вторую
    параллельно, а ведём к уже идущей. Строки заводятся по деталям,
    у которых сейчас есть ячейка внутри этой кассетницы: у пустых ячеек
    сверять нечего.
    """
    cabinet = get_object_or_404(Cabinet, pk=pk)
    existing = InventorySession.objects.filter(
        cabinet=cabinet, status=InventorySession.STATUS_IN_PROGRESS
    ).first()
    if existing:
        messages.info(request, f'По кассетнице {cabinet.number} уже идёт инвентаризация — продолжаем её')
        return redirect('inventory_count', pk=existing.pk)

    with transaction.atomic():
        session = InventorySession.objects.create(cabinet=cabinet, started_by=request.user)
        cells = cabinet.cells.prefetch_related('parts').order_by('row_number', 'cell_row')
        new_lines = [
            InventorySessionLine(session=session, part=part, cell=cell, expected_quantity=part.current_stock)
            for cell in cells
            for part in cell.parts.all()
        ]
        InventorySessionLine.objects.bulk_create(new_lines)

    if not new_lines:
        messages.warning(
            request, f'В кассетнице {cabinet.number} нет ни одной размещённой детали — сверять нечего')
    else:
        messages.success(
            request,
            f'Инвентаризация кассетницы {cabinet.number} начата, деталей к пересчёту: {len(new_lines)}'
        )
    return redirect('inventory_count', pk=session.pk)


@login_required
def inventory_count(request, pk):
    """Ввод фактически посчитанного количества по каждой строке сессии.

    Учётное количество (снимок на начало сессии) показывается для
    справки; сохранение не пересчитывает и не применяет ничего — это
    делает следующий шаг, `inventory_confirm`.
    """
    session = get_object_or_404(
        InventorySession.objects.select_related('cabinet'), pk=pk)
    if not session.is_in_progress:
        return redirect('inventory_detail', pk=session.pk)

    lines = session.lines.select_related('part', 'cell').order_by(
        'cell__row_number', 'cell__cell_row', 'part__part_number')

    if request.method == 'POST':
        with transaction.atomic():
            for line in lines:
                raw = request.POST.get(f'counted_{line.pk}', '').strip()
                if raw == '':
                    continue
                try:
                    value = int(raw)
                except ValueError:
                    continue
                if value < 0:
                    continue
                line.counted_quantity = value
                line.save(update_fields=['counted_quantity'])
        messages.success(request, 'Количества сохранены')
        return redirect('inventory_confirm', pk=session.pk)

    return render(request, 'core/inventory/count.html', {
        'session': session,
        'lines': lines,
    })


@login_required
def inventory_confirm(request, pk):
    """Показывает расхождения, посчитанные против ЖИВОГО остатка (не
    против учтённого на начало сессии — см. docstring
    `InventorySessionLine`), и по подтверждению применяет их.

    Избыток без заполненного комментария блокирует применение целиком —
    ни одно движение не создаётся, пока причина не указана по каждой
    строке с превышением.
    """
    session = get_object_or_404(
        InventorySession.objects.select_related('cabinet'), pk=pk)
    if not session.is_in_progress:
        return redirect('inventory_detail', pk=session.pk)

    counted_lines = list(
        session.lines.select_related('part', 'cell')
        .filter(counted_quantity__isnull=False)
        .order_by('cell__row_number', 'cell__cell_row', 'part__part_number')
    )
    uncounted = session.lines.filter(counted_quantity__isnull=True).count()

    rows = [
        {
            'line': line,
            'live_stock': line.part.current_stock,
            'discrepancy': line.counted_quantity - line.part.current_stock,
            'drifted': line.part.current_stock != line.expected_quantity,
        }
        for line in counted_lines
    ]

    if request.method == 'POST':
        comment_values = {
            row['line'].pk: request.POST.get(f"comment_{row['line'].pk}", '').strip()
            for row in rows if row['discrepancy'] > 0
        }
        missing_comment_ids = {pk_ for pk_, text in comment_values.items() if not text}

        if missing_comment_ids:
            for row in rows:
                row['comment_value'] = comment_values.get(row['line'].pk, '')
            messages.error(
                request,
                'Укажите причину избытка по каждой отмеченной строке — без неё заявка не применена'
            )
            return render(request, 'core/inventory/confirm.html', {
                'session': session, 'rows': rows, 'uncounted': uncounted,
                'missing_comment_ids': missing_comment_ids,
            })

        to_apply = [(row['line'], row['discrepancy']) for row in rows if row['discrepancy'] != 0]
        for line, _ in to_apply:
            if line.pk in comment_values:
                line.comment = comment_values[line.pk]

        with transaction.atomic():
            for line, discrepancy in to_apply:
                _apply_inventory_discrepancy(line, discrepancy, request.user)
            session.status = InventorySession.STATUS_COMPLETED
            session.completed_at = timezone.now()
            session.completed_by = request.user
            session.save(update_fields=['status', 'completed_at', 'completed_by'])

        if to_apply:
            messages.success(request, f'Инвентаризация завершена. Применено расхождений: {len(to_apply)}')
        else:
            messages.success(request, 'Инвентаризация завершена. Расхождений не найдено')
        return redirect('inventory_detail', pk=session.pk)

    return render(request, 'core/inventory/confirm.html', {
        'session': session, 'rows': rows, 'uncounted': uncounted,
    })


@login_required
def inventory_list(request):
    """История сессий инвентаризации по всем кассетницам — для аудита."""
    sessions = (
        InventorySession.objects
        .select_related('cabinet', 'started_by', 'completed_by')
        .annotate(
            surplus_count=Count(
                'lines', filter=Q(lines__movement__movement_type='incoming'), distinct=True),
            deficit_count=Count(
                'lines', filter=Q(lines__movement__movement_type='outgoing'), distinct=True),
        )
    )
    return render(request, 'core/inventory/list.html', {
        'sessions': sessions,
        'cabinets': Cabinet.objects.all(),
    })


@login_required
def inventory_detail(request, pk):
    """Одна сессия инвентаризации целиком — для истории/аудита, и как
    экран продолжения незавершённой сессии."""
    session = get_object_or_404(
        InventorySession.objects.select_related('cabinet', 'started_by', 'completed_by'), pk=pk)
    lines = session.lines.select_related('part', 'cell', 'movement').order_by(
        'cell__row_number', 'cell__cell_row', 'part__part_number')
    return render(request, 'core/inventory/detail.html', {
        'session': session,
        'lines': lines,
    })


@login_required
def inventory_delete(request, pk):
    """Удаление черновика инвентаризации: только пока сессия в процессе
    и по ней ещё ничего не применено — иначе это уже аудиторский след."""
    session = get_object_or_404(InventorySession.objects.select_related('cabinet'), pk=pk)
    if not session.can_be_deleted:
        messages.error(request, 'Эту сессию удалить нельзя: по ней уже применены движения склада')
        return redirect('inventory_detail', pk=session.pk)

    if request.method == 'POST':
        cabinet_number = session.cabinet.number
        session.delete()
        messages.success(request, f'Черновик инвентаризации кассетницы {cabinet_number} удалён')
        return redirect('inventory_list')

    return render(request, 'core/inventory/delete.html', {'session': session})


def _label_specs(part):
    """Строка характеристик для этикетки: тип и номиналы через точку."""
    return ' · '.join(value for value in (part.component_type, part.specs_display) if value)


def _grouped_specs(parts):
    """Что у набора общее и чем детали в нём различаются.

    Повторять «0.125Вт» у каждого номинала незачем — в 26 мм ширины из-за
    этого не помещаются сами номиналы. Общее выносим в строку рядом
    с заголовком, под заголовком оставляем только различия — списком,
    по элементу на деталь.
    """
    values = [[value for value in part.specs_display.split(', ') if value] for part in parts]
    common = [value for value in values[0] if all(value in row for row in values[1:])]
    distinct = [
        ', '.join(value for value in row if value not in common) or part.part_number
        for part, row in zip(parts, values)
    ]
    # Если после вычитания общего номиналы совпадают, различить детали по ним
    # нельзя: так вышло с россыпью резисторов, где заполнена только мощность
    # («2Вт, 2Вт, 2Вт»), а номинал живёт в артикуле. Тогда перечисляем артикулы
    if len(set(distinct)) < len(distinct):
        distinct = [part.part_number for part in parts]
    return ', '.join(common), distinct


def _cell_label_hidden_top():
    """Сколько миллиметров этикетки ячейки закрыто выступом.

    Отрицательное и заведомо бессмысленное отбрасывается: значение правит
    человек на странице настроек, а этикетка с закрытой половиной — это
    уже не этикетка.
    """
    try:
        value = int(envfile.setting('LABEL_CELL_HIDDEN_TOP_MM', 0))
    except (TypeError, ValueError):
        return 0
    return min(max(value, 0), 12)


def _cell_label(cell, base_url):
    """Данные одной этикетки ячейки.

    Поля те же, что и у этикетки детали, и печатаются они тем же шаблоном:
    на пакете и на ячейке одна и та же деталь должна выглядеть одинаково,
    иначе кладовщик каждый раз ищет глазами, где что написано.

    Если все детали в ячейке одного типа (набор номиналов одного типа —
    резисторы, конденсаторы и т.д.), сверху пишется «Набор резисторов»,
    рядом — то, что у них общее, а ниже столбцами номиналы. Разнотипную
    ячейку подписываем перечнем артикулов: общего имени у неё нет.

    Перечисление возвращается списком, а не строкой: на этикетке он рисуется
    сеткой в два столбца. Сплошная строка переносилась посреди номинала —
    «10кОм, 4.7к» и «Ом» на следующей строке, — и понять, где кончается
    одна деталь и начинается другая, было нельзя.
    """
    parts = list(cell.parts.all())
    component_types = {p.component_type for p in parts if p.component_type}
    packages = {p.package for p in parts if p.package}
    applications = {p.application for p in parts if p.application}
    grouped = len(parts) > 1 and len(component_types) == 1
    items = []
    # Применимость на ячейку — только если она у всех одна: «Otis» на ячейке,
    # где половина деталей от ABB, вводит в заблуждение
    application = applications.pop() if len(applications) == 1 else ''

    if not parts:
        title, specs, description, package = '', '', 'Ячейка пуста', ''
    elif len(parts) == 1:
        part = parts[0]
        title, specs = part.part_number, _label_specs(part)
        description, package = part.label_text, part.package
    elif grouped:
        # В нижний регистр — только первое слово: «Диод Шоттки» просклоняется
        # в «Диодов Шоттки», а .lower() всей строки стёр бы заглавную у
        # «Шоттки» — она у эпонима не двигается вслед за типом.
        plural = plural_genitive(component_types.copy().pop())
        head, sep, rest = plural.partition(' ')
        title = f'Набор {head.lower()}{sep}{rest}'
        # Номинал — то, чем детали набора различаются; если его не заполнили,
        # различить их можно только по артикулу
        specs, items = _grouped_specs(parts)
        description = ''
        # Корпус общий на всю ячейку — только если он у всех один: «0805»
        # на ячейке, где половина деталей в 1206, хуже, чем ничего
        package = packages.pop() if len(packages) == 1 else ''
    else:
        title = 'Разные детали'
        specs = ''
        description = ''
        items = [p.part_number for p in parts]
        package = packages.pop() if len(packages) == 1 else ''

    payload = qr_payload('c', cell.pk)
    link = qr_link(base_url, 'c', cell.pk)
    return {
        'cell': cell,
        'cell_parts': parts,
        'title': title,
        'specs': specs,
        'description': description,
        'items': items,
        'package': package,
        'application': application,
        'address': cell.address,
        # Ссылка вместо простого адреса: сканирование сразу открывает
        # содержимое, а не показывает строку, которую потом ищут вручную
        'qr_url': link,
        'qr_payload': payload,
        'qr_img': generate_qr_image(payload),
        # Сколько сверху закрыто выступом кассетницы. Только у ячейки:
        # на пакете с деталью выступа нет, и пустая полоса пропала бы зря
        'hidden_top': _cell_label_hidden_top(),
    }


@login_required
def storage_cell_label(request, pk):
    """Печать этикетки одной ячейки."""
    cell = get_object_or_404(
        StorageCell.objects.select_related('cabinet').prefetch_related('parts'), pk=pk)
    base_url = label_base_url(request)
    context = _cell_label(cell, base_url)
    context['qr_base'] = base_url
    context['qr_warning'] = qr_length_warning([context['qr_payload']])
    return render(request, 'core/storage_cells/label.html', context)


def _order_equipment_label(order, roe, position, base_url):
    """Данные одной этикетки оборудования в заказе.

    Общее для одиночной печати и для пачки (`repair_order_labels_batch`) —
    те же поля, тот же QR, чтобы одна и та же единица оборудования не могла
    напечататься по-разному в зависимости от того, откуда её печатали.

    Ссылка на заказ вместо текста «LT-2026-08-001/1». Текст читался
    человеком, но сканирование им ничего не давало: заказ всё равно искали
    руками. Плата за ссылку — код вырастает с 21 до 25–29 модулей
    (сколько именно, зависит от длины LABEL_BASE_URL), поэтому место под
    него освобождено: эмблема с этикетки убрана, код печатается сам по себе.
    """
    payload = qr_payload('u', roe.pk)
    link = qr_link(base_url, 'u', roe.pk)
    return {
        'order': order,
        'roe': roe,
        'position': position,
        'qr_payload': payload,
        'qr_img': generate_qr_image(payload),
        'qr_url': link,
    }


@login_required
def repair_order_equipment_label(request, order_pk, roe_pk):
    """Печать этикетки оборудования в контексте заказа (43x25 мм).
    Номер позиции — порядковый номер оборудования в списке заказа."""
    order = get_object_or_404(RepairOrder, pk=order_pk)
    roe_list = list(
        order.order_equipments.select_related('equipment__model').order_by('id')
    )
    roe = get_object_or_404(RepairOrderEquipment, pk=roe_pk, repair_order=order)
    position = next(i for i, r in enumerate(roe_list, start=1) if r.pk == roe.pk)

    base_url = label_base_url(request)
    context = _order_equipment_label(order, roe, position, base_url)
    context['qr_base'] = base_url
    context['qr_warning'] = qr_length_warning([context['qr_payload']])
    return render(request, 'core/repair_orders/equipment_label.html', context)


@login_required
def repair_order_labels_batch(request):
    """Пачка этикеток оборудования: по отмеченным заказам или по всему
    текущему отбору списка. Этикетка печатается не на заказ, а на единицу
    оборудования внутри него — заказ с тремя единицами даёт три этикетки."""
    ids = _selected_ids(request)
    if ids:
        orders = RepairOrder.objects.filter(pk__in=ids)
    else:
        orders, _ = _filter_orders(request)

    # prefetch_related(None) снимает набор из _filter_orders (он тянет
    # оборудование только для поиска по серийнику) — иначе Django ругается
    # на два разных queryset для одного и того же related-пути
    orders = list(
        orders.prefetch_related(None).prefetch_related(
            Prefetch(
                'order_equipments',
                queryset=RepairOrderEquipment.objects.select_related('equipment__model').order_by('id'),
            )
        ).order_by('-date_received')
    )

    base_url = label_base_url(request)
    labels = []
    for order in orders:
        for position, roe in enumerate(order.order_equipments.all(), start=1):
            if len(labels) >= MAX_LABELS_PER_BATCH:
                break
            labels.append(_order_equipment_label(order, roe, position, base_url))
        if len(labels) >= MAX_LABELS_PER_BATCH:
            break

    return render(request, 'core/repair_orders/labels_batch.html', {
        'labels': labels,
        'layout': _batch_layout(request),
        'limit': MAX_LABELS_PER_BATCH,
        'qr_base': base_url,
        'qr_warning': qr_length_warning(label['qr_payload'] for label in labels),
    })


@login_required
def repair_order_equipment_labels(request, pk):
    """Этикетки на оборудование одного заказа: на отмеченные единицы либо,
    если ничего не отмечено, на все.

    Отдельно от `repair_order_labels_batch`: та печатает по отмеченным
    заказам целиком из списка, а отсюда, из карточки заказа, выбирают
    единицы внутри одного заказа — печатать заново весь заказ ради одной
    переклеенной наклейки не нужно.

    Номер позиции берётся по месту единицы в полном списке заказа, а не
    по месту в отобранном: наклейка на вторую единицу должна остаться
    «/2» и тогда, когда печатают её одну.
    """
    order = get_object_or_404(RepairOrder, pk=pk)
    roe_list = list(
        order.order_equipments.select_related('equipment__model').order_by('id')
    )

    selected = {value for value in request.GET.getlist('roe') if value.isdigit()}
    base_url = label_base_url(request)
    labels = [
        _order_equipment_label(order, roe, position, base_url)
        for position, roe in enumerate(roe_list, start=1)
        if not selected or str(roe.pk) in selected
    ][:MAX_LABELS_PER_BATCH]

    return render(request, 'core/repair_orders/labels_batch.html', {
        'labels': labels,
        'layout': _batch_layout(request),
        'limit': MAX_LABELS_PER_BATCH,
        'qr_base': base_url,
        'qr_warning': qr_length_warning(label['qr_payload'] for label in labels),
        'order': order,
        'back_url': reverse('repair_order_detail', args=[order.pk]),
    })


# ==================== ОТЧЁТЫ ====================

@login_required
def reports(request):
    return render(request, 'core/reports/index.html')


def _purchase_plan_parts():
    """Детали ниже минимального остатка — общее для отчёта и его выгрузки."""
    return (
        SparePart.objects
        .below_minimum()
        .prefetch_related('storage_cells')
        .order_by('part_number')
    )


@login_required
def report_purchase_plan(request):
    """План закупок — детали ниже минимального остатка."""
    parts = list(_purchase_plan_parts())
    return render(request, 'core/reports/purchase_plan.html', {
        'parts': parts,
        **_purchase_plan_totals(parts),
    })


def _purchase_plan_totals(parts):
    """Во что обойдётся закупка и по скольким деталям цена неизвестна.

    Считаем в Python, а не запросом: список короткий, а цена может быть
    пустой, и складывать её с нулём нельзя — «ноль» и «неизвестно» это
    разные вещи, и вторую надо назвать вслух.
    """
    priced = [part for part in parts if part.price is not None]
    return {
        'plan_total': sum(part.purchase_cost for part in priced),
        'without_price': len(parts) - len(priced),
    }


@login_required
def report_purchase_plan_export(request):
    """План закупок в Excel — файл отправляют поставщику."""
    headers = [
        'Артикул', 'Название', 'Тип', 'Характеристики',
        'Остаток', 'Мин. остаток', 'Не хватает',
        'Цена, ₽', 'Сумма, ₽',
        'Срок поставки, дней', 'Поставщик', 'Ячейка',
    ]
    parts = list(_purchase_plan_parts())
    rows = [
        [
            part.part_number, part.name, part.component_type, part.specs_display,
            part.current_stock, part.min_stock, part.stock_deficit,
            part.price, part.purchase_cost,
            part.lead_time_days, part.preferred_supplier,
            part.current_cell.address if part.current_cell else '',
        ]
        for part in parts
    ]
    # Итог в самой книге: файл уходит поставщику и начальству, и сумму
    # из него достают в первую очередь
    if rows:
        totals = _purchase_plan_totals(parts)
        rows.append(['Итого'] + [''] * 7 + [totals['plan_total']] + [''] * 3)
    wb = build_workbook('План закупок', headers, rows)
    return xlsx_response(wb, f'План закупок {timezone.localdate():%Y-%m-%d}.xlsx')


def _filter_movements(request):
    """Фильтрация движений склада (общая для журнала и его выгрузки)."""
    part_id = request.GET.get('part', '')
    movement_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    movements = StockMovement.objects.select_related('part', 'repair_order', 'created_by')

    # Значения приходят из адресной строки, и её правят руками: нечисловой
    # номер детали или дата вроде «вчера» роняли бы страницу с ошибкой базы,
    # поэтому неразобранное значение просто не применяется как фильтр.
    if part_id.isdigit():
        movements = movements.filter(part_id=int(part_id))
    if movement_type in dict(StockMovement.MOVEMENT_TYPE_CHOICES):
        movements = movements.filter(movement_type=movement_type)
    if parse_date(date_from):
        movements = movements.filter(movement_date__date__gte=date_from)
    if parse_date(date_to):
        movements = movements.filter(movement_date__date__lte=date_to)

    return movements.order_by('-movement_date'), {
        'part_id': part_id,
        'movement_type': movement_type,
        'date_from': date_from,
        'date_to': date_to,
    }


@login_required
def report_stock_movements(request):
    """Журнал движений запчастей."""
    movements, filter_context = _filter_movements(request)

    paginator = Paginator(movements, 50)
    page = request.GET.get('page')

    # Отбор по детали — общий выбор детали, а не список всего каталога:
    # подписать выбранное нужно только её одной
    part_id = filter_context['part_id']
    selected_part = SparePart.objects.filter(pk=part_id).first() if part_id.isdigit() else None
    return render(request, 'core/reports/stock_movements.html', {
        'movements': paginator.get_page(page),
        'selected_part': selected_part,
        'movement_types': StockMovement.MOVEMENT_TYPE_CHOICES,
        **filter_context,
    })


@login_required
def report_stock_movements_export(request):
    """Журнал движений в Excel — с учётом фильтров, выставленных на странице."""
    movements, _ = _filter_movements(request)

    headers = [
        'Дата', 'Артикул', 'Название', 'Тип', 'Количество',
        'Цена, ₽', 'Сумма, ₽',
        'Документ', 'Заказ', 'Сотрудник', 'Примечания',
    ]
    rows = [
        [
            excel_datetime(m.movement_date),
            m.part.part_number, m.part.name,
            m.get_movement_type_display(), m.quantity,
            m.unit_price, m.total_price,
            m.document_number,
            m.repair_order.order_number if m.repair_order else '',
            m.created_by.full_name if m.created_by else '',
            m.notes,
        ]
        for m in movements
    ]
    wb = build_workbook('Движения', headers, rows)
    return xlsx_response(wb, f'Журнал движений {timezone.localdate():%Y-%m-%d}.xlsx')


def _debtor_orders():
    """Неоплаченные и частично оплаченные заказы — общее для отчёта и выгрузки.

    Условие «должник» живёт в `RepairOrderQuerySet`: то же самое нужно
    напоминаниям, и расходиться эти определения не должны.
    """
    return (
        RepairOrder.objects.with_debt()
        .select_related('client')
        .prefetch_related('order_equipments__equipment__model')
        .order_by('-date_received')
    )


def _total_debt(orders):
    """Сумма долга одним запросом.

    Складываем не стоимости, а остатки: часть денег заказчик мог уже внести,
    и итог из полных стоимостей завышал бы задолженность.

    Стоимости и оплаты берутся подзапросами, а не двумя Sum по соединениям:
    соединение размножило бы строки, и заказ с тремя платежами попал бы
    в итог трижды.
    """
    totals = orders.aggregate(
        cost=Sum(_order_cost_subquery()),
        paid=Sum(_order_paid_subquery()),
    )
    return (totals['cost'] or 0) - (totals['paid'] or 0)


def _order_cost_subquery():
    return Coalesce(
        Subquery(
            RepairOrderEquipment.objects
            .filter(repair_order=OuterRef('pk'))
            .values('repair_order').annotate(total=Sum('repair_cost')).values('total')[:1]
        ),
        Value(Decimal('0')),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )


def _order_paid_subquery():
    return Coalesce(
        Subquery(
            Payment.objects
            .filter(repair_order=OuterRef('pk'))
            .values('repair_order').annotate(total=Sum('amount')).values('total')[:1]
        ),
        Value(Decimal('0')),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )


@login_required
def report_debtors(request):
    """Задолженности по заказам."""
    orders = _debtor_orders()
    total_debt = _total_debt(orders)

    # Когда заказчику в последний раз напоминали. Без этой колонки на вопрос
    # «мы им вообще писали?» отвечать нечем, кроме как листать очередь
    orders = orders.annotate(
        last_reminder=Max(
            'notifications__created_at',
            filter=Q(notifications__event='debt_reminder'),
        )
    )

    return render(request, 'core/reports/debtors.html', {
        'orders': orders,
        'total_debt': total_debt,
        'overdue_days': envfile.setting('DEBT_OVERDUE_DAYS', 14),
    })


@login_required
def report_debtors_export(request):
    """Задолженности в Excel — для сверки с бухгалтерией."""
    orders = _debtor_orders()

    headers = [
        '№ заказа', 'Дата приёма', 'Заказчик', 'ИНН', 'Оборудование',
        'Статус ремонта', 'Статус оплаты', '№ счёта', 'Дата счёта',
        'Сумма, ₽', 'Оплачено, ₽', 'Остаток, ₽',
    ]
    rows = [
        [
            order.order_number,
            excel_datetime(order.date_received),
            order.client.name,
            order.client.inn,
            ', '.join(str(roe.equipment) for roe in order.order_equipments.all()),
            order.get_status_display(),
            order.get_payment_status_display(),
            order.invoice_number,
            order.invoice_date,
            order.total_repair_cost,
            order.paid_amount,
            order.debt,
        ]
        for order in orders
    ]
    # Итог в самой книге: иначе бухгалтеру пришлось бы складывать столбец
    # вручную, а именно эта цифра из отчёта и нужна
    if rows:
        rows.append(['Итого'] + [''] * (len(headers) - 2) + [_total_debt(orders)])

    wb = build_workbook('Задолженности', headers, rows)
    return xlsx_response(wb, f'Задолженности {timezone.localdate():%Y-%m-%d}.xlsx')


# ==================== Аналитика по срокам и загрузке ремонта ====================
#
# У заказа нет поля «ответственный инженер» — есть только автор смены
# статуса (`OrderStatusHistory.changed_by`). Везде ниже «инженер» — это
# суррогат по этому автору, а не отдельно назначенный человек. Если
# в модель когда-нибудь добавят настоящее поле ответственного, всю эту
# группу функций нужно будет заменить на использование поля напрямую,
# а не достраивать поверх суррогата.

TERMINAL_STATUSES = ('shipped', 'unrepairable', 'partially_repaired')


def _repair_analytics_period(request):
    """Диапазон дат отчёта — по умолчанию последние 30 дней."""
    default_to = timezone.localdate()
    default_from = default_to - timedelta(days=30)
    date_from = parse_date(request.GET.get('date_from', '')) or default_from
    date_to = parse_date(request.GET.get('date_to', '')) or default_to
    return date_from, date_to


def _completed_orders_in_period(date_from, date_to):
    """Заказы, завершённые (отгружены или признаны неремонтопригодными)
    в периоде — {id заказа: запись истории о завершении}.

    Момент завершения берём из истории статусов, а не из полей заказа
    (`date_completed` / `shipping_date`): они не всегда заполнены
    единообразно для обоих терминальных статусов, а история переходов —
    общий надёжный источник для обоих случаев. Если заказ побывал
    в терминальном статусе несколько раз, берётся самая свежая запись.
    """
    start = timezone.make_aware(datetime.combine(date_from, time.min))
    end = timezone.make_aware(datetime.combine(date_to, time.max))
    entries = (
        OrderStatusHistory.objects
        .filter(status__in=TERMINAL_STATUSES, changed_at__gte=start, changed_at__lte=end)
        .order_by('order_id', '-changed_at')
    )
    latest = {}
    for entry in entries:
        latest.setdefault(entry.order_id, entry)
    return latest


def _order_assignees(order_ids):
    """Сотрудник-суррогат «ответственного инженера» для каждого заказа
    из `order_ids` — {id заказа: id сотрудника}, без записи для заказов
    без суррогата.

    Правило (см. заголовок раздела): тот, кто последним перевёл заказ
    в статус «Ремонт»; если такого перехода не было — тот, кто последним
    перевёл в «Диагностика» (ремонт признан невозможным без попытки
    чинить). Если нет и такого перехода — у заказа нет персонального
    инженера в этой статистике, но он участвует в общих средних без
    разбивки по человеку.
    """
    if not order_ids:
        return {}
    entries = (
        OrderStatusHistory.objects
        .filter(order_id__in=order_ids, status__in=('repair', 'diagnostic'),
                changed_by__isnull=False)
        .order_by('changed_at')
        .values('order_id', 'status', 'changed_by_id')
    )
    repair_by_order = {}
    diagnostic_by_order = {}
    for entry in entries:
        # Записи идут по возрастанию времени, так что более позднее
        # присвоение в словарь всегда затирает более раннее — в итоге
        # остаётся именно последний переход в каждый из двух статусов
        target = repair_by_order if entry['status'] == 'repair' else diagnostic_by_order
        target[entry['order_id']] = entry['changed_by_id']
    return {
        order_id: repair_by_order[order_id] if order_id in repair_by_order
        else diagnostic_by_order[order_id]
        for order_id in order_ids
        if order_id in repair_by_order or order_id in diagnostic_by_order
    }


def _order_fault_types(order_ids):
    """{id заказа: {id типовой неисправности: название}} — только для
    единиц оборудования, где неисправность выбрана из справочника
    (Этап 3, `RepairOrderEquipment.faults`). Заказ, где указан только
    свободный текст («Другое»), в разбивку по типу не попадает."""
    if not order_ids:
        return {}
    result = defaultdict(dict)
    roe_qs = (
        RepairOrderEquipment.objects
        .filter(repair_order_id__in=order_ids)
        .prefetch_related('faults')
    )
    for roe in roe_qs:
        for fault in roe.faults.all():
            result[roe.repair_order_id][fault.id] = str(fault)
    return result


def _repair_analytics_data(date_from, date_to, viewer=None):
    """Показатели отчёта за период.

    `viewer=None` — полный отчёт по всем сотрудникам (только для роли
    `admin`). `viewer=<Employee>` — все показатели ограничены заказами,
    где этот сотрудник — инженер-суррогат (см. `_order_assignees`);
    у остальных ролей своих коллег в отчёте не видно вовсе.
    """
    completions = _completed_orders_in_period(date_from, date_to)
    order_ids = list(completions.keys())
    orders = {
        o.pk: o for o in
        RepairOrder.objects.filter(pk__in=order_ids).only('pk', 'date_received')
    }
    assignees = _order_assignees(order_ids)

    if viewer is not None:
        order_ids = [oid for oid in order_ids if assignees.get(oid) == viewer.pk]

    fault_types = _order_fault_types(order_ids)

    durations = []
    by_employee = defaultdict(list)
    by_fault = defaultdict(list)
    fault_names = {}

    for order_id in order_ids:
        order = orders.get(order_id)
        if order is None or order.date_received is None:
            continue
        completion = completions[order_id]
        duration = (completion.changed_at - order.date_received).total_seconds() / 86400
        durations.append(duration)

        assignee_id = assignees.get(order_id)
        if assignee_id:
            by_employee[assignee_id].append(duration)

        for fault_id, name in fault_types.get(order_id, {}).items():
            by_fault[fault_id].append(duration)
            fault_names[fault_id] = name

    def _avg(values):
        return round(sum(values) / len(values), 1) if values else None

    employees = {e.pk: e for e in Employee.objects.filter(pk__in=by_employee.keys())}
    by_employee_rows = sorted(
        (
            {'employee': employees[emp_id], 'orders': len(vals), 'avg_days': _avg(vals)}
            for emp_id, vals in by_employee.items() if emp_id in employees
        ),
        key=lambda row: row['employee'].full_name,
    )
    by_fault_rows = sorted(
        (
            {'fault_type': fault_names[fault_id], 'orders': len(vals), 'avg_days': _avg(vals)}
            for fault_id, vals in by_fault.items()
        ),
        key=lambda row: row['fault_type'],
    )

    return {
        'total_orders': len(durations),
        'avg_days': _avg(durations),
        'by_employee': by_employee_rows,
        'by_fault_type': by_fault_rows,
    }


def _current_load():
    """Текущая загрузка — {id сотрудника: число открытых заказов сейчас}.

    «Открытый» — статус из `RepairOrder.OPEN_STATUSES`. Заказ считается
    закреплённым за сотрудником, который последним оставил запись в его
    истории статусов, независимо от того, что именно эта запись меняла
    (статус ремонта, статус оплаты или ничего из этого): кто последним
    трогал заказ, тот сейчас его и ведёт. Тот же суррогат «ответственного
    инженера», что и у остальных функций этого раздела, только по
    другому правилу — здесь неважно, в какой статус был переход.
    """
    open_ids = list(RepairOrder.objects.open().values_list('pk', flat=True))
    if not open_ids:
        return Counter()
    entries = (
        OrderStatusHistory.objects
        .filter(order_id__in=open_ids)
        .order_by('order_id', '-changed_at')
        .values('order_id', 'changed_by_id')
    )
    latest_author = {}
    for entry in entries:
        latest_author.setdefault(entry['order_id'], entry['changed_by_id'])
    counts = Counter(latest_author.values())
    counts.pop(None, None)
    return counts


@login_required
def report_repair_analytics(request):
    """Аналитика по срокам ремонта и загрузке инженеров.

    Роль `admin` видит показатели по всем сотрудникам; любая другая роль —
    только свои собственные (не отчёт целиком и не показатели коллег).
    «Инженер» здесь — суррогат, см. докстринг раздела выше.
    """
    date_from, date_to = _repair_analytics_period(request)
    everyone = request.user.allows('reports_all_engineers')
    viewer = None if everyone else request.user

    data = _repair_analytics_data(date_from, date_to, viewer=viewer)

    load_counts = _current_load()
    if everyone:
        load_rows = sorted(
            (
                {'employee': e, 'count': load_counts.get(e.pk, 0)}
                for e in Employee.objects.filter(is_active=True)
            ),
            key=lambda row: (-row['count'], row['employee'].full_name),
        )
    else:
        load_rows = [{'employee': request.user, 'count': load_counts.get(request.user.pk, 0)}]

    return render(request, 'core/reports/repair_analytics.html', {
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'sees_everyone': everyone,
        'total_orders': data['total_orders'],
        'avg_days': data['avg_days'],
        'by_employee': data['by_employee'],
        'by_fault_type': data['by_fault_type'],
        'load_rows': load_rows,
    })


@login_required
def report_repair_analytics_export(request):
    """Та же аналитика в Excel — тремя листами (по инженерам, по типу
    неисправности, по текущей загрузке). Права доступа те же, что
    и у страницы отчёта."""
    date_from, date_to = _repair_analytics_period(request)
    everyone = request.user.allows('reports_all_engineers')
    viewer = None if everyone else request.user

    data = _repair_analytics_data(date_from, date_to, viewer=viewer)

    load_counts = _current_load()
    if everyone:
        load_rows = [
            (e.full_name, load_counts.get(e.pk, 0))
            for e in Employee.objects.filter(is_active=True).order_by('full_name')
        ]
    else:
        load_rows = [(request.user.full_name, load_counts.get(request.user.pk, 0))]

    wb = build_workbook(
        'По инженерам',
        ['Сотрудник', 'Заказов', 'Среднее время ремонта, дней'],
        [
            [row['employee'].full_name, row['orders'], row['avg_days']]
            for row in data['by_employee']
        ],
    )
    add_sheet(
        wb, 'По типу неисправности',
        ['Тип неисправности', 'Заказов', 'Среднее время ремонта, дней'],
        [
            [row['fault_type'], row['orders'], row['avg_days']]
            for row in data['by_fault_type']
        ],
    )
    add_sheet(
        wb, 'Текущая загрузка',
        ['Сотрудник', 'Открытых заказов'],
        [[name, count] for name, count in load_rows],
    )

    return xlsx_response(wb, f'Аналитика ремонта {timezone.localdate():%Y-%m-%d}.xlsx')


# ==================== ПРИБЫЛЬ ПО ЗАКАЗАМ ====================

def _profit_period(request):
    """Диапазон дат отчёта о прибыли — по умолчанию последние 30 дней,
    как и у аналитики ремонта (`_repair_analytics_period`)."""
    default_to = timezone.localdate()
    default_from = default_to - timedelta(days=30)
    date_from = parse_date(request.GET.get('date_from', '')) or default_from
    date_to = parse_date(request.GET.get('date_to', '')) or default_to
    return date_from, date_to


def _profit_data(date_from, date_to):
    """Выручка, себестоимость деталей и прибыль за период — итог и разбивка
    по заказчику.

    Выручка считается по дате поступления денег (`Payment.payment_date`),
    себестоимость деталей — по дате её записи (`OrderCost.created_at`,
    заводится в момент списания детали заказа, см. `_use_repair_order_part`).
    Это два независимых события одного заказа, и они не обязаны попасть
    в один период — платёж и списание детали по одному заказу обычно
    и происходят в разные дни; отчёт складывает то, что случилось
    в периоде, для каждого из событий по отдельности, а не подбирает
    себестоимость под период платежа или наоборот.

    Себестоимость известна не всегда (`OrderCost.amount` может быть пустым
    — см. модель). Известная и неизвестная часть считаются раздельно, как
    и в плане закупок (`_purchase_plan_totals`): сумма по известным записям
    плюс отдельно количество записей без суммы. Прибыль здесь — прибыль
    по известной себестоимости; если неизвестных записей в периоде нет,
    она точна, если есть — реальная прибыль ниже показанной на неизвестную
    часть, и это явно видно по count'у, а не скрыто молчаливым `None`
    на весь отчёт (в отличие от `RepairOrder.parts_cost` для одного заказа,
    где именно такое молчаливое скрытие и было бы вводящим в заблуждение).
    """
    payments = Payment.objects.filter(payment_date__gte=date_from, payment_date__lte=date_to)
    costs = OrderCost.objects.filter(
        category='parts', created_at__date__gte=date_from, created_at__date__lte=date_to,
    )

    revenue_total = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    known_cost_total = costs.filter(amount__isnull=False).aggregate(
        total=Sum('amount'))['total'] or Decimal('0')
    unknown_cost_count = costs.filter(amount__isnull=True).count()

    by_client = defaultdict(lambda: {
        'revenue': Decimal('0'), 'known_cost': Decimal('0'), 'unknown_cost_count': 0,
    })
    client_names = {}

    for payment in payments.select_related('repair_order__client'):
        client = payment.repair_order.client
        by_client[client.pk]['revenue'] += payment.amount
        client_names[client.pk] = client.name

    for cost in costs.select_related('repair_order__client'):
        client = cost.repair_order.client
        if cost.amount is None:
            by_client[client.pk]['unknown_cost_count'] += 1
        else:
            by_client[client.pk]['known_cost'] += cost.amount
        client_names[client.pk] = client.name

    rows = sorted(
        (
            {
                'client': client_names[client_id],
                'revenue': values['revenue'],
                'known_cost': values['known_cost'],
                'unknown_cost_count': values['unknown_cost_count'],
                'profit': values['revenue'] - values['known_cost'],
            }
            for client_id, values in by_client.items()
        ),
        key=lambda row: row['client'],
    )

    return {
        'revenue_total': revenue_total,
        'known_cost_total': known_cost_total,
        'unknown_cost_count': unknown_cost_count,
        'profit_total': revenue_total - known_cost_total,
        'rows': rows,
    }


@permission_required('reports_profit')
def report_profit(request):
    """Прибыль по заказам за период: поступившие деньги минус себестоимость
    деталей, списанных по конкретным партиям прихода (не по средней
    и не по текущей цене детали), плюс разбивка по заказчику.

    Право — `reports_profit`, как у остальных финансовых разделов
    (`bank_statement` у поступлений, `invoices_send` у счетов): это деньги
    заказа, складу и мастеру они не нужны.
    """
    date_from, date_to = _profit_period(request)
    data = _profit_data(date_from, date_to)
    return render(request, 'core/reports/profit.html', {
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        **data,
    })


@permission_required('reports_profit')
def report_profit_export(request):
    """Тот же отчёт в Excel — двумя листами (итог за период, разбивка
    по заказчику). Права доступа те же, что и у страницы отчёта."""
    date_from, date_to = _profit_period(request)
    data = _profit_data(date_from, date_to)

    wb = build_workbook(
        'Итог за период',
        ['Показатель', 'Значение'],
        [
            ['Выручка (поступившие платежи), ₽', data['revenue_total']],
            ['Себестоимость деталей, известная часть, ₽', data['known_cost_total']],
            ['Списаний деталей без известной себестоимости, шт.', data['unknown_cost_count']],
            ['Прибыль по известной себестоимости, ₽', data['profit_total']],
        ],
    )
    add_sheet(
        wb, 'По заказчикам',
        ['Заказчик', 'Выручка, ₽', 'Себестоимость (известная), ₽',
         'Списаний без цены, шт.', 'Прибыль (известная), ₽'],
        [
            [row['client'], row['revenue'], row['known_cost'],
             row['unknown_cost_count'], row['profit']]
            for row in data['rows']
        ],
    )
    return xlsx_response(wb, f'Прибыль {date_from:%Y-%m-%d} - {date_to:%Y-%m-%d}.xlsx')


# ==================== AJAX: Создание из формы заказа ====================

@login_required
def ajax_equipment_model_list(request):
    """AJAX-получение списка моделей оборудования вместе с их исполнениями.

    Исполнения идут здесь же, а не отдельным запросом: моделей в справочнике
    десятки, ответ и так небольшой, а лишний поход на сервер при каждом
    выборе модели — это ещё одна вещь, которая не придёт, когда со связью
    плохо.
    """
    models = EquipmentModel.objects.prefetch_related('versions').order_by('name')
    return JsonResponse({'success': True, 'models': [
        {
            'id': model.pk,
            'name': model.name,
            'versions': [
                {'id': version.pk, 'name': version.name}
                for version in model.versions.all()
            ],
        }
        for model in models
    ]})


@login_required
@require_POST
def ajax_equipment_model_create(request):
    """AJAX-создание модели оборудования из формы заказа."""
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Название модели обязательно'})

    model, created = EquipmentModel.objects.get_or_create(name=name)
    if not created:
        return JsonResponse({'success': False, 'error': 'Модель с таким названием уже существует'})

    return JsonResponse({
        'success': True,
        'id': model.id,
        'name': model.name,
        'message': f'Модель "{model.name}" создана'
    })


@login_required
@require_POST
def ajax_equipment_create(request):
    """AJAX-создание оборудования из формы заказа."""
    model_id = request.POST.get('model_id')
    serial_number = request.POST.get('serial_number', '').strip()

    if not model_id:
        return JsonResponse({'success': False, 'error': 'Выберите модель'})
    if not serial_number:
        return JsonResponse({'success': False, 'error': 'Укажите серийный номер'})

    try:
        model = EquipmentModel.objects.get(pk=model_id)
    except EquipmentModel.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Модель не найдена'})

    if Equipment.objects.filter(serial_number=serial_number).exists():
        return JsonResponse({'success': False, 'error': 'Оборудование с таким серийным номером уже существует'})

    # Похожий серийник — не ошибка, а повод переспросить: «БУАД-1234» и
    # «буад 1234» это, скорее всего, одна и та же единица, набранная разными
    # людьми по-разному. Решает человек: если создать вторую запись молча,
    # история ремонтов разъедется на две и найти её будет нечем.
    if request.POST.get('confirmed') != '1':
        similar = Equipment.find_similar(serial_number)
        if similar.exists():
            return JsonResponse({
                'success': False,
                'similar': [
                    {
                        'id': eq.id,
                        'name': str(eq),
                        'serial_number': eq.serial_number,
                        'orders_count': eq.repair_orders.count(),
                        'history_url': reverse('equipment_history', args=[eq.id]),
                    }
                    for eq in similar[:5]
                ],
            })

    # Заказчик приходит из формы заказа, где он уже выбран. Если ещё
    # не выбран — оставляем пустым и проставим при сохранении заказа
    # (RepairOrder.assign_equipment_owners); выдумывать владельца нельзя.
    client = None
    client_id = request.POST.get('client_id')
    if client_id:
        client = Client.objects.filter(pk=client_id).first()

    # Исполнение — только своё у этой модели: чужое означало бы, что
    # у прибора обозначение от другой модели.
    version = None
    version_id = request.POST.get('version_id')
    if version_id:
        version = model.versions.filter(pk=version_id).first()

    equipment = Equipment.objects.create(
        model=model, serial_number=serial_number,
        version=version, current_client=client
    )

    return JsonResponse({
        'success': True,
        'id': equipment.id,
        'name': str(equipment),
        'message': f'Оборудование "{equipment}" создано'
    })


@login_required
def ajax_equipment_history_summary(request, pk):
    """Краткая сводка прошлых ремонтов — для подсказки в форме заказа."""
    equipment = get_object_or_404(Equipment.objects.select_related('model'), pk=pk)
    visits = equipment.repair_history()
    count = visits.count()
    last = visits.first()

    # Гарантия — самое важное, что можно сказать при приёме: если единица
    # вернулась в её пределах, это повторное обращение по тому же ремонту,
    # и решать по оплате нужно до того, как заказ оформлен
    warranty = equipment.active_warranty()

    return JsonResponse({
        'equipment': str(equipment),
        'orders_count': count,
        'last_order_number': last.repair_order.order_number if last else '',
        'last_date': (
            timezone.localtime(last.repair_order.date_received).strftime('%d.%m.%Y')
            if last else ''
        ),
        'history_url': reverse('equipment_history', args=[equipment.pk]),
        'under_warranty': warranty is not None,
        'warranty_until': (
            timezone.localtime(warranty.warranty_until).strftime('%d.%m.%Y')
            if warranty else ''
        ),
        'warranty_order_number': warranty.repair_order.order_number if warranty else '',
        'warranty_order_url': (
            reverse('repair_order_detail', args=[warranty.repair_order_id])
            if warranty else ''
        ),
    })


@login_required
def ajax_equipment_faults(request, pk):
    """Типовые неисправности модели этой единицы оборудования — для выбора
    в строке заказа. Список зависит от того, какое оборудование выбрано
    в конкретной строке, и меняется вместе с ним, поэтому тянется отдельным
    запросом, а не единым списком по всем моделям сразу."""
    equipment = get_object_or_404(Equipment.objects.select_related('model'), pk=pk)
    faults = FaultType.objects.filter(equipment_model_id=equipment.model_id).order_by('name')
    return JsonResponse({
        'success': True,
        'faults': [{'id': fault.id, 'name': fault.name} for fault in faults],
    })


@login_required
@require_POST
def ajax_client_create(request):
    """AJAX-создание заказчика из формы заказа."""
    name = request.POST.get('name', '').strip()
    inn = request.POST.get('inn', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    contact_person = request.POST.get('contact_person', '').strip()

    if not name:
        return JsonResponse({'success': False, 'error': 'Название заказчика обязательно'})

    if Client.objects.filter(name=name).exists():
        return JsonResponse({'success': False, 'error': 'Заказчик с таким названием уже существует'})

    client = Client.objects.create(
        name=name,
        inn=inn,
        phone=phone,
        email=email,
        contact_person=contact_person
    )

    return JsonResponse({
        'success': True,
        'id': client.id,
        'name': client.name,
        'message': f'Заказчик "{client.name}" создан'
    })


# ==================== АДМИНИСТРИРОВАНИЕ ====================

@permission_required('admin_access')
def admin_users(request):
    users = Employee.objects.all().order_by('full_name')
    return render(request, 'core/admin/users.html', {'users': users})


@permission_required('admin_access')
def admin_user_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Пользователь создан')
            return redirect('admin_users')
    else:
        form = EmployeeForm()
    return render(request, 'core/admin/user_form.html', {'form': form, 'title': 'Новый пользователь'})


@permission_required('admin_access')
def admin_user_edit(request, pk):
    user = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=user)
        if form.is_valid():
            try:
                _save_keeping_admin(form)
            except LastAdminError:
                messages.error(request, _NO_ADMIN_LEFT)
            else:
                messages.success(request, 'Пользователь обновлён')
                return redirect('admin_users')
    else:
        form = EmployeeForm(instance=user)
    return render(request, 'core/admin/user_form.html', {'form': form, 'title': 'Редактирование пользователя', 'user_obj': user})


# ==================== ДОЛЖНОСТИ И ПРАВА (v2.98.0) ====================
#
# Права заводит администратор, а не программист: до v2.98.0 ролей было
# ровно четыре, они были зашиты в код, и пятая требовала правки исходников.

_NO_ADMIN_LEFT = (
    'Не сохранено: после этой правки в программе не осталось бы ни одного '
    'действующего сотрудника с полным доступом. Сначала дайте полный доступ '
    'кому-то ещё.'
)


@contextmanager
def _admin_access_kept():
    """Не дать правке отобрать последний полный доступ.

    Проверяем **по факту**, уже после записи и внутри транзакции, а не
    предсказыванием «что будет, если»: предсказание однажды окажется
    неверным, и владелец останется снаружи собственной программы — чинить
    это пришлось бы из консоли на самом Pi. То же соображение, что
    и у `_keep_current_host` на странице настроек, только цена ошибки выше.

    Сравниваем с тем, что было **до** правки: если доступа не было
    и так (например, единственного администратора уволили раньше),
    запрет только мешал бы чинить положение — «нельзя, потому что уже
    сломано» это не защита, а ловушка.
    """
    had = admin_access_exists()
    yield
    if had and not admin_access_exists():
        raise LastAdminError


def _save_keeping_admin(form):
    """Сохранить форму, откатив правку, которая отобрала бы полный доступ."""
    with transaction.atomic(), _admin_access_kept():
        return form.save()


@permission_required('admin_access')
def admin_positions(request):
    """Должности со списком выданных прав и числом занятых."""
    positions = (
        Position.objects.prefetch_related('permissions')
        .annotate(people=Count('employees'))
    )
    return render(request, 'core/admin/positions.html', {'positions': positions})


@permission_required('admin_access')
def admin_position_create(request):
    if request.method == 'POST':
        form = PositionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Должность заведена')
            return redirect('admin_positions')
    else:
        form = PositionForm()
    return render(request, 'core/admin/position_form.html', {
        'form': form, 'title': 'Новая должность',
        'sections': permissions_by_section(),
    })


@permission_required('admin_access')
def admin_position_edit(request, pk):
    position = get_object_or_404(Position, pk=pk)
    if request.method == 'POST':
        form = PositionForm(request.POST, instance=position)
        if form.is_valid():
            try:
                _save_keeping_admin(form)
            except LastAdminError:
                messages.error(request, _NO_ADMIN_LEFT)
            else:
                messages.success(request, 'Должность обновлена')
                return redirect('admin_positions')
    else:
        form = PositionForm(instance=position)
    return render(request, 'core/admin/position_form.html', {
        'form': form, 'title': 'Правка должности', 'position': position,
        'sections': permissions_by_section(),
    })


@permission_required('admin_access')
def admin_position_delete(request, pk):
    """Удаление должности.

    Занятую удалить нельзя — и не потому, что «нельзя», а потому что
    иначе люди на ней молча лишились бы прав (`PROTECT` у поля). Страница
    называет тех, кто на ней числится: переводить их всё равно придётся,
    и лучше знать сразу, кого.
    """
    position = get_object_or_404(Position, pk=pk)
    holders = list(position.employees.order_by('full_name'))

    if request.method == 'POST':
        if holders:
            messages.error(
                request,
                'Не удалено: должность занята. Сначала переведите людей '
                'на другую должность.'
            )
            return redirect('admin_positions')
        try:
            with transaction.atomic(), _admin_access_kept():
                position.delete()
        except LastAdminError:
            messages.error(request, _NO_ADMIN_LEFT)
            return redirect('admin_positions')
        messages.success(request, 'Должность удалена')
        return redirect('admin_positions')

    return render(request, 'core/admin/position_delete.html', {
        'position': position, 'holders': holders,
    })






@permission_required('admin_access')
def admin_notifications(request):
    """Очередь оповещений: что ушло, что не ушло и почему.

    Без этой страницы на вопрос «почему заказчик не получил письмо» отвечать
    нечем: очередь живёт в базе, а журнал отправки — в логах systemd.
    """
    status = request.GET.get('status', '')
    queue = Notification.objects.select_related('repair_order', 'part')
    if status in dict(Notification.STATUS_CHOICES):
        queue = queue.filter(status=status)

    paginator = Paginator(queue.order_by('-created_at'), 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    counts = {
        item['status']: item['count']
        for item in Notification.objects.values('status').annotate(count=Count('id'))
    }
    status_tabs = [
        {'code': code, 'label': label, 'count': counts.get(code, 0)}
        for code, label in Notification.STATUS_CHOICES
    ]

    return render(request, 'core/admin/notifications.html', {
        'notifications': page_obj,
        'status_filter': status,
        'status_tabs': status_tabs,
        'found_count': paginator.count,
        'sending_enabled': envfile.setting('NOTIFICATIONS_ENABLED', False),
        'clients_enabled': envfile.setting('NOTIFY_CLIENTS', False),
        'low_stock_enabled': envfile.setting('NOTIFY_LOW_STOCK', True),
        'channels': [
            {
                'name': 'MAX',
                'configured': messengers.max_is_configured(),
                'enabled': envfile.setting('NOTIFY_MAX', False)
                           and messengers.max_is_configured(),
            },
            {
                'name': 'Telegram',
                'configured': messengers.telegram_is_configured(),
                'enabled': envfile.setting('NOTIFY_TELEGRAM', False)
                           and messengers.telegram_is_configured(),
            },
        ],
    })


@permission_required('admin_access')
@require_POST
def admin_notification_retry(request, pk):
    """Вернуть неудачное оповещение в очередь.

    Счётчик попыток обнуляется: обычно причина неудачи — не письмо, а почта
    (пароль приложения, отвалившийся канал), и после починки прошлые попытки
    ни о чём не говорят.
    """
    notification = get_object_or_404(Notification, pk=pk)
    notification.status = 'pending'
    notification.attempts = 0
    notification.last_error = ''
    notification.save(update_fields=['status', 'attempts', 'last_error'])

    messages.success(request, f'Оповещение для {notification.recipient} возвращено в очередь')
    return redirect('admin_notifications')


# ==================== НАСТРОЙКИ ====================

@permission_required('admin_access')
def admin_settings(request):
    """Страница настроек: правимое — здесь, секреты — только состоянием.

    Хранилище — файл `.env` на самом Pi, а не база: базу увозит ночная
    выгрузка в облако. Правка действует сразу, без перезапуска, —
    за исключением того, что Django читает при старте; такие настройки
    помечены прямо на странице.

    Секретов на странице нет и быть не может. Она открывается с любого
    устройства в Tailscale, а по локальному адресу браузер идёт
    без сертификата — токен ушёл бы открытым текстом. Их вводят у Pi:
    manage.py setsecret. Здесь видно только «задан, столько-то знаков».
    """
    return render(request, 'core/admin/settings.html', {
        'sections': envfile.editable_sections(_settings_notes()),
        'secrets': [envfile.describe_secret(name)
                    for name in envfile.SECRET_NAMES],
        'env_path': envfile.path(),
        'env_exists': envfile.exists(),
        # Ничего не хранится: «нужен перезапуск» выводится из того,
        # что в файле лежит не то, с чем программа запускалась
        'restart_needed': [envfile.title_of(name)
                           for name in envfile.restart_needed()],
        'restart_available': restarter.is_available(),
        'restart_pending': restarter.pending(),
        'history': SettingChange.objects.select_related('changed_by')[:20],
    })


def _settings_notes():
    """Живые приписки под полями настроек.

    Без них частота загрузки выписки — число в вакууме: понять,
    работает ли расписание вообще, можно только по журналу systemd.
    Привязка по имени настройки, а не по названию раздела: имя —
    опознавательный знак, а названия разделов переписывают.
    """
    notes = {}
    fetched = tbank.last_fetch_at()
    if fetched:
        notes['TBANK_STATEMENT_INTERVAL_MINUTES'] = (
            'Выписка последний раз загружалась %s.'
            % timezone.localtime(fetched).strftime('%d.%m.%Y в %H:%M')
        )
    elif tbank.is_configured():
        notes['TBANK_STATEMENT_INTERVAL_MINUTES'] = (
            'Выписку ещё не тянули ни разу. Если расписание установлено, '
            'первая загрузка случится на ближайшем тике.'
        )
    return notes


@permission_required('admin_access')
@require_POST
def admin_settings_save(request):
    """Записать правимые настройки в файл.

    Пишется только то, что действительно изменилось: каждая запись
    перекладывает файл целиком и оставляет копию, и делать это ради
    неизменившегося поля незачем. Заодно в журнале не появляется
    правок, которых не было.
    """
    # Невыбранный флажок в запрос не попадает вовсе, и отличить «снял
    # галочку» от «поля на форме не было» по одному его отсутствию нельзя.
    # Поэтому форма присылает список показанных флажков: молча выключить
    # оповещения неполный запрос не должен
    shown_flags = set(request.POST.getlist('flag'))
    saved, failed = [], []
    kept_host = None
    for name, row in envfile.EDITABLE_BY_NAME.items():
        if row['kind'] == 'flag':
            if name not in shown_flags:
                continue
            value = 'True' if request.POST.get(name) else 'False'
        else:
            if name not in request.POST:
                continue
            value = (request.POST.get(name) or '').strip()
            if name == 'ALLOWED_HOSTS':
                value, kept_host = _keep_current_host(request, value)
        if value == envfile.as_text(envfile.setting(name, '')):
            continue
        try:
            envfile.set_value(name, value)
        except envfile.EnvFileError as error:
            failed.append(str(error))
            continue
        saved.append(name)
        SettingChange.objects.create(name=name, changed_by=request.user)

    for text in failed:
        messages.error(request, text)
    if kept_host:
        messages.warning(
            request,
            'Адрес %s возвращён в список: вы работаете по нему, и без него '
            'следующий же переход упёрся бы в «400 Bad Request». Убрать его '
            'можно, открыв программу по другому адресу.' % kept_host
        )
    if saved:
        restart = [name for name in saved
                   if envfile.EDITABLE_BY_NAME[name]['restart']]
        messages.success(
            request, 'Сохранено настроек: %d. Действуют сразу.' % len(saved)
        )
        if restart:
            messages.warning(
                request,
                'Эти настройки подхватятся только после перезапуска службы: '
                '%s. Кнопка внизу страницы.' % ', '.join(
                    envfile.EDITABLE_BY_NAME[name]['title'] for name in restart
                )
            )
    elif not failed:
        messages.info(request, 'Менять было нечего — всё осталось как было.')
    return redirect('admin_settings')


def _keep_current_host(request, value):
    """Не дать вычеркнуть адрес, по которому админ сейчас работает.

    `ALLOWED_HOSTS` — единственная настройка, ошибка в которой закрывает
    программу целиком: Django ответит «400 Bad Request» на каждый запрос,
    и починить это можно будет только по SSH. Поэтому свой адрес
    возвращается в список молча, а сказано об этом вслух.

    Убрать его всё-таки можно — открыв программу по другому адресу
    из списка. Тогда «свой» будет уже другой.
    """
    host = request.get_host().split(':')[0]
    hosts = [part.strip() for part in value.split(',') if part.strip()]
    if not hosts or host in hosts or '*' in hosts:
        return value, None
    return ','.join([*hosts, host]), host


@permission_required('admin_access')
@require_POST
def admin_settings_restart(request):
    """Заявка на перезапуск службы.

    Сам перезапуск делает служба systemd от root — приложение работает
    от своего пользователя и root не имеет намеренно (`core/restarter.py`).
    Об обрыве связи и о возвращении её обратно скажет обычная полоса
    вверху страницы: своего слежения за ходом перезапуска заводить
    не надо.
    """
    if not restarter.is_available():
        messages.error(
            request,
            'Перезапуск из программы не настроен: на Raspberry Pi нет '
            'скрипта /usr/local/sbin/lifteam-restart. Порядок установки — '
            'в DEPLOY.md. Пока его нет, перезапуск делается по SSH: '
            'sudo systemctl restart lifteam'
        )
        return redirect('admin_settings')

    restarter.request_restart(requested_by=request.user.username)
    messages.info(
        request,
        'Заявка на перезапуск подана. Связь на несколько секунд оборвётся — '
        'полоса вверху страницы скажет, когда она вернётся.'
    )
    return redirect('admin_settings')


@permission_required('admin_access')
@require_POST
def admin_settings_check(request):
    """Проверить связь со службой. Отвечает JSON — страница не уезжает.

    Проверку держит одно место (`core/selfcheck.py`): команда у Pi
    и эта кнопка обязаны говорить об одной службе одно и то же.
    """
    name = (request.POST.get('name') or '').strip().upper()
    if name not in envfile.SECRET_NAMES:
        return JsonResponse({'state': 'fail', 'message': 'Неизвестная служба.'},
                            status=400)
    result = selfcheck.check(name)
    return JsonResponse({'state': result.state, 'message': result.message})


# ==================== ОБНОВЛЕНИЕ ПРИЛОЖЕНИЯ ====================

@permission_required('admin_access')
def admin_update(request):
    """Страница обновления. Само обновление выполняет служба systemd от root —
    приложение лишь оставляет заявку (подробности в core/updater.py)."""
    if not updater.is_git_checkout():
        return render(request, 'core/admin/update.html', {'not_a_repo': True})

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'check':
            ok, output = updater.fetch_remote()
            if ok:
                count = len(updater.pending_updates())
                messages.success(
                    request,
                    f'Доступно обновлений: {count}' if count else 'Установлена последняя версия'
                )
            else:
                messages.error(request, f'Не удалось проверить обновления: {output}')

        elif action in ('update', 'rollback'):
            if updater.update_in_progress():
                messages.warning(request, 'Обновление уже выполняется')
            else:
                target = 'latest' if action == 'update' else request.POST.get('target', '')
                try:
                    updater.request_update(target, requested_by=request.user.username)
                except ValueError as exc:
                    messages.error(request, str(exc))
                else:
                    messages.info(request, 'Обновление запущено, приложение перезапустится')
        return redirect('admin_update')

    return render(request, 'core/admin/update.html', {
        'current': updater.current_version(),
        'pending': updater.pending_updates(),
        'history': updater.recent_history(),
        'status': updater.read_status(),
    })


@permission_required('admin_access')
def admin_update_status(request):
    """Ход обновления для опроса со страницы."""
    return JsonResponse(updater.read_status() or {'state': 'idle', 'message': '', 'log': []})


# ==================== ПЕЧАТНЫЕ АКТЫ ====================

# Печать через браузер, а не сборка PDF на сервере: диалог печати сохраняет
# в PDF сам, а библиотека рендеринга на Raspberry Pi — это лишние зависимости
# и лишние мегабайты ради того, что уже умеет каждый браузер.

def _act_context(pk):
    """Общее для обоих актов: заказ, оборудование и шапка с реквизитами."""
    order = get_object_or_404(
        RepairOrder.objects.select_related('client'), pk=pk
    )
    return {
        'order': order,
        # Юрлицо заказа, а не всегда основное: документы должны быть
        # от того же лица, от которого выставлен счёт
        'organization': order.legal_entity(),
        'order_equipments': list(
            order.order_equipments.select_related('equipment__model').order_by('id')
        ),
        'today': timezone.localdate(),
    }


@login_required
def repair_order_act_receive(request, pk):
    """Акт приёма оборудования в ремонт.

    Печатается, когда оборудование привезли: в нём пломбы и состояние
    на момент приёма — то, о чём потом спорят, если что-то не так.
    """
    return render(request, 'core/repair_orders/act_receive.html', _act_context(pk))


@login_required
def repair_order_act_complete(request, pk):
    """Акт выполненных работ.

    Печатается при выдаче: что сделали, сколько стоит, до какого числа
    действует гарантия.
    """
    context = _act_context(pk)
    order = context['order']
    context['details'] = list(order.details.select_related('part'))
    context['warranty_until'] = (
        add_months(order.date_completed, warranty_months())
        if order.date_completed and warranty_months() else None
    )
    return render(request, 'core/repair_orders/act_complete.html', context)


def _order_equipment(order_pk, roe_pk):
    """Единица оборудования в заказе — с проверкой, что она из этого заказа.

    Иначе по подобранному адресу открылась бы чужая позиция, и акт ушёл бы
    заказчику с серийником из другого заказа.
    """
    return get_object_or_404(
        RepairOrderEquipment.objects.select_related(
            'equipment__model', 'repair_order__client'
        ),
        pk=roe_pk, repair_order_id=order_pk,
    )


@login_required
def repair_order_unit_detail(request, order_pk, roe_pk):
    """Страница одной единицы оборудования в заказе.

    Всё об этом приборе в этой работе, в одном месте: что привезли, что
    нашли при диагностике, что сделали, какие детали ушли, какие пломбы
    стоят, где документы и папка со снимками.

    Страница, а не окно. Окна в программе рисует Bootstrap, а он приходит
    из интернета и регулярно не приезжает; сюда же ведёт наклейка
    с прибора, ссылку на страницу можно дать в переписке и открыть
    в соседней вкладке. Всё это окно не умеет.

    С v2.95.0 здесь же и дефектация: своей страницы у неё больше нет.
    Мастер с платой в руках заполнял диагноз на одном экране, работы
    на другом, а список того, что осталось, читал на третьем.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)
    order = order_equipment.repair_order
    equipment = order_equipment.equipment

    # Позиция в заказе печатается на наклейке, и человек с коробкой
    # в руках ищет по ней. Считаем по тому же порядку, что и в списке
    # единиц на карточке заказа
    position = list(
        order.order_equipments.order_by('id').values_list('pk', flat=True)
    ).index(order_equipment.pk) + 1

    # Гарантия по прошлым ремонтам этой же железки: текущий заказ
    # исключён, иначе он сам попадал бы в «повторное обращение»
    previous = Equipment.warranty_map([equipment], exclude_order_id=order.pk)

    # Цена по прайсу — предложение, а не подстановка: мастер видит её рядом
    # с полем оценки и решает сам. Молча вписывать нельзя — он подписывает
    # акт. Переехало сюда вместе с самой дефектацией
    line = order_equipment.price_list_line

    details = list(order_equipment.details.select_related('part').order_by('id'))

    return render(request, 'core/repair_orders/unit_detail.html', {
        'order': order,
        'order_equipment': order_equipment,
        # Раздел правится своей формой и своей кнопкой: сохранение одного
        # раздела не должно трогать поля другого — прибор в работе бывает
        # у двоих сразу
        'diagnosis_form': UnitDiagnosisForm(instance=order_equipment),
        'repair_form': UnitRepairForm(instance=order_equipment),
        'equipment': equipment,
        'position': position,
        'previous_warranty': previous.get(equipment.pk),
        'faults': order_equipment.faults.all(),
        'price_line': line,
        'price_source': str(line.price_list) if line else '',
        # Детали, списанные именно на эту единицу. Списанные «на заказ
        # целиком» сюда не попадают намеренно: программа не знает,
        # в какую железку они ушли, и приписывать их наугад нельзя
        'details': details,
        # Намеченные, но ещё не взятые со склада — ради кнопки «списать
        # весь список». Считаем здесь, а не в шаблоне: `details` уже
        # выбраны из базы, и второй запрос ради того же незачем
        'planned_details': [line for line in details if line.is_planned],
        # Форма списания на самой странице: единицу выбирать не надо,
        # она известна — в этом и выигрыш перед формой на карточке заказа
        'detail_form': RepairOrderDetailForm(),
        'materials': equipment.model.materials_for(equipment.version),
        'tech_cards': equipment.model.tech_cards.select_related('fault_type'),
        'repairs_count': equipment.repair_orders.count(),
    })


@login_required
def repair_order_unit_remove(request, order_pk, roe_pk):
    """Убрать единицу из заказа.

    Отдельная страница подтверждения, а не окно: убрать прибор из заказа
    — это потерять по нему записанные работы, дефектацию и пломбы,
    и переспросить тут дешевле, чем восстанавливать.

    **Детали, списанные на эту единицу, со склада не возвращаются.**
    Их и правда взяли и потратили; связь с единицей просто обнуляется,
    и списание становится «на заказ целиком». Возврат на склад — своё
    действие, и делать его молча за человека нельзя.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)

    if request.method == 'POST':
        title = str(order_equipment.equipment)
        order_equipment.delete()
        messages.success(request, f'{title} убрано из заказа')
        return redirect('repair_order_detail', pk=order_pk)

    return render(request, 'core/repair_orders/unit_remove.html', {
        'order': order_equipment.repair_order,
        'order_equipment': order_equipment,
        'details_count': order_equipment.details.count(),
    })


# Разделы страницы единицы и формы, которыми они правятся (с v2.95.0).
# Имя раздела приходит в запросе, и незнакомое не принимается: без этого
# списка подсунутое имя означало бы «сохрани неизвестно что».
UNIT_SECTION_FORMS = {
    'diagnosis': UnitDiagnosisForm,
    'repair': UnitRepairForm,
}


@login_required
@require_POST
def repair_order_unit_edit(request, order_pk, roe_pk):
    """Сохранить один раздел страницы единицы.

    Только POST: страница показывается своим представлением, а сюда
    приходит форма того раздела, который правили. Возвращаемся туда же,
    к нему: мастер записывает работы по одному прибору и берётся
    за следующий, а не уходит в заказ целиком.

    Раздел приходит полем `section`. Сохраняется **только он**: одна
    форма на всю страницу означала бы, что открывший её утром и
    сохранивший вечером затирает всё, что за день вписал сосед.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)
    section = request.POST.get('section', '')
    form_class = UNIT_SECTION_FORMS.get(section)
    if form_class is None:
        messages.error(request, 'Не сохранено: неизвестный раздел страницы')
        return redirect('repair_order_unit_detail',
                        order_pk=order_pk, roe_pk=roe_pk)

    form = form_class(request.POST, instance=order_equipment)
    if form.is_valid():
        saved = form.save()
        if section == 'diagnosis':
            # Дата акта — день первой записи, а не день печати: пустая
            # означала «печатай сегодняшнее число», и акт, заполненный
            # в понедельник, к среде становился средой. Цену прайса
            # запоминаем тогда же — только здесь известно, от чего мастер
            # отступил, договариваясь с заказчиком
            saved.stamp_defect_act_date()
            saved.freeze_list_price()
        messages.success(request, 'Записано')
    else:
        messages.error(request, 'Не сохранено: проверьте отмеченные поля')

    url = reverse('repair_order_unit_detail', args=[order_pk, roe_pk])
    return redirect(f'{url}#{section}' if section else url)


@login_required
def repair_order_defect_act_edit(request, order_pk, roe_pk):
    """Бывшая страница дефектации — теперь указатель на её место.

    С v2.95.0 дефектацию заполняют в разделе «Приём и диагностика»
    на странице самой единицы: своей страницы у неё больше нет. Адрес
    оставлен переадресацией, а не убран совсем, — на него ведут закладки
    в браузерах мастеров и ссылки, разосланные в переписке; 404 на них
    выглядел бы как пропавшая дефектация.

    Это не второе место правки, а вывеска: своей формы здесь нет.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)
    url = reverse('repair_order_unit_detail',
                  args=[order_equipment.repair_order_id, order_equipment.pk])
    return redirect(f'{url}#diagnosis')


@login_required
@require_POST
def repair_order_unit_disk_folder(request, order_pk, roe_pk):
    """Завести на Яндекс.Диске папку под снимки этой единицы.

    Одно нажатие вместо похода в веб-интерфейс Диска и ручной вставки
    ссылки: путь считает одно место (`yadisk.unit_path`), поэтому снимки
    одного ремонта не могут уехать в папку другого.

    Папка, которая уже есть, ошибкой не считается — ссылка просто
    записывается заново. Так же ведёт себя повторное нажатие двумя
    мастерами почти одновременно.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)

    reason = yadisk.unconfigured_reason()
    if reason:
        messages.error(request, f'Папка не создана. {reason}')
        return redirect('repair_order_detail', pk=order_pk)

    try:
        url = yadisk.ensure_unit_folder(order_equipment)
    except yadisk.YandexDiskError as exc:
        messages.error(request, f'Папка не создана: {exc}')
        return redirect('repair_order_detail', pk=order_pk)

    # Ссылку записываем и тогда, когда папка уже была: поле могли
    # очистить руками, а путь от этого не изменился
    order_equipment.yandex_disk_folder = url
    order_equipment.save(update_fields=['yandex_disk_folder'])
    messages.success(
        request,
        f'Папка на Яндекс.Диске готова: {yadisk.unit_path(order_equipment)}'
    )
    return redirect(f"{reverse('repair_order_detail', args=[order_pk])}#unit-{roe_pk}")


@login_required
@require_POST
def repair_order_unit_disk_folder_set(request, order_pk, roe_pk):
    """Вписать ссылку на папку Диска руками — рядом с кнопкой «завести
    папку» (с v2.93.0).

    На Диске ничего не создаёт и не проверяет: это для двух случаев,
    когда автоматический путь не годится — папку завели раньше программы,
    или снимки переложили в другое место.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)
    form = UnitDiskFolderForm(request.POST)
    if form.is_valid():
        order_equipment.yandex_disk_folder = form.cleaned_data['url']
        order_equipment.save(update_fields=['yandex_disk_folder'])
        messages.success(request, 'Ссылка на папку сохранена.')
    else:
        messages.error(request, 'Ссылка не похожа на настоящую — не сохранено.')
    return redirect('repair_order_unit_detail', order_pk=order_pk, roe_pk=roe_pk)


@login_required
def repair_order_act_defect(request, order_pk, roe_pk):
    """Акт дефектации оборудования.

    Печатается по итогам диагностики: что нашли внутри, какие коды ошибок
    в памяти устройства, гарантийный ли случай и во сколько обойдётся
    ремонт. Один акт на одну единицу — заказчик по нему решает, чинить ли,
    а решает он по каждой единице отдельно.
    """
    order_equipment = _order_equipment(order_pk, roe_pk)
    return render(request, 'core/repair_orders/act_defect.html', {
        'order': order_equipment.repair_order,
        'order_equipment': order_equipment,
        'organization': order_equipment.repair_order.legal_entity(),
        'act_date': order_equipment.defect_act_date or timezone.localdate(),
    })


@login_required
def repair_order_quote_edit(request, pk):
    """Условия коммерческого предложения и строки по единицам."""
    order = get_object_or_404(RepairOrder.objects.select_related('client'), pk=pk)

    if request.method == 'POST':
        form = QuoteForm(request.POST, instance=order)
        formset = QuoteLineFormSet(request.POST, instance=order)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Коммерческое предложение сохранено')
            return redirect('repair_order_quote', pk=order.pk)
        messages.error(request, 'Предложение не сохранено: проверьте отмеченные поля')
    else:
        today = timezone.localdate()
        form = QuoteForm(instance=order, initial={
            'quote_date': order.quote_date or today,
            'quote_valid_until': order.quote_valid_until or today + timedelta(
                days=envfile.setting('QUOTE_VALID_DAYS', 14)),
        })
        formset = QuoteLineFormSet(instance=order)

    return render(request, 'core/repair_orders/quote_form.html', {
        'order': order,
        'form': form,
        'formset': formset,
    })


@login_required
def repair_order_quote(request, pk):
    """Коммерческое предложение на А4.

    Печатается по итогам дефектации: что предлагаем сделать, во сколько
    обойдётся, в какие сроки и на каких условиях.
    """
    order = get_object_or_404(RepairOrder.objects.select_related('client'), pk=pk)
    return render(request, 'core/repair_orders/quote.html', {
        'order': order,
        'organization': order.legal_entity(),
        'rows': order.quote_rows(),
        'total': order.quote_total,
        'quote_date': order.quote_date or timezone.localdate(),
    })


# ==================== СЧЁТ ЧЕРЕЗ API БАНКА ====================

# Единственное место в программе, которое что-то создаёт в банке. Поэтому:
# отправка только по нажатию человека, только со страницы, где он видит
# всё, что уйдёт, и только при включённом выключателе того банка,
# который выбран. По расписанию счета не выставляются нигде и никогда.
#
# Банков два, и с v2.50.0 представление не знает, с каким именно работает:
# всё идёт через общий интерфейс из `core/invoicing.py`. Разницу между
# банками знают только их собственные модули.


def _invoice_organization(provider_code):
    """Юрлицо, от которого уйдёт счёт этого банка.

    Возвращает пару «юрлицо, закреплено ли оно за банком». Если за банком
    юрлица нет, берётся основное, а вторым значением приходит False —
    чтобы страница честно сказала бухгалтеру, чьи реквизиты подставлены
    и по какому ряду считается номер. Отказывать в выставлении из-за
    этого не за что: реквизиты юрлица в запрос к банку не уходят вовсе,
    документ банк рисует по своему счёту. А вот ряд номеров окажется
    чужим — и об этом надо предупредить.
    """
    bound = Organization.for_provider(provider_code)
    if bound is not None:
        return bound, True
    return Organization.get_solo(), False


def _invoice_payload(provider, order, form_data):
    """Тело запроса к банку — то самое, что показано на странице."""
    client = order.client
    return provider.build_invoice(
        number=form_data['invoice_number'],
        items=order.invoice_items(),
        payer={'name': client.name, 'inn': client.inn, 'kpp': client.kpp},
        emails=form_data['emails'],
        invoice_date=form_data['invoice_date'],
        due_date=form_data['due_date'],
    )


@permission_required('invoices_send')
def repair_order_invoice(request, pk):
    """Выставление счёта заказчику через API банка.

    GET показывает, что именно уйдёт в банк; POST отправляет. Разделение
    не формальность: счёт уходит заказчику, и «подтверждаю» должно означать
    согласие с тем, что человек видел, а не с тем, что программа придумала.
    """
    order = get_object_or_404(RepairOrder.objects.select_related('client'), pk=pk)
    items = order.invoice_items()

    if request.method == 'POST':
        form = InvoiceSendForm(request.POST)
        if form.is_valid():
            return _send_invoice(request, order, form)
        chosen = (request.POST.get('provider')
                  or invoicing.default_provider_for(request.user))
    else:
        chosen = invoicing.default_provider_for(request.user)
        today = timezone.localdate()
        form = InvoiceSendForm(initial={
            'provider': chosen,
            'invoice_number': order.invoice_number or RepairOrder.next_invoice_number(),
            'invoice_date': today,
            'due_date': today + timedelta(days=invoicing.due_days(chosen)),
            'emails': order.client.email,
        })

    organization, bound = _invoice_organization(chosen)
    try:
        provider = invoicing.get_provider(chosen)
        enabled = provider.invoice_enabled()
        configured = provider.is_configured()
        missing = provider.missing_settings()
    except invoicing.InvoiceError:
        provider, enabled, configured, missing = None, False, False, []

    return render(request, 'core/bank/invoice.html', {
        'order': order,
        'form': form,
        'items': items,
        'items_total': sum(Decimal(str(item['price'])) for item in items),
        'provider': provider,
        'provider_label': invoicing.provider_label(chosen),
        'organization': organization,
        'organization_is_bound': bound,
        'enabled': enabled,
        'configured': configured,
        'missing_settings': missing,
    })


def _send_invoice(request, order, form):
    """Отправка счёта в банк и запись следов в заказе."""
    chosen = form.cleaned_data['provider']

    try:
        provider = invoicing.get_provider(chosen)
        payload = _invoice_payload(provider, order, form.cleaned_data)
        response = provider.send_invoice(payload, emails=form.cleaned_data['emails'])
    except invoicing.InvoiceError as exc:
        # Ошибку храним в заказе, а не только в сообщении на экране:
        # человек уйдёт со страницы, а причина отказа понадобится потом
        order.invoice_error = str(exc)[:500]
        order.save(update_fields=['invoice_error'])
        messages.error(request, f'Счёт не выставлен. {exc}')
        return redirect('repair_order_invoice', pk=order.pk)

    order.invoice_number = form.cleaned_data['invoice_number']
    order.invoice_date = form.cleaned_data['invoice_date']
    order.invoice_sent_at = timezone.now()
    order.invoice_pdf_url = provider.invoice_pdf_url(response)
    order.invoice_external_id = provider.external_id(response)
    order.invoice_provider = chosen
    order.invoice_error = ''
    order.save(update_fields=[
        'invoice_number', 'invoice_date', 'invoice_sent_at',
        'invoice_pdf_url', 'invoice_external_id', 'invoice_provider',
        'invoice_error',
    ])

    OrderStatusHistory.objects.create(
        order=order,
        changed_by=request.user,
        notes=f'Выставлен счёт № {order.invoice_number} через {provider.label}'
              + (f', отправлен: {", ".join(form.cleaned_data["emails"])}'
                 if form.cleaned_data['emails'] else ' (без отправки по почте)'),
    )

    # Счёт выставлен, а письмо не ушло — это не отказ, и выдавать одно
    # за другое нельзя: счёт в банке уже есть, и повторять его не надо
    for failure in (response.get('emailErrors') or ()) if isinstance(response, dict) else ():
        messages.warning(request, f'Счёт выставлен, но письмо не ушло — {failure}')

    messages.success(request, f'Счёт № {order.invoice_number} выставлен')
    return redirect('repair_order_detail', pk=order.pk)


# ==================== ПОСТУПЛЕНИЯ ИЗ БАНКА ====================

# Выписка только читается, деньги по заказам разносит человек. Автоматика
# здесь ограничена подсказкой: ошибочно разнесённое поступление ищут потом
# неделю, а нажать кнопку — секунда.

@permission_required('bank_statement')
def bank_operations(request):
    """Поступления из выписки Т-Банка и подсказки, к каким они заказам."""
    status = request.GET.get('status', 'new')
    if status not in dict(BankOperation.STATUS_CHOICES):
        status = 'new'

    operations = list(
        BankOperation.objects.filter(status=status)
        .select_related('payment__repair_order', 'processed_by')[:200]
    )
    # Подсказки считаем только для неразнесённых: у разнесённых заказ уже
    # известен, и лишние запросы к базе там ни к чему
    rows = [
        {'operation': operation,
         'suggestions': operation.guess_orders()[:5] if status == 'new' else []}
        for operation in operations
    ]

    counts = dict(
        BankOperation.objects.values_list('status')
        .annotate(total=Count('id')).values_list('status', 'total')
    )

    return render(request, 'core/bank/operations.html', {
        'rows': rows,
        'status': status,
        'tabs': [
            {'value': value, 'label': label, 'count': counts.get(value, 0)}
            for value, label in BankOperation.STATUS_CHOICES
        ],
        'configured': tbank.is_configured(),
        'last_fetch_at': tbank.last_fetch_at(),
    })


@permission_required('bank_statement')
@require_POST
def bank_statement_fetch(request):
    """«Загрузить сейчас» — тот же путь, что у таймера, но по нажатию.

    Промежуток из настроек здесь не спрашивается (как и у ручного запуска
    команды): нажавший кнопку хочет выписку сейчас, а не ждать оставшуюся
    часть промежутка.
    """
    if not tbank.is_configured():
        messages.error(request, 'Т-Банк не настроен: пустой TBANK_TOKEN')
        return redirect('bank_operations')

    try:
        result = tbank.fetch_and_store()
    except tbank.TBankError as exc:
        messages.error(request, f'Выписка не получена: {exc}')
        return redirect('bank_operations')

    messages.success(
        request,
        f'Выписка с {result["date_from"]:%d.%m.%Y} по {result["date_to"]:%d.%m.%Y}: '
        f'операций {result["total"]}, поступлений {result["incoming"]}, '
        f'новых {result["added"]}.'
    )
    return redirect('bank_operations')


@permission_required('bank_statement')
@require_POST
def bank_operation_apply(request, pk):
    """Записать поступление оплатой по выбранному заказу."""
    operation = get_object_or_404(BankOperation, pk=pk)
    if operation.status == 'applied':
        messages.warning(request, 'Это поступление уже разнесено')
        return redirect('bank_operations')

    order = get_object_or_404(RepairOrder, pk=request.POST.get('order') or 0)

    with transaction.atomic():
        payment = Payment.objects.create(
            repair_order=order,
            amount=operation.amount,
            payment_date=operation.operation_date or timezone.localdate(),
            note=_bank_payment_note(operation),
            created_by=request.user,
        )
        operation.payment = payment
        operation.status = 'applied'
        operation.processed_by = request.user
        operation.processed_at = timezone.now()
        operation.save(update_fields=['payment', 'status', 'processed_by', 'processed_at'])

    order.refresh_from_db()
    OrderStatusHistory.objects.create(
        order=order,
        payment_status=order.payment_status,
        changed_by=request.user,
        notes=f'Разнесено поступление из банка {operation.amount_text} ₽'
              + (f' от {operation.counterparty}' if operation.counterparty else ''),
    )

    messages.success(
        request,
        f'{operation.amount_text} ₽ разнесены на {order.order_number}. '
        + (f'Остаток: {order.debt} ₽' if order.debt else 'Заказ оплачен полностью')
    )
    return redirect('bank_operations')


@permission_required('bank_statement')
@require_POST
def bank_operation_skip(request, pk):
    """Пометить поступление как не относящееся к заказам.

    Возвраты, переводы между своими счетами, поступления не за ремонт —
    их не разносят, но и висеть в списке они не должны.
    """
    operation = get_object_or_404(BankOperation, pk=pk)
    if operation.status == 'applied':
        messages.error(request, 'Сначала отмените разнесение')
        return redirect('bank_operations')

    operation.status = 'skipped'
    operation.processed_by = request.user
    operation.processed_at = timezone.now()
    operation.save(update_fields=['status', 'processed_by', 'processed_at'])
    messages.success(request, 'Поступление убрано из списка')
    return redirect('bank_operations')


@permission_required('bank_statement')
@require_POST
def bank_operation_reset(request, pk):
    """Вернуть поступление в неразнесённые.

    Созданная оплата при этом удаляется — иначе деньги остались бы
    по заказу и одновременно ждали разнесения.
    """
    operation = get_object_or_404(BankOperation, pk=pk)
    payment = operation.payment
    if payment is not None:
        # Оплата уходит — сигнал pre_delete сам вернёт поступление в «новые»
        payment.delete()
        messages.success(request, 'Разнесение отменено, оплата по заказу удалена')
    else:
        operation.status = 'new'
        operation.processed_by = None
        operation.processed_at = None
        operation.save(update_fields=['status', 'processed_by', 'processed_at'])
        messages.success(request, 'Поступление возвращено в список')
    return redirect('bank_operations')


def _bank_payment_note(operation):
    """Примечание к оплате: откуда деньги. Обрезано под длину поля."""
    parts = ['Выписка Т-Банка']
    if operation.document_number:
        parts.append(f'п/п {operation.document_number}')
    if operation.counterparty:
        parts.append(operation.counterparty)
    return ', '.join(parts)[:255]


@permission_required('admin_access')
def admin_organization(request):
    """Справочник своих юрлиц: реквизиты, подписи, банк для счетов.

    Записей может быть несколько — бухгалтеров двое, и работают они
    от разных юрлиц. Без указания, какую правим, открывается основная:
    так же, как открывалась единственная запись до v2.50.0.
    """
    organization = _chosen_organization(request)
    if request.method == 'POST':
        form = OrganizationForm(request.POST, instance=organization)
        if form.is_valid():
            saved = form.save()
            messages.success(request, 'Реквизиты сохранены')
            return redirect(f"{reverse('admin_organization')}?id={saved.pk}")
        messages.error(request, 'Реквизиты не сохранены: проверьте отмеченные поля')
    else:
        form = OrganizationForm(instance=organization)

    return render(request, 'core/admin/organization.html', {
        'form': form,
        'organization': organization,
        'organizations': list(Organization.objects.all()),
        'is_new': organization.pk is None,
    })


def _chosen_organization(request):
    """Какое юрлицо правим. Без указания — основное, как было раньше."""
    if request.GET.get('new'):
        return Organization()
    chosen = request.GET.get('id') or request.POST.get('id')
    if chosen:
        return get_object_or_404(Organization, pk=chosen)
    return Organization.get_solo()


# ==================== КОРОТКИЕ АДРЕСА ДЛЯ QR-КОДОВ ====================

# Длина ссылки определяет размер QR: каждый лишний десяток символов —
# это следующая версия кода, больше модулей и мельче каждый из них.
# При печати 12 мм разница между 25 и 29 модулями означает 3,8 против
# 3,3 точек принтера на модуль, то есть уверенное чтение или пограничное.

@login_required
def short_part(request, pk):
    """Короткий адрес детали для QR: /p/<id>/"""
    get_object_or_404(SparePart, pk=pk)
    return redirect('part_detail', pk=pk)


@login_required
def short_cell(request, pk):
    """Короткий адрес ячейки для QR: /c/<id>/
    Открывает сетку на нужной кассетнице и сразу показывает содержимое ячейки."""
    cell = get_object_or_404(StorageCell, pk=pk)
    return redirect(f"{reverse('storage_cell_grid')}?cabinet={cell.cabinet.number}&open_cell={cell.pk}")


@login_required
def short_order(request, pk):
    """Короткий адрес заказа для QR: /o/<id>/

    Ведёт в карточку заказа. Номер позиции в ссылку не входит: он крупно
    напечатан на самой этикетке, а каждый лишний символ — это модули кода,
    которых на 43x25 мм и так впритык.
    """
    get_object_or_404(RepairOrder, pk=pk)
    return redirect('repair_order_detail', pk=pk)


@login_required
def short_order_equipment(request, pk):
    """Короткий адрес единицы оборудования в заказе: /u/<id>/

    Это код с наклейки на самом приборе, и ведёт он на страницу этой
    единицы: у мастера в руках плата, и ему нужно всё про неё — что
    привезли, что нашли, что сделано, какие детали ушли, где документы.

    До v2.61.0 наклейка вела на заказ целиком, а какая это единица,
    подсказывал только номер позиции рядом: в код помещался адрес
    сервера, и на позицию знаков не оставалось. С v2.61.0 она вела
    в карточку заказа к нужной строке — тоже не от хорошей жизни:
    страницы единицы тогда просто не было.
    """
    roe = get_object_or_404(
        RepairOrderEquipment.objects.select_related('repair_order'), pk=pk
    )
    return redirect('repair_order_unit_detail',
                    order_pk=roe.repair_order_id, roe_pk=roe.pk)


@login_required
def short_equipment(request, pk):
    """Короткий адрес единицы оборудования для QR: /e/<id>/

    Ведёт в историю ремонтов, а не в карточку редактирования: этикетку
    сканируют, когда железка в руках и нужно вспомнить, что с ней уже
    делали и на гарантии ли она. Редактировать её в этот момент незачем.
    """
    get_object_or_404(Equipment, pk=pk)
    return redirect('equipment_history', pk=pk)


# ==================== ПРАЙСЫ ====================


@login_required
def price_list_index(request):
    """Все прайсы: базовый и по заказчикам."""
    lists = (
        PriceList.objects.select_related('client')
        .annotate(lines_count=Count('lines'))
        .order_by('client__name')
    )
    base = next((p for p in lists if p.is_base), None)
    return render(request, 'core/prices/index.html', {
        'base': base,
        'client_lists': [p for p in lists if not p.is_base],
        'clients_without_price': Client.objects.filter(price_list__isnull=True).order_by('name'),
    })


def _price_list_copy_source(request):
    """Прайс-образец для `?copy_from=<номер>` — или None.

    Тот же приём, что у типовых неисправностей и карточек деталей:
    образец только заполняет форму, и до нажатия «Сохранить» в базе
    не появляется ничего.
    """
    copy_from = request.GET.get('copy_from')
    if not copy_from:
        return None
    return PriceList.objects.filter(pk=copy_from).first()


def _price_list_back_url(request):
    """Куда вернуться из формы прайса.

    `?next=` приходит с карточки заказчика, когда прайс завели или
    открыли прямо оттуда — без него список прайсов, как и раньше.
    """
    next_url = request.GET.get('next') or request.POST.get('next') or ''
    if not url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        next_url = ''
    return next_url or reverse('price_list_index'), next_url


@login_required
def price_list_create(request):
    back_url, next_url = _price_list_back_url(request)
    if request.method == 'POST':
        form = PriceListForm(request.POST)
        formset = PriceListLineFormSet(request.POST, prefix='lines')
        if form.is_valid() and formset.is_valid():
            try:
                price_list = form.save()
            except ValidationError as error:
                messages.error(request, '; '.join(error.messages))
            else:
                formset.instance = price_list
                formset.save()
                messages.success(request, f'{price_list} сохранён')
                return redirect('price_list_edit', pk=price_list.pk)
        else:
            messages.error(request, 'Прайс не сохранён: проверьте отмеченные поля')
    else:
        source = _price_list_copy_source(request)
        initial_client = request.GET.get('client') or None
        form = PriceListForm(initial={'client': initial_client,
                                      'note': source.note if source else ''})
        lines = [
            {'equipment_type': line.equipment_type_id,
             'complexity': line.complexity,
             'price': line.price}
            for line in source.lines.all()
        ] if source else []
        formset = PriceListLineFormSet(
            prefix='lines', queryset=PriceListLine.objects.none(), initial=lines
        )
        formset.extra = len(lines) + 1
    return render(request, 'core/prices/form.html', {
        'form': form, 'formset': formset, 'title': 'Новый прайс',
        'copy_source': _price_list_copy_source(request),
        'back_url': back_url, 'next_url': next_url,
    })


@login_required
def price_list_edit(request, pk):
    price_list = get_object_or_404(PriceList.objects.select_related('client'), pk=pk)
    back_url, next_url = _price_list_back_url(request)
    if request.method == 'POST':
        form = PriceListForm(request.POST, instance=price_list)
        formset = PriceListLineFormSet(request.POST, instance=price_list, prefix='lines')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f'{price_list} сохранён')
            return redirect('price_list_edit', pk=price_list.pk)
        messages.error(request, 'Изменения не сохранены: проверьте отмеченные поля')
    else:
        form = PriceListForm(instance=price_list)
        formset = PriceListLineFormSet(instance=price_list, prefix='lines')
    return render(request, 'core/prices/form.html', {
        'form': form, 'formset': formset, 'price_list': price_list,
        'back_url': back_url, 'next_url': next_url,
        'title': str(price_list),
    })


@login_required
def price_list_delete(request, pk):
    price_list = get_object_or_404(PriceList, pk=pk)
    if request.method == 'POST':
        name = str(price_list)
        price_list.delete()
        messages.success(request, f'{name} удалён')
        return redirect('price_list_index')
    return render(request, 'core/prices/delete.html', {'price_list': price_list})


# ==================== СКАНИРОВАНИЕ ====================

# Сканер — это клавиатура: он «набирает» содержимое кода и жмёт Enter.
# Ловит это `core/static/js/scanner.js`, разбирает — `core/scanning.py`,
# а здесь только страница «что это за коробка на полке» и ответ на вопрос
# «что лежит за этим кодом».
#
# Съёмка камерой браузера здесь невозможна и не задумана: камеру браузер
# отдаёт только в защищённом окружении (HTTPS), а программа открывается
# по обычному http на имени в Tailscale.


@login_required
def scan_page(request):
    """Страница «Сканирование»: поднести любой код и увидеть, что это.

    Нужна не для действия, а для ответа на вопрос: коробка с полки,
    наклейка на пакете, приехавший прибор — что это и что с ним делать.
    Разобранные коды копятся списком тут же, чтобы проверить несколько
    штук подряд, не теряя предыдущие ответы.
    """
    return render(request, 'core/scan.html')


def _scan_part(part):
    cell = part.current_cell
    return {
        # Где деталь лежит — отдельными полями, а не только строкой для
        # человека: по ним сетка кассетниц открывает нужную кассетницу
        # и подсвечивает ячейку, не заставляя искать её глазами. Тот же
        # приём, что `equipment_id` у оборудования.
        'cell_id': cell.pk if cell else None,
        'cabinet_number': cell.cabinet.number if cell else None,
        'title': part.part_number,
        'subtitle': part.name,
        'lines': [
            {'label': 'Характеристики', 'value': _label_specs(part)},
            {'label': 'Корпус', 'value': part.package},
            {'label': 'Остаток', 'value': f'{part.current_stock} шт (минимум {part.min_stock})'},
            {'label': 'Ячейка', 'value': cell.address if cell else 'не назначена'},
        ],
        'actions': [
            {'label': 'Карточка детали', 'url': reverse('part_detail', args=[part.pk])},
            {'label': 'Приход', 'url': reverse('part_stock_incoming', args=[part.pk])},
            {'label': 'Расход', 'url': reverse('part_stock_outgoing', args=[part.pk])},
        ],
    }


def _scan_cell(cell):
    parts = list(cell.parts.all())
    return {
        'title': cell.address,
        'subtitle': str(cell.cabinet),
        # Что лежит в ячейке — не только строкой для человека, но и
        # номерами: страница единицы кладёт отсканированную ячейку
        # в список деталей, а для этого ей нужен номер самой детали.
        # Тот же приём, что `cell_id` у детали.
        'parts': [
            {'id': part.pk, 'part_number': part.part_number, 'name': part.name}
            for part in parts
        ],
        'lines': [
            {'label': 'Кассетница', 'value': str(cell.cabinet)},
            {'label': 'Деталей в ячейке', 'value': str(len(parts))},
            {'label': 'Что лежит',
             'value': ', '.join(part.part_number for part in parts) or 'пусто'},
        ],
        'actions': [
            {'label': 'Открыть в кассетнице',
             'url': f"{reverse('storage_cell_grid')}?cabinet={cell.cabinet.number}&open_cell={cell.pk}"},
            {'label': 'Этикетка', 'url': reverse('storage_cell_label', args=[cell.pk])},
        ],
    }


def _scan_equipment(equipment):
    return {
        # Номер самой единицы: по нему экран приёма заказа ставит
        # отсканированное в строку. У кода вида `e` он совпадает с номером
        # в коде, у `u` — нет, поэтому поле есть у обоих.
        'equipment_id': equipment.pk,
        'title': equipment.serial_number,
        'subtitle': equipment.model.full_name,
        'lines': [
            {'label': 'Модель', 'value': equipment.model.name},
            {'label': 'Версия', 'value': equipment.version.name if equipment.version else ''},
            {'label': 'Заказчик',
             'value': equipment.current_client.name if equipment.current_client else 'не указан'},
            {'label': 'Ремонтов', 'value': str(equipment.repair_orders.count())},
        ],
        'actions': [
            {'label': 'История ремонтов', 'url': reverse('equipment_history', args=[equipment.pk])},
            {'label': 'Карточка', 'url': reverse('equipment_edit', args=[equipment.pk])},
        ],
    }


def _scan_order(order):
    return {
        'title': order.order_number,
        'subtitle': order.client.name,
        'lines': [
            {'label': 'Статус', 'value': order.get_status_display()},
            {'label': 'Оплата', 'value': order.get_payment_status_display()},
            {'label': 'Принят', 'value': timezone.localtime(order.date_received).strftime('%d.%m.%Y')},
            {'label': 'Единиц оборудования', 'value': str(order.order_equipments.count())},
        ],
        'actions': [
            {'label': 'Открыть заказ', 'url': reverse('repair_order_detail', args=[order.pk])},
        ],
    }


def _scan_order_equipment(roe):
    """Единица в заказе: что это за прибор и что с ним в этой работе."""
    order = roe.repair_order
    return {
        'equipment_id': roe.equipment_id,
        'title': roe.equipment.designation,
        'subtitle': f'{order.order_number} — {order.client.name}',
        'lines': [
            {'label': 'Серийный номер', 'value': roe.equipment.serial_number},
            {'label': 'Неисправность', 'value': roe.fault_description},
            {'label': 'Выполнено', 'value': roe.work_performed},
            {'label': 'Статус заказа', 'value': order.get_status_display()},
        ],
        'actions': [
            {'label': 'Открыть в заказе',
             'url': reverse('repair_order_detail', args=[order.pk]) + f'#unit-{roe.pk}'},
            {'label': 'История этой единицы',
             'url': reverse('equipment_history', args=[roe.equipment_id])},
        ],
    }


# Что искать и чем описывать — по видам кода. Модель и сборщик рядом,
# чтобы новый вид добавлялся одной строкой, а не правкой в трёх местах.
_SCAN_KINDS = {
    'part': (SparePart, _scan_part),
    'cell': (StorageCell, _scan_cell),
    'equipment': (Equipment, _scan_equipment),
    'order': (RepairOrder, _scan_order),
    'order_equipment': (RepairOrderEquipment, _scan_order_equipment),
}


@login_required
def scan_resolve(request):
    """Что стоит за отсканированным кодом — в JSON, для страницы сканирования.

    Разбор — общий (`core/scanning.py`), тот же, что проверяется тестами:
    вид кода определяется видом пути, а не адресом сервера. Наклейка могла
    быть напечатана с другой основой (`LABEL_BASE_URL`), и читаться она
    обязана всё равно.

    Неузнанный код — это ответ «не наш», а не догадка: неверно угаданный
    номер открыл бы чужую карточку, и человек со сканером в руках этого
    не заметил бы.
    """
    payload = (request.GET.get('code') or '').strip()
    scan = scanning.decode(payload)

    if not scan:
        return JsonResponse({
            'recognized': False,
            'payload': payload[:200],
            'message': 'Это не код LiftTeam. Подойдут наклейки программы — '
                       'деталь, ячейка, оборудование, заказ.',
        })

    model, describe = _SCAN_KINDS[scan['kind']]
    obj = model.objects.filter(pk=scan['id']).first()
    label = scanning.kind_label(scan['kind'])

    if obj is None:
        return JsonResponse({
            'recognized': True,
            'found': False,
            'kind': scan['kind'],
            'kind_label': label,
            'id': scan['id'],
            'message': f'{label} №{scan["id"]} в программе не найдена — '
                       f'запись удалили или наклейка от другой базы.',
        })

    payload_data = describe(obj)
    payload_data.update({
        'recognized': True,
        'found': True,
        'kind': scan['kind'],
        'kind_label': label,
        'id': scan['id'],
    })
    # Пустые строки не показываем: «Версия: —» ничего не добавляет
    payload_data['lines'] = [line for line in payload_data['lines'] if line['value']]
    return JsonResponse(payload_data)


def label_base_url(request):
    """Основа для ссылок в QR-кодах.

    Берётся из LABEL_BASE_URL; на Raspberry Pi там прошит адрес в Tailscale.
    Смысл настройки в том, чтобы код не зависел от того, откуда открыли
    страницу печати: иначе две одинаковые с виду наклейки вели бы в разные
    места — напечатанная из офиса на локальный адрес, напечатанная снаружи
    на адрес Tailscale.

    Пустое значение возвращает прежнее поведение: адрес берётся из запроса.
    """
    configured = envfile.setting('LABEL_BASE_URL', '')
    return configured.rstrip('/') if configured else request.build_absolute_uri('/').rstrip('/')


def qr_payload(prefix, pk):
    """То, что лежит внутри QR-кода: `u/123`, `p/45`.

    Адреса сервера здесь нет намеренно. Сканируют, когда программа уже
    открыта, и адрес в коде не нужен ей ни для чего — зато он занимал
    три четверти содержимого: `http://lifteam.taile9b605.ts.net/o/123` это
    38 знаков против пяти. На этикетке заказа, где код всего 9,57 мм, это
    разница между 29 модулями (2,6 точки принтера на модуль, впритык
    к пределу читаемости) и 21 модулем (3,6 точки, спокойный запас).

    Второе следствие важнее первого: внутренний адрес программы больше
    не печатается на наклейках, которые уезжают к заказчику.

    Разбор со сканера (`core/scanning.py`) принимает и то и другое, так
    что наклейки, напечатанные раньше, читаются по-прежнему.
    """
    return f'{prefix}/{pk}'


def qr_link(base_url, prefix, pk):
    """Полный адрес — для человека, а не для кода.

    Показывается под кодом на странице печати, чтобы ошибку в адресе было
    видно на экране, а не со сканером у стеллажа. В сам код не уходит.
    """
    return f'{base_url}/{qr_payload(prefix, pk)}'


# Сколько знаков помещается в QR, не переводя его на следующую версию.
# Границы при уровне коррекции M: до 14 знаков — 21 модуль, 15–26 — 25,
# 27–42 — 29. Содержимое теперь короткое (`u/123` — пять знаков), и в
# 21 модуль оно укладывается с запасом на пятизначные номера. Проверка
# оставлена сторожем на будущее: если в код когда-нибудь решат положить
# что-то ещё, беда всплывёт до печати сотни наклеек, а не после.
QR_MAX_CHARS = 14


def qr_length_warning(payloads):
    """Предупреждение, если содержимое не помещается в QR нужного размера.

    Считается по факту, а не по настройке: номер записи растёт сам собой.
    Сказать об этом надо до печати — после того как этикетки наклеены
    на сотню пакетов, чинить нечего.
    """
    longest = max((str(payload) for payload in payloads), key=len, default='')
    if len(longest) <= QR_MAX_CHARS:
        return ''
    return (
        f'Содержимое QR длиннее {QR_MAX_CHARS} знаков ({len(longest)}): {longest}. '
        'Код станет мельче и может не читаться сканером — особенно на этикетке '
        'заказа, где он всего 9,6 мм.'
    )


def _part_label(part, base_url):
    """Данные одной этикетки детали.

    Набор полей общий с этикеткой ячейки (`_cell_label`) — печатаются они
    одним шаблоном. Общий он должен быть **целиком**, включая поля,
    которых у детали не бывает: пачка этикеток передаёт их в шаблон
    поимённо (`items=label.items`), а у словаря Python есть собственный
    метод `items()`. Не окажись ключа — шаблонизатор нашёл бы метод,
    вызвал его и напечатал на наклейке весь словарь: «('part',
    <SparePart: …>), ('qr_payload', 'p/3473')…». Ровно это и печаталось
    до v2.93.0 на каждой этикетке детали, отправленной пачкой.
    """
    payload = qr_payload('p', part.pk)
    link = qr_link(base_url, 'p', part.pk)
    cell = part.current_cell
    return {
        'part': part,
        'cell': cell,
        # Перечня у детали нет: он бывает только у ячейки с несколькими
        # деталями. Но ключ обязан быть — почему, сказано выше
        'items': (),
        'title': part.part_number,
        'specs': _label_specs(part),
        'description': part.label_text,
        'package': part.package,
        'application': part.application,
        'address': cell.address if cell else 'нет ячейки',
        'qr_url': link,
        'qr_payload': payload,
        'qr_img': generate_qr_image(payload),
    }


@login_required
def part_label(request, pk):
    """Этикетка детали — для наклейки на пакет."""
    part = get_object_or_404(SparePart, pk=pk)
    base_url = label_base_url(request)
    context = _part_label(part, base_url)
    context['qr_base'] = base_url
    context['qr_warning'] = qr_length_warning([context['qr_payload']])
    return render(request, 'core/parts/label.html', context)


# Верхняя граница пачки. Столько этикеток — уже полтора метра ленты; больше
# за один раз не печатают, а страница с тысячей QR-картинок открывалась бы
# на планшете минуту и могла не влезть в память.
MAX_LABELS_PER_BATCH = 200


def _batch_layout(request):
    """Раскладка пачки: рулон (по этикетке на страницу) или лист A4."""
    return 'a4' if request.GET.get('layout') == 'a4' else 'roll'


def _selected_ids(request):
    """Отмеченные записи из формы списка."""
    return [value for value in request.GET.getlist('ids') if value.isdigit()]


@login_required
def part_labels_batch(request):
    """Пачка этикеток деталей: по отмеченным или по всему текущему отбору."""
    ids = _selected_ids(request)
    if ids:
        parts = SparePart.objects.filter(pk__in=ids)
    else:
        parts, _ = _filter_parts(request.GET)

    parts = list(parts.prefetch_related('storage_cells').order_by('part_number')[:MAX_LABELS_PER_BATCH])
    base_url = label_base_url(request)

    labels = [_part_label(part, base_url) for part in parts]

    return render(request, 'core/parts/labels_batch.html', {
        'labels': labels,
        'layout': _batch_layout(request),
        'limit': MAX_LABELS_PER_BATCH,
        'qr_base': base_url,
        'qr_warning': qr_length_warning(label['qr_payload'] for label in labels),
    })


@login_required
def storage_cell_labels_batch(request):
    """Пачка этикеток ячеек: по отмеченным, либо по кассетнице целиком.

    Кассетница целиком — обычный случай: этикетки клеят на все ячейки сразу,
    когда собирают новую или переклеивают старые.
    """
    ids = _selected_ids(request)
    cells = StorageCell.objects.select_related('cabinet').prefetch_related('parts')

    if ids:
        cells = cells.filter(pk__in=ids)
    else:
        cabinet = request.GET.get('cabinet', '')
        cells = cells.filter(cabinet__number=int(cabinet)) if cabinet.isdigit() else cells.none()

    # Пустые ячейки обычно подписывать незачем: адрес и так виден в сетке,
    # а лента тратится
    only_filled = request.GET.get('only_filled') == '1'
    cells = list(cells.order_by('cabinet__number', 'row_number', 'cell_row')[:MAX_LABELS_PER_BATCH])
    if only_filled:
        cells = [cell for cell in cells if cell.parts.all()]

    base_url = label_base_url(request)

    labels = [_cell_label(cell, base_url) for cell in cells]

    return render(request, 'core/storage_cells/labels_batch.html', {
        'labels': labels,
        'qr_base': base_url,
        'qr_warning': qr_length_warning(label['qr_payload'] for label in labels),
        'layout': _batch_layout(request),
        'only_filled': only_filled,
        'cabinet': request.GET.get('cabinet', ''),
        'limit': MAX_LABELS_PER_BATCH,
    })
