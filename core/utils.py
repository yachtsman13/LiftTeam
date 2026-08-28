"""
Утилиты для LiftTeam v2.59.0.
"""
import barcode
from barcode.writer import ImageWriter
from datetime import date, datetime
from io import BytesIO
from urllib.parse import quote
import base64
import json

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def excel_datetime(value):
    """Дата и время в виде, пригодном для Excel.

    Excel не хранит часовой пояс, и openpyxl отказывается записывать значение
    с ним. Переводим в местное время и снимаем пояс — иначе выгрузка падала бы
    на первой же строке журнала движений.
    """
    if value is None:
        return None
    if timezone.is_aware(value):
        value = timezone.localtime(value)
        value = value.replace(tzinfo=None)
    # Доли секунды в журнале движений не нужны и только удлиняют ячейку
    return value.replace(microsecond=0)


def build_workbook(sheet_title, headers, rows):
    """Книга Excel с одним листом: жирная шапка, закреплённая при прокрутке,
    ширина колонок по содержимому."""
    wb = openpyxl.Workbook()
    add_sheet(wb, sheet_title, headers, rows, ws=wb.active)
    return wb


def add_sheet(wb, sheet_title, headers, rows, ws=None):
    """Ещё один лист в уже существующей книге, с тем же оформлением, что
    и у листа из `build_workbook` (жирная шапка, закреплена при прокрутке,
    ширина колонок по содержимому).

    Нужен, когда одна выгрузка — это несколько таблиц сразу с разными
    столбцами (например, аналитика ремонта: по инженерам, по типу
    неисправности, по загрузке — свести их в одну таблицу нельзя, у них
    разный состав колонок).

    `ws` — использовать конкретный лист вместо создания нового; так
    `build_workbook` заполняет самый первый лист книги, а не добавляет
    лишний пустой.
    """
    if ws is None:
        ws = wb.create_sheet()
    # Excel не принимает названия листов длиннее 31 символа
    ws.title = sheet_title[:31]

    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'

    for row in rows:
        ws.append(list(row))

    # Даты в привычном виде: без этого Excel показывает их по-американски
    # (2026-08-12 06:19:30), а отчёт читают в офисе
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD.MM.YYYY HH:MM'
            elif isinstance(cell.value, date):
                cell.number_format = 'DD.MM.YYYY'

    for index, column in enumerate(ws.columns, start=1):
        longest = max(
            (len(str(cell.value)) for cell in column if cell.value is not None),
            default=0,
        )
        # Нижняя граница — чтобы заголовок не сжимался, верхняя — чтобы
        # длинное примечание не растягивало колонку на весь экран
        ws.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), 50)

    return ws


def xlsx_response(workbook, filename):
    """HTTP-ответ с книгой Excel.

    Имя файла передаётся дважды: в ASCII-виде для старых клиентов и в кодировке
    UTF-8 по RFC 5987 — иначе русское название сохраняется как набор символов.
    """
    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    encoded = quote(filename)
    response['Content-Disposition'] = (
        f'attachment; filename="{encoded}"; filename*=UTF-8\'\'{encoded}'
    )
    workbook.save(response)
    return response


def generate_barcode_image(text, width=300, height=100):
    """Генерация штрихкода Code 128 из текста адреса ячейки.
    Оптимизировано для этикеток 43x25 мм (термопринтер XP-365)."""
    try:
        code128 = barcode.get_barcode_class('code128')
        barcode_obj = code128(text, writer=ImageWriter())
        buffer = BytesIO()
        barcode_obj.write(buffer, options={
            'write_text': True,
            'module_width': 0.25,
            'module_height': 12,
            'font_size': 9,
            'text_distance': 2,
            'quiet_zone': 2,
        })
        buffer.seek(0)
        img_base64 = base64.b64encode(buffer.read()).decode('utf-8')
        return f"data:image/png;base64,{img_base64}"
    except Exception:
        return None


def generate_qr_image(data, size=200):
    """QR-код для этикетки 43×25 мм.

    Уровень восстановления — **H**, самый высокий: код читается,
    даже если повреждено около 30 % его площади. Это не запас
    на всякий случай, а прямой ответ на то, как живут наклейки
    на оборудовании: в машинном отделении прибор задевают, трут
    и ставят на него что попало.

    Взято даром. Содержимое кода короткое (`u/123`, `p/42`, `c/7` —
    с v2.61.0 адреса сервера в нём нет), и при переходе с прежнего
    уровня M на H размер не изменился вовсе: как был 21 модуль,
    так и остался. Платить пришлось бы только за длинное содержимое:
    полная ссылка выросла бы с 29 модулей до 37, то есть перестала бы
    помещаться. За этим следит `QrErrorCorrectionTests`.
    """
    try:
        import qrcode
        from qrcode.image.styledpil import StyledPilImage

        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=6,
            border=1,
        )
        qr.add_data(data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        img_base64 = base64.b64encode(buffer.read()).decode('utf-8')
        return f"data:image/png;base64,{img_base64}"
    except ImportError:
        # Fallback: штрихкод вместо QR
        return generate_barcode_image(data[:50])
    except Exception:
        return None




