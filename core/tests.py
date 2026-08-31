"""
Тесты для LiftTeam. Покрывают наиболее рискованную бизнес-логику:
генерацию номера заказа, движение склада, ячейки с несколькими деталями,
импорт/экспорт Excel, права доступа по ролям, маршрутизацию,
настройку SQLite и резервное копирование.
"""
import base64
import datetime
import io
import os
import stat
from decimal import Decimal
import json
import re
import sqlite3
import tempfile
from pathlib import Path
from unittest.mock import patch
from urllib import error as urllib_error
from urllib import parse as urllib_parse

import openpyxl
from channels.db import database_sync_to_async
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError, connection, transaction
from django.test import SimpleTestCase, TestCase, TransactionTestCase, override_settings
from django.test import Client as TestClient
from django.test.utils import CaptureQueriesContext
from django.test import RequestFactory
from django.template.loader import render_to_string
from django.test.utils import CaptureQueriesContext
from django.urls import resolve, reverse
from django.utils import timezone

from django import forms

from . import (
    envfile, invoicing, messengers, net, notifications, restarter, scanning,
    selfcheck, tbank, tochka, views, webhooks, yadisk,
)
from .forms import OrganizationForm, RepairOrderEquipmentForm
from django.core.files.uploadedfile import SimpleUploadedFile

from .images import MAX_UPLOAD_BYTES
from .utils import generate_qr_image
from .forms import UnitEditForm
from .models import (
    BankOperation, Cabinet, Client as ClientModel, Employee, Equipment, EquipmentModel,
    PriceList, PriceListLine, available_equipment_for_order,
    EquipmentMaterial, EquipmentType, EquipmentVersion,
    FaultType, FaultTypePart, InventorySession, InventorySessionLine, Notification, Organization,
    OrderCost, OrderStatusHistory, Payment,
    RepairOrder, RepairOrderDetail, RepairOrderEquipment, SparePart, StockAllocation, StockMovement,
    SettingChange, StorageCell, TechCard, TechCardStep, WebhookDelivery,
    complexity_css,
    parse_layout, plural_genitive, format_spec,
)


class FakeChannelLayer:
    """Подменяет Redis/InMemory channel layer, чтобы посчитать реальное число
    отправленных WebSocket-уведомлений (регрессия дублирования, см. signals.py/views.py)."""
    def __init__(self):
        self.calls = []

    async def group_send(self, group, message):
        self.calls.append((group, message))


class OrderNumberGenerationTests(TestCase):
    def setUp(self):
        self.client_obj = ClientModel.objects.create(name='Заказчик 1')

    def test_format_and_uniqueness(self):
        order1 = RepairOrder.objects.create(client=self.client_obj)
        order2 = RepairOrder.objects.create(client=self.client_obj)

        self.assertRegex(order1.order_number, r'^LT-\d{4}-\d{2}-\d{3}$')
        self.assertNotEqual(order1.order_number, order2.order_number)

        seq1 = int(order1.order_number.rsplit('-', 1)[-1])
        seq2 = int(order2.order_number.rsplit('-', 1)[-1])
        self.assertEqual(seq2, seq1 + 1)


class StockMovementTests(TestCase):
    def setUp(self):
        self.user = Employee.objects.create_user(
            username='warehouse1', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.part = SparePart.objects.create(
            part_number='P-001', name='Тестовая деталь', current_stock=10, min_stock=2
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.user)

    def test_incoming_updates_stock_and_sends_single_notification(self):
        fake_layer = FakeChannelLayer()
        with patch('core.views.get_channel_layer', return_value=fake_layer), \
             patch('core.signals.get_channel_layer', return_value=fake_layer):
            self.client_http.post(
                f'/parts/{self.part.pk}/stock-incoming/',
                {'quantity': 5, 'document_number': '', 'notes': ''}
            )

        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 15)
        self.assertEqual(self.part.movements.count(), 1)
        # Регрессия: раньше уведомление слалось дважды (сигнал + ручной вызов)
        self.assertEqual(len(fake_layer.calls), 1)

    def test_outgoing_updates_stock_and_sends_single_notification(self):
        fake_layer = FakeChannelLayer()
        with patch('core.views.get_channel_layer', return_value=fake_layer), \
             patch('core.signals.get_channel_layer', return_value=fake_layer):
            self.client_http.post(
                f'/parts/{self.part.pk}/stock-outgoing/',
                {'quantity': 3, 'reason': 'consumption', 'notes': '', 'document_number': ''}
            )

        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 7)
        self.assertEqual(len(fake_layer.calls), 1)


class RepairOrderAddDetailTests(TestCase):
    def setUp(self):
        self.user = Employee.objects.create_user(
            username='manager1', full_name='Менеджер', password='pass', role='repair_manager'
        )
        self.client_obj = ClientModel.objects.create(name='Заказчик 2')
        self.order = RepairOrder.objects.create(client=self.client_obj)
        self.part = SparePart.objects.create(part_number='P-002', name='Деталь 2', current_stock=10, min_stock=2)
        self.client_http = TestClient()
        self.client_http.force_login(self.user)

    def test_add_detail_decrements_stock_and_logs_history(self):
        self.client_http.post(
            f'/repair-orders/{self.order.pk}/add-detail/',
            {'part': self.part.pk, 'quantity_used': 4}
        )
        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 6)
        self.assertEqual(self.order.details.count(), 1)
        self.assertTrue(self.order.status_history.filter(notes__icontains='Добавлена деталь').exists())


class RepairOrderFormErrorTests(TestCase):
    """Регрессия: форма молча не сохранялась. Ошибки полей «Стоимость ремонта»
    и «Папка на Яндекс.Диске» не выводились нигде — страница перезагружалась
    без единого сообщения, и пользователь считал заказ созданным.

    С v2.57.0 этих двух полей на приёме нет вовсе: стоимость считают после
    диагностики, папку заводят по ходу работы. Проверка переехала туда, где
    поля теперь живут, — на страницу редактирования. Она и была про то, что
    ошибка поля видна, а не про то, на какой она странице.
    """

    def setUp(self):
        self.user = Employee.objects.create_superuser(
            username='order_form', full_name='Тест', password='pass'
        )
        self.client_obj = ClientModel.objects.create(name='Заказчик формы')
        self.model = EquipmentModel.objects.create(name='Модель формы')
        self.equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-FORM-1'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.user)

    def _post(self, **overrides):
        """Приём: полей работы в форме нет, и посылать их незачем."""
        data = {
            'client': self.client_obj.pk,
            'fault_description': '',
            'equipments-TOTAL_FORMS': '1',
            'equipments-INITIAL_FORMS': '0',
            'equipments-MIN_NUM_FORMS': '0',
            'equipments-MAX_NUM_FORMS': '1000',
            'equipments-0-equipment': '',
            'equipments-0-fault_description': '',
            'equipments-0-initial_condition': '',
        }
        data.update(overrides)
        return self.client_http.post('/repair-orders/create/', data)

    def _order_with_equipment(self):
        order = RepairOrder.objects.create(client=self.client_obj)
        oe = RepairOrderEquipment.objects.create(
            repair_order=order, equipment=self.equipment
        )
        return order, oe

    def _post_unit(self, order, oe, **overrides):
        """Поля прибора правятся на его странице: страницы правки заказа
        больше нет, её работу забрали карточка заказа и страница единицы."""
        data = {
            'fault_description': '',
            'initial_condition': '',
            'work_performed': '',
            'seal_numbers': '',
            'repair_cost': '',
            'yandex_disk_folder': '',
        }
        data.update(overrides)
        return self.client_http.post(
            reverse('repair_order_unit_edit', args=[order.pk, oe.pk]),
            data, follow=True,
        )

    def test_valid_submission_creates_order(self):
        response = self._post()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(RepairOrder.objects.count(), 1)

    def test_invalid_repair_cost_reports_error_and_saves_nothing(self):
        order, oe = self._order_with_equipment()
        response = self._post_unit(order, oe, repair_cost='15 000 руб')

        oe.refresh_from_db()
        self.assertIsNone(oe.repair_cost)
        texts = [str(m) for m in response.context['messages']]
        self.assertTrue(any('Не сохранено' in text for text in texts), texts)

    def test_invalid_yandex_link_reports_error_and_saves_nothing(self):
        """«Папка на диске» словами открывать нечем, а молча сохранённая
        строка выглядит как рабочая ссылка."""
        order, oe = self._order_with_equipment()
        response = self._post_unit(order, oe, yandex_disk_folder='папка на диске')

        oe.refresh_from_db()
        self.assertEqual(oe.yandex_disk_folder, '')
        texts = [str(m) for m in response.context['messages']]
        self.assertTrue(any('Не сохранено' in text for text in texts), texts)


class RepairOrderIntakeFormTests(TestCase):
    """Приём заказа спрашивает то, что в этот момент известно, и только это.

    Прибор привезли: известно, от кого он, что с ним со слов заказчика
    и в каком он виде приехал. Счёта, оплаты, диагноза, стоимости,
    выполненных работ и номеров пломб в этот момент не существует —
    счёт выставляют после согласования, пломбы ставят при выдаче.
    Ни одно поле при этом не потеряно: у каждого своё место —
    диагноз и оценка на странице дефектации, работы, пломбы
    и стоимость на странице единицы, счёт при его выставлении.
    """

    INTAKE_ONLY = ['invoice_number', 'invoice_date', 'payment_status']
    LATER_PER_UNIT = ['work_performed', 'seal_numbers', 'faults',
                      'repair_cost', 'yandex_disk_folder']

    def setUp(self):
        self.user = Employee.objects.create_superuser(
            username='intake', full_name='Приёмщик', password='pass'
        )
        self.client_obj = ClientModel.objects.create(name='Заказчик приёма')
        self.model = EquipmentModel.objects.create(name='Модель приёма')
        self.equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-INTAKE-1'
        )
        self.http = TestClient()
        self.http.force_login(self.user)

    def test_intake_asks_only_what_is_known_on_arrival(self):
        page = self.http.get('/repair-orders/create/')
        form_fields = list(page.context['form'].fields)
        self.assertEqual(form_fields, ['client'])
        # id, DELETE и repair_order — служебные поля набора форм, их
        # добавляет Django, а не мы.
        unit_fields = [name for name in page.context['formset'].empty_form.fields
                       if name not in ('id', 'DELETE', 'repair_order')]
        self.assertEqual(
            unit_fields, ['equipment', 'fault_description', 'initial_condition']
        )

    def test_later_fields_are_not_on_the_intake_page_at_all(self):
        """Не спрятаны стилями, а не отрисованы: спрятанное поле всё равно
        уходит на сервер и всё равно занимает место в разметке."""
        body = self.http.get('/repair-orders/create/').content.decode()
        for name in self.INTAKE_ONLY:
            with self.subTest(field=name):
                self.assertNotIn('name="%s"' % name, body)
        for name in self.LATER_PER_UNIT:
            with self.subTest(field=name):
                self.assertNotIn('equipments-0-%s' % name, body)
                self.assertNotIn('equipments-__prefix__-%s' % name, body)

    def test_no_field_was_lost_when_intake_was_shortened(self):
        """Урезан приём, а не заказ: каждое поле заполняется позже,
        и у каждого есть своё место. После разбора страницы правки
        заказа (v2.79.0) этих мест три."""
        order = RepairOrder.objects.create(client=self.client_obj)
        roe = RepairOrderEquipment.objects.create(
            repair_order=order, equipment=self.equipment
        )

        # Поля прибора — на его странице
        unit_fields = list(UnitEditForm(instance=roe).fields)
        for name in self.LATER_PER_UNIT:
            with self.subTest(field=name):
                self.assertIn(name, unit_fields)

        # Счёт — при выставлении, статус оплаты — своей формой на карточке
        card = self.http.get(
            reverse('repair_order_detail', args=[order.pk])
        ).content.decode()
        self.assertIn(reverse('repair_order_change_payment_status', args=[order.pk]), card)
        self.assertIn(reverse('repair_order_invoice', args=[order.pk]), card)

    def test_intake_says_where_the_other_fields_went(self):
        """Иначе мастер решит, что поля пропали, и пойдёт их искать."""
        body = self.http.get('/repair-orders/create/').content.decode()

        self.assertIn('на странице дефектации', body)
        self.assertIn('на странице единицы', body)

    def test_intake_saves_the_order_and_the_unit(self):
        response = self.http.post('/repair-orders/create/', {
            'client': self.client_obj.pk,
            'equipments-TOTAL_FORMS': '1',
            'equipments-INITIAL_FORMS': '0',
            'equipments-MIN_NUM_FORMS': '0',
            'equipments-MAX_NUM_FORMS': '1000',
            'equipments-0-equipment': self.equipment.pk,
            'equipments-0-fault_description': 'Гудит и не запускается',
            'equipments-0-initial_condition': 'Корпус целый, пломбы на месте',
        })
        self.assertEqual(response.status_code, 302)
        order = RepairOrder.objects.get()
        # Оплата и счёт остаются в состоянии «ещё не было», а не пустыми
        # по недосмотру: значение берётся из модели, а не из формы.
        self.assertEqual(order.payment_status, 'unpaid')
        self.assertEqual(order.invoice_number, '')
        self.assertIsNone(order.invoice_date)
        unit = order.order_equipments.get()
        self.assertEqual(unit.equipment, self.equipment)
        self.assertEqual(unit.fault_description, 'Гудит и не запускается')
        self.assertEqual(unit.initial_condition, 'Корпус целый, пломбы на месте')

    def test_intake_ignores_fields_it_does_not_ask_for(self):
        """Подсунутое в запрос поле работы на приёме не оседает в базе.

        Форма их не объявляет, значит и не принимает: иначе «урезали»
        значило бы только «убрали с экрана».
        """
        response = self.http.post('/repair-orders/create/', {
            'client': self.client_obj.pk,
            'fault_description': '',
            'payment_status': 'paid',
            'invoice_number': 'СЧ-1',
            'equipments-TOTAL_FORMS': '1',
            'equipments-INITIAL_FORMS': '0',
            'equipments-MIN_NUM_FORMS': '0',
            'equipments-MAX_NUM_FORMS': '1000',
            'equipments-0-equipment': self.equipment.pk,
            'equipments-0-fault_description': '',
            'equipments-0-initial_condition': '',
            'equipments-0-repair_cost': '9999',
            'equipments-0-work_performed': 'Ничего не делали',
        })
        self.assertEqual(response.status_code, 302)
        order = RepairOrder.objects.get()
        self.assertEqual(order.payment_status, 'unpaid')
        self.assertEqual(order.invoice_number, '')
        unit = order.order_equipments.get()
        self.assertIsNone(unit.repair_cost)
        self.assertEqual(unit.work_performed, '')


class StorageCellMultiPartTests(TestCase):
    def setUp(self):
        self.admin = Employee.objects.create_superuser(username='admin_t', full_name='Админ', password='pass')
        self.part1 = SparePart.objects.create(part_number='CELL-1', name='Деталь A')
        self.part2 = SparePart.objects.create(part_number='CELL-2', name='Деталь B')
        self.cell1 = StorageCell.objects.create(cabinet=Cabinet.objects.get_or_create(number=1)[0], row_number=1, cell_row=1)
        self.cell2 = StorageCell.objects.create(cabinet=Cabinet.objects.get_or_create(number=1)[0], row_number=1, cell_row=2)
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def test_cell_holds_multiple_parts(self):
        self.cell1.parts.add(self.part1, self.part2)
        self.assertEqual(self.cell1.parts.count(), 2)

    def test_assign_cell_moves_part_from_previous_cell(self):
        self.cell1.parts.add(self.part1)
        self.client_http.post(f'/parts/{self.part1.pk}/assign-cell/', {'cell_id': self.cell2.pk})

        self.assertEqual(self.cell1.parts.count(), 0)
        self.assertEqual(list(self.cell2.parts.all()), [self.part1])

    def test_assign_cell_does_not_evict_other_parts_from_target(self):
        self.cell2.parts.add(self.part2)
        self.client_http.post(f'/parts/{self.part1.pk}/assign-cell/', {'cell_id': self.cell2.pk})

        self.assertEqual(self.cell2.parts.count(), 2)

    def test_add_part_endpoint(self):
        self.client_http.post(f'/storage-cells/{self.cell1.pk}/add-part/', {'part_id': self.part1.pk})
        self.client_http.post(f'/storage-cells/{self.cell1.pk}/add-part/', {'part_id': self.part2.pk})

        self.assertEqual(self.cell1.parts.count(), 2)

    def test_add_part_evicts_from_previous_cell(self):
        self.cell1.parts.add(self.part1)
        self.client_http.post(f'/storage-cells/{self.cell2.pk}/add-part/', {'part_id': self.part1.pk})

        self.assertEqual(self.cell1.parts.count(), 0)
        self.assertEqual(self.cell2.parts.count(), 1)

    def test_remove_part_endpoint(self):
        self.cell1.parts.add(self.part1)
        self.client_http.post(f'/storage-cells/{self.cell1.pk}/remove-part/', {'part_id': self.part1.pk})

        self.assertEqual(self.cell1.parts.count(), 0)

    def test_move_endpoint_allows_target_cell_with_other_parts(self):
        self.cell1.parts.add(self.part1)
        self.cell2.parts.add(self.part2)

        resp = self.client_http.post('/storage-cells/move/', {
            'from_cell': self.cell1.pk, 'to_cell': self.cell2.pk, 'part_id': self.part1.pk,
        })

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['success'], True)
        self.assertEqual(self.cell2.parts.count(), 2)
        self.assertEqual(self.cell1.parts.count(), 0)


class ExcelImportExportTests(TestCase):
    def setUp(self):
        self.admin = Employee.objects.create_superuser(username='admin_x', full_name='Админ', password='pass')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def _make_xlsx(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['part_number', 'name', 'component_type', 'voltage'])
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'import.xlsx'
        return buf

    def test_import_parses_text_values_with_units(self):
        """Регрессия: строковое значение с единицей измерения ('5 В') раньше
        роняло весь импорт с NameError из-за отсутствующего `import re`."""
        f = self._make_xlsx([['IMP-1', 'Импортированная деталь', 'Резистор', '5 В']])
        resp = self.client_http.post('/parts/import/', {'file': f, 'update_existing': False}, follow=True)

        self.assertEqual(resp.status_code, 200)
        part = SparePart.objects.get(part_number='IMP-1')
        self.assertEqual(float(part.voltage), 5.0)

    def test_import_requires_mandatory_columns(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name'])
        ws.append(['Без артикула'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'bad.xlsx'

        resp = self.client_http.post('/parts/import/', {'file': buf, 'update_existing': False})
        self.assertEqual(SparePart.objects.count(), 0)
        self.assertEqual(resp.status_code, 302)

    def test_export_returns_xlsx_with_header(self):
        SparePart.objects.create(part_number='EXP-1', name='Экспортируемая деталь', voltage=12)

        resp = self.client_http.get('/parts/export/')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual([c.value for c in ws[1]][:2], ['part_number', 'name'])
        self.assertEqual(ws.max_row, 2)

    def test_export_respects_search_filter(self):
        SparePart.objects.create(part_number='EXP-A', name='Совпадёт')
        SparePart.objects.create(part_number='EXP-B', name='Не совпадёт')

        resp = self.client_http.get('/parts/export/?q=EXP-A')
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertEqual(wb.active.max_row, 2)  # заголовок + 1 деталь

    def test_package_survives_import_and_export(self):
        """Выгрузка принимается обратно импортом, поэтому новая колонка
        обязана быть в обеих."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['part_number', 'name', 'component_type', 'package'])
        ws.append(['PKG-1', 'Стабилизатор', 'Микросхема', 'TO-220'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'import.xlsx'

        self.client_http.post('/parts/import/', {'file': buf, 'update_existing': False})
        self.assertEqual(SparePart.objects.get(part_number='PKG-1').package, 'TO-220')

        resp = self.client_http.get('/parts/export/?q=PKG-1')
        exported = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        row = dict(zip([c.value for c in exported[1]], [c.value for c in exported[2]]))
        self.assertEqual(row['package'], 'TO-220')

    def test_import_then_export_round_trip_preserves_values(self):
        f = self._make_xlsx([['RT-1', 'Round trip', 'Диод', '3,3']])
        self.client_http.post('/parts/import/', {'file': f, 'update_existing': False})

        resp = self.client_http.get('/parts/export/?q=RT-1')
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        headers = [c.value for c in wb.active[1]]
        row = dict(zip(headers, [c.value for c in wb.active[2]]))
        self.assertEqual(row['part_number'], 'RT-1')
        self.assertEqual(float(row['voltage']), 3.3)


class RolePermissionTests(TestCase):
    def setUp(self):
        self.warehouse_user = Employee.objects.create_user(
            username='wh1', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.manager_user = Employee.objects.create_user(
            username='rm1', full_name='Менеджер', password='pass', role='repair_manager'
        )
        self.part = SparePart.objects.create(part_number='PERM-1', name='Деталь')

    def test_warehouse_can_delete_part(self):
        client = TestClient()
        client.force_login(self.warehouse_user)
        resp = client.get(f'/parts/{self.part.pk}/delete/')
        self.assertEqual(resp.status_code, 200)

    def test_repair_manager_cannot_delete_part(self):
        client = TestClient()
        client.force_login(self.manager_user)
        resp = client.get(f'/parts/{self.part.pk}/delete/', follow=True)
        self.assertRedirects(resp, '/')

    def test_non_admin_cannot_access_user_management(self):
        client = TestClient()
        client.force_login(self.warehouse_user)
        resp = client.get('/management/users/', follow=True)
        self.assertRedirects(resp, '/')


class EquipmentHistoryTests(TestCase):
    """История ремонтов одной физической единицы по серийному номеру.

    Оборудование уже было выделено в отдельную сущность с уникальным
    серийником, поэтому история собирается связью, а не поиском по строке.
    Открытым оставался разъезд истории из-за разного написания одного
    и того же номера разными сотрудниками."""

    def setUp(self):
        self.user = Employee.objects.create_superuser(
            username='hist_user', full_name='Тест', password='pass'
        )
        self.model = EquipmentModel.objects.create(name='Экодрайв')
        self.client_obj = ClientModel.objects.create(name='ООО Лифт')
        self.equipment = Equipment.objects.create(model=self.model, serial_number='БУАД-1234')
        self.client_http = TestClient()
        self.client_http.force_login(self.user)

    def _make_order(self, equipment=None):
        order = RepairOrder.objects.create(client=self.client_obj)
        RepairOrderEquipment.objects.create(
            repair_order=order, equipment=equipment or self.equipment
        )
        return order

    def test_normalization_ignores_case_and_separators(self):
        for written in ['буад 1234', 'БУАД_1234', ' БУАД-1234 ', 'бУаД1234']:
            self.assertEqual(
                Equipment.normalize_serial(written),
                self.equipment.serial_normalized,
                f'не сошлось на варианте {written!r}',
            )

    def test_normalized_is_filled_on_save(self):
        equipment = Equipment.objects.create(model=self.model, serial_number='ec-99 x')
        self.assertEqual(equipment.serial_normalized, 'EC99X')

    def test_saved_serial_keeps_original_spelling(self):
        """Нормализация нужна для поиска, но печатать надо то, что набрали."""
        equipment = Equipment.objects.create(model=self.model, serial_number='ЭД-77/2')
        equipment.refresh_from_db()
        self.assertEqual(equipment.serial_number, 'ЭД-77/2')

    def test_find_similar_matches_other_spelling(self):
        other = Equipment.objects.create(model=self.model, serial_number='буад 1234!')
        found = Equipment.find_similar('БУАД-1234')
        self.assertIn(other, found)
        self.assertIn(self.equipment, found)

    def test_find_similar_can_exclude_itself(self):
        found = Equipment.find_similar(self.equipment.serial_number, exclude_pk=self.equipment.pk)
        self.assertNotIn(self.equipment, found)

    def test_find_similar_ignores_empty_serial(self):
        self.assertEqual(list(Equipment.find_similar('   ')), [])

    def test_history_page_lists_all_visits(self):
        first = self._make_order()
        second = self._make_order()

        response = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, first.order_number)
        self.assertContains(response, second.order_number)
        self.assertEqual(response.context['visits_count'], 2)

    def test_history_excludes_other_equipment_orders(self):
        other = Equipment.objects.create(model=self.model, serial_number='ДРУГОЙ-1')
        foreign_order = self._make_order(equipment=other)
        self._make_order()

        response = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')

        self.assertNotContains(response, foreign_order.order_number)

    def test_history_marks_parts_as_order_wide_for_multi_unit_orders(self):
        """Детали списываются на заказ целиком. Когда единиц в заказе
        несколько, нельзя утверждать, что деталь ставили именно в эту."""
        order = self._make_order()
        second_unit = Equipment.objects.create(model=self.model, serial_number='ВТОРОЙ-1')
        RepairOrderEquipment.objects.create(repair_order=order, equipment=second_unit)

        response = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')

        self.assertFalse(response.context['visit_rows'][0]['parts_are_exact'])

    def test_history_marks_parts_exact_for_single_unit_order(self):
        self._make_order()
        response = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')
        self.assertTrue(response.context['visit_rows'][0]['parts_are_exact'])

    def test_empty_history_is_not_an_error(self):
        response = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['visits_count'], 0)

    def test_search_finds_equipment_by_different_spelling(self):
        response = self.client_http.get('/equipment/', {'q': 'буад 1234'})
        self.assertContains(response, 'БУАД-1234')

    def test_history_summary_reports_previous_visits(self):
        self._make_order()
        response = self.client_http.get(
            f'/ajax/equipment/{self.equipment.pk}/history-summary/'
        )

        data = response.json()
        self.assertEqual(data['orders_count'], 1)
        self.assertTrue(data['history_url'].endswith(f'/equipment/{self.equipment.pk}/history/'))

    def test_history_summary_is_zero_for_new_equipment(self):
        response = self.client_http.get(
            f'/ajax/equipment/{self.equipment.pk}/history-summary/'
        )
        self.assertEqual(response.json()['orders_count'], 0)

    def test_creating_similar_serial_warns_instead_of_duplicating(self):
        response = self.client_http.post('/ajax/equipment/create/', {
            'model_id': self.model.pk,
            'serial_number': 'буад 1234',
        })

        data = response.json()
        self.assertFalse(data['success'])
        self.assertEqual(len(data['similar']), 1)
        self.assertEqual(data['similar'][0]['serial_number'], 'БУАД-1234')
        # Запись не создана: сотрудник ещё не решил, та же это единица или нет
        self.assertEqual(Equipment.objects.count(), 1)

    def test_confirmed_creation_proceeds_despite_similarity(self):
        response = self.client_http.post('/ajax/equipment/create/', {
            'model_id': self.model.pk,
            'serial_number': 'буад 1234',
            'confirmed': '1',
        })

        self.assertTrue(response.json()['success'])
        self.assertEqual(Equipment.objects.count(), 2)

    def test_exact_duplicate_is_still_rejected_outright(self):
        """Полное совпадение — не повод переспрашивать, это просто дубль."""
        response = self.client_http.post('/ajax/equipment/create/', {
            'model_id': self.model.pk,
            'serial_number': 'БУАД-1234',
            'confirmed': '1',
        })

        data = response.json()
        self.assertFalse(data['success'])
        self.assertNotIn('similar', data)
        self.assertEqual(Equipment.objects.count(), 1)

    def test_unrelated_serial_creates_without_warning(self):
        response = self.client_http.post('/ajax/equipment/create/', {
            'model_id': self.model.pk,
            'serial_number': 'СОВСЕМ-ДРУГОЙ',
        })

        self.assertTrue(response.json()['success'])
        self.assertEqual(Equipment.objects.count(), 2)

    def test_history_requires_login(self):
        anonymous = TestClient()
        response = anonymous.get(f'/equipment/{self.equipment.pk}/history/')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])


class LabelTests(TestCase):
    """Две разные этикетки: на пакет с деталью и на саму ячейку.
    Раньше кнопка в списке деталей печатала этикетку ячейки, что для
    наклейки на пакет неверно — на ней не было самой детали."""

    def setUp(self):
        self.user = Employee.objects.create_superuser(
            username='label_user', full_name='Тест', password='pass'
        )
        self.part = SparePart.objects.create(
            part_number='LBL-1', name='Деталь для этикетки',
            component_type='Резисторы', current_stock=5,
        )
        self.other = SparePart.objects.create(part_number='LBL-2', name='Соседняя деталь')
        self.cell = StorageCell.objects.create(cabinet=Cabinet.objects.get_or_create(number=2)[0], row_number=3, cell_row=4)
        self.cell.parts.add(self.part, self.other)
        self.client_http = TestClient()
        self.client_http.force_login(self.user)

    def test_part_label_shows_only_that_part(self):
        response = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'LBL-1')
        # Соседняя деталь из той же ячейки на этикетке пакета быть не должна
        self.assertNotContains(response, 'LBL-2')

    def test_part_label_shows_cell_address(self):
        response = self.client_http.get(f'/parts/{self.part.pk}/label/')
        self.assertContains(response, self.cell.address)

    def test_cell_label_lists_all_parts(self):
        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.cell.address)

    def test_short_urls_redirect_to_pages(self):
        part_response = self.client_http.get(f'/p/{self.part.pk}/')
        self.assertRedirects(part_response, f'/parts/{self.part.pk}/')

        cell_response = self.client_http.get(f'/c/{self.cell.pk}/')
        self.assertRedirects(
            cell_response,
            f'/storage-cells/?cabinet={self.cell.cabinet.number}&open_cell={self.cell.pk}',
        )

    def test_short_urls_require_login(self):
        anonymous = TestClient()
        response = anonymous.get(f'/p/{self.part.pk}/')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])

    def test_label_base_url_setting_overrides_request_host(self):
        """Адрес в коде не должен зависеть от того, откуда печатали."""
        from core.views import label_base_url

        request = TestClient().get('/').wsgi_request
        with self.settings(LABEL_BASE_URL='http://100.108.92.92/'):
            self.assertEqual(label_base_url(request), 'http://100.108.92.92')

    def test_part_label_shows_the_package_next_to_the_article(self):
        """Корпус — то, по чему деталь опознают в руках."""
        self.part.package = 'TO-220'
        self.part.save(update_fields=['package'])

        response = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertContains(response, 'TO-220')
        self.assertContains(response, 'label-package')

    def test_part_label_prints_the_description(self):
        self.part.description = 'Ставится в блок питания БУАД'
        self.part.save(update_fields=['description'])

        response = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertContains(response, 'Ставится в блок питания БУАД')

    def test_part_label_falls_back_to_the_name(self):
        """Описание заполняют не всегда, а пустая строка на этикетке
        не нужна никому."""
        self.assertEqual(self.part.description, '')

        response = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertContains(response, 'Деталь для этикетки')

    def test_part_label_loads_the_font_fitting_script(self):
        """Без него длинный артикул обрезался бы многоточием."""
        for page in (f'/parts/{self.part.pk}/label/', '/parts/labels/?ids=%d' % self.part.pk):
            with self.subTest(page=page):
                response = self.client_http.get(page)
                self.assertContains(response, 'label-fit.js')
                self.assertContains(response, 'data-fit')

    def test_short_urls_work_without_the_trailing_slash(self):
        """Именно эта форма попадает в QR — она на символ короче."""
        part_response = self.client_http.get(f'/p/{self.part.pk}')
        self.assertRedirects(part_response, f'/parts/{self.part.pk}/')

        cell_response = self.client_http.get(f'/c/{self.cell.pk}')
        self.assertRedirects(
            cell_response,
            f'/storage-cells/?cabinet={self.cell.cabinet.number}&open_cell={self.cell.pk}',
        )

    def test_the_code_carries_no_server_address(self):
        """В коде — только вид и номер. Адрес сервера там не нужен: сканируют,
        когда программа уже открыта, а места он занимал три четверти."""
        from core.views import qr_link, qr_payload

        self.assertEqual(qr_payload('u', 123), 'u/123')
        self.assertEqual(qr_link('http://100.108.92.92', 'u', 123),
                         'http://100.108.92.92/u/123')

    @override_settings(LABEL_BASE_URL='http://100.108.92.92')
    def test_every_label_type_keeps_the_address_out_of_the_code(self):
        """Все этикетки, а не только те, о которых вспомнили."""
        equipment_model = EquipmentModel.objects.create(name='БУАД-QR')
        equipment = Equipment.objects.create(model=equipment_model, serial_number='QR-1')
        client = ClientModel.objects.create(name='Заказчик QR')
        order = RepairOrder.objects.create(
            order_number='LT-QR-1', client=client, date_received=datetime.date.today()
        )
        roe = RepairOrderEquipment.objects.create(repair_order=order, equipment=equipment)

        pages = {
            f'/parts/{self.part.pk}/label/': f'p/{self.part.pk}',
            f'/storage-cells/{self.cell.pk}/label/': f'c/{self.cell.pk}',
            f'/repair-orders/{order.pk}/equipment/{roe.pk}/label/': f'u/{roe.pk}',
        }
        for page, payload in pages.items():
            with self.subTest(page=page):
                with patch('core.views.generate_qr_image') as qr:
                    qr.return_value = 'data:image/png;base64,x'
                    response = self.client_http.get(page)

                # В коде — только вид и номер, адреса там нет
                self.assertEqual(qr.call_args[0][0], payload)
                # А настроенный адрес показывается человеку на экране,
                # чтобы ошибку в нём было видно до печати
                self.assertEqual(response.context['qr_url'],
                                 f'http://100.108.92.92/{payload}')


class PiSettingsLabelTests(TestCase):
    """Настройки Raspberry Pi: адрес этикеток и то, что по нему пускают."""

    def _load(self, **env):
        """Загружает settings_pi с заданным окружением.

        Модуль настроек читает переменные при импорте, поэтому его
        приходится перезагружать, а не подменять значения после.
        """
        import importlib
        from unittest.mock import patch as patch_env

        environment = {'SECRET_KEY': 'test-key-for-settings'}
        environment.update(env)
        with patch_env.dict('os.environ', environment, clear=False):
            module = importlib.import_module('lifteam.settings_pi')
            return importlib.reload(module)

    def test_labels_point_at_the_tailscale_name_by_default(self):
        """Имя, а не адрес 100.x: этикетка — бумажка, адрес в ней впечатан
        навсегда, а 100.x меняется при заведении узла заново."""
        settings_pi = self._load(ALLOWED_HOSTS='192.168.1.50,localhost')

        self.assertEqual(settings_pi.LABEL_BASE_URL,
                         'http://lifteam.taile9b605.ts.net')

    def test_the_name_does_not_reach_the_code_at_all(self):
        """Длина имени на размер кода больше не влияет: адреса в коде нет.

        Раньше она влияла напрямую, и опасение «имя длиннее адреса, код
        помельчает» было осмысленным. Теперь в коде лежит `u/1234`,
        и настройка на него не действует никак."""
        settings_pi = self._load(ALLOWED_HOSTS='localhost')

        self.assertTrue(settings_pi.LABEL_BASE_URL)
        self.assertEqual(views.qr_length_warning([views.qr_payload('u', 1234)]), '')

    def test_the_scanned_address_is_allowed(self):
        """Иначе отсканированный код приводит на «400 Bad Request»."""
        settings_pi = self._load(ALLOWED_HOSTS='192.168.1.50,localhost')

        self.assertIn('lifteam.taile9b605.ts.net', settings_pi.ALLOWED_HOSTS)
        self.assertIn('http://lifteam.taile9b605.ts.net',
                      settings_pi.CSRF_TRUSTED_ORIGINS)

    def test_a_custom_address_is_allowed_too(self):
        """Адрес можно переопределить в .env — например, дописав порт,
        если nginx не ставили."""
        settings_pi = self._load(LABEL_BASE_URL='http://100.64.0.7:8000')

        self.assertEqual(settings_pi.LABEL_BASE_URL, 'http://100.64.0.7:8000')
        # В ALLOWED_HOSTS порт не указывается, в доверенных источниках — да
        self.assertIn('100.64.0.7', settings_pi.ALLOWED_HOSTS)
        self.assertIn('http://100.64.0.7:8000', settings_pi.CSRF_TRUSTED_ORIGINS)

    def test_an_empty_address_falls_back_to_the_print_page(self):
        """Пустая настройка — прежнее поведение: адрес берётся из запроса."""
        settings_pi = self._load(LABEL_BASE_URL='')

        self.assertEqual(settings_pi.LABEL_BASE_URL, '')
        self.assertNotIn('100.108.92.92', settings_pi.ALLOWED_HOSTS)


class UpdaterTests(TestCase):
    """Обновление через интерфейс. Приложение не имеет прав root и лишь
    оставляет заявку — проверяем, что в заявку нельзя подсунуть что угодно
    и что попасть на страницу может только администратор."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='upd_admin', full_name='Админ', password='pass'
        )
        self.warehouse = Employee.objects.create_user(
            username='upd_wh', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)

    def test_page_requires_admin_role(self):
        client = TestClient()
        client.force_login(self.warehouse)
        response = client.get('/management/update/', follow=True)
        self.assertRedirects(response, '/')

    def test_page_open_for_admin(self):
        client = TestClient()
        client.force_login(self.admin)
        self.assertEqual(client.get('/management/update/').status_code, 200)

    def test_status_endpoint_requires_admin(self):
        client = TestClient()
        client.force_login(self.warehouse)
        response = client.get('/management/update/status/', follow=True)
        self.assertRedirects(response, '/')

    def test_request_update_rejects_command_injection(self):
        from core import updater

        for evil in ('; rm -rf /', '$(whoami)', '../../etc/passwd', 'main', 'HEAD', ''):
            with self.subTest(target=evil):
                with self.assertRaises(ValueError):
                    updater.request_update(evil)

    def test_request_update_accepts_valid_targets(self):
        from core import updater

        tmp_path = Path(self.tmp.name)
        with patch.object(updater, 'REQUEST_FILE', tmp_path / '.update-request'), \
             patch.object(updater, 'STATUS_FILE', tmp_path / '.update-status'):
            updater.request_update('latest', requested_by='upd_admin')
            written = json.loads((tmp_path / '.update-request').read_text(encoding='utf-8'))

        self.assertEqual(written['target'], 'latest')
        self.assertEqual(written['requested_by'], 'upd_admin')

    def test_request_update_accepts_commit_hash(self):
        from core import updater

        tmp_path = Path(self.tmp.name)
        with patch.object(updater, 'REQUEST_FILE', tmp_path / '.update-request'), \
             patch.object(updater, 'STATUS_FILE', tmp_path / '.update-status'):
            updater.request_update('a1b2c3d')
            written = json.loads((tmp_path / '.update-request').read_text(encoding='utf-8'))

        self.assertEqual(written['target'], 'a1b2c3d')

    def test_no_request_file_written_for_invalid_target(self):
        """Заявка не должна появляться, если значение отвергнуто."""
        from core import updater

        tmp_path = Path(self.tmp.name)
        request_file = tmp_path / '.update-request'
        with patch.object(updater, 'REQUEST_FILE', request_file), \
             patch.object(updater, 'STATUS_FILE', tmp_path / '.update-status'):
            with self.assertRaises(ValueError):
                updater.request_update('; reboot')

        self.assertFalse(request_file.exists())


class UrlRoutingTests(TestCase):
    def test_management_users_not_shadowed_by_django_admin(self):
        """Регрессия: /admin/users/ раньше перехватывался встроенной админкой Django
        (path('admin/', admin.site.urls) match'ился раньше core.urls)."""
        match = resolve('/management/users/')
        self.assertEqual(match.func.__name__, 'admin_users')


class SqliteConfigurationTests(TestCase):
    def test_busy_timeout_applied_to_connection(self):
        """Без увеличенного таймаута одновременная работа даёт 'database is locked'."""
        with connection.cursor() as cursor:
            timeout = cursor.execute('PRAGMA busy_timeout;').fetchone()[0]
        self.assertEqual(timeout, 30000)

    def test_handler_ignores_non_sqlite_backends(self):
        """На PostgreSQL PRAGMA-команды недопустимы — обработчик обязан выйти сразу."""
        from .signals import configure_sqlite

        fake_connection = type('FakeConnection', (), {
            'vendor': 'postgresql',
            'cursor': lambda self: (_ for _ in ()).throw(
                AssertionError('cursor() не должен вызываться для PostgreSQL')
            ),
        })()
        configure_sqlite(sender=None, connection=fake_connection)


class BackupRestoreTests(TestCase):
    """Проверяют, что копии действительно восстановимы, а не просто создаются."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.tmp_path = Path(self.tmp.name)
        self.source_db = self.tmp_path / 'source.sqlite3'

        con = sqlite3.connect(self.source_db)
        con.execute('CREATE TABLE parts (id INTEGER PRIMARY KEY, name TEXT);')
        con.executemany(
            'INSERT INTO parts (name) VALUES (?);',
            [('Резистор',), ('Конденсатор',), ('Диод',)],
        )
        con.commit()
        con.close()

    def _run_backup(self, **kwargs):
        db_config = settings.DATABASES['default']
        with patch.dict(db_config, {'NAME': str(self.source_db)}):
            call_command('backup_db', output=str(self.tmp_path / 'backups'), verbosity=0, **kwargs)

    def test_backup_is_created_and_restorable(self):
        self._run_backup()

        backups = list((self.tmp_path / 'backups').glob('db_*.sqlite3'))
        self.assertEqual(len(backups), 1)

        con = sqlite3.connect(backups[0])
        try:
            self.assertEqual(con.execute('PRAGMA integrity_check;').fetchone()[0], 'ok')
            names = [row[0] for row in con.execute('SELECT name FROM parts ORDER BY id;')]
        finally:
            con.close()
        self.assertEqual(names, ['Резистор', 'Конденсатор', 'Диод'])

    def test_backup_reflects_data_written_after_previous_backup(self):
        self._run_backup()

        con = sqlite3.connect(self.source_db)
        con.execute("INSERT INTO parts (name) VALUES ('Стабилитрон');")
        con.commit()
        con.close()

        self._run_backup()

        backups = sorted((self.tmp_path / 'backups').glob('db_*.sqlite3'))
        # Вторая копия за ту же секунду не должна затирать первую
        self.assertEqual(len(backups), 2)

        counts = []
        for path in backups:
            con = sqlite3.connect(path)
            try:
                counts.append(con.execute('SELECT COUNT(*) FROM parts;').fetchone()[0])
            finally:
                con.close()
        self.assertEqual(counts, [3, 4])

    def test_gzip_backup_is_compressed_and_readable(self):
        self._run_backup(gzip=True)

        backups = list((self.tmp_path / 'backups').glob('db_*.sqlite3.gz'))
        self.assertEqual(len(backups), 1)
        self.assertFalse(list((self.tmp_path / 'backups').glob('db_*.sqlite3')))

    def test_retention_keeps_only_requested_number(self):
        from core.management.commands.backup_db import Command

        backup_dir = self.tmp_path / 'retention'
        backup_dir.mkdir()
        for i in range(6):
            (backup_dir / f'db_2026-01-0{i}_120000.sqlite3').write_bytes(b'x')

        removed = Command()._prune(backup_dir, keep=2)

        self.assertEqual(removed, 4)
        remaining = sorted(p.name for p in backup_dir.glob('db_*.sqlite3'))
        self.assertEqual(remaining, [
            'db_2026-01-04_120000.sqlite3',
            'db_2026-01-05_120000.sqlite3',
        ])

    def test_retention_ignores_sqlite_sidecar_files(self):
        """Файлы -wal/-shm не должны занимать места в квоте хранения копий."""
        from core.management.commands.backup_db import Command

        backup_dir = self.tmp_path / 'sidecars'
        backup_dir.mkdir()
        for i in range(3):
            base = backup_dir / f'db_2026-01-0{i}_120000.sqlite3'
            base.write_bytes(b'x')
            base.with_name(base.name + '-wal').write_bytes(b'x')
            base.with_name(base.name + '-shm').write_bytes(b'x')

        removed = Command()._prune(backup_dir, keep=2)

        self.assertEqual(removed, 1)
        remaining = sorted(
            p.name for p in backup_dir.glob('db_*.sqlite3')
            if not p.name.endswith(('-wal', '-shm'))
        )
        self.assertEqual(remaining, [
            'db_2026-01-01_120000.sqlite3',
            'db_2026-01-02_120000.sqlite3',
        ])
        # служебные файлы удалённой копии тоже убраны
        self.assertFalse((backup_dir / 'db_2026-01-00_120000.sqlite3-wal').exists())

    def test_retention_disabled_keeps_everything(self):
        from core.management.commands.backup_db import Command

        backup_dir = self.tmp_path / 'keep-all'
        backup_dir.mkdir()
        for i in range(3):
            (backup_dir / f'db_2026-01-0{i}_120000.sqlite3').write_bytes(b'x')

        self.assertEqual(Command()._prune(backup_dir, keep=0), 0)
        self.assertEqual(len(list(backup_dir.glob('db_*.sqlite3'))), 3)

    def test_restore_refuses_corrupted_backup(self):
        """Повреждённая копия не должна затирать рабочую базу."""
        from django.core.management.base import CommandError

        corrupted = self.tmp_path / 'corrupted.sqlite3'
        corrupted.write_bytes(b'this is not a database')

        with self.assertRaises(CommandError):
            call_command('restore_db', str(corrupted), yes=True, verbosity=0)


class ReportExportTests(TestCase):
    """Выгрузка отчётов в Excel: состав строк, учёт фильтров, итог по долгам."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_rep', full_name='Админ отчётов', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def _sheet(self, url):
        resp = self.client_http.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        return openpyxl.load_workbook(io.BytesIO(resp.content)).active

    def _rows(self, ws):
        """Строки листа без шапки, в виде словарей по названиям колонок."""
        headers = [c.value for c in ws[1]]
        return [
            dict(zip(headers, [c.value for c in row]))
            for row in ws.iter_rows(min_row=2)
        ]

    # --- План закупок ---

    def test_purchase_plan_export_lists_only_parts_below_minimum(self):
        SparePart.objects.create(part_number='LOW-1', name='Не хватает', current_stock=1, min_stock=5)
        SparePart.objects.create(part_number='OK-1', name='В норме', current_stock=10, min_stock=5)

        rows = self._rows(self._sheet('/reports/purchase-plan/export/'))

        # Последняя строка — «Итого»: файл уходит поставщику и начальству,
        # и сумму из него достают в первую очередь
        self.assertEqual([r['Артикул'] for r in rows], ['LOW-1', 'Итого'])
        self.assertEqual(rows[0]['Не хватает'], 4)

    def test_purchase_plan_export_totals_the_prices(self):
        SparePart.objects.create(part_number='LOW-P1', name='С ценой',
                                 current_stock=0, min_stock=4, price=Decimal('25.00'))
        SparePart.objects.create(part_number='LOW-P2', name='Без цены',
                                 current_stock=0, min_stock=3)

        rows = self._rows(self._sheet('/reports/purchase-plan/export/'))

        self.assertEqual(float(rows[0]['Сумма, ₽']), 100.0)
        self.assertIsNone(rows[1]['Сумма, ₽'])  # цену не заполняли — и суммы нет
        self.assertEqual(float(rows[-1]['Сумма, ₽']), 100.0)

    def test_purchase_plan_export_shows_cell_address(self):
        part = SparePart.objects.create(part_number='LOW-2', name='Деталь', current_stock=0, min_stock=3)
        cell = StorageCell.objects.create(cabinet=Cabinet.objects.get_or_create(number=1)[0], row_number=2, cell_row=3)
        cell.parts.add(part)

        rows = self._rows(self._sheet('/reports/purchase-plan/export/'))
        self.assertEqual(rows[0]['Ячейка'], 'К1-Р2-Я3')

    def test_purchase_plan_export_requires_login(self):
        resp = TestClient().get('/reports/purchase-plan/export/')
        self.assertEqual(resp.status_code, 302)

    # --- Журнал движений ---

    def test_movements_export_writes_timezone_aware_dates(self):
        """Регрессия: openpyxl отказывается записывать дату с часовым поясом,
        а все движения хранятся именно такими."""
        part = SparePart.objects.create(part_number='MOV-1', name='Деталь', current_stock=5)
        StockMovement.objects.create(
            part=part, quantity=3, movement_type='incoming',
            notes='Первый приход', created_by=self.admin
        )

        rows = self._rows(self._sheet('/reports/stock-movements/export/'))

        self.assertEqual(len(rows), 1)
        self.assertIsNotNone(rows[0]['Дата'])
        self.assertEqual(rows[0]['Тип'], 'Приход')
        self.assertEqual(rows[0]['Сотрудник'], 'Админ отчётов')
        self.assertEqual(rows[0]['Примечания'], 'Первый приход')

    def test_movements_export_respects_type_filter(self):
        part = SparePart.objects.create(part_number='MOV-2', name='Деталь', current_stock=5)
        StockMovement.objects.create(part=part, quantity=1, movement_type='incoming')
        StockMovement.objects.create(part=part, quantity=2, movement_type='outgoing')

        rows = self._rows(self._sheet('/reports/stock-movements/export/?type=outgoing'))

        self.assertEqual([r['Количество'] for r in rows], [2])

    def test_movements_export_respects_part_filter(self):
        first = SparePart.objects.create(part_number='MOV-3', name='Первая')
        second = SparePart.objects.create(part_number='MOV-4', name='Вторая')
        StockMovement.objects.create(part=first, quantity=1, movement_type='incoming')
        StockMovement.objects.create(part=second, quantity=1, movement_type='incoming')

        rows = self._rows(self._sheet(f'/reports/stock-movements/export/?part={first.pk}'))
        self.assertEqual([r['Артикул'] for r in rows], ['MOV-3'])

    def test_movement_filters_ignore_unparsable_values(self):
        """Адрес правят руками; мусор в фильтре не должен ронять страницу."""
        part = SparePart.objects.create(part_number='MOV-5', name='Деталь')
        StockMovement.objects.create(part=part, quantity=1, movement_type='incoming')

        page = self.client_http.get('/reports/stock-movements/?part=abc&date_from=вчера&type=что-то')
        self.assertEqual(page.status_code, 200)

        rows = self._rows(self._sheet('/reports/stock-movements/export/?part=abc&date_from=вчера'))
        self.assertEqual(len(rows), 1)

    # --- Задолженности ---

    def _debtor_order(self, cost, payment_status='unpaid'):
        client_obj = ClientModel.objects.create(name=f'Должник {cost}')
        order = RepairOrder.objects.create(client=client_obj, payment_status=payment_status)
        model = EquipmentModel.objects.create(name=f'Модель {cost}')
        equipment = Equipment.objects.create(model=model, serial_number=f'SN-{cost}')
        RepairOrderEquipment.objects.create(
            repair_order=order, equipment=equipment, repair_cost=cost
        )
        return order

    def test_debtors_export_excludes_paid_orders(self):
        self._debtor_order(1000)
        self._debtor_order(2000, payment_status='paid')

        rows = self._rows(self._sheet('/reports/debtors/export/'))

        # одна строка долга плюс строка «Итого»
        self.assertEqual(len(rows), 2)
        self.assertEqual(float(rows[0]['Сумма, ₽']), 1000.0)

    def test_debtors_export_ends_with_total(self):
        self._debtor_order(1500)
        self._debtor_order(2500, payment_status='partially_paid')

        rows = self._rows(self._sheet('/reports/debtors/export/'))
        total_row = rows[-1]

        self.assertEqual(total_row['№ заказа'], 'Итого')
        # Итог по остаткам, а не по стоимостям: часть денег могла уже прийти
        self.assertEqual(float(total_row['Остаток, ₽']), 4000.0)

    def test_debtors_export_subtracts_payments(self):
        order = self._debtor_order(2000)
        Payment.objects.create(repair_order=order, amount=500)

        rows = self._rows(self._sheet('/reports/debtors/export/'))

        self.assertEqual(float(rows[0]['Оплачено, ₽']), 500.0)
        self.assertEqual(float(rows[0]['Остаток, ₽']), 1500.0)
        self.assertEqual(float(rows[-1]['Остаток, ₽']), 1500.0)

    def test_debtors_export_is_empty_without_debts(self):
        ws = self._sheet('/reports/debtors/export/')
        self.assertEqual(ws.max_row, 1)  # только шапка, без строки «Итого»

    def test_report_pages_offer_export(self):
        self._debtor_order(100)
        SparePart.objects.create(part_number='LOW-3', name='Деталь', current_stock=0, min_stock=1)

        for url in ('/reports/purchase-plan/', '/reports/stock-movements/', '/reports/debtors/'):
            with self.subTest(url=url):
                resp = self.client_http.get(url)
                self.assertEqual(resp.status_code, 200)
                self.assertIn('export/', resp.content.decode())

    def test_total_debt_matches_sum_of_orders(self):
        """Итог считается одним запросом — он должен совпадать с суммой по заказам."""
        from core.views import _debtor_orders, _total_debt

        self._debtor_order(700)
        self._debtor_order(300, payment_status='partially_paid')

        orders = _debtor_orders()
        self.assertEqual(
            _total_debt(orders),
            sum(order.total_repair_cost for order in orders)
        )


class PaginationFilterTests(TestCase):
    """Переход на следующую страницу не должен терять выставленные фильтры."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_pag', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def test_parts_pagination_keeps_voltage_filter(self):
        """Регрессия: ссылка на следующую страницу собиралась вручную и не
        включала фильтры по напряжению, току и остатку — на второй странице
        показывались детали, не проходящие фильтр."""
        for i in range(30):
            SparePart.objects.create(part_number=f'PAG-{i:03d}', name='Деталь', voltage=12)

        resp = self.client_http.get('/parts/?voltage_from=5')
        content = resp.content.decode()

        self.assertIn('page=2', content)
        self.assertIn('voltage_from=5', content)

    def test_movements_pagination_keeps_date_filter(self):
        part = SparePart.objects.create(part_number='PAG-MOV', name='Деталь')
        for _ in range(60):
            StockMovement.objects.create(part=part, quantity=1, movement_type='incoming')

        resp = self.client_http.get('/reports/stock-movements/?date_from=2020-01-01')
        content = resp.content.decode()

        self.assertIn('page=2', content)
        self.assertIn('date_from=2020-01-01', content)


class WarrantyTests(TestCase):
    """Гарантия на ремонт: отсчёт от даты завершения заказа."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_war', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.client_obj = ClientModel.objects.create(name='Заказчик')
        self.model = EquipmentModel.objects.create(name='БУАД-3')
        self.equipment = Equipment.objects.create(model=self.model, serial_number='W-001')

    def _repair(self, completed_at, equipment=None):
        """Завершённый ремонт единицы с заданной датой завершения."""
        order = RepairOrder.objects.create(client=self.client_obj, status='shipped')
        # date_completed заполняется сменой статуса, в тесте ставим напрямую
        RepairOrder.objects.filter(pk=order.pk).update(date_completed=completed_at)
        order.refresh_from_db()
        return RepairOrderEquipment.objects.create(
            repair_order=order, equipment=equipment or self.equipment, repair_cost=1000
        )

    # --- арифметика срока ---

    def test_add_months_keeps_day_of_month(self):
        from core.models import add_months
        from datetime import datetime

        self.assertEqual(
            add_months(datetime(2026, 8, 12), 12), datetime(2027, 8, 12)
        )

    def test_add_months_clamps_to_last_day(self):
        """29 февраля плюс год — 28 февраля, а не ошибка."""
        from core.models import add_months
        from datetime import datetime

        self.assertEqual(
            add_months(datetime(2024, 2, 29), 12), datetime(2025, 2, 28)
        )

    def test_warranty_uses_calendar_year_not_365_days(self):
        """В високосный год гарантия должна кончаться в ту же дату, а не на день раньше."""
        completed = timezone.make_aware(datetime.datetime(2024, 3, 1, 12, 0))
        visit = self._repair(completed)

        self.assertEqual(timezone.localtime(visit.warranty_until).date(),
                         datetime.date(2025, 3, 1))

    # --- границы ---

    def test_recent_repair_is_under_warranty(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=30))

        self.assertTrue(visit.is_under_warranty)
        self.assertGreater(visit.warranty_days_left, 300)

    def test_repair_just_inside_the_year_is_still_covered(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=364))
        self.assertTrue(visit.is_under_warranty)

    def test_repair_older_than_the_year_is_not_covered(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=400))

        self.assertFalse(visit.is_under_warranty)
        self.assertLess(visit.warranty_days_left, 0)

    def test_unfinished_order_has_no_warranty_yet(self):
        order = RepairOrder.objects.create(client=self.client_obj)
        visit = RepairOrderEquipment.objects.create(
            repair_order=order, equipment=self.equipment
        )

        self.assertIsNone(visit.warranty_until)
        self.assertFalse(visit.is_under_warranty)
        self.assertIsNone(visit.warranty_days_left)

    @override_settings(WARRANTY_MONTHS=0)
    def test_zero_months_disables_warranty(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=1))

        self.assertIsNone(visit.warranty_until)
        self.assertEqual(Equipment.warranty_map([self.equipment]), {})

    @override_settings(WARRANTY_MONTHS=6)
    def test_period_is_configurable(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=200))
        self.assertFalse(visit.is_under_warranty)

    # --- поиск действующей гарантии ---

    def test_active_warranty_picks_the_latest_repair(self):
        self._repair(timezone.now() - datetime.timedelta(days=300))
        recent = self._repair(timezone.now() - datetime.timedelta(days=10))

        found = self.equipment.active_warranty()
        self.assertEqual(found.pk, recent.pk)

    def test_active_warranty_is_none_when_all_repairs_are_old(self):
        self._repair(timezone.now() - datetime.timedelta(days=500))
        self.assertIsNone(self.equipment.active_warranty())

    def test_active_warranty_can_exclude_current_order(self):
        """На странице заказа его собственная гарантия не должна выглядеть
        как гарантия по прошлому ремонту."""
        visit = self._repair(timezone.now() - datetime.timedelta(days=10))

        self.assertIsNotNone(self.equipment.active_warranty())
        self.assertIsNone(
            self.equipment.active_warranty(exclude_order_id=visit.repair_order_id)
        )

    def test_warranty_map_covers_many_units_in_one_query(self):
        other = Equipment.objects.create(model=self.model, serial_number='W-002')
        cold = Equipment.objects.create(model=self.model, serial_number='W-003')
        self._repair(timezone.now() - datetime.timedelta(days=10))
        self._repair(timezone.now() - datetime.timedelta(days=20), equipment=other)
        self._repair(timezone.now() - datetime.timedelta(days=500), equipment=cold)

        units = [self.equipment, other, cold]
        with self.assertNumQueries(1):
            found = Equipment.warranty_map(units)

        self.assertEqual(set(found), {self.equipment.pk, other.pk})

    # --- интерфейс ---

    def test_history_page_shows_active_warranty(self):
        self._repair(timezone.now() - datetime.timedelta(days=10))

        resp = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')
        self.assertContains(resp, 'На гарантии до')

    def test_history_page_marks_expired_warranty(self):
        self._repair(timezone.now() - datetime.timedelta(days=500))

        resp = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')
        self.assertNotContains(resp, 'На гарантии до')
        self.assertContains(resp, 'истекла')

    def test_equipment_list_shows_warranty_badge(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=10))
        expiry = timezone.localtime(visit.warranty_until).strftime('%d.%m.%Y')

        resp = self.client_http.get('/equipment/')
        self.assertContains(resp, f'до {expiry}')

    def test_equipment_list_leaves_uncovered_units_empty(self):
        self._repair(timezone.now() - datetime.timedelta(days=500))

        resp = self.client_http.get('/equipment/')
        self.assertIsNone(resp.context['equipment'][0].warranty)

    def test_intake_hint_reports_warranty(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=10))

        resp = self.client_http.get(
            f'/ajax/equipment/{self.equipment.pk}/history-summary/'
        )
        data = resp.json()

        self.assertTrue(data['under_warranty'])
        self.assertEqual(data['warranty_order_number'], visit.repair_order.order_number)
        self.assertTrue(data['warranty_until'])

    def test_intake_hint_without_warranty(self):
        self._repair(timezone.now() - datetime.timedelta(days=500))

        data = self.client_http.get(
            f'/ajax/equipment/{self.equipment.pk}/history-summary/'
        ).json()

        self.assertFalse(data['under_warranty'])
        self.assertEqual(data['warranty_until'], '')

    def test_order_page_flags_repeat_visit_under_warranty(self):
        """Регрессия смысла: гарантия по прошлому заказу должна быть видна
        в новом, иначе повторный ремонт примут как обычный платный."""
        old = self._repair(timezone.now() - datetime.timedelta(days=10))

        new_order = RepairOrder.objects.create(client=self.client_obj)
        RepairOrderEquipment.objects.create(
            repair_order=new_order, equipment=self.equipment
        )

        resp = self.client_http.get(f'/repair-orders/{new_order.pk}/')
        self.assertContains(resp, old.repair_order.order_number)

    def test_order_page_does_not_flag_its_own_warranty_as_previous(self):
        visit = self._repair(timezone.now() - datetime.timedelta(days=10))

        resp = self.client_http.get(f'/repair-orders/{visit.repair_order.pk}/')
        shown = resp.context['order_equipments'][0]

        self.assertIsNone(shown.previous_warranty)
        self.assertTrue(shown.is_under_warranty)


class OrderSearchTests(TestCase):
    """Расширенный поиск заказов: по всем полям, которые сотрудник может помнить."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_search', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.alpha = ClientModel.objects.create(name='ООО Альфа', inn='7700000001')
        self.beta = ClientModel.objects.create(name='ООО Бета', inn='7700000002')
        self.buad = EquipmentModel.objects.create(name='БУАД')
        self.eco = EquipmentModel.objects.create(name='Экодрайв')

        self.order_a = RepairOrder.objects.create(
            client=self.alpha, status='repair', payment_status='unpaid',
            invoice_number='СЧ-100', tracking_number='TRK-777',
        )
        RepairOrderEquipment.objects.create(
            repair_order=self.order_a,
            equipment=Equipment.objects.create(model=self.buad, serial_number='БУАД-1234'),
            fault_description='Не запускается двигатель',
        )

        self.order_b = RepairOrder.objects.create(
            client=self.beta, status='shipped', payment_status='paid',
            invoice_number='СЧ-200',
        )
        RepairOrderEquipment.objects.create(
            repair_order=self.order_b,
            equipment=Equipment.objects.create(model=self.eco, serial_number='ECO-9'),
            fault_description='Скрип при подъёме',
        )

    def _found(self, query=''):
        resp = self.client_http.get(f'/repair-orders/{query}')
        self.assertEqual(resp.status_code, 200)
        return {order.pk for order in resp.context['orders']}

    # --- текстовый поиск ---

    def test_search_by_order_number(self):
        self.assertEqual(self._found(f'?q={self.order_a.order_number}'), {self.order_a.pk})

    def test_search_by_client_name(self):
        self.assertEqual(self._found('?q=Альфа'), {self.order_a.pk})

    def test_search_by_client_inn(self):
        self.assertEqual(self._found('?q=7700000002'), {self.order_b.pk})

    def test_search_by_invoice_number(self):
        self.assertEqual(self._found('?q=СЧ-200'), {self.order_b.pk})

    def test_search_by_tracking_number(self):
        self.assertEqual(self._found('?q=TRK-777'), {self.order_a.pk})

    def test_search_by_unit_fault_description(self):
        """Неисправность записана у единицы: общего описания у заказа
        больше нет — его заполняли вместо описания по прибору."""
        self.assertEqual(self._found('?q=двигатель'), {self.order_a.pk})
        self.assertEqual(self._found('?q=скрип'), {self.order_b.pk})

    def test_search_by_equipment_model(self):
        self.assertEqual(self._found('?q=Экодрайв'), {self.order_b.pk})

    def test_search_by_serial_in_another_spelling(self):
        """«буад 1234» должен находить заказ с «БУАД-1234» — как в оборудовании."""
        self.assertEqual(self._found('?q=буад 1234'), {self.order_a.pk})

    def test_search_finds_nothing_for_unknown_text(self):
        self.assertEqual(self._found('?q=такого+нет'), set())

    def test_order_with_several_units_is_listed_once(self):
        """Регрессия: соединение с оборудованием размножало строки заказа."""
        RepairOrderEquipment.objects.create(
            repair_order=self.order_a,
            equipment=Equipment.objects.create(model=self.buad, serial_number='БУАД-5678'),
        )
        resp = self.client_http.get('?q=Альфа'.join(['/repair-orders/', '']))
        listed = [o.pk for o in resp.context['orders']]

        self.assertEqual(listed.count(self.order_a.pk), 1)

    # --- фильтры ---

    def test_filter_by_status(self):
        self.assertEqual(self._found('?status=repair'), {self.order_a.pk})

    def test_filter_by_payment_status(self):
        self.assertEqual(self._found('?payment_status=paid'), {self.order_b.pk})

    def test_filter_by_client(self):
        self.assertEqual(self._found(f'?client={self.beta.pk}'), {self.order_b.pk})

    def test_filter_by_equipment_model(self):
        self.assertEqual(self._found(f'?model={self.buad.pk}'), {self.order_a.pk})

    def test_filter_by_date_range(self):
        today = timezone.localdate().isoformat()
        self.assertEqual(
            self._found(f'?date_from={today}&date_to={today}'),
            {self.order_a.pk, self.order_b.pk},
        )

    def test_date_range_excludes_other_days(self):
        tomorrow = (timezone.localdate() + datetime.timedelta(days=1)).isoformat()
        self.assertEqual(self._found(f'?date_from={tomorrow}'), set())

    def test_filters_combine(self):
        self.assertEqual(
            self._found(f'?q=Альфа&status=repair&payment_status=unpaid'),
            {self.order_a.pk},
        )

    def test_contradictory_filters_find_nothing(self):
        self.assertEqual(self._found('?q=Альфа&status=shipped'), set())

    # --- устойчивость и мелочи интерфейса ---

    def test_garbage_in_filters_does_not_break_the_page(self):
        found = self._found('?client=abc&model=xyz&date_from=вчера&status=нет-такого')
        self.assertEqual(found, {self.order_a.pk, self.order_b.pk})

    def test_reset_link_shown_only_with_active_filters(self):
        with_filter = self.client_http.get('/repair-orders/?status=repair')
        without = self.client_http.get('/repair-orders/')

        self.assertTrue(with_filter.context['filters_active'])
        self.assertFalse(without.context['filters_active'])

    def test_found_count_counts_all_matches_not_just_the_page(self):
        for i in range(30):
            RepairOrder.objects.create(client=self.alpha, status='accepted')

        resp = self.client_http.get('/repair-orders/?status=accepted')
        self.assertEqual(resp.context['found_count'], 30)
        self.assertEqual(len(resp.context['orders']), 25)

    def test_pagination_keeps_filters(self):
        for i in range(30):
            RepairOrder.objects.create(client=self.alpha, status='accepted')

        content = self.client_http.get('/repair-orders/?status=accepted').content.decode()
        self.assertIn('status=accepted', content)
        self.assertIn('page=2', content)


class CyrillicSearchTests(TestCase):
    """Поиск без учёта регистра для кириллицы.

    Встроенный LIKE в SQLite приводит к одному регистру только латиницу,
    поэтому «скрип» не находил «Скрип». Заменённая функция like() чинит это
    сразу везде, где используется icontains, — проверяем на разных разделах.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_cyr', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def test_lowercase_query_finds_capitalised_client(self):
        ClientModel.objects.create(name='ООО Лифтсервис')

        resp = self.client_http.get('/clients/?q=лифтсервис')
        self.assertEqual(len(resp.context['clients']), 1)

    def test_uppercase_query_finds_lowercase_part(self):
        SparePart.objects.create(part_number='C-1', name='конденсатор плёночный')

        resp = self.client_http.get('/parts/?q=КОНДЕНСАТОР')
        self.assertEqual(len(resp.context['parts']), 1)

    def test_mixed_case_finds_equipment_model(self):
        model = EquipmentModel.objects.create(name='Экодрайв')
        Equipment.objects.create(model=model, serial_number='E-1')

        resp = self.client_http.get('/equipment/?q=эКоДрАйВ')
        self.assertEqual(len(resp.context['equipment']), 1)

    def test_latin_search_still_works(self):
        SparePart.objects.create(part_number='RES-10', name='Resistor')

        resp = self.client_http.get('/parts/?q=resistor')
        self.assertEqual(len(resp.context['parts']), 1)

    def test_percent_in_query_is_not_a_wildcard(self):
        """Регрессия замены LIKE: символы шаблона внутри запроса Django
        экранирует, и они должны искаться буквально."""
        SparePart.objects.create(part_number='P-1', name='Резистор 5% точность')
        SparePart.objects.create(part_number='P-2', name='Резистор обычный')

        resp = self.client_http.get('/parts/?q=5%25')
        self.assertEqual([p.part_number for p in resp.context['parts']], ['P-1'])

    def test_underscore_in_query_is_not_a_wildcard(self):
        SparePart.objects.create(part_number='A_1', name='С подчёркиванием')
        SparePart.objects.create(part_number='AX1', name='Без него')

        resp = self.client_http.get('/parts/?q=A_1')
        self.assertEqual([p.part_number for p in resp.context['parts']], ['A_1'])

    def test_like_helper_handles_null(self):
        from core.signals import _unicode_like

        self.assertFalse(_unicode_like('%текст%', None))
        self.assertFalse(_unicode_like(None, 'текст'))


class EquipmentShortLinkTests(TestCase):
    """Короткий адрес единицы оборудования.

    Отдельную этикетку оборудования больше не печатают — нужную наклейку
    даёт заказ. Сам адрес оставлен: коды с уже наклеенных этикеток должны
    открываться и дальше.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_lbl', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        model = EquipmentModel.objects.create(name='БУАД')
        self.equipment = Equipment.objects.create(model=model, serial_number='БУАД-1234')

    def test_short_url_opens_repair_history(self):
        resp = self.client_http.get(f'/e/{self.equipment.pk}/')
        self.assertRedirects(resp, f'/equipment/{self.equipment.pk}/history/')

    def test_short_url_requires_login(self):
        resp = TestClient().get(f'/e/{self.equipment.pk}/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])

    def test_short_url_404_for_missing_equipment(self):
        resp = self.client_http.get('/e/999999/')
        self.assertEqual(resp.status_code, 404)

    def test_the_label_page_is_gone(self):
        """Печать этикетки оборудования вне заказа убрана в v2.26.0."""
        resp = self.client_http.get(f'/equipment/{self.equipment.pk}/label/')
        self.assertEqual(resp.status_code, 404)


class OrderLabelLinkTests(TestCase):
    """Этикетка оборудования в заказе: ссылка в QR вместо текста «LT-…/1»."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_olbl', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        client_obj = ClientModel.objects.create(name='Заказчик')
        model = EquipmentModel.objects.create(name='БУАД-3')
        self.order = RepairOrder.objects.create(client=client_obj)
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='БУАД-1234'),
        )

    def _label_url(self):
        return f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/label/'

    def test_short_url_opens_the_order(self):
        resp = self.client_http.get(f'/o/{self.order.pk}/')
        self.assertRedirects(resp, f'/repair-orders/{self.order.pk}/')

    def test_short_url_requires_login(self):
        resp = TestClient().get(f'/o/{self.order.pk}/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])

    def test_short_url_404_for_missing_order(self):
        self.assertEqual(self.client_http.get('/o/999999/').status_code, 404)

    def _encoded(self):
        with patch('core.views.generate_qr_image') as qr:
            qr.return_value = 'data:image/png;base64,x'
            self.client_http.get(self._label_url())
        return qr.call_args[0][0]

    def test_label_encodes_the_unit_not_the_whole_order(self):
        """Наклейка на приборе — про этот прибор. Раньше она вела на заказ
        целиком, и какая это единица, подсказывал только номер позиции:
        в код помещался адрес сервера, и на позицию знаков не оставалось."""
        encoded = self._encoded()

        self.assertEqual(encoded, f'u/{self.roe.pk}')
        # Прежнее содержимое — «LT-2026-08-001/1» — сканированием
        # не открывало ничего
        self.assertNotIn(self.order.order_number, encoded)

    @override_settings(LABEL_BASE_URL='http://192.168.1.50')
    def test_the_configured_address_never_reaches_the_code(self):
        """Внутренний адрес программы не печатается на наклейках, которые
        уезжают к заказчику. И код от длины адреса больше не зависит."""
        self.assertEqual(self._encoded(), f'u/{self.roe.pk}')

    def test_the_unit_code_opens_the_page_of_that_unit(self):
        """Наклейка на приборе ведёт на страницу этой единицы.

        До v2.76.0 она вела в карточку заказа к нужной строке — не от
        хорошей жизни: страницы единицы тогда просто не было, и всё
        про прибор лежало в трёх местах.
        """
        resp = self.client_http.get(f'/u/{self.roe.pk}/')

        self.assertRedirects(
            resp,
            f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/')

    def test_the_unit_row_carries_the_anchor(self):
        """Иначе переход приводит в начало страницы, и мастер ищет
        свою строку глазами."""
        resp = self.client_http.get(f'/repair-orders/{self.order.pk}/')

        self.assertContains(resp, f'id="unit-{self.roe.pk}"')

    def test_the_history_of_that_unit_is_two_clicks_away(self):
        """История ремонтов переехала со строки заказа на страницу
        единицы: в строке она нужна редко, а место занимала всегда."""
        row = self.client_http.get(f'/repair-orders/{self.order.pk}/')
        self.assertContains(
            row, f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/')

        unit = self.client_http.get(
            f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/')
        self.assertContains(unit, f'/equipment/{self.roe.equipment_id}/history/')

    def test_the_unit_code_requires_login(self):
        resp = TestClient().get(f'/u/{self.roe.pk}/')

        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])

    def test_the_unit_code_404_for_a_missing_unit(self):
        self.assertEqual(self.client_http.get('/u/999999/').status_code, 404)

    def test_position_still_printed_on_the_label(self):
        """Позиция в ссылку не входит, поэтому она обязана остаться на бумаге —
        в строке с номером заказа, вида «LT-2026-08-001/1»."""
        resp = self.client_http.get(self._label_url())
        content = resp.content.decode()

        self.assertContains(resp, self.order.order_number)
        self.assertIn('/1</span>', content)

    def test_the_label_has_no_emblem(self):
        """Кольцо с надписями занимало место вокруг кода и ограничивало его."""
        content = self.client_http.get(self._label_url()).content.decode()

        self.assertNotIn('<svg', content)
        self.assertNotIn('label-logo', content)

    def test_the_company_and_the_site_stay_on_the_label(self):
        """Эмблемы нет, но по этикетке должно быть видно, чья это железка."""
        response = self.client_http.get(self._label_url())

        self.assertContains(response, 'LIFT TEAM')
        self.assertContains(response, 'LIFTTEAM.RU')

    def test_the_phones_are_the_current_ones(self):
        response = self.client_http.get(self._label_url())

        self.assertContains(response, '+7 964 524 84 00')
        self.assertContains(response, '+7 977 760 10 89')
        self.assertNotContains(response, '282-40-31')

    def test_the_qr_is_the_common_size(self):
        """Размер кода один на всех этикетках: 12,3 мм."""
        response = self.client_http.get(self._label_url())

        self.assertContains(response, '12.3mm')


class OrderExportTests(TestCase):
    """Выгрузка списка заказов в Excel — с учётом фильтров страницы."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_oexp', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.alpha = ClientModel.objects.create(name='ООО Альфа', inn='7700000001')
        model = EquipmentModel.objects.create(name='БУАД')

        self.paid = RepairOrder.objects.create(
            client=self.alpha, status='shipped', payment_status='paid',
            invoice_number='СЧ-1',
        )
        RepairOrderEquipment.objects.create(
            repair_order=self.paid,
            equipment=Equipment.objects.create(model=model, serial_number='SN-1'),
            repair_cost=1000,
        )

        self.unpaid = RepairOrder.objects.create(
            client=self.alpha, status='repair', payment_status='unpaid',
        )
        for serial, cost in (('SN-2', 2000), ('SN-3', 500)):
            RepairOrderEquipment.objects.create(
                repair_order=self.unpaid,
                equipment=Equipment.objects.create(model=model, serial_number=serial),
                repair_cost=cost,
            )

    def _rows(self, query=''):
        resp = self.client_http.get(f'/repair-orders/export/{query}')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        headers = [c.value for c in ws[1]]
        return [dict(zip(headers, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]

    def test_export_lists_all_orders_with_total(self):
        rows = self._rows()

        self.assertEqual(len(rows), 3)  # два заказа плюс «Итого»
        self.assertEqual(rows[-1]['№ заказа'], 'Итого')
        self.assertEqual(float(rows[-1]['Сумма, ₽']), 3500.0)

    def test_export_respects_filters(self):
        rows = self._rows('?payment_status=unpaid')

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['№ заказа'], self.unpaid.order_number)
        self.assertEqual(float(rows[0]['Сумма, ₽']), 2500.0)

    def test_export_sums_all_units_of_an_order(self):
        """Регрессия: соединение с оборудованием при фильтрах могло задвоить сумму."""
        rows = self._rows(f'?q=Альфа&status=repair')

        self.assertEqual(len(rows), 2)
        self.assertEqual(float(rows[0]['Сумма, ₽']), 2500.0)

    def test_export_lists_every_unit_of_an_order(self):
        rows = self._rows('?payment_status=unpaid')
        self.assertIn('SN-2', rows[0]['Оборудование'])
        self.assertIn('SN-3', rows[0]['Оборудование'])

    def test_empty_result_has_no_total_row(self):
        ws_rows = self._rows('?q=такого-заказа-нет')
        self.assertEqual(ws_rows, [])

    def test_export_requires_login(self):
        resp = TestClient().get('/repair-orders/export/')
        self.assertEqual(resp.status_code, 302)

    def test_order_list_page_offers_export(self):
        resp = self.client_http.get('/repair-orders/')
        self.assertContains(resp, '/repair-orders/export/')


class EquipmentListFilterTests(TestCase):
    """Поиск по заказчику, фильтр по гарантии и выгрузка списка оборудования."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_eqf', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.alpha = ClientModel.objects.create(name='ООО Альфа')
        model = EquipmentModel.objects.create(name='БУАД')

        self.covered = Equipment.objects.create(
            model=model, serial_number='SN-COVERED', current_client=self.alpha
        )
        self.expired = Equipment.objects.create(model=model, serial_number='SN-EXPIRED')
        self.never = Equipment.objects.create(model=model, serial_number='SN-NEVER')

        self._repair(self.covered, days_ago=10)
        self._repair(self.expired, days_ago=500)

    def _repair(self, equipment, days_ago):
        order = RepairOrder.objects.create(
            client=self.alpha, status='shipped'
        )
        RepairOrder.objects.filter(pk=order.pk).update(
            date_completed=timezone.now() - datetime.timedelta(days=days_ago)
        )
        return RepairOrderEquipment.objects.create(
            repair_order=order, equipment=equipment, repair_cost=100
        )

    def _serials(self, query=''):
        resp = self.client_http.get(f'/equipment/{query}')
        self.assertEqual(resp.status_code, 200)
        return {eq.serial_number for eq in resp.context['equipment']}

    def test_search_by_current_client(self):
        self.assertEqual(self._serials('?q=Альфа'), {'SN-COVERED'})

    def test_filter_shows_only_covered_equipment(self):
        self.assertEqual(self._serials('?warranty=active'), {'SN-COVERED'})

    def test_filter_shows_equipment_without_warranty(self):
        self.assertEqual(self._serials('?warranty=expired'), {'SN-EXPIRED', 'SN-NEVER'})

    def test_no_filter_shows_everything(self):
        self.assertEqual(len(self._serials()), 3)

    def test_unknown_filter_value_is_ignored(self):
        self.assertEqual(len(self._serials('?warranty=что-то')), 3)

    @override_settings(WARRANTY_MONTHS=0)
    def test_filter_does_nothing_when_warranty_is_off(self):
        self.assertEqual(len(self._serials('?warranty=active')), 3)

    def test_equipment_repaired_twice_is_listed_once(self):
        self._repair(self.covered, days_ago=20)
        self.assertEqual(len(self._serials('?warranty=active')), 1)

    def test_export_includes_warranty_and_repair_count(self):
        self._repair(self.covered, days_ago=20)

        resp = self.client_http.get('/equipment/export/?warranty=active')
        self.assertEqual(resp.status_code, 200)
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        headers = [c.value for c in ws[1]]
        row = dict(zip(headers, [c.value for c in ws[2]]))

        self.assertEqual(row['Серийный номер'], 'SN-COVERED')
        self.assertEqual(row['Текущий заказчик'], 'ООО Альфа')
        self.assertEqual(row['Всего ремонтов'], 2)
        # openpyxl читает ячейку с датой обратно как datetime, сравниваем по дате
        self.assertEqual(
            row['Гарантия до'].date(),
            timezone.localtime(self.covered.active_warranty().warranty_until).date(),
        )
        self.assertEqual(ws.max_row, 2)  # шапка и одна единица

    def test_export_requires_login(self):
        self.assertEqual(TestClient().get('/equipment/export/').status_code, 302)


class ClientExportTests(TestCase):
    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_cexp', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.alpha = ClientModel.objects.create(name='ООО Альфа', inn='7700000001')
        ClientModel.objects.create(name='ООО Бета', inn='7700000002')

        model = EquipmentModel.objects.create(name='БУАД')
        for cost, payment in ((1000, 'unpaid'), (2000, 'paid')):
            order = RepairOrder.objects.create(client=self.alpha, payment_status=payment)
            RepairOrderEquipment.objects.create(
                repair_order=order,
                equipment=Equipment.objects.create(model=model, serial_number=f'SN-{cost}'),
                repair_cost=cost,
            )

    def _rows(self, query=''):
        resp = self.client_http.get(f'/clients/export/{query}')
        self.assertEqual(resp.status_code, 200)
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        headers = [c.value for c in ws[1]]
        return [dict(zip(headers, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]

    def test_export_counts_orders_and_debt(self):
        rows = {r['Название']: r for r in self._rows()}

        self.assertEqual(rows['ООО Альфа']['Заказов'], 2)
        # В долг попадает только неоплаченный заказ
        self.assertEqual(float(rows['ООО Альфа']['Долг, ₽']), 1000.0)
        self.assertEqual(rows['ООО Бета']['Заказов'], 0)
        self.assertEqual(float(rows['ООО Бета']['Долг, ₽']), 0.0)

    def test_export_respects_search(self):
        rows = self._rows('?q=Бета')
        self.assertEqual([r['Название'] for r in rows], ['ООО Бета'])

    def test_export_requires_login(self):
        self.assertEqual(TestClient().get('/clients/export/').status_code, 302)


class StockStateTests(TestCase):
    """Три состояния остатка вместо двух несогласованных определений.

    Раньше «мало на складе» считалось по-разному: сетка кассетниц красила
    ячейку по «остаток <= минимума», а дашборд, план закупок и список деталей
    отбирали строго ниже минимума. Деталь ровно на минимуме выглядела
    дефицитной, но в план закупок не попадала.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_stock', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.below = SparePart.objects.create(
            part_number='ST-BELOW', name='Дефицит', current_stock=2, min_stock=5
        )
        self.at_min = SparePart.objects.create(
            part_number='ST-AT', name='На минимуме', current_stock=5, min_stock=5
        )
        self.ok = SparePart.objects.create(
            part_number='ST-OK', name='В норме', current_stock=9, min_stock=5
        )
        self.no_min = SparePart.objects.create(
            part_number='ST-NOMIN', name='Без минимума', current_stock=0, min_stock=0
        )

    # --- состояние детали ---

    def test_states(self):
        self.assertEqual(self.below.stock_state, 'below')
        self.assertEqual(self.at_min.stock_state, 'at_minimum')
        self.assertEqual(self.ok.stock_state, 'ok')

    def test_zero_minimum_is_not_a_shortage(self):
        """Минимум не задан — деталь не считается ни дефицитной, ни на минимуме."""
        self.assertEqual(self.no_min.stock_state, 'ok')

    def test_is_below_min_stock_matches_the_state(self):
        self.assertTrue(self.below.is_below_min_stock())
        self.assertFalse(self.at_min.is_below_min_stock())

    # --- отбор ---

    def test_queryset_below_minimum(self):
        found = SparePart.objects.below_minimum().values_list('part_number', flat=True)
        self.assertEqual(set(found), {'ST-BELOW'})

    def test_queryset_at_minimum(self):
        found = SparePart.objects.at_minimum().values_list('part_number', flat=True)
        self.assertEqual(set(found), {'ST-AT'})

    def test_purchase_plan_covers_only_shortage(self):
        resp = self.client_http.get('/reports/purchase-plan/')
        listed = {p.part_number for p in resp.context['parts']}
        self.assertEqual(listed, {'ST-BELOW'})

    # --- дашборд ---

    def test_dashboard_separates_shortage_from_at_minimum(self):
        resp = self.client_http.get('/')

        self.assertEqual(resp.context['low_stock_count'], 1)
        self.assertEqual(resp.context['at_minimum_count'], 1)
        self.assertEqual(
            {p.part_number for p in resp.context['low_stock_parts']}, {'ST-BELOW'}
        )
        self.assertEqual(
            {p.part_number for p in resp.context['at_minimum_parts']}, {'ST-AT'}
        )

    def test_dashboard_shows_both_groups(self):
        content = self.client_http.get('/').content.decode()
        self.assertIn('ST-BELOW', content)
        self.assertIn('ST-AT', content)
        self.assertIn('на минимуме', content)

    # --- ячейки ---

    def _cell_with(self, *parts):
        cell = StorageCell.objects.create(
            cabinet=Cabinet.objects.get_or_create(number=1)[0], row_number=1, cell_row=StorageCell.objects.count() + 1
        )
        cell.parts.add(*parts)
        return cell

    def test_cell_with_shortage_is_red(self):
        self.assertEqual(self._cell_with(self.below).get_status(), 'low_stock')

    def test_cell_at_minimum_is_its_own_status(self):
        self.assertEqual(self._cell_with(self.at_min).get_status(), 'at_minimum')

    def test_cell_shows_the_worst_state_of_its_parts(self):
        self.assertEqual(self._cell_with(self.at_min, self.below).get_status(), 'low_stock')

    def test_normal_and_empty_cells(self):
        self.assertEqual(self._cell_with(self.ok).get_status(), 'normal')
        self.assertEqual(self._cell_with().get_status(), 'free')


class PartStockFilterTests(TestCase):
    """Фильтры списка деталей: состояние остатка, диапазоны, устойчивость."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_pf', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        SparePart.objects.create(part_number='PF-BELOW', name='Дефицит',
                                 current_stock=1, min_stock=5, voltage=12, current=2, power=1)
        SparePart.objects.create(part_number='PF-AT', name='На минимуме',
                                 current_stock=5, min_stock=5, voltage=24, current=5, power=10)
        SparePart.objects.create(part_number='PF-OK', name='В норме',
                                 current_stock=50, min_stock=5, voltage=48, current=10, power=100)

    def _found(self, query=''):
        resp = self.client_http.get(f'/parts/{query}')
        self.assertEqual(resp.status_code, 200)
        return {p.part_number for p in resp.context['parts']}

    def test_filter_by_package(self):
        """Корпус — такой же признак отбора, как тип: 0805 вместо 1206
        на плату не встанет, как бы ни совпадали характеристики."""
        SparePart.objects.filter(part_number='PF-OK').update(package='TO-220')
        SparePart.objects.filter(part_number='PF-AT').update(package='0805')

        self.assertEqual(self._found('?package=TO-220'), {'PF-OK'})

    def test_search_finds_the_package(self):
        SparePart.objects.filter(part_number='PF-OK').update(package='SOT-23')

        self.assertEqual(self._found('?q=SOT-23'), {'PF-OK'})

    def test_only_shortage(self):
        self.assertEqual(self._found('?stock_state=below'), {'PF-BELOW'})

    def test_only_at_minimum(self):
        self.assertEqual(self._found('?stock_state=at_minimum'), {'PF-AT'})

    def test_attention_covers_both(self):
        self.assertEqual(self._found('?stock_state=attention'), {'PF-BELOW', 'PF-AT'})

    def test_old_below_min_link_still_works(self):
        """Ссылки вида ?below_min=1 могли остаться в закладках."""
        self.assertEqual(self._found('?below_min=1'), {'PF-BELOW'})

    def test_stock_range(self):
        self.assertEqual(self._found('?stock_from=2&stock_to=10'), {'PF-AT'})

    def test_current_range_now_reaches_the_view(self):
        self.assertEqual(self._found('?current_from=4&current_to=6'), {'PF-AT'})

    def test_power_range_is_supported(self):
        self.assertEqual(self._found('?power_from=50'), {'PF-OK'})

    def test_ranges_combine_with_state(self):
        self.assertEqual(self._found('?stock_state=attention&voltage_from=20'), {'PF-AT'})

    def test_garbage_in_range_does_not_break_the_page(self):
        """Регрессия: голый float() на «5 В» ронял страницу целиком."""
        self.assertEqual(len(self._found('?voltage_from=5 В&stock_from=много')), 3)

    def test_comma_decimal_is_accepted(self):
        self.assertEqual(self._found('?voltage_from=23,5&voltage_to=24,5'), {'PF-AT'})

    def test_export_uses_the_same_filters(self):
        resp = self.client_http.get('/parts/export/?stock_state=below')
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        self.assertEqual(ws.max_row, 2)  # шапка и одна деталь
        self.assertEqual(ws.cell(row=2, column=1).value, 'PF-BELOW')


class EquipmentHistoryExportTests(TestCase):
    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_hexp', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.client_obj = ClientModel.objects.create(name='ООО Альфа')
        model = EquipmentModel.objects.create(name='БУАД')
        self.equipment = Equipment.objects.create(model=model, serial_number='SN-100')
        self.other = Equipment.objects.create(model=model, serial_number='SN-200')

        self.order = RepairOrder.objects.create(
            client=self.client_obj, status='shipped'
        )
        RepairOrder.objects.filter(pk=self.order.pk).update(
            date_completed=timezone.now() - datetime.timedelta(days=5)
        )
        self.visit = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment,
            fault_description='Не крутится', seal_numbers='П-1, П-2',
            initial_condition='Корпус цел', repair_cost=3000,
        )
        part = SparePart.objects.create(part_number='D-1', name='Диод', current_stock=10)
        RepairOrderDetail.objects.create(repair_order=self.order, part=part, quantity_used=2)

    def _rows(self, equipment=None):
        target = equipment or self.equipment
        resp = self.client_http.get(f'/equipment/{target.pk}/history/export/')
        self.assertEqual(resp.status_code, 200)
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        headers = [c.value for c in ws[1]]
        return [dict(zip(headers, [c.value for c in row])) for row in ws.iter_rows(min_row=2)]

    def test_export_contains_the_visit(self):
        row = self._rows()[0]

        self.assertEqual(row['№ заказа'], self.order.order_number)
        self.assertEqual(row['Заказчик'], 'ООО Альфа')
        self.assertEqual(row['Неисправность'], 'Не крутится')
        self.assertEqual(row['Номера пломб'], 'П-1, П-2')
        self.assertEqual(float(row['Стоимость, ₽']), 3000.0)
        self.assertIn('Диод x2', row['Детали'])

    def test_parts_are_marked_as_order_wide_for_multi_unit_orders(self):
        RepairOrderEquipment.objects.create(repair_order=self.order, equipment=self.other)

        row = self._rows()[0]
        self.assertIn('на весь заказ', row['Детали'])

    def test_export_covers_only_the_requested_equipment(self):
        self.assertEqual(self._rows(self.other), [])

    def test_warranty_column_filled_for_completed_order(self):
        row = self._rows()[0]
        self.assertIsNotNone(row['Гарантия до'])

    def test_history_page_offers_export(self):
        resp = self.client_http.get(f'/equipment/{self.equipment.pk}/history/')
        self.assertContains(resp, f'/equipment/{self.equipment.pk}/history/export/')

    def test_export_requires_login(self):
        resp = TestClient().get(f'/equipment/{self.equipment.pk}/history/export/')
        self.assertEqual(resp.status_code, 302)


class BatchLabelTests(TestCase):
    """Печать этикеток пачкой: отбор, ограничение, раскладка."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_batch', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.resistor = SparePart.objects.create(
            part_number='B-RES', name='Резистор', component_type='Резистор', current_stock=5
        )
        self.diode = SparePart.objects.create(
            part_number='B-DIO', name='Диод', component_type='Диод', current_stock=5
        )

        self.filled = StorageCell.objects.create(cabinet=Cabinet.objects.get_or_create(number=1)[0], row_number=1, cell_row=1)
        self.filled.parts.add(self.resistor)
        self.empty = StorageCell.objects.create(cabinet=Cabinet.objects.get_or_create(number=1)[0], row_number=1, cell_row=2)
        StorageCell.objects.create(cabinet=Cabinet.objects.get_or_create(number=2)[0], row_number=1, cell_row=1)

    def _labels(self, url):
        resp = self.client_http.get(url)
        self.assertEqual(resp.status_code, 200)
        return resp, resp.context['labels']

    # --- детали ---

    def test_selected_parts_only(self):
        _, labels = self._labels(f'/parts/labels/?ids={self.diode.pk}')

        self.assertEqual([item['part'].part_number for item in labels], ['B-DIO'])

    def test_several_selected_parts(self):
        _, labels = self._labels(f'/parts/labels/?ids={self.diode.pk}&ids={self.resistor.pk}')
        self.assertEqual(len(labels), 2)

    def test_without_selection_falls_back_to_the_current_filter(self):
        """Кнопка «на все отобранные» передаёт условия списка, а не номера."""
        _, labels = self._labels('/parts/labels/?q=Диод')

        self.assertEqual([item['part'].part_number for item in labels], ['B-DIO'])

    def test_without_selection_and_filter_takes_everything(self):
        _, labels = self._labels('/parts/labels/')
        self.assertEqual(len(labels), 2)

    def test_garbage_in_ids_is_ignored(self):
        _, labels = self._labels('/parts/labels/?ids=abc')
        # мусор отброшен, отбора нет — печатается всё
        self.assertEqual(len(labels), 2)

    def test_every_label_has_its_own_qr(self):
        _, labels = self._labels('/parts/labels/')
        codes = {item['qr_img'] for item in labels}

        self.assertEqual(len(codes), 2)
        self.assertTrue(all(code for code in codes))

    def test_batch_is_capped(self):
        from core.views import MAX_LABELS_PER_BATCH

        SparePart.objects.bulk_create([
            SparePart(part_number=f'CAP-{i:04d}', name='Деталь')
            for i in range(MAX_LABELS_PER_BATCH + 5)
        ])
        _, labels = self._labels('/parts/labels/')

        self.assertEqual(len(labels), MAX_LABELS_PER_BATCH)

    # --- ячейки ---

    def test_whole_cabinet(self):
        _, labels = self._labels('/storage-cells/labels/?cabinet=1')
        self.assertEqual({item['cell'].address for item in labels}, {'К1-Р1-Я1', 'К1-Р1-Я2'})

    def test_only_filled_cells(self):
        _, labels = self._labels('/storage-cells/labels/?cabinet=1&only_filled=1')
        self.assertEqual([item['cell'].address for item in labels], ['К1-Р1-Я1'])

    def test_cabinet_filter_excludes_other_cabinets(self):
        _, labels = self._labels('/storage-cells/labels/?cabinet=2')
        self.assertEqual([item['cell'].address for item in labels], ['К2-Р1-Я1'])

    def test_selected_cells_win_over_cabinet(self):
        _, labels = self._labels(f'/storage-cells/labels/?ids={self.empty.pk}&cabinet=2')
        self.assertEqual([item['cell'].address for item in labels], ['К1-Р1-Я2'])

    def test_without_cabinet_nothing_is_printed(self):
        """Иначе одна опечатка в адресе вывела бы этикетки на все 768 ячеек."""
        _, labels = self._labels('/storage-cells/labels/')
        self.assertEqual(labels, [])

    def test_cell_label_keeps_grouping_logic(self):
        """Пачка и одиночная печать используют одну сборку данных."""
        second = SparePart.objects.create(
            part_number='B-RES2', name='Резистор 2', component_type='Резистор'
        )
        self.filled.parts.add(second)

        _, labels = self._labels('/storage-cells/labels/?cabinet=1&only_filled=1')
        self.assertEqual(labels[0]['title'], 'Набор резисторов')

    # --- раскладка и доступ ---

    def test_roll_layout_by_default(self):
        resp, _ = self._labels('/parts/labels/')

        self.assertEqual(resp.context['layout'], 'roll')
        self.assertContains(resp, 'size: 43mm 25mm')
        self.assertContains(resp, 'page-break-after: always')

    def test_a4_layout(self):
        resp, _ = self._labels('/parts/labels/?layout=a4')

        self.assertEqual(resp.context['layout'], 'a4')
        self.assertContains(resp, 'size: A4')
        self.assertNotContains(resp, 'page-break-after: always')

    def test_unknown_layout_falls_back_to_roll(self):
        resp, _ = self._labels('/parts/labels/?layout=что-то')
        self.assertEqual(resp.context['layout'], 'roll')

    def test_batch_pages_require_login(self):
        anon = TestClient()
        self.assertEqual(anon.get('/parts/labels/').status_code, 302)
        self.assertEqual(anon.get('/storage-cells/labels/').status_code, 302)

    def test_entry_points_are_on_the_pages(self):
        parts_page = self.client_http.get('/parts/')
        grid_page = self.client_http.get('/storage-cells/?cabinet=1')

        self.assertContains(parts_page, '/parts/labels/')
        self.assertContains(grid_page, '/storage-cells/labels/')


class StockSocketTests(TestCase):
    """Канал живых обновлений остатка: доступ и рассылка.

    Серверная часть существовала с самого начала, но к ней никто
    не подключался — сообщения уходили в пустоту, а сам сокет пускал любого.
    """

    def setUp(self):
        self.user = Employee.objects.create_user(
            username='ws_user', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.part = SparePart.objects.create(
            part_number='WS-1', name='Деталь', current_stock=10, min_stock=3
        )

    def _communicator(self, user=None):
        from channels.testing import WebsocketCommunicator
        from lifteam.asgi import websocket_urlpatterns
        from channels.routing import URLRouter

        communicator = WebsocketCommunicator(URLRouter(websocket_urlpatterns), '/ws/stock/')
        # AuthMiddlewareStack в тестах подменяем прямой подстановкой:
        # проверяем поведение потребителя, а не саму сессионную прослойку
        communicator.scope['user'] = user
        return communicator

    async def _connect(self, user):
        communicator = self._communicator(user)
        connected, _ = await communicator.connect()
        return communicator, connected

    def test_anonymous_connection_is_refused(self):
        from asgiref.sync import async_to_sync
        from django.contrib.auth.models import AnonymousUser

        async def run():
            communicator, connected = await self._connect(AnonymousUser())
            await communicator.disconnect()
            return connected

        self.assertFalse(async_to_sync(run)())

    def test_connection_without_user_is_refused(self):
        from asgiref.sync import async_to_sync

        async def run():
            communicator, connected = await self._connect(None)
            await communicator.disconnect()
            return connected

        self.assertFalse(async_to_sync(run)())

    def test_authenticated_user_gets_updates(self):
        from asgiref.sync import async_to_sync
        from channels.layers import get_channel_layer

        async def run():
            communicator, connected = await self._connect(self.user)
            greeting = await communicator.receive_json_from()

            layer = get_channel_layer()
            await layer.group_send('stock_updates', {
                'type': 'stock_update',
                'data': {'part_id': self.part.pk, 'part_number': 'WS-1',
                         'current_stock': 2, 'min_stock': 3},
            })
            update = await communicator.receive_json_from()
            await communicator.disconnect()
            return connected, greeting, update

        connected, greeting, update = async_to_sync(run)()

        self.assertTrue(connected)
        self.assertEqual(greeting['type'], 'connection_established')
        self.assertEqual(update['type'], 'stock_update')
        self.assertEqual(update['data']['current_stock'], 2)


class StockWatchMarkupTests(TestCase):
    """Разметка, по которой скрипт находит цифры остатка на странице."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_ws', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)
        self.part = SparePart.objects.create(
            part_number='WS-2', name='Деталь', current_stock=7, min_stock=3
        )

    def test_part_list_marks_stock(self):
        resp = self.client_http.get('/parts/')
        self.assertContains(resp, f'data-stock-part="{self.part.pk}"')

    def test_part_detail_marks_stock(self):
        resp = self.client_http.get(f'/parts/{self.part.pk}/')
        self.assertContains(resp, f'data-stock-part="{self.part.pk}"')

    def test_script_and_toast_holder_are_on_the_page(self):
        content = self.client_http.get('/parts/').content.decode()
        self.assertIn('js/stock-updates.js', content)
        self.assertIn('id="stockToasts"', content)

    def test_reconnect_helper_is_loaded_before_the_stock_script(self):
        """Логика переподключения вынесена в общий файл и должна грузиться
        раньше того, кто ей пользуется, — иначе живых обновлений не будет."""
        content = self.client_http.get('/parts/').content.decode()
        self.assertIn('js/ws-connection.js', content)
        self.assertLess(content.index('js/ws-connection.js'), content.index('js/stock-updates.js'))


class PresenceStatusTests(TestCase):
    """Кто сейчас в сети — расчёт статуса от отметки активности.

    Сокет здесь не нужен: «в сети» — это функция от last_seen и таймаута,
    и проверять её быстрее и честнее прямо на модели.
    """

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='presence_user', full_name='Кладовщик', password='pass', role='warehouse'
        )

    def _seen(self, seconds_ago):
        Employee.objects.filter(pk=self.employee.pk).update(
            last_seen=timezone.now() - datetime.timedelta(seconds=seconds_ago)
        )
        self.employee.refresh_from_db()

    def test_recent_activity_is_online(self):
        self._seen(30)
        self.assertTrue(self.employee.is_online)

    def test_activity_past_timeout_is_offline(self):
        self._seen(settings.PRESENCE_TIMEOUT_SECONDS + 1)
        self.assertFalse(self.employee.is_online)

    def test_never_seen_is_offline(self):
        self.assertIsNone(self.employee.last_seen)
        self.assertFalse(self.employee.is_online)

    def test_two_missed_heartbeats_are_still_online(self):
        """Смысл выбранного таймаута: два подряд пропущенных сигнала —
        ещё не разрыв. Иначе индикатор мигал бы на каждой заминке связи."""
        heartbeat = settings.PRESENCE_TIMEOUT_SECONDS // 3
        self._seen(heartbeat * 2 + 1)
        self.assertTrue(self.employee.is_online)

    def test_touch_presence_writes_without_touching_other_fields(self):
        Employee.objects.filter(pk=self.employee.pk).update(full_name='Переименован')
        self.employee.touch_presence()

        fresh = Employee.objects.get(pk=self.employee.pk)
        self.assertIsNotNone(fresh.last_seen)
        self.assertTrue(fresh.is_online)
        # Отметка пишется запросом к набору, поэтому чужая правка не затирается
        self.assertEqual(fresh.full_name, 'Переименован')

    @override_settings(PRESENCE_TIMEOUT_SECONDS=10)
    def test_timeout_is_configurable(self):
        self._seen(30)
        self.assertFalse(self.employee.is_online)


class PresencePageTests(TestCase):
    """Страница «Кто на связи»: доступ и содержимое."""

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='presence_page', full_name='Кладовщик Иванов', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.employee)

    def test_anonymous_is_redirected_to_login(self):
        anon = TestClient()
        resp = anon.get('/presence/')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])

    def test_non_admin_role_can_see_presence(self):
        """Присутствие видно всем вошедшим: это не надзор, а способ
        не идти через лабораторию к пустому терминалу."""
        resp = self.client_http.get('/presence/')
        self.assertEqual(resp.status_code, 200)
        self.assertIn(self.employee, list(resp.context['employees']))

    def test_page_shows_online_state_from_the_database(self):
        Employee.objects.filter(pk=self.employee.pk).update(last_seen=timezone.now())
        content = self.client_http.get('/presence/').content.decode()
        self.assertIn('presence-online', content)
        self.assertIn('в сети', content)

    def test_inactive_employee_is_not_listed(self):
        gone = Employee.objects.create_user(
            username='presence_gone', full_name='Уволенный', password='pass', is_active=False
        )
        resp = self.client_http.get('/presence/')
        self.assertNotIn(gone, list(resp.context['employees']))

    def test_presence_script_is_loaded_on_every_page(self):
        """Отметку ставит любая открытая страница, а не только эта:
        иначе мастер, весь день сидящий в заказах, числился бы ушедшим."""
        content = self.client_http.get('/repair-orders/').content.decode()
        self.assertIn('js/presence.js', content)


class PresenceSocketTests(TestCase):
    """Канал присутствия: доступ, отметка активности, рассылка списка."""

    def setUp(self):
        self.user = Employee.objects.create_user(
            username='presence_ws', full_name='Кладовщик', password='pass', role='warehouse'
        )

    def _communicator(self, user=None):
        from channels.testing import WebsocketCommunicator
        from lifteam.asgi import websocket_urlpatterns
        from channels.routing import URLRouter

        communicator = WebsocketCommunicator(URLRouter(websocket_urlpatterns), '/ws/presence/')
        # Как и в тестах остатков: проверяем потребителя, а не сессионную
        # прослойку, поэтому пользователя подставляем прямо в scope
        communicator.scope['user'] = user
        return communicator

    def test_anonymous_connection_is_refused(self):
        from asgiref.sync import async_to_sync
        from django.contrib.auth.models import AnonymousUser

        async def run():
            communicator = self._communicator(AnonymousUser())
            connected, _ = await communicator.connect()
            await communicator.disconnect()
            return connected

        self.assertFalse(async_to_sync(run)())

    def test_connection_without_user_is_refused(self):
        from asgiref.sync import async_to_sync

        async def run():
            communicator = self._communicator(None)
            connected, _ = await communicator.connect()
            await communicator.disconnect()
            return connected

        self.assertFalse(async_to_sync(run)())

    def test_live_connection_makes_employee_online(self):
        from asgiref.sync import async_to_sync

        async def run():
            communicator = self._communicator(self.user)
            connected, _ = await communicator.connect()
            greeting = await communicator.receive_json_from()
            roster = await communicator.receive_json_from()
            await communicator.disconnect()
            return connected, greeting, roster

        connected, greeting, roster = async_to_sync(run)()

        self.assertTrue(connected)
        self.assertEqual(greeting['type'], 'connection_established')
        # Период сигнала задаёт сервер, чтобы он не разошёлся с таймаутом
        self.assertEqual(greeting['heartbeat_seconds'], settings.PRESENCE_TIMEOUT_SECONDS // 3)
        self.assertEqual(roster['type'], 'presence_roster')

        entry = [row for row in roster['roster'] if row['id'] == self.user.pk][0]
        self.assertTrue(entry['online'])
        self.assertEqual(entry['full_name'], 'Кладовщик')

        self.user.refresh_from_db()
        self.assertTrue(self.user.is_online)

    def test_heartbeat_refreshes_the_mark(self):
        from asgiref.sync import async_to_sync

        stale = timezone.now() - datetime.timedelta(seconds=settings.PRESENCE_TIMEOUT_SECONDS + 60)

        async def run():
            communicator = self._communicator(self.user)
            await communicator.connect()
            await communicator.receive_json_from()  # приветствие
            await communicator.receive_json_from()  # список
            # Состарим отметку, будто соединение молчало дольше таймаута
            await database_sync_to_async(
                Employee.objects.filter(pk=self.user.pk).update
            )(last_seen=stale)

            await communicator.send_json_to({'action': 'ping'})
            pong = await communicator.receive_json_from()
            roster = await communicator.receive_json_from()
            await communicator.disconnect()
            return pong, roster

        pong, roster = async_to_sync(run)()

        self.assertEqual(pong['type'], 'pong')
        self.assertEqual(roster['type'], 'presence_roster')
        self.user.refresh_from_db()
        self.assertTrue(self.user.is_online)

    def test_drop_without_logout_expires_by_timeout_only(self):
        """Обрыв связи не гасит индикатор мгновенно — обрыв неотличим
        от заминки в сети. Статус снимает истёкший таймаут, и только он.
        """
        from asgiref.sync import async_to_sync

        async def run():
            communicator = self._communicator(self.user)
            await communicator.connect()
            await communicator.receive_json_from()
            await communicator.receive_json_from()
            await communicator.disconnect()

        async_to_sync(run)()

        self.user.refresh_from_db()
        self.assertTrue(self.user.is_online)

        # ...но и бесконечно «в сети» никто не висит
        Employee.objects.filter(pk=self.user.pk).update(
            last_seen=timezone.now() - datetime.timedelta(seconds=settings.PRESENCE_TIMEOUT_SECONDS + 1)
        )
        self.user.refresh_from_db()
        self.assertFalse(self.user.is_online)


class DashboardStatusStatsTests(TestCase):
    """Разбивка заказов по статусам на дашборде.

    Данные считались с самого начала, но в шаблон не попадали — на экране
    их не было вовсе.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_stats', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        client_obj = ClientModel.objects.create(name='Заказчик')
        for status in ('repair', 'repair', 'shipped'):
            RepairOrder.objects.create(client=client_obj, status=status)

    def test_counts_by_status(self):
        resp = self.client_http.get('/')
        counts = {item['code']: item['count'] for item in resp.context['status_stats']}

        self.assertEqual(counts['repair'], 2)
        self.assertEqual(counts['shipped'], 1)

    def test_statuses_without_orders_are_shown_as_zero(self):
        """Иначе набор плиток менялся бы от загрузки к загрузке."""
        resp = self.client_http.get('/')
        codes = [item['code'] for item in resp.context['status_stats']]

        self.assertEqual(codes, [code for code, _ in RepairOrder.STATUS_CHOICES])
        counts = {item['code']: item['count'] for item in resp.context['status_stats']}
        self.assertEqual(counts['diagnostic'], 0)

    def test_tiles_link_to_the_filtered_order_list(self):
        content = self.client_http.get('/').content.decode()
        self.assertIn('/repair-orders/?status=repair', content)

    def test_link_actually_filters(self):
        resp = self.client_http.get('/repair-orders/?status=repair')
        self.assertEqual(len(resp.context['orders']), 2)


class GridLiveUpdateMarkupTests(TestCase):
    """Сетка кассетниц отдаёт данные, по которым скрипт перекрашивает ячейки."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_grid', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        call_command('init_cells', verbosity=0)
        self.part = SparePart.objects.create(
            part_number='GRID-1', name='Деталь', current_stock=10, min_stock=2
        )
        cell = StorageCell.objects.filter(cabinet__number=1).first()
        cell.parts.add(self.part)

    def test_grid_exposes_cells_data_to_the_script(self):
        content = self.client_http.get('/storage-cells/?cabinet=1').content.decode()

        self.assertIn('window.CELLS_DATA = CELLS_DATA', content)
        self.assertIn('js/stock-updates.js', content)

    def test_cells_carry_their_id(self):
        content = self.client_http.get('/storage-cells/?cabinet=1').content.decode()
        self.assertIn('data-cell-id=', content)


@override_settings(NOTIFY_CLIENTS=True)
class OrderNotificationTests(TestCase):
    """Оповещения заказчику о смене статуса заказа."""

    def setUp(self):
        self.user = Employee.objects.create_superuser(
            username='admin_notify', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.user)

        self.client_obj = ClientModel.objects.create(
            name='ООО Альфа', email='client@example.com'
        )
        self.order = RepairOrder.objects.create(client=self.client_obj)
        model = EquipmentModel.objects.create(name='БУАД')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-1'),
        )

    def _change_status(self, status):
        return self.client_http.post(
            f'/repair-orders/{self.order.pk}/change-status/',
            {'new_status': status, 'notes': ''},
        )

    def test_ready_for_shipment_queues_a_letter(self):
        self._change_status('ready_for_shipment')

        note = Notification.objects.get(event='order_status')
        self.assertEqual(note.recipient, 'client@example.com')
        self.assertEqual(note.status, 'pending')
        self.assertIn(self.order.order_number, note.subject)
        self.assertIn('SN-1', note.body)
        self.assertEqual(note.repair_order, self.order)

    def test_internal_statuses_do_not_bother_the_client(self):
        """«Диагностика» и «ремонт» заказчику ничего не говорят."""
        self._change_status('diagnostic')
        self._change_status('repair')

        self.assertEqual(Notification.objects.count(), 0)

    def test_shipped_letter_carries_the_tracking_number(self):
        self.order.tracking_number = 'TRK-77'
        self.order.save(update_fields=['tracking_number'])
        self._change_status('shipped')

        self.assertIn('TRK-77', Notification.objects.get().body)

    def test_client_without_email_gets_nothing(self):
        self.client_obj.email = ''
        self.client_obj.save(update_fields=['email'])
        self._change_status('ready_for_shipment')

        self.assertEqual(Notification.objects.count(), 0)

    @override_settings(NOTIFY_CLIENTS=False)
    def test_disabled_client_letters_are_not_queued(self):
        """Переписка с внешними людьми не должна включиться сама собой."""
        self._change_status('ready_for_shipment')
        self.assertEqual(Notification.objects.count(), 0)

    def test_status_change_still_works_and_is_logged(self):
        """Оповещение не должно мешать основному делу."""
        self._change_status('ready_for_shipment')
        self.order.refresh_from_db()

        self.assertEqual(self.order.status, 'ready_for_shipment')
        self.assertTrue(self.order.status_history.exists())


@override_settings(NOTIFY_CLIENTS=True)
class UnrepairableStatusTests(TestCase):
    """Статус «Ремонт невозможен»: переход, дашборд, должники, письмо."""

    def setUp(self):
        self.user = Employee.objects.create_superuser(
            username='admin_unrepairable', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.user)

        self.client_obj = ClientModel.objects.create(
            name='ООО Гамма', email='gamma@example.com'
        )
        self.order = RepairOrder.objects.create(client=self.client_obj)
        model = EquipmentModel.objects.create(name='Привод дверей')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-UNREP'),
        )

    def _change_status(self, status):
        return self.client_http.post(
            f'/repair-orders/{self.order.pk}/change-status/',
            {'new_status': status, 'notes': ''},
        )

    def test_order_can_be_marked_unrepairable(self):
        self._change_status('unrepairable')
        self.order.refresh_from_db()

        self.assertEqual(self.order.status, 'unrepairable')
        self.assertTrue(
            self.order.status_history.filter(status='unrepairable').exists()
        )

    def test_unrepairable_orders_are_not_active(self):
        response = self.client_http.get('/')
        before = response.context['active_orders']

        self._change_status('unrepairable')

        response = self.client_http.get('/')
        self.assertEqual(response.context['active_orders'], before - 1)

    def test_unrepairable_orders_are_not_debtors(self):
        """Ремонт не сделан, счёт не выставлен — это не долг заказчика."""
        self._change_status('unrepairable')
        self.order.refresh_from_db()

        self.assertEqual(self.order.payment_status, 'unpaid')
        self.assertNotIn(self.order, RepairOrder.objects.with_debt())

        response = self.client_http.get('/')
        self.assertNotIn(self.order, response.context['debtors'])

    def test_client_is_notified(self):
        self._change_status('unrepairable')

        note = Notification.objects.get(event='order_status')
        self.assertEqual(note.recipient, 'gamma@example.com')
        self.assertIn(self.order.order_number, note.subject)
        self.assertIn(self.order.order_number, note.body)


class LowStockNotificationTests(TestCase):
    """Оповещения сотрудникам о дефиците."""

    def setUp(self):
        self.warehouse = Employee.objects.create_user(
            username='wh', full_name='Кладовщик', password='pass',
            role='warehouse', email='wh@example.com',
        )
        Employee.objects.create_user(
            username='acc', full_name='Бухгалтер', password='pass',
            role='accountant', email='acc@example.com',
        )
        Employee.objects.create_user(
            username='wh_noemail', full_name='Без почты', password='pass',
            role='warehouse',
        )
        self.part = SparePart.objects.create(
            part_number='LOW-1', name='Резистор', current_stock=10, min_stock=5
        )

    def _spend(self, quantity):
        self.part.current_stock -= quantity
        self.part.save(update_fields=['current_stock'])

    def test_shortage_notifies_warehouse_only(self):
        self._spend(6)

        recipients = set(Notification.objects.values_list('recipient', flat=True))
        self.assertEqual(recipients, {'wh@example.com'})

    def test_letter_contains_what_is_needed_to_order(self):
        self.part.preferred_supplier = 'Чип и Дип'
        self.part.lead_time_days = 14
        self.part.save()
        self._spend(6)

        body = Notification.objects.first().body
        self.assertIn('LOW-1', body)
        self.assertIn('Чип и Дип', body)
        self.assertIn('14', body)

    def test_stock_above_minimum_notifies_nobody(self):
        self._spend(1)
        self.assertEqual(Notification.objects.count(), 0)

    def test_repeated_write_offs_do_not_spam(self):
        """При разборе заказа списывают по несколько деталей подряд."""
        self._spend(6)
        self._spend(1)
        self._spend(1)

        self.assertEqual(Notification.objects.count(), 1)

    @override_settings(NOTIFY_LOW_STOCK_COOLDOWN_HOURS=0)
    def test_cooldown_can_be_switched_off(self):
        self._spend(6)
        self._spend(1)

        self.assertEqual(Notification.objects.count(), 2)

    @override_settings(NOTIFY_LOW_STOCK=False)
    def test_can_be_disabled(self):
        self._spend(6)
        self.assertEqual(Notification.objects.count(), 0)


class SendNotificationsCommandTests(TestCase):
    """Команда отправки: выключатель, повторы, просроченные."""

    def setUp(self):
        self.note = Notification.objects.create(
            event='low_stock', recipient='wh@example.com',
            subject='Дефицит', body='Текст',
        )

    def _run(self, **options):
        from io import StringIO
        out = StringIO()
        call_command('send_notifications', stdout=out, stderr=StringIO(), **options)
        return out.getvalue()

    @override_settings(NOTIFICATIONS_ENABLED=False)
    def test_nothing_is_sent_while_disabled(self):
        from django.core import mail

        output = self._run()
        self.note.refresh_from_db()

        self.assertEqual(len(mail.outbox), 0)
        self.assertEqual(self.note.status, 'pending')
        self.assertIn('выключена', output)

    @override_settings(NOTIFICATIONS_ENABLED=True,
                       EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_sends_and_marks_as_sent(self):
        from django.core import mail

        self._run()
        self.note.refresh_from_db()

        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ['wh@example.com'])
        self.assertEqual(self.note.status, 'sent')
        self.assertIsNotNone(self.note.sent_at)

    @override_settings(NOTIFICATIONS_ENABLED=True,
                       EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_dry_run_sends_nothing(self):
        from django.core import mail

        output = self._run(dry_run=True)
        self.note.refresh_from_db()

        self.assertEqual(len(mail.outbox), 0)
        self.assertEqual(self.note.status, 'pending')
        self.assertIn('проверка', output)

    @override_settings(NOTIFICATIONS_ENABLED=True, NOTIFICATIONS_MAX_ATTEMPTS=2)
    def test_failure_is_retried_then_given_up(self):
        with patch('core.management.commands.send_notifications.EmailMessage.send',
                   side_effect=OSError('SMTP молчит')):
            self._run()
            self.note.refresh_from_db()
            self.assertEqual(self.note.status, 'pending')
            self.assertEqual(self.note.attempts, 1)
            self.assertIn('SMTP молчит', self.note.last_error)

            self._run()
            self.note.refresh_from_db()

        self.assertEqual(self.note.status, 'failed')
        self.assertEqual(self.note.attempts, 2)

    @override_settings(NOTIFICATIONS_ENABLED=True, NOTIFICATIONS_MAX_AGE_HOURS=24,
                       EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_stale_notifications_are_skipped(self):
        """Иначе после включения отправки уедет месячная пачка новостей."""
        from django.core import mail

        Notification.objects.filter(pk=self.note.pk).update(
            created_at=timezone.now() - datetime.timedelta(days=3)
        )
        self._run()
        self.note.refresh_from_db()

        self.assertEqual(len(mail.outbox), 0)
        self.assertEqual(self.note.status, 'skipped')
        self.assertIn('Старше', self.note.last_error)


class PartPriceTests(TestCase):
    """Закупочные цены: план закупок, стоимость запаса, себестоимость ремонта."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_price', full_name='Админ', password='pass')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.part = SparePart.objects.create(
            part_number='PR-1', name='Конденсатор', current_stock=2, min_stock=10,
            price=Decimal('150.00'))
        self.no_price = SparePart.objects.create(
            part_number='PR-2', name='Без цены', current_stock=0, min_stock=5)

    def test_purchase_cost_covers_only_the_shortage(self):
        """Закупать надо недостающее, а не весь минимум."""
        self.assertEqual(self.part.stock_deficit, 8)
        self.assertEqual(self.part.purchase_cost, Decimal('1200.00'))

    def test_stock_value_uses_what_is_on_the_shelf(self):
        self.assertEqual(self.part.stock_value, Decimal('300.00'))

    def test_a_part_without_a_price_has_no_sums(self):
        """Пустая цена — не ноль: «неизвестно» и «бесплатно» разные вещи."""
        self.assertIsNone(self.no_price.price)
        self.assertIsNone(self.no_price.purchase_cost)
        self.assertIsNone(self.no_price.stock_value)

    def test_the_plan_totals_what_it_can_and_names_the_rest(self):
        resp = self.client_http.get('/reports/purchase-plan/')

        self.assertEqual(resp.context['plan_total'], Decimal('1200.00'))
        self.assertEqual(resp.context['without_price'], 1)
        self.assertContains(resp, 'без учёта 1 поз. без цены')

    def test_the_plan_export_carries_price_and_total(self):
        resp = self.client_http.get('/reports/purchase-plan/export/')
        ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        headers = rows[0]

        self.assertIn('Цена, ₽', headers)
        self.assertIn('Сумма, ₽', headers)
        self.assertEqual(rows[-1][0], 'Итого')
        self.assertEqual(float(rows[-1][headers.index('Сумма, ₽')]), 1200.0)

    def test_incoming_with_a_price_updates_the_part(self):
        """Вводить цену дважды — в приходе и в карточке — никто не будет."""
        self.client_http.post(f'/parts/{self.part.pk}/stock-incoming/', {
            'quantity': 5, 'unit_price': '175.50', 'document_number': 'ТН-7', 'notes': '',
        })
        self.part.refresh_from_db()

        self.assertEqual(self.part.price, Decimal('175.50'))
        self.assertEqual(self.part.current_stock, 7)

    def test_incoming_without_a_price_keeps_the_old_one(self):
        self.client_http.post(f'/parts/{self.part.pk}/stock-incoming/', {
            'quantity': 5, 'unit_price': '', 'document_number': '', 'notes': '',
        })
        self.part.refresh_from_db()

        self.assertEqual(self.part.price, Decimal('150.00'))

    def test_the_delivery_price_stays_in_the_history(self):
        """На вопрос «почему деталь подорожала вдвое» отвечает история приходов."""
        self.client_http.post(f'/parts/{self.part.pk}/stock-incoming/', {
            'quantity': 4, 'unit_price': '200.00', 'document_number': '', 'notes': '',
        })
        movement = StockMovement.objects.get(part=self.part, movement_type='incoming')

        self.assertEqual(movement.unit_price, Decimal('200.00'))
        self.assertEqual(movement.total_price, Decimal('800.00'))

    def test_parts_used_in_an_order_show_their_cost(self):
        order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Цена'))
        detail = RepairOrderDetail.objects.create(
            repair_order=order, part=self.part, quantity_used=3)
        RepairOrderDetail.objects.create(
            repair_order=order, part=self.no_price, quantity_used=1)

        self.assertEqual(detail.cost, Decimal('450.00'))

        resp = self.client_http.get(f'/repair-orders/{order.pk}/')
        self.assertEqual(resp.context['details_cost'], Decimal('450.00'))

    def test_the_cost_price_never_reaches_the_client_act(self):
        """Себестоимость — внутренняя цифра; заказчику идёт стоимость работ."""
        order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Акт-цена'))
        RepairOrderDetail.objects.create(
            repair_order=order, part=self.part, quantity_used=3)

        resp = self.client_http.get(f'/repair-orders/{order.pk}/act/complete/')

        self.assertContains(resp, 'Конденсатор')      # деталь названа
        self.assertNotContains(resp, '450')            # а цена — нет
        self.assertNotContains(resp, 'Себестоимость')

    def test_price_survives_import_and_export(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['part_number', 'name', 'component_type', 'price'])
        ws.append(['PR-IMP', 'Импортированная', 'Резистор', '99.90'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'import.xlsx'

        self.client_http.post('/parts/import/', {'file': buf, 'update_existing': False})
        self.assertEqual(SparePart.objects.get(part_number='PR-IMP').price,
                         Decimal('99.90'))

        resp = self.client_http.get('/parts/export/?q=PR-IMP')
        exported = openpyxl.load_workbook(io.BytesIO(resp.content)).active
        row = dict(zip([c.value for c in exported[1]], [c.value for c in exported[2]]))
        self.assertEqual(float(row['price']), 99.90)


class ActTests(TestCase):
    """Печатные акты: приёма оборудования и выполненных работ."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_act', full_name='Админ', password='pass')
        self.manager = Employee.objects.create_user(
            username='mgr_act', full_name='Менеджер', password='pass',
            role='repair_manager')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.organization = Organization.get_solo()
        self.organization.name = 'ООО «Лифт Тим»'
        self.organization.inn = '7701234567'
        self.organization.signatory_position = 'Директор'
        self.organization.signatory_name = 'Петров П. П.'
        self.organization.save()

        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Городские лифты»', inn='5001'),
        )
        model = EquipmentModel.objects.create(name='БУАД-акт')
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-ACT'),
            fault_description='Не запускается привод',
            seal_numbers='П-0012',
            initial_condition='Разъём окислен',
            repair_cost=7200,
        )

    def _receive(self):
        return self.client_http.get(f'/repair-orders/{self.order.pk}/act/receive/')

    def _complete(self):
        return self.client_http.get(f'/repair-orders/{self.order.pk}/act/complete/')

    def test_receive_act_carries_seals_and_condition(self):
        """Ради этих двух граф акт и подписывают: о них потом спорят."""
        resp = self._receive()

        self.assertContains(resp, 'П-0012')
        self.assertContains(resp, 'Разъём окислен')
        self.assertContains(resp, 'SN-ACT')

    def test_both_acts_carry_the_company_details(self):
        for resp in (self._receive(), self._complete()):
            self.assertContains(resp, 'ООО «Лифт Тим»')
            self.assertContains(resp, '7701234567')
            self.assertContains(resp, 'Петров П. П.')

    def test_acts_name_both_parties(self):
        resp = self._receive()

        self.assertContains(resp, 'МУП «Городские лифты»')
        self.assertContains(resp, 'Исполнитель')
        self.assertContains(resp, 'Заказчик')

    def test_completion_act_prints_the_work_not_the_fault(self):
        """Подписывать «выполнено» под описанием поломки нельзя."""
        self.roe.work_performed = 'Заменён силовой ключ, настроен привод'
        self.roe.save(update_fields=['work_performed'])

        resp = self._complete()

        self.assertContains(resp, 'Заменён силовой ключ')

    def test_without_recorded_work_the_fault_is_marked_as_fixed(self):
        resp = self._complete()

        self.assertContains(resp, 'Устранена неисправность')

    def test_completion_act_shows_the_total(self):
        resp = self._complete()

        self.assertContains(resp, '7200')

    def test_completion_act_shows_the_warranty_date(self):
        self.order.date_completed = timezone.now()
        self.order.save(update_fields=['date_completed'])

        resp = self._complete()

        self.assertContains(resp, 'Гарантия на выполненные работы')

    def test_an_unfinished_order_warns_instead_of_inventing_a_date(self):
        self.assertIsNone(self.order.date_completed)

        resp = self._complete()

        self.assertNotContains(resp, 'Гарантия на выполненные работы')
        self.assertContains(resp, 'Заказ ещё не завершён')

    def test_completion_act_lists_replaced_parts(self):
        part = SparePart.objects.create(part_number='ACT-1', name='Конденсатор',
                                        current_stock=10)
        RepairOrderDetail.objects.create(repair_order=self.order, part=part,
                                         quantity_used=2)
        resp = self._complete()

        self.assertContains(resp, 'Конденсатор')
        self.assertContains(resp, 'ACT-1')

    def test_acts_print_even_without_company_details(self):
        """Пустая шапка — не повод не напечатать: допишут от руки."""
        Organization.objects.all().delete()

        resp = self._receive()

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Реквизиты не заполнены')

    def test_acts_require_login(self):
        resp = TestClient().get(f'/repair-orders/{self.order.pk}/act/receive/')

        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])


class DefectActTests(TestCase):
    """Акт дефектации: что нашли при диагностике и во сколько это обойдётся."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_defect', full_name='Админ', password='pass')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        organization = Organization.get_solo()
        organization.name = 'ООО «Лифт Тим»'
        organization.signatory_position = 'Директор'
        organization.signatory_name = 'Петров П. П.'
        organization.save()

        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО «Ремонт»'),
        )
        self.model = EquipmentModel.objects.create(
            name='Emotron DSV35-40-028 Lift', kind='Преобразователь частоты')
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=self.model, serial_number='SN-DEF'),
            fault_description='Не включается',
        )

    def _act_url(self):
        return f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/act/defect/'

    def _edit_url(self):
        return f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/defect/'

    def _fill(self, **overrides):
        data = {
            'defect_act_date': '2026-05-06',
            'diagnosis': 'Выявлен выход из строя IGBT модуля и его обвязки.',
            'error_codes': '«Десат» — короткое замыкание между фазами\n'
                           '«EOL» — переезд зоны полного открытия',
            'warranty_case': 'non_warranty',
            'non_warranty_reason': 'перепадами напряжения в питающей сети',
            'estimated_cost': '14000',
        }
        data.update(overrides)
        return self.client_http.post(self._edit_url(), data)

    def test_the_generic_type_stands_before_the_model(self):
        """«Тип: Преобразователь частоты Emotron …» — так набран их бланк."""
        self.assertEqual(self.model.full_name,
                         'Преобразователь частоты Emotron DSV35-40-028 Lift')

    def test_a_model_without_a_generic_type_prints_as_is(self):
        model = EquipmentModel.objects.create(name='БУАД-3')

        self.assertEqual(model.full_name, 'БУАД-3')

    def test_saving_stays_on_the_form(self):
        """Сохранение больше не уводит на печатный акт.

        Пока уводило, правка одной строки диагноза стоила ухода
        на документ и возвращения назад. Документ открывается своей
        кнопкой, когда он нужен.
        """
        resp = self._fill()

        self.assertRedirects(
            resp,
            reverse('repair_order_defect_act_edit',
                    args=[self.order.pk, self.roe.pk]),
        )
        self.roe.refresh_from_db()
        self.assertEqual(self.roe.estimated_cost, Decimal('14000'))
        self.assertEqual(self.roe.defect_act_date, datetime.date(2026, 5, 6))

    def test_the_act_is_opened_by_its_own_button(self):
        """И рядом с ней сказано, что акт печатает сохранённое: молчание
        здесь означало бы документ заказчику без того, что мастер
        только что вписал."""
        html = self.client_http.get(
            reverse('repair_order_defect_act_edit',
                    args=[self.order.pk, self.roe.pk])
        ).content.decode()

        self.assertIn('Открыть акт', html)
        self.assertIn(self._act_url(), html)
        self.assertIn('несохранённые правки', html)
        self.assertNotIn('Сохранить и открыть акт', html)

    def test_the_act_carries_diagnosis_codes_and_estimate(self):
        self._fill()

        resp = self.client_http.get(self._act_url())

        self.assertContains(resp, 'IGBT модуля')
        self.assertContains(resp, '«Десат»')
        self.assertContains(resp, '«EOL»')
        self.assertContains(resp, 'Случай не является гарантийным')
        self.assertContains(resp, 'перепадами напряжения')
        # Разряды разделены неразрывным пробелом: иначе печать перенесёт
        # строку посреди суммы
        self.assertContains(resp, '14 000 рублей')
        self.assertContains(resp, 'Преобразователь частоты Emotron DSV35-40-028 Lift')
        self.assertContains(resp, 'SN-DEF')

    def test_the_estimate_is_grouped_by_thousands(self):
        self.roe.estimated_cost = Decimal('54000.00')

        self.assertEqual(self.roe.estimated_cost_text, '54 000')

    def test_no_estimate_means_no_text(self):
        self.assertIsNone(self.roe.estimated_cost_text)

    def test_error_codes_are_split_by_lines(self):
        self.roe.error_codes = ' «F2340» — КЗ в модуле \n\n «EOC» — КЗ на выходе '
        self.roe.save(update_fields=['error_codes'])

        self.assertEqual(self.roe.error_code_lines,
                         ['«F2340» — КЗ в модуле', '«EOC» — КЗ на выходе'])

    def test_a_warranty_case_does_not_demand_money(self):
        """Гарантийному случаю оценка ремонта в акте не место."""
        self._fill(warranty_case='warranty', estimated_cost='')

        resp = self.client_http.get(self._act_url())

        self.assertContains(resp, 'Случай является гарантийным')
        self.assertNotContains(resp, 'Случай не является гарантийным')
        self.assertNotContains(resp, 'Ориентировочная стоимость')

    def test_a_warranty_case_drops_the_prefilled_reason(self):
        """Заготовка причины не должна противоречить самому акту."""
        self._fill(warranty_case='warranty')

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.non_warranty_reason, '')

    def test_an_undecided_case_says_nothing_about_the_warranty(self):
        self._fill(warranty_case='', estimated_cost='')

        resp = self.client_http.get(self._act_url())

        self.assertNotContains(resp, 'гарантийным')

    def test_an_empty_act_warns_instead_of_printing_a_blank(self):
        resp = self.client_http.get(self._act_url())

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Результаты диагностики не заполнены')

    def test_without_a_date_the_act_is_dated_today(self):
        self._fill(defect_act_date='')

        resp = self.client_http.get(self._act_url())

        self.assertContains(resp, timezone.localdate().strftime('%d.%m.%Y'))

    def test_the_prefilled_reason_is_the_usual_wording(self):
        resp = self.client_http.get(self._edit_url())

        self.assertContains(resp, 'естественной деградацией электронных компонентов')

    def test_equipment_from_another_order_is_not_reachable(self):
        """Иначе в акт попал бы серийник из чужого заказа."""
        other = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО «Другой»'))

        resp = self.client_http.get(
            f'/repair-orders/{other.pk}/equipment/{self.roe.pk}/act/defect/')

        self.assertEqual(resp.status_code, 404)

    def test_the_act_requires_login(self):
        resp = TestClient().get(self._act_url())

        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])

    def test_the_order_card_offers_to_fill_the_act(self):
        resp = self.client_http.get(f'/repair-orders/{self.order.pk}/')

        self.assertContains(resp, self._edit_url())

    def test_a_filled_act_is_linked_from_the_order_card(self):
        self._fill()

        resp = self.client_http.get(f'/repair-orders/{self.order.pk}/')

        self.assertContains(resp, self._act_url())

    def test_known_generic_types_are_suggested_in_the_model_form(self):
        resp = self.client_http.get('/equipment/models/create/')

        self.assertContains(resp, 'equipment-kinds')
        self.assertContains(resp, 'Преобразователь частоты')


class OrganizationTests(TestCase):
    """Реквизиты фирмы: одна запись, править может только администратор."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_org', full_name='Админ', password='pass')
        self.accountant = Employee.objects.create_user(
            username='buh_org', full_name='Бухгалтер', password='pass',
            role='accountant')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def test_the_record_is_created_on_first_use(self):
        self.assertEqual(Organization.objects.count(), 0)

        Organization.get_solo()
        Organization.get_solo()

        self.assertEqual(Organization.objects.count(), 1)

    def test_admin_can_save_the_details(self):
        resp = self.client_http.post('/management/organization/', {
            'name': 'ООО «Лифт Тим»', 'inn': '7701234567', 'kpp': '770101001',
            'address': 'Москва', 'phone': '+7 925 282-40-31', 'email': '',
            'signatory_position': 'Директор', 'signatory_name': 'Петров П. П.',
        }, follow=True)

        self.assertEqual(resp.status_code, 200)
        organization = Organization.get_solo()
        self.assertEqual(organization.name, 'ООО «Лифт Тим»')
        self.assertIn('ИНН 7701234567', organization.details_line)

    def test_others_cannot_edit_the_details(self):
        other = TestClient()
        other.force_login(self.accountant)

        resp = other.get('/management/organization/')

        self.assertEqual(resp.status_code, 302)


class PaymentTests(TestCase):
    """Оплаты частями. До них «оплачено» было только статусом, и при
    частичной оплате долгом считалась вся стоимость ремонта."""

    def setUp(self):
        self.accountant = Employee.objects.create_user(
            username='buh_pay', full_name='Бухгалтер', password='pass',
            role='accountant')
        self.manager = Employee.objects.create_user(
            username='mgr_pay', full_name='Менеджер', password='pass',
            role='repair_manager')
        self.client_http = TestClient()
        self.client_http.force_login(self.accountant)

        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Плательщик'),
            payment_status='unpaid',
        )
        model = EquipmentModel.objects.create(name='БУАД-оплата')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-PAY'),
            repair_cost=10000,
        )

    def _pay(self, amount, client=None):
        return (client or self.client_http).post(
            f'/repair-orders/{self.order.pk}/payments/add/',
            {'amount': amount, 'payment_date': datetime.date.today().isoformat(),
             'note': 'п/п 1'},
        )

    def test_partial_payment_leaves_the_remainder_as_debt(self):
        self._pay('4000')
        self.order.refresh_from_db()

        self.assertEqual(self.order.paid_amount, Decimal('4000'))
        self.assertEqual(self.order.debt, Decimal('6000'))
        self.assertEqual(self.order.payment_status, 'partially_paid')

    def test_full_payment_closes_the_order(self):
        """Иначе заказ висел бы в должниках после последнего платежа."""
        self._pay('4000')
        self._pay('6000')
        self.order.refresh_from_db()

        self.assertEqual(self.order.debt, 0)
        self.assertEqual(self.order.payment_status, 'paid')
        self.assertNotIn(self.order, RepairOrder.objects.with_debt())

    def test_overpayment_still_counts_as_paid(self):
        self._pay('12000')
        self.order.refresh_from_db()

        self.assertEqual(self.order.debt, 0)
        self.assertEqual(self.order.payment_status, 'paid')

    def test_deleting_a_payment_brings_the_debt_back(self):
        """Суммы вбивают руками, опечатка в разряде — обычное дело."""
        self._pay('10000')
        payment = Payment.objects.get()

        self.client_http.post(
            f'/repair-orders/{self.order.pk}/payments/{payment.pk}/delete/')
        self.order.refresh_from_db()

        self.assertEqual(self.order.paid_amount, 0)
        self.assertEqual(self.order.debt, Decimal('10000'))
        self.assertEqual(self.order.payment_status, 'unpaid')

    def test_a_manually_paid_order_has_no_debt(self):
        """Статус ставит человек, видевший платёжку, — он главнее арифметики."""
        self.order.payment_status = 'paid'
        self.order.save(update_fields=['payment_status'])

        self.assertEqual(self.order.paid_amount, 0)
        self.assertEqual(self.order.debt, 0)

    def test_payment_is_recorded_in_the_history(self):
        self._pay('4000')

        note = self.order.status_history.first().notes
        self.assertIn('4000', note)
        self.assertIn('п/п 1', note)

    def test_a_negative_amount_is_rejected(self):
        self._pay('-500')

        self.assertEqual(Payment.objects.count(), 0)

    def test_only_accounting_may_enter_payments(self):
        manager_client = TestClient()
        manager_client.force_login(self.manager)

        resp = self._pay('1000', client=manager_client)

        self.assertEqual(resp.status_code, 302)
        self.assertEqual(Payment.objects.count(), 0)


class DebtQuerySetTests(TestCase):
    """Что считается просроченным долгом."""

    def setUp(self):
        self.client_obj = ClientModel.objects.create(name='ООО Должник')
        self.today = datetime.date.today()

    def _order(self, **kwargs):
        order = RepairOrder.objects.create(client=self.client_obj, **kwargs)
        model, _ = EquipmentModel.objects.get_or_create(name='БУАД-долг')
        RepairOrderEquipment.objects.create(
            repair_order=order,
            equipment=Equipment.objects.create(
                model=model, serial_number=f'SN-{order.pk}'),
            repair_cost=1000,
        )
        return order

    def test_paid_orders_are_not_debts(self):
        self._order(payment_status='paid',
                    invoice_date=self.today - datetime.timedelta(days=90))

        self.assertEqual(RepairOrder.objects.with_debt().count(), 0)

    def test_partially_paid_is_a_debt(self):
        self._order(payment_status='partially_paid')

        self.assertEqual(RepairOrder.objects.with_debt().count(), 1)

    def test_unrepairable_is_not_a_debt(self):
        """По неремонтопригодному оборудованию счёт не выставляют."""
        self._order(payment_status='unpaid', status='unrepairable')

        self.assertEqual(RepairOrder.objects.with_debt().count(), 0)

    @override_settings(DEBT_OVERDUE_DAYS=14)
    def test_fresh_invoice_is_not_overdue_yet(self):
        self._order(payment_status='unpaid',
                    invoice_date=self.today - datetime.timedelta(days=3))

        self.assertEqual(RepairOrder.objects.overdue().count(), 0)

    @override_settings(DEBT_OVERDUE_DAYS=14)
    def test_old_invoice_is_overdue(self):
        order = self._order(payment_status='unpaid',
                            invoice_date=self.today - datetime.timedelta(days=20))

        self.assertEqual(list(RepairOrder.objects.overdue()), [order])
        self.assertEqual(order.days_overdue, 20)
        self.assertTrue(order.is_overdue)

    @override_settings(DEBT_OVERDUE_DAYS=14)
    def test_a_zero_sum_order_is_not_chased(self):
        """Письмо «оплатите 0 ₽» позорнее, чем отсутствие письма."""
        order = RepairOrder.objects.create(
            client=self.client_obj, payment_status='unpaid',
            invoice_date=self.today - datetime.timedelta(days=30))
        model, _ = EquipmentModel.objects.get_or_create(name='БУАД-ноль')
        RepairOrderEquipment.objects.create(
            repair_order=order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-ZERO'),
            repair_cost=0,
        )

        self.assertEqual(RepairOrder.objects.overdue().count(), 0)
        # В отчёте он при этом остаётся: это недозаполненная карточка
        self.assertIn(order, RepairOrder.objects.with_debt())

    def test_without_an_invoice_nothing_is_overdue(self):
        """Пока счёт не выставлен, требовать оплату не за что."""
        order = self._order(payment_status='unpaid', invoice_date=None)

        self.assertEqual(RepairOrder.objects.overdue().count(), 0)
        self.assertEqual(order.days_overdue, 0)
        self.assertFalse(order.is_overdue)

    @override_settings(DEBT_OVERDUE_DAYS=14)
    def test_a_fully_paid_order_leaves_the_overdue_list(self):
        order = self._order(payment_status='unpaid',
                            invoice_date=self.today - datetime.timedelta(days=30))
        self.assertEqual(RepairOrder.objects.overdue().count(), 1)

        Payment.objects.create(repair_order=order, amount=1000)

        self.assertEqual(RepairOrder.objects.overdue().count(), 0)

    @override_settings(DEBT_OVERDUE_DAYS=14)
    def test_a_partly_paid_order_stays_but_owes_less(self):
        order = self._order(payment_status='unpaid',
                            invoice_date=self.today - datetime.timedelta(days=30))
        Payment.objects.create(repair_order=order, amount=400)
        order.refresh_from_db()

        self.assertEqual(list(RepairOrder.objects.overdue()), [order])
        self.assertEqual(order.debt, Decimal('600'))

    def test_shipped_without_an_invoice_is_found_separately(self):
        """Это не долг заказчика, а свой недосмотр — но видеть его надо."""
        order = self._order(payment_status='unpaid', invoice_date=None,
                            status='shipped')
        self._order(payment_status='unpaid', invoice_date=None, status='repair')

        self.assertEqual(list(RepairOrder.objects.shipped_without_invoice()), [order])


class DebtReminderTests(TestCase):
    """Напоминания заказчику и сводка бухгалтерии."""

    def setUp(self):
        Employee.objects.create_user(
            username='buh', full_name='Бухгалтер', password='pass',
            role='accountant', email='buh@example.com',
        )
        Employee.objects.create_user(
            username='sklad_d', full_name='Кладовщик', password='pass',
            role='warehouse', email='sklad@example.com',
        )
        self.client_obj = ClientModel.objects.create(
            name='ООО Должник', email='debtor@example.com')
        self.order = RepairOrder.objects.create(
            client=self.client_obj, payment_status='unpaid',
            invoice_number='42',
            invoice_date=datetime.date.today() - datetime.timedelta(days=30),
        )
        model = EquipmentModel.objects.create(name='БУАД-напоминание')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-D1'),
            repair_cost=15000,
        )

    def _run(self, **options):
        out = io.StringIO()
        call_command('debt_reminders', stdout=out, stderr=io.StringIO(), **options)
        return out.getvalue()

    def _queued(self, event):
        return list(Notification.objects.filter(event=event))

    @override_settings(NOTIFY_DEBT_DIGEST=True)
    def test_digest_goes_to_accounting_not_to_the_warehouse(self):
        """Долги — дело бухгалтерии; кладовщику этот список ни к чему."""
        self._run()

        recipients = {n.recipient for n in self._queued('debt_digest')}
        self.assertEqual(recipients, {'buh@example.com'})

    @override_settings(NOTIFY_DEBT_DIGEST=True)
    def test_digest_lists_the_order_and_the_total(self):
        self._run()

        body = self._queued('debt_digest')[0].body
        self.assertIn(self.order.order_number, body)
        self.assertIn('ООО Должник', body)
        self.assertIn(notifications.money(15000), body)
        self.assertIn('просрочено 30 дней', body)

    @override_settings(NOTIFY_DEBT_DIGEST=True)
    def test_digest_counts_the_remainder_not_the_full_price(self):
        """Регрессия: при частичной оплате требовали всю стоимость."""
        Payment.objects.create(repair_order=self.order, amount=5000)

        self._run()
        note = self._queued('debt_digest')[0]

        self.assertIn(notifications.money(10000), note.body)
        self.assertIn(notifications.money(5000), note.body)
        self.assertIn(notifications.money(10000), note.subject)

    @override_settings(NOTIFY_CLIENTS=True, NOTIFY_DEBTS=True)
    def test_the_client_is_asked_for_the_remainder_only(self):
        Payment.objects.create(repair_order=self.order, amount=5000)

        self._run()
        body = self._queued('debt_reminder')[0].body

        self.assertIn(f'Остаток к оплате: {notifications.money(10000)}', body)
        self.assertIn(notifications.money(5000), body)

    @override_settings(NOTIFY_CLIENTS=True, NOTIFY_DEBTS=True,
                       NOTIFY_DEBT_DIGEST=True)
    def test_a_settled_order_is_left_alone(self):
        Payment.objects.create(repair_order=self.order, amount=15000)

        self._run()

        self.assertEqual(Notification.objects.count(), 0)

    @override_settings(NOTIFY_DEBT_DIGEST=True, DEBT_DIGEST_COOLDOWN_DAYS=7)
    def test_digest_is_not_repeated_every_day(self):
        """Команда запускается ежедневно, сводка каждый день не нужна."""
        self._run()
        self._run()

        self.assertEqual(len(self._queued('debt_digest')), 1)

    @override_settings(NOTIFY_DEBT_DIGEST=True)
    def test_shipped_without_an_invoice_gets_into_the_digest(self):
        without = RepairOrder.objects.create(
            client=self.client_obj, payment_status='unpaid', status='shipped')
        self._run()

        body = self._queued('debt_digest')[0].body
        self.assertIn('счёт не выставлен', body)
        self.assertIn(without.order_number, body)

    @override_settings(NOTIFY_CLIENTS=True, NOTIFY_DEBTS=True)
    def test_client_reminder_carries_the_invoice_and_the_sum(self):
        self._run()

        reminders = self._queued('debt_reminder')
        self.assertEqual([n.recipient for n in reminders], ['debtor@example.com'])
        self.assertIn('42', reminders[0].body)
        self.assertIn(notifications.money(15000), reminders[0].body)

    @override_settings(NOTIFY_CLIENTS=True, NOTIFY_DEBTS=False)
    def test_client_reminders_are_off_by_default(self):
        """Требование денег от лица фирмы не должно начаться само собой."""
        self._run()

        self.assertEqual(self._queued('debt_reminder'), [])

    @override_settings(NOTIFY_CLIENTS=False, NOTIFY_DEBTS=True)
    def test_the_general_client_switch_also_holds_them(self):
        self._run()

        self.assertEqual(self._queued('debt_reminder'), [])

    @override_settings(NOTIFY_CLIENTS=True, NOTIFY_DEBTS=True,
                       DEBT_REMINDER_COOLDOWN_DAYS=7)
    def test_the_same_client_is_not_pestered_daily(self):
        self._run()
        self._run()

        self.assertEqual(len(self._queued('debt_reminder')), 1)

    @override_settings(NOTIFY_CLIENTS=True, NOTIFY_DEBTS=True)
    def test_a_client_without_an_email_is_skipped_quietly(self):
        self.client_obj.email = ''
        self.client_obj.save(update_fields=['email'])

        self._run()

        self.assertEqual(self._queued('debt_reminder'), [])

    @override_settings(NOTIFY_CLIENTS=True, NOTIFY_DEBTS=True,
                       NOTIFY_DEBT_DIGEST=True)
    def test_dry_run_queues_nothing(self):
        output = self._run(dry_run=True)

        self.assertEqual(Notification.objects.count(), 0)
        self.assertIn('проверка', output)
        self.assertIn(self.order.order_number, output)

    @override_settings(DEBT_OVERDUE_DAYS=14, NOTIFY_CLIENTS=True, NOTIFY_DEBTS=True,
                       NOTIFY_DEBT_DIGEST=True)
    def test_a_fresh_invoice_is_left_alone(self):
        RepairOrder.objects.filter(pk=self.order.pk).update(
            invoice_date=datetime.date.today())

        self._run()

        self.assertEqual(Notification.objects.count(), 0)


class OrderOverdueTests(TestCase):
    """Оповещения о заказах, зависших в одном статусе дольше порога (SLA)."""

    ORDER_OVERDUE_DAYS = {
        'accepted': 2, 'diagnostic': 3, 'repair': 7, 'ready_for_shipment': 5,
    }

    def setUp(self):
        self.manager = Employee.objects.create_user(
            username='rem', full_name='Менеджер по ремонту', password='pass',
            role='repair_manager', email='rem@example.com',
        )
        Employee.objects.create_user(
            username='sklad_o', full_name='Кладовщик', password='pass',
            role='warehouse', email='sklad_o@example.com',
        )
        self.client_obj = ClientModel.objects.create(name='ООО Клиент')

    def _order(self, status, days_ago):
        """Заказ, застрявший в `status` уже `days_ago` дней назад.

        Создание заказа уже заводит запись истории само (сигнал
        `create_status_history`) — сдвигаем в прошлое именно её,
        а не добавляем свою: `order_last_touched` берёт самую свежую
        запись, и вторая, недвинутая, забила бы первую. `changed_at`
        задним числом ставится через queryset.update(), потому что
        auto_now_add применяется только при создании.
        """
        order = RepairOrder.objects.create(client=self.client_obj, status=status)
        when = timezone.now() - datetime.timedelta(days=days_ago)
        OrderStatusHistory.objects.filter(order=order).update(changed_at=when)
        return order

    def _queued(self):
        return list(Notification.objects.filter(event='order_overdue'))

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS)
    def test_an_order_younger_than_the_threshold_is_left_alone(self):
        order = self._order('accepted', days_ago=1)

        self.assertIsNone(notifications.notify_order_overdue(order))
        self.assertEqual(self._queued(), [])

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS)
    def test_an_overdue_order_is_notified_once_not_on_every_check(self):
        order = self._order('diagnostic', days_ago=4)

        notifications.notify_order_overdue(order)
        notifications.notify_order_overdue(order)

        self.assertEqual(len(self._queued()), 1)

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS, ORDER_OVERDUE_ESCALATION_DAYS=7)
    def test_escalation_repeats_after_the_extra_interval_not_before(self):
        # Заказ давно застрял (дольше, чем окно эскалации), чтобы порог по
        # статусу заведомо не мешал проверке самой эскалации
        order = self._order('diagnostic', days_ago=20)
        notifications.notify_order_overdue(order)
        first = self._queued()[0]

        # До эскалации остался день — повторно писать ещё рано
        Notification.objects.filter(pk=first.pk).update(
            created_at=timezone.now() - datetime.timedelta(days=6))
        self.assertIsNone(notifications.notify_order_overdue(order))
        self.assertEqual(len(self._queued()), 1)

        # А неделя с прошлого письма уже прошла — пора написать снова
        Notification.objects.filter(pk=first.pk).update(
            created_at=timezone.now() - datetime.timedelta(days=8))
        notifications.notify_order_overdue(order)
        self.assertEqual(len(self._queued()), 2)

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS)
    def test_different_statuses_apply_different_thresholds(self):
        """4 дня в диагностике (порог 3) — уже просрочка; 4 дня в ремонте
        (порог 7) — ещё нет."""
        diagnostic_order = self._order('diagnostic', days_ago=4)
        repair_order = self._order('repair', days_ago=4)

        self.assertTrue(notifications.notify_order_overdue(diagnostic_order))
        self.assertIsNone(notifications.notify_order_overdue(repair_order))

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS)
    def test_terminal_statuses_have_no_threshold(self):
        order = self._order('shipped', days_ago=60)

        self.assertIsNone(notifications.notify_order_overdue(order))
        self.assertEqual(self._queued(), [])

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS)
    def test_only_repair_manager_and_admin_are_notified_not_the_warehouse(self):
        order = self._order('diagnostic', days_ago=4)

        notifications.notify_order_overdue(order)

        recipients = {n.recipient for n in self._queued()}
        self.assertEqual(recipients, {'rem@example.com'})

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS, NOTIFY_ORDER_OVERDUE=False)
    def test_the_whole_check_can_be_switched_off(self):
        order = self._order('diagnostic', days_ago=10)

        self.assertIsNone(notifications.notify_order_overdue(order))

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS)
    def test_the_command_queues_the_overdue_order(self):
        self._order('diagnostic', days_ago=4)

        call_command('order_overdue_check', stdout=io.StringIO(), stderr=io.StringIO())

        self.assertEqual(len(self._queued()), 1)

    @override_settings(ORDER_OVERDUE_DAYS=ORDER_OVERDUE_DAYS)
    def test_dry_run_queues_nothing(self):
        order = self._order('diagnostic', days_ago=4)

        out = io.StringIO()
        call_command('order_overdue_check', '--dry-run', stdout=out, stderr=io.StringIO())

        self.assertEqual(Notification.objects.count(), 0)
        self.assertIn(order.order_number, out.getvalue())

    def test_open_orders_exclude_shipped_and_unrepairable(self):
        open_order = self._order('repair', days_ago=1)
        self._order('shipped', days_ago=1)
        self._order('unrepairable', days_ago=1)

        self.assertEqual(list(RepairOrder.objects.open()), [open_order])


class RepairAnalyticsTests(TestCase):
    """Отчёт «Аналитика ремонта»: средние сроки, загрузка, разбивка по
    типу неисправности, права доступа.

    «Инженер» в отчёте — суррогат вместо отсутствующего поля
    «ответственный»: сотрудник, который последним перевёл заказ в статус
    «Ремонт», а если такого перехода не было — в «Диагностика» (см.
    `core.views._order_assignees`).
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_ra', full_name='Админ', password='pass')
        self.manager1 = Employee.objects.create_user(
            username='rm1_ra', full_name='Иванов', password='pass', role='repair_manager')
        self.manager2 = Employee.objects.create_user(
            username='rm2_ra', full_name='Петров', password='pass', role='repair_manager')
        Employee.objects.create_user(
            username='sklad_ra', full_name='Кладовщик', password='pass', role='warehouse')
        self.client_http = TestClient()

        self.model = EquipmentModel.objects.create(name='БУАД-аналитика')
        self.client_obj = ClientModel.objects.create(name='ООО Аналитика')
        self.fault_a = FaultType.objects.create(equipment_model=self.model, name='Не включается')
        self.fault_b = FaultType.objects.create(equipment_model=self.model, name='Перегрев')

        self._serial = 0

    def _equipment(self):
        self._serial += 1
        return Equipment.objects.create(model=self.model, serial_number=f'SN-RA-{self._serial}')

    def _make_order(self, received_days_ago, transitions, faults=()):
        """Заказ с управляемой историей: `transitions` — список
        (статус, сколько дней назад, кто сменил), от старой записи
        к новой. Дата приёма и `changed_at` каждой записи истории
        (включая автосозданную при создании) ставятся задним числом —
        auto_now_add применяет их только при создании, а дальше это
        делается через queryset.update()."""
        order = RepairOrder.objects.create(client=self.client_obj, status='accepted')
        when_received = timezone.now() - datetime.timedelta(days=received_days_ago)
        RepairOrder.objects.filter(pk=order.pk).update(date_received=when_received)
        OrderStatusHistory.objects.filter(order=order).update(changed_at=when_received)

        final_status = 'accepted'
        for status, days_ago, changed_by in transitions:
            entry = OrderStatusHistory.objects.create(
                order=order, status=status, changed_by=changed_by)
            when = timezone.now() - datetime.timedelta(days=days_ago)
            OrderStatusHistory.objects.filter(pk=entry.pk).update(changed_at=when)
            final_status = status

        RepairOrder.objects.filter(pk=order.pk).update(status=final_status)

        roe = RepairOrderEquipment.objects.create(repair_order=order, equipment=self._equipment())
        if faults:
            roe.faults.set(faults)

        order.refresh_from_db()
        return order

    def _build_dataset(self):
        # Заказ 1: диагностика → ремонт → отгружен. Инженер — тот, кто вёл
        # ремонт (manager1), а не тот, кто отгрузил (manager2).
        # Приём 10 дн. назад, отгрузка 1 дн. назад — длительность 9 дней.
        self.order1 = self._make_order(
            received_days_ago=10,
            transitions=[
                ('diagnostic', 9, self.manager1),
                ('repair', 8, self.manager1),
                ('shipped', 1, self.manager2),
            ],
            faults=[self.fault_a],
        )
        # Заказ 2: до ремонта не дошло — диагностика → ремонт невозможен.
        # Инженер — тот, кто вёл диагностику. Приём 6 дн., завершение 2 дн.
        # назад — длительность 4 дня. Неисправность не из справочника.
        self.order2 = self._make_order(
            received_days_ago=6,
            transitions=[
                ('diagnostic', 5, self.manager2),
                ('unrepairable', 2, self.manager2),
            ],
        )
        # Заказ 3: отгружен без диагностики и ремонта вовсе (редкий, но
        # возможный случай) — своего инженера в статистике не имеет.
        # Приём 20 дн., отгрузка 15 дн. назад — длительность 5 дней.
        self.order3 = self._make_order(
            received_days_ago=20,
            transitions=[('shipped', 15, self.manager1)],
        )

    def _get(self, employee, url='/reports/repair-analytics/', **params):
        self.client_http.force_login(employee)
        return self.client_http.get(url, params)

    def test_average_repair_duration_is_computed_correctly(self):
        self._build_dataset()

        resp = self._get(self.admin, date_from='2000-01-01')

        self.assertEqual(resp.context['total_orders'], 3)
        # (9 + 4 + 5) / 3 = 6.0
        self.assertEqual(resp.context['avg_days'], 6.0)

    def test_duration_is_grouped_by_surrogate_engineer(self):
        self._build_dataset()

        resp = self._get(self.admin, date_from='2000-01-01')

        by_employee = {row['employee'].pk: row for row in resp.context['by_employee']}
        self.assertEqual(by_employee[self.manager1.pk]['orders'], 1)
        self.assertEqual(by_employee[self.manager1.pk]['avg_days'], 9.0)
        self.assertEqual(by_employee[self.manager2.pk]['orders'], 1)
        self.assertEqual(by_employee[self.manager2.pk]['avg_days'], 4.0)

    def test_an_order_without_a_repair_or_diagnostic_transition_has_no_engineer(self):
        """Заказ 3 не входит ни в чью персональную разбивку, но входит
        в общее среднее (проверено выше — там total_orders включает все три)."""
        self._build_dataset()

        resp = self._get(self.admin, date_from='2000-01-01')

        assigned_orders = sum(row['orders'] for row in resp.context['by_employee'])
        self.assertEqual(assigned_orders, 2)  # заказ 3 не учтён ни у кого

    def test_fault_type_breakdown_only_counts_orders_with_a_chosen_fault_type(self):
        self._build_dataset()

        resp = self._get(self.admin, date_from='2000-01-01')

        by_fault = {row['fault_type']: row for row in resp.context['by_fault_type']}
        self.assertEqual(list(by_fault.keys()), [str(self.fault_a)])
        self.assertEqual(by_fault[str(self.fault_a)]['orders'], 1)
        self.assertEqual(by_fault[str(self.fault_a)]['avg_days'], 9.0)

    def test_period_filter_excludes_orders_completed_outside_it(self):
        self._build_dataset()

        # Только заказ 1 (отгружен 1 день назад) попадает в узкое окно —
        # заказ 2 (завершён 2 дня назад) уже за его границей
        today = timezone.localdate()
        resp = self._get(
            self.admin,
            date_from=(today - datetime.timedelta(days=1)).isoformat(),
            date_to=today.isoformat(),
        )

        self.assertEqual(resp.context['total_orders'], 1)
        self.assertEqual(resp.context['avg_days'], 9.0)

    def test_current_load_counts_open_orders_by_the_last_person_who_touched_them(self):
        self._make_order(
            received_days_ago=4,
            transitions=[('diagnostic', 3, self.manager1), ('repair', 2, self.manager1)],
        )
        self._make_order(
            received_days_ago=2,
            transitions=[('diagnostic', 1, self.manager2)],
        )
        # Только что создан, статус ремонта не менялся вовсе — не в чьей
        # загрузке не отражается (у автосозданной записи истории нет автора)
        RepairOrder.objects.create(client=self.client_obj, status='accepted')

        resp = self._get(self.admin)

        load = {row['employee'].pk: row['count'] for row in resp.context['load_rows']}
        self.assertEqual(load[self.manager1.pk], 1)
        self.assertEqual(load[self.manager2.pk], 1)

    def test_current_load_ignores_completed_orders(self):
        self._build_dataset()  # все три завершены (отгружены/неремонтопригодны)

        resp = self._get(self.admin)

        load = {row['employee'].pk: row['count'] for row in resp.context['load_rows']}
        self.assertEqual(load.get(self.manager1.pk, 0), 0)
        self.assertEqual(load.get(self.manager2.pk, 0), 0)

    def test_admin_sees_the_whole_report(self):
        self._build_dataset()

        resp = self._get(self.admin, date_from='2000-01-01')

        self.assertTrue(resp.context['is_admin'])
        self.assertEqual(resp.context['total_orders'], 3)
        self.assertEqual(len(resp.context['by_employee']), 2)

    def test_a_non_admin_role_sees_only_their_own_figures(self):
        """Менеджер по ремонту не должен видеть ни отчёт целиком, ни
        показатели коллеги — только свой заказ 1 (manager2 к нему не
        имеет отношения, хотя и отгружал его)."""
        self._build_dataset()

        resp = self._get(self.manager1, date_from='2000-01-01')

        self.assertFalse(resp.context['is_admin'])
        self.assertEqual(resp.context['total_orders'], 1)
        self.assertEqual(resp.context['avg_days'], 9.0)

    def test_a_non_admin_role_does_not_see_a_colleagues_load(self):
        self._make_order(
            received_days_ago=4,
            transitions=[('diagnostic', 3, self.manager1), ('repair', 2, self.manager1)],
        )
        self._make_order(
            received_days_ago=2,
            transitions=[('diagnostic', 1, self.manager2)],
        )

        resp = self._get(self.manager1)

        self.assertEqual(len(resp.context['load_rows']), 1)
        self.assertEqual(resp.context['load_rows'][0]['employee'], self.manager1)
        self.assertEqual(resp.context['load_rows'][0]['count'], 1)

    def test_a_role_with_no_repairs_of_its_own_sees_an_empty_personal_report(self):
        self._build_dataset()
        warehouse = Employee.objects.get(username='sklad_ra')

        resp = self._get(warehouse, date_from='2000-01-01')

        self.assertFalse(resp.context['is_admin'])
        self.assertEqual(resp.context['total_orders'], 0)

    def test_the_report_page_loads_without_error(self):
        self._build_dataset()

        resp = self._get(self.admin)

        self.assertEqual(resp.status_code, 200)

    def test_export_has_three_sheets_with_the_expected_data(self):
        self._build_dataset()
        self.client_http.force_login(self.admin)

        resp = self.client_http.get('/reports/repair-analytics/export/', {'date_from': '2000-01-01'})

        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertEqual(
            set(wb.sheetnames),
            {'По инженерам', 'По типу неисправности', 'Текущая загрузка'},
        )

        engineers = {row[0].value: row[1].value for row in wb['По инженерам'].iter_rows(min_row=2)}
        self.assertEqual(engineers.get('Иванов'), 1)

        faults = {row[0].value: row[1].value for row in wb['По типу неисправности'].iter_rows(min_row=2)}
        self.assertEqual(faults.get(str(self.fault_a)), 1)


class DebtReportTests(TestCase):
    """Отчёт «Задолженности»: просрочка и след от напоминаний."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_debt', full_name='Админ', password='pass')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        client_obj = ClientModel.objects.create(name='ООО Должник')
        self.order = RepairOrder.objects.create(
            client=client_obj, payment_status='unpaid',
            invoice_date=datetime.date.today() - datetime.timedelta(days=30))
        model = EquipmentModel.objects.create(name='БУАД-отчёт')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-R1'),
            repair_cost=2500,
        )

    @override_settings(DEBT_OVERDUE_DAYS=14)
    def test_overdue_is_marked(self):
        resp = self.client_http.get('/reports/debtors/')

        self.assertContains(resp, 'просрочен 30 дн.')

    def test_last_reminder_is_shown(self):
        """Иначе на вопрос «мы им вообще писали?» отвечать нечем."""
        Notification.objects.create(
            event='debt_reminder', recipient='debtor@example.com',
            subject='Оплата', body='Текст', repair_order=self.order,
        )
        resp = self.client_http.get('/reports/debtors/')

        # localdate, а не date.today(): вторая берёт дату по часам сервера
        # (UTC), а last_reminder приводится к Europe/Moscow — и с 21:00 UTC
        # до полуночи эти две даты разные, отчего тест падал каждую ночь
        self.assertEqual(
            timezone.localtime(resp.context['orders'][0].last_reminder).date(),
            timezone.localdate())

    def test_the_report_shows_the_remainder(self):
        Payment.objects.create(repair_order=self.order, amount=1000)

        resp = self.client_http.get('/reports/debtors/')

        self.assertEqual(resp.context['total_debt'], Decimal('1500'))
        self.assertEqual(resp.context['orders'][0].debt, Decimal('1500'))

    def test_the_total_is_not_multiplied_by_several_payments(self):
        """Соединение с оплатами могло посчитать заказ трижды."""
        for _ in range(3):
            Payment.objects.create(repair_order=self.order, amount=100)

        resp = self.client_http.get('/reports/debtors/')

        self.assertEqual(resp.context['total_debt'], Decimal('2200'))

    def test_total_is_not_doubled_by_the_reminder_join(self):
        """Регрессия: соединение с очередью могло посчитать сумму дважды."""
        for _ in range(3):
            Notification.objects.create(
                event='debt_reminder', recipient='debtor@example.com',
                subject='Оплата', body='Текст', repair_order=self.order,
            )
        resp = self.client_http.get('/reports/debtors/')

        self.assertEqual(resp.context['total_debt'], 2500)
        self.assertEqual(len(resp.context['orders']), 1)


class MaxTransportTests(TestCase):
    """Отправка в MAX: что именно уходит в сеть и как разбирается ответ."""

    def _fake_urlopen(self, body='{"message": {}}', calls=None):
        """Подменяет сеть, запоминая запрос."""
        class _Response:
            def __enter__(self_inner):
                return self_inner

            def __exit__(self_inner, *args):
                return False

            def read(self_inner):
                return body.encode('utf-8')

        def _open(req, timeout=None):
            if calls is not None:
                calls.append(req)
            return _Response()

        return _open

    @override_settings(MAX_BOT_TOKEN='secret-token',
                       MAX_API_URL='https://platform-api2.max.ru')
    def test_personal_message_goes_to_user_id(self):
        calls = []
        with patch('core.messengers.request.urlopen', self._fake_urlopen(calls=calls)):
            messengers.send_max_message('user:842910', 'Дефицит')

        request_sent = calls[0]
        self.assertEqual(
            request_sent.full_url,
            'https://platform-api2.max.ru/messages?user_id=842910',
        )
        self.assertEqual(json.loads(request_sent.data.decode()), {'text': 'Дефицит'})

    @override_settings(MAX_BOT_TOKEN='secret-token')
    def test_token_goes_in_the_header_without_bearer(self):
        """MAX ждёт голый токен: с префиксом Bearer запрос отвергается."""
        calls = []
        with patch('core.messengers.request.urlopen', self._fake_urlopen(calls=calls)):
            messengers.send_max_message('user:1', 'текст')

        self.assertEqual(calls[0].get_header('Authorization'), 'secret-token')

    @override_settings(MAX_BOT_TOKEN='secret-token')
    def test_group_chat_goes_to_chat_id(self):
        calls = []
        with patch('core.messengers.request.urlopen', self._fake_urlopen(calls=calls)):
            messengers.send_max_message('chat:-98765', 'текст')

        self.assertIn('chat_id=-98765', calls[0].full_url)

    @override_settings(MAX_BOT_TOKEN='')
    def test_without_token_it_refuses(self):
        with self.assertRaises(messengers.MaxError):
            messengers.send_max_message('user:1', 'текст')

    @override_settings(MAX_BOT_TOKEN='secret-token')
    def test_unknown_recipient_format_is_rejected(self):
        with self.assertRaises(messengers.MaxError):
            messengers.send_max_message('wh@example.com', 'текст')

    @override_settings(MAX_BOT_TOKEN='secret-token')
    def test_error_inside_a_200_response_is_noticed(self):
        """MAX умеет отвечать «200 OK» с ошибкой в теле."""
        body = '{"code": "not.found", "message": "chat not found"}'
        with patch('core.messengers.request.urlopen', self._fake_urlopen(body=body)):
            with self.assertRaises(messengers.MaxError) as caught:
                messengers.send_max_message('user:1', 'текст')

        self.assertIn('chat not found', str(caught.exception))

    @override_settings(MAX_BOT_TOKEN='secret-token')
    def test_http_error_becomes_a_readable_reason(self):
        import urllib.error

        failure = urllib.error.HTTPError(
            'https://platform-api2.max.ru/messages', 401, 'Unauthorized',
            {}, io.BytesIO(b'{"message": "invalid access_token"}')
        )
        with patch('core.messengers.request.urlopen', side_effect=failure):
            with self.assertRaises(messengers.MaxError) as caught:
                messengers.send_max_message('user:1', 'текст')

        self.assertIn('401', str(caught.exception))
        self.assertIn('invalid access_token', str(caught.exception))


class MaxQueueTests(TestCase):
    """Кому ставится оповещение в MAX при дефиците детали."""

    def setUp(self):
        Employee.objects.create_user(
            username='wh_max', full_name='Кладовщик', password='pass',
            role='warehouse', email='wh@example.com', max_user_id='842910',
        )
        Employee.objects.create_user(
            username='wh_nomax', full_name='Без MAX', password='pass',
            role='warehouse', email='wh2@example.com',
        )
        Employee.objects.create_user(
            username='acc_max', full_name='Бухгалтер', password='pass',
            role='accountant', email='acc@example.com', max_user_id='111',
        )
        self.part = SparePart.objects.create(
            part_number='MAX-1', name='Резистор', current_stock=10, min_stock=5
        )

    def _spend(self):
        self.part.current_stock -= 6
        self.part.save(update_fields=['current_stock'])

    def _max_recipients(self):
        return set(
            Notification.objects.filter(channel='max').values_list('recipient', flat=True)
        )

    @override_settings(NOTIFY_MAX=True, MAX_BOT_TOKEN='t', MAX_GROUP_CHAT_ID='')
    def test_shortage_reaches_those_who_gave_their_id(self):
        self._spend()

        self.assertEqual(self._max_recipients(), {'user:842910'})
        # Почта при этом никуда не делась
        self.assertEqual(
            set(Notification.objects.filter(channel='email').values_list('recipient', flat=True)),
            {'wh@example.com', 'wh2@example.com'},
        )

    @override_settings(NOTIFY_MAX=True, MAX_BOT_TOKEN='t', MAX_GROUP_CHAT_ID='-98765')
    def test_group_chat_replaces_personal_messages(self):
        """Три одинаковых сообщения подряд — это не оповещение, а шум."""
        self._spend()

        self.assertEqual(self._max_recipients(), {'chat:-98765'})

    @override_settings(NOTIFY_MAX=False, MAX_BOT_TOKEN='t')
    def test_disabled_channel_queues_nothing(self):
        self._spend()
        self.assertEqual(Notification.objects.filter(channel='max').count(), 0)

    @override_settings(NOTIFY_MAX=True, MAX_BOT_TOKEN='')
    def test_without_a_token_the_channel_is_silent(self):
        self._spend()
        self.assertEqual(Notification.objects.filter(channel='max').count(), 0)

    @override_settings(NOTIFY_MAX=True, MAX_BOT_TOKEN='t', MAX_GROUP_CHAT_ID='')
    def test_message_carries_the_subject_inside_the_text(self):
        """У сообщения в мессенджере нет отдельного поля темы."""
        self._spend()
        note = Notification.objects.get(channel='max')

        self.assertTrue(note.body.startswith(note.subject))
        self.assertIn('MAX-1', note.body)


class PersonalNotificationChoiceTests(TestCase):
    """Личный выбор канала (`notify_by_*`) поверх глобального включения."""

    def setUp(self):
        self.opted_out = Employee.objects.create_user(
            username='wh_optout', full_name='Отключил MAX', password='pass',
            role='warehouse', email='optout@example.com', max_user_id='555',
            notify_by_max=False,
        )
        self.opted_in = Employee.objects.create_user(
            username='wh_optin', full_name='Не отключал', password='pass',
            role='warehouse', email='optin@example.com', max_user_id='777',
        )
        self.part = SparePart.objects.create(
            part_number='PNC-1', name='Резистор', current_stock=10, min_stock=5
        )

    def _spend(self):
        self.part.current_stock -= 6
        self.part.save(update_fields=['current_stock'])

    @override_settings(NOTIFY_MAX=True, MAX_BOT_TOKEN='t', MAX_GROUP_CHAT_ID='')
    def test_opting_out_silences_the_channel_even_with_an_id_on_file(self):
        """Личный выбор перевешивает и глобальное включение, и заполненный ID."""
        self._spend()

        recipients = set(
            Notification.objects.filter(channel='max').values_list('recipient', flat=True)
        )
        self.assertEqual(recipients, {'user:777'})

    @override_settings(NOTIFY_MAX=True, MAX_BOT_TOKEN='t', MAX_GROUP_CHAT_ID='')
    def test_an_empty_id_still_means_silence_regardless_of_the_choice(self):
        """Поведение без ID не меняется: раньше молчал, молчит и теперь."""
        self.opted_in.max_user_id = ''
        self.opted_in.save(update_fields=['max_user_id'])
        self._spend()

        self.assertEqual(Notification.objects.filter(channel='max').count(), 0)


class MyNotificationsPageTests(TestCase):
    """Страница «Мои оповещения»: каждый видит и правит только свои три поля."""

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='my_notif', full_name='Сотрудник', password='pass',
            role='warehouse',
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.employee)

    def test_shows_current_settings(self):
        response = self.client_http.get('/my-notifications/')

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['form'].initial.get('notify_by_max', True))

    def test_saves_own_choice(self):
        response = self.client_http.post('/my-notifications/', {
            'notify_by_email': '',
            'notify_by_max': 'on',
            'notify_by_telegram': '',
        })
        self.assertEqual(response.status_code, 302)

        self.employee.refresh_from_db()
        self.assertFalse(self.employee.notify_by_email)
        self.assertTrue(self.employee.notify_by_max)
        self.assertFalse(self.employee.notify_by_telegram)

    def test_does_not_touch_role_or_password(self):
        """Своя форма не даёт поменять то, что не про личный выбор канала."""
        self.client_http.post('/my-notifications/', {
            'notify_by_email': 'on', 'notify_by_max': 'on', 'notify_by_telegram': 'on',
            'role': 'admin',
        })

        self.employee.refresh_from_db()
        self.assertEqual(self.employee.role, 'warehouse')


class AdminEditsStaffNotificationChoiceTests(TestCase):
    """Администратор может поменять флаги оповещений другому сотруднику."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_notif_edit', full_name='Админ', password='pass'
        )
        self.staff = Employee.objects.create_user(
            username='staff_notif_edit', full_name='Кладовщик', password='pass',
            role='warehouse', email='staff@example.com',
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def test_admin_can_turn_channels_off_for_another_employee(self):
        response = self.client_http.post(
            f'/management/users/{self.staff.pk}/edit/',
            {
                'username': self.staff.username,
                'full_name': self.staff.full_name,
                'email': self.staff.email,
                'max_user_id': '', 'telegram_chat_id': '',
                'role': 'warehouse', 'is_active': 'on',
                'notify_by_email': '',
                'notify_by_max': '',
                'notify_by_telegram': '',
            },
        )
        self.assertEqual(response.status_code, 302)

        self.staff.refresh_from_db()
        self.assertFalse(self.staff.notify_by_email)
        self.assertFalse(self.staff.notify_by_max)
        self.assertFalse(self.staff.notify_by_telegram)


class MaxDeliveryTests(TestCase):
    """Команда отправки: оповещение MAX не должно уходить почтой и наоборот."""

    def setUp(self):
        self.letter = Notification.objects.create(
            event='low_stock', recipient='wh@example.com',
            subject='Дефицит', body='Текст',
        )
        self.message = Notification.objects.create(
            event='low_stock', channel='max', recipient='user:842910',
            subject='Дефицит', body='Дефицит\n\nТекст',
        )

    def _run(self):
        call_command('send_notifications', stdout=io.StringIO(), stderr=io.StringIO())

    @override_settings(NOTIFICATIONS_ENABLED=True, MAX_BOT_TOKEN='t',
                       EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_each_channel_uses_its_own_transport(self):
        from django.core import mail

        with patch('core.messengers.send_max_message') as send_max:
            self._run()

        self.letter.refresh_from_db()
        self.message.refresh_from_db()

        send_max.assert_called_once_with('user:842910', 'Дефицит\n\nТекст')
        self.assertEqual([m.to for m in mail.outbox], [['wh@example.com']])
        self.assertEqual(self.letter.status, 'sent')
        self.assertEqual(self.message.status, 'sent')

    @override_settings(NOTIFICATIONS_ENABLED=True, MAX_BOT_TOKEN='t',
                       NOTIFICATIONS_MAX_ATTEMPTS=2,
                       EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_a_broken_messenger_does_not_stop_the_post(self):
        from django.core import mail

        with patch('core.messengers.send_max_message',
                   side_effect=messengers.MaxError('MAX недоступен')):
            self._run()

        self.letter.refresh_from_db()
        self.message.refresh_from_db()

        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(self.letter.status, 'sent')
        self.assertEqual(self.message.status, 'pending')
        self.assertIn('MAX недоступен', self.message.last_error)

    @override_settings(NOTIFICATIONS_ENABLED=True, MAX_BOT_TOKEN='t')
    def test_a_max_only_batch_does_not_open_smtp(self):
        """Иначе очередь из одних сообщений упиралась бы в почтовый сервер."""
        self.letter.delete()

        with patch('core.messengers.send_max_message'), \
             patch('core.management.commands.send_notifications.get_connection') as connect:
            self._run()

        connect.assert_not_called()


class TelegramTransportTests(TestCase):
    """Отправка в Telegram: что уходит в сеть и как разбирается ответ."""

    def _fake_urlopen(self, body='{"ok": true, "result": {}}', calls=None):
        class _Response:
            def __enter__(self_inner):
                return self_inner

            def __exit__(self_inner, *args):
                return False

            def read(self_inner):
                return body.encode('utf-8')

        def _open(req, timeout=None):
            if calls is not None:
                calls.append(req)
            return _Response()

        return _open

    @override_settings(TELEGRAM_BOT_TOKEN='123:ABC',
                       TELEGRAM_API_URL='https://api.telegram.org')
    def test_token_is_part_of_the_address(self):
        """У Telegram токен в адресе, а не в заголовке, — в отличие от MAX."""
        calls = []
        with patch('core.messengers.request.urlopen', self._fake_urlopen(calls=calls)):
            messengers.send_telegram_message('842910', 'Дефицит')

        self.assertEqual(
            calls[0].full_url,
            'https://api.telegram.org/bot123:ABC/sendMessage',
        )
        self.assertEqual(
            json.loads(calls[0].data.decode()),
            {'chat_id': '842910', 'text': 'Дефицит'},
        )

    @override_settings(TELEGRAM_BOT_TOKEN='123:ABC')
    def test_group_looks_the_same_as_a_person(self):
        """В Telegram и человек, и группа — это chat_id."""
        calls = []
        with patch('core.messengers.request.urlopen', self._fake_urlopen(calls=calls)):
            messengers.send_telegram_message('-1001234567890', 'текст')

        self.assertEqual(
            json.loads(calls[0].data.decode())['chat_id'], '-1001234567890'
        )

    @override_settings(TELEGRAM_BOT_TOKEN='')
    def test_without_token_it_refuses(self):
        with self.assertRaises(messengers.TelegramError):
            messengers.send_telegram_message('1', 'текст')

    @override_settings(TELEGRAM_BOT_TOKEN='123:ABC')
    def test_empty_recipient_is_rejected(self):
        with self.assertRaises(messengers.TelegramError):
            messengers.send_telegram_message('  ', 'текст')

    @override_settings(TELEGRAM_BOT_TOKEN='123:ABC')
    def test_refusal_inside_a_200_response_is_noticed(self):
        """«ok: false» приходит и с кодом 200 — например, если бота заблокировали."""
        body = '{"ok": false, "error_code": 403, "description": "bot was blocked by the user"}'
        with patch('core.messengers.request.urlopen', self._fake_urlopen(body=body)):
            with self.assertRaises(messengers.TelegramError) as caught:
                messengers.send_telegram_message('1', 'текст')

        self.assertIn('bot was blocked', str(caught.exception))

    @override_settings(TELEGRAM_BOT_TOKEN='123:ABC')
    def test_unreachable_telegram_becomes_a_readable_reason(self):
        """Из России Telegram может быть попросту недоступен."""
        import urllib.error

        with patch('core.messengers.request.urlopen',
                   side_effect=urllib.error.URLError('Network is unreachable')):
            with self.assertRaises(messengers.TelegramError) as caught:
                messengers.send_telegram_message('1', 'текст')

        self.assertIn('Telegram недоступен', str(caught.exception))


class TelegramQueueTests(TestCase):
    """Кому ставится оповещение в Telegram при дефиците детали."""

    def setUp(self):
        Employee.objects.create_user(
            username='wh_tg', full_name='Кладовщик', password='pass',
            role='warehouse', email='wh@example.com', telegram_chat_id='842910',
        )
        Employee.objects.create_user(
            username='wh_notg', full_name='Без Telegram', password='pass',
            role='warehouse', email='wh2@example.com',
        )
        self.part = SparePart.objects.create(
            part_number='TG-1', name='Резистор', current_stock=10, min_stock=5
        )

    def _spend(self):
        self.part.current_stock -= 6
        self.part.save(update_fields=['current_stock'])

    def _recipients(self):
        return set(
            Notification.objects.filter(channel='telegram')
            .values_list('recipient', flat=True)
        )

    @override_settings(NOTIFY_TELEGRAM=True, TELEGRAM_BOT_TOKEN='123:ABC',
                       TELEGRAM_GROUP_CHAT_ID='')
    def test_shortage_reaches_those_who_gave_their_id(self):
        self._spend()
        self.assertEqual(self._recipients(), {'842910'})

    @override_settings(NOTIFY_TELEGRAM=True, TELEGRAM_BOT_TOKEN='123:ABC',
                       TELEGRAM_GROUP_CHAT_ID='-1001234567890')
    def test_group_replaces_personal_messages(self):
        self._spend()
        self.assertEqual(self._recipients(), {'-1001234567890'})

    @override_settings(NOTIFY_TELEGRAM=False, TELEGRAM_BOT_TOKEN='123:ABC')
    def test_disabled_channel_queues_nothing(self):
        self._spend()
        self.assertEqual(Notification.objects.filter(channel='telegram').count(), 0)

    @override_settings(NOTIFY_TELEGRAM=True, TELEGRAM_BOT_TOKEN='')
    def test_without_a_token_the_channel_is_silent(self):
        self._spend()
        self.assertEqual(Notification.objects.filter(channel='telegram').count(), 0)

    @override_settings(NOTIFY_TELEGRAM=True, TELEGRAM_BOT_TOKEN='123:ABC',
                       TELEGRAM_GROUP_CHAT_ID='',
                       NOTIFY_MAX=True, MAX_BOT_TOKEN='t', MAX_GROUP_CHAT_ID='-9')
    def test_channels_do_not_interfere(self):
        """Оба мессенджера и почта работают одновременно и независимо."""
        self._spend()

        self.assertEqual(self._recipients(), {'842910'})
        self.assertEqual(
            set(Notification.objects.filter(channel='max').values_list('recipient', flat=True)),
            {'chat:-9'},
        )
        self.assertEqual(Notification.objects.filter(channel='email').count(), 2)


class TelegramDeliveryTests(TestCase):
    """Команда отправки: Telegram идёт своим транспортом."""

    def setUp(self):
        self.message = Notification.objects.create(
            event='low_stock', channel='telegram', recipient='842910',
            subject='Дефицит', body='Дефицит\n\nТекст',
        )

    def _run(self):
        call_command('send_notifications', stdout=io.StringIO(), stderr=io.StringIO())

    @override_settings(NOTIFICATIONS_ENABLED=True, TELEGRAM_BOT_TOKEN='123:ABC')
    def test_message_goes_through_telegram_and_not_the_post(self):
        with patch('core.messengers.send_telegram_message') as send, \
             patch('core.management.commands.send_notifications.get_connection') as connect:
            self._run()

        self.message.refresh_from_db()

        send.assert_called_once_with('842910', 'Дефицит\n\nТекст')
        connect.assert_not_called()
        self.assertEqual(self.message.status, 'sent')

    @override_settings(NOTIFICATIONS_ENABLED=True, TELEGRAM_BOT_TOKEN='123:ABC',
                       NOTIFICATIONS_MAX_ATTEMPTS=2)
    def test_failure_is_recorded_and_retried(self):
        with patch('core.messengers.send_telegram_message',
                   side_effect=messengers.TelegramError('Telegram недоступен')):
            self._run()

        self.message.refresh_from_db()

        self.assertEqual(self.message.status, 'pending')
        self.assertEqual(self.message.attempts, 1)
        self.assertIn('Telegram недоступен', self.message.last_error)


class MaxRecipientDisplayTests(TestCase):
    """«user:842910» в списке оповещений ни о чём не говорит."""

    def test_known_employee_is_shown_by_name(self):
        Employee.objects.create_user(
            username='wh_disp', full_name='Иванов И.И.', password='pass',
            role='warehouse', max_user_id='842910',
        )
        note = Notification.objects.create(
            event='low_stock', channel='max', recipient='user:842910',
            subject='Дефицит', body='Текст',
        )

        self.assertEqual(note.recipient_display, 'Иванов И.И. (MAX)')

    def test_unknown_id_is_still_readable(self):
        note = Notification.objects.create(
            event='low_stock', channel='max', recipient='user:1',
            subject='Дефицит', body='Текст',
        )
        self.assertIn('1', note.recipient_display)

    def test_email_is_shown_as_is(self):
        note = Notification.objects.create(
            event='low_stock', recipient='wh@example.com',
            subject='Дефицит', body='Текст',
        )
        self.assertEqual(note.recipient_display, 'wh@example.com')

    def test_telegram_employee_is_shown_by_name(self):
        Employee.objects.create_user(
            username='wh_tg_disp', full_name='Сидоров С.С.', password='pass',
            role='warehouse', telegram_chat_id='842910',
        )
        note = Notification.objects.create(
            event='low_stock', channel='telegram', recipient='842910',
            subject='Дефицит', body='Текст',
        )

        self.assertEqual(note.recipient_display, 'Сидоров С.С. (Telegram)')

    def test_telegram_group_is_told_apart_by_the_minus(self):
        """Идентификатор группы в Telegram отрицательный."""
        note = Notification.objects.create(
            event='low_stock', channel='telegram', recipient='-1001234567890',
            subject='Дефицит', body='Текст',
        )

        self.assertIn('чат', note.recipient_display)


class NotificationAdminPageTests(TestCase):
    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_np', full_name='Админ', password='pass'
        )
        self.staff = Employee.objects.create_user(
            username='wh_np', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.failed = Notification.objects.create(
            event='low_stock', recipient='wh@example.com', subject='Дефицит',
            body='Текст', status='failed', attempts=5, last_error='SMTP молчит',
        )

    def test_page_lists_the_queue(self):
        resp = self.client_http.get('/management/notifications/')

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'wh@example.com')
        self.assertContains(resp, 'SMTP молчит')

    def test_max_row_shows_the_channel_and_the_person(self):
        Employee.objects.create_user(
            username='wh_max_page', full_name='Петров П.П.', password='pass',
            role='warehouse', max_user_id='842910',
        )
        Notification.objects.create(
            event='low_stock', channel='max', recipient='user:842910',
            subject='Дефицит', body='Текст',
        )
        resp = self.client_http.get('/management/notifications/')

        self.assertContains(resp, 'MAX')
        self.assertContains(resp, 'Петров П.П.')

    def test_status_filter(self):
        Notification.objects.create(
            event='low_stock', recipient='other@example.com',
            subject='В очереди', body='Текст',
        )
        resp = self.client_http.get('/management/notifications/?status=failed')

        self.assertEqual([n.pk for n in resp.context['notifications']], [self.failed.pk])

    def test_retry_returns_it_to_the_queue(self):
        self.client_http.post(f'/management/notifications/{self.failed.pk}/retry/')
        self.failed.refresh_from_db()

        self.assertEqual(self.failed.status, 'pending')
        self.assertEqual(self.failed.attempts, 0)
        self.assertEqual(self.failed.last_error, '')

    def test_only_admin_gets_in(self):
        staff_client = TestClient()
        staff_client.force_login(self.staff)

        resp = staff_client.get('/management/notifications/')
        self.assertEqual(resp.status_code, 302)

    def test_page_warns_when_sending_is_off(self):
        with override_settings(NOTIFICATIONS_ENABLED=False):
            resp = self.client_http.get('/management/notifications/')
        self.assertFalse(resp.context['sending_enabled'])


class TBankTransportTests(TestCase):
    """Обращение к банку: что уходит в сеть и как разбирается ответ."""

    def _fake_urlopen(self, body='{"operations": []}', calls=None):
        class _Response:
            def __enter__(self_inner):
                return self_inner

            def __exit__(self_inner, *args):
                return False

            def read(self_inner):
                return body.encode('utf-8')

        def _open(req, timeout=None):
            if calls is not None:
                calls.append(req)
            return _Response()

        return _open

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='40802810700006096146',
                       TBANK_API_URL='https://business.tbank.ru/openapi')
    def test_statement_request_carries_account_and_period(self):
        calls = []
        with patch('core.tbank.request.urlopen', self._fake_urlopen(calls=calls)):
            tbank.get_statement(datetime.date(2026, 8, 1), datetime.date(2026, 8, 13))

        sent = calls[0]
        self.assertIn('/openapi/api/v1/statement?', sent.full_url)
        self.assertIn('accountNumber=40802810700006096146', sent.full_url)
        # Банк требует не дату, а момент времени: одну дату он отвергает
        # с «Value '2026-07-29' is not a valid date-time (schema: query.from)»
        self.assertIn('from=2026-08-01T00%3A00%3A00', sent.full_url)
        self.assertIn('till=2026-08-13T23%3A59%3A59', sent.full_url)

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='40802810700006096146')
    def test_token_goes_as_bearer(self):
        calls = []
        with patch('core.tbank.request.urlopen', self._fake_urlopen(calls=calls)):
            tbank.get_statement(datetime.date(2026, 8, 1), datetime.date(2026, 8, 2))

        self.assertEqual(calls[0].get_header('Authorization'), 'Bearer secret')

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='')
    def test_statement_without_an_account_says_so(self):
        with self.assertRaises(tbank.TBankError) as caught:
            tbank.get_statement(datetime.date(2026, 8, 1), datetime.date(2026, 8, 2))

        self.assertIn('TBANK_ACCOUNT', str(caught.exception))

    @override_settings(TBANK_TOKEN='')
    def test_without_a_token_the_channel_is_off(self):
        self.assertFalse(tbank.is_configured())

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='123')
    def test_an_expired_token_is_named_not_just_numbered(self):
        def _fail(req, timeout=None):
            raise urllib_error.HTTPError(req.full_url, 401, 'Unauthorized', {}, io.BytesIO(b'{}'))

        with patch('core.tbank.request.urlopen', _fail):
            with self.assertRaises(tbank.TBankError) as caught:
                tbank.get_statement(datetime.date(2026, 8, 1), datetime.date(2026, 8, 2))

        self.assertIn('TBANK_TOKEN', str(caught.exception))


class TBankStatementPeriodTests(TestCase):
    """Границы периода выписки. Живой банк отверг одну дату — здесь
    закреплено то, что он принял."""

    def test_a_plain_date_is_refused_by_the_bank_so_we_send_a_moment(self):
        moment = tbank._moment(datetime.date(2026, 7, 29))

        self.assertTrue(moment.startswith('2026-07-29T00:00:00'))
        self.assertNotEqual(moment, '2026-07-29')

    def test_the_offset_is_local_and_not_utc(self):
        """Граница периода — это границы рабочего дня. От полуночи UTC
        при нашем UTC+3 период начинался бы с трёх ночи предыдущего дня,
        и на краях выписки операции терялись бы."""
        with override_settings(TIME_ZONE='Europe/Moscow', USE_TZ=True):
            moment = tbank._moment(datetime.date(2026, 7, 29))

        self.assertTrue(moment.endswith('+03:00'), moment)

    def test_the_end_of_the_period_is_the_end_of_the_day(self):
        """`till`, судя по названию, включающий: полночь следующего дня
        захватила бы лишний день целиком."""
        moment = tbank._moment(datetime.date(2026, 7, 29), end_of_day=True)

        self.assertTrue(moment.startswith('2026-07-29T23:59:59'), moment)

    def test_a_datetime_is_taken_as_it_is(self):
        given = timezone.make_aware(
            datetime.datetime(2026, 7, 29, 14, 30, 5)
        )

        self.assertEqual(tbank._moment(given), given.isoformat(timespec='seconds'))

    def test_the_offset_survives_the_query_string(self):
        """Плюс в адресе без кодирования читается как пробел, и банк
        снова ответил бы «не date-time»."""
        query = urllib_parse.urlencode(
            {'from': tbank._moment(datetime.date(2026, 7, 29))}
        )

        self.assertIn('%2B03%3A00', query)
        self.assertNotIn('+03', query)


class TBankParsingTests(TestCase):
    """Разбор выписки. Имена полей у банка не одни и те же, отсюда терпимость."""

    def test_operations_are_found_under_any_known_key(self):
        for key in ('operations', 'items', 'data', 'result', 'transactions'):
            payload = {key: [{'id': '1'}]}

            self.assertEqual(len(tbank.operation_list(payload)), 1, key)

    def test_a_bare_list_is_a_list_of_operations(self):
        self.assertEqual(len(tbank.operation_list([{'id': '1'}, {'id': '2'}])), 2)

    def test_amount_is_read_from_a_number_a_string_or_an_object(self):
        cases = [
            {'id': '1', 'typeOfOperation': 'Credit', 'amount': 14000},
            {'id': '1', 'typeOfOperation': 'Credit', 'amount': '14 000,00'},
            {'id': '1', 'typeOfOperation': 'Credit', 'amount': {'value': '14000.00'}},
        ]
        for raw in cases:
            self.assertEqual(tbank.parse_operation(raw)['amount'], Decimal('14000'), raw)

    def test_date_is_read_with_or_without_time(self):
        for value in ('2026-08-13', '2026-08-13T10:20:30+03:00', '13.08.2026'):
            parsed = tbank.parse_operation({'id': '1', 'operationDate': value})

            self.assertEqual(parsed['operation_date'], datetime.date(2026, 8, 13), value)

    def test_counterparty_is_read_from_a_nested_object_too(self):
        parsed = tbank.parse_operation({
            'id': '1',
            'counterParty': {'name': 'ООО «ЛИФТПРОЕКТ»', 'inn': '9722051089'},
        })

        self.assertEqual(parsed['counterparty'], 'ООО «ЛИФТПРОЕКТ»')
        self.assertEqual(parsed['counterparty_inn'], '9722051089')

    def test_only_incoming_money_is_taken(self):
        payload = {'operations': [
            {'id': 'in', 'typeOfOperation': 'Credit', 'amount': 100},
            {'id': 'out', 'typeOfOperation': 'Debit', 'amount': 100},
        ]}

        found = tbank.incoming_operations(payload)

        self.assertEqual([item['external_id'] for item in found], ['in'])

    def test_without_a_direction_a_negative_amount_is_an_expense(self):
        payload = {'operations': [
            {'id': 'in', 'amount': 100},
            {'id': 'out', 'amount': -100},
        ]}

        found = tbank.incoming_operations(payload)

        self.assertEqual([item['external_id'] for item in found], ['in'])

    def test_an_operation_without_an_id_is_dropped(self):
        """Без идентификатора её не отличить от такой же в следующей выписке."""
        payload = {'operations': [{'typeOfOperation': 'Credit', 'amount': 100}]}

        self.assertEqual(tbank.incoming_operations(payload), [])


class BankOperationTests(TestCase):
    """Поступления из выписки и разнесение их по заказам."""

    def setUp(self):
        self.accountant = Employee.objects.create_user(
            username='buh_bank', full_name='Бухгалтер', password='pass',
            role='accountant')
        self.warehouse = Employee.objects.create_user(
            username='sklad_bank', full_name='Кладовщик', password='pass',
            role='warehouse')
        self.client_http = TestClient()
        self.client_http.force_login(self.accountant)

        self.customer = ClientModel.objects.create(
            name='ООО «ЛИФТПРОЕКТ»', inn='9722051089')
        self.order = RepairOrder.objects.create(
            client=self.customer, invoice_number='942', invoice_date=datetime.date(2026, 8, 1))
        model = EquipmentModel.objects.create(name='Emotron-банк')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-BANK'),
            repair_cost=Decimal('14000'),
        )

    def _operation(self, **overrides):
        data = {
            'external_id': 'op-1',
            'operation_date': datetime.date(2026, 8, 10),
            'amount': Decimal('14000'),
            'purpose': 'Оплата по счету 942 от 01.08.2026, без НДС',
            'counterparty': 'ООО «ЛИФТПРОЕКТ»',
            'counterparty_inn': '9722051089',
            'document_number': '178',
        }
        data.update(overrides)
        return BankOperation.objects.create(**data)

    def test_the_order_is_guessed_by_the_invoice_number(self):
        operation = self._operation()

        self.assertEqual(operation.guess_orders(), [self.order])

    def test_the_order_number_in_the_purpose_wins_over_the_inn(self):
        """Точное попадание не должно тонуть среди прочих долгов заказчика."""
        other = RepairOrder.objects.create(client=self.customer)
        operation = self._operation(
            purpose=f'Оплата по заказу {other.order_number}')

        self.assertEqual(operation.guess_orders(), [other])

    def test_an_invoice_number_with_a_suffix_still_matches(self):
        operation = self._operation(purpose='Оплата счет № 942-01 за ремонт')

        self.assertEqual(operation.guess_orders(), [self.order])

    def test_without_any_hint_debts_of_the_payer_are_offered(self):
        operation = self._operation(purpose='Оплата за услуги')

        self.assertEqual(operation.guess_orders(), [self.order])

    def test_a_stranger_gets_no_suggestions(self):
        operation = self._operation(purpose='Возврат средств', counterparty_inn='7700000000')

        self.assertEqual(operation.guess_orders(), [])

    def test_applying_creates_a_payment_and_closes_the_debt(self):
        operation = self._operation()

        resp = self.client_http.post(
            f'/bank/operations/{operation.pk}/apply/', {'order': self.order.pk})

        self.assertRedirects(resp, '/bank/operations/')
        operation.refresh_from_db()
        self.order.refresh_from_db()
        self.assertEqual(operation.status, 'applied')
        self.assertEqual(operation.payment.amount, Decimal('14000'))
        self.assertEqual(operation.payment.payment_date, datetime.date(2026, 8, 10))
        self.assertEqual(self.order.paid_amount, Decimal('14000'))
        self.assertEqual(self.order.payment_status, 'paid')

    def test_the_payment_note_says_where_the_money_came_from(self):
        operation = self._operation()

        self.client_http.post(
            f'/bank/operations/{operation.pk}/apply/', {'order': self.order.pk})

        operation.refresh_from_db()
        self.assertIn('Выписка Т-Банка', operation.payment.note)
        self.assertIn('178', operation.payment.note)

    def test_the_same_money_cannot_be_applied_twice(self):
        operation = self._operation()
        self.client_http.post(
            f'/bank/operations/{operation.pk}/apply/', {'order': self.order.pk})

        self.client_http.post(
            f'/bank/operations/{operation.pk}/apply/', {'order': self.order.pk})

        self.order.refresh_from_db()
        self.assertEqual(self.order.payments.count(), 1)
        self.assertEqual(self.order.paid_amount, Decimal('14000'))

    def test_cancelling_removes_the_payment_and_returns_the_operation(self):
        operation = self._operation()
        self.client_http.post(
            f'/bank/operations/{operation.pk}/apply/', {'order': self.order.pk})

        self.client_http.post(f'/bank/operations/{operation.pk}/reset/')

        operation.refresh_from_db()
        self.order.refresh_from_db()
        self.assertEqual(operation.status, 'new')
        self.assertIsNone(operation.payment)
        self.assertEqual(self.order.paid_amount, 0)

    def test_deleting_the_payment_frees_the_operation(self):
        """Иначе деньги в банке есть, по заказу их нет, и никто не заметит."""
        operation = self._operation()
        self.client_http.post(
            f'/bank/operations/{operation.pk}/apply/', {'order': self.order.pk})
        operation.refresh_from_db()

        operation.payment.delete()

        operation.refresh_from_db()
        self.assertEqual(operation.status, 'new')
        self.assertIsNone(operation.payment)

    def test_money_not_related_to_orders_can_be_put_aside(self):
        operation = self._operation(purpose='Перевод между своими счетами')

        self.client_http.post(f'/bank/operations/{operation.pk}/skip/')

        operation.refresh_from_db()
        self.assertEqual(operation.status, 'skipped')

    def test_an_applied_operation_is_not_skipped_silently(self):
        operation = self._operation()
        self.client_http.post(
            f'/bank/operations/{operation.pk}/apply/', {'order': self.order.pk})

        self.client_http.post(f'/bank/operations/{operation.pk}/skip/')

        operation.refresh_from_db()
        self.assertEqual(operation.status, 'applied')

    def test_the_list_shows_the_suggestion(self):
        self._operation()

        resp = self.client_http.get('/bank/operations/')

        self.assertContains(resp, self.order.order_number)
        self.assertContains(resp, 'ООО «ЛИФТПРОЕКТ»')

    def test_the_warehouse_does_not_see_the_bank(self):
        """Выписка по расчётному счёту — не то, что нужно кладовщику."""
        self.client_http.force_login(self.warehouse)

        resp = self.client_http.get('/bank/operations/')

        self.assertEqual(resp.status_code, 302)

    def test_the_dashboard_warns_about_unapplied_money(self):
        self._operation()

        resp = self.client_http.get('/')

        self.assertEqual(resp.context['unapplied_operations'], 1)

    def test_the_dashboard_stays_quiet_for_the_warehouse(self):
        """Кладовщик деньги не разносит — и знать о них ему незачем."""
        self._operation()
        self.client_http.force_login(self.warehouse)

        resp = self.client_http.get('/')

        self.assertFalse(resp.context['unapplied_operations'])


class TBankStatementCommandTests(TestCase):
    """Команда загрузки выписки."""

    PAYLOAD = {'operations': [
        {'id': 'op-1', 'typeOfOperation': 'Credit', 'amount': 14000,
         'operationDate': '2026-08-10', 'paymentPurpose': 'Оплата по счету 942',
         'payerName': 'ООО «ЛИФТПРОЕКТ»', 'payerInn': '9722051089',
         'documentNumber': '178'},
        {'id': 'op-2', 'typeOfOperation': 'Debit', 'amount': 500,
         'operationDate': '2026-08-11', 'paymentPurpose': 'Комиссия'},
    ]}

    def _run(self, **options):
        out = io.StringIO()
        with patch('core.tbank.get_statement', return_value=self.PAYLOAD):
            call_command('tbank_statement', stdout=out, stderr=out, **options)
        return out.getvalue()

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='123')
    def test_only_incoming_operations_are_stored(self):
        self._run()

        self.assertEqual(BankOperation.objects.count(), 1)
        operation = BankOperation.objects.get()
        self.assertEqual(operation.external_id, 'op-1')
        self.assertEqual(operation.amount, Decimal('14000'))
        self.assertEqual(operation.counterparty_inn, '9722051089')

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='123')
    def test_running_twice_does_not_duplicate_money(self):
        self._run()
        self._run()

        self.assertEqual(BankOperation.objects.count(), 1)

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='123')
    def test_dry_run_stores_nothing(self):
        output = self._run(dry_run=True)

        self.assertEqual(BankOperation.objects.count(), 0)
        self.assertIn('проверка', output)

    @override_settings(TBANK_TOKEN='')
    def test_without_a_token_the_command_says_so_and_stops(self):
        out = io.StringIO()
        call_command('tbank_statement', stdout=out)

        self.assertIn('не настроен', out.getvalue())
        self.assertEqual(BankOperation.objects.count(), 0)

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='123')
    def test_a_bank_failure_is_reported_without_a_traceback(self):
        out = io.StringIO()
        with patch('core.tbank.get_statement',
                   side_effect=tbank.TBankError('Т-Банк недоступен')):
            call_command('tbank_statement', stdout=out, stderr=out)

        self.assertIn('Выписка не получена', out.getvalue())


class TBankInvoiceBuildingTests(TestCase):
    """Сборка счёта: что именно уходит в банк."""

    def setUp(self):
        self.customer = ClientModel.objects.create(
            name='ООО «ЛИФТПРОЕКТ»', inn='9722051089', kpp='772201001',
            email='buh@liftproekt.ru')
        self.order = RepairOrder.objects.create(client=self.customer)
        model = EquipmentModel.objects.create(
            name='EkoDrive-2.2-1.0', kind='Устройство управления дверьми лифта')
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='13593'),
            work_performed='Ремонт импульсного блока питания,\nзамена транзисторов',
            repair_cost=Decimal('14000'),
        )

    def test_the_line_repeats_the_wording_used_in_hand_written_invoices(self):
        self.assertEqual(
            self.roe.invoice_line,
            'Ремонт Устройство управления дверьми лифта EkoDrive-2.2-1.0 SN:13593 '
            '(Ремонт импульсного блока питания, замена транзисторов)'
        )

    def test_line_breaks_do_not_leak_into_the_invoice(self):
        """Перенос строки разорвал бы ячейку таблицы в PDF банка."""
        self.assertNotIn('\n', self.roe.invoice_line)

    def test_equipment_without_a_price_is_not_a_line(self):
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=EquipmentModel.objects.create(name='Без цены'),
                serial_number='NOPRICE'),
        )

        items = self.order.invoice_items()

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]['price'], 14000.0)

    @override_settings(TBANK_INVOICE_UNIT='шт.', TBANK_INVOICE_VAT='None')
    def test_items_carry_the_unit_and_the_vat_mode(self):
        item = self.order.invoice_items()[0]

        self.assertEqual(item['unit'], 'шт.')
        self.assertEqual(item['vat'], 'None')
        self.assertEqual(item['amount'], 1)

    @override_settings(TBANK_ACCOUNT='40802810700006096146')
    def test_the_payload_matches_the_documented_shape(self):
        payload = tbank.build_invoice(
            number='943',
            items=self.order.invoice_items(),
            payer={'name': self.customer.name, 'inn': self.customer.inn,
                   'kpp': self.customer.kpp},
            emails=['buh@liftproekt.ru'],
            invoice_date=datetime.date(2026, 8, 13),
            due_date=datetime.date(2026, 8, 27),
        )

        self.assertEqual(payload['invoiceNumber'], '943')
        self.assertEqual(payload['invoiceDate'], '2026-08-13')
        self.assertEqual(payload['dueDate'], '2026-08-27')
        self.assertEqual(payload['accountNumber'], '40802810700006096146')
        self.assertEqual(payload['payer'],
                         {'name': 'ООО «ЛИФТПРОЕКТ»', 'inn': '9722051089',
                          'kpp': '772201001'})
        self.assertEqual(payload['contacts'], [{'email': 'buh@liftproekt.ru'}])
        self.assertEqual(len(payload['items']), 1)

    def test_empty_payer_fields_are_not_sent(self):
        """Пустая строка в реквизитах счёта выглядит как ошибка."""
        payload = tbank.build_invoice(
            number='1', items=[{'name': 'x', 'price': 1, 'unit': 'шт.',
                                'vat': 'None', 'amount': 1}],
            payer={'name': 'ИП Петров', 'inn': '770000000000', 'kpp': ''},
        )

        self.assertEqual(payload['payer'], {'name': 'ИП Петров', 'inn': '770000000000'})

    def test_without_recipients_there_are_no_contacts(self):
        payload = tbank.build_invoice(
            number='1', items=[{'name': 'x'}], emails=[])

        self.assertNotIn('contacts', payload)


class TBankInvoiceNumberTests(TestCase):
    """Номер счёта. Банк его не выдаёт — он приходит от нас."""

    def setUp(self):
        self.customer = ClientModel.objects.create(name='ООО «Счётчик»')

    def _order(self, invoice_number=''):
        return RepairOrder.objects.create(
            client=self.customer, invoice_number=invoice_number)

    @override_settings(TBANK_INVOICE_NUMBER_START=1)
    def test_the_next_number_follows_the_biggest_one_seen(self):
        self._order('942')
        self._order('906')

        self.assertEqual(RepairOrder.next_invoice_number(), '943')

    @override_settings(TBANK_INVOICE_NUMBER_START=1)
    def test_non_numeric_numbers_are_not_guessed_from(self):
        """«942-01» и «б/н» в сквозной ряд не встают."""
        self._order('942-01')
        self._order('б/н')
        self._order('900')

        self.assertEqual(RepairOrder.next_invoice_number(), '901')

    @override_settings(TBANK_INVOICE_NUMBER_START=664)
    def test_the_starting_number_is_respected_on_an_empty_base(self):
        self.assertEqual(RepairOrder.next_invoice_number(), '664')

    @override_settings(TBANK_INVOICE_NUMBER_START=664)
    def test_the_start_does_not_pull_the_series_backwards(self):
        self._order('943')

        self.assertEqual(RepairOrder.next_invoice_number(), '944')


class TBankInvoiceSendingTests(TestCase):
    """Отправка счёта. Единственное, что программа создаёт в банке."""

    def setUp(self):
        self.accountant = Employee.objects.create_user(
            username='buh_inv', full_name='Бухгалтер', password='pass',
            role='accountant')
        self.manager = Employee.objects.create_user(
            username='mgr_inv', full_name='Менеджер', password='pass',
            role='repair_manager')
        self.client_http = TestClient()
        self.client_http.force_login(self.accountant)

        self.customer = ClientModel.objects.create(
            name='ООО «ЛИФТПРОЕКТ»', inn='9722051089', email='buh@liftproekt.ru')
        self.order = RepairOrder.objects.create(client=self.customer)
        model = EquipmentModel.objects.create(name='Emotron-счёт')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-INV'),
            repair_cost=Decimal('54000'),
        )

    def _url(self):
        return f'/repair-orders/{self.order.pk}/invoice/'

    def _post(self, **overrides):
        data = {
            'provider': 'tbank',
            'invoice_number': '943',
            'invoice_date': '2026-08-13',
            'due_date': '2026-08-27',
            'emails': 'buh@liftproekt.ru',
        }
        data.update(overrides)
        return self.client_http.post(self._url(), data)

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_sending_records_the_number_and_the_pdf(self):
        with patch('core.tbank.send_invoice',
                   return_value={'pdfUrl': 'https://example.org/1.pdf'}) as send:
            resp = self._post()

        self.assertRedirects(resp, f'/repair-orders/{self.order.pk}/')
        self.order.refresh_from_db()
        self.assertEqual(self.order.invoice_number, '943')
        self.assertEqual(self.order.invoice_date, datetime.date(2026, 8, 13))
        self.assertEqual(self.order.invoice_pdf_url, 'https://example.org/1.pdf')
        self.assertIsNotNone(self.order.invoice_sent_at)
        self.assertEqual(self.order.invoice_error, '')
        self.assertEqual(send.call_count, 1)

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_what_is_sent_is_what_was_shown(self):
        with patch('core.tbank.send_invoice', return_value={}) as send:
            self._post()

        payload = send.call_args[0][0]
        self.assertEqual(payload['invoiceNumber'], '943')
        self.assertEqual(payload['payer']['inn'], '9722051089')
        self.assertEqual(payload['contacts'], [{'email': 'buh@liftproekt.ru'}])
        self.assertEqual(payload['items'][0]['price'], 54000.0)

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_the_send_is_recorded_in_the_order_history(self):
        with patch('core.tbank.send_invoice', return_value={}):
            self._post()

        last = self.order.status_history.order_by('-id').first()
        self.assertIn('943', last.notes)
        self.assertEqual(last.changed_by, self.accountant)

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_a_refusal_is_kept_and_nothing_is_marked_as_sent(self):
        with patch('core.tbank.send_invoice',
                   side_effect=tbank.TBankError('Т-Банк отказал: нет прав')):
            self._post()

        self.order.refresh_from_db()
        self.assertIn('нет прав', self.order.invoice_error)
        self.assertIsNone(self.order.invoice_sent_at)
        self.assertEqual(self.order.invoice_number, '')

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=False)
    def test_sending_is_off_by_default(self):
        """Отправка документа заказчику не должна включаться сама."""
        self.assertFalse(tbank.invoice_enabled())
        with self.assertRaises(tbank.TBankError) as caught:
            tbank.send_invoice({'invoiceNumber': '1', 'items': [{'name': 'x'}]})

        self.assertIn('выключено', str(caught.exception))

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_an_invoice_without_lines_is_refused_before_the_network(self):
        with self.assertRaises(tbank.TBankError) as caught:
            tbank.send_invoice({'invoiceNumber': '1', 'items': []})

        self.assertIn('ни одной позиции', str(caught.exception))

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_an_invoice_without_a_number_is_refused_before_the_network(self):
        with self.assertRaises(tbank.TBankError) as caught:
            tbank.send_invoice({'items': [{'name': 'x'}]})

        self.assertIn('номер счёта', str(caught.exception).lower())

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_a_refusal_inside_a_200_response_is_still_a_refusal(self):
        def _fake(req, timeout=None):
            class _R:
                def __enter__(self_inner):
                    return self_inner

                def __exit__(self_inner, *args):
                    return False

                def read(self_inner):
                    return b'{"errorMessage": "Counterparty not found"}'
            return _R()

        with patch('core.tbank.request.urlopen', _fake):
            with self.assertRaises(tbank.TBankError) as caught:
                tbank.send_invoice({'invoiceNumber': '1', 'items': [{'name': 'x'}]})

        self.assertIn('Counterparty not found', str(caught.exception))

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_the_request_is_a_post_with_json(self):
        calls = []

        def _fake(req, timeout=None):
            calls.append(req)

            class _R:
                def __enter__(self_inner):
                    return self_inner

                def __exit__(self_inner, *args):
                    return False

                def read(self_inner):
                    return b'{"pdfUrl": "https://example.org/x.pdf"}'
            return _R()

        with patch('core.tbank.request.urlopen', _fake):
            tbank.send_invoice({'invoiceNumber': '943', 'items': [{'name': 'Ремонт'}]})

        sent = calls[0]
        self.assertEqual(sent.method, 'POST')
        self.assertIn('/openapi/api/v1/invoice/send', sent.full_url)
        self.assertEqual(sent.get_header('Content-type'), 'application/json')
        self.assertEqual(sent.get_header('Authorization'), 'Bearer secret')
        # Кириллица уходит как есть, а не escape-последовательностями
        self.assertIn('Ремонт', sent.data.decode('utf-8'))

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_a_broken_recipient_stops_the_whole_send(self):
        """Иначе счёт уйдёт не туда, а человек будет уверен, что отправил."""
        with patch('core.tbank.send_invoice') as send:
            resp = self._post(emails='buh@liftproekt.ru, кривой-адрес')

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(send.call_count, 0)
        self.assertContains(resp, 'Непохоже на адрес почты')

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_a_due_date_before_the_invoice_date_is_refused(self):
        with patch('core.tbank.send_invoice') as send:
            resp = self._post(due_date='2026-08-01')

        self.assertEqual(send.call_count, 0)
        self.assertContains(resp, 'раньше даты счёта')

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_the_page_shows_the_lines_and_the_total(self):
        resp = self.client_http.get(self._url())

        self.assertContains(resp, 'SN-INV')
        # Запятая, а не точка: локаль ru-ru
        self.assertContains(resp, '54000,00')
        self.assertContains(resp, 'ООО «ЛИФТПРОЕКТ»')

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=False)
    def test_the_page_says_why_the_button_is_dead(self):
        resp = self.client_http.get(self._url())

        self.assertContains(resp, 'TBANK_INVOICE_ENABLED')
        self.assertContains(resp, 'disabled')

    @override_settings(TBANK_TOKEN='secret', TBANK_INVOICE_ENABLED=True)
    def test_a_repeat_send_is_warned_about(self):
        self.order.invoice_sent_at = timezone.now()
        self.order.invoice_number = '943'
        self.order.save(update_fields=['invoice_sent_at', 'invoice_number'])

        resp = self.client_http.get(self._url())

        self.assertContains(resp, 'уже выставлен счёт')

    def test_the_repair_manager_does_not_issue_invoices(self):
        self.client_http.force_login(self.manager)

        resp = self.client_http.get(self._url())

        self.assertEqual(resp.status_code, 302)


class QuoteTests(TestCase):
    """Коммерческое предложение по заказу."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_quote', full_name='Админ', password='pass')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        organization = Organization.get_solo()
        organization.name = 'ИП Петров П. П.'
        organization.inn = '772600000000'
        organization.ogrn = '324774600000000'
        organization.address = '117042, Москва, ул. Примерная, д. 1'
        organization.bank_name = 'АО «ТБанк»'
        organization.bank_bik = '044525974'
        organization.bank_account = '40802810700006096146'
        organization.corr_account = '30101810145250000974'
        organization.tax_note = 'Без НДС, применяется УСН, ПСН'
        organization.signatory_position = 'Индивидуальный предприниматель'
        organization.signatory_name = 'Петров П. П.'
        organization.save()

        self.customer = ClientModel.objects.create(
            name='ООО ГК «Промресурс»',
            address='305048, г. Курск, проспект Дружбы, д. 9А')
        self.order = RepairOrder.objects.create(client=self.customer)
        model = EquipmentModel.objects.create(
            name='EkoDrive-2.3-1.3', kind='Привод дверей')
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='44223'),
        )

    def _edit_url(self):
        return f'/repair-orders/{self.order.pk}/quote/edit/'

    def _print_url(self):
        return f'/repair-orders/{self.order.pk}/quote/'

    def _fill(self, **overrides):
        data = {
            'quote_subject': 'на ремонт приводов дверей EkoDrive-2.3-1.3',
            'quote_date': '2026-02-06',
            'quote_valid_until': '2026-02-20',
            'quote_lead_time': '3-14',
            'quote_payment_terms': '100% предоплата покупателем',
            'quote_delivery_terms': 'до склада Покупателя включена в Предложение',
            'order_equipments-TOTAL_FORMS': '1',
            'order_equipments-INITIAL_FORMS': '1',
            'order_equipments-MIN_NUM_FORMS': '0',
            'order_equipments-MAX_NUM_FORMS': '1000',
            'order_equipments-0-id': str(self.roe.pk),
            'order_equipments-0-proposed_work':
                'Ремонт импульсного блока питания,\nзамена транзисторов ВКО, ВКЗ, РЕВ',
            'order_equipments-0-repair_complexity': 'simple',
            'order_equipments-0-estimated_cost': '6000',
        }
        data.update(overrides)
        return self.client_http.post(self._edit_url(), data)

    def test_filling_the_quote_opens_it(self):
        resp = self._fill()

        self.assertRedirects(resp, self._print_url())
        self.order.refresh_from_db()
        self.roe.refresh_from_db()
        self.assertEqual(self.order.quote_date, datetime.date(2026, 2, 6))
        self.assertEqual(self.roe.estimated_cost, Decimal('6000'))
        self.assertEqual(self.roe.repair_complexity, 'simple')

    def test_the_quote_carries_the_letterhead_and_the_bank_details(self):
        self._fill()

        resp = self.client_http.get(self._print_url())

        self.assertContains(resp, 'ИП Петров П. П.')
        self.assertContains(resp, 'ОГРН/ОГРНИП 324774600000000')
        self.assertContains(resp, 'АО «ТБанк» БИК 044525974')
        self.assertContains(resp, '40802810700006096146')
        self.assertContains(resp, 'Без НДС, применяется УСН, ПСН')

    def test_the_quote_names_the_addressee_with_the_address(self):
        self._fill()

        resp = self.client_http.get(self._print_url())

        self.assertContains(resp, 'ООО ГК «Промресурс»')
        self.assertContains(resp, 'проспект Дружбы')

    def test_the_table_repeats_the_shape_of_their_own_quotes(self):
        self._fill()

        resp = self.client_http.get(self._print_url())

        self.assertContains(resp, 'Ремонт импульсного блока питания')
        self.assertContains(resp, '44223')
        self.assertContains(resp, 'Простой')
        self.assertContains(resp, '3-14')
        self.assertContains(resp, '100% предоплата покупателем')
        self.assertContains(resp, 'действительно до 20.02.2026')

    def test_line_breaks_do_not_break_the_table_cell(self):
        self.roe.proposed_work = 'Ремонт блока питания,\nзамена транзисторов'
        self.roe.save(update_fields=['proposed_work'])

        self.assertEqual(self.roe.quote_line,
                         'Ремонт блока питания, замена транзисторов')

    def test_proposed_work_is_not_the_same_as_performed_work(self):
        """Предложение пишут до согласия, выполненные работы — после ремонта."""
        self.roe.work_performed = 'Заменён блок питания'
        self.roe.proposed_work = 'Ремонт блока питания'
        self.roe.save(update_fields=['work_performed', 'proposed_work'])

        self.assertEqual(self.roe.quote_line, 'Ремонт блока питания')

    def test_without_proposed_work_the_performed_one_is_used(self):
        self.roe.work_performed = 'Заменён блок питания'
        self.roe.save(update_fields=['work_performed'])

        self.assertEqual(self.roe.quote_line, 'Заменён блок питания')

    def test_with_nothing_written_the_line_is_still_meaningful(self):
        self.assertEqual(self.roe.quote_line, 'Ремонт Привод дверей EkoDrive-2.3-1.3')

    def test_the_estimate_is_preferred_over_the_final_cost(self):
        """Предложение делают до ремонта: оценка ближе к разговору с заказчиком."""
        self.roe.repair_cost = Decimal('9000')
        self.roe.estimated_cost = Decimal('6000')
        self.roe.save(update_fields=['repair_cost', 'estimated_cost'])

        self.assertEqual(self.roe.quote_price, Decimal('6000'))

    def test_without_an_estimate_the_final_cost_is_used(self):
        self.roe.repair_cost = Decimal('9000')
        self.roe.save(update_fields=['repair_cost'])

        self.assertEqual(self.roe.quote_price, Decimal('9000'))

    def test_a_line_without_a_price_is_printed_with_a_dash_not_dropped(self):
        """Потерять строку молча хуже, чем напечатать её без суммы."""
        second = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=EquipmentModel.objects.create(name='Без оценки'),
                serial_number='NOEST'),
        )
        self.roe.estimated_cost = Decimal('6000')
        self.roe.save(update_fields=['estimated_cost'])

        rows = self.order.quote_rows()

        self.assertEqual(len(rows), 2)
        self.assertIsNone(rows[1]['price'])
        self.assertEqual(second.quote_price, None)

    def test_the_total_skips_unpriced_lines(self):
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=EquipmentModel.objects.create(name='Без оценки 2'),
                serial_number='NOEST2'),
        )
        self.roe.estimated_cost = Decimal('6000')
        self.roe.save(update_fields=['estimated_cost'])

        self.assertEqual(self.order.quote_total, Decimal('6000'))

    def test_an_empty_quote_warns_instead_of_printing_a_blank(self):
        resp = self.client_http.get(self._print_url())

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Условия предложения не заполнены')

    def test_a_validity_date_before_the_quote_date_is_refused(self):
        resp = self._fill(quote_valid_until='2026-01-01')

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Срок действия раньше даты предложения')
        self.order.refresh_from_db()
        self.assertIsNone(self.order.quote_date)

    def test_the_order_card_leads_to_filling_while_the_quote_is_empty(self):
        resp = self.client_http.get(f'/repair-orders/{self.order.pk}/')

        self.assertContains(resp, self._edit_url())

    def test_a_filled_quote_is_linked_from_the_order_card(self):
        self._fill()

        resp = self.client_http.get(f'/repair-orders/{self.order.pk}/')

        self.assertContains(resp, self._print_url())

    def test_a_missing_client_address_is_pointed_out_before_printing(self):
        self.customer.address = ''
        self.customer.save(update_fields=['address'])

        resp = self.client_http.get(self._edit_url())

        self.assertContains(resp, 'не заполнен адрес')

    def test_the_client_form_keeps_the_address(self):
        resp = self.client_http.post('/clients/create/', {
            'name': 'ООО «Адресное»', 'inn': '', 'kpp': '',
            'address': '305048, г. Курск, проспект Дружбы, д. 9А',
            'contact_person': '', 'phone': '', 'email': '',
        }, follow=True)

        self.assertEqual(resp.status_code, 200)
        created = ClientModel.objects.get(name='ООО «Адресное»')
        self.assertEqual(created.address, '305048, г. Курск, проспект Дружбы, д. 9А')

    def test_the_organization_form_keeps_the_bank_details(self):
        resp = self.client_http.post('/management/organization/', {
            'name': 'ИП Петров П. П.', 'inn': '772600000000', 'kpp': '',
            'ogrn': '324774600000000', 'address': 'Москва', 'phone': '', 'email': '',
            'signatory_position': 'ИП', 'signatory_name': 'Петров П. П.',
            'bank_name': 'АО «ТБанк»', 'bank_bik': '044525974',
            'bank_account': '40802810700006096146',
            'corr_account': '30101810145250000974',
            'tax_note': 'Без НДС, применяется УСН',
        }, follow=True)

        self.assertEqual(resp.status_code, 200)
        organization = Organization.get_solo()
        self.assertEqual(organization.bank_bik, '044525974')
        self.assertEqual(organization.ogrn, '324774600000000')

    def test_the_bank_lines_are_empty_when_nothing_is_filled(self):
        Organization.objects.all().delete()

        self.assertEqual(Organization.get_solo().bank_lines, [])

    def test_the_quote_requires_login(self):
        resp = TestClient().get(self._print_url())

        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login/', resp['Location'])


class QrLinkLengthTests(TestCase):
    """Длина ссылки в QR. Этикетка — бумажка, чинить её после печати нечем."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_qr', full_name='Админ', password='pass')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.part = SparePart.objects.create(
            part_number='QR-1', name='Резистор', current_stock=5)
        self.cell = StorageCell.objects.create(
            cabinet=Cabinet.objects.get_or_create(number=1)[0], row_number=1, cell_row=1)
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО «QR»'))
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=EquipmentModel.objects.create(name='QR-модель'),
                serial_number='QR-SN'),
        )

    def test_the_threshold_matches_the_qr_version_boundary(self):
        """15-й знак переводит код с 21 модуля на 25.

        Содержимое теперь короткое («u/123» — пять знаков), и в 21 модуль
        оно укладывается с запасом даже на пятизначные номера. Проверка
        оставлена сторожем: если в код когда-нибудь решат положить что-то
        ещё, беда всплывёт до печати сотни наклеек, а не после.
        """
        self.assertEqual(views.QR_MAX_CHARS, 14)

    def test_a_short_payload_says_nothing(self):
        self.assertEqual(views.qr_length_warning(['u/1']), '')

    def test_a_payload_exactly_at_the_limit_is_still_fine(self):
        self.assertEqual(views.qr_length_warning(['h' * 14]), '')

    def test_one_character_over_the_limit_warns(self):
        warning = views.qr_length_warning(['h' * 15])

        self.assertIn('15', warning)
        # Настройка адреса тут больше ни при чём: её в коде нет
        self.assertNotIn('LABEL_BASE_URL', warning)

    def test_the_longest_payload_on_the_page_decides(self):
        """Длина растёт с номером записи, а не от настройки."""
        warning = views.qr_length_warning(['короткая', 'h' * 50])

        self.assertIn('50', warning)

    def test_no_payloads_at_all_is_not_a_problem(self):
        self.assertEqual(views.qr_length_warning([]), '')

    @override_settings(LABEL_BASE_URL='http://lifteam.taile9b605.ts.net')
    def test_the_magic_dns_name_fits(self):
        for url in (f'/parts/{self.part.pk}/label/',
                    f'/storage-cells/{self.cell.pk}/label/',
                    f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/label/'):
            resp = self.client_http.get(url)

            self.assertEqual(resp.status_code, 200, url)
            self.assertEqual(resp.context['qr_warning'], '', url)

    @override_settings(LABEL_BASE_URL='http://' + 'x' * 60 + '.example.org')
    def test_even_an_absurd_address_leaves_the_code_alone(self):
        """Прежде длинная настройка делала код мельче и вызывала
        предупреждение. Теперь адрес в код не попадает вовсе — и это
        сильнее предупреждения: беды нет, а не о ней сообщают."""
        for url in (f'/parts/{self.part.pk}/label/',
                    '/parts/labels/?ids=%d' % self.part.pk,
                    '/storage-cells/labels/?cabinet=1'):
            resp = self.client_http.get(url)

            self.assertEqual(resp.status_code, 200, url)
            self.assertEqual(resp.context['qr_warning'], '', url)

    @override_settings(LABEL_BASE_URL='http://' + 'x' * 60 + '.example.org')
    def test_the_code_itself_stays_short(self):
        with patch('core.views.generate_qr_image') as qr:
            qr.return_value = 'data:image/png;base64,x'
            self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertEqual(qr.call_args[0][0], f'p/{self.part.pk}')

    @override_settings(LABEL_BASE_URL='http://lifteam.taile9b605.ts.net')
    def test_the_link_shown_to_a_human_still_carries_the_address(self):
        """Ошибку в адресе иначе видно только со сканером у стеллажа."""
        resp = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertEqual(
            resp.context['qr_url'],
            f'http://lifteam.taile9b605.ts.net/p/{self.part.pk}')

    def test_the_warning_is_not_printed_on_the_sticker(self):
        """Печатать предупреждение на наклейку 43 мм незачем."""
        base = (settings.BASE_DIR / 'core/templates/core/_qr_warning.html'
                ).read_text(encoding='utf-8')

        self.assertIn('no-print', base)


class QrLinkVisibilityTests(TestCase):
    """Куда ведёт код, должно быть видно на экране, а не только со сканером."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_qrvis', full_name='Админ', password='pass')
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.part = SparePart.objects.create(
            part_number='VIS-1', name='Диод', current_stock=5)
        self.cell = StorageCell.objects.create(
            cabinet=Cabinet.objects.get_or_create(number=2)[0], row_number=1, cell_row=1)
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО «Видно»'))
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=EquipmentModel.objects.create(name='VIS-модель'),
                serial_number='VIS-SN'),
        )

    @override_settings(LABEL_BASE_URL='http://lifteam.taile9b605.ts.net')
    def test_single_label_pages_show_the_exact_link(self):
        pages = {
            f'/parts/{self.part.pk}/label/':
                f'http://lifteam.taile9b605.ts.net/p/{self.part.pk}',
            f'/storage-cells/{self.cell.pk}/label/':
                f'http://lifteam.taile9b605.ts.net/c/{self.cell.pk}',
            f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/label/':
                f'http://lifteam.taile9b605.ts.net/u/{self.roe.pk}',
        }
        for url, expected in pages.items():
            resp = self.client_http.get(url)

            self.assertContains(resp, expected, msg_prefix=url)

    @override_settings(LABEL_BASE_URL='http://lifteam.taile9b605.ts.net')
    def test_batch_pages_show_the_base(self):
        for url in (f'/parts/labels/?ids={self.part.pk}',
                    '/storage-cells/labels/?cabinet=2'):
            resp = self.client_http.get(url)

            self.assertEqual(resp.context['qr_base'],
                             'http://lifteam.taile9b605.ts.net', url)
            self.assertContains(resp, 'Код ведёт на', msg_prefix=url)

    @override_settings(LABEL_BASE_URL='http://192.168.1.50')
    def test_a_local_address_left_in_env_is_visible_on_the_page(self):
        """Именно так это и ловится: настройка из .env перебивает умолчание."""
        resp = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertContains(resp, 'http://192.168.1.50/p/')

    @override_settings(LABEL_BASE_URL='http://lifteam.taile9b605.ts.net')
    def test_the_link_is_not_printed_on_the_sticker(self):
        resp = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertContains(resp, 'text-muted small no-print')


class CabinetTests(TestCase):
    """Кассетницы: своя раскладка у каждой, ряды разной ширины."""

    def setUp(self):
        self.warehouse = Employee.objects.create_user(
            username='sklad_cab', full_name='Кладовщик', password='pass',
            role='warehouse')
        self.manager = Employee.objects.create_user(
            username='mgr_cab', full_name='Менеджер', password='pass',
            role='repair_manager')
        self.client_http = TestClient()
        self.client_http.force_login(self.warehouse)

    def _create(self, **overrides):
        data = {'number': 1, 'name': 'Резисторы', 'note': '',
                'layout': '8, 8, 8, 4, 4'}
        data.update(overrides)
        return self.client_http.post('/storage-cells/cabinets/create/', data)

    def test_a_cabinet_gets_exactly_the_cells_of_its_layout(self):
        self._create()

        cabinet = Cabinet.objects.get(number=1)
        self.assertEqual(cabinet.layout(), [8, 8, 8, 4, 4])
        self.assertEqual(cabinet.cell_count, 32)

    def test_rows_may_differ_in_width(self):
        """Ряд из четырёх — это четыре крупных ящика, а не четыре из восьми."""
        self._create(layout='8, 2')

        cabinet = Cabinet.objects.get(number=1)
        self.assertEqual(cabinet.layout(), [8, 2])
        self.assertEqual(
            sorted(cabinet.cells.filter(row_number=2).values_list('cell_row', flat=True)),
            [1, 2])

    def test_the_layout_is_parsed_from_any_separator(self):
        for text in ('8,8,4', '8 8 4', '8, 8, 4', '8х8х4'):
            self.assertEqual(parse_layout(text), [8, 8, 4], text)

    def test_the_address_follows_the_cabinet_number(self):
        self._create(number=7)
        cell = Cabinet.objects.get(number=7).cells.get(row_number=1, cell_row=1)

        self.assertEqual(cell.address, 'К7-Р1-Я1')

    def test_the_next_number_is_suggested(self):
        self._create(number=3)

        self.assertEqual(Cabinet.next_number(), 4)

    def test_growing_the_layout_adds_cells_and_keeps_the_old_ones(self):
        self._create(layout='4')
        cabinet = Cabinet.objects.get(number=1)
        kept = cabinet.cells.get(row_number=1, cell_row=1)

        self.client_http.post(f'/storage-cells/cabinets/{cabinet.pk}/edit/', {
            'number': 1, 'name': 'Резисторы', 'note': '', 'layout': '4, 4'})

        self.assertEqual(cabinet.layout(), [4, 4])
        self.assertTrue(StorageCell.objects.filter(pk=kept.pk).exists())

    def test_shrinking_removes_empty_cells(self):
        self._create(layout='4, 4')
        cabinet = Cabinet.objects.get(number=1)

        self.client_http.post(f'/storage-cells/cabinets/{cabinet.pk}/edit/', {
            'number': 1, 'name': '', 'note': '', 'layout': '4'})

        self.assertEqual(cabinet.layout(), [4])
        self.assertEqual(cabinet.cell_count, 4)

    def test_shrinking_refuses_to_drop_cells_with_parts(self):
        """Иначе деталь на складе есть, а где она лежит — уже неизвестно."""
        self._create(layout='4, 4')
        cabinet = Cabinet.objects.get(number=1)
        part = SparePart.objects.create(part_number='CAB-1', name='Резистор',
                                        current_stock=10)
        cabinet.cells.get(row_number=2, cell_row=1).parts.add(part)

        resp = self.client_http.post(f'/storage-cells/cabinets/{cabinet.pk}/edit/', {
            'number': 1, 'name': '', 'note': '', 'layout': '4'})

        self.assertContains(resp, 'К1-Р2-Я1')
        self.assertEqual(cabinet.layout(), [4, 4])

    def test_an_empty_cabinet_can_be_deleted(self):
        self._create()
        cabinet = Cabinet.objects.get(number=1)

        self.client_http.post(f'/storage-cells/cabinets/{cabinet.pk}/delete/')

        self.assertFalse(Cabinet.objects.filter(pk=cabinet.pk).exists())
        self.assertEqual(StorageCell.objects.count(), 0)

    def test_a_cabinet_with_parts_is_not_deleted(self):
        self._create()
        cabinet = Cabinet.objects.get(number=1)
        part = SparePart.objects.create(part_number='CAB-2', name='Диод',
                                        current_stock=5)
        cabinet.cells.first().parts.add(part)

        self.client_http.post(f'/storage-cells/cabinets/{cabinet.pk}/delete/')

        self.assertTrue(Cabinet.objects.filter(pk=cabinet.pk).exists())

    def test_an_empty_layout_is_refused(self):
        resp = self._create(layout='')

        self.assertIn('layout', resp.context['form'].errors)
        self.assertEqual(Cabinet.objects.count(), 0)

    def test_too_many_cells_in_a_row_are_refused(self):
        resp = self._create(layout='100')

        self.assertIn('layout', resp.context['form'].errors)
        self.assertEqual(Cabinet.objects.count(), 0)

    def test_the_number_stays_unique(self):
        self._create(number=1)

        resp = self._create(number=1)

        self.assertEqual(Cabinet.objects.count(), 1)
        self.assertEqual(resp.status_code, 200)

    def test_the_grid_draws_rows_by_their_real_width(self):
        self._create(layout='8, 2')

        resp = self.client_http.get('/storage-cells/?cabinet=1')

        # Восемь в ряду — по 12,5% ширины, два — по 50%
        self.assertContains(resp, '12.5000%')
        self.assertContains(resp, '50.0000%')

    def test_the_width_is_written_with_a_dot_not_a_comma(self):
        """Локаль ru-ru печатает дробь через запятую, а «calc(12,5% - 4px)» —
        невалидный CSS: ячейки остались бы без ширины."""
        self._create(layout='8, 2')

        resp = self.client_http.get('/storage-cells/?cabinet=1')

        self.assertNotContains(resp, 'calc(12,5')

    def test_the_grid_survives_an_empty_base(self):
        """Свежая установка без кассетниц не должна падать."""
        resp = self.client_http.get('/storage-cells/')

        self.assertEqual(resp.status_code, 200)

    def test_the_grid_falls_back_to_the_first_cabinet(self):
        self._create(number=5)

        resp = self.client_http.get('/storage-cells/?cabinet=99')

        self.assertEqual(resp.context['cabinet'].number, 5)

    def test_the_repair_manager_does_not_rebuild_the_warehouse(self):
        self.client_http.force_login(self.manager)

        resp = self.client_http.get('/storage-cells/cabinets/')

        self.assertEqual(resp.status_code, 302)

    def test_the_list_shows_the_layout_and_what_is_occupied(self):
        self._create(layout='4, 4')
        cabinet = Cabinet.objects.get(number=1)
        part = SparePart.objects.create(part_number='CAB-3', name='Кондёр',
                                        current_stock=3)
        cabinet.cells.first().parts.add(part)

        resp = self.client_http.get('/storage-cells/cabinets/')

        self.assertContains(resp, '4, 4')
        self.assertContains(resp, 'Резисторы')

    def test_init_cells_creates_cabinets_with_the_given_layout(self):
        out = io.StringIO()
        call_command('init_cells', cabinets=2, layout='3,3', stdout=out)

        self.assertEqual(Cabinet.objects.count(), 2)
        self.assertEqual(Cabinet.objects.get(number=1).layout(), [3, 3])
        self.assertEqual(StorageCell.objects.count(), 12)

    def test_init_cells_does_not_recut_existing_cabinets(self):
        """Раскладку могли поменять руками — повторный запуск её не трогает."""
        self._create(number=1, layout='2')
        call_command('init_cells', cabinets=1, layout='8,8', stdout=io.StringIO())

        self.assertEqual(Cabinet.objects.get(number=1).layout(), [2])


class CommonLabelLayoutTests(TestCase):
    """Этикетка детали и этикетка ячейки печатаются одним шаблоном.

    Раньше разметок было две, и одна и та же деталь на пакете и на ячейке
    выглядела по-разному: где артикул, где название, адрес то сверху,
    то снизу.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_common_label', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.cabinet = Cabinet.objects.create(number=7)
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.first()

        self.part = SparePart.objects.create(
            part_number='CL-1', name='Резистор 10к', component_type='Резистор',
            package='0805', resistance=10, resistance_unit='кОм',
            description='Ставится в блок питания БУАД',
        )
        self.cell.parts.add(self.part)

    def _both_pages(self):
        return (
            f'/parts/{self.part.pk}/label/',
            f'/storage-cells/{self.cell.pk}/label/',
        )

    def test_both_labels_use_the_same_template(self):
        for page in self._both_pages():
            with self.subTest(page=page):
                response = self.client_http.get(page)
                self.assertTemplateUsed(response, 'core/_label_part.html')
                self.assertTemplateUsed(response, 'core/_label_part_styles.html')

    def test_a_single_part_prints_the_same_on_both(self):
        """Ячейка с одной деталью и пакет с ней же — одно и то же содержимое."""
        for page in self._both_pages():
            with self.subTest(page=page):
                response = self.client_http.get(page)
                self.assertEqual(response.context['title'], 'CL-1')
                self.assertEqual(response.context['package'], '0805')
                self.assertEqual(response.context['address'], self.cell.address)
                self.assertEqual(
                    response.context['description'], 'Ставится в блок питания БУАД'
                )

    def test_the_package_is_printed_below_and_stands_out(self):
        """По корпусу подбирают замену: 0805 вместо 1206 на плату не встанет."""
        for page in self._both_pages():
            with self.subTest(page=page):
                response = self.client_http.get(page)
                self.assertContains(response, 'label-package')
                self.assertContains(response, '0805')

    def test_the_specs_go_next_to_the_article(self):
        response = self.client_http.get(self._both_pages()[0])

        self.assertEqual(response.context['specs'], 'Резистор · 10кОм')

    def test_a_name_repeating_the_article_is_not_printed_twice(self):
        """Артикул уже стоит сверху и самым крупным шрифтом."""
        twin = SparePart.objects.create(part_number='CL-TWIN', name='CL-TWIN',
                                        component_type='Диод')

        response = self.client_http.get(f'/parts/{twin.pk}/label/')

        self.assertEqual(response.context['title'], 'CL-TWIN')
        self.assertEqual(response.context['description'], '')

    def test_a_part_without_a_cell_says_so(self):
        loose = SparePart.objects.create(part_number='CL-LOOSE', name='Ничья деталь')

        response = self.client_http.get(f'/parts/{loose.pk}/label/')

        self.assertEqual(response.context['address'], 'нет ячейки')

    def test_a_set_of_one_type_is_named_by_the_type(self):
        """Перечислять одинаковые названия на этикетке незачем — важны номиналы."""
        for number, value in (('CL-2', 4.7), ('CL-3', 1)):
            self.cell.parts.add(SparePart.objects.create(
                part_number=number, name=f'Резистор {value}к',
                component_type='Резистор', package='0805',
                resistance=value, resistance_unit='кОм',
            ))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.context['title'], 'Набор резисторов')
        self.assertEqual(response.context['items'], ['10кОм', '4.7кОм', '1кОм'])
        # Корпус общий на всю ячейку — он у всех один
        self.assertEqual(response.context['package'], '0805')

    def test_what_a_set_has_in_common_is_printed_once(self):
        """«0.125Вт» у каждого номинала — это шум, а место на этикетке жёсткое."""
        self.part.power, self.part.power_unit = 0.125, 'Вт'
        self.part.save(update_fields=['power', 'power_unit'])
        for number, value in (('CS-1', 4.7), ('CS-2', 22)):
            self.cell.parts.add(SparePart.objects.create(
                part_number=number, name=f'Резистор {value}к',
                component_type='Резистор', package='0805',
                resistance=value, resistance_unit='кОм',
                power=0.125, power_unit='Вт',
            ))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.context['specs'], '0.125Вт')
        self.assertEqual(response.context['items'], ['10кОм', '4.7кОм', '22кОм'])

    def test_values_of_two_kinds_stay_in_one_cell_of_the_grid(self):
        """У конденсатора различаются два числа сразу; сплошной строкой
        они слились бы в список, где не понять, что к чему относится."""
        self.cell.parts.clear()
        for number, farad, volt in (('CD-1', 47, 35), ('CD-2', 1000, 50)):
            self.cell.parts.add(SparePart.objects.create(
                part_number=number, name=f'Конденсатор {number}',
                component_type='Конденсатор',
                capacitance=farad, capacitance_unit='мкФ',
                voltage=volt, voltage_unit='В',
            ))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.context['items'], ['35В, 47мкФ', '50В, 1000мкФ'])

    def test_a_set_falls_back_to_articles_when_the_values_repeat(self):
        """У россыпи резисторов заполнена только мощность, а номинал живёт
        в артикуле: «2Вт, 2Вт, 2Вт» не различает ничего."""
        self.cell.parts.clear()
        for number in ('RS-1', 'RS-2', 'RS-3'):
            self.cell.parts.add(SparePart.objects.create(
                part_number=number, name=f'Резисторы {number}',
                component_type='Резистор', power=2, power_unit='Вт',
            ))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.context['specs'], '2Вт')
        self.assertEqual(response.context['items'], ['RS-1', 'RS-2', 'RS-3'])

    def test_a_mixed_set_keeps_the_common_package_out(self):
        """«0805» на ячейке, где половина деталей в 1206, хуже, чем ничего."""
        self.cell.parts.add(SparePart.objects.create(
            part_number='CL-4', name='Резистор 1к', component_type='Резистор',
            package='1206', resistance=1, resistance_unit='кОм',
        ))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.context['package'], '')

    def test_different_types_are_listed_by_article(self):
        self.cell.parts.add(SparePart.objects.create(
            part_number='CL-D', name='Диод', component_type='Диод',
        ))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.context['title'], 'Разные детали')
        self.assertEqual(response.context['items'], ['CL-1', 'CL-D'])

    def test_a_list_of_parts_is_printed_as_a_grid(self):
        """Сплошная строка переносилась посреди номинала: «10кОм, 4.7к»
        и «Ом» на следующей строке."""
        self.cell.parts.add(SparePart.objects.create(
            part_number='CG-1', name='Диод', component_type='Диод'))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertContains(response, 'class="label-items"')
        self.assertContains(response, '<span class="label-item">CL-1</span>', html=False)
        self.assertContains(response, '<span class="label-item">CG-1</span>', html=False)

    def test_a_single_part_has_no_grid(self):
        """У одной детали перечислять нечего — печатается описание."""
        response = self.client_http.get(f'/parts/{self.part.pk}/label/')

        self.assertNotContains(response, 'class="label-items"')
        self.assertContains(response, 'label-description')

    def test_an_empty_cell_says_it_is_empty(self):
        empty = self.cabinet.cells.last()

        response = self.client_http.get(f'/storage-cells/{empty.pk}/label/')

        self.assertEqual(response.context['title'], '')
        self.assertEqual(response.context['description'], 'Ячейка пуста')
        self.assertEqual(response.context['address'], empty.address)

    def test_every_label_carries_the_same_qr_size(self):
        """Размер кода один на всех этикетках — сканер подносят одинаково."""
        pages = self._both_pages() + (
            f'/parts/labels/?ids={self.part.pk}',
            f'/storage-cells/labels/?cabinet={self.cabinet.number}',
        )
        for page in pages:
            with self.subTest(page=page):
                self.assertContains(self.client_http.get(page), '12.3mm')

    def test_the_cell_label_fits_its_font(self):
        """Длинное описание должно ужиматься, а не обрезаться на полуслове."""
        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertContains(response, 'label-fit.js')


class PluralGenitiveTests(TestCase):
    """«Набор резисторов» — форму выводим правилами: список типов
    компонентов ведёт человек, готового справочника в программе нет."""

    def test_common_component_types(self):
        cases = {
            'Резистор': 'Резисторов',
            'Диод': 'Диодов',
            'Транзистор': 'Транзисторов',
            'Конденсатор': 'Конденсаторов',
            'Предохранитель': 'Предохранителей',
            'Микросхема': 'Микросхем',
            'Батарея': 'Батарей',
            'Реле': 'Реле',
        }
        for word, expected in cases.items():
            with self.subTest(word=word):
                self.assertEqual(plural_genitive(word), expected)

    def test_an_empty_type_gives_an_empty_string(self):
        self.assertEqual(plural_genitive(''), '')
        self.assertEqual(plural_genitive(None), '')


class SpecFormattingTests(TestCase):
    """Характеристики хранятся с шестью знаками после точки, и Django
    дописывает нули при каждой записи. «0.150000 А» читается как точность
    до микроампера, которой ни у кого нет."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_spec', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)
        self.part = SparePart.objects.create(
            part_number='SP-1', name='Диод', component_type='Диод',
            current=Decimal('0.15'), current_unit='А',
            voltage=Decimal('100'), voltage_unit='В',
        )

    def test_format_spec_drops_the_trailing_zeros(self):
        cases = {
            Decimal('0.150000'): '0.15',
            Decimal('100.000000'): '100',
            Decimal('4700.000000'): '4700',
            Decimal('0.125000'): '0.125',
            Decimal('2.200000'): '2.2',
        }
        for value, expected in cases.items():
            with self.subTest(value=value):
                self.assertEqual(format_spec(value), expected)

    def test_format_spec_of_nothing_is_an_empty_string(self):
        self.assertEqual(format_spec(None), '')

    def test_the_part_card_shows_the_short_form(self):
        response = self.client_http.get(f'/parts/{self.part.pk}/')

        self.assertContains(response, '0.15')
        self.assertNotContains(response, '0.150000')

    def test_the_edit_form_shows_the_short_form(self):
        """Иначе при каждом сохранении в поле остаётся «0.150000»."""
        response = self.client_http.get(f'/parts/{self.part.pk}/edit/')

        self.assertContains(response, 'value="0.15"')
        self.assertNotContains(response, 'value="0.150000"')

    def test_the_label_shows_the_short_form(self):
        self.assertEqual(self.part.specs_display, '100В, 0.15А')


class ImportUnitTests(TestCase):
    """Единица измерения без значения — мусор: «Ом» у диода не значит ничего."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_import_unit', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def _import(self, rows, update_existing=False):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['part_number', 'name', 'component_type', 'voltage', 'voltage_unit',
                   'resistance', 'resistance_unit', 'capacitance', 'capacitance_unit']
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, '') for header in headers])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'parts.xlsx'
        data = {'file': buffer}
        if update_existing:
            data['update_existing'] = 'on'
        return self.client_http.post('/parts/import/', data, follow=True)

    def test_a_unit_without_a_value_is_not_imported(self):
        """В присланных файлах единицы проставлены во всех строках подряд,
        независимо от того, есть ли что измерять."""
        self._import([{
            'part_number': 'IU-1', 'name': 'Диод', 'component_type': 'Диод',
            'voltage': 100, 'voltage_unit': 'В',
            'resistance_unit': 'Ом', 'capacitance_unit': 'Ф',
        }])

        part = SparePart.objects.get(part_number='IU-1')
        self.assertEqual(part.voltage_unit, 'В')
        self.assertEqual(part.resistance_unit, '')
        self.assertEqual(part.capacitance_unit, '')

    def test_an_empty_unit_does_not_wipe_the_one_typed_by_hand(self):
        SparePart.objects.create(
            part_number='IU-2', name='Резистор', component_type='Резистор',
            resistance=Decimal('10'), resistance_unit='кОм',
        )

        self._import([{
            'part_number': 'IU-2', 'name': 'Резистор', 'component_type': 'Резистор',
            'resistance': 10,
        }], update_existing=True)

        self.assertEqual(SparePart.objects.get(part_number='IU-2').resistance_unit, 'кОм')


class PartBulkDeleteTests(TestCase):
    """Удаление отмеченных деталей списком: после загрузки каталога лишние
    позиции удаляют десятками."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_bulk_del', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)
        self.parts = [
            SparePart.objects.create(part_number=f'BD-{index}', name=f'Деталь {index}',
                                     component_type='Диод', current_stock=index)
            for index in range(3)
        ]

    def _ids(self, parts):
        return '&'.join(f'ids={part.pk}' for part in parts)

    def test_the_page_lists_what_will_be_deleted(self):
        response = self.client_http.get(f'/parts/delete-selected/?{self._ids(self.parts[:2])}')

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'BD-0')
        self.assertContains(response, 'BD-1')
        self.assertNotContains(response, 'BD-2')

    def test_nothing_is_deleted_by_opening_the_page(self):
        """GET только показывает — удаляет подтверждение."""
        self.client_http.get(f'/parts/delete-selected/?{self._ids(self.parts)}')

        self.assertEqual(SparePart.objects.count(), 3)

    def test_confirmation_deletes_the_selected_parts(self):
        response = self.client_http.post(
            '/parts/delete-selected/', {'ids': [self.parts[0].pk, self.parts[2].pk]}
        )

        self.assertRedirects(response, '/parts/')
        self.assertEqual(
            list(SparePart.objects.values_list('part_number', flat=True)), ['BD-1']
        )

    def test_the_page_warns_about_stock_and_orders(self):
        """Удаление уносит историю движений и записи о деталях в заказах."""
        order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Тест', inn='7700000123')
        )
        RepairOrderDetail.objects.create(
            repair_order=order, part=self.parts[1], quantity_used=2
        )

        response = self.client_http.get(f'/parts/delete-selected/?{self._ids(self.parts)}')

        self.assertEqual(response.context['in_orders'], [self.parts[1]])
        self.assertEqual(response.context['in_stock'], self.parts[1:])

    def test_an_empty_selection_deletes_nothing(self):
        response = self.client_http.get('/parts/delete-selected/')

        self.assertEqual(response.context['parts'], [])
        self.assertEqual(SparePart.objects.count(), 3)

    def test_the_list_has_the_button(self):
        response = self.client_http.get('/parts/')

        self.assertContains(response, 'formaction="/parts/delete-selected/"')

    def test_a_repair_manager_cannot_delete_parts(self):
        """Склад ведут кладовщики; удаление списком тем более не для всех."""
        manager = Employee.objects.create_user(
            username='manager_bulk', full_name='Мастер', password='pass', role='repair_manager'
        )
        client = TestClient()
        client.force_login(manager)

        response = client.post('/parts/delete-selected/', {'ids': [self.parts[0].pk]})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(SparePart.objects.count(), 3)


class RepairOrderLabelsBatchTests(TestCase):
    """Пачка этикеток заказов: печатается не на заказ, а на единицу
    оборудования внутри него — заказ с двумя единицами даёт две этикетки."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_orders_lbl', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        client_obj = ClientModel.objects.create(name='ООО Дельта')
        model = EquipmentModel.objects.create(name='БУАД-5')

        self.order1 = RepairOrder.objects.create(client=client_obj)
        RepairOrderEquipment.objects.create(
            repair_order=self.order1,
            equipment=Equipment.objects.create(model=model, serial_number='SN-101'),
        )
        RepairOrderEquipment.objects.create(
            repair_order=self.order1,
            equipment=Equipment.objects.create(model=model, serial_number='SN-102'),
        )

        self.order2 = RepairOrder.objects.create(client=client_obj)
        RepairOrderEquipment.objects.create(
            repair_order=self.order2,
            equipment=Equipment.objects.create(model=model, serial_number='SN-201'),
        )

    def _ids(self, orders):
        return '&'.join(f'ids={order.pk}' for order in orders)

    def test_checked_orders_get_one_label_per_unit(self):
        response = self.client_http.get(f'/repair-orders/labels/?{self._ids([self.order1, self.order2])}')

        self.assertEqual(len(response.context['labels']), 3)

    def test_no_duplicate_or_missing_labels(self):
        response = self.client_http.get(f'/repair-orders/labels/?{self._ids([self.order1])}')
        serials = sorted(label['roe'].equipment.serial_number for label in response.context['labels'])

        self.assertEqual(serials, ['SN-101', 'SN-102'])

    def test_no_selection_falls_back_to_the_current_filter(self):
        """Без отметок — печать по тому же отбору, что и в списке (та же
        логика, что уже используется у этикеток деталей)."""
        response = self.client_http.get(f'/repair-orders/labels/?q={self.order2.order_number}')

        self.assertEqual(len(response.context['labels']), 1)
        self.assertEqual(response.context['labels'][0]['order'], self.order2)

    def test_positions_are_counted_per_order(self):
        response = self.client_http.get(f'/repair-orders/labels/?{self._ids([self.order1])}')
        positions = {
            label['roe'].equipment.serial_number: label['position']
            for label in response.context['labels']
        }

        self.assertEqual(positions['SN-101'], 1)
        self.assertEqual(positions['SN-102'], 2)

    def test_query_count_does_not_grow_with_selected_orders(self):
        """Больше отмеченных заказов не должно означать пропорционально
        больше обращений к базе — иначе полсотни этикеток тянули бы
        полсотню отдельных запросов вместо select_related/prefetch_related."""
        model = self.order1.order_equipments.first().equipment.model
        extra_order = RepairOrder.objects.create(client=self.order1.client)
        RepairOrderEquipment.objects.create(
            repair_order=extra_order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-301'),
        )

        with CaptureQueriesContext(connection) as few:
            self.client_http.get(f'/repair-orders/labels/?{self._ids([self.order1])}')
        with CaptureQueriesContext(connection) as many:
            self.client_http.get(f'/repair-orders/labels/?{self._ids([self.order1, self.order2, extra_order])}')

        self.assertEqual(len(few.captured_queries), len(many.captured_queries))

    def test_the_list_has_the_buttons(self):
        response = self.client_http.get('/repair-orders/')

        self.assertContains(response, 'action="/repair-orders/labels/"')
        self.assertContains(response, 'formaction="/repair-orders/bulk-status/"')


@override_settings(NOTIFY_CLIENTS=True)
class RepairOrderBulkStatusTests(TestCase):
    """Массовая отгрузка отмеченных заказов — единственный переход,
    разрешённый массовой смене статуса: «Готов к отгрузке» → «Отгружен»."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_bulk_status', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        client_obj = ClientModel.objects.create(name='ООО Эпсилон', email='eps@example.com')
        model = EquipmentModel.objects.create(name='БУАД-7')

        self.ready1 = RepairOrder.objects.create(client=client_obj, status='ready_for_shipment')
        RepairOrderEquipment.objects.create(
            repair_order=self.ready1,
            equipment=Equipment.objects.create(model=model, serial_number='SN-501'),
        )
        self.ready2 = RepairOrder.objects.create(client=client_obj, status='ready_for_shipment')
        RepairOrderEquipment.objects.create(
            repair_order=self.ready2,
            equipment=Equipment.objects.create(model=model, serial_number='SN-502'),
        )
        self.in_repair = RepairOrder.objects.create(client=client_obj, status='repair')

    def _ids(self, orders):
        return [order.pk for order in orders]

    def test_confirmation_page_changes_nothing(self):
        response = self.client_http.get(
            '/repair-orders/bulk-status/', {'ids': self._ids([self.ready1, self.in_repair])}
        )

        self.assertEqual(response.status_code, 200)
        self.ready1.refresh_from_db()
        self.in_repair.refresh_from_db()
        self.assertEqual(self.ready1.status, 'ready_for_shipment')
        self.assertEqual(self.in_repair.status, 'repair')
        self.assertEqual(response.context['eligible'], [self.ready1])
        self.assertEqual([row['order'] for row in response.context['ineligible']], [self.in_repair])

    def test_opening_the_confirmation_page_does_not_ship_anything(self):
        """Действие не должно выполняться без прохождения шага подтверждения."""
        self.client_http.get(
            '/repair-orders/bulk-status/', {'ids': self._ids([self.ready1, self.ready2])}
        )

        self.assertFalse(RepairOrder.objects.filter(status='shipped').exists())

    def test_post_ships_the_eligible_orders(self):
        response = self.client_http.post(
            '/repair-orders/bulk-status/', {'ids': self._ids([self.ready1, self.ready2])}
        )

        self.ready1.refresh_from_db()
        self.ready2.refresh_from_db()
        self.assertEqual(self.ready1.status, 'shipped')
        self.assertEqual(self.ready2.status, 'shipped')
        self.assertIsNotNone(self.ready1.shipping_date)
        self.assertIsNotNone(self.ready1.date_completed)
        self.assertEqual(response.context['applied'], [self.ready1, self.ready2])

    def test_each_shipped_order_gets_its_own_history_entry(self):
        self.client_http.post(
            '/repair-orders/bulk-status/', {'ids': self._ids([self.ready1, self.ready2])}
        )

        self.assertTrue(self.ready1.status_history.filter(status='shipped').exists())
        self.assertTrue(self.ready2.status_history.filter(status='shipped').exists())

    def test_each_shipped_order_is_notified(self):
        self.client_http.post(
            '/repair-orders/bulk-status/', {'ids': self._ids([self.ready1, self.ready2])}
        )

        self.assertEqual(
            Notification.objects.filter(
                event='order_status', repair_order__in=[self.ready1, self.ready2]
            ).count(),
            2,
        )

    def test_partial_failure_skips_ineligible_orders_with_a_reason(self):
        response = self.client_http.post(
            '/repair-orders/bulk-status/', {'ids': self._ids([self.ready1, self.in_repair])}
        )

        self.ready1.refresh_from_db()
        self.in_repair.refresh_from_db()
        self.assertEqual(self.ready1.status, 'shipped')
        self.assertEqual(self.in_repair.status, 'repair')
        self.assertEqual(response.context['applied'], [self.ready1])
        skipped = response.context['skipped']
        self.assertEqual(len(skipped), 1)
        self.assertEqual(skipped[0]['order'], self.in_repair)
        self.assertIn('Ремонт', skipped[0]['reason'])

    def test_an_empty_selection_ships_nothing(self):
        response = self.client_http.post('/repair-orders/bulk-status/', {})

        self.assertEqual(response.status_code, 302)
        self.assertFalse(RepairOrder.objects.filter(status='shipped').exists())

    def test_the_confirmation_page_lists_who_will_be_skipped_and_why(self):
        response = self.client_http.get(
            '/repair-orders/bulk-status/', {'ids': self._ids([self.ready1, self.in_repair])}
        )

        self.assertContains(response, self.in_repair.order_number)
        self.assertContains(response, 'Ремонт')


class ApplicationFieldTests(TestCase):
    """Применимость — отдельное поле, а не строка в описании: её печатают
    на этикетке своим местом и по ней ищут."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_appl', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)
        self.cabinet = Cabinet.objects.create(number=9)
        self.cabinet.apply_layout([2])
        self.cell = self.cabinet.cells.first()
        self.part = SparePart.objects.create(
            part_number='AP-1', name='Оптопара', component_type='Оптрон',
            package='DIP-8', application='Otis',
        )
        self.cell.parts.add(self.part)

    def test_the_label_prints_it_between_the_package_and_the_address(self):
        for page in (f'/parts/{self.part.pk}/label/', f'/storage-cells/{self.cell.pk}/label/'):
            with self.subTest(page=page):
                response = self.client_http.get(page)
                self.assertEqual(response.context['application'], 'Otis')
                content = response.content.decode()
                self.assertLess(content.index('label-package"'), content.index('label-application"'))
                self.assertLess(content.index('label-application"'), content.index('label-address"'))

    def test_a_cell_shows_it_only_when_every_part_agrees(self):
        """«Otis» на ячейке, где половина деталей от ABB, вводит в заблуждение."""
        self.cell.parts.add(SparePart.objects.create(
            part_number='AP-2', name='Драйвер', component_type='Оптрон', application='ABB',
        ))

        response = self.client_http.get(f'/storage-cells/{self.cell.pk}/label/')

        self.assertEqual(response.context['application'], '')

    def test_the_card_and_the_form_show_it(self):
        card = self.client_http.get(f'/parts/{self.part.pk}/')
        form = self.client_http.get(f'/parts/{self.part.pk}/edit/')

        self.assertContains(card, 'Otis')
        self.assertContains(form, 'name="application"')
        self.assertContains(form, 'value="Otis"')

    def test_import_and_export_carry_the_field(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['part_number', 'name', 'component_type', 'application'])
        ws.append(['AP-3', 'Реле', 'Реле', 'Altivar'])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer.name = 'parts.xlsx'
        self.client_http.post('/parts/import/', {'file': buffer}, follow=True)

        self.assertEqual(SparePart.objects.get(part_number='AP-3').application, 'Altivar')

        export = self.client_http.get('/parts/export/')
        sheet = openpyxl.load_workbook(io.BytesIO(export.content)).active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn('application', headers)
        values = [row[headers.index('application')] for row in sheet.iter_rows(min_row=2, values_only=True)]
        self.assertIn('Altivar', values)

    def test_the_migration_splits_it_out_of_the_description(self):
        """Так её записывали каталоги, из которых детали загружали:
        «Оптопара DIP-8 | Применение: Otis»."""
        from importlib import import_module

        from django.apps import apps as installed_apps

        migration = import_module('core.migrations.0023_spare_part_application')
        legacy = SparePart.objects.create(
            part_number='AP-OLD', name='Оптопара', component_type='Оптрон',
            description='Оптопара с логическим выходом | Применение: ABB',
        )

        migration.split_application(installed_apps, None)

        legacy.refresh_from_db()
        self.assertEqual(legacy.application, 'ABB')
        self.assertEqual(legacy.description, 'Оптопара с логическим выходом')


class TestRunnerTests(SimpleTestCase):
    """Быстрый хешер паролей — только для тестов и нигде больше."""

    def test_tests_run_with_the_fast_hasher(self):
        from core.test_runner import FAST_HASHER

        self.assertEqual(settings.PASSWORD_HASHERS, [FAST_HASHER])

    def test_the_project_settings_do_not_weaken_hashing(self):
        """Если ускорение однажды перенесут в настройки, пароли сотрудников
        станут храниться этим хешером по-настоящему."""
        from lifteam import settings as project_settings

        self.assertFalse(hasattr(project_settings, 'PASSWORD_HASHERS'))


class OrderEditLabelButtonTests(TestCase):
    """Этикетку печатают сразу после правки серийника или пломб —
    возвращаться для этого в карточку заказа незачем."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_edit_label', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        model = EquipmentModel.objects.create(name='БУАД-3')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Лифт', inn='7700000456')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-EDIT-1'),
        )

    def test_a_saved_unit_offers_the_label(self):
        """Кнопка этикетки стоит в строке карточки заказа и на странице
        единицы; страницы правки заказа больше нет."""
        label = f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/label/'

        card = self.client_http.get(f'/repair-orders/{self.order.pk}/')
        unit = self.client_http.get(
            f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/'
        )

        self.assertContains(card, label)
        self.assertContains(unit, label)

    def test_a_new_order_has_nothing_to_print_yet(self):
        """У несохранённой единицы нет ни номера, ни ссылки для кода."""
        response = self.client_http.get('/repair-orders/create/')

        self.assertNotContains(response, '/label/')

    def test_the_label_itself_still_opens(self):
        response = self.client_http.get(
            f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/label/'
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'SN-EDIT-1')


class NetworkErrorHintTests(TestCase):
    """«CERTIFICATE_VERIFY_FAILED» само по себе не говорит, что чинить."""

    def test_a_certificate_failure_names_the_missing_root(self):
        from core.net import explain

        reason = ('[SSL: CERTIFICATE_VERIFY_FAILED] certificate verify failed: '
                  'unable to get local issuer certificate (_ssl.c:1006)')

        result = explain(reason)

        self.assertIn('CERTIFICATE_VERIFY_FAILED', result)
        self.assertIn('НУЦ Минцифры', result)
        self.assertIn('DEPLOY.md', result)

    def test_other_reasons_are_left_alone(self):
        from core.net import explain

        self.assertEqual(explain('[Errno -3] Temporary failure in name resolution'),
                         '[Errno -3] Temporary failure in name resolution')

    @override_settings(MAX_BOT_TOKEN='secret-token')
    def test_the_messenger_error_carries_the_hint(self):
        """Текст попадает в очередь оповещений — там его и читают."""
        from urllib import error

        broken = error.URLError('[SSL: CERTIFICATE_VERIFY_FAILED] certificate verify failed')
        with patch('core.messengers.request.urlopen', side_effect=broken):
            with self.assertRaises(messengers.MaxError) as caught:
                messengers.send_max_message('user:1', 'текст')

        self.assertIn('НУЦ Минцифры', str(caught.exception))

    @override_settings(TBANK_TOKEN='secret-token')
    def test_the_bank_error_carries_the_hint(self):
        from urllib import error

        from core import tbank

        broken = error.URLError('[SSL: CERTIFICATE_VERIFY_FAILED] certificate verify failed')
        with patch('core.tbank.request.urlopen', side_effect=broken):
            with self.assertRaises(tbank.TBankError) as caught:
                tbank.get_accounts()

        self.assertIn('НУЦ Минцифры', str(caught.exception))


class DryRunOutputTests(TestCase):
    """`--dry-run` ничего не отправляет, и сказать об этом нужно словами:
    один раз это стоило вечера поисков причины, по которой «письма не уходят»."""

    def setUp(self):
        self.note = Notification.objects.create(
            event='low_stock', recipient='wh@example.com',
            subject='Дефицит', body='Текст',
        )

    @override_settings(NOTIFICATIONS_ENABLED=True,
                       EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_the_dry_run_says_that_nothing_was_sent(self):
        from django.core import mail

        out = io.StringIO()
        call_command('send_notifications', dry_run=True, stdout=out, stderr=io.StringIO())
        output = out.getvalue()

        self.assertIn('НИЧЕГО НЕ ОТПРАВЛЕНО', output)
        self.assertIn('без --dry-run', output)
        self.assertEqual(len(mail.outbox), 0)
        self.note.refresh_from_db()
        self.assertEqual(self.note.status, 'pending')


class FaultTemplateApplyTests(TestCase):
    """Типовая неисправность предлагает рецепт деталей — заказ дополняется
    им, а не переписывается заново."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_faults', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.model_a = EquipmentModel.objects.create(name='БУАД-3')
        self.model_b = EquipmentModel.objects.create(name='ШУНЛ-2')

        self.part1 = SparePart.objects.create(part_number='FP-1', name='Транзистор', current_stock=10)
        self.part2 = SparePart.objects.create(part_number='FP-2', name='Резистор', current_stock=5)
        self.part3 = SparePart.objects.create(part_number='FP-3', name='Конденсатор', current_stock=2)

        self.fault1 = FaultType.objects.create(equipment_model=self.model_a, name='Не запускается')
        FaultTypePart.objects.create(fault_type=self.fault1, part=self.part1, quantity=2)
        FaultTypePart.objects.create(fault_type=self.fault1, part=self.part2, quantity=1)

        # part2 сознательно повторяется в обоих рецептах — проверка слияния
        self.fault2 = FaultType.objects.create(equipment_model=self.model_a, name='Шумит')
        FaultTypePart.objects.create(fault_type=self.fault2, part=self.part2, quantity=3)
        FaultTypePart.objects.create(fault_type=self.fault2, part=self.part3, quantity=1)

        self.fault_other_model = FaultType.objects.create(equipment_model=self.model_b, name='Неисправность другой модели')
        FaultTypePart.objects.create(fault_type=self.fault_other_model, part=self.part1, quantity=1)

        self.equipment = Equipment.objects.create(model=self.model_a, serial_number='SN-FAULT-1')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Лифт', inn='7700000789')
        )
        self.roe = RepairOrderEquipment.objects.create(repair_order=self.order, equipment=self.equipment)

    def _apply(self, fault_ids):
        return self.client_http.post(
            f'/repair-orders/{self.order.pk}/apply-fault-template/',
            {'fault_ids': fault_ids},
        )

    def test_applying_one_fault_creates_its_recipe(self):
        response = self._apply([self.fault1.pk])
        data = response.json()

        self.assertTrue(data['success'])
        details = {d.part_id: d.quantity_used for d in self.order.details.all()}
        self.assertEqual(details, {self.part1.pk: 2, self.part2.pk: 1})

        self.part1.refresh_from_db()
        self.part2.refresh_from_db()
        self.assertEqual(self.part1.current_stock, 8)
        self.assertEqual(self.part2.current_stock, 4)

    def test_applying_adds_to_an_existing_manual_line_without_replacing_it(self):
        """Ручная строка по part1 уже есть — применение шаблона добавляет
        свою собственную, а не трогает и не сливается с существующей."""
        RepairOrderDetail.objects.create(repair_order=self.order, part=self.part1, quantity_used=5)

        self._apply([self.fault1.pk])

        details = list(self.order.details.all())
        self.assertEqual(len(details), 3)
        part1_lines = sorted(d.quantity_used for d in details if d.part_id == self.part1.pk)
        self.assertEqual(part1_lines, [2, 5])

    def test_several_faults_merge_the_same_part_into_one_line(self):
        """part2 встречается в обоих рецептах (1 и 3) — в заказе должна
        появиться одна строка на 4, а не две частичные."""
        response = self._apply([self.fault1.pk, self.fault2.pk])
        data = response.json()

        self.assertTrue(data['success'])
        details = list(self.order.details.all())
        part2_lines = [d for d in details if d.part_id == self.part2.pk]
        self.assertEqual(len(part2_lines), 1)
        self.assertEqual(part2_lines[0].quantity_used, 4)
        self.assertEqual(
            {d.part_id: d.quantity_used for d in details},
            {self.part1.pk: 2, self.part2.pk: 4, self.part3.pk: 1}
        )

    def test_a_shortage_is_reported_but_the_detail_is_still_added(self):
        fault3 = FaultType.objects.create(equipment_model=self.model_a, name='Сильный дефицит')
        FaultTypePart.objects.create(fault_type=fault3, part=self.part3, quantity=10)  # остаток всего 2

        response = self._apply([fault3.pk])
        data = response.json()

        self.assertTrue(data['success'])
        self.assertIn('не хватило', data['message'])
        self.part3.refresh_from_db()
        self.assertEqual(self.part3.current_stock, -8)
        self.assertEqual(self.order.details.get(part=self.part3).quantity_used, 10)

    def test_applying_nothing_fails_without_touching_the_order(self):
        response = self._apply([])
        data = response.json()

        self.assertFalse(data['success'])
        self.assertEqual(self.order.details.count(), 0)

    def test_the_unit_page_lists_only_this_equipments_model_faults(self):
        roe = self.order.order_equipments.get(equipment=self.equipment)
        response = self.client_http.get(
            f'/repair-orders/{self.order.pk}/equipment/{roe.pk}/'
        )

        self.assertContains(response, self.fault1.name)
        self.assertContains(response, self.fault2.name)
        self.assertNotContains(response, self.fault_other_model.name)

    def test_the_faults_ajax_endpoint_is_scoped_to_the_equipment_model(self):
        response = self.client_http.get(f'/ajax/equipment/{self.equipment.pk}/faults/')
        data = response.json()

        names = {f['name'] for f in data['faults']}
        self.assertEqual(names, {self.fault1.name, self.fault2.name})

    def test_the_equipment_form_rejects_a_fault_from_another_model(self):
        # order= обязателен: без него единицы этого же заказа считаются
        # занятыми в другом и в список выбора не попадают
        form = RepairOrderEquipmentForm(order=self.order, data={
            'equipment': self.equipment.pk,
            'fault_description': '',
            'faults': [self.fault_other_model.pk],
            'work_performed': '', 'seal_numbers': '', 'initial_condition': '',
            'repair_cost': '', 'yandex_disk_folder': '',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('faults', form.errors)

    def test_the_other_option_is_not_a_database_choice(self):
        """«Другое» в списке — сигнал для интерфейса, а не id неисправности:
        его выбор не должен ронять валидацию формы."""
        form = RepairOrderEquipmentForm(order=self.order, data={
            'equipment': self.equipment.pk,
            'fault_description': 'своими словами, чего в справочнике ещё нет',
            'faults': ['other'],
            'work_performed': '', 'seal_numbers': '', 'initial_condition': '',
            'repair_cost': '', 'yandex_disk_folder': '',
        })

        self.assertTrue(form.is_valid())
        self.assertEqual(list(form.cleaned_data['faults']), [])


class FaultTypeAdminTests(TestCase):
    """Справочник типовых неисправностей — права те же, что у EquipmentModel:
    создание и правка открыты любому авторизованному, удаление — только
    складу и мастеру."""

    def setUp(self):
        self.model = EquipmentModel.objects.create(name='БУАД-9')
        self.part = SparePart.objects.create(part_number='FTP-1', name='Диод')
        self.fault = FaultType.objects.create(equipment_model=self.model, name='Проверочная неисправность')

    @staticmethod
    def _formset_data(prefix='parts', total=0):
        return {
            f'{prefix}-TOTAL_FORMS': str(total),
            f'{prefix}-INITIAL_FORMS': '0',
            f'{prefix}-MIN_NUM_FORMS': '0',
            f'{prefix}-MAX_NUM_FORMS': '1000',
        }

    def test_a_repair_manager_can_create_a_fault_type(self):
        manager = Employee.objects.create_user(
            username='manager_faults', full_name='Мастер', password='pass', role='repair_manager'
        )
        client = TestClient()
        client.force_login(manager)

        data = {'equipment_model': self.model.pk, 'name': 'Новая неисправность', 'description': ''}
        data.update(self._formset_data())
        response = client.post('/faults/create/', data)

        self.assertEqual(response.status_code, 302)
        self.assertTrue(FaultType.objects.filter(name='Новая неисправность').exists())

    def test_creating_a_fault_type_together_with_its_recipe(self):
        manager = Employee.objects.create_user(
            username='manager_recipe', full_name='Мастер', password='pass', role='repair_manager'
        )
        client = TestClient()
        client.force_login(manager)

        data = {'equipment_model': self.model.pk, 'name': 'С рецептом', 'description': ''}
        data.update(self._formset_data(total=1))
        data['parts-0-part'] = self.part.pk
        data['parts-0-quantity'] = '3'
        response = client.post('/faults/create/', data)

        self.assertEqual(response.status_code, 302)
        fault = FaultType.objects.get(name='С рецептом')
        self.assertEqual(fault.parts.count(), 1)
        self.assertEqual(fault.parts.first().quantity, 3)

    def test_an_accountant_cannot_delete_a_fault_type(self):
        accountant = Employee.objects.create_user(
            username='accountant_faults', full_name='Бухгалтер', password='pass', role='accountant'
        )
        client = TestClient()
        client.force_login(accountant)

        response = client.post(f'/faults/{self.fault.pk}/delete/')

        self.assertEqual(response.status_code, 302)
        self.assertTrue(FaultType.objects.filter(pk=self.fault.pk).exists())

    def test_warehouse_can_delete_a_fault_type(self):
        warehouse = Employee.objects.create_user(
            username='warehouse_faults', full_name='Кладовщик', password='pass', role='warehouse'
        )
        client = TestClient()
        client.force_login(warehouse)

        response = client.post(f'/faults/{self.fault.pk}/delete/')

        self.assertRedirects(response, '/faults/')
        self.assertFalse(FaultType.objects.filter(pk=self.fault.pk).exists())

    def test_an_anonymous_user_is_sent_to_login(self):
        client = TestClient()

        response = client.get('/faults/create/')

        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)


class StockAllocationTests(TestCase):
    """Распределение расхода по партиям прихода (FIFO): каждый расход должен
    ссылаться на конкретную партию, из которой он физически списан."""

    def setUp(self):
        self.part = SparePart.objects.create(part_number='BATCH-1', name='Диод')

    def _incoming(self, quantity, price):
        return StockMovement.objects.create(
            part=self.part, quantity=quantity, movement_type='incoming', unit_price=price
        )

    def _outgoing(self, quantity):
        return StockMovement.objects.create(
            part=self.part, quantity=quantity, movement_type='outgoing'
        )

    def test_a_single_batch_covers_the_request(self):
        batch = self._incoming(5, Decimal('10'))
        outgoing = self._outgoing(3)

        fully_covered = StockAllocation.allocate(outgoing)

        self.assertTrue(fully_covered)
        allocations = list(outgoing.allocations.all())
        self.assertEqual(len(allocations), 1)
        self.assertEqual(allocations[0].incoming, batch)
        self.assertEqual(allocations[0].quantity, 3)
        self.assertEqual(batch.remaining_in_batch, 2)

    def test_a_request_bigger_than_the_oldest_batch_spills_into_the_next(self):
        """Запрошенное количество больше самой старой партии — распределяется
        по датам поступления, старейшая партия расходуется первой."""
        old_batch = self._incoming(5, Decimal('10'))
        new_batch = self._incoming(5, Decimal('20'))
        outgoing = self._outgoing(8)

        fully_covered = StockAllocation.allocate(outgoing)

        self.assertTrue(fully_covered)
        by_batch = {a.incoming_id: a.quantity for a in outgoing.allocations.all()}
        self.assertEqual(by_batch, {old_batch.pk: 5, new_batch.pk: 3})
        self.assertEqual(old_batch.remaining_in_batch, 0)
        self.assertEqual(new_batch.remaining_in_batch, 2)

    def test_insufficient_stock_across_all_batches_leaves_a_remainder(self):
        """Партий вместе не хватает — известная часть распределяется как
        обычно, а на недостачу партии не хватает ни одной."""
        batch = self._incoming(5, Decimal('10'))
        outgoing = self._outgoing(8)

        fully_covered = StockAllocation.allocate(outgoing)

        self.assertFalse(fully_covered)
        allocations = list(outgoing.allocations.all())
        self.assertEqual(len(allocations), 1)
        self.assertEqual(allocations[0].incoming, batch)
        self.assertEqual(allocations[0].quantity, 5)

    def test_a_later_call_in_the_same_transaction_does_not_see_the_batch_as_still_full(self):
        """Два расхода подряд (как при нескольких вызовах
        `_use_repair_order_part` в одной транзакции) не должны оба списаться
        с полного остатка партии — второй обязан увидеть то, что уже забрал
        первый."""
        batch = self._incoming(5, Decimal('10'))

        with transaction.atomic():
            first = self._outgoing(3)
            StockAllocation.allocate(first)
            second = self._outgoing(2)
            StockAllocation.allocate(second)

        self.assertEqual(batch.remaining_in_batch, 0)
        self.assertEqual(StockAllocation.objects.filter(incoming=batch).count(), 2)
        self.assertEqual(
            sorted(StockAllocation.objects.filter(incoming=batch).values_list('quantity', flat=True)),
            [2, 3],
        )


class OrderCostFromAllocationsTests(TestCase):
    """`_use_repair_order_part` заводит `OrderCost` (category='parts') по
    фактически задействованным партиям — не по средней/текущей цене."""

    def setUp(self):
        self.user = Employee.objects.create_user(
            username='wh_cost', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.user)
        self.client_obj = ClientModel.objects.create(name='ООО Себестоимость')
        self.order = RepairOrder.objects.create(client=self.client_obj)
        self.part = SparePart.objects.create(part_number='COST-1', name='Стабилизатор')

    def _incoming(self, quantity, price):
        return StockMovement.objects.create(
            part=self.part, quantity=quantity, movement_type='incoming', unit_price=price
        )

    def _add_detail(self, quantity):
        return self.client_http.post(
            f'/repair-orders/{self.order.pk}/add-detail/',
            {'part': self.part.pk, 'quantity_used': quantity},
        )

    def test_cost_from_a_single_priced_batch(self):
        self._incoming(10, Decimal('50'))

        self._add_detail(4)

        cost = OrderCost.objects.get(repair_order=self.order, category='parts')
        self.assertEqual(cost.amount, Decimal('200'))

    def test_cost_split_across_two_differently_priced_batches(self):
        self._incoming(3, Decimal('50'))
        self._incoming(10, Decimal('80'))

        self._add_detail(5)

        cost = OrderCost.objects.get(repair_order=self.order, category='parts')
        # 3 шт. по 50 из первой партии + 2 шт. по 80 из второй
        self.assertEqual(cost.amount, Decimal('150') + Decimal('160'))

    def test_cost_is_unknown_when_a_batch_has_no_price(self):
        self._incoming(10, None)

        self._add_detail(4)

        cost = OrderCost.objects.get(repair_order=self.order, category='parts')
        self.assertIsNone(cost.amount)

    def test_cost_is_unknown_not_partial_when_stock_runs_short(self):
        """Недостаточно остатка — известная часть распределяется, но сумма
        затраты — None целиком, а не по покрытой части."""
        self._incoming(3, Decimal('50'))

        self._add_detail(5)  # спишется в минус, предупреждение уже проверено в других тестах

        cost = OrderCost.objects.get(repair_order=self.order, category='parts')
        self.assertIsNone(cost.amount)
        # Известная часть партий всё равно распределена
        self.assertEqual(StockAllocation.objects.filter(outgoing__repair_order=self.order).count(), 1)

    def test_repeated_calls_in_one_transaction_do_not_double_dip_a_batch(self):
        """Применение шаблона неисправности несколькими деталями подряд
        вызывает `_use_repair_order_part` несколько раз в одной транзакции —
        распределение по партиям должно учитывать уже списанное предыдущим
        вызовом, а не увидеть партию как ещё полную."""
        self._incoming(5, Decimal('10'))

        with transaction.atomic():
            views._use_repair_order_part(self.order, self.part, 3, self.user, 'первый вызов')
            views._use_repair_order_part(self.order, self.part, 2, self.user, 'второй вызов')

        costs = list(OrderCost.objects.filter(repair_order=self.order, category='parts'))
        self.assertEqual(len(costs), 2)
        self.assertEqual({c.amount for c in costs}, {Decimal('30'), Decimal('20')})

    def test_manual_stock_outgoing_allocates_but_creates_no_order_cost(self):
        """Ручное списание не привязано к заказу — партии распределяются
        (для будущей себестоимости), но `OrderCost` не заводится."""
        self._incoming(10, Decimal('50'))

        self.client_http.post(
            f'/parts/{self.part.pk}/stock-outgoing/',
            {'quantity': 4, 'reason': 'consumption', 'notes': '', 'document_number': ''},
        )

        movement = StockMovement.objects.get(part=self.part, movement_type='outgoing')
        self.assertEqual(movement.allocations.count(), 1)
        self.assertEqual(movement.allocations.first().quantity, 4)
        self.assertEqual(OrderCost.objects.count(), 0)


class RepairOrderPartsCostAndProfitTests(TestCase):
    """`RepairOrder.parts_cost` / `.profit` — прибыль = поступившие платежи
    минус себестоимость списанных деталей (не выставленная стоимость
    ремонта)."""

    def setUp(self):
        self.client_obj = ClientModel.objects.create(name='ООО Прибыль')
        self.order = RepairOrder.objects.create(client=self.client_obj)

    def test_an_order_with_no_cost_records_has_zero_parts_cost(self):
        self.assertEqual(self.order.parts_cost, Decimal('0'))
        self.assertEqual(self.order.profit, Decimal('0'))

    def test_parts_cost_sums_known_amounts(self):
        OrderCost.objects.create(repair_order=self.order, category='parts', amount=Decimal('100'))
        OrderCost.objects.create(repair_order=self.order, category='parts', amount=Decimal('50'))

        self.assertEqual(self.order.parts_cost, Decimal('150'))

    def test_a_single_unknown_cost_record_makes_the_whole_parts_cost_unknown(self):
        OrderCost.objects.create(repair_order=self.order, category='parts', amount=Decimal('100'))
        OrderCost.objects.create(repair_order=self.order, category='parts', amount=None)

        self.assertIsNone(self.order.parts_cost)
        self.assertIsNone(self.order.profit)

    def test_profit_is_payments_minus_parts_cost_not_the_quoted_repair_price(self):
        model = EquipmentModel.objects.create(name='БУАД-профит')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-PROFIT'),
            repair_cost=Decimal('10000'),
        )
        Payment.objects.create(repair_order=self.order, amount=Decimal('4000'))
        Payment.objects.create(repair_order=self.order, amount=Decimal('1000'))
        OrderCost.objects.create(repair_order=self.order, category='parts', amount=Decimal('1200'))

        # Оплачено всего 5000 из выставленных 10000 — прибыль считается
        # именно от поступивших денег, а не от total_repair_cost
        self.assertEqual(self.order.paid_amount, Decimal('5000'))
        self.assertEqual(self.order.profit, Decimal('3800'))


class ProfitReportTests(TestCase):
    """Отчёт «Прибыль по заказам»: агрегаты за период и разбивка по
    заказчику, доступ только бухгалтерии."""

    def setUp(self):
        self.accountant = Employee.objects.create_user(
            username='buh_profit', full_name='Бухгалтер', password='pass', role='accountant'
        )
        self.warehouse = Employee.objects.create_user(
            username='wh_profit', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.accountant)

        self.today = datetime.date.today()

        self.client_a = ClientModel.objects.create(name='ООО Альфа')
        self.client_b = ClientModel.objects.create(name='ООО Бета')
        self.order_a = RepairOrder.objects.create(client=self.client_a)
        self.order_b = RepairOrder.objects.create(client=self.client_b)

        Payment.objects.create(repair_order=self.order_a, amount=Decimal('5000'), payment_date=self.today)
        Payment.objects.create(repair_order=self.order_b, amount=Decimal('3000'), payment_date=self.today)
        OrderCost.objects.create(repair_order=self.order_a, category='parts', amount=Decimal('1200'))
        OrderCost.objects.create(repair_order=self.order_b, category='parts', amount=None)

    def test_role_without_accounting_access_is_denied(self):
        client = TestClient()
        client.force_login(self.warehouse)

        response = client.get('/reports/profit/', follow=True)

        self.assertRedirects(response, '/')

    def test_an_anonymous_user_is_sent_to_login(self):
        response = TestClient().get('/reports/profit/')

        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_totals_and_the_unknown_cost_count(self):
        response = self.client_http.get('/reports/profit/')

        self.assertEqual(response.context['revenue_total'], Decimal('8000'))
        self.assertEqual(response.context['known_cost_total'], Decimal('1200'))
        self.assertEqual(response.context['unknown_cost_count'], 1)
        self.assertEqual(response.context['profit_total'], Decimal('6800'))

    def test_breakdown_by_client(self):
        response = self.client_http.get('/reports/profit/')

        rows = {row['client']: row for row in response.context['rows']}
        self.assertEqual(rows['ООО Альфа']['revenue'], Decimal('5000'))
        self.assertEqual(rows['ООО Альфа']['profit'], Decimal('3800'))
        self.assertEqual(rows['ООО Бета']['revenue'], Decimal('3000'))
        self.assertEqual(rows['ООО Бета']['unknown_cost_count'], 1)
        # Известная часть себестоимости у Беты — 0, платёж известен полностью
        self.assertEqual(rows['ООО Бета']['profit'], Decimal('3000'))

    def test_a_period_outside_the_range_is_excluded(self):
        old_date = self.today - datetime.timedelta(days=90)
        Payment.objects.create(repair_order=self.order_a, amount=Decimal('9999'), payment_date=old_date)

        response = self.client_http.get('/reports/profit/')

        # 90 дней назад — за пределами окна по умолчанию (30 дней)
        self.assertEqual(response.context['revenue_total'], Decimal('8000'))

    def test_export_is_accountant_only(self):
        client = TestClient()
        client.force_login(self.warehouse)

        response = client.get('/reports/profit/export/', follow=True)

        self.assertRedirects(response, '/')

    def test_export_produces_a_workbook_with_two_sheets(self):
        response = self.client_http.get('/reports/profit/export/')

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(wb.sheetnames, ['Итог за период', 'По заказчикам'])


class InventorySessionTests(TestCase):
    """Инвентаризация кассетницы: старт по одной кассетнице (в проекте нет
    понятия «весь склад» или «зона»), ввод посчитанного, применение
    расхождений. Недостача — обычное списание, распределённое по партиям
    FIFO (`StockAllocation.allocate`), как и любой другой расход в проекте
    — ручного выбора партии под недостачу сознательно нет. Избыток —
    приход без цены (находка — не покупка) с обязательным комментарием."""

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='wh_inv', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.employee)

        self.cabinet = Cabinet.objects.create(number=901, name='Инвентаризация')
        self.cabinet.apply_layout([3])
        self.cell_a, self.cell_b, self.cell_c = list(self.cabinet.cells.order_by('cell_row'))

        # current_stock=10/5 соответствует одной партии прихода на каждую —
        # так expected_quantity на старте сессии совпадает с тем, что реально
        # можно списать/распределить по партиям
        self.part_a = SparePart.objects.create(
            part_number='INV-A', name='Деталь A', current_stock=10, min_stock=0)
        self.part_b = SparePart.objects.create(
            part_number='INV-B', name='Деталь B', current_stock=5, min_stock=0)
        self.cell_a.parts.add(self.part_a)
        self.cell_b.parts.add(self.part_b)
        # cell_c намеренно остаётся пустой — деталей в ней нет, строки инвентаризации быть не должно

        StockMovement.objects.create(
            part=self.part_a, quantity=10, movement_type='incoming', unit_price=Decimal('3.00'))
        StockMovement.objects.create(
            part=self.part_b, quantity=5, movement_type='incoming', unit_price=Decimal('7.00'))

    def _start(self, follow=True):
        response = self.client_http.post(f'/inventory/start/{self.cabinet.pk}/', follow=follow)
        session = InventorySession.objects.filter(cabinet=self.cabinet, status='in_progress').first()
        return session, response

    def _count(self, session, counted):
        data = {f'counted_{line.pk}': str(qty) for line, qty in counted}
        return self.client_http.post(f'/inventory/{session.pk}/count/', data, follow=True)

    def _confirm(self, session, comments=None):
        data = {}
        if comments:
            for line, text in comments.items():
                data[f'comment_{line.pk}'] = text
        return self.client_http.post(f'/inventory/{session.pk}/confirm/', data, follow=True)

    # --- старт сессии ---

    def test_anonymous_user_is_redirected_to_login(self):
        response = TestClient().get('/inventory/')

        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_start_creates_lines_only_for_parts_placed_in_a_cell(self):
        session, response = self._start()

        self.assertRedirects(response, f'/inventory/{session.pk}/count/')
        parts_in_lines = set(session.lines.values_list('part__part_number', flat=True))
        self.assertEqual(parts_in_lines, {'INV-A', 'INV-B'})

    def test_expected_quantity_is_a_snapshot_of_current_stock_at_start(self):
        session, _ = self._start()

        line_a = session.lines.get(part=self.part_a)
        self.assertEqual(line_a.expected_quantity, 10)

    def test_starting_again_redirects_to_the_already_running_session(self):
        """Не даём открыть вторую параллельную сессию на ту же кассетницу —
        ведём к уже идущей."""
        first_session, _ = self._start()

        response = self.client_http.post(f'/inventory/start/{self.cabinet.pk}/', follow=False)

        self.assertEqual(InventorySession.objects.filter(cabinet=self.cabinet).count(), 1)
        self.assertRedirects(response, f'/inventory/{first_session.pk}/count/')

    # --- расхождение как число ---

    def test_discrepancy_is_counted_minus_expected(self):
        session, _ = self._start()
        line_a = session.lines.get(part=self.part_a)

        line_a.counted_quantity = 7
        self.assertEqual(line_a.discrepancy, -3)

        line_a.counted_quantity = 13
        self.assertEqual(line_a.discrepancy, 3)

        line_a.counted_quantity = None
        self.assertIsNone(line_a.discrepancy)

    # --- применение недостачи ---

    def test_deficit_allocates_from_the_oldest_batch_and_reflects_its_price(self):
        """Недостача списывается как обычный расход и уходит по FIFO —
        сначала со старейшей партии; себестоимость расхождения — это
        себестоимость именно задействованной партии."""
        part = SparePart.objects.create(part_number='INV-FIFO', name='Деталь FIFO', current_stock=8, min_stock=0)
        self.cell_c.parts.add(part)
        old_batch = StockMovement.objects.create(
            part=part, quantity=5, movement_type='incoming', unit_price=Decimal('10'))
        StockMovement.objects.create(
            part=part, quantity=3, movement_type='incoming', unit_price=Decimal('20'))

        session, _ = self._start()
        line = session.lines.get(part=part)
        self._count(session, [(line, 3)])  # посчитано 3 при учтённых 8 → недостача 5

        response = self._confirm(session)
        self.assertEqual(response.status_code, 200)

        part.refresh_from_db()
        line.refresh_from_db()
        self.assertEqual(part.current_stock, 3)
        movement = line.movement
        self.assertIsNotNone(movement)
        self.assertEqual(movement.movement_type, 'outgoing')
        self.assertEqual(movement.quantity, 5)
        self.assertIn('Инвентаризация', movement.notes)

        allocations = {a.incoming_id: a.quantity for a in movement.allocations.all()}
        self.assertEqual(allocations, {old_batch.pk: 5})  # только старая партия, новую не трогает
        self.assertEqual(views._cost_from_allocations(movement), Decimal('50'))  # 5 шт. по цене старой партии

    def test_deficit_does_not_require_a_comment(self):
        session, _ = self._start()
        line_a = session.lines.get(part=self.part_a)
        self._count(session, [(line_a, 6)])  # недостача 4

        response = self._confirm(session)  # без комментария

        self.assertEqual(response.status_code, 200)
        session.refresh_from_db()
        self.assertEqual(session.status, InventorySession.STATUS_COMPLETED)
        line_a.refresh_from_db()
        self.assertIsNotNone(line_a.movement)

    # --- применение избытка ---

    def test_surplus_creates_incoming_movement_without_price_and_does_not_touch_part_price(self):
        self.part_b.price = Decimal('7.00')
        self.part_b.save(update_fields=['price'])

        session, _ = self._start()
        line_b = session.lines.get(part=self.part_b)
        self._count(session, [(line_b, 8)])  # излишек 3

        response = self._confirm(session, comments={line_b: 'Найден лишний пакет на полке'})
        self.assertEqual(response.status_code, 200)

        self.part_b.refresh_from_db()
        line_b.refresh_from_db()
        self.assertEqual(self.part_b.current_stock, 8)
        self.assertEqual(self.part_b.price, Decimal('7.00'))  # цену не тронули — находка не покупка
        movement = line_b.movement
        self.assertEqual(movement.movement_type, 'incoming')
        self.assertIsNone(movement.unit_price)
        self.assertIn('Найден лишний пакет на полке', movement.notes)
        self.assertIn('Инвентаризация', movement.notes)

    def test_surplus_without_a_comment_blocks_the_whole_apply(self):
        """Избыток без причины не проходит — и не проходит целиком: даже
        строка с недостачей в той же сессии не применяется, пока не
        заполнен комментарий у строки с избытком."""
        session, _ = self._start()
        line_a = session.lines.get(part=self.part_a)
        line_b = session.lines.get(part=self.part_b)
        self._count(session, [(line_a, 6), (line_b, 8)])  # a: недостача 4, b: излишек 3

        response = self._confirm(session)  # комментарий для b не передан

        self.assertEqual(response.status_code, 200)
        session.refresh_from_db()
        self.assertEqual(session.status, InventorySession.STATUS_IN_PROGRESS)
        line_a.refresh_from_db()
        line_b.refresh_from_db()
        self.assertIsNone(line_a.movement)
        self.assertIsNone(line_b.movement)
        self.part_a.refresh_from_db()
        self.part_b.refresh_from_db()
        self.assertEqual(self.part_a.current_stock, 10)
        self.assertEqual(self.part_b.current_stock, 5)
        self.assertFalse(StockMovement.objects.filter(notes__icontains='Инвентаризация').exists())

        # Заполнив комментарий, ту же сессию можно применить следующим запросом
        ok_response = self._confirm(session, comments={line_b: 'Пересортица при прошлой поставке'})
        self.assertEqual(ok_response.status_code, 200)
        session.refresh_from_db()
        self.assertEqual(session.status, InventorySession.STATUS_COMPLETED)

    # --- расхождение против живого остатка, а не снимка на старте (снос по времени) ---

    def test_apply_uses_live_stock_at_confirm_time_not_the_stale_start_snapshot(self):
        """Между стартом сессии и подтверждением остаток мог измениться
        другим движением (второй сотрудник на складе). Показанное и
        применённое расхождение должно быть против ЖИВОГО остатка —
        после применения current_stock равен ровно посчитанному, а не
        тому, что предполагал снимок на начало сессии."""
        session, _ = self._start()
        line_a = session.lines.get(part=self.part_a)
        self.assertEqual(line_a.expected_quantity, 10)

        self._count(session, [(line_a, 10)])  # сотрудник вводит число, совпадающее со снимком

        # Пока считали, кто-то успел списать 2 шт. этой же детали в заказ —
        # против снимка (10) расхождения нет, но против живого остатка (8) есть
        self.part_a.current_stock = 8
        self.part_a.save(update_fields=['current_stock'])

        confirm_page = self.client_http.get(f'/inventory/{session.pk}/confirm/')
        row = next(r for r in confirm_page.context['rows'] if r['line'].pk == line_a.pk)
        self.assertTrue(row['drifted'])
        self.assertEqual(row['discrepancy'], 2)  # 10 (посчитано) − 8 (живой остаток)

        blocked = self._confirm(session)  # излишек против живого остатка, комментария ещё нет
        session.refresh_from_db()
        self.assertEqual(session.status, InventorySession.STATUS_IN_PROGRESS)

        applied = self._confirm(session, comments={line_a: 'Расхождение из-за параллельного списания'})
        self.assertEqual(applied.status_code, 200)

        self.part_a.refresh_from_db()
        line_a.refresh_from_db()
        session.refresh_from_db()
        self.assertEqual(self.part_a.current_stock, 10)  # ровно посчитанное
        self.assertEqual(session.status, InventorySession.STATUS_COMPLETED)
        self.assertEqual(line_a.movement.movement_type, 'incoming')
        self.assertEqual(line_a.movement.quantity, 2)  # разница против живого (8), а не против снимка (10)

    # --- непосчитанные строки и завершение сессии ---

    def test_uncounted_lines_stay_untouched_and_session_still_completes(self):
        session, _ = self._start()
        line_a = session.lines.get(part=self.part_a)
        line_b = session.lines.get(part=self.part_b)
        self._count(session, [(line_a, 10)])  # только A посчитана, расхождения нет; B не тронута

        response = self._confirm(session)

        self.assertEqual(response.status_code, 200)
        session.refresh_from_db()
        self.assertEqual(session.status, InventorySession.STATUS_COMPLETED)
        line_b.refresh_from_db()
        self.assertIsNone(line_b.counted_quantity)
        self.assertIsNone(line_b.movement)
        self.part_b.refresh_from_db()
        self.assertEqual(self.part_b.current_stock, 5)

    # --- удаление черновика ---

    def test_draft_session_can_be_deleted_before_anything_is_applied(self):
        session, _ = self._start()

        response = self.client_http.post(f'/inventory/{session.pk}/delete/', follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(InventorySession.objects.filter(pk=session.pk).exists())

    def test_session_with_applied_movements_cannot_be_deleted(self):
        session, _ = self._start()
        line_a = session.lines.get(part=self.part_a)
        self._count(session, [(line_a, 6)])
        self._confirm(session)
        session.refresh_from_db()
        self.assertEqual(session.status, InventorySession.STATUS_COMPLETED)

        response = self.client_http.post(f'/inventory/{session.pk}/delete/', follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(InventorySession.objects.filter(pk=session.pk).exists())

    # --- история ---

    def test_history_list_reports_surplus_and_deficit_line_counts(self):
        session, _ = self._start()
        line_a = session.lines.get(part=self.part_a)
        line_b = session.lines.get(part=self.part_b)
        self._count(session, [(line_a, 6), (line_b, 8)])
        self._confirm(session, comments={line_b: 'Излишек при пересчёте'})

        response = self.client_http.get('/inventory/')

        self.assertEqual(response.status_code, 200)
        row = next(s for s in response.context['sessions'] if s.pk == session.pk)
        self.assertEqual(row.deficit_count, 1)
        self.assertEqual(row.surplus_count, 1)


class ConnectionResilienceTests(TestCase):
    """Поведение при обрыве связи с сервером.

    Сама полоса и защита форм живут в браузере, поэтому здесь проверяется
    то, что можно проверить на сервере: разметка есть на любой странице,
    скрипты подключены в рабочем порядке, а запросы страниц идут через
    общий слой связи. Само поведение проверялось в браузере.
    """

    STATIC_JS = Path(__file__).resolve().parent / 'static' / 'js'

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='conn_user', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.employee)

    def test_banner_markup_is_on_every_page(self):
        """Полоса о связи нужна везде: обрыв застаёт человека на той
        странице, на которой он работает, а не на специальной."""
        for url in ('/', '/repair-orders/', '/parts/'):
            with self.subTest(url=url):
                content = self.client_http.get(url).content.decode()
                self.assertIn('id="connectionBanner"', content)
                self.assertIn('id="connectionBannerText"', content)
                self.assertIn('id="connectionRetry"', content)

    def test_banner_is_hidden_without_help_from_bootstrap(self):
        """Bootstrap приходит из интернета, и при неполадках со связью его
        может не оказаться — то есть ровно тогда, когда полоса и работает.
        Спрячься она классом d-none, пустая красная полоса висела бы
        поверх страницы постоянно."""
        content = self.client_http.get('/parts/').content.decode()
        banner = content[content.index('id="connectionBanner"'):]
        banner = banner[:banner.index('</div>')]
        self.assertIn('hidden', banner)
        self.assertNotIn('d-none', banner)
        self.assertIn('[hidden] { display: none !important; }', content)

        source = (self.STATIC_JS / 'connection-status.js').read_text(encoding='utf-8')
        self.assertNotIn('d-none', source)

    def test_banner_script_loads_after_the_shared_connection_layer(self):
        """Полоса подписывается на состояние связи, поэтому общий слой
        должен быть загружен раньше неё, а она — раньше тех, кто этим
        слоем пользуется."""
        content = self.client_http.get('/repair-orders/').content.decode()
        self.assertIn('js/connection-status.js', content)
        self.assertLess(
            content.index('js/ws-connection.js'), content.index('js/connection-status.js')
        )
        self.assertLess(
            content.index('js/connection-status.js'), content.index('js/stock-updates.js')
        )
        self.assertLess(
            content.index('js/connection-status.js'), content.index('js/presence.js')
        )

    def test_only_one_place_in_the_program_opens_a_websocket(self):
        """Переподключение после обрыва — общее. Второй new WebSocket
        означал бы вторую логику повторов, которую починят не всю."""
        opening = [
            path.name for path in self.STATIC_JS.glob('*.js')
            if 'new WebSocket(' in path.read_text(encoding='utf-8')
        ]
        self.assertEqual(opening, ['ws-connection.js'])

    def test_shared_layer_exposes_connection_state_and_request_helpers(self):
        source = (self.STATIC_JS / 'ws-connection.js').read_text(encoding='utf-8')
        for name in ('open:', 'watch:', 'isOffline:', 'fetch:', 'errorText:'):
            with self.subTest(name=name):
                self.assertIn(name, source)

    def test_offline_guard_holds_back_only_posting_forms(self):
        """Фильтры и поиск ходят методом GET и данных не меняют — держать
        их незачем; удерживается только то, что что-то записывает."""
        source = (self.STATIC_JS / 'connection-status.js').read_text(encoding='utf-8')
        self.assertIn("!== 'post'", source)
        self.assertIn('data-offline-ignore', source)
        # Повтор — по кнопке, а не сам собой: очереди отправки в программе нет
        self.assertNotIn('setInterval', source)

    def test_pages_with_background_requests_use_the_shared_helper(self):
        """Молчаливый сбой запроса из окна — то, ради чего всё это
        делалось: раньше в ответ показывался текст исключения браузера."""
        cabinet = Cabinet.objects.create(number=1, name='К1')
        cabinet.apply_layout([4])
        grid = self.client_http.get('/storage-cells/').content.decode()
        self.assertIn('LiftTeamWS.fetch(', grid)
        self.assertNotIn('Ошибка сети:', grid)

    def test_repair_order_form_uses_the_shared_helper(self):
        admin = Employee.objects.create_superuser(
            username='conn_admin', full_name='Админ', password='pass'
        )
        staff = TestClient()
        staff.force_login(admin)
        form = staff.get('/repair-orders/create/').content.decode()
        self.assertIn('LiftTeamWS.fetch(', form)
        self.assertNotIn("alert('Ошибка: ' + err)", form)


class OrderEquipmentLabelsSelectionTests(TestCase):
    """Печать этикеток из карточки заказа: на всё оборудование либо
    на отмеченное."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_eq_labels', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        client_obj = ClientModel.objects.create(name='ООО Эпсилон')
        model = EquipmentModel.objects.create(name='БУАД-7')
        self.order = RepairOrder.objects.create(client=client_obj)
        self.units = [
            RepairOrderEquipment.objects.create(
                repair_order=self.order,
                equipment=Equipment.objects.create(model=model, serial_number=f'SN-30{i}'),
            )
            for i in range(1, 4)
        ]

    def test_without_selection_prints_every_unit(self):
        response = self.http.get(f'/repair-orders/{self.order.pk}/labels/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['labels']), 3)

    def test_selection_prints_only_marked_units(self):
        response = self.http.get(
            f'/repair-orders/{self.order.pk}/labels/',
            {'roe': [self.units[1].pk]},
        )
        labels = response.context['labels']
        self.assertEqual(len(labels), 1)
        self.assertEqual(labels[0]['roe'].pk, self.units[1].pk)

    def test_position_keeps_place_in_order_not_in_selection(self):
        """Наклейка на вторую единицу остаётся «/2», даже когда печатают
        её одну: номер — это место в заказе, а не в отобранном списке."""
        response = self.http.get(
            f'/repair-orders/{self.order.pk}/labels/',
            {'roe': [self.units[2].pk]},
        )
        self.assertEqual(response.context['labels'][0]['position'], 3)

    def test_detail_page_offers_checkboxes_and_print_button(self):
        page = self.http.get(f'/repair-orders/{self.order.pk}/').content.decode()
        self.assertIn('equipmentLabelsForm', page)
        self.assertIn('equipmentCheckAll', page)
        self.assertEqual(page.count('class="form-check-input equipment-check"'), 3)

    def test_broken_selection_does_not_crash(self):
        response = self.http.get(
            f'/repair-orders/{self.order.pk}/labels/', {'roe': ['не число']}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['labels']), 3)


class OrderEquipmentFormsetRowTests(TestCase):
    """Новая строка оборудования берётся из заготовки Django, а не
    копированием строки со страницы.

    Копирование ломало два случая сразу: со второго нажатия в новом заказе
    копия уходила с именами первой строки, и Django сохранял только одно
    значение — в заказе оставалось не всё оборудование; а при правке заказа
    копия уносила ещё и скрытый id сохранённой единицы и переписывала её
    собой.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_formset_row', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        client_obj = ClientModel.objects.create(name='ООО Дзета')
        model = EquipmentModel.objects.create(name='БУАД-9')
        self.order = RepairOrder.objects.create(client=client_obj)
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-401'),
        )

    def test_create_page_carries_empty_form_template(self):
        page = self.http.get('/repair-orders/create/').content.decode()
        self.assertIn('id="equipmentEmptyForm"', page)
        self.assertIn('equipments-__prefix__-equipment', page)

    def test_row_markup_lives_in_one_place(self):
        """Заготовка и строки на странице — один и тот же файл: разойдись
        они, добавленная строка отличалась бы от уже стоящих."""
        form = (settings.BASE_DIR / 'core/templates/core/repair_orders/form.html').read_text(encoding='utf-8')
        self.assertEqual(form.count('_equipment_row.html'), 2)
        self.assertNotIn('cloneNode', form)

    def test_empty_form_row_has_no_saved_id(self):
        """У заготовки скрытый id пуст — иначе новая единица переписала бы
        собой уже сохранённую."""
        page = self.http.get('/repair-orders/create/').content.decode()
        template = page.split('id="equipmentEmptyForm"')[1]
        self.assertIn('equipments-__prefix__-id', template)
        self.assertNotIn('equipments-__prefix__-id" value="', template)


class PrintingChromeTests(TestCase):
    """На бумагу уходит только документ.

    Страница этикетки задаёт размер листа 43 мм, браузер считал это узким
    экраном и включал мобильную вёрстку — на наклейку выезжала кнопка меню.
    А скрытые диалоги прячет Bootstrap, который грузится из интернета:
    при неполадках со связью он не приезжает, и окно уходило на печать.
    """

    def setUp(self):
        self.base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')

    def test_mobile_layout_is_limited_to_screen(self):
        self.assertIn('@media screen and (max-width: 768px)', self.base)
        self.assertNotIn('@media (max-width: 768px)', self.base)

    def test_print_hides_menu_button_and_backdrop(self):
        self.assertIn('.sidebar-toggle, .sidebar-backdrop { display: none !important; }', self.base)

    def test_print_hides_dialogs_without_bootstrap(self):
        self.assertIn('.modal, .modal-backdrop', self.base)
# ======================= ДВА БАНКА, ДВА ЮРЛИЦА =======================


class OrganizationDirectoryTests(TestCase):
    """Справочник юрлиц. До v2.50.0 запись была одна на всю программу."""

    def test_the_first_record_becomes_the_default_one(self):
        """Иначе печатным актам нечего ставить в шапку."""
        first = Organization.objects.create(name='ООО «Первое»')

        first.refresh_from_db()
        self.assertTrue(first.is_default)

    def test_only_one_record_stays_the_default(self):
        """«Основных» два — это два разных бланка на один заказ."""
        first = Organization.objects.create(name='ООО «Первое»')
        second = Organization.objects.create(name='ИП Второй', is_default=True)

        first.refresh_from_db()
        second.refresh_from_db()
        self.assertFalse(first.is_default)
        self.assertTrue(second.is_default)
        self.assertEqual(Organization.objects.filter(is_default=True).count(), 1)

    def test_get_solo_returns_the_default_one(self):
        """Печатные документы зовут его по-старому и должны получать то же."""
        Organization.objects.create(name='ООО «Первое»')
        chosen = Organization.objects.create(name='ИП Второй', is_default=True)

        self.assertEqual(Organization.get_solo().pk, chosen.pk)

    def test_a_bank_points_at_exactly_one_entity(self):
        Organization.objects.create(name='ООО «Первое»', provider='tbank')

        self.assertEqual(Organization.for_provider('tbank').name, 'ООО «Первое»')
        self.assertIsNone(Organization.for_provider('tochka'))

    def test_the_same_bank_cannot_be_taken_twice(self):
        """Иначе по банку нельзя понять, чьи реквизиты ставить в счёт."""
        Organization.objects.create(name='ООО «Первое»', provider='tbank')
        form = OrganizationForm({'name': 'ИП Второй', 'provider': 'tbank'})

        self.assertFalse(form.is_valid())
        self.assertIn('уже закреплён', form.errors['provider'][0])

    def test_entities_without_a_bank_do_not_collide(self):
        Organization.objects.create(name='ООО «Первое»')
        form = OrganizationForm({'name': 'ИП Второй', 'provider': ''})

        self.assertTrue(form.is_valid(), form.errors)


class InvoiceNumberingIsUntouchedTests(TestCase):
    """Ряд номеров один на всю программу и при двух юрлицах.

    Разводить его по юрлицам владелец не захотел: за сквозной нумерацией
    всё равно следит человек, и два ряда он свёл бы труднее одного.
    """

    def setUp(self):
        self.customer = ClientModel.objects.create(name='ООО «Заказчик»')
        Organization.objects.create(name='ООО «Первое»', provider='tbank',
                                    is_default=True)
        Organization.objects.create(name='ИП Второй', provider='tochka')

    def test_the_series_is_shared_by_both_entities(self):
        RepairOrder.objects.create(client=self.customer, invoice_number='943')

        self.assertEqual(RepairOrder.next_invoice_number(), '944')

    def test_the_method_still_takes_no_arguments(self):
        """Подпись не менялась: её зовут из представления счёта."""
        self.assertEqual(RepairOrder.next_invoice_number(), '1')


class OrderLegalEntityTests(TestCase):
    """От какого юрлица идут печатные документы заказа."""

    def setUp(self):
        self.customer = ClientModel.objects.create(name='ООО «Заказчик»')
        self.first = Organization.objects.create(
            name='ООО «Первое»', inn='7701234567', provider='tbank',
            is_default=True)
        self.second = Organization.objects.create(
            name='ИП Второй', inn='772600000000', provider='tochka')

    def test_an_order_without_an_invoice_uses_the_default_entity(self):
        """Обычное состояние нового заказа и всех заказов до v2.50.0."""
        order = RepairOrder.objects.create(client=self.customer)

        self.assertEqual(order.legal_entity().pk, self.first.pk)

    def test_an_order_follows_the_bank_that_issued_its_invoice(self):
        """Иначе заказчик получит по одной работе документы от двух фирм."""
        order = RepairOrder.objects.create(client=self.customer,
                                           invoice_provider='tochka')

        self.assertEqual(order.legal_entity().pk, self.second.pk)

    def test_a_bank_without_an_entity_falls_back_to_the_default(self):
        """Пустая шапка на документе хуже, чем чужая: пустую отправят как есть."""
        self.second.provider = ''
        self.second.save()
        order = RepairOrder.objects.create(client=self.customer,
                                           invoice_provider='tochka')

        self.assertEqual(order.legal_entity().pk, self.first.pk)


class PrintedDocumentsFollowTheInvoiceEntityTests(TestCase):
    """Акты и предложение печатаются от юрлица счёта, а не всегда основного."""

    def setUp(self):
        self.staff = Employee.objects.create_user(
            username='print_staff', full_name='Сотрудник', password='pass',
            role='repair_manager')
        self.http = TestClient()
        self.http.force_login(self.staff)

        self.first = Organization.objects.create(
            name='ООО «Первое»', inn='7701234567', provider='tbank',
            is_default=True)
        self.second = Organization.objects.create(
            name='ИП Второй', inn='772600000000', provider='tochka')

        self.customer = ClientModel.objects.create(name='ООО «Заказчик»')
        self.order = RepairOrder.objects.create(client=self.customer)
        model = EquipmentModel.objects.create(name='Emotron-печать')
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-PR'),
            repair_cost=Decimal('1000'),
        )

    def _pages(self):
        return [
            f'/repair-orders/{self.order.pk}/act/receive/',
            f'/repair-orders/{self.order.pk}/act/complete/',
            f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/act/defect/',
            f'/repair-orders/{self.order.pk}/quote/',
        ]

    def test_by_default_every_document_uses_the_default_entity(self):
        """Прежнее поведение: до v2.50.0 юрлицо было одно."""
        for url in self._pages():
            with self.subTest(url=url):
                resp = self.http.get(url)
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp.context['organization'].pk, self.first.pk)

    def test_every_document_follows_the_bank_of_the_invoice(self):
        self.order.invoice_provider = 'tochka'
        self.order.save(update_fields=['invoice_provider'])

        for url in self._pages():
            with self.subTest(url=url):
                resp = self.http.get(url)
                self.assertEqual(resp.context['organization'].pk, self.second.pk)
                self.assertContains(resp, 'ИП Второй')


class InvoiceProviderInterfaceTests(SimpleTestCase):
    """Общий интерфейс. Всё выше банка не должно знать, какой это банк."""

    def test_both_banks_are_available_by_code(self):
        self.assertEqual(invoicing.get_provider('tbank').label, 'Т-Банк')
        self.assertEqual(invoicing.get_provider('tochka').label, 'Точка Банк')

    def test_an_unknown_bank_is_refused_and_not_guessed(self):
        """Молчаливая подстановка — это чужие реквизиты в чужом документе."""
        with self.assertRaises(invoicing.InvoiceError):
            invoicing.get_provider('sberbank')

    def test_the_prefill_falls_back_when_the_employee_has_no_bank(self):
        self.assertEqual(
            invoicing.default_provider_for(Employee(username='x', full_name='x')),
            'tbank')

    def test_the_prefill_uses_the_employee_choice(self):
        someone = Employee(username='x', full_name='x', default_provider='tochka')

        self.assertEqual(invoicing.default_provider_for(someone), 'tochka')


@override_settings(TOCHKA_TOKEN='jwt-token-value-0123456789',
                   TOCHKA_CUSTOMER_CODE='300000092',
                   TOCHKA_ACCOUNT_ID='40802810/044525104',
                   TOCHKA_INVOICE_ENABLED=True)
class TochkaInvoiceTests(TestCase):
    """Точка Банк. Схема запроса сверена по двум SDK, см. шапку core/tochka.py."""

    ITEMS = [{'name': 'Ремонт привода SN:1', 'price': 54000.0,
              'unit': 'шт.', 'vat': 'None', 'amount': 1}]

    def _built(self, **kwargs):
        return tochka.build_invoice(
            number='943', items=self.ITEMS,
            payer={'name': 'ООО «ЛИФТПРОЕКТ»', 'inn': '9722051089',
                   'kpp': '772201001'},
            invoice_date=datetime.date(2026, 8, 13),
            due_date=datetime.date(2026, 8, 27), **kwargs)

    def test_the_request_carries_the_fields_the_bank_documents(self):
        data = self._built()['Data']

        self.assertEqual(data['accountId'], '40802810/044525104')
        self.assertEqual(data['customerCode'], '300000092')
        self.assertEqual(data['SecondSide']['taxCode'], '9722051089')
        self.assertEqual(data['SecondSide']['type'], 'company')
        invoice = data['Content']['Invoice']
        self.assertEqual(invoice['number'], '943')
        self.assertEqual(invoice['date'], '2026-08-13')
        self.assertEqual(invoice['paymentExpiryDate'], '2026-08-27')
        self.assertEqual(invoice['Positions'][0]['positionName'],
                         'Ремонт привода SN:1')
        self.assertEqual(invoice['Positions'][0]['ndsKind'], 'without_nds')
        self.assertEqual(invoice['totalAmount'], 54000.0)

    def test_a_twelve_digit_inn_is_a_sole_trader(self):
        self.assertEqual(tochka.counterpart_type('772600000000'), 'ip')
        self.assertEqual(tochka.counterpart_type('9722051089'), 'company')

    def test_empty_payer_fields_are_not_sent_at_all(self):
        """Пустая строка в реквизитах счёта выглядит как ошибка."""
        built = tochka.build_invoice(number='1', items=self.ITEMS,
                                     payer={'inn': '9722051089'})

        self.assertNotIn('kpp', built['Data']['SecondSide'])
        self.assertNotIn('secondSideName', built['Data']['SecondSide'])

    def test_the_tbank_unit_setting_does_not_leak_into_tochka(self):
        """Списки допустимых единиц у банков разные."""
        with override_settings(TBANK_INVOICE_UNIT='услуга.',
                               TOCHKA_INVOICE_UNIT='шт.'):
            built = self._built()

        position = built['Data']['Content']['Invoice']['Positions'][0]
        self.assertEqual(position['unitCode'], 'шт.')

    @override_settings(TOCHKA_INVOICE_AMOUNTS_AS_STRING=True)
    def test_amounts_can_be_sent_as_strings(self):
        """Вид сумм по документации не подтверждён — отсюда переключатель."""
        invoice = self._built()['Data']['Content']['Invoice']

        self.assertEqual(invoice['totalAmount'], '54000.00')

    @override_settings(TOCHKA_INVOICE_NDS='nds_25')
    def test_an_unknown_vat_rate_is_refused_before_the_network(self):
        with self.assertRaises(tochka.TochkaError) as caught:
            self._built()

        self.assertIn('nds_25', str(caught.exception))

    def test_sending_returns_the_document_id(self):
        with patch('core.tochka._call',
                   return_value={'Data': {'documentId': 'a1b2c3'}}):
            sent = tochka.send_invoice(self._built())

        self.assertEqual(tochka.document_id(sent), 'a1b2c3')

    def test_a_refusal_in_the_body_is_still_a_refusal(self):
        """Ответ 200 ещё не значит «выставлено»."""
        answer = {'errors': [{'message': 'нет прав на выставление'}]}
        with patch('core.tochka._call', return_value=answer):
            with self.assertRaises(tochka.TochkaError) as caught:
                tochka.send_invoice(self._built())

        self.assertIn('нет прав', str(caught.exception))

    def test_an_answer_without_a_document_id_is_not_taken_as_success(self):
        with patch('core.tochka._call', return_value={'Data': {}}):
            with self.assertRaises(tochka.TochkaError) as caught:
                tochka.send_invoice(self._built())

        self.assertIn('идентификатор', str(caught.exception))

    def test_a_timeout_is_named_in_words(self):
        with patch('core.tochka.request.urlopen', side_effect=TimeoutError()):
            with self.assertRaises(tochka.TochkaError) as caught:
                tochka.send_invoice(self._built())

        self.assertIn('не ответила вовремя', str(caught.exception))

    def test_an_unreachable_bank_is_named_in_words(self):
        with patch('core.tochka.request.urlopen',
                   side_effect=urllib_error.URLError('Connection refused')):
            with self.assertRaises(tochka.TochkaError) as caught:
                tochka.send_invoice(self._built())

        self.assertIn('недоступна', str(caught.exception))

    @override_settings(TOCHKA_INVOICE_ENABLED=False)
    def test_sending_is_off_by_default(self):
        """Отправка документа заказчику не должна включаться сама."""
        self.assertFalse(tochka.invoice_enabled())
        with self.assertRaises(tochka.TochkaError) as caught:
            tochka.send_invoice(self._built())

        self.assertIn('выключено', str(caught.exception))

    @override_settings(TOCHKA_TOKEN='')
    def test_the_missing_settings_are_named(self):
        self.assertIn('TOCHKA_TOKEN', tochka.missing_settings())
        self.assertFalse(tochka.is_configured())

    def test_creating_an_invoice_is_never_retried(self):
        """Повтор с неясным исходом первой попытки — два счёта заказчику."""
        with patch('core.tochka.request.urlopen',
                   side_effect=TimeoutError()) as opened:
            with self.assertRaises(tochka.TochkaError):
                tochka.send_invoice(self._built())

        self.assertEqual(opened.call_count, 1)

    def test_reading_the_status_is_retried(self):
        with patch('core.tochka.request.urlopen',
                   side_effect=TimeoutError()) as opened:
            with self.assertRaises(tochka.TochkaError):
                tochka.payment_status('a1b2c3')

        self.assertEqual(opened.call_count, tochka.READ_RETRIES + 1)

    def test_no_pdf_link_is_invented(self):
        """У Точки ссылки на PDF нет: файл отдаётся по токену."""
        self.assertEqual(tochka.invoice_pdf_url({'Data': {'documentId': 'x'}}), '')


class BankSecretsInLogsTests(TestCase):
    """Токен в журнале живёт до ротации и уезжает в резервную копию."""

    TOKEN = 'jwt-token-value-0123456789'

    @override_settings(TOCHKA_TOKEN=TOKEN, TOCHKA_CUSTOMER_CODE='300000092',
                       TOCHKA_ACCOUNT_ID='40802810/044525104')
    def test_the_tochka_token_never_reaches_the_log(self):
        with patch('core.tochka.request.urlopen', side_effect=TimeoutError()):
            with self.assertLogs('core.tochka', level='INFO') as captured:
                with self.assertRaises(tochka.TochkaError):
                    tochka.payment_status('a1b2c3')

        written = chr(10).join(captured.output)
        self.assertNotIn(self.TOKEN, written)
        self.assertIn('***', written)

    @override_settings(TBANK_TOKEN=TOKEN)
    def test_the_tbank_token_never_reaches_the_log(self):
        with patch('core.tbank.request.urlopen', side_effect=TimeoutError()):
            with self.assertLogs('core.tbank', level='INFO') as captured:
                with self.assertRaises(tbank.TBankError):
                    tbank.get_accounts()

        written = chr(10).join(captured.output)
        self.assertNotIn(self.TOKEN, written)
        self.assertIn('***', written)

    def test_short_values_are_not_blanked_out(self):
        """Замена строки из двух символов испортила бы сообщение."""
        self.assertEqual(net.redact('счёт 40802810 не найден', 'ok'),
                         'счёт 40802810 не найден')


class InvoiceProviderChoiceOnTheFormTests(TestCase):
    """Банк на форме счёта: подставлен, но виден и правится."""

    TOCHKA_ON = {
        'TOCHKA_TOKEN': 'jwt-token-value-0123456789',
        'TOCHKA_CUSTOMER_CODE': '300000092',
        'TOCHKA_ACCOUNT_ID': '40802810/044525104',
        'TOCHKA_INVOICE_ENABLED': True,
    }

    def setUp(self):
        self.tochka_accountant = Employee.objects.create_user(
            username='buh_tochka', full_name='Бухгалтер Точки', password='pass',
            role='accountant', default_provider='tochka')
        self.tbank_accountant = Employee.objects.create_user(
            username='buh_tbank', full_name='Бухгалтер Т-Банка', password='pass',
            role='accountant', default_provider='tbank')
        self.manager = Employee.objects.create_user(
            username='mgr_prov', full_name='Менеджер', password='pass',
            role='repair_manager')

        self.first = Organization.objects.create(
            name='ООО «Первое»', inn='7701234567', provider='tbank',
            is_default=True)
        self.second = Organization.objects.create(
            name='ИП Второй', inn='772600000000', provider='tochka')

        self.customer = ClientModel.objects.create(
            name='ООО «ЛИФТПРОЕКТ»', inn='9722051089', email='buh@liftproekt.ru')
        self.order = RepairOrder.objects.create(client=self.customer)
        model = EquipmentModel.objects.create(name='Emotron-провайдер')
        RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-PRV'),
            repair_cost=Decimal('54000'),
        )
        self.http = TestClient()

    def _url(self):
        return f'/repair-orders/{self.order.pk}/invoice/'

    def test_the_field_is_prefilled_from_the_employee(self):
        self.http.force_login(self.tochka_accountant)

        resp = self.http.get(self._url())

        self.assertEqual(resp.context['form'].initial['provider'], 'tochka')

    def test_the_other_accountant_gets_the_other_bank(self):
        self.http.force_login(self.tbank_accountant)

        resp = self.http.get(self._url())

        self.assertEqual(resp.context['form'].initial['provider'], 'tbank')

    def test_the_field_stays_visible_and_editable(self):
        """Бухгалтеры подменяют друг друга: спрятать поле нельзя."""
        self.http.force_login(self.tochka_accountant)

        resp = self.http.get(self._url())
        field = resp.context['form'].fields['provider']

        self.assertNotIsInstance(field.widget, forms.HiddenInput)
        self.assertFalse(field.disabled)
        self.assertEqual([code for code, _ in field.choices], ['tbank', 'tochka'])

    def test_the_prefilled_bank_can_be_changed_to_the_other_one(self):
        self.http.force_login(self.tbank_accountant)

        with override_settings(**self.TOCHKA_ON):
            with patch('core.tochka._call',
                       return_value={'Data': {'documentId': 'a1b2c3'}}):
                self.http.post(self._url(), {
                    'provider': 'tochka', 'invoice_number': '7',
                    'invoice_date': '2026-08-13', 'due_date': '2026-08-27',
                    'emails': '',
                })

        self.order.refresh_from_db()
        self.assertEqual(self.order.invoice_provider, 'tochka')
        self.assertEqual(self.order.invoice_external_id, 'a1b2c3')

    def test_the_requisites_come_from_the_entity_of_the_chosen_bank(self):
        """Не от основного юрлица: у второго банка своё, и ИНН другой."""
        self.http.force_login(self.tochka_accountant)

        resp = self.http.get(self._url())

        self.assertEqual(resp.context['organization'].pk, self.second.pk)
        self.assertEqual(resp.context['organization'].inn, '772600000000')
        self.assertTrue(resp.context['organization_is_bound'])

    def test_the_other_bank_shows_the_other_entity(self):
        self.http.force_login(self.tbank_accountant)

        resp = self.http.get(self._url())

        self.assertEqual(resp.context['organization'].pk, self.first.pk)
        self.assertEqual(resp.context['organization'].inn, '7701234567')

    def test_a_bank_without_an_entity_is_explained_not_crashed(self):
        self.second.provider = ''
        self.second.save()
        self.http.force_login(self.tochka_accountant)

        resp = self.http.get(self._url())

        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.context['organization_is_bound'])
        self.assertContains(resp, 'не закреплено ни одно юрлицо')

    def test_a_refusal_is_kept_on_the_order_and_nothing_is_marked_as_sent(self):
        """Человек уйдёт со страницы, а причина отказа понадобится потом."""
        self.http.force_login(self.tochka_accountant)

        with override_settings(**self.TOCHKA_ON):
            with patch('core.tochka._call',
                       side_effect=tochka.TochkaError('Точка отказала: нет прав')):
                self.http.post(self._url(), {
                    'provider': 'tochka', 'invoice_number': '7',
                    'invoice_date': '2026-08-13', 'due_date': '2026-08-27',
                    'emails': '',
                })

        self.order.refresh_from_db()
        self.assertIn('нет прав', self.order.invoice_error)
        self.assertIsNone(self.order.invoice_sent_at)
        self.assertEqual(self.order.invoice_number, '')
        self.assertEqual(self.order.invoice_provider, '')

    def test_the_repair_manager_still_does_not_issue_invoices(self):
        self.http.force_login(self.manager)

        self.assertEqual(self.http.get(self._url()).status_code, 302)

    def test_the_admin_still_passes(self):
        boss = Employee.objects.create_superuser(
            username='boss_prov', full_name='Админ', password='pass')
        self.http.force_login(boss)

        self.assertEqual(self.http.get(self._url()).status_code, 200)


class OrderFormStartupTests(TestCase):
    """Первоначальная настройка формы заказа ждёт разбора страницы.

    Скрипт формы стоит в теле страницы, а общие скрипты подключены ниже,
    за содержимым. Выполняясь сразу, он обращался к LiftTeamWS, которого
    в этот момент ещё нет: на странице правки заказа, где оборудование
    уже выбрано, подсказка «эта единица уже обслуживалась» падала
    с ошибкой и не показывалась вовсе. На новом заказе это было незаметно —
    там выбирать ещё нечего.
    """

    def setUp(self):
        self.form = (settings.BASE_DIR / 'core/templates/core/repair_orders/form.html').read_text(
            encoding='utf-8'
        )

    def test_startup_waits_for_dom(self):
        startup = self.form.index("document.querySelectorAll('select[name*=\"-equipment\"]').forEach(showEquipmentHistory)")
        handler = self.form.index("document.addEventListener('DOMContentLoaded'")
        self.assertLess(handler, startup,
                        'первоначальная настройка должна быть внутри DOMContentLoaded')

    def test_shared_scripts_load_before_extra_js(self):
        """Общие скрипты подключены раньше блока extra_js — иначе перенос
        настройки в него ничего бы не дал."""
        base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')
        self.assertLess(base.index('js/ws-connection.js'), base.index('{% block extra_js %}'))


class WebhookIntakeTests(TestCase):
    """Приём уведомлений от банков — дверь, которая смотрит в интернет.

    У Т-Банка подписи нет как таковой: подлинность держится на адресе
    отправителя и на заголовке авторизации, который мы сами задаём банку.
    Обе проверки обязательны, и тесты закрепляют именно это — ни одна
    сама по себе дверь не открывает. У Точки проверять нечем, и она
    отказывает всегда.
    """

    TBANK_URL = '/webhooks/tbank/'
    TOCHKA_URL = '/webhooks/tochka/'

    # Один из шести адресов, перечисленных в документации Т-Банка
    BANK_IP = '212.233.80.7'
    # Заголовок авторизации придумывает владелец и сообщает банку.
    # В тестах — заведомо выдуманная строка
    CREDENTIAL = 'Bearer test-credential-not-a-real-one'

    def setUp(self):
        self.http = TestClient()

    def post(self, url, body=b'{"invoiceId": "inv-1", "status": "PAID"}', **extra):
        return self.http.post(url, data=body, content_type='application/json', **extra)

    def bank_post(self, url=None, **extra):
        """Запрос, у которого верно всё, кроме того, что меняет тест."""
        extra.setdefault('REMOTE_ADDR', self.BANK_IP)
        extra.setdefault('HTTP_AUTHORIZATION', self.CREDENTIAL)
        return self.post(url or self.TBANK_URL, **extra)

    def assertNothingStored(self):
        self.assertEqual(WebhookDelivery.objects.count(), 0)
        self.assertEqual(Payment.objects.count(), 0)

    # --- Выключено ---

    def test_disabled_provider_answers_503(self):
        """По умолчанию приём выключен у обоих банков."""
        for url in (self.TBANK_URL, self.TOCHKA_URL):
            with self.subTest(url=url):
                self.assertEqual(self.post(url).status_code, 503)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_each_provider_is_switched_on_separately(self):
        """Включение одного банка не открывает второй."""
        self.assertEqual(self.bank_post(self.TOCHKA_URL).status_code, 503)
        self.assertEqual(self.bank_post(self.TBANK_URL).status_code, 200)

    # --- Точка: проверять нечем ---

    @override_settings(WEBHOOKS_TOCHKA_ENABLED=True, WEBHOOKS_TOCHKA_SECRET='что угодно')
    def test_tochka_refuses_even_when_enabled(self):
        """У Точки нет ни адресов, ни схемы заверения — принимать нельзя."""
        response = self.bank_post(self.TOCHKA_URL)

        self.assertEqual(response.status_code, 403)
        self.assertNothingStored()

    def test_tochka_refusal_says_what_is_missing(self):
        """Отказ должен называть недостающее, иначе он неотличим от ошибки."""
        verifier = webhooks.get_verifier(invoicing.TOCHKA)

        with self.assertRaises(webhooks.WebhookNotVerifiable) as caught:
            verifier.verify(None, b'{}')

        message = str(caught.exception)
        self.assertIn('адрес', message)
        self.assertIn('документаци', message)

    def test_unknown_provider_code_is_an_error(self):
        with self.assertRaises(webhooks.WebhookError):
            webhooks.get_verifier('sberbank')

    # --- Т-Банк: обе проверки обязательны ---

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET='')
    def test_empty_credential_setting_refuses(self):
        """Пустой секрет — отказ, а не «проверим только по адресу».

        Один список адресов слишком слаб, чтобы остаться единственной
        проверкой: адрес подделывают на пути к нам.
        """
        response = self.bank_post()

        self.assertEqual(response.status_code, 403)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL,
                       WEBHOOKS_TBANK_IPS=[])
    def test_empty_address_list_refuses(self):
        """Пустой список адресов — тоже отказ: проверять не по чему."""
        self.assertEqual(self.bank_post().status_code, 403)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_address_outside_the_bank_list_is_refused(self):
        response = self.bank_post(REMOTE_ADDR='198.51.100.9')

        self.assertEqual(response.status_code, 403)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_missing_credential_is_refused(self):
        response = self.post(self.TBANK_URL, REMOTE_ADDR=self.BANK_IP)

        self.assertEqual(response.status_code, 403)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_wrong_credential_is_refused(self):
        wrong_ones = ('Bearer other', self.CREDENTIAL[:-1],
                      self.CREDENTIAL.replace('Bearer ', ''))
        for wrong in wrong_ones:
            with self.subTest(wrong=wrong):
                response = self.bank_post(HTTP_AUTHORIZATION=wrong)
                self.assertEqual(response.status_code, 403)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_right_address_and_credential_together_open_the_door(self):
        self.assertEqual(self.bank_post().status_code, 200)

    # --- Адрес отправителя за обратным прокси ---

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_forwarded_header_from_an_untrusted_address_is_ignored(self):
        """`X-Forwarded-For` ставит кто угодно — с чужого адреса он не читается."""
        response = self.bank_post(REMOTE_ADDR='198.51.100.9',
                                  HTTP_X_FORWARDED_FOR=self.BANK_IP)

        self.assertEqual(response.status_code, 403)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_behind_the_proxy_the_last_hop_is_the_sender(self):
        """Через свой nginx настоящий адрес — последнее звено заголовка."""
        response = self.bank_post(REMOTE_ADDR='127.0.0.1',
                                  HTTP_X_FORWARDED_FOR=self.BANK_IP)

        self.assertEqual(response.status_code, 200)

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_the_left_side_of_the_forwarded_header_cannot_lie(self):
        """Отправитель прислал свой `X-Forwarded-For`, nginx дописал справа
        настоящий адрес — читается именно он, а не подставленный."""
        response = self.bank_post(
            REMOTE_ADDR='127.0.0.1',
            HTTP_X_FORWARDED_FOR=f'{self.BANK_IP}, 198.51.100.9',
        )

        self.assertEqual(response.status_code, 403)
        self.assertNothingStored()

    def test_client_address_reads_remote_addr_without_a_proxy(self):
        request = RequestFactory().post(
            '/webhooks/tbank/', REMOTE_ADDR='198.51.100.9',
            HTTP_X_FORWARDED_FOR='212.233.80.7')

        self.assertEqual(webhooks.client_address(request), '198.51.100.9')

    def test_client_address_falls_back_to_the_proxy_without_a_header(self):
        request = RequestFactory().post('/webhooks/tbank/', REMOTE_ADDR='127.0.0.1')

        self.assertEqual(webhooks.client_address(request), '127.0.0.1')

    # --- Метод ---

    @override_settings(WEBHOOKS_TBANK_ENABLED=True)
    def test_only_post_is_accepted(self):
        for method in ('get', 'put', 'delete', 'patch'):
            with self.subTest(method=method):
                response = getattr(self.http, method)(self.TBANK_URL)
                self.assertEqual(response.status_code, 405)
        self.assertNothingStored()

    def test_method_is_checked_even_when_disabled(self):
        self.assertEqual(self.http.get(self.TBANK_URL).status_code, 405)

    # --- Размер тела ---

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_MAX_BODY_BYTES=100)
    def test_oversized_body_is_rejected_before_parsing(self):
        """Отказ идёт до разбора: разбирать мегабайт мусора незачем."""
        with patch.object(webhooks, 'parse_payload',
                          side_effect=AssertionError('тело не должно разбираться')):
            response = self.bank_post(body=b'x' * 500)

        self.assertEqual(response.status_code, 413)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_MAX_BODY_BYTES=10000000,
                       DATA_UPLOAD_MAX_MEMORY_SIZE=50)
    def test_body_cut_short_by_django_is_also_413(self):
        """Заголовок длины может и соврать — тогда чтение обрывает сам Django."""
        response = self.bank_post(body=b'x' * 500)

        self.assertEqual(response.status_code, 413)
        self.assertNothingStored()

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_MAX_BODY_BYTES=100,
                       WEBHOOKS_TBANK_SECRET=CREDENTIAL)
    def test_body_within_limit_reaches_the_check(self):
        self.assertEqual(self.bank_post().status_code, 200)

    # --- Общий список адресов поверх банковского ---

    @override_settings(WEBHOOKS_TBANK_ENABLED=True, WEBHOOKS_TBANK_SECRET=CREDENTIAL,
                       WEBHOOKS_ALLOWED_IPS=['203.0.113.7'])
    def test_the_extra_list_narrows_the_bank_list(self):
        self.assertEqual(self.bank_post().status_code, 403)
        self.assertNothingStored()

    # --- Ни входа, ни CSRF-токена ---

    def test_no_login_required(self):
        """Ответ 503, а не переадресация на страницу входа: банк не входит."""
        response = self.post(self.TBANK_URL)

        self.assertEqual(response.status_code, 503)
        self.assertNotIn('login', response.get('Location', ''))

    @override_settings(WEBHOOKS_TBANK_ENABLED=True)
    def test_no_csrf_token_required(self):
        """С включённой проверкой CSRF запрос без токена доходит до проверки
        подлинности — и получает отказ от неё, а не от CSRF."""
        strict = TestClient(enforce_csrf_checks=True)

        response = strict.post(self.TBANK_URL, data=b'{}',
                               content_type='application/json')

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.content, b'forbidden')

    # --- Журнал ---

    @override_settings(WEBHOOKS_TBANK_ENABLED=True,
                       WEBHOOKS_TBANK_SECRET='webhook-secret-0123456789',
                       TBANK_TOKEN='tbank-token-0123456789')
    def test_secrets_never_reach_the_log(self):
        with self.assertLogs('core.webhooks', level='WARNING') as captured:
            self.post(self.TBANK_URL,
                      body=b'{"secret": "webhook-secret-0123456789"}',
                      REMOTE_ADDR='198.51.100.9',
                      HTTP_AUTHORIZATION='Bearer tbank-token-0123456789')

        written = chr(10).join(captured.output)
        self.assertNotIn('webhook-secret-0123456789', written)
        self.assertNotIn('tbank-token-0123456789', written)

    @override_settings(WEBHOOKS_TBANK_ENABLED=True)
    def test_the_log_says_enough_to_understand_the_refusal(self):
        with self.assertLogs('core.webhooks', level='WARNING') as captured:
            self.post(self.TBANK_URL, body=b'{"a": 1}', REMOTE_ADDR='198.51.100.9')

        written = chr(10).join(captured.output)
        self.assertIn('tbank', written)
        self.assertIn('198.51.100.9', written)
        self.assertIn('размер тела 8 байт', written)

    @override_settings(WEBHOOKS_TBANK_ENABLED=True)
    def test_the_body_is_not_written_to_the_log(self):
        """В теле чужие данные, а разбирать отказы можно и без него."""
        with self.assertLogs('core.webhooks', level='WARNING') as captured:
            self.post(self.TBANK_URL, body='{"payer": "ООО Ромашка"}'.encode('utf-8'))

        self.assertNotIn('Ромашка', chr(10).join(captured.output))

    # --- Защита от повтора ---

    def test_the_same_delivery_cannot_be_stored_twice(self):
        """Ограничение уникальности — то, что не даст разнести оплату дважды."""
        body = b'{"invoiceId": "inv-9", "status": "PAID"}'
        digest = webhooks.body_hash(body)
        WebhookDelivery.objects.create(
            provider=invoicing.TBANK, dedup_key=digest, body_hash=digest,
            body=body.decode(),
        )

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                WebhookDelivery.objects.create(
                    provider=invoicing.TBANK, dedup_key=digest, body_hash=digest,
                    body=body.decode(),
                )

        self.assertEqual(WebhookDelivery.objects.count(), 1)

    def test_the_same_body_from_two_banks_is_two_deliveries(self):
        """Ключ уникален внутри банка: одинаковое тело от разных банков —
        разные события."""
        body = b'{"invoiceId": "inv-8", "status": "PAID"}'
        digest = webhooks.body_hash(body)
        for code in (invoicing.TBANK, invoicing.TOCHKA):
            WebhookDelivery.objects.create(provider=code, dedup_key=digest,
                                           body_hash=digest, body=body.decode())

        self.assertEqual(WebhookDelivery.objects.count(), 2)

    def test_dedup_key_of_a_tbank_event_is_the_invoice_id(self):
        """Банк предупреждает, что уведомления могут дублироваться;
        `invoiceId` у события один и тот же."""
        body = b'{"invoiceId": "3fa85f64-5717", "status": "PAID"}'
        verifier = webhooks.get_verifier(invoicing.TBANK)

        key = webhooks.dedup_key(verifier, webhooks.parse_payload(body),
                                 webhooks.body_hash(body))

        self.assertEqual(key, '3fa85f64-5717')

    def test_dedup_key_falls_back_to_the_body_hash(self):
        """Без пригодного идентификатора доставка различается по хешу тела."""
        body = b'{"event": "paid"}'
        verifier = webhooks.get_verifier(invoicing.TOCHKA)
        digest = webhooks.body_hash(body)

        key = webhooks.dedup_key(verifier, webhooks.parse_payload(body), digest)

        self.assertEqual(key, digest)
        self.assertEqual(len(digest), 64)

    def test_body_hash_is_counted_over_the_exact_bytes(self):
        self.assertNotEqual(webhooks.body_hash(b'{"a":1}'), webhooks.body_hash(b'{"a": 1}'))

    def test_broken_json_does_not_raise(self):
        """Разбор тела не должен падать: тело может оказаться чем угодно."""
        self.assertEqual(webhooks.parse_payload(b'not json at all'), {})
        self.assertEqual(webhooks.parse_payload(b'[1, 2, 3]'), {})
        self.assertEqual(webhooks.parse_payload(b'\xff\xfe'), {})


@override_settings(WEBHOOKS_TBANK_ENABLED=True,
                   WEBHOOKS_TBANK_SECRET='Bearer test-credential-not-a-real-one')
class WebhookInvoicePaidTests(TestCase):
    """Событие Т-Банка «счёт оплачен»: что программа с ним делает.

    Главное здесь — чего она НЕ делает. В уведомлении нет суммы, поэтому
    ни `Payment`, ни статус оплаты заказа оно не трогает: сумма, взятая
    с потолка, — это неверный долг у заказчика. Уведомление ставит отметку
    и зовёт бухгалтера внести поступление по выписке.
    """

    URL = '/webhooks/tbank/'
    BANK_IP = '212.233.80.7'
    CREDENTIAL = 'Bearer test-credential-not-a-real-one'
    INVOICE_ID = '3fa85f64-5717-4562-b3fc-2c963f66afa6'

    def setUp(self):
        self.http = TestClient()
        self.accountant = Employee.objects.create_user(
            username='buh_wh', full_name='Бухгалтер', password='pass',
            role='accountant', email='buh@example.com',
        )
        self.client_obj = ClientModel.objects.create(name='ООО Плательщик')
        self.order = RepairOrder.objects.create(
            client=self.client_obj, payment_status='unpaid',
            invoice_number='77', invoice_date=datetime.date.today(),
            invoice_provider=invoicing.TBANK,
            invoice_external_id=self.INVOICE_ID,
        )

    def deliver(self, invoice_id=None, status='PAID', body=None):
        if body is None:
            body = json.dumps({
                'invoiceId': invoice_id or self.INVOICE_ID,
                'status': status,
            }).encode('utf-8')
        return self.http.post(self.URL, data=body, content_type='application/json',
                              REMOTE_ADDR=self.BANK_IP,
                              HTTP_AUTHORIZATION=self.CREDENTIAL)

    def notices(self):
        return list(Notification.objects.filter(event='invoice_paid'))

    def test_paid_event_stamps_the_order_and_calls_the_accountant(self):
        response = self.deliver()

        self.assertEqual(response.status_code, 200)
        self.order.refresh_from_db()
        self.assertIsNotNone(self.order.invoice_paid_at)
        self.assertEqual([n.recipient for n in self.notices()], ['buh@example.com'])
        self.assertEqual(self.notices()[0].repair_order_id, self.order.pk)

    def test_no_money_is_posted_by_the_notification(self):
        """Суммы в уведомлении нет — значит, и оплаты в программе нет."""
        self.deliver()

        self.order.refresh_from_db()
        self.assertEqual(Payment.objects.count(), 0)
        self.assertEqual(self.order.payment_status, 'unpaid')
        self.assertEqual(self.order.paid_amount, 0)

    def test_the_notice_says_the_accountant_still_has_to_post_the_receipt(self):
        self.deliver()

        self.assertIn('по выписке', self.notices()[0].body)

    def test_the_delivery_is_recorded_as_processed(self):
        self.deliver()

        delivery = WebhookDelivery.objects.get()
        self.assertEqual(delivery.provider, invoicing.TBANK)
        self.assertEqual(delivery.event_id, self.INVOICE_ID)
        self.assertEqual(delivery.dedup_key, self.INVOICE_ID)
        self.assertEqual(delivery.status, WebhookDelivery.STATUS_PROCESSED)
        self.assertIn(self.order.order_number, delivery.result)
        self.assertIsNotNone(delivery.processed_at)

    def test_the_same_delivery_twice_changes_nothing_the_second_time(self):
        """Банк прямо предупреждает, что доставка может повториться."""
        first = self.deliver()
        second = self.deliver()

        self.assertEqual((first.status_code, second.status_code), (200, 200))
        self.assertEqual(WebhookDelivery.objects.count(), 1)
        self.assertEqual(len(self.notices()), 1)

    def test_an_unknown_invoice_is_recorded_and_not_retried(self):
        """Банк доставил всё верно — это мы не знаем такого счёта.
        Значит, 200: повторять доставку ему незачем."""
        response = self.deliver(invoice_id='чужой-счёт-1')

        self.assertEqual(response.status_code, 200)
        delivery = WebhookDelivery.objects.get()
        self.assertEqual(delivery.status, WebhookDelivery.STATUS_UNMATCHED)
        self.assertEqual(self.notices(), [])
        self.order.refresh_from_db()
        self.assertIsNone(self.order.invoice_paid_at)

    def test_an_unmatched_delivery_points_at_the_likely_cause(self):
        """Самая вероятная причина — у заказов не сохранён идентификатор
        счёта; без подсказки разбираться пришлось бы с нуля."""
        self.order.invoice_external_id = ''
        self.order.invoice_sent_at = timezone.now()
        self.order.save(update_fields=['invoice_external_id', 'invoice_sent_at'])

        self.deliver()

        self.assertIn('идентификатор счёта не сохранён',
                      WebhookDelivery.objects.get().result)

    def test_an_invoice_of_another_bank_is_not_matched(self):
        """Идентификаторы счетов разных банков совпадать не обязаны,
        и отмечать заказ Точки по уведомлению Т-Банка нельзя."""
        self.order.invoice_provider = invoicing.TOCHKA
        self.order.save(update_fields=['invoice_provider'])

        self.deliver()

        self.assertEqual(WebhookDelivery.objects.get().status,
                         WebhookDelivery.STATUS_UNMATCHED)
        self.order.refresh_from_db()
        self.assertIsNone(self.order.invoice_paid_at)

    def test_a_status_other_than_paid_is_refused(self):
        """Других значений у этого события не бывает — значит, это не оно."""
        response = self.deliver(status='CANCELLED')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(WebhookDelivery.objects.count(), 0)
        self.order.refresh_from_db()
        self.assertIsNone(self.order.invoice_paid_at)

    def test_a_body_without_an_invoice_id_is_refused(self):
        response = self.deliver(body=b'{"status": "PAID"}')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(WebhookDelivery.objects.count(), 0)

    def test_malformed_json_is_refused_and_stored_nowhere(self):
        response = self.deliver(body=b'not json at all')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(WebhookDelivery.objects.count(), 0)
        self.assertEqual(self.notices(), [])

    def test_an_unexpected_failure_lets_the_bank_retry(self):
        """Иначе повтор банка отбросился бы как дубль, а оповещение
        об оплате пропало бы совсем."""
        with patch.object(webhooks, 'apply_event', side_effect=RuntimeError('сбой')):
            with self.assertLogs('core.webhooks', level='WARNING'):
                broken = self.deliver()

        self.assertEqual(broken.status_code, 500)
        self.assertEqual(WebhookDelivery.objects.count(), 0)
        self.order.refresh_from_db()
        self.assertIsNone(self.order.invoice_paid_at)

        # Повтор банка проходит целиком
        self.assertEqual(self.deliver().status_code, 200)
        self.order.refresh_from_db()
        self.assertIsNotNone(self.order.invoice_paid_at)
        self.assertEqual(len(self.notices()), 1)

    def test_the_stamp_is_shown_on_the_order_page(self):
        """Отметка должна быть видна, и рядом с ней — оговорка про выписку."""
        self.deliver()
        self.http.force_login(self.accountant)

        response = self.http.get(f'/repair-orders/{self.order.pk}/')

        self.assertContains(response, 'Банк сообщил об оплате счёта')
        self.assertContains(response, 'по выписке')

    def test_a_repeat_after_the_log_was_cleared_does_not_notify_twice(self):
        """Если журнал доставок чистили, отметка всё равно ставится один раз."""
        self.deliver()
        WebhookDelivery.objects.all().delete()

        response = self.deliver()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(self.notices()), 1)
        self.assertEqual(WebhookDelivery.objects.get().status,
                         WebhookDelivery.STATUS_PROCESSED)


class TBankInvoiceExternalIdTests(SimpleTestCase):
    """Идентификатор счёта из ответа Т-Банка.

    Он нужен ровно затем, чтобы уведомление об оплате нашло заказ:
    в уведомлении приезжает только он. Имя поля в ответе банка
    не подтверждено, поэтому читаются несколько правдоподобных, а годным
    считается только значение вида UUID — так догадка отказывает молча,
    а не подставляет в заказ что попало.
    """

    def test_a_uuid_is_taken(self):
        value = tbank.invoice_external_id(
            {'invoiceId': '3fa85f64-5717-4562-b3fc-2c963f66afa6', 'pdfUrl': 'http://x'})

        self.assertEqual(value, '3fa85f64-5717-4562-b3fc-2c963f66afa6')

    def test_other_field_names_are_tried_too(self):
        for key in ('invoice_id', 'documentId', 'id'):
            with self.subTest(key=key):
                value = tbank.invoice_external_id(
                    {key: '3fa85f64-5717-4562-b3fc-2c963f66afa6'})
                self.assertEqual(value, '3fa85f64-5717-4562-b3fc-2c963f66afa6')

    def test_anything_that_is_not_a_uuid_is_ignored(self):
        """Не тот ключ — лучше пусто, чем чужое значение в поле счёта."""
        for response in ({'id': '943'}, {'id': ''}, {'documentId': 'счёт номер два'},
                         {'pdfUrl': 'https://example.ru/invoice.pdf'}, {}, None):
            with self.subTest(response=response):
                self.assertEqual(tbank.invoice_external_id(response), '')

    def test_the_provider_passes_it_through(self):
        provider = invoicing.get_provider(invoicing.TBANK)

        value = provider.external_id({'invoiceId': '3fa85f64-5717-4562-b3fc-2c963f66afa6'})

        self.assertEqual(value, '3fa85f64-5717-4562-b3fc-2c963f66afa6')


class PartSearchEndpointTests(TestCase):
    """Поиск детали для выбора из формы.

    Отбор здесь тот же, что у списка радиодеталей, и это главное:
    вторая, чуть иначе написанная фильтрация означала бы, что выбор детали
    в заказе находит не то же, что склад.
    """

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='search_user', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.employee)

        self.cabinet = Cabinet.objects.create(number=7, name='Поиск')
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.order_by('cell_row').first()

        self.resistor = SparePart.objects.create(
            part_number='R-100', name='Резистор 100 Ом', component_type='Резистор',
            package='0805', resistance=Decimal('100'), resistance_unit='Ом',
            current_stock=25, min_stock=5,
        )
        self.capacitor = SparePart.objects.create(
            part_number='C-220', name='Конденсатор 220 мкФ', component_type='Конденсатор',
            package='DIP', capacitance=Decimal('220'), capacitance_unit='мкФ',
            current_stock=0, min_stock=3,
        )
        self.transistor = SparePart.objects.create(
            part_number='T-500', name='Транзистор мощный', component_type='Транзистор',
            current_stock=4, min_stock=0,
        )
        self.cell.parts.add(self.resistor)

    def _search(self, **params):
        return self.client_http.get('/parts/search/', params).json()

    def test_requires_login(self):
        """Каталог склада — не для гостя: остальные складские страницы
        закрыты так же."""
        response = TestClient().get('/parts/search/')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])

    def test_free_text_finds_by_article_and_by_name(self):
        by_article = self._search(q='R-100')
        self.assertEqual([row['id'] for row in by_article['results']], [self.resistor.pk])

        by_name = self._search(q='Транзистор мощный')
        self.assertEqual([row['id'] for row in by_name['results']], [self.transistor.pk])

    def test_filters_by_component_type(self):
        data = self._search(component_type='Конденсатор')
        self.assertEqual([row['id'] for row in data['results']], [self.capacitor.pk])

    def test_in_stock_filter_hides_what_is_not_on_the_shelf(self):
        """«Только в наличии» — про то, лежит деталь на полке или нет,
        а не про близость к минимальному остатку."""
        data = self._search(in_stock='1')
        found = {row['id'] for row in data['results']}
        self.assertIn(self.resistor.pk, found)
        self.assertIn(self.transistor.pk, found)
        self.assertNotIn(self.capacitor.pk, found)

    def test_row_carries_stock_and_cell(self):
        """Остаток и адрес ячейки — то, по чему решают, идти к полке
        или заказывать. Без них выбор просто отодвигает вопрос."""
        row = self._search(q='R-100')['results'][0]
        self.assertEqual(row['part_number'], 'R-100')
        self.assertEqual(row['name'], 'Резистор 100 Ом')
        self.assertEqual(row['stock'], 25)
        self.assertEqual(row['min_stock'], 5)
        self.assertEqual(row['cell'], self.cell.address)
        self.assertEqual(row['specs'], '100Ом')
        self.assertEqual(row['stock_state'], 'ok')

    def test_part_without_a_cell_says_so_instead_of_breaking(self):
        row = self._search(q='T-500')['results'][0]
        self.assertEqual(row['cell'], '')

    def test_result_cap_is_reported(self):
        """Показано не всё — об этом надо сказать прямо, иначе человек
        решит, что остальных деталей в программе нет."""
        SparePart.objects.bulk_create([
            SparePart(part_number=f'MASS-{index:03d}', name='Массовая деталь', current_stock=1)
            for index in range(60)
        ])
        data = self._search(q='Массовая')

        self.assertEqual(len(data['results']), views.PART_SEARCH_LIMIT)
        self.assertEqual(data['total'], 60)
        self.assertTrue(data['limited'])
        self.assertEqual(data['limit'], views.PART_SEARCH_LIMIT)

    def test_short_result_is_not_marked_as_capped(self):
        data = self._search(q='R-100')
        self.assertFalse(data['limited'])
        self.assertEqual(data['total'], 1)

    def test_excluded_parts_do_not_show_up(self):
        """В ячейке уже лежащую деталь предлагать в неё же незачем."""
        data = self._search(exclude=str(self.resistor.pk))
        self.assertNotIn(self.resistor.pk, {row['id'] for row in data['results']})

    def test_lookup_by_id_returns_the_label_of_an_already_chosen_part(self):
        data = self._search(id=str(self.capacitor.pk))
        self.assertEqual([row['id'] for row in data['results']], [self.capacitor.pk])
        self.assertEqual(data['results'][0]['label'], 'C-220 — Конденсатор 220 мкФ')

    def test_component_types_are_sent_only_when_asked(self):
        """Список типов выбору нужен один раз, на первое открытие: тянуть
        его на каждую букву — лишний запрос к базе на каждое нажатие."""
        self.assertNotIn('component_types', self._search(q='R'))
        self.assertEqual(
            self._search(with_types='1')['component_types'],
            ['Конденсатор', 'Резистор', 'Транзистор'],
        )

    def test_search_uses_the_same_filtering_as_the_parts_list(self):
        """Одна фильтрация на список, выгрузку и выбор детали: разойдись
        они, «найдено 3» на складе не совпало бы с тем, что предлагает
        выбор детали в заказе."""
        params = {'q': 'Резистор', 'component_type': 'Резистор', 'stock_from': '10'}
        found_in_picker = {row['id'] for row in self._search(**params)['results']}
        listed = self.client_http.get('/parts/', params).context['parts']
        self.assertEqual(found_in_picker, {part.pk for part in listed})


class PartPickerEverywhereTests(TestCase):
    """Выбор детали стоит везде, где выбирают деталь.

    Проверяется по каждой странице отдельно: страница, тихо вернувшаяся
    к списку на весь каталог, — это ровно то, на что жаловался владелец.
    """

    STATIC = Path(__file__).resolve().parent / 'static'

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='picker_admin', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.part = SparePart.objects.create(
            part_number='PK-1', name='Деталь выбора', component_type='Резистор',
            current_stock=10, min_stock=1,
        )
        self.cabinet = Cabinet.objects.create(number=3, name='Выбор')
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.order_by('cell_row').first()

        self.model = EquipmentModel.objects.create(name='БУАД-выбор')
        self.fault = FaultType.objects.create(equipment_model=self.model, name='Не включается')
        FaultTypePart.objects.create(fault_type=self.fault, part=self.part, quantity=2)

        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Выбор', inn='7700000111')
        )

    def _picker_code(self):
        """Сам код выбора детали, без пояснений в комментариях: в них
        как раз и написано, чего здесь быть не должно."""
        source = (self.STATIC / 'js' / 'part-picker.js').read_text(encoding='utf-8')
        source = re.sub(r'/\*.*?\*/', '', source, flags=re.S)
        return re.sub(r'^\s*//.*$', '', source, flags=re.M)

    def _pages(self):
        return {
            'заказ': f'/repair-orders/{self.order.pk}/',
            'типовая неисправность': f'/faults/{self.fault.pk}/edit/',
            'сетка кассетниц': '/storage-cells/',
            'журнал движений': '/reports/stock-movements/',
        }

    def test_every_page_carries_the_picker(self):
        for title, url in self._pages().items():
            with self.subTest(page=title):
                content = self.client_http.get(url).content.decode()
                self.assertIn('data-part-picker', content)
                self.assertIn('part-picker-value', content)
                self.assertIn('/parts/search/', content)

    def test_no_page_renders_the_whole_catalogue_as_a_list(self):
        """Каталог целиком в разметке — то, ради чего всё это делалось:
        листать несколько сотен деталей в поисках одной невозможно."""
        for title, url in self._pages().items():
            with self.subTest(page=title):
                content = self.client_http.get(url).content.decode()
                self.assertNotIn(f'value="{self.part.pk}">PK-1', content)
                self.assertNotIn('PK-1 — Деталь выбора</option>', content)

    def test_grid_page_no_longer_ships_the_catalogue_to_the_browser(self):
        """Раньше сетка увозила в браузер весь каталог отдельным списком
        ALL_PARTS — на каждое открытие кассетницы."""
        content = self.client_http.get('/storage-cells/').content.decode()
        self.assertNotIn('ALL_PARTS', content)
        self.assertNotIn('addPartSelect', content)

    def test_field_names_are_the_ones_the_views_already_expect(self):
        """Скрытое поле подставляется под прежним именем — принимающие
        представления и формы об этой замене не знают."""
        order_page = self.client_http.get(f'/repair-orders/{self.order.pk}/').content.decode()
        self.assertIn('name="part"', order_page)

        fault_page = self.client_http.get(f'/faults/{self.fault.pk}/edit/').content.decode()
        self.assertIn('name="parts-0-part"', fault_page)

        grid = self.client_http.get('/storage-cells/').content.decode()
        self.assertIn('name="part_id"', grid)

        report = self.client_http.get('/reports/stock-movements/').content.decode()
        self.assertIn('name="part"', report)

    def test_already_chosen_part_is_shown_by_name(self):
        """Рецепт неисправности открывается с уже выбранными деталями —
        человек должен видеть, что выбрано, а не пустое поле."""
        content = self.client_http.get(f'/faults/{self.fault.pk}/edit/').content.decode()
        self.assertIn(f'value="{self.part.pk}"', content)
        self.assertIn('PK-1 — Деталь выбора', content)

    def test_adding_a_part_to_an_order_still_works(self):
        response = self.client_http.post(
            f'/repair-orders/{self.order.pk}/add-detail/',
            {'part': self.part.pk, 'quantity_used': 3},
        )
        self.assertEqual(response.status_code, 302)
        detail = self.order.details.get()
        self.assertEqual(detail.part_id, self.part.pk)
        self.assertEqual(detail.quantity_used, 3)

    def test_adding_a_part_to_a_cell_still_works(self):
        response = self.client_http.post(
            f'/storage-cells/{self.cell.pk}/add-part/', {'part_id': self.part.pk}
        )
        self.assertTrue(response.json()['success'])
        self.assertIn(self.part, self.cell.parts.all())

    def test_saving_a_fault_recipe_still_works(self):
        other = SparePart.objects.create(part_number='PK-2', name='Вторая деталь', current_stock=1)
        response = self.client_http.post(f'/faults/{self.fault.pk}/edit/', {
            'equipment_model': self.model.pk,
            'name': 'Не включается',
            'description': '',
            'parts-TOTAL_FORMS': '2',
            'parts-INITIAL_FORMS': '1',
            'parts-MIN_NUM_FORMS': '0',
            'parts-MAX_NUM_FORMS': '1000',
            'parts-0-id': self.fault.parts.get().pk,
            'parts-0-fault_type': self.fault.pk,
            'parts-0-part': self.part.pk,
            'parts-0-quantity': '2',
            'parts-1-part': other.pk,
            'parts-1-quantity': '5',
        })
        self.assertEqual(response.status_code, 302)
        recipe = {line.part_id: line.quantity for line in self.fault.parts.all()}
        self.assertEqual(recipe, {self.part.pk: 2, other.pk: 5})

    def test_report_filter_by_part_still_works(self):
        StockMovement.objects.create(
            part=self.part, movement_type='incoming', quantity=5,
        )
        page = self.client_http.get('/reports/stock-movements/', {'part': self.part.pk})
        self.assertEqual([m.part_id for m in page.context['movements']], [self.part.pk])
        self.assertEqual(page.context['selected_part'], self.part)

    def test_new_recipe_row_gets_its_own_field_numbers(self):
        """Кнопка «Добавить деталь» в рецепте копировала первую строку,
        а имена полей давала ей по номеру последней. Новая строка приходила
        с именами первой, и рецепт не сохранялся вовсе — форма отвечала
        «Деталь: обязательное поле» на заполненную деталь."""
        source = (Path(__file__).resolve().parent / 'templates' / 'core' /
                  'faults' / 'form.html').read_text(encoding='utf-8')
        self.assertIn('rows[rows.length - 1]', source)
        self.assertNotIn("'parts-' + (partsFormNum - 1) + '-'", source)
        # Номер записи от образца в новой строке означал бы правку старой
        self.assertIn('input[name$="-id"]', source)

    def test_picker_does_not_lean_on_bootstrap_javascript(self):
        """Bootstrap приезжает из интернета, и дважды уже не приезжал.
        Выбор детали обязан работать без него."""
        # Проверяем готовую разметку, а не исходник шаблона: в исходнике
        # `data-bs-` встречается в пояснении, почему его здесь нет
        markup = render_to_string('core/_part_picker.html', {'name': 'part'})
        self.assertNotIn('data-bs-', markup)
        self.assertNotIn('class="modal', markup)
        # Панель прячется атрибутом hidden — за ним в base.html закреплено
        # display: none !important, и оно не зависит от чужих стилей
        self.assertIn('hidden', markup)

        source = self._picker_code()
        self.assertNotIn('bootstrap.', source)
        self.assertNotIn('data-bs-', source)

    def test_picker_asks_the_server_through_the_shared_connection_layer(self):
        """Голый fetch спрятал бы обрыв связи под «ничего не найдено»."""
        source = self._picker_code()
        self.assertIn('LiftTeamWS.fetch(', source)
        self.assertNotIn('window.fetch(', source)
        self.assertNotIn('= fetch(', source)

    def test_picker_files_are_loaded_on_every_page(self):
        """Выбор подключается одной вставкой шаблона: искать, где ещё
        дописать скрипт, при добавлении пятого места не придётся."""
        content = self.client_http.get('/parts/').content.decode()
        self.assertIn('js/part-picker.js', content)
        self.assertIn('css/part-picker.css', content)
        self.assertLess(content.index('js/ws-connection.js'), content.index('js/part-picker.js'))

    def test_keyboard_works_without_a_mouse(self):
        """Мастера работают быстро, и выбор, требующий мыши, раздражает
        сильнее прежнего длинного списка."""
        source = (self.STATIC / 'js' / 'part-picker.js').read_text(encoding='utf-8')
        for key in ('ArrowDown', 'ArrowUp', 'Enter', 'Escape'):
            with self.subTest(key=key):
                self.assertIn(key, source)

    def test_empty_state_and_cap_notice_are_spelled_out(self):
        source = (self.STATIC / 'js' / 'part-picker.js').read_text(encoding='utf-8')
        self.assertIn('Ничего не найдено — уточните запрос', source)
        self.assertIn('Показаны первые ', source)


class WebhookTBankAddressListTests(SimpleTestCase):
    """Шесть адресов Т-Банка перечислены дважды — в настройках и в nginx.

    Списки обязаны совпадать: nginx отсекает чужой запрос до приложения,
    приложение проверяет то же самое ещё раз. Разойдясь молча, они дали бы
    либо глухую стену для настоящих уведомлений, либо дыру на прокси.
    """

    # Из документации Т-Банка, раздел «Вебхуки» → «Подключение».
    # Сеть 91.194.226.0/23 оттуда же сюда не входит: она относится
    # к товарному кредитованию покупателей, которым программа не пользуется
    DOCUMENTED = (
        '212.233.80.7',
        '91.218.132.2',
        '91.194.226.234',
        '91.194.226.235',
        '91.194.226.250',
        '91.194.226.251',
    )

    def directives(self, text):
        """Только строки настройки, без объяснений в комментариях.

        Иначе тест ловил бы упоминание адреса в тексте «этот адрес сюда
        НЕ вписан» и считал его вписанным.
        """
        return chr(10).join(
            line for line in text.splitlines() if not line.strip().startswith('#')
        )

    def conf(self, name):
        return (Path(settings.BASE_DIR) / 'deploy' / name).read_text(encoding='utf-8')

    def nginx_conf(self):
        return self.directives(self.conf('nginx-lifteam-hooks.conf'))

    def test_the_documented_addresses_are_the_default(self):
        self.assertEqual(tuple(settings.WEBHOOKS_TBANK_IPS_DOCUMENTED), self.DOCUMENTED)

    def test_nginx_lets_through_exactly_the_same_addresses(self):
        allowed = set(re.findall(r'allow\s+([0-9.]+);', self.nginx_conf()))

        self.assertEqual(allowed, set(self.DOCUMENTED),
                         'списки адресов в настройках и в nginx разошлись')

    def test_nginx_closes_everything_else_under_the_prefix(self):
        conf = self.nginx_conf()
        hooks = conf[conf.index('location /webhooks/'):]

        self.assertIn('deny all;', hooks)

    def test_the_consumer_credit_network_is_not_allowed(self):
        """91.194.226.0/23 — это 510 адресов чужого продукта."""
        self.assertNotIn('91.194.226.0/23', self.nginx_conf())
        self.assertNotIn('91.194.226.0/23', ','.join(settings.WEBHOOKS_TBANK_IPS))

    def test_the_application_block_is_not_open_to_the_world(self):
        """Заголовок Host подделывается — приложение держит на замке адрес."""
        conf = self.directives(self.conf('nginx-lifteam.conf'))

        self.assertNotIn('server_name _;', conf)
        self.assertIn('allow 100.64.0.0/10;', conf)
        self.assertIn('deny all;', conf)


class WebhookUrlIsolationTests(TestCase):
    """Наружу открыт ровно один каталог адресов — `/webhooks/`.

    nginx на Raspberry Pi пускает снаружи всё, что начинается с этого
    префикса, и отвечает 404 на остальное. Поэтому любой посторонний
    маршрут, случайно оказавшийся под префиксом, оказался бы открыт
    интернету, а вьюха приёма за его пределами — недоступна банку.
    """

    WEBHOOK_MODULE = 'core.webhook_views'
    PREFIX = 'webhooks/'

    def routes(self):
        """Все маршруты программы: полный путь и то, что его обслуживает."""
        from django.urls import get_resolver
        from django.urls.resolvers import URLResolver

        def walk(resolver, prefix=''):
            for pattern in resolver.url_patterns:
                path = prefix + str(pattern.pattern)
                if isinstance(pattern, URLResolver):
                    yield from walk(pattern, path)
                else:
                    yield path, pattern.callback

        return list(walk(get_resolver()))

    def test_nothing_but_webhooks_lives_under_the_prefix(self):
        strangers = [
            path for path, view in self.routes()
            if path.startswith(self.PREFIX)
            and getattr(view, '__module__', '') != self.WEBHOOK_MODULE
        ]

        self.assertEqual(strangers, [], 'под /webhooks/ попал посторонний маршрут')

    def test_webhook_views_live_nowhere_else(self):
        outside = [
            path for path, view in self.routes()
            if getattr(view, '__module__', '') == self.WEBHOOK_MODULE
            and not path.startswith(self.PREFIX)
        ]

        self.assertEqual(outside, [], 'вьюха приёма оказалась вне /webhooks/')

    def test_both_banks_are_routed(self):
        paths = {path for path, view in self.routes()
                 if getattr(view, '__module__', '') == self.WEBHOOK_MODULE}

        self.assertEqual(paths, {'webhooks/tbank/', 'webhooks/tochka/'})

    def test_an_unknown_bank_under_the_prefix_is_404(self):
        response = TestClient().post('/webhooks/sberbank/', data=b'{}',
                                     content_type='application/json')

        self.assertEqual(response.status_code, 404)

    def test_the_prefix_itself_is_not_a_page(self):
        self.assertEqual(TestClient().get('/webhooks/').status_code, 404)


class EquipmentOwnerFromOrderTests(TestCase):
    """Оборудование, заведённое из заказа, получает владельца.

    Оборудование почти всегда заводят прямо из формы заказа, где заказчик
    уже выбран, — а владелец при этом не проставлялся вовсе. Поле
    `current_client` заполнялось только через отдельную форму
    редактирования оборудования, куда почти никто не заходит.
    """

    def setUp(self):
        self.user = Employee.objects.create_superuser(
            username='owner_test', full_name='Приёмщик', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.user)
        self.client_obj = ClientModel.objects.create(name='МУП «Городские лифты»')
        self.other = ClientModel.objects.create(name='ООО «Лифтсервис»')
        self.model = EquipmentModel.objects.create(name='БУАД-7-31')

    def _create_equipment(self, serial, client=None):
        data = {'model_id': self.model.pk, 'serial_number': serial, 'confirmed': '1'}
        if client is not None:
            data['client_id'] = client.pk
        return self.http.post('/ajax/equipment/create/', data)

    def test_equipment_created_from_an_order_gets_its_client(self):
        response = self._create_equipment('SN-OWN-1', self.client_obj)
        self.assertTrue(response.json()['success'])
        self.assertEqual(
            Equipment.objects.get(serial_number='SN-OWN-1').current_client,
            self.client_obj,
        )

    def test_without_a_chosen_client_nothing_is_invented(self):
        """Заказчика ещё не выбрали — владелец остаётся пустым.

        Подставить сюда «какого-нибудь» заказчика хуже, чем оставить
        пусто: пустое поле видно, а неверное — нет.
        """
        self._create_equipment('SN-OWN-2')
        self.assertIsNone(
            Equipment.objects.get(serial_number='SN-OWN-2').current_client
        )

    def test_saving_the_order_fills_owners_left_empty(self):
        """Единицу могли завести до того, как выбрали заказчика, или взять
        из справочника, где владельца никогда не было."""
        equipment = Equipment.objects.create(model=self.model, serial_number='SN-OWN-3')
        order = RepairOrder.objects.create(client=self.client_obj)
        RepairOrderEquipment.objects.create(repair_order=order, equipment=equipment)

        self.assertEqual(order.assign_equipment_owners(), 1)
        equipment.refresh_from_db()
        self.assertEqual(equipment.current_client, self.client_obj)

    def test_an_existing_owner_is_never_overwritten(self):
        """Прибор может приехать от другого предприятия. Менять ли владельца
        в этом случае — вопрос к владельцу программы (PLAN.md), и пока он
        не решён, молча переписывать чужую запись нельзя."""
        equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-OWN-4', current_client=self.other
        )
        order = RepairOrder.objects.create(client=self.client_obj)
        RepairOrderEquipment.objects.create(repair_order=order, equipment=equipment)

        self.assertEqual(order.assign_equipment_owners(), 0)
        equipment.refresh_from_db()
        self.assertEqual(equipment.current_client, self.other)

    def test_an_order_without_a_client_touches_nothing(self):
        """Заказчик у заказа обязателен, так что в жизни этого не случается.
        Но подстраховка на месте: без заказчика проставлять нечего,
        и в базу за этим ходить незачем."""
        self.assertEqual(RepairOrder(client=None).assign_equipment_owners(), 0)

    def test_the_whole_way_through_the_intake_form(self):
        """Тот же путь, которым пользуются: приём заказа целиком."""
        equipment = Equipment.objects.create(model=self.model, serial_number='SN-OWN-6')
        response = self.http.post('/repair-orders/create/', {
            'client': self.client_obj.pk,
            'fault_description': '',
            'equipments-TOTAL_FORMS': '1',
            'equipments-INITIAL_FORMS': '0',
            'equipments-MIN_NUM_FORMS': '0',
            'equipments-MAX_NUM_FORMS': '1000',
            'equipments-0-equipment': equipment.pk,
            'equipments-0-fault_description': 'не открывает двери',
            'equipments-0-initial_condition': '',
        })
        self.assertEqual(response.status_code, 302)
        equipment.refresh_from_db()
        self.assertEqual(equipment.current_client, self.client_obj)


class MobileMenuSingleColumnTests(SimpleTestCase):
    """Меню — один столбец, прокрутка вверх-вниз.

    Регрессия v2.56.1: `.nav` от Bootstrap даёт flex-wrap: wrap, класс
    `flex-column` — направление в столбец, а прокрутка тогда же ограничила
    меню по высоте. Колонка плюс перенос плюс предел по высоте = нижние
    разделы уезжают во второй столбец, и меню двигается вбок.
    """

    def setUp(self):
        self.base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')
        self.nav = re.search(r'\.sidebar-nav \{([^}]*)\}', self.base).group(1)

    def test_sections_never_wrap_into_a_second_column(self):
        self.assertIn('flex-wrap: nowrap', self.nav)

    def test_the_column_is_our_own_rule_not_bootstrap(self):
        """Bootstrap приходит из интернета: без него `flex-column`
        не значит ничего, и меню рассыпалось бы в строку."""
        self.assertIn('flex-direction: column', self.nav)
        self.assertIn('display: flex', self.nav)

    def test_there_is_room_below_for_the_address_bar(self):
        """На телефоне снизу всплывает строка адреса и накрывает
        последний раздел."""
        gap = int(re.search(r'padding-bottom: (\d+)px', self.nav).group(1))
        self.assertGreaterEqual(gap, 20)

    def test_the_list_looks_like_a_list_without_bootstrap(self):
        self.assertIn('list-style: none', self.nav)
        self.assertIn('padding-left: 0', self.nav)


class BrowserTabIconTests(SimpleTestCase):
    """Значок вкладки — логотип компании."""

    def setUp(self):
        self.base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')

    def test_the_tab_shows_the_logo(self):
        self.assertIn("img/favicon-32.png", self.base)
        self.assertIn('rel="icon"', self.base)

    def test_the_phone_home_screen_gets_a_proper_icon(self):
        self.assertIn('rel="apple-touch-icon"', self.base)
        self.assertIn("img/favicon-180.png", self.base)

    def test_the_browser_bar_takes_the_brand_colour(self):
        self.assertIn('name="theme-color"', self.base)

    def test_the_icons_exist_and_are_small(self):
        """Уменьшены заранее: сам логотип 1168x1168 и весит 28 КБ —
        возить его на каждой странице ради значка незачем."""
        for name, limit in (('favicon-32.png', 10_000), ('favicon-180.png', 60_000)):
            with self.subTest(файл=name):
                path = settings.BASE_DIR / 'core/static/img' / name
                self.assertTrue(path.exists(), '%s не найден' % name)
                self.assertLess(path.stat().st_size, limit)

    def test_the_original_logo_is_not_used_as_the_icon(self):
        icons = re.findall(r'<link rel="(?:icon|apple-touch-icon)"[^>]*>', self.base)
        self.assertTrue(icons)
        for tag in icons:
            self.assertNotIn('lift_team_logo.png', tag)


class BusyEquipmentTests(TestCase):
    """Прибор, лежащий в другом незакрытом заказе, к приёму не предлагается.

    Одна и та же железка не может одновременно стоять на двух верстаках.
    А вот отремонтированный и отгруженный в списке остаётся: он у заказчика
    и приехать снова может — ради этого история ремонтов и ведётся
    по единице.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_busy', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.client_obj = ClientModel.objects.create(name='МУП «Лифты»')
        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.busy = Equipment.objects.create(model=self.model, serial_number='SN-BUSY')
        self.free = Equipment.objects.create(model=self.model, serial_number='SN-FREE')
        self.returned = Equipment.objects.create(model=self.model, serial_number='SN-BACK')

        self.other_order = RepairOrder.objects.create(
            client=self.client_obj, status='repair'
        )
        RepairOrderEquipment.objects.create(
            repair_order=self.other_order, equipment=self.busy
        )
        shipped = RepairOrder.objects.create(client=self.client_obj, status='shipped')
        RepairOrderEquipment.objects.create(repair_order=shipped, equipment=self.returned)

        self.order = RepairOrder.objects.create(client=self.client_obj)

    def test_equipment_in_an_open_order_is_not_offered(self):
        available = list(available_equipment_for_order())

        self.assertIn(self.free, available)
        self.assertNotIn(self.busy, available)

    def test_a_shipped_order_releases_the_equipment(self):
        """Прибор вернулся к заказчику и может приехать снова — иначе
        историю ремонтов по единице вести было бы не на чем."""
        self.assertIn(self.returned, available_equipment_for_order())

    def test_the_order_being_edited_keeps_its_own_equipment(self):
        """Иначе форма правки потеряла бы то, что в ней уже выбрано."""
        RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.free
        )
        self.order.status = 'repair'
        self.order.save(update_fields=['status'])

        available = list(available_equipment_for_order(self.order))

        self.assertIn(self.free, available)
        self.assertNotIn(self.busy, available)

    def test_the_order_form_offers_only_free_equipment(self):
        page = self.http.get('/repair-orders/create/')
        choices = list(page.context['formset'].empty_form.fields['equipment'].queryset)

        self.assertIn(self.free, choices)
        self.assertNotIn(self.busy, choices)

    def test_a_units_own_order_does_not_make_it_busy(self):
        """`exclude_order` для того и нужен: без него собственные единицы
        заказа считались бы занятыми в нём же."""
        RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.free
        )
        self.order.status = 'diagnostic'
        self.order.save(update_fields=['status'])

        self.assertIn(self.free, available_equipment_for_order(self.order))
        self.assertNotIn(self.free, available_equipment_for_order(None))

    def test_the_order_card_offers_only_free_equipment(self):
        page = self.http.get('/repair-orders/%d/' % self.order.pk)
        available = list(page.context['available_equipment'])

        self.assertIn(self.free, available)
        self.assertNotIn(self.busy, available)

    def test_accepting_a_busy_unit_is_refused_and_names_the_order(self):
        """Без названия заказа непонятно, где прибор искать."""
        response = self.http.post('/repair-orders/%d/add-unit/' % self.order.pk, {
            'equipment': self.busy.pk,
        })

        self.assertEqual(self.order.order_equipments.count(), 0)
        messages = [str(m) for m in response.wsgi_request._messages]
        self.assertTrue(
            any(self.other_order.order_number in m for m in messages), messages
        )


class VersionDesignationJoinTests(TestCase):
    """Обозначение исполнения складывается без своего разделителя.

    Разделитель хранится внутри самого обозначения («.4», «-1.1»), потому
    что на изделиях он произвольный. Пока дефис был прошит в сборке, к нему
    добавлялся ещё один из обозначения — в списке выбора выходило
    «БУАД-7-31-.4».
    """

    def setUp(self):
        self.model = EquipmentModel.objects.create(name='БУАД-7-31')

    def test_a_dash_is_not_added_in_front(self):
        version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )

        self.assertEqual(str(version), 'БУАД-7-31.4')

    def test_the_separator_from_the_directory_is_kept_as_is(self):
        version = EquipmentVersion.objects.create(
            equipment_model=EquipmentModel.objects.create(name='EcoDrive-2.3'),
            name='-1.1',
        )

        self.assertEqual(str(version), 'EcoDrive-2.3-1.1')

    def test_the_list_and_the_documents_agree(self):
        """В списке выбора и в акте должно стоять одно и то же обозначение."""
        version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-JOIN-1', version=version
        )

        self.assertEqual(str(version), equipment.designation)


class EquipmentVersionOnQuickCreateTests(TestCase):
    """Исполнение указывается сразу при заведении оборудования из заказа.

    Раньше окно спрашивало только модель и серийный номер, и обозначение
    приходилось дописывать потом в карточке — а печатается оно в актах
    и на наклейке, то есть уезжает к заказчику неполным.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_qc_version', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        self.other_model = EquipmentModel.objects.create(name='EcoDrive-2.3')
        self.alien = EquipmentVersion.objects.create(
            equipment_model=self.other_model, name='-1.1'
        )

    def _create(self, **extra):
        data = {'model_id': self.model.pk, 'serial_number': 'SN-QC-1', 'confirmed': '1'}
        data.update(extra)
        return self.http.post('/ajax/equipment/create/', data)

    def test_the_model_list_carries_the_versions(self):
        """Исполнения приезжают вместе с моделями: лишний поход на сервер
        при каждом выборе модели — это ещё одна вещь, которая не придёт,
        когда со связью плохо."""
        data = self.http.get('/ajax/equipment-model/list/').json()

        by_name = {m['name']: m for m in data['models']}
        self.assertEqual([v['name'] for v in by_name['БУАД-7-31']['versions']], ['.4'])
        self.assertEqual([v['name'] for v in by_name['EcoDrive-2.3']['versions']], ['-1.1'])

    def test_the_version_is_saved_with_the_equipment(self):
        response = self._create(version_id=self.version.pk)

        self.assertTrue(response.json()['success'])
        equipment = Equipment.objects.get(serial_number='SN-QC-1')
        self.assertEqual(equipment.version, self.version)
        self.assertEqual(equipment.designation, 'БУАД-7-31.4')

    def test_without_a_version_nothing_is_invented(self):
        self._create()

        self.assertIsNone(Equipment.objects.get(serial_number='SN-QC-1').version)

    def test_a_version_of_another_model_is_refused(self):
        """Иначе у прибора оказалось бы обозначение от чужой модели."""
        self._create(version_id=self.alien.pk)

        self.assertIsNone(Equipment.objects.get(serial_number='SN-QC-1').version)

    def test_the_modal_asks_for_it(self):
        """Окно заведения оборудования осталось на приёме заказа —
        страницы правки заказа больше нет."""
        page = self.http.get('/repair-orders/create/').content.decode()

        self.assertIn('id="ceVersionSelect"', page)
        # прячется атрибутом, а не классом Bootstrap: он приходит из интернета
        self.assertIn('id="ceVersionRow" hidden', page)
        self.assertIn('row.hidden = versions.length === 0', page)


class ListFiltersSurviveBackTests(TestCase):
    """«Назад» с карточки возвращает к списку вместе с отбором.

    «Назад» в программе — обычная ссылка, а не кнопка браузера: она вела
    на список без параметров, и отбор терялся. Искали «BAV», открыли
    деталь, вернулись — и снова весь каталог, ищи заново.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_back', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)
        self.part = SparePart.objects.create(part_number='BAV70', name='Диод')
        self.equipment = Equipment.objects.create(
            model=EquipmentModel.objects.create(name='БУАД-7-31'),
            serial_number='SN-BACK-1',
        )

    def test_the_part_card_remembers_the_search(self):
        self.http.get('/parts/?search=BAV')

        page = self.http.get('/parts/%d/' % self.part.pk)

        self.assertEqual(page.context['back_url'], '/parts/?search=BAV')

    def test_the_equipment_card_remembers_the_search(self):
        self.http.get('/equipment/?search=%D0%91%D0%A3%D0%90%D0%94')

        page = self.http.get('/equipment/%d/edit/' % self.equipment.pk)

        self.assertIn('search=', page.context['back_url'])

    def test_the_repair_history_remembers_it_too(self):
        """На историю попадают и сканером, и из заказа — «Назад» обязан
        вести туда же, куда и обычно."""
        self.http.get('/equipment/?search=SN-BACK')

        page = self.http.get('/equipment/%d/history/' % self.equipment.pk)

        self.assertIn('search=SN-BACK', page.context['back_url'])

    def test_without_a_search_the_link_is_plain(self):
        self.http.get('/parts/')

        page = self.http.get('/parts/%d/' % self.part.pk)

        self.assertEqual(page.context['back_url'], '/parts/')

    def test_the_card_opened_first_still_has_a_way_back(self):
        """На карточку попадают и по ссылке из письма, и сканером —
        списка до этого могло не быть вовсе."""
        page = self.http.get('/parts/%d/' % self.part.pk)

        self.assertEqual(page.context['back_url'], '/parts/')
        self.assertContains(page, 'Назад')


class PlannedPartTests(TestCase):
    """Отложенное списание: деталь нужна, но со склада ещё не взята.

    Так записывают то, чего на полке не оказалось, и то, что мастер наметил,
    ещё не вскрыв прибор. Пока деталь не списана, остаток она не трогает
    и заказу ничего не стоит.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='admin_planned', full_name='Мастер', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.part = SparePart.objects.create(
            part_number='R-10', name='Резистор', current_stock=0,
            price=Decimal('15.00')
        )
        StockMovement.objects.create(
            part=self.part, quantity=10, movement_type='incoming',
            unit_price=Decimal('12.00')
        )
        self.part.current_stock = 10
        self.part.save(update_fields=['current_stock'])

        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.unit = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=EquipmentModel.objects.create(name='БУАД-7-31'),
                serial_number='SN-PLAN-1',
            ),
        )

    def _plan(self, quantity=3):
        return self.http.post('/repair-orders/%d/add-detail/' % self.order.pk, {
            'part': self.part.pk, 'quantity_used': quantity, 'plan': '1',
        })

    def test_planning_does_not_touch_the_stock(self):
        self._plan()

        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 10)
        self.assertFalse(StockMovement.objects.filter(movement_type='outgoing').exists())
        detail = self.order.details.get()
        self.assertTrue(detail.is_planned)

    def test_a_planned_part_costs_the_order_nothing_yet(self):
        """Ноль тут был бы не лучше: он попал бы в сумму как настоящая
        цифра, хотя со склада деталь не брали."""
        self._plan()

        detail = self.order.details.get()
        self.assertIsNone(detail.cost)
        self.assertFalse(self.order.costs.exists())

    def test_planning_keeps_the_unit_it_belongs_to(self):
        """В заказе одна единица — привязка очевидна и проставляется сама,
        как и при обычном списании."""
        self._plan()

        self.assertEqual(self.order.details.get().order_equipment, self.unit)

    def test_writing_it_off_goes_the_usual_way(self):
        """Тот же путь, что у обычного списания: партии, затрата, история.
        Второго кода для этого заводить нельзя — однажды разойдётся."""
        self._plan()
        detail = self.order.details.get()

        self.http.post('/repair-orders/%d/details/%d/write-off/' % (
            self.order.pk, detail.pk))

        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 7)
        written = self.order.details.get()
        self.assertFalse(written.is_planned)
        self.assertEqual(written.quantity_used, 3)
        self.assertEqual(written.order_equipment, self.unit)
        self.assertIsNotNone(written.movement)
        self.assertEqual(self.order.costs.get().amount, Decimal('36.00'))

    def test_the_written_off_line_replaces_the_planned_one(self):
        """Не вторая строка рядом с первой: деталь одна."""
        self._plan()
        detail = self.order.details.get()

        self.http.post('/repair-orders/%d/details/%d/write-off/' % (
            self.order.pk, detail.pk))

        self.assertEqual(self.order.details.count(), 1)

    def test_cancelling_leaves_no_trace_on_the_stock(self):
        """Со склада деталь не брали — и возвращать нечего."""
        self._plan()
        detail = self.order.details.get()

        self.http.post('/repair-orders/%d/details/%d/cancel/' % (
            self.order.pk, detail.pk))

        self.assertFalse(self.order.details.exists())
        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 10)
        self.assertFalse(StockMovement.objects.filter(movement_type='outgoing').exists())

    def test_a_planned_part_is_not_returnable(self):
        self._plan()
        detail = self.order.details.get()

        self.assertFalse(detail.returnable)
        response = self.http.post(
            '/repair-orders/%d/details/%d/return/' % (self.order.pk, detail.pk),
            {'quantity': '1'}
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(self.order.details.get().quantity_used, 3)

    def test_a_written_off_part_is_not_written_off_twice(self):
        self.http.post('/repair-orders/%d/add-detail/' % self.order.pk, {
            'part': self.part.pk, 'quantity_used': 2,
        })
        detail = self.order.details.get()

        response = self.http.post('/repair-orders/%d/details/%d/write-off/' % (
            self.order.pk, detail.pk))

        self.assertEqual(response.status_code, 404)
        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 8)

    def test_planning_more_than_the_shelf_holds_is_allowed(self):
        """Именно для этого случая отложенное списание и нужно: детали нет,
        а записать её надо, чтобы не забыть заказать."""
        self._plan(25)

        detail = self.order.details.get()
        self.assertTrue(detail.is_planned)
        self.assertFalse(detail.enough_in_stock)
        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 10)

    def test_the_card_shows_the_planned_line_apart(self):
        self._plan()
        page = self.http.get('/repair-orders/%d/' % self.order.pk)

        self.assertContains(page, 'запланирована')
        self.assertContains(page, 'не списана')
        self.assertEqual(page.context['details_cost'], 0)

    def test_the_card_offers_planning_next_to_adding(self):
        page = self.http.get('/repair-orders/%d/' % self.order.pk)

        self.assertContains(page, 'name="plan"')


class ReturnPartToBatchTests(TestCase):
    """Возврат детали из заказа — в те самые партии, из которых её брали.

    Вернуть «куда-нибудь» нельзя: себестоимость считается по партиям (FIFO),
    и неверный возврат разъедет её так, что заметят через полгода в отчёте
    о прибыли. Поэтому партии разбираются в обратном порядке — сначала та,
    из которой брали последней: это отменяет списание ровно наоборот тому,
    как оно делалось, и сохраняет свойство FIFO, что старые партии
    израсходованы раньше новых.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='admin_return', full_name='Мастер', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.part = SparePart.objects.create(
            part_number='C-220', name='Конденсатор', current_stock=0
        )
        # Две партии по разной цене: без этого возврат «не в ту» партию
        # ничем не отличался бы от возврата в правильную
        self.old_batch = StockMovement.objects.create(
            part=self.part, quantity=3, movement_type='incoming',
            unit_price=Decimal('10.00')
        )
        self.new_batch = StockMovement.objects.create(
            part=self.part, quantity=5, movement_type='incoming',
            unit_price=Decimal('20.00')
        )
        self.part.current_stock = 8
        self.part.save(update_fields=['current_stock'])

        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )

    def _write_off(self, quantity):
        views._use_repair_order_part(
            self.order, self.part, quantity, self.employee, 'списание'
        )
        self.part.refresh_from_db()
        return self.order.details.get()

    def _batches(self):
        self.old_batch.refresh_from_db()
        self.new_batch.refresh_from_db()
        return self.old_batch.remaining_in_batch, self.new_batch.remaining_in_batch

    def _order_cost(self):
        amounts = list(self.order.costs.values_list('amount', flat=True))
        if any(a is None for a in amounts):
            return None
        return sum(amounts)

    def test_the_newest_batch_is_unwound_first(self):
        """Списали 5: три из старой партии и две из новой. Возвращаем две —
        они обязаны вернуться в новую, иначе очередь FIFO перевернётся."""
        detail = self._write_off(5)
        self.assertEqual(self._batches(), (0, 3))

        detail.return_to_stock(2, self.employee)

        self.assertEqual(self._batches(), (0, 5))

    def test_the_stock_and_the_order_agree_after_a_partial_return(self):
        detail = self._write_off(5)

        detail.return_to_stock(2, self.employee)

        self.part.refresh_from_db()
        detail.refresh_from_db()
        self.assertEqual(self.part.current_stock, 5)
        self.assertEqual(detail.quantity_used, 3)

    def test_the_cost_of_the_order_drops_by_exactly_what_came_back(self):
        """3 по 10 плюс 2 по 20 — это 70. Вернули две по 20, осталось 30,
        то есть ровно три штуки из старой партии."""
        detail = self._write_off(5)
        self.assertEqual(self._order_cost(), Decimal('70.00'))

        detail.return_to_stock(2, self.employee)

        self.assertEqual(self._order_cost(), Decimal('30.00'))

    def test_the_correction_is_a_new_record_not_an_edit(self):
        """Правка прошлой записи стёрла бы след, а по затратам считают
        прибыль."""
        detail = self._write_off(5)
        detail.return_to_stock(2, self.employee)

        amounts = sorted(self.order.costs.values_list('amount', flat=True))
        self.assertEqual(amounts, [Decimal('-40.00'), Decimal('70.00')])

    def test_returning_everything_removes_the_line_but_keeps_the_journal(self):
        """Со склада деталь уходила и возвращалась — в журнале это видно,
        даже когда в заказе её больше нет."""
        detail = self._write_off(5)

        detail.return_to_stock(5, self.employee)

        self.assertFalse(self.order.details.exists())
        self.assertEqual(self._batches(), (3, 5))
        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 8)
        types = sorted(StockMovement.objects.filter(
            repair_order=self.order).values_list('movement_type', flat=True))
        self.assertEqual(types, ['outgoing', 'return'])

    def test_a_return_is_not_a_new_batch(self):
        """Приход — это партия, из которой потом списывают. Возвращённая
        деталь в новую партию не превращается: она вернулась в свою."""
        detail = self._write_off(5)
        detail.return_to_stock(2, self.employee)

        batches = StockMovement.objects.filter(
            part=self.part, movement_type='incoming'
        )
        self.assertEqual(batches.count(), 2)
        returned = StockMovement.objects.get(movement_type='return')
        self.assertIsNone(returned.remaining_in_batch)

    def test_more_than_was_written_off_is_refused(self):
        detail = self._write_off(3)

        with self.assertRaises(ValidationError):
            detail.return_to_stock(4, self.employee)

        detail.refresh_from_db()
        self.assertEqual(detail.quantity_used, 3)
        self.assertEqual(self._batches(), (0, 5))

    def test_nothing_at_all_is_refused(self):
        detail = self._write_off(3)

        with self.assertRaises(ValidationError):
            detail.return_to_stock(0, self.employee)

    def test_an_old_write_off_cannot_be_returned_to_its_batch(self):
        """У списаний до появления возвратов связи с расходом нет, и
        в какие партии возвращать — неизвестно. Догадка тут хуже отказа."""
        detail = self._write_off(3)
        detail.movement = None
        detail.save(update_fields=['movement'])

        self.assertFalse(detail.returnable)
        with self.assertRaises(ValidationError):
            detail.return_to_stock(1, self.employee)

    def test_a_part_written_off_into_the_red_returns_without_a_known_cost(self):
        """Часть списывалась без партии — её себестоимость и тогда была
        неизвестна. Ноль вместо неизвестного соврал бы."""
        detail = self._write_off(10)          # партий всего на 8
        self.assertIsNone(self._order_cost())

        detail.return_to_stock(10, self.employee)

        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 8)
        self.assertEqual(self._batches(), (3, 5))
        self.assertIsNone(self.order.costs.order_by('-pk').first().amount)

    def test_a_batch_without_a_price_makes_the_whole_return_unknown(self):
        StockMovement.objects.filter(pk=self.new_batch.pk).update(unit_price=None)
        detail = self._write_off(5)

        detail.return_to_stock(2, self.employee)

        self.assertIsNone(self.order.costs.order_by('-pk').first().amount)

    def test_the_page_returns_the_part(self):
        detail = self._write_off(5)

        response = self.http.post(
            '/repair-orders/%d/details/%d/return/' % (self.order.pk, detail.pk),
            {'quantity': '2'}
        )

        self.assertRedirects(response, '/repair-orders/%d/' % self.order.pk)
        detail.refresh_from_db()
        self.assertEqual(detail.quantity_used, 3)

    def test_a_detail_from_another_order_is_not_returned(self):
        """Иначе подставленный номер вернул бы чужую деталь."""
        detail = self._write_off(5)
        other = RepairOrder.objects.create(client=self.order.client)

        response = self.http.post(
            '/repair-orders/%d/details/%d/return/' % (other.pk, detail.pk),
            {'quantity': '2'}
        )

        self.assertEqual(response.status_code, 404)
        detail.refresh_from_db()
        self.assertEqual(detail.quantity_used, 5)

    def test_returning_is_not_a_get(self):
        detail = self._write_off(5)

        response = self.http.get(
            '/repair-orders/%d/details/%d/return/' % (self.order.pk, detail.pk)
        )

        self.assertEqual(response.status_code, 405)
        detail.refresh_from_db()
        self.assertEqual(detail.quantity_used, 5)


class PartsPerUnitTests(TestCase):
    """Деталь списывается в конкретную железку, а не «в заказ вообще».

    В заказе с одним прибором привязка очевидна и проставляется сама;
    с несколькими программа не угадывает — деталь остаётся общей, пока
    мастер не укажет. Так решил владелец: неверная привязка хуже пустой,
    потому что по ней потом считают, во сколько обошёлся ремонт.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_parts_unit', full_name='Мастер', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.first = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=self.model, serial_number='SN-U1'),
        )
        self.part = SparePart.objects.create(
            part_number='C-100', name='Конденсатор', current_stock=50
        )

    def _second_unit(self):
        return RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(model=self.model, serial_number='SN-U2'),
        )

    def _add(self, **extra):
        data = {'part': self.part.pk, 'quantity_used': 2}
        data.update(extra)
        return self.http.post('/repair-orders/%d/add-detail/' % self.order.pk, data)

    def test_one_unit_in_the_order_gets_the_part_by_itself(self):
        self._add()

        detail = self.order.details.get()
        self.assertEqual(detail.order_equipment, self.first)

    def test_several_units_leave_the_part_on_the_order(self):
        """Программа не угадывает, в который из пяти приборов поставили
        конденсатор."""
        self._second_unit()

        self._add()

        detail = self.order.details.get()
        self.assertIsNone(detail.order_equipment)

    def test_the_master_can_name_the_unit(self):
        second = self._second_unit()

        self._add(order_equipment=second.pk)

        detail = self.order.details.get()
        self.assertEqual(detail.order_equipment, second)

    def test_a_unit_from_another_order_is_not_accepted(self):
        """Иначе подставленный номер увёл бы деталь в чужой ремонт."""
        other_order = RepairOrder.objects.create(client=self.order.client)
        alien = RepairOrderEquipment.objects.create(
            repair_order=other_order,
            equipment=Equipment.objects.create(model=self.model, serial_number='SN-ALIEN'),
        )
        self._second_unit()

        self._add(order_equipment=alien.pk)

        detail = self.order.details.get()
        self.assertIsNone(detail.order_equipment)

    def test_the_stock_is_written_off_the_same_way_as_before(self):
        """Привязка — это только запись о том, куда ушла деталь. Склад,
        партии и себестоимость считаются как считались."""
        self._add()

        self.part.refresh_from_db()
        self.assertEqual(self.part.current_stock, 48)
        movement = StockMovement.objects.get(part=self.part, movement_type='outgoing')
        self.assertEqual(movement.repair_order, self.order)

    def test_removing_a_unit_keeps_the_write_off(self):
        """Деталь со склада списана — терять запись о ней нельзя,
        даже если единицу убрали из заказа."""
        self._add()
        self.first.delete()

        detail = self.order.details.get()
        self.assertIsNone(detail.order_equipment)
        self.assertEqual(detail.quantity_used, 2)

    def test_the_recipe_puts_parts_into_the_unit_it_came_from(self):
        """Кнопку «применить шаблон» нажимают в строке единицы — значит
        и детали её."""
        second = self._second_unit()
        fault = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы'
        )
        FaultTypePart.objects.create(fault_type=fault, part=self.part, quantity=3)

        self.http.post('/repair-orders/%d/apply-fault-template/' % self.order.pk, {
            'fault_ids': [fault.pk], 'equipment_id': second.equipment_id,
        })

        detail = self.order.details.get()
        self.assertEqual(detail.order_equipment, second)
        self.assertEqual(detail.quantity_used, 3)

    def test_the_card_asks_only_when_there_is_a_choice(self):
        # Ищем сам список, а не имя поля где угодно: имя встречается
        # и в скрипте кнопки «списать деталь на эту единицу»
        page = self.http.get('/repair-orders/%d/' % self.order.pk)
        self.assertNotContains(page, '<select name="order_equipment"')

        self._second_unit()
        page = self.http.get('/repair-orders/%d/' % self.order.pk)
        self.assertContains(page, '<select name="order_equipment"')
        self.assertContains(page, 'На заказ целиком')

    def test_old_write_offs_read_as_on_the_order(self):
        """Списания до появления привязки лежат пустыми, и это честно:
        выдумывать за них нельзя."""
        self._second_unit()
        self._add()

        page = self.http.get('/repair-orders/%d/' % self.order.pk)

        self.assertContains(page, 'на заказ целиком')


class PriceListTests(TestCase):
    """Прайсы: базовый и по заказчикам.

    Базовый отвечает на вопрос «сколько это стоит вообще», прайс заказчика
    уточняет его там, где договорились иначе. Внутри прайса строка
    со сложностью вытесняет строку без неё — тот же приём, что и у рецепта
    деталей: уточнение бьёт общее, но только там, где заведено.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_prices', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.drive = EquipmentType.objects.create(name='Привод дверей')
        self.vfd = EquipmentType.objects.create(name='Преобразователь частоты')
        self.client_obj = ClientModel.objects.create(name='МУП «Лифты»')
        self.base = PriceList.base()
        PriceListLine.objects.create(
            price_list=self.base, equipment_type=self.drive, price=Decimal('8000')
        )
        PriceListLine.objects.create(
            price_list=self.base, equipment_type=self.drive,
            complexity='complex', price=Decimal('15000')
        )

    def test_the_base_price_answers_when_the_client_has_no_list(self):
        line = PriceList.line_for(self.client_obj, self.drive, '')
        self.assertEqual(line.price, Decimal('8000'))
        self.assertTrue(line.price_list.is_base)

    def test_complexity_beats_the_general_row(self):
        line = PriceList.line_for(None, self.drive, 'complex')
        self.assertEqual(line.price, Decimal('15000'))

    def test_the_clients_own_list_beats_the_base_one(self):
        own = PriceList.objects.create(client=self.client_obj)
        PriceListLine.objects.create(
            price_list=own, equipment_type=self.drive,
            complexity='complex', price=Decimal('12000')
        )

        line = PriceList.line_for(self.client_obj, self.drive, 'complex')

        self.assertEqual(line.price, Decimal('12000'))
        self.assertEqual(line.price_list, own)

    def test_the_client_list_need_not_repeat_the_base_one(self):
        """Заказчику завели цену только на сложный ремонт. На простой она
        должна прийти из базового, а не пропасть."""
        own = PriceList.objects.create(client=self.client_obj)
        PriceListLine.objects.create(
            price_list=own, equipment_type=self.drive,
            complexity='complex', price=Decimal('12000')
        )

        line = PriceList.line_for(self.client_obj, self.drive, 'simple')

        self.assertEqual(line.price, Decimal('8000'))
        self.assertTrue(line.price_list.is_base)

    def test_a_type_without_a_price_gives_nothing(self):
        """Пусто — значит пусто: выдумывать цену нельзя, по ней
        разговаривают с заказчиком."""
        self.assertIsNone(PriceList.line_for(None, self.vfd, ''))

    def test_without_an_equipment_type_there_is_nothing_to_look_up(self):
        self.assertIsNone(PriceList.line_for(None, None, 'complex'))

    def test_the_base_price_list_is_the_only_one(self):
        """Второй базовый превратил бы поиск цены в лотерею: какой из двух
        подхватится, зависело бы от порядка записей."""
        with self.assertRaises(ValidationError):
            PriceList.objects.create()

    def test_the_base_price_list_appears_when_first_needed(self):
        PriceList.objects.all().delete()

        created = PriceList.base()

        self.assertTrue(created.is_base)
        self.assertEqual(PriceList.objects.count(), 1)
        self.assertEqual(PriceList.base(), created)


class PriceInDefectActTests(TestCase):
    """Цена предлагается мастеру после диагностики и замораживается в заказе."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_price_act', full_name='Мастер', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.type = EquipmentType.objects.create(name='Привод дверей')
        self.model = EquipmentModel.objects.create(
            name='БУАД-7-31', equipment_type=self.type
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-PRICE-1'
            ),
        )
        base = PriceList.base()
        PriceListLine.objects.create(
            price_list=base, equipment_type=self.type, price=Decimal('8000')
        )

    def _act_url(self):
        return '/repair-orders/%d/equipment/%d/defect/' % (self.order.pk, self.roe.pk)

    def _save(self, cost):
        return self.http.post(self._act_url(), {
            'defect_act_date': '', 'diagnosis': 'вздулись конденсаторы',
            'error_codes': '', 'non_warranty_reason': '', 'estimated_cost': cost,
        })

    def test_the_price_is_offered_not_filled_in(self):
        """Акт подписывает мастер — вписать за него программа не вправе."""
        page = self.http.get(self._act_url())

        self.assertEqual(page.context['price_line'].price, Decimal('8000'))
        self.assertContains(page, 'По прайсу')
        self.assertContains(page, 'Подставить')
        # но в поле по-прежнему пусто
        self.assertIsNone(page.context['form'].initial.get('estimated_cost'))

    def test_saving_freezes_the_price_of_the_day(self):
        self._save('9500')

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.estimated_cost, Decimal('9500'))
        self.assertEqual(self.roe.list_price, Decimal('8000'))
        self.assertTrue(self.roe.price_differs_from_list)

    def test_a_later_change_of_the_price_list_leaves_the_order_alone(self):
        """Иначе правка прайса задним числом меняла бы то, о чём
        договорились."""
        self._save('8000')
        PriceListLine.objects.update(price=Decimal('11000'))

        self._save('8000')

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.list_price, Decimal('8000'))
        self.assertFalse(self.roe.price_differs_from_list)

    def test_no_price_in_the_list_freezes_nothing(self):
        """Незаполненное поле честнее выдуманного нуля."""
        PriceListLine.objects.all().delete()

        self._save('7000')

        self.roe.refresh_from_db()
        self.assertIsNone(self.roe.list_price)
        self.assertFalse(self.roe.price_differs_from_list)

    def test_the_page_says_when_there_is_no_price_for_this_type(self):
        PriceListLine.objects.all().delete()
        page = self.http.get(self._act_url())

        self.assertIsNone(page.context['price_line'])
        self.assertContains(page, 'В прайсе нет цены')

    def test_a_model_without_a_type_says_so(self):
        self.model.equipment_type = None
        self.model.save()

        page = self.http.get(self._act_url())

        self.assertIsNone(page.context['price_line'])
        self.assertContains(page, 'не указан тип оборудования')


class PriceListPagesTests(TestCase):
    """Страницы прайсов: список, правка, копирование."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_price_pages', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)
        self.type = EquipmentType.objects.create(name='Привод дверей')
        self.first = ClientModel.objects.create(name='МУП «Лифты»')
        self.second = ClientModel.objects.create(name='ООО «Подъём»')
        self.own = PriceList.objects.create(client=self.first)
        PriceListLine.objects.create(
            price_list=self.own, equipment_type=self.type,
            complexity='complex', price=Decimal('12000')
        )

    def test_the_index_lists_the_base_and_the_client_lists(self):
        base = PriceList.base()
        page = self.http.get('/price-lists/')

        self.assertEqual(page.context['base'], base)
        self.assertIn(self.own, page.context['client_lists'])

    def test_copying_fills_the_form_and_saves_nothing_yet(self):
        """Тот же приём, что у типовых неисправностей и карточек деталей:
        до нажатия «Сохранить» в базе не появляется ничего."""
        before = PriceList.objects.count()
        page = self.http.get('/price-lists/create/?copy_from=%d' % self.own.pk)

        self.assertEqual(PriceList.objects.count(), before)
        self.assertEqual(page.context['copy_source'], self.own)
        # Набор форм модели кладёт переданное в initial_extra: initial
        # у него занят данными уже сохранённых строк, а их тут нет
        initial = page.context['formset'].initial_extra
        self.assertEqual(len(initial), 1)
        self.assertEqual(initial[0]['price'], Decimal('12000'))

    def test_a_client_can_have_only_one_price_list(self):
        """Второй прайс у того же заказчика — это две цены на одно и то же."""
        page = self.http.get('/price-lists/create/')
        offered = list(page.context['form'].fields['client'].queryset)

        self.assertIn(self.second, offered)
        self.assertNotIn(self.first, offered)

    def test_the_client_card_leads_to_the_price_list(self):
        page = self.http.get('/clients/%d/edit/' % self.first.pk)

        self.assertContains(page, '/price-lists/%d/edit/' % self.own.pk)

    def test_the_client_card_offers_to_start_one(self):
        page = self.http.get('/clients/%d/edit/' % self.second.pk)

        self.assertContains(page, 'Завести прайс')

    def test_deleting_a_price_list_leaves_agreed_prices_alone(self):
        """Цена заморожена в заказе, а не берётся из прайса каждый раз."""
        model = EquipmentModel.objects.create(name='БУАД-7-31', equipment_type=self.type)
        order = RepairOrder.objects.create(client=self.first)
        roe = RepairOrderEquipment.objects.create(
            repair_order=order,
            equipment=Equipment.objects.create(model=model, serial_number='SN-DEL-1'),
            estimated_cost=Decimal('12000'), list_price=Decimal('12000'),
        )

        self.http.post('/price-lists/%d/delete/' % self.own.pk)

        roe.refresh_from_db()
        self.assertEqual(roe.list_price, Decimal('12000'))
        self.assertEqual(roe.estimated_cost, Decimal('12000'))


class AddUnitToOrderTests(TestCase):
    """Приём по одной единице с печатью наклейки сразу.

    Так это и происходит у стола: коробку поставили, описали, наклеили
    ярлык, взяли следующую. Пока заказ сохранялся целиком, ярлыки
    печатались пачкой в конце, и их раскладывали по коробкам, сверяя
    номер позиции.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_add_unit', full_name='Приёмщик', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.client_obj = ClientModel.objects.create(name='МУП «Лифты»')
        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.first = Equipment.objects.create(model=self.model, serial_number='SN-ADD-1')
        self.second = Equipment.objects.create(model=self.model, serial_number='SN-ADD-2')
        self.order = RepairOrder.objects.create(client=self.client_obj)
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.first
        )

    def _add(self, equipment, **extra):
        data = {'equipment': equipment.pk if equipment else ''}
        data.update(extra)
        return self.http.post('/repair-orders/%d/add-unit/' % self.order.pk, data)

    def test_the_unit_is_added_and_the_label_opens(self):
        """Переход на наклейку, а не открытие окна из скрипта: окно,
        открытое не по щелчку человека, браузер часто не показывает."""
        response = self._add(self.second, fault_description='не открывает двери',
                             initial_condition='корпус целый')

        added = self.order.order_equipments.get(equipment=self.second)
        self.assertRedirects(
            response,
            '/repair-orders/%d/equipment/%d/label/' % (self.order.pk, added.pk)
        )
        self.assertEqual(added.fault_description, 'не открывает двери')
        self.assertEqual(added.initial_condition, 'корпус целый')

    def test_the_owner_is_filled_in_on_the_way(self):
        self._add(self.second)

        self.second.refresh_from_db()
        self.assertEqual(self.second.current_client, self.client_obj)

    def test_the_same_unit_twice_is_refused_out_loud(self):
        """Молчание человек примет за несработавшую кнопку и нажмёт ещё раз."""
        response = self._add(self.first, fault_description='повтор')

        self.assertEqual(self.order.order_equipments.count(), 1)
        self.assertRedirects(response, '/repair-orders/%d/' % self.order.pk)
        messages = [str(m) for m in response.wsgi_request._messages]
        self.assertTrue(any('уже в этом заказе' in m for m in messages), messages)

    def test_nothing_chosen_adds_nothing(self):
        response = self._add(None)

        self.assertEqual(self.order.order_equipments.count(), 1)
        self.assertRedirects(response, '/repair-orders/%d/' % self.order.pk)

    def test_a_missing_equipment_adds_nothing(self):
        """Номер из чужой базы не должен молча создавать пустую строку."""
        response = self.http.post(
            '/repair-orders/%d/add-unit/' % self.order.pk, {'equipment': '999999'}
        )

        self.assertEqual(self.order.order_equipments.count(), 1)
        self.assertRedirects(response, '/repair-orders/%d/' % self.order.pk)

    def test_adding_is_not_a_get(self):
        """Переход по ссылке или перезагрузка страницы не должны заводить
        оборудование в заказ."""
        response = self.http.get('/repair-orders/%d/add-unit/' % self.order.pk)

        self.assertEqual(response.status_code, 405)
        self.assertEqual(self.order.order_equipments.count(), 1)

    def test_adding_requires_login(self):
        response = TestClient().post(
            '/repair-orders/%d/add-unit/' % self.order.pk, {'equipment': self.second.pk}
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])
        self.assertEqual(self.order.order_equipments.count(), 1)

    def test_the_list_offers_only_what_is_not_in_the_order(self):
        """Иначе первым делом предлагается то, что уже принято."""
        page = self.http.get('/repair-orders/%d/' % self.order.pk)
        available = list(page.context['available_equipment'])

        self.assertIn(self.second, available)
        self.assertNotIn(self.first, available)

    def test_the_card_carries_the_intake_form_and_the_scanner(self):
        page = self.http.get('/repair-orders/%d/' % self.order.pk).content.decode()

        self.assertIn('id="addUnitForm"', page)
        self.assertIn("kinds: ['equipment', 'order_equipment']", page)
        self.assertIn("name: 'Карточка заказа'", page)
        # обрыв связи должен подниматься общей полосой, а не молчать
        self.assertIn('LiftTeamWS.fetch', page)


class IntakeScanTests(TestCase):
    """Скан наклейки, уже наклеенной на приборе, при приёме заказа.

    То, ради чего коды и заводились: прибор приезжает второй раз, на нём
    наклейка с прошлого ремонта — вместо того чтобы искать серийник глазами
    и набирать руками, кладовщик подносит сканер.
    """

    FORM = 'core/templates/core/repair_orders/form.html'

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_intake_scan', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-SCAN-1'
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment
        )
        self.form = (settings.BASE_DIR / self.FORM).read_text(encoding='utf-8')

    def _function(self, name):
        """Тело одной функции из скрипта страницы.

        Именно одной: срез «от имени и до конца файла» захватывал соседние
        функции, и проверка «здесь нет d-none» ловила чужой код.
        """
        body = self.form.split('function %s' % name)[1]
        return body.split('\n}')[0]

    def _resolve(self, code):
        return self.http.get('/scan/resolve/?code=%s' % code).json()

    def test_the_equipment_label_says_which_unit_it_is(self):
        data = self._resolve('e/%d' % self.equipment.pk)

        self.assertTrue(data['found'])
        self.assertEqual(data['equipment_id'], self.equipment.pk)

    def test_an_old_order_label_leads_to_the_same_equipment(self):
        """Наклейка прошлого заказа — про ту же железку.

        Номер в коде там свой (строка заказа), поэтому одного `id` мало:
        без отдельного поля экран приёма поставил бы в заказ не то.
        """
        # Номера должны заведомо разойтись, иначе проверка ничего не значит:
        # без этих записей счётчики единиц и строк заказа идут вровень
        for n in range(3):
            Equipment.objects.create(model=self.model, serial_number='SN-GAP-%d' % n)
        other = Equipment.objects.create(model=self.model, serial_number='SN-SCAN-2')
        roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=other
        )

        data = self._resolve('u/%d' % roe.pk)

        self.assertTrue(data['found'])
        self.assertEqual(data['equipment_id'], other.pk)
        self.assertNotEqual(roe.pk, other.pk)
        self.assertEqual(data['id'], roe.pk)

    def test_both_kinds_are_accepted_on_the_order_page(self):
        """Наклейка бывает и на самом приборе, и от прошлого заказа —
        экран обязан принимать обе, иначе одна из них молча не сработает."""
        self.assertIn("kinds: ['equipment', 'order_equipment']", self.form)
        self.assertIn("name: 'Заказ на ремонт'", self.form)

    def test_the_scan_goes_through_the_connection_layer(self):
        """Голый fetch показал бы обрыв связи как «не найдено» вместо общей
        полосы вверху страницы."""
        handler = self._function('registerIntakeScanner')
        self.assertIn('LiftTeamWS.fetch', handler)

    def test_the_highlight_does_not_depend_on_bootstrap(self):
        """Bootstrap приходит из интернета: подсветки не было бы ровно
        тогда, когда со сканером стоят у стола."""
        placing = self._function('placeScannedEquipment')
        self.assertNotIn('d-none', placing)
        self.assertNotIn('data-bs-', placing)
        self.assertIn('is-scanned', self.form)
        self.assertIn('.equipment-row.is-scanned', self.form)

    def test_a_repeat_scan_is_refused_out_loud(self):
        """Молчаливого бездействия быть не должно: человек решит,
        что не сработал сканер, и поднесёт код ещё раз."""
        placing = self._function('placeScannedEquipment')
        self.assertIn('уже в заказе', placing)
        self.assertIn('ok: false', placing)

    def test_the_choice_is_applied_through_an_event(self):
        """На событии висят подсказка о прошлых ремонтах, список типовых
        неисправностей и пересчёт занятых вариантов. Прямое присваивание
        значения ничего из этого не запустило бы."""
        placing = self._function('placeScannedEquipment')
        self.assertIn("dispatchEvent(new Event('change'", placing)


class EquipmentVersionInDocumentsTests(TestCase):
    """Исполнение печатается слитно с названием модели: «БУАД-7-31.4».

    Заводя версии, решили, что в документы они не идут. Владелец решение
    изменил (v2.58.0), а форму уточнил: не приписка «, исп. 1.1» отдельно,
    а одно обозначение одной строкой — так, как написано на изделии.
    Разделитель хранится внутри обозначения версии, потому что на изделиях
    он произвольный: «БУАД-7-31.4», но «EcoDrive-2.3-1.1».
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_version_docs', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(
            name='БУАД-7-31', kind='Привод дверей'
        )
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4', note='алюминиевый корпус'
        )
        self.equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-VER-1', version=self.version
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Подъём', inn='7700000902')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment,
            diagnosis='вздулись конденсаторы',
        )

    def _pages(self):
        return {
            'акт приёма': '/repair-orders/%d/act/receive/' % self.order.pk,
            'акт выполненных работ': '/repair-orders/%d/act/complete/' % self.order.pk,
            'акт дефектации': '/repair-orders/%d/equipment/%d/act/defect/' % (
                self.order.pk, self.roe.pk),
            'этикетка': '/repair-orders/%d/equipment/%d/label/' % (
                self.order.pk, self.roe.pk),
        }

    def test_the_designation_is_one_string(self):
        """Склейка без своих правил: что завели в справочнике,
        то и печатается."""
        self.assertEqual(self.equipment.version_suffix, '.4')
        self.assertEqual(self.equipment.designation, 'БУАД-7-31.4')
        self.assertEqual(self.equipment.full_designation, 'Привод дверей БУАД-7-31.4')

    def test_every_document_and_the_label_show_it_joined(self):
        for name, url in self._pages().items():
            with self.subTest(документ=name):
                body = self.http.get(url).content.decode()
                self.assertIn('БУАД-7-31.4', body, '%s: обозначения нет' % name)
                # именно слитно: прежней приписки «исп.» быть не должно
                self.assertNotIn('исп.', body.lower())

    def test_the_generic_word_only_reaches_the_defect_act(self):
        """«Привод дверей» на наклейке 43 мм съел бы место, а короткое
        обозначение и так узнают. Полностью изделие называют в акте
        дефектации, и только там."""
        bodies = {name: self.http.get(url).content.decode()
                  for name, url in self._pages().items()}
        self.assertIn('Привод дверей БУАД-7-31.4', bodies['акт дефектации'])
        for name in ('акт приёма', 'акт выполненных работ', 'этикетка'):
            with self.subTest(документ=name):
                self.assertNotIn('Привод дверей', bodies[name])

    def test_the_note_stays_out_of_everything(self):
        """«алюминиевый корпус» — заметка мастеру у стола, а не заказчику."""
        for name, url in self._pages().items():
            with self.subTest(документ=name):
                self.assertNotIn('алюминиевый корпус', self.http.get(url).content.decode())

    def test_without_a_version_nothing_is_printed(self):
        """Пусто — значит пусто: «.0» вместо незаполненного поля
        не выдумывается."""
        self.equipment.version = None
        self.equipment.save(update_fields=['version'])
        self.assertEqual(self.equipment.version_suffix, '')
        self.assertEqual(self.equipment.designation, 'БУАД-7-31')
        for name, url in self._pages().items():
            with self.subTest(документ=name):
                self.assertIn('БУАД-7-31', self.http.get(url).content.decode())

    def test_the_separator_comes_from_the_directory_not_from_a_rule(self):
        """На изделиях разделители разные, и программа своего не добавляет."""
        other_model = EquipmentModel.objects.create(name='EcoDrive-2.3')
        other_version = EquipmentVersion.objects.create(
            equipment_model=other_model, name='-1.1'
        )
        other = Equipment.objects.create(
            model=other_model, serial_number='SN-VER-2', version=other_version
        )
        self.assertEqual(other.designation, 'EcoDrive-2.3-1.1')

    def test_lists_show_the_designation_too(self):
        """Два одинаковых прибора разных исполнений должны различаться
        в списке выбора, а не выглядеть одинаково."""
        self.assertEqual(str(self.equipment), 'БУАД-7-31.4 — SN-VER-1')


class EquipmentLabelPhonesTests(SimpleTestCase):
    """Телефоны на этикетке крупнее: на печати 5 пунктов почти не читались.

    По ним звонят, а не разглядывают их, поэтому размер важнее, чем
    свободное место под остальное.
    """

    STYLES = 'core/templates/core/repair_orders/_label_order_equipment_styles.html'
    MARKUP = 'core/templates/core/repair_orders/_label_order_equipment.html'

    def setUp(self):
        self.styles = (settings.BASE_DIR / self.STYLES).read_text(encoding='utf-8')
        self.markup = (settings.BASE_DIR / self.MARKUP).read_text(encoding='utf-8')

    def test_phones_are_bigger_than_the_rest_of_the_footer(self):
        phones = re.search(r'\.label-phones \{([^}]*)\}', self.styles).group(1)
        company = re.search(r'\.label-company \{([^}]*)\}', self.styles).group(1)
        size = lambda rule: float(re.search(r'font-size: ([\d.]+)pt', rule).group(1))
        self.assertGreaterEqual(size(phones), 6.5)
        # Название и сайт тоже подросли: строка освободилась там, где было
        # отдельное «Исп.». Но телефоны остаются крупнее — по ним звонят.
        self.assertGreaterEqual(size(company), 6)
        self.assertGreater(size(phones), size(company))

    def test_phones_stay_bold(self):
        phones = re.search(r'\.label-phones \{([^}]*)\}', self.styles).group(1)
        self.assertIn('font-weight: bold', phones)

    def test_a_number_never_breaks_in_the_middle(self):
        """Перенос допускается между номерами, но не внутри номера:
        иначе на наклейке остаётся «+7 964 524» и «84 00»."""
        self.assertIn('.label-phones span { white-space: nowrap; }', self.styles)
        self.assertIn('<span>+7 964 524 84 00</span>', self.markup)
        self.assertIn('<span>+7 977 760 10 89</span>', self.markup)


class FaultTextInDocumentsTests(TestCase):
    """Короткое название типовой неисправности не попадает в документы
    никогда: в акт дефектации и в предложение идёт только полное описание,
    а за ним — то, что дописал мастер."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_fault_docs', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(name='EkoDrive 2.0', kind='Преобразователь частоты')
        self.equipment = Equipment.objects.create(model=self.model, serial_number='SN-DOC-1')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Подъём', inn='7700000901')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment,
            diagnosis='вздулись конденсаторы в цепи питания',
            proposed_work='Замена силовой платы',
            estimated_cost=Decimal('12000.00'),
        )
        self.dry = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            description='Высыхание электролитических конденсаторов звена постоянного тока',
        )
        self.burnt = FaultType.objects.create(
            equipment_model=self.model, name='сгорел IGBT',
            description='Пробой силового модуля IGBT',
        )

    def _act(self):
        return self.client_http.get(
            f'/repair-orders/{self.order.pk}/equipment/{self.roe.pk}/act/defect/'
        )

    def _quote(self):
        return self.client_http.get(f'/repair-orders/{self.order.pk}/quote/')

    def test_without_typical_faults_the_documents_are_unchanged(self):
        """Заказ без выбранных неисправностей печатается ровно как раньше:
        в акте — записанная диагностика, в предложении — предлагаемые работы."""
        self.assertEqual(self.roe.diagnosis_document_text, self.roe.diagnosis)
        self.assertEqual(self.roe.quote_line, 'Замена силовой платы')
        self.assertEqual(self.roe.fault_document_lines, [])

        self.assertContains(self._act(), 'вздулись конденсаторы в цепи питания')
        self.assertContains(self._quote(), 'Замена силовой платы')

    def test_the_full_descriptions_reach_both_documents(self):
        self.roe.faults.set([self.dry, self.burnt])

        act = self._act()
        quote = self._quote()
        for response in (act, quote):
            self.assertContains(response, 'Высыхание электролитических конденсаторов')
            self.assertContains(response, 'Пробой силового модуля IGBT')

    def test_the_short_names_never_reach_the_documents(self):
        self.roe.faults.set([self.dry, self.burnt])

        self.assertNotContains(self._act(), 'высохли конденсаторы')
        self.assertNotContains(self._act(), 'сгорел IGBT')
        self.assertNotContains(self._quote(), 'высохли конденсаторы')
        self.assertNotContains(self._quote(), 'сгорел IGBT')

    def test_the_free_text_follows_the_descriptions(self):
        self.roe.faults.set([self.dry])

        self.assertEqual(
            self.roe.diagnosis_document_text,
            'Высыхание электролитических конденсаторов звена постоянного тока\n'
            'вздулись конденсаторы в цепи питания',
        )
        self.assertEqual(
            self.roe.quote_line,
            'Высыхание электролитических конденсаторов звена постоянного тока\n'
            'Замена силовой платы',
        )

    def test_a_fault_without_a_description_adds_nothing(self):
        """Подставлять вместо пустого описания короткое название нельзя —
        тогда цеховой жаргон уехал бы заказчику."""
        nameless = FaultType.objects.create(
            equipment_model=self.model, name='глючит по-непонятному'
        )
        self.roe.faults.set([nameless])

        self.assertEqual(self.roe.fault_document_lines, [])
        self.assertEqual(self.roe.diagnosis_document_text, self.roe.diagnosis)
        self.assertNotContains(self._act(), 'глючит по-непонятному')

    def test_an_empty_unit_still_prints_a_dash(self):
        """Ни описаний, ни диагностики — в акте прочерк, как и раньше."""
        self.roe.diagnosis = ''
        self.roe.save(update_fields=['diagnosis'])

        self.assertEqual(self.roe.diagnosis_document_text, '')
        self.assertFalse(self.roe.has_defect_act)

    def test_the_quote_falls_back_to_the_model_name(self):
        """Ни неисправностей, ни текста — прежняя строка «Ремонт …»."""
        self.roe.proposed_work = ''
        self.roe.save(update_fields=['proposed_work'])

        self.assertEqual(self.roe.quote_line, 'Ремонт Преобразователь частоты EkoDrive 2.0')


class RepairComplexityDerivationTests(TestCase):
    """Сложность — свойство неисправности; у единицы она из них выводится,
    но остаётся правимой вручную."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_complexity', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(name='БУАД-11')
        self.equipment = Equipment.objects.create(model=self.model, serial_number='SN-CX-1')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Лифтсервис', inn='7700000902')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment
        )
        self.simple_fault = FaultType.objects.create(
            equipment_model=self.model, name='окислился разъём',
            description='Окисление контактов разъёма', complexity='simple',
        )
        self.complex_fault = FaultType.objects.create(
            equipment_model=self.model, name='слетела прошивка',
            description='Повреждение прошивки процессора', complexity='complex',
        )

    def test_a_new_fault_is_simple_unless_said_otherwise(self):
        self.assertEqual(
            FaultType.objects.create(equipment_model=self.model, name='ещё одна').complexity,
            'simple',
        )

    def test_without_faults_nothing_is_derived(self):
        self.assertEqual(self.roe.derived_complexity, '')
        self.assertEqual(self.roe.effective_complexity, '')
        self.assertEqual(self.roe.effective_complexity_display, '')
        self.assertFalse(self.roe.complexity_is_derived)

    def test_only_simple_faults_make_a_simple_repair(self):
        self.roe.faults.set([self.simple_fault])

        self.assertEqual(self.roe.derived_complexity, 'simple')
        self.assertTrue(self.roe.complexity_is_derived)
        self.assertEqual(self.roe.effective_complexity_display, 'Простой')

    def test_one_complex_fault_makes_the_whole_repair_complex(self):
        self.roe.faults.set([self.simple_fault, self.complex_fault])

        self.assertEqual(self.roe.derived_complexity, 'complex')
        self.assertEqual(self.roe.effective_complexity_display, 'Сложный')

    def test_a_manual_value_survives_the_derivation(self):
        """Проставленное руками не перебивается выведенным: мастер видит
        прибор, а справочник — нет."""
        self.roe.faults.set([self.simple_fault, self.complex_fault])
        self.roe.repair_complexity = 'simple'
        self.roe.save(update_fields=['repair_complexity'])

        # Из неисправностей вышло бы «сложный» — сложна хотя бы одна
        self.assertEqual(self.roe.derived_complexity, 'complex')
        self.assertEqual(self.roe.effective_complexity, 'simple')
        self.assertEqual(self.roe.effective_complexity_display, 'Простой')
        self.assertFalse(self.roe.complexity_is_derived)

    def test_the_quote_prints_the_derived_value(self):
        self.roe.faults.set([self.complex_fault])

        row = self.order.quote_rows()[0]
        self.assertEqual(row['complexity'], 'Сложный')
        self.assertContains(
            self.client_http.get(f'/repair-orders/{self.order.pk}/quote/'), 'Сложный'
        )

    def test_the_screen_says_where_the_value_came_from(self):
        self.roe.faults.set([self.complex_fault])
        derived = self.client_http.get(f'/repair-orders/{self.order.pk}/quote/edit/')
        self.assertContains(derived, 'Из неисправностей')
        self.assertNotContains(derived, 'Задано вручную')

        self.roe.repair_complexity = 'simple'
        self.roe.save(update_fields=['repair_complexity'])
        manual = self.client_http.get(f'/repair-orders/{self.order.pk}/quote/edit/')
        self.assertContains(manual, 'Задано вручную')
        self.assertNotContains(manual, 'Из неисправностей')


class FaultRecipeVersionTests(TestCase):
    """Строка рецепта с версией заменяет общую только у этого исполнения
    и только по своей детали."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_recipe_version', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(name='EkoDrive 2.0')
        self.v11 = EquipmentVersion.objects.create(
            equipment_model=self.model, name='1.1', note='алюминиевый корпус'
        )
        self.v07 = EquipmentVersion.objects.create(equipment_model=self.model, name='0.7')

        self.capacitor = SparePart.objects.create(
            part_number='RV-1', name='Конденсатор', current_stock=100
        )
        self.resistor = SparePart.objects.create(
            part_number='RV-2', name='Резистор', current_stock=100
        )

        self.fault = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            description='Высыхание конденсаторов',
        )
        FaultTypePart.objects.create(fault_type=self.fault, part=self.capacitor, quantity=3)
        FaultTypePart.objects.create(fault_type=self.fault, part=self.resistor, quantity=1)
        FaultTypePart.objects.create(
            fault_type=self.fault, part=self.capacitor, quantity=5, version=self.v11
        )

        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Высота', inn='7700000903')
        )

    @staticmethod
    def _as_dict(lines):
        return {line.part_id: line.quantity for line in lines}

    def test_the_base_recipe_applies_when_the_version_is_unknown(self):
        self.assertEqual(
            self._as_dict(self.fault.recipe_lines()),
            {self.capacitor.pk: 3, self.resistor.pk: 1},
        )

    def test_the_override_wins_for_its_own_version(self):
        self.assertEqual(
            self._as_dict(self.fault.recipe_lines(self.v11)),
            {self.capacitor.pk: 5, self.resistor.pk: 1},
        )

    def test_another_version_keeps_the_base_line(self):
        self.assertEqual(
            self._as_dict(self.fault.recipe_lines(self.v07)),
            {self.capacitor.pk: 3, self.resistor.pk: 1},
        )

    def _apply_for(self, equipment):
        payload = {'fault_ids': [self.fault.pk]}
        if equipment is not None:
            payload['equipment_id'] = equipment.pk
        return self.client_http.post(
            f'/repair-orders/{self.order.pk}/apply-fault-template/', payload
        )

    def test_applying_to_a_unit_of_that_version_writes_the_override(self):
        unit = Equipment.objects.create(
            model=self.model, serial_number='SN-RV-11', version=self.v11
        )
        RepairOrderEquipment.objects.create(repair_order=self.order, equipment=unit)

        self.assertTrue(self._apply_for(unit).json()['success'])
        written = {d.part_id: d.quantity_used for d in self.order.details.all()}
        self.assertEqual(written, {self.capacitor.pk: 5, self.resistor.pk: 1})

    def test_applying_to_a_unit_without_a_version_writes_the_base_line(self):
        unit = Equipment.objects.create(model=self.model, serial_number='SN-RV-NONE')
        RepairOrderEquipment.objects.create(repair_order=self.order, equipment=unit)

        self.assertTrue(self._apply_for(unit).json()['success'])
        written = {d.part_id: d.quantity_used for d in self.order.details.all()}
        self.assertEqual(written, {self.capacitor.pk: 3, self.resistor.pk: 1})

    def test_a_version_only_line_does_not_leak_to_other_versions(self):
        """Уточнение по детали, которой нет в общем рецепте, достаётся
        только своей версии."""
        extra = SparePart.objects.create(part_number='RV-3', name='Дроссель', current_stock=10)
        FaultTypePart.objects.create(
            fault_type=self.fault, part=extra, quantity=2, version=self.v07
        )

        self.assertNotIn(extra.pk, self._as_dict(self.fault.recipe_lines(self.v11)))
        self.assertNotIn(extra.pk, self._as_dict(self.fault.recipe_lines()))
        self.assertEqual(self._as_dict(self.fault.recipe_lines(self.v07))[extra.pk], 2)

    def test_the_recipe_rejects_a_version_of_another_model(self):
        other_model = EquipmentModel.objects.create(name='ШУНЛ-7')
        stranger = EquipmentVersion.objects.create(equipment_model=other_model, name='2.0')

        data = {
            'equipment_model': self.model.pk, 'name': 'С чужой версией',
            'description': '', 'complexity': 'simple',
            'parts-TOTAL_FORMS': '1', 'parts-INITIAL_FORMS': '0',
            'parts-MIN_NUM_FORMS': '0', 'parts-MAX_NUM_FORMS': '1000',
            'parts-0-part': self.capacitor.pk, 'parts-0-quantity': '1',
            'parts-0-version': stranger.pk,
        }
        response = self.client_http.post('/faults/create/', data)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(FaultType.objects.filter(name='С чужой версией').exists())


class ReferenceCopyTests(TestCase):
    """«Создать копию» открывает форму создания, заполненную образцом.
    До нажатия «Сохранить» в справочнике не появляется ничего."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_copy', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

        self.model = EquipmentModel.objects.create(name='EkoDrive 2.0')
        self.version = EquipmentVersion.objects.create(equipment_model=self.model, name='1.1')
        self.capacitor = SparePart.objects.create(
            part_number='CP-1', name='Конденсатор 470 мкФ', component_type='Конденсатор',
            package='радиальный', min_stock=5, application='Otis',
            description='Электролитический', current_stock=42,
        )
        self.resistor = SparePart.objects.create(part_number='CP-2', name='Резистор')

        self.fault = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            description='Высыхание конденсаторов', complexity='complex',
        )
        FaultTypePart.objects.create(fault_type=self.fault, part=self.capacitor, quantity=3)
        FaultTypePart.objects.create(
            fault_type=self.fault, part=self.resistor, quantity=2, version=self.version
        )

    # --- Типовая неисправность ---

    def test_copying_a_fault_fills_the_form_from_the_source(self):
        response = self.client_http.get(f'/faults/create/?copy_from={self.fault.pk}')

        self.assertEqual(response.status_code, 200)
        form = response.context['form']
        self.assertEqual(form.initial['equipment_model'], self.model.pk)
        self.assertIn('копия', form.initial['name'])
        self.assertEqual(form.initial['description'], 'Высыхание конденсаторов')
        self.assertEqual(form.initial['complexity'], 'complex')

    def test_copying_a_fault_brings_the_whole_recipe(self):
        response = self.client_http.get(f'/faults/create/?copy_from={self.fault.pk}')

        forms = response.context['formset'].forms
        self.assertEqual(
            [(form.initial['part'], form.initial['quantity'], form.initial['version'])
             for form in forms[:2]],
            [(self.capacitor.pk, 3, None), (self.resistor.pk, 2, self.version.pk)],
        )
        # Последняя строка пустая — чтобы в копию можно было дописать деталь
        self.assertEqual(len(forms), 3)
        self.assertNotIn('part', forms[2].initial)

    def test_opening_a_copy_writes_nothing(self):
        faults_before = FaultType.objects.count()
        lines_before = FaultTypePart.objects.count()

        self.client_http.get(f'/faults/create/?copy_from={self.fault.pk}')

        self.assertEqual(FaultType.objects.count(), faults_before)
        self.assertEqual(FaultTypePart.objects.count(), lines_before)

    def test_saving_the_copy_creates_a_second_fault_with_its_recipe(self):
        data = {
            'equipment_model': self.model.pk, 'name': 'высохли конденсаторы (копия)',
            'description': 'Высыхание конденсаторов', 'complexity': 'complex',
            'parts-TOTAL_FORMS': '2', 'parts-INITIAL_FORMS': '0',
            'parts-MIN_NUM_FORMS': '0', 'parts-MAX_NUM_FORMS': '1000',
            'parts-0-part': self.capacitor.pk, 'parts-0-quantity': '3', 'parts-0-version': '',
            'parts-1-part': self.resistor.pk, 'parts-1-quantity': '2',
            'parts-1-version': self.version.pk,
        }
        response = self.client_http.post(f'/faults/create/?copy_from={self.fault.pk}', data)

        self.assertEqual(response.status_code, 302)
        copy = FaultType.objects.get(name='высохли конденсаторы (копия)')
        self.assertEqual(copy.complexity, 'complex')
        self.assertEqual(copy.parts.count(), 2)
        # Образец при этом не тронут
        self.assertEqual(self.fault.parts.count(), 2)

    def test_a_copy_of_a_missing_fault_is_just_an_empty_form(self):
        response = self.client_http.get('/faults/create/?copy_from=999999')

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.context['copy_source'])

    # --- Карточка детали ---

    def test_copying_a_part_fills_the_form_but_not_the_article(self):
        response = self.client_http.get(f'/parts/create/?copy_from={self.capacitor.pk}')

        self.assertEqual(response.status_code, 200)
        initial = response.context['form'].initial
        self.assertEqual(initial['name'], 'Конденсатор 470 мкФ')
        self.assertEqual(initial['component_type'], 'Конденсатор')
        self.assertEqual(initial['application'], 'Otis')
        self.assertNotIn('part_number', initial)

    def test_opening_a_part_copy_writes_nothing(self):
        before = SparePart.objects.count()

        self.client_http.get(f'/parts/create/?copy_from={self.capacitor.pk}')

        self.assertEqual(SparePart.objects.count(), before)

    def test_saving_a_part_copy_starts_from_an_empty_shelf(self):
        response = self.client_http.post('/parts/create/', {
            'part_number': 'CP-1-COPY', 'name': 'Конденсатор 470 мкФ',
            'component_type': 'Конденсатор', 'package': 'радиальный',
            'min_stock': '5', 'lead_time_days': '0',
            'application': 'Otis', 'description': 'Электролитический',
        })

        self.assertEqual(response.status_code, 302)
        copy = SparePart.objects.get(part_number='CP-1-COPY')
        self.assertEqual(copy.current_stock, 0)
        self.capacitor.refresh_from_db()
        self.assertEqual(self.capacitor.current_stock, 42)


class EquipmentTypeDirectoryTests(TestCase):
    """Тип оборудования — справочник со своими экранами; права как
    у моделей: правка всем авторизованным, удаление складу и мастеру."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_types', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)
        self.drive = EquipmentType.objects.create(name='Привод дверей')

    def test_the_list_opens(self):
        self.assertContains(self.client_http.get('/equipment/types/'), 'Привод дверей')

    def test_creating_and_editing(self):
        self.client_http.post('/equipment/types/create/', {
            'name': 'Преобразователь частоты', 'description': '',
        })
        created = EquipmentType.objects.get(name='Преобразователь частоты')

        self.client_http.post(f'/equipment/types/{created.pk}/edit/', {
            'name': 'Преобразователь частоты', 'description': 'ПЧ',
        })
        created.refresh_from_db()

        self.assertEqual(created.description, 'ПЧ')

    def test_deleting_a_type_leaves_the_models_alone(self):
        model = EquipmentModel.objects.create(name='EkoDrive 2.0', equipment_type=self.drive)

        response = self.client_http.post(f'/equipment/types/{self.drive.pk}/delete/')

        self.assertRedirects(response, '/equipment/types/')
        model.refresh_from_db()
        self.assertIsNone(model.equipment_type_id)

    def test_an_accountant_cannot_delete_a_type(self):
        accountant = Employee.objects.create_user(
            username='accountant_types', full_name='Бухгалтер', password='pass', role='accountant'
        )
        client = TestClient()
        client.force_login(accountant)

        response = client.post(f'/equipment/types/{self.drive.pk}/delete/')

        self.assertEqual(response.status_code, 302)
        self.assertTrue(EquipmentType.objects.filter(pk=self.drive.pk).exists())

    def test_an_anonymous_user_is_sent_to_login(self):
        response = TestClient().get('/equipment/types/')

        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response.url)

    def test_the_type_is_not_the_generic_word_of_the_defect_act(self):
        """`kind` печатается в акте перед названием модели, тип — нет.
        Смешивать их нельзя: у модели «Преобразователь частоты Emotron»
        родовое слово уже в названии, и тип его туда не добавляет."""
        model = EquipmentModel.objects.create(
            name='Преобразователь частоты Emotron',
            equipment_type=EquipmentType.objects.create(name='Преобразователь частоты'),
        )

        self.assertEqual(model.full_name, 'Преобразователь частоты Emotron')


class EquipmentVersionDirectoryTests(TestCase):
    """Версия модели — справочник; у единицы она необязательна."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_versions', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)
        self.model = EquipmentModel.objects.create(name='EkoDrive 2.0')
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='1.1', note='алюминиевый корпус'
        )

    def test_the_list_shows_the_note(self):
        response = self.client_http.get('/equipment/versions/')

        self.assertContains(response, '1.1')
        self.assertContains(response, 'алюминиевый корпус')

    def test_creating_a_version(self):
        self.client_http.post('/equipment/versions/create/', {
            'equipment_model': self.model.pk, 'name': '0.7', 'note': 'пластиковый корпус',
        })

        self.assertTrue(
            EquipmentVersion.objects.filter(equipment_model=self.model, name='0.7').exists()
        )

    def test_two_versions_of_one_model_cannot_repeat(self):
        with self.assertRaises(IntegrityError):
            EquipmentVersion.objects.create(equipment_model=self.model, name='1.1')

    def test_deleting_a_version_leaves_the_units_alone(self):
        unit = Equipment.objects.create(
            model=self.model, serial_number='SN-VD-1', version=self.version
        )

        response = self.client_http.post(f'/equipment/versions/{self.version.pk}/delete/')

        self.assertRedirects(response, '/equipment/versions/')
        unit.refresh_from_db()
        self.assertIsNone(unit.version_id)

    def test_an_accountant_cannot_delete_a_version(self):
        accountant = Employee.objects.create_user(
            username='accountant_versions', full_name='Бухгалтер', password='pass', role='accountant'
        )
        client = TestClient()
        client.force_login(accountant)

        response = client.post(f'/equipment/versions/{self.version.pk}/delete/')

        self.assertEqual(response.status_code, 302)
        self.assertTrue(EquipmentVersion.objects.filter(pk=self.version.pk).exists())

    def test_the_equipment_form_rejects_a_version_of_another_model(self):
        other = EquipmentModel.objects.create(name='ШУНЛ-5')

        response = self.client_http.post('/equipment/create/', {
            'model': other.pk, 'version': self.version.pk,
            'serial_number': 'SN-VD-2', 'manufacture_date': '', 'current_client': '',
        })

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Equipment.objects.filter(serial_number='SN-VD-2').exists())


class OptionalTypeAndVersionTests(TestCase):
    """Тип у модели, версия и дата изготовления у единицы — необязательны:
    без них всё работает от начала до конца, как работало раньше."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='admin_optional', full_name='Админ', password='pass'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.admin)

    def test_a_model_without_a_type_and_a_unit_without_a_version(self):
        self.client_http.post('/equipment/models/create/', {'name': 'БУАД-77', 'kind': ''})
        model = EquipmentModel.objects.get(name='БУАД-77')
        self.assertIsNone(model.equipment_type_id)

        self.client_http.post('/equipment/create/', {
            'model': model.pk, 'version': '', 'serial_number': 'SN-OPT-1',
            'manufacture_date': '', 'current_client': '',
        })
        unit = Equipment.objects.get(serial_number='SN-OPT-1')
        self.assertIsNone(unit.version_id)
        self.assertIsNone(unit.manufacture_date)

        order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Без версий', inn='7700000904')
        )
        roe = RepairOrderEquipment.objects.create(
            repair_order=order, equipment=unit, diagnosis='не включается'
        )

        act = self.client_http.get(
            f'/repair-orders/{order.pk}/equipment/{roe.pk}/act/defect/'
        )
        quote = self.client_http.get(f'/repair-orders/{order.pk}/quote/')

        self.assertContains(act, 'не включается')
        self.assertEqual(quote.status_code, 200)

    def test_the_manufacture_date_is_kept_when_it_was_read_off_the_box(self):
        model = EquipmentModel.objects.create(name='БУАД-78')
        self.client_http.post('/equipment/create/', {
            'model': model.pk, 'version': '', 'serial_number': 'SN-OPT-2',
            'manufacture_date': '2019-04-01', 'current_client': '',
        })

        self.assertEqual(
            Equipment.objects.get(serial_number='SN-OPT-2').manufacture_date,
            datetime.date(2019, 4, 1),
        )


class MigrationOnLiveDataTests(TransactionTestCase):
    """Миграция накатывается на базу, в которой уже есть модели, версии
    которых никто не заводил, оборудование, неисправности и заказы.

    На Raspberry Pi лежат настоящие данные заказчиков, и обновление
    не должно ни падать, ни выдумывать значения: всё новое приходит
    пустым или со значением по умолчанию.
    """

    # Откатываемся к состоянию до появления типов, версий и сложности —
    # именно на такой базе стояла программа у владельца.
    BEFORE = '0033_invoice_paid_notice'

    available_apps = None

    @property
    def AFTER(self):
        """Последняя миграция, а не та, ради которой тест писали.

        Имя выводится, а не записано: после отката проверки идут текущими
        классами моделей, и схема обязана совпадать с ними. Пока здесь
        стояло имя одной миграции, тест ломался при каждом новом поле
        в уже существующей таблице — не потому, что миграция плохая,
        а потому что база оставалась на полпути.
        """
        from django.db.migrations.loader import MigrationLoader

        names = sorted(
            name for app, name in MigrationLoader(None, ignore_no_migrations=True).disk_migrations
            if app == 'core'
        )
        return names[-1]

    def _migrate(self, target):
        from django.db import connection as db_connection
        from django.db.migrations.executor import MigrationExecutor

        executor = MigrationExecutor(db_connection)
        executor.loader.build_graph()
        state = executor.migrate([('core', target)])
        return state

    def tearDown(self):
        # База остаётся на последней миграции — следующие тесты работают
        # с актуальной схемой
        self._migrate(self.AFTER)

    def test_existing_rows_survive_and_get_empty_values(self):
        old_state = self._migrate(self.BEFORE)

        OldModel = old_state.apps.get_model('core', 'EquipmentModel')
        OldEquipment = old_state.apps.get_model('core', 'Equipment')
        OldFault = old_state.apps.get_model('core', 'FaultType')
        OldPart = old_state.apps.get_model('core', 'SparePart')
        OldRecipe = old_state.apps.get_model('core', 'FaultTypePart')
        OldClient = old_state.apps.get_model('core', 'Client')
        OldOrder = old_state.apps.get_model('core', 'RepairOrder')
        OldOrderEquipment = old_state.apps.get_model('core', 'RepairOrderEquipment')

        model = OldModel.objects.create(name='EkoDrive 2.0', kind='Преобразователь частоты')
        unit = OldEquipment.objects.create(model=model, serial_number='SN-MIGRATE-1')
        fault = OldFault.objects.create(
            equipment_model=model, name='высохли конденсаторы', description='Высыхание'
        )
        part = OldPart.objects.create(part_number='MG-1', name='Конденсатор')
        OldRecipe.objects.create(fault_type=fault, part=part, quantity=3)
        order = OldOrder.objects.create(
            client=OldClient.objects.create(name='ООО Миграция', inn='7700000905'),
            order_number='M-1',
        )
        OldOrderEquipment.objects.create(repair_order=order, equipment=unit)

        self._migrate(self.AFTER)

        migrated_model = EquipmentModel.objects.get(name='EkoDrive 2.0')
        migrated_unit = Equipment.objects.get(serial_number='SN-MIGRATE-1')
        migrated_fault = FaultType.objects.get(name='высохли конденсаторы')

        # Ничего не выдумано: тип пуст, версии нет, дата изготовления пуста
        self.assertIsNone(migrated_model.equipment_type_id)
        self.assertEqual(migrated_model.kind, 'Преобразователь частоты')
        self.assertIsNone(migrated_unit.version_id)
        self.assertIsNone(migrated_unit.manufacture_date)
        self.assertEqual(EquipmentType.objects.count(), 0)
        self.assertEqual(EquipmentVersion.objects.count(), 0)

        # Сложность у старых неисправностей — «простой», как у поля
        # по умолчанию; рецепт остаётся общим для всех исполнений
        self.assertEqual(migrated_fault.complexity, 'simple')
        self.assertEqual(migrated_fault.parts.count(), 1)
        self.assertIsNone(migrated_fault.parts.first().version_id)

        # Заказ и его оборудование на месте
        self.assertEqual(RepairOrder.objects.get(order_number='M-1').order_equipments.count(), 1)


class ScanDecodeTests(SimpleTestCase):
    """Разбор того, что пришло со сканера.

    Главное здесь — что вид кода определяется видом пути, а не адресом
    сервера: наклейки печатались с той основы, что стояла в LABEL_BASE_URL
    на момент печати, и вчерашняя наклейка обязана читаться сегодня.
    Второе главное — что неузнанный код именно неузнан, а не угадан:
    угаданный номер открыл бы чужую карточку, и человек со сканером
    в руках этого не заметил бы.
    """

    def test_all_four_kinds(self):
        for prefix, kind in (('p', 'part'), ('c', 'cell'), ('e', 'equipment'), ('o', 'order')):
            with self.subTest(prefix=prefix):
                self.assertEqual(
                    scanning.decode(f'http://lifteam.taile9b605.ts.net/{prefix}/12'),
                    {'kind': kind, 'id': 12},
                )

    def test_host_does_not_matter(self):
        """Наклейка напечатана с другого адреса — читается всё равно."""
        for payload in (
            'http://lifteam.taile9b605.ts.net/p/7',
            'http://192.168.1.50:8000/p/7',
            'https://lifteam.example.org/p/7',
            'http://100.83.12.4/p/7',
        ):
            with self.subTest(payload=payload):
                self.assertEqual(scanning.decode(payload), {'kind': 'part', 'id': 7})

    def test_with_and_without_trailing_slash(self):
        self.assertEqual(scanning.decode('http://pi/o/3'), {'kind': 'order', 'id': 3})
        self.assertEqual(scanning.decode('http://pi/o/3/'), {'kind': 'order', 'id': 3})

    def test_bare_path_and_whitespace(self):
        """Со сканера приходит и голый путь, и лишние пробелы вокруг."""
        self.assertEqual(scanning.decode('/c/5'), {'kind': 'cell', 'id': 5})
        self.assertEqual(scanning.decode('c/5'), {'kind': 'cell', 'id': 5})
        self.assertEqual(scanning.decode('  \n/c/5/\t '), {'kind': 'cell', 'id': 5})
        self.assertEqual(scanning.decode('﻿/c/5'), {'kind': 'cell', 'id': 5})

    def test_query_and_fragment_are_ignored(self):
        self.assertEqual(scanning.decode('http://pi/p/9?from=label'), {'kind': 'part', 'id': 9})
        self.assertEqual(scanning.decode('http://pi/p/9#top'), {'kind': 'part', 'id': 9})

    def test_not_ours_is_not_guessed(self):
        """Всё, что не одна из четырёх форм, — чужой код, а не догадка."""
        for payload in (
            '',
            None,
            '4607034170264',                      # штрихкод товара из магазина
            'https://example.com/',
            'http://pi/parts/12/',                # длинный адрес той же детали
            '/p/',
            '/p/abc',
            '/x/12',
            '/p/12/edit/',
            '/p/12 /c/13',
            'p12',
            'BEGIN:VCARD',
        ):
            with self.subTest(payload=payload):
                self.assertIsNone(scanning.decode(payload))

    def test_kind_label_is_russian(self):
        self.assertEqual(scanning.kind_label('part'), 'Радиодеталь')
        self.assertEqual(scanning.kind_label('nonsense'), 'Неизвестный вид')


class ScanPageTests(TestCase):
    """Страница «Сканирование» и разбор кода на сервере."""

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='scan_user', full_name='Кладовщик', password='pass', role='warehouse'
        )
        self.client_http = TestClient()
        self.client_http.force_login(self.employee)

        self.part = SparePart.objects.create(
            part_number='SCAN-R1', name='Резистор для скана',
            component_type='Резистор', current_stock=8, min_stock=2,
        )
        self.cabinet = Cabinet.objects.create(number=11, name='Сканы')
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.order_by('cell_row').first()
        self.cell.parts.add(self.part)

        self.model = EquipmentModel.objects.create(name='БУАД-скан')
        self.equipment = Equipment.objects.create(model=self.model, serial_number='SN-SCAN-1')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='ООО Скан', inn='7700000222')
        )

    def _resolve(self, code):
        return self.client_http.get('/scan/resolve/', {'code': code}).json()

    def test_page_requires_login(self):
        for url in ('/scan/', '/scan/resolve/'):
            with self.subTest(url=url):
                response = TestClient().get(url)
                self.assertEqual(response.status_code, 302)
                self.assertIn('/login/', response['Location'])

    def test_page_opens(self):
        response = self.client_http.get('/scan/')
        self.assertEqual(response.status_code, 200)

    def test_resolves_all_four_kinds(self):
        cases = {
            f'/p/{self.part.pk}': ('part', 'SCAN-R1'),
            f'/c/{self.cell.pk}': ('cell', self.cell.address),
            f'/e/{self.equipment.pk}': ('equipment', 'SN-SCAN-1'),
            f'/o/{self.order.pk}': ('order', self.order.order_number),
        }
        for code, (kind, title) in cases.items():
            with self.subTest(code=code):
                data = self._resolve(code)
                self.assertTrue(data['recognized'])
                self.assertTrue(data['found'])
                self.assertEqual(data['kind'], kind)
                self.assertEqual(data['title'], title)
                self.assertTrue(data['actions'])

    def test_foreign_host_resolves(self):
        """Наклейка напечатана с другой основой — код всё равно наш."""
        data = self._resolve(f'http://lifteam.taile9b605.ts.net/p/{self.part.pk}')
        self.assertTrue(data['found'])
        self.assertEqual(data['title'], 'SCAN-R1')

    def test_slash_at_the_end_does_not_matter(self):
        without = self._resolve(f'/o/{self.order.pk}')
        with_slash = self._resolve(f'/o/{self.order.pk}/')
        self.assertEqual(without, with_slash)

    def test_missing_object_is_reported_not_invented(self):
        data = self._resolve('/p/999999')
        self.assertTrue(data['recognized'])
        self.assertFalse(data['found'])
        self.assertEqual(data['kind'], 'part')
        self.assertIn('не найдена', data['message'])

    def test_unrecognized_payload(self):
        data = self._resolve('4607034170264')
        self.assertFalse(data['recognized'])
        self.assertNotIn('kind', data)
        self.assertIn('не код LiftTeam', data['message'])

    def test_part_answer_carries_stock_and_cell(self):
        """По ответу решают, идти к полке или заказывать, — значит остаток
        и адрес ячейки обязательны."""
        data = self._resolve(f'/p/{self.part.pk}')
        values = {line['label']: line['value'] for line in data['lines']}
        self.assertIn('8 шт', values['Остаток'])
        self.assertEqual(values['Ячейка'], self.cell.address)

    def test_short_routes_lead_where_the_scan_says(self):
        """Скан без подписанного экрана открывает объект — теми же
        короткими адресами, что записаны в QR."""
        response = self.client_http.get(f'/p/{self.part.pk}')
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], f'/parts/{self.part.pk}/')


class ScanLayerWiringTests(SimpleTestCase):
    """Слой сканера подключён на каждой странице и не зависит от Bootstrap.

    Bootstrap приезжает из интернета и не приезжал уже дважды; сканером
    работают с занятыми руками, и молчащая полоса подтверждения — это
    способ завести неверные данные, не заметив этого.
    """

    STATIC = Path(__file__).resolve().parent / 'static'
    TEMPLATES = Path(__file__).resolve().parent / 'templates' / 'core'

    def _code(self, path):
        """Сам код, без комментариев: в них как раз и написано, чего
        здесь быть не должно."""
        source = path.read_text(encoding='utf-8')
        source = re.sub(r'/\*.*?\*/', '', source, flags=re.S)
        return re.sub(r'^\s*//.*$', '', source, flags=re.M)

    def test_loaded_from_base_template(self):
        base = (self.TEMPLATES / 'base.html').read_text(encoding='utf-8')
        self.assertIn("js/scanner.js", base)
        self.assertIn("css/scanner.css", base)

    def test_no_bootstrap_js(self):
        for name in ('static/js/scanner.js', 'templates/core/scan.html'):
            with self.subTest(file=name):
                path = Path(__file__).resolve().parent / name
                code = self._code(path)
                self.assertNotIn('data-bs-', code)
                self.assertNotIn('bootstrap.', code)
                self.assertNotIn('new bootstrap', code)

    def test_requests_go_through_the_connection_layer(self):
        """Голый fetch означал бы, что обрыв связи выглядит как «не найдено»
        вместо общей полосы вверху страницы."""
        code = self._code(self.TEMPLATES / 'scan.html')
        self.assertIn('LiftTeamWS.fetch', code)
        self.assertNotIn('window.fetch(', code)

    def test_scanner_knows_the_same_kinds_as_the_server(self):
        """Разбор в браузере и на сервере обязан совпадать: расхождение
        означало бы, что скан открывает не то, что показала страница."""
        code = self._code(self.STATIC / 'js' / 'scanner.js')
        for prefix, kind in scanning.KINDS.items():
            with self.subTest(prefix=prefix):
                self.assertIn(f"{prefix}: '{kind}'", code)
                self.assertIn(f"{kind}: '", code)   # и человеческое название

    def test_the_letter_for_a_kind_is_derived_not_listed_twice(self):
        """Приставка для перехода выводится из того же списка видов.

        Пока она была отдельным списком, при добавлении вида его забыли
        обновить: код `u/20` разбирался верно, а переход вёл
        на /undefined/20/ — «страница не найдена» вместо принятой единицы.
        Две таблицы одного и того же расходятся молча, и замечают это
        со сканером в руках у стеллажа.
        """
        code = self._code(self.STATIC / 'js' / 'scanner.js')

        self.assertIn('KIND_PREFIX[KINDS[letter]] = letter', code)
        # именно выводится, а не записана вторым списком
        self.assertNotRegex(code, r'KIND_PREFIX\s*=\s*\{\s*\w')

    def test_every_kind_has_a_page_to_open(self):
        """Приставка есть — значит и адрес по ней должен открываться.
        Иначе скан приводит на «страница не найдена»."""
        for prefix in scanning.KINDS:
            with self.subTest(prefix=prefix):
                match = resolve(f'/{prefix}/1/')
                self.assertTrue(match.func)

    def test_repeat_and_speed_thresholds_are_stated(self):
        """Пороги — предмет этого слоя: скан узнаётся по средней скорости
        набора, а повтор в течение секунды считается одним сканом.

        Средняя, а не каждый промежуток по отдельности: в браузере видно,
        как заминка посреди кода разрывала его пополам и деталь
        превращалась в «это не наш код»."""
        code = self._code(self.STATIC / 'js' / 'scanner.js')
        self.assertIn('MAX_AVERAGE_MS = 30', code)
        self.assertIn('REPEAT_MS = 1000', code)
        self.assertIn('payload.length * MAX_AVERAGE_MS', code)


class LeaveDialogWithoutBootstrapTests(TestCase):
    """Окно «Несохранённые изменения» не должно запирать человека в форме,
    когда Bootstrap не приехал из интернета.

    Переход по ссылке отменяется вызовом preventDefault ещё до показа окна.
    Раньше следом шло обращение к `bootstrap.Modal`, и без Bootstrap оно
    падало с ошибкой: окно не появлялось, а переход уже был отменён —
    нажатие на ссылку не делало ровно ничего. Сама разметка окна при этом
    проступала текстом внизу каждой страницы: прячет её правило Bootstrap,
    которого тоже не было.
    """

    def setUp(self):
        self.base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')

    def test_leave_dialog_falls_back_without_bootstrap(self):
        self.assertIn("typeof bootstrap === 'undefined'", self.base)

    def test_dialog_hidden_by_own_rule(self):
        self.assertIn('.modal { display: none; }', self.base)

    def test_closing_goes_through_one_guarded_helper(self):
        """Обе кнопки закрывают окно одним защищённым помощником, а не
        обращаются к Bootstrap напрямую."""
        self.assertIn('function hideLeaveDialog()', self.base)
        self.assertNotIn('bootstrap.Modal.getInstance(modalEl).hide()', self.base)


# ==================== СПИСКИ И МЕНЮ НА ТЕЛЕФОНЕ (v2.56.0) ====================


class MobileListTableTests(SimpleTestCase):
    """Строчки в списках помещаются на экран телефона.

    Две отдельные беды, и обе проверяются здесь. Первая: широкая таблица
    растягивала страницу шире экрана, потому что прокрутку ей давал
    .table-responsive — класс Bootstrap, а Bootstrap приходит из интернета
    и в лаборатории регулярно не приезжает. Вторая: прокрутка и сама
    по себе не ответ — семь столбцов на 390 точках не читаются, и строка
    на узком экране становится карточкой с названиями столбцов у значений.
    """

    TEMPLATES = 'core/templates/core/'
    CSS = 'core/static/css/list-table.css'

    # Печатные формы: у них своя таблица act-table и свой лист бумаги.
    # Карточки им не нужны и вредны — это единственное исключение,
    # и список исключений намеренно перечислен здесь целиком.
    PRINT_ONLY = {
        'repair_orders/act_complete.html',
        'repair_orders/act_receive.html',
        'repair_orders/quote.html',
    }

    @property
    def PAGES(self):
        """Все экранные шаблоны с таблицей-шапкой, найденные в дереве.

        Список не записан руками намеренно: новый список, заведённый
        обычной широкой таблицей, обязан ронять тесты сам, без того чтобы
        кто-то вспомнил дописать его сюда.
        """
        root = settings.BASE_DIR / self.TEMPLATES
        pages = {}
        for path in sorted(root.rglob('*.html')):
            rel = path.relative_to(root).as_posix()
            if rel in self.PRINT_ONLY:
                continue
            text = path.read_text(encoding='utf-8')
            for table in re.findall(r'<table[^>]*>.*?</table>', text, re.S):
                if '<thead' in table:
                    pages[rel] = rel
                    break
        return pages

    def _text(self, rel):
        return (settings.BASE_DIR / self.TEMPLATES / rel).read_text(encoding='utf-8')

    def _stacked_tables(self, rel):
        """Куски шаблона от <table ... list-table ...> до </table>."""
        text = self._text(rel)
        return [(m.start(), m.group(0)) for m in
                re.finditer(r'<table[^>]*\blist-table\b[^>]*>.*?</table>', text, re.S)]

    def test_every_list_page_stacks_its_rows(self):
        for page, rel in self.PAGES.items():
            with self.subTest(page=page):
                self.assertTrue(self._stacked_tables(rel),
                                '%s: у таблицы нет class="list-table"' % page)

    def test_every_list_page_has_its_own_scroll_wrapper(self):
        """Подстраховка на случай, когда карточки не спасают: коробка
        со своей прокруткой, а не класс Bootstrap."""
        for page, rel in self.PAGES.items():
            with self.subTest(page=page):
                text = self._text(rel)
                for start, _ in self._stacked_tables(rel):
                    before = text[:start]
                    wrapper = before[before.rindex('<div'):]
                    self.assertIn('table-scroll', wrapper,
                                  '%s: таблица не в коробке с прокруткой' % page)

    def test_every_data_cell_carries_its_column_name(self):
        """Без data-label ячейка в карточке остаётся значением без подписи:
        «19.08.2026» само по себе не говорит, дата это приёма или отгрузки."""
        allowed = ('data-label=', 'col-check', 'col-actions', 'colspan=')
        for page, rel in self.PAGES.items():
            with self.subTest(page=page):
                for _, table in self._stacked_tables(rel):
                    body = re.search(r'<tbody[^>]*>.*?</tbody>', table, re.S)
                    self.assertIsNotNone(body, '%s: у таблицы нет <tbody>' % page)
                    # (?:[^>{]|\{%.*?%\})* — ячейка вида
                    # <td{% if ... %} class="..."{% endif %} data-label="...">
                    for cell in re.findall(r'<td((?:[^>{]|\{%.*?%\})*)>', body.group(0)):
                        self.assertTrue(
                            any(mark in cell for mark in allowed),
                            '%s: ячейка без названия столбца: %s' % (page, cell),
                        )

    def test_scroll_and_stacking_rules_are_our_own(self):
        """Правила лежат в своём файле, а не берутся у Bootstrap: он
        приходит из интернета, и когда не приезжает, класс .table-responsive
        не значит ничего."""
        css = (settings.BASE_DIR / self.CSS).read_text(encoding='utf-8')
        self.assertIn('.table-scroll', css)
        self.assertIn('overflow-x: auto', css)
        self.assertIn('.list-table', css)
        self.assertIn('attr(data-label)', css)
        # своё правило, а не переопределение чужого класса
        self.assertNotIn('.table-responsive {', css)
        # ничего от поведения Bootstrap
        self.assertNotIn('data-bs-', css)

    def test_desktop_layout_is_untouched(self):
        """Карточки — только для узкого экрана. И только для экрана:
        страница этикетки шириной 43 мм для браузера тоже узкая."""
        css = (settings.BASE_DIR / self.CSS).read_text(encoding='utf-8')
        self.assertIn('@media screen and (max-width: 768px)', css)
        self.assertNotIn('@media (max-width: 768px)', css)
        # всё, что превращает строку в карточку, живёт внутри этого условия
        head, _, tail = css.partition('@media screen and (max-width: 768px)')
        self.assertNotIn('display: block', head)
        self.assertIn('display: block', tail)

    def test_stylesheet_is_linked_on_every_page(self):
        base = (settings.BASE_DIR / self.TEMPLATES / 'base.html').read_text(encoding='utf-8')
        self.assertIn("css/list-table.css", base)


class MobileListTableRenderTests(TestCase):
    """То же самое, но на живых страницах: разметка доезжает до браузера."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='mobile_admin', full_name='Админ', password='pass'
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.client_obj = ClientModel.objects.create(name='ООО «Телефон»')
        self.model = EquipmentModel.objects.create(name='Шкаф телефонный')
        self.equipment = Equipment.objects.create(model=self.model, serial_number='SN-M-1')
        self.order = RepairOrder.objects.create(client=self.client_obj)
        RepairOrderEquipment.objects.create(repair_order=self.order, equipment=self.equipment)
        part = SparePart.objects.create(part_number='M-1', name='Деталь', current_stock=1, min_stock=0)
        # Журнал движений без единого движения рисует только строку
        # «ничего не найдено», а проверяется здесь именно строка записи.
        StockMovement.objects.create(part=part, quantity=1, movement_type='incoming')

    def test_pages_carry_the_stacking_markup(self):
        pages = {
            '/repair-orders/': '№ заказа',
            '/repair-orders/%d/' % self.order.pk: 'Серийный номер',
            '/parts/': 'Артикул',
            '/reports/stock-movements/': 'Деталь',
            '/reports/debtors/': None,
            '/inventory/': None,
            '/equipment/': 'Серийный номер',
        }
        for url, label in pages.items():
            with self.subTest(url=url):
                page = self.http.get(url).content.decode()
                self.assertIn('table-scroll', page)
                self.assertIn('list-table', page)
                if label:
                    self.assertIn('data-label="%s"' % label, page)


class MobileMenuButtonTests(SimpleTestCase):
    """Кнопка меню — логотип на объёмной площадке (v2.56.0).

    Имя класса .sidebar-toggle прежнее: на него ссылаются правила печати
    в base.html и в шаблонах этикеток и актов. Переименование оставило бы
    кнопку меню на наклейке.
    """

    def setUp(self):
        self.base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')
        self.button = self.base.split('class="sidebar-toggle"')[1].split('</button>')[0]

    def test_button_shows_the_logo(self):
        """Именно логотип, а не заглушка.

        lift_team_logo.svg — это синий квадрат с буквами «LT», нарисованный
        когда-то вместо картинки. Настоящий логотип, круглая эмблема
        с микросхемой, лежит в lift_team_logo.png — он же стоит на входе,
        в шапке меню и в актах. Кнопка обязана показывать его.
        """
        self.assertIn('img/lift_team_logo.png', self.button)
        self.assertNotIn('lift_team_logo.svg', self.button)
        self.assertIn('sidebar-toggle-mark', self.button)

    def test_button_keeps_its_accessible_label(self):
        self.assertIn('aria-label="Меню"', self.button)

    def test_missing_image_leaves_a_visible_button(self):
        """Иначе лаборатория остаётся без единственного способа открыть меню.

        Запасных видов два: сначала PNG вместо SVG, потом буквы вместо
        картинки вовсе. Подмена написана прямо в onerror, а не в общем
        скрипте: ошибка загрузки случается раньше, чем доходит очередь
        до скриптов внизу страницы.
        """
        self.assertIn('onerror=', self.button)
        self.assertIn('sidebar-toggle-letters', self.button)
        self.assertIn('>LT<', self.button)
        self.assertIn('.sidebar-toggle-letters', self.base)

    def rule(self, selector, must_contain):
        """Тело правила по селектору.

        Одно и то же имя написано в base.html несколько раз — своё правило
        для печати, для компьютера и для телефона, — поэтому нужное
        отбирается по тому, что внутри него стоит, а не по порядку.
        """
        bodies = re.findall(r'%s\s*\{([^}]*)\}' % re.escape(selector), self.base)
        found = [body for body in bodies if must_contain in body]
        self.assertTrue(found, 'правило %s с «%s» не найдено' % (selector, must_contain))
        return found[0]

    def test_touch_target_stays_big_enough(self):
        rule = self.rule('.sidebar-toggle', 'width:')
        for side in ('width', 'height'):
            size = int(re.search(r'\b%s: (\d+)px' % side, rule).group(1))
            self.assertGreaterEqual(size, 44, 'кнопка меньше пальца')

    def test_button_looks_pressed_at_the_touch_not_at_the_release(self):
        self.assertIn('.sidebar-toggle.is-pressed', self.base)
        self.assertIn("addEventListener('pointerdown'", self.base)
        self.assertIn("addEventListener('pointerup'", self.base)

    def test_press_animation_respects_the_system_setting(self):
        self.assertIn('@media (prefers-reduced-motion: reduce)', self.base)
        quiet = self.base.split('@media (prefers-reduced-motion: reduce)')[1]
        self.assertIn('transition: none', quiet)
        self.assertIn('transform: none', quiet)

    def test_button_stays_above_the_sidebar_and_its_backdrop(self):
        toggle = self.rule('.sidebar-toggle', 'z-index:')
        sidebar = self.rule('.sidebar', 'z-index:')
        backdrop = self.rule('.sidebar-backdrop.show', 'z-index:')
        self.assertGreater(int(re.search(r'z-index: (\d+)', toggle).group(1)),
                           int(re.search(r'z-index: (\d+)', sidebar).group(1)))
        self.assertGreater(int(re.search(r'z-index: (\d+)', toggle).group(1)),
                           int(re.search(r'z-index: (\d+)', backdrop).group(1)))

    def test_button_needs_no_bootstrap(self):
        """Bootstrap приходит из интернета; открыть меню надо и без него."""
        self.assertNotIn('data-bs-', self.button)

    def test_printing_templates_still_know_the_button_by_name(self):
        """Печатные шаблоны прячут кнопку по имени класса. Переименуй его —
        и логотип уедет на наклейку 43 мм."""
        printing = [
            'core/templates/core/parts/label.html',
            'core/templates/core/parts/labels_batch.html',
            'core/templates/core/storage_cells/label.html',
            'core/templates/core/storage_cells/labels_batch.html',
            'core/templates/core/repair_orders/labels_batch.html',
            'core/templates/core/repair_orders/_act_styles.html',
        ]
        for rel in printing:
            with self.subTest(template=rel):
                text = (settings.BASE_DIR / rel).read_text(encoding='utf-8')
                self.assertIn('.sidebar-toggle', text)

    def test_open_menu_scrolls_itself_and_not_the_page(self):
        """Палец на открытом меню двигает разделы, а не страницу под ним.

        Три условия, и каждое закрывает свой случай: страница под меню
        замирает целиком, затемнение не ловит прокрутку, а список
        разделов, докрученный до конца, не передаёт движение наружу.
        """
        self.assertIn('body.sidebar-open { overflow: hidden; }', self.base)
        self.assertIn('.sidebar-backdrop { touch-action: none; }', self.base)
        self.assertIn('overscroll-behavior: contain', self.base)
        # и класс действительно снимается при закрытии, а не только вешается
        self.assertIn("classList.toggle('sidebar-open', open)", self.base)

    def test_the_logo_is_the_button_and_not_a_second_one(self):
        """В шапке открытого меню логотип не рисуется отдельной картинкой.

        Пока их было две, кнопка вставала поверх шапки и накрывала
        логотип — владелец увидел ровно это. Логотип и кнопка должны
        быть одним и тем же.
        """
        mobile = self.base.split('@media screen and (max-width: 768px)')[1]
        self.assertIn('.sidebar .navbar-brand img { display: none; }', mobile)
        # и название начинается правее кнопки, а не под ней
        self.assertRegex(mobile, r'\.sidebar \.navbar-brand \{[^}]*padding-left')


class MobileSidebarScrollTests(SimpleTestCase):
    """Меню прокручивается, и до нижних разделов можно дойти.

    Панель закреплена (position: fixed). Пока у неё стояло min-height:
    100vh, меню длиннее экрана уезжало вниз за его край: страница под
    панелью двигается сама по себе, а панель — нет, и «Оповещения»,
    «Реквизиты» и «Обновление» на телефоне были недоступны вовсе.
    """

    def setUp(self):
        self.base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')

    def rule(self, selector):
        found = re.findall(r'(?<![\w-])%s\s*\{([^}]*)\}' % re.escape(selector), self.base)
        self.assertTrue(found, 'правило %s не найдено' % selector)
        return found[0]

    def test_sidebar_is_exactly_one_screen_tall(self):
        sidebar = self.rule('.sidebar')
        self.assertIn('height: 100vh', sidebar)
        self.assertNotIn('min-height: 100vh', sidebar)

    def test_list_of_sections_scrolls_inside_the_sidebar(self):
        nav = self.rule('.sidebar-nav')
        self.assertIn('overflow-y: auto', nav)
        # без min-height: 0 ячейка flex не сжимается и не прокручивается
        self.assertIn('min-height: 0', nav)

    def test_header_and_signature_stay_put(self):
        fixed = self.rule('.sidebar .navbar-brand,\n        .sidebar-user')
        self.assertIn('flex: 0 0 auto', fixed)

    def test_scrolling_does_not_depend_on_bootstrap(self):
        """Раскладка в столбец написана своими правилами.

        В разметке стоят и классы Bootstrap (d-flex flex-column,
        flex-grow-1), но держаться за них нельзя: Bootstrap приходит
        из интернета, и вместе с ним пропала бы прокрутка меню.
        """
        sidebar = self.rule('.sidebar')
        self.assertIn('display: flex', sidebar)
        self.assertIn('flex-direction: column', sidebar)
        self.assertIn('class="nav flex-column flex-grow-1 sidebar-nav"', self.base)
        self.assertIn('sidebar-user', self.base)


# ==================== СКАН ПО СКЛАДСКИМ ЭКРАНАМ (v2.67.0) ====================


class WarehouseScanScreensTests(TestCase):
    """Скан на складских экранах: пересчёт, сетка кассетниц, карточка детали.

    До этого подписаны были только заказы. На складе сканер полезен больше:
    в кассетнице двести ячеек, строки в таблице пересчёта отсортированы
    не так, как детали лежат на полке, и артикул глазами ищут дольше,
    чем считают сами детали.

    Проверяется не поведение в браузере (его проверить нечем), а то, что
    подписка на месте, принимает нужные виды кодов, ходит через слой связи
    и не полагается на Bootstrap. Каждое из трёх уже ломалось в этом
    проекте по отдельности.
    """

    TEMPLATES = Path(__file__).resolve().parent / 'templates' / 'core'

    SCREENS = {
        'inventory/count.html': ("kinds: ['part', 'cell']", "name: 'Инвентаризация'"),
        'storage_cells/grid.html': ("kinds: ['part', 'cell']", "name: 'Кассетницы'"),
        'parts/detail.html': ("kinds: ['cell']", "name: 'Карточка детали'"),
    }

    def setUp(self):
        self.employee = Employee.objects.create_user(
            username='scan_warehouse', full_name='Кладовщик',
            password='pass', role='warehouse',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.part = SparePart.objects.create(
            part_number='SCR-100n', name='Конденсатор 100 нФ',
            component_type='Конденсатор', current_stock=40, min_stock=5,
        )
        self.cabinet = Cabinet.objects.create(number=41, name='Сканы склада')
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.order_by('cell_row').first()
        self.cell.parts.add(self.part)

    def source(self, name):
        return (self.TEMPLATES / name).read_text(encoding='utf-8')

    def test_every_warehouse_screen_is_subscribed(self):
        """Подписка на месте и объявляет ровно те виды, которые берёт.

        Вид, которого экран не ждёт, слой отказывает вслух и предлагает
        обычный переход, — но только если знает, чего экран ждёт.
        """
        for name, (kinds, screen) in self.SCREENS.items():
            with self.subTest(screen=name):
                code = self.source(name)
                self.assertIn('LiftTeamScanner.register', code)
                self.assertIn(kinds, code)
                self.assertIn(screen, code)

    def test_scans_go_through_the_connection_layer(self):
        """Голый fetch показал бы обрыв связи как «деталь не найдена»
        вместо общей полосы вверху страницы."""
        for name in ('storage_cells/grid.html',):
            with self.subTest(screen=name):
                code = self.source(name)
                self.assertIn('LiftTeamWS.fetch', code)
                self.assertNotIn('window.fetch(', code)

    def test_the_grid_survives_without_bootstrap(self):
        """Окна сетки рисует Bootstrap, а он приходит из интернета.

        Пока `new bootstrap.Modal` стоял голым, он ронял весь обработчик
        готовности страницы — вместе с окнами переставали работать режим
        перемещения, переход с наклейки ячейки и подписка на сканер,
        стоящие в той же функции ниже.
        """
        code = self.source('storage_cells/grid.html')
        self.assertIn('cellInfoModal = null', code)
        self.assertIn('if (cellInfoModal) cellInfoModal.show();', code)
        self.assertIn('if (moveCabinetModal) moveCabinetModal.show();', code)
        # А содержимое ячейки называется словами, не только в окне
        self.assertIn("return 'Ячейка ' + data.address + ': ' + what;", code)

    def test_count_rows_carry_the_numbers_the_scan_looks_up(self):
        """Строка пересчёта ищется по номеру, а не по напечатанному тексту:
        артикул может оказаться куском названия соседней детали."""
        session = InventorySession.objects.create(
            cabinet=self.cabinet, started_by=self.employee
        )
        line = InventorySessionLine.objects.create(
            session=session, part=self.part, cell=self.cell,
            expected_quantity=self.part.current_stock,
        )

        html = self.http.get(reverse('inventory_count', args=[session.pk])).content.decode()

        self.assertIn('data-part="%d"' % self.part.pk, html)
        self.assertIn('data-cell="%d"' % self.cell.pk, html)
        self.assertIn('data-title="%s' % self.part.part_number, html)
        self.assertIn('counted_%d' % line.pk, html)

    def test_scan_answer_says_where_the_part_lies(self):
        """Сетка открывает кассетницу и подсвечивает ячейку по этим полям.

        Строки для человека («Ячейка: A-1-1») для этого не годятся: адрес
        пришлось бы разбирать обратно, а разбор напечатанного — это способ
        подсветить не ту ячейку.
        """
        data = self.http.get('/scan/resolve/?code=p/%d' % self.part.pk).json()

        self.assertTrue(data['found'])
        self.assertEqual(data['cell_id'], self.cell.pk)
        self.assertEqual(data['cabinet_number'], self.cabinet.number)

    def test_a_part_without_a_cell_says_so_instead_of_guessing(self):
        homeless = SparePart.objects.create(
            part_number='SCR-NOCELL', name='Не размещённая', current_stock=1,
        )

        data = self.http.get('/scan/resolve/?code=p/%d' % homeless.pk).json()

        self.assertTrue(data['found'])
        self.assertIsNone(data['cell_id'])
        self.assertIsNone(data['cabinet_number'])

    def test_the_grid_opens_without_a_single_cabinet(self):
        """Свежая установка: кассетниц ещё нет.

        Набор ячеек вставляется в скрипт как есть, и пока его не было,
        получалось `const CELLS_DATA = ;` — синтаксическая ошибка, гасившая
        весь скрипт страницы разом.
        """
        Cabinet.objects.all().delete()

        html = self.http.get(reverse('storage_cell_grid')).content.decode()

        self.assertIn('const CELLS_DATA = {};', html)

    def test_moving_a_part_by_scan_reports_the_real_outcome(self):
        """Перекладка отдаёт исход наружу: сказать «переложено», когда
        перекладка не удалась, хуже, чем не сказать ничего."""
        code = self.source('storage_cells/grid.html')

        self.assertIn('return LiftTeamWS.fetch', code)
        self.assertIn('return performMoveToCell(', code)
        self.assertIn("ok: false, text: 'Переложить в ячейку '", code)

    def test_the_part_card_does_not_move_the_part_by_itself(self):
        """Скан выбирает ячейку, но не назначает её.

        Назначение снимает деталь с прежней ячейки. Сделать это по одному
        поднесению кода, без подтверждения, — способ потерять деталь
        на складе из-за случайно задетого сканера.
        """
        code = self.source('parts/detail.html')

        self.assertIn('select.value = id;', code)
        self.assertNotIn('.submit()', code)
        self.assertIn('нажмите «Назначить»', code)


# ==================== МАТЕРИАЛЫ МОДЕЛИ (v2.68.0) ====================


class EquipmentMaterialTests(TestCase):
    """Схемы, инструкции и методики — ссылками у модели.

    Ссылка, а не файл: материалы правят на Диске, и копия в программе
    разошлась бы с оригиналом молча, а схема в приличном разрешении весит
    десятки мегабайт — программа живёт на карте памяти Raspberry Pi.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='materials_admin', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        self.other_version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.7'
        )

        self.common = EquipmentMaterial.objects.create(
            equipment_model=self.model, kind='manual',
            title='Инструкция по настройке',
            url='https://disk.yandex.ru/d/manual',
        )
        self.for_version = EquipmentMaterial.objects.create(
            equipment_model=self.model, version=self.version, kind='scheme',
            title='Схема исполнения .4',
            url='https://disk.yandex.ru/d/scheme-4',
        )

    def test_a_version_gets_the_common_material_too(self):
        """Вытеснения нет, и это отличие от рецепта деталей намеренное:
        общая инструкция нужна и тогда, когда у исполнения своя схема."""
        materials = set(self.model.materials_for(self.version))

        self.assertEqual(materials, {self.common, self.for_version})

    def test_another_version_does_not_see_a_foreign_scheme(self):
        materials = list(self.model.materials_for(self.other_version))

        self.assertEqual(materials, [self.common])

    def test_a_unit_without_a_version_sees_only_the_common_ones(self):
        """Исполнение не указано — уточнений не показываем: схема
        не того исполнения хуже, чем её отсутствие."""
        materials = list(self.model.materials_for(None))

        self.assertEqual(materials, [self.common])

    def test_a_version_of_another_model_is_refused(self):
        """Материал с чужим исполнением не показался бы никогда: отбор
        идёт от модели той единицы, что лежит на столе."""
        other_model = EquipmentModel.objects.create(name='EkoDrive 2.0')
        alien = EquipmentVersion.objects.create(
            equipment_model=other_model, name='-1.1'
        )
        material = EquipmentMaterial(
            equipment_model=self.model, version=alien,
            title='Чужая схема', url='https://example.com/x',
        )

        with self.assertRaises(ValidationError) as caught:
            material.full_clean()

        self.assertIn('version', caught.exception.error_dict)


class EquipmentMaterialScreensTests(TestCase):
    """Где материалы заводят и где их видят."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='materials_screens', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        self.equipment = Equipment.objects.create(
            model=self.model, version=self.version, serial_number='SN-MAT-1',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment
        )

    def _edit_url(self):
        return reverse('equipment_model_edit', args=[self.model.pk])

    def _defect_url(self):
        return reverse('repair_order_defect_act_edit',
                       args=[self.order.pk, self.roe.pk])

    def test_the_new_model_page_has_no_materials(self):
        """Пока модели нет, вешать ссылки не на что — и страница так
        и говорит, вместо пустой таблицы, которая ничего не сохранит."""
        html = self.http.get(reverse('equipment_model_create')).content.decode()

        self.assertNotIn('name="materials-TOTAL_FORMS"', html)
        self.assertIn('добавляются после сохранения', html)

    def test_a_material_is_added_from_the_model_card(self):
        response = self.http.post(self._edit_url(), {
            'name': self.model.name, 'kind': '', 'equipment_type': '',
            'materials-TOTAL_FORMS': '1', 'materials-INITIAL_FORMS': '0',
            'materials-MIN_NUM_FORMS': '0', 'materials-MAX_NUM_FORMS': '1000',
            'materials-0-kind': 'scheme',
            'materials-0-title': 'Схема принципиальная',
            'materials-0-url': 'https://disk.yandex.ru/d/scheme',
            'materials-0-version': str(self.version.pk),
            'materials-0-note': 'лист 2 — блок питания',
        })

        self.assertEqual(response.status_code, 302)
        material = self.model.materials.get()
        self.assertEqual(material.title, 'Схема принципиальная')
        self.assertEqual(material.version, self.version)

    def test_a_broken_link_saves_nothing(self):
        """Не ссылка — не материал: «схема на диске» словами открывать
        нечем, а молча сохранённая строка выглядит как рабочая кнопка."""
        response = self.http.post(self._edit_url(), {
            'name': self.model.name, 'kind': '', 'equipment_type': '',
            'materials-TOTAL_FORMS': '1', 'materials-INITIAL_FORMS': '0',
            'materials-MIN_NUM_FORMS': '0', 'materials-MAX_NUM_FORMS': '1000',
            'materials-0-kind': 'scheme',
            'materials-0-title': 'Схема',
            'materials-0-url': 'схема лежит у Николая',
            'materials-0-version': '',
            'materials-0-note': '',
        })

        self.assertEqual(response.status_code, 200)
        self.assertFalse(self.model.materials.exists())

    def test_only_this_models_versions_are_offered(self):
        """Чужое исполнение в выборе означало бы материал, который
        не покажется никогда."""
        other_model = EquipmentModel.objects.create(name='EkoDrive 2.0')
        alien = EquipmentVersion.objects.create(
            equipment_model=other_model, name='-1.1'
        )

        html = self.http.get(self._edit_url()).content.decode()

        self.assertIn('>%s</option>' % self.version, html)
        self.assertNotIn(str(alien), html)

    def test_the_diagnostics_page_shows_the_materials(self):
        """То, ради чего всё и заводилось: мастер сидит с прибором
        и актом, и схема нужна ему ровно в эту минуту."""
        EquipmentMaterial.objects.create(
            equipment_model=self.model, version=self.version, kind='scheme',
            title='Схема исполнения .4', url='https://disk.yandex.ru/d/s4',
        )
        EquipmentMaterial.objects.create(
            equipment_model=self.model, kind='manual',
            title='Инструкция', url='https://disk.yandex.ru/d/manual',
        )

        html = self.http.get(self._defect_url()).content.decode()

        self.assertIn('Схема исполнения .4', html)
        self.assertIn('Инструкция', html)
        self.assertIn('https://disk.yandex.ru/d/s4', html)

    def test_materials_of_another_version_stay_away(self):
        other = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.7'
        )
        EquipmentMaterial.objects.create(
            equipment_model=self.model, version=other, kind='scheme',
            title='Схема исполнения .7', url='https://disk.yandex.ru/d/s7',
        )

        html = self.http.get(self._defect_url()).content.decode()

        self.assertNotIn('Схема исполнения .7', html)

    def test_no_materials_says_where_to_add_them(self):
        """Пустая карточка без объяснения выглядит как поломка."""
        html = self.http.get(self._defect_url()).content.decode()

        self.assertIn('материалов пока нет', html)
        self.assertIn(self._edit_url(), html)

    def test_the_scheme_comes_first(self):
        """По алфавиту кода схема оказывалась последней, а за столом
        её открывают первой."""
        EquipmentMaterial.objects.create(
            equipment_model=self.model, kind='other',
            title='Прочее', url='https://example.com/o',
        )
        EquipmentMaterial.objects.create(
            equipment_model=self.model, kind='manual',
            title='Инструкция', url='https://example.com/m',
        )
        scheme = EquipmentMaterial.objects.create(
            equipment_model=self.model, kind='scheme',
            title='Схема', url='https://example.com/s',
        )

        self.assertEqual(self.model.materials_for(None).first(), scheme)

    def test_links_open_in_a_new_tab(self):
        """Акт заполнен наполовину — уводить с него страницу нельзя."""
        EquipmentMaterial.objects.create(
            equipment_model=self.model, kind='manual',
            title='Инструкция', url='https://disk.yandex.ru/d/manual',
        )

        html = self.http.get(self._defect_url()).content.decode()

        self.assertIn('target="_blank"', html)
        self.assertIn('rel="noopener"', html)


# ==================== ТЕХНОЛОГИЧЕСКИЕ КАРТЫ (v2.69.0) ====================


class TechCardTests(TestCase):
    """Карта — «как это делают руками»: разобрать корпус, проверить
    на стенде, заменить высохшие конденсаторы.

    Привязана к модели всегда, к неисправности — по желанию: так решил
    владелец, и это ответ на давний открытый вопрос. «Как разобрать
    корпус» не про поломку, а про прибор.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='techcard_admin', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        self.other_version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.7'
        )
        self.fault = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            description='Электролитические конденсаторы потеряли ёмкость.',
        )

    def _card(self, **kwargs):
        kwargs.setdefault('equipment_model', self.model)
        kwargs.setdefault('title', 'Разборка корпуса')
        return TechCard.objects.create(**kwargs)

    def test_a_card_without_a_fault_is_allowed(self):
        """То, ради чего вопрос и задавался: «как разобрать корпус»
        не привязана ни к какой поломке."""
        card = self._card()

        card.full_clean()
        self.assertIsNone(card.fault_type)

    def test_a_fault_of_another_model_is_refused(self):
        """Карта висела бы у одного прибора, а предлагалась при поломке
        другого."""
        other = EquipmentModel.objects.create(name='EkoDrive 2.0')
        alien = FaultType.objects.create(equipment_model=other, name='иная')
        card = TechCard(
            equipment_model=self.model, fault_type=alien, title='Чужая'
        )

        with self.assertRaises(ValidationError) as caught:
            card.full_clean()

        self.assertIn('fault_type', caught.exception.error_dict)

    def test_general_cards_come_before_fault_ones(self):
        """Прибор разбирают раньше, чем чинят."""
        self._card(fault_type=self.fault, title='Замена конденсаторов')
        general = self._card(title='Разборка корпуса')

        self.assertEqual(self.model.tech_cards.first(), general)

    def test_a_step_of_a_version_shows_only_for_it(self):
        card = self._card()
        common = TechCardStep.objects.create(
            card=card, number=10, text='Снять крышку'
        )
        special = TechCardStep.objects.create(
            card=card, number=20, text='Отвернуть четыре винта',
            version=self.version,
        )

        self.assertEqual(list(card.steps_for(self.version)), [common, special])
        self.assertEqual(list(card.steps_for(self.other_version)), [common])
        self.assertEqual(list(card.steps_for(None)), [common])

    def test_a_step_version_of_another_model_is_refused(self):
        other = EquipmentModel.objects.create(name='EkoDrive 2.0')
        alien = EquipmentVersion.objects.create(
            equipment_model=other, name='-1.1'
        )
        card = self._card()
        step = TechCardStep(card=card, number=10, text='шаг', version=alien)

        with self.assertRaises(ValidationError) as caught:
            step.full_clean()

        self.assertIn('version', caught.exception.error_dict)

    def test_a_deleted_fault_does_not_take_the_card_with_it(self):
        """Неисправность убрали из справочника — описанная процедура
        никуда не делась, и терять её нельзя."""
        card = self._card(fault_type=self.fault, title='Замена конденсаторов')

        self.fault.delete()

        card.refresh_from_db()
        self.assertIsNone(card.fault_type)


class TechCardScreensTests(TestCase):
    """Где карты заводят, где печатают и где видят при работе."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='techcard_screens', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        self.other_version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.7'
        )
        self.equipment = Equipment.objects.create(
            model=self.model, version=self.version, serial_number='SN-TC-1',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment
        )
        self.card = TechCard.objects.create(
            equipment_model=self.model, title='Разборка корпуса',
            purpose='Перед любой работой внутри.',
        )
        self.common_step = TechCardStep.objects.create(
            card=self.card, number=10, text='Снять крышку',
            caution='сначала разрядить конденсаторы',
        )
        self.version_step = TechCardStep.objects.create(
            card=self.card, number=20, text='Отвернуть четыре винта',
            version=self.version,
        )

    def _defect_url(self):
        return reverse('repair_order_defect_act_edit',
                       args=[self.order.pk, self.roe.pk])

    def test_a_card_is_created_with_its_steps(self):
        response = self.http.post(reverse('tech_card_create'), {
            'equipment_model': str(self.model.pk), 'fault_type': '',
            'title': 'Проверка на стенде', 'purpose': '',
            'steps-TOTAL_FORMS': '1', 'steps-INITIAL_FORMS': '0',
            'steps-MIN_NUM_FORMS': '0', 'steps-MAX_NUM_FORMS': '1000',
            'steps-0-number': '10', 'steps-0-text': 'Подключить питание',
            'steps-0-version': '', 'steps-0-caution': '',
        })

        self.assertEqual(response.status_code, 302)
        card = TechCard.objects.get(title='Проверка на стенде')
        self.assertEqual(card.steps.get().text, 'Подключить питание')

    def test_a_new_card_offers_no_versions_yet(self):
        """Исполнения берутся у модели, а модель выбирают в этой же форме:
        предлагать чужие исполнения нельзя — шаг с ним не покажется никогда."""
        html = self.http.get(reverse('tech_card_create')).content.decode()

        self.assertNotIn(str(self.version), html)

    def test_editing_offers_only_this_models_versions(self):
        other_model = EquipmentModel.objects.create(name='EkoDrive 2.0')
        alien = EquipmentVersion.objects.create(
            equipment_model=other_model, name='-1.1'
        )

        html = self.http.get(
            reverse('tech_card_edit', args=[self.card.pk])
        ).content.decode()

        self.assertIn('>%s</option>' % self.version, html)
        self.assertNotIn(str(alien), html)

    def test_the_printed_card_shows_only_asked_for_steps(self):
        url = reverse('tech_card_detail', args=[self.card.pk])

        for_version = self.http.get(url, {'version': self.version.pk}).content.decode()
        for_other = self.http.get(url, {'version': self.other_version.pk}).content.decode()
        bare = self.http.get(url).content.decode()

        self.assertIn('Отвернуть четыре винта', for_version)
        self.assertNotIn('Отвернуть четыре винта', for_other)
        self.assertNotIn('Отвернуть четыре винта', bare)
        for html in (for_version, for_other, bare):
            self.assertIn('Снять крышку', html)

    def test_a_caution_is_printed_with_a_word_not_only_a_colour(self):
        """На бумаге цвет уходит в серый, и предостережение нужно узнавать
        по слову, а не по оттенку."""
        html = self.http.get(
            reverse('tech_card_detail', args=[self.card.pk])
        ).content.decode()

        self.assertIn('Внимание: сначала разрядить конденсаторы', html)

    def test_an_unknown_version_says_so_instead_of_pretending(self):
        """Молчание тут означало бы, что мастер считает карту полной,
        а половины шагов не видит."""
        other_model = EquipmentModel.objects.create(name='EkoDrive 2.0')
        alien = EquipmentVersion.objects.create(
            equipment_model=other_model, name='-1.1'
        )

        html = self.http.get(
            reverse('tech_card_detail', args=[self.card.pk]),
            {'version': alien.pk},
        ).content.decode()

        self.assertIn('Такого исполнения у модели', html)
        self.assertNotIn('Отвернуть четыре винта', html)

    def test_diagnostics_links_the_card_with_the_units_version(self):
        """Без исполнения в ссылке мастер увидел бы половину карты
        и не понял бы, что половины нет."""
        html = self.http.get(self._defect_url()).content.decode()

        self.assertIn('Разборка корпуса', html)
        self.assertIn(
            '%s?version=%d' % (
                reverse('tech_card_detail', args=[self.card.pk]), self.version.pk
            ),
            html,
        )

    def test_diagnostics_without_cards_says_where_to_add_them(self):
        TechCard.objects.all().delete()

        html = self.http.get(self._defect_url()).content.decode()

        self.assertIn('карт пока нет', html)
        self.assertIn(reverse('tech_card_create'), html)


# ============ СНИМКИ К ШАГАМ КАРТЫ И РЕЗЕРВНАЯ КОПИЯ (v2.70.0) ============


def _photo_bytes(size=(4000, 3000), fmt='JPEG', mode='RGB'):
    """Снимок нужного размера — как приезжает с телефона."""
    from PIL import Image
    buffer = io.BytesIO()
    Image.new(mode, size, 'red').save(buffer, format=fmt)
    return buffer.getvalue()


@override_settings(MEDIA_ROOT=tempfile.mkdtemp(prefix='lifteam-test-media-'))
class TechCardImageTests(TestCase):
    """Снимок к шагу — единственный файл, который программа хранит у себя.

    Схемы и инструкции по-прежнему ссылками: их правят на Диске, и копия
    разошлась бы с оригиналом. Снимок шага никто не правит — он показывает,
    как это выглядит на столе.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='techcard_photo', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.card = TechCard.objects.create(
            equipment_model=self.model, title='Разборка корпуса'
        )
        self.step = TechCardStep.objects.create(
            card=self.card, number=10, text='Снять крышку'
        )

    def _post(self, upload):
        return self.http.post(
            reverse('tech_card_edit', args=[self.card.pk]),
            {
                'equipment_model': str(self.model.pk), 'fault_type': '',
                'title': self.card.title, 'purpose': '',
                'steps-TOTAL_FORMS': '1', 'steps-INITIAL_FORMS': '1',
                'steps-MIN_NUM_FORMS': '0', 'steps-MAX_NUM_FORMS': '1000',
                'steps-0-id': str(self.step.pk), 'steps-0-card': str(self.card.pk),
                'steps-0-number': '10', 'steps-0-text': 'Снять крышку',
                'steps-0-version': '', 'steps-0-caution': '',
                'steps-0-image': upload,
            },
        )

    def test_a_phone_photo_is_shrunk_before_it_is_stored(self):
        """С телефона приезжает четыре мегабайта, а программа живёт
        на карте памяти и выгружает копии по домашнему каналу."""
        from PIL import Image
        original = _photo_bytes()
        upload = SimpleUploadedFile('IMG_0042.jpg', original, 'image/jpeg')

        response = self._post(upload)

        self.assertEqual(response.status_code, 302)
        self.step.refresh_from_db()
        self.assertTrue(self.step.image)
        with Image.open(self.step.image.path) as stored:
            self.assertEqual(max(stored.size), 1600)
        self.assertLess(self.step.image.size, len(original))

    def test_a_png_stays_a_png(self):
        """PNG обычно приходит скриншотом схемы: перегон в JPEG размыл бы
        подписи на выводах."""
        upload = SimpleUploadedFile(
            'scheme.png', _photo_bytes(size=(2000, 1200), fmt='PNG'), 'image/png'
        )

        self._post(upload)

        self.step.refresh_from_db()
        self.assertTrue(self.step.image.name.endswith('.png'))

    def test_a_small_photo_is_not_enlarged(self):
        from PIL import Image
        upload = SimpleUploadedFile(
            'small.jpg', _photo_bytes(size=(800, 600)), 'image/jpeg'
        )

        self._post(upload)

        self.step.refresh_from_db()
        with Image.open(self.step.image.path) as stored:
            self.assertEqual(stored.size, (800, 600))

    def test_a_huge_file_is_refused_with_a_reason(self):
        """Схема на сорок мегабайт — не снимок шага, и её место
        на Диске ссылкой.

        Порог на время проверки занижен: настоящие пятнадцать мегабайт
        пришлось бы собирать из шума, а проверяется здесь правило,
        а не число.
        """
        upload = SimpleUploadedFile(
            'scheme.jpg', _photo_bytes(size=(1200, 900)), 'image/jpeg'
        )

        with patch('core.forms.MAX_UPLOAD_BYTES', 1024):
            response = self._post(upload)

        self.assertEqual(response.status_code, 200)
        self.step.refresh_from_db()
        self.assertFalse(self.step.image)
        self.assertContains(response, 'вешаются ссылкой в карточке модели')

    def test_the_limit_is_a_phone_photo_not_a_scheme(self):
        """Порог — про снимок с телефона: он укладывается с запасом,
        а схема на десятки мегабайт нет, и её место ссылкой."""
        self.assertEqual(MAX_UPLOAD_BYTES, 15 * 1024 * 1024)

    def test_saving_the_card_again_does_not_touch_the_photo(self):
        """Поле не трогали — приходит уже сохранённый файл, а не загрузка:
        уменьшать его второй раз незачем."""
        self._post(SimpleUploadedFile('IMG.jpg', _photo_bytes(), 'image/jpeg'))
        self.step.refresh_from_db()
        stored_name, stored_size = self.step.image.name, self.step.image.size

        self.http.post(reverse('tech_card_edit', args=[self.card.pk]), {
            'equipment_model': str(self.model.pk), 'fault_type': '',
            'title': self.card.title, 'purpose': '',
            'steps-TOTAL_FORMS': '1', 'steps-INITIAL_FORMS': '1',
            'steps-MIN_NUM_FORMS': '0', 'steps-MAX_NUM_FORMS': '1000',
            'steps-0-id': str(self.step.pk), 'steps-0-card': str(self.card.pk),
            'steps-0-number': '10', 'steps-0-text': 'Снять крышку и отложить',
            'steps-0-version': '', 'steps-0-caution': '',
        })

        self.step.refresh_from_db()
        self.assertEqual(self.step.text, 'Снять крышку и отложить')
        self.assertEqual(self.step.image.name, stored_name)
        self.assertEqual(self.step.image.size, stored_size)

    def test_the_photo_is_printed_on_the_card(self):
        self._post(SimpleUploadedFile('IMG.jpg', _photo_bytes(), 'image/jpeg'))
        self.step.refresh_from_db()

        html = self.http.get(
            reverse('tech_card_detail', args=[self.card.pk])
        ).content.decode()

        self.assertIn(self.step.image.url, html)
        self.assertIn('step-image', html)

    def test_the_form_can_carry_files_at_all(self):
        """Без enctype снимки не доехали бы до сервера вовсе, а форма
        сохранилась бы молча — как будто картинку не выбирали."""
        html = self.http.get(
            reverse('tech_card_edit', args=[self.card.pk])
        ).content.decode()

        self.assertIn('enctype="multipart/form-data"', html)


class MediaInBackupTests(SimpleTestCase):
    """Снимки уходят в резервную копию вместе с базой.

    В базе только имя файла. Копия базы без снимков восстановит карты
    без картинок, и заметно это станет у стола, когда картинка понадобится.
    """

    @property
    def script(self):
        return (settings.BASE_DIR / 'deploy/backup.sh').read_text(encoding='utf-8')

    def test_media_is_uploaded_too(self):
        self.assertIn('rclone copy "${APP_DIR}/media"', self.script)

    def test_media_goes_with_copy_not_sync(self):
        """sync зеркалит удаления: файл, удалённый здесь по ошибке,
        исчез бы и в облаке — то есть ровно тогда, когда копия и нужна."""
        self.assertNotIn('rclone sync', self.script)

    def test_a_missing_media_directory_is_not_an_error(self):
        """Карт со снимками ещё не завели — это не повод падать
        и оставлять базу без выгрузки."""
        self.assertIn('if [[ -d "${APP_DIR}/media" ]]; then', self.script)

    def test_an_already_configured_install_needs_no_edits(self):
        """Выгрузка снимков включается сама там, где выгрузка базы уже
        настроена: править юнит на работающей установке никто не пойдёт."""
        self.assertIn(
            'RCLONE_MEDIA_REMOTE="${LIFTEAM_RCLONE_MEDIA_REMOTE:-'
            '${RCLONE_REMOTE:+${RCLONE_REMOTE}/media}}"',
            self.script,
        )

    def test_a_saved_message_does_not_go_to_paper(self):
        """Печатают сразу после сохранения, и «Карта сохранена» уходило
        на бумагу первой строкой, над самим документом. Правило своё,
        не бутстраповское: Bootstrap приходит из интернета."""
        base = (settings.BASE_DIR / 'core/templates/core/base.html').read_text(encoding='utf-8')

        self.assertIn("alert-dismissible fade show no-print", base)
        self.assertIn('.no-print { display: none !important; }', base)

    def test_nginx_serves_media(self):
        """Снимок, который некому отдать, на странице карты не появится."""
        conf = (settings.BASE_DIR / 'deploy/nginx-lifteam.conf').read_text(encoding='utf-8')

        self.assertIn('location /media/', conf)
        self.assertIn('alias /opt/lifteam/media/;', conf)


# ============ ГОТОВНОСТЬ ЕДИНИЦЫ И ЧЕК-ЛИСТ (v2.71.0) ============


class UnitReadinessTests(TestCase):
    """Что по каждому прибору сделано, а что осталось.

    Заказ один, а приборов в нём пять, и идут они не в ногу. Статус заказа
    об этом не говорит ничего: перед отгрузкой приходилось открывать
    единицы по очереди и вспоминать, что по ним не заполнено.

    Отдельного поля состояния у единицы нет намеренно: любое такое поле
    проставляют руками и забывают, и оно начинает врать.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='readiness_admin', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-READY-1'
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment
        )

    def _codes(self, roe=None):
        return [check['code'] for check in (roe or self.roe).readiness_pending]

    def _finish(self, roe=None):
        """Заполнить всё, что чек-лист спрашивает."""
        roe = roe or self.roe
        roe.diagnosis = 'Высохли конденсаторы в цепи питания.'
        roe.work_performed = 'Заменены конденсаторы C12, C13.'
        # Стоимость по факту, а не оценка из дефектации: по ней считается
        # сумма заказа и выставляется счёт
        roe.repair_cost = Decimal('4500.00')
        roe.save()
        return roe

    def test_a_fresh_unit_is_not_ready_and_says_why(self):
        self.assertFalse(self.roe.is_ready)
        # Запланированных деталей нет — и требовать по ним нечего:
        # пункт есть в списке проверок, но выполнен. Порядок — по ходу
        # работы: вскрыли, починили, взяли детали, посчитали
        self.assertEqual(self._codes(), ['defect_act', 'work', 'repair_cost'])
        self.assertEqual(len(self.roe.readiness()), 4)

    def test_a_filled_unit_is_ready(self):
        self._finish()

        self.assertTrue(self.roe.is_ready)
        self.assertEqual(self.roe.readiness_label, 'Готов')

    def test_a_warranty_repair_is_not_asked_for_a_price(self):
        """«Назначьте стоимость» на гарантийном ремонте — ложная тревога,
        а к ложным тревогам привыкают и перестают читать все остальные."""
        self.roe.warranty_case = 'warranty'
        self.roe.diagnosis = 'Отказал драйвер.'
        self.roe.work_performed = 'Заменён драйвер.'
        self.roe.save()

        self.assertNotIn('repair_cost', [c['code'] for c in self.roe.readiness()])
        self.assertTrue(self.roe.is_ready)

    def test_a_non_warranty_repair_is_asked_for_a_price(self):
        self.roe.warranty_case = 'non_warranty'
        self.roe.save()

        self.assertIn('repair_cost', self._codes())

    def test_a_planned_part_keeps_the_unit_unfinished(self):
        """Деталь числится нужной по ремонту, но со склада не взята —
        значит работа не закончена, чем бы ни были заполнены поля."""
        self._finish()
        part = SparePart.objects.create(
            part_number='C-100u', name='Конденсатор 100 мкФ', current_stock=0
        )
        RepairOrderDetail.objects.create(
            repair_order=self.order, order_equipment=self.roe,
            part=part, quantity_used=2, is_planned=True,
        )

        self.roe.refresh_from_db()
        self.assertEqual(self._codes(), ['planned_parts'])
        self.assertEqual(self.roe.readiness_label, 'Осталось 1')

    def test_a_written_off_part_does_not_keep_it_unfinished(self):
        self._finish()
        part = SparePart.objects.create(
            part_number='C-100u', name='Конденсатор 100 мкФ', current_stock=10
        )
        RepairOrderDetail.objects.create(
            repair_order=self.order, order_equipment=self.roe,
            part=part, quantity_used=2, is_planned=False,
        )

        self.roe.refresh_from_db()
        self.assertTrue(self.roe.is_ready)

    def test_whitespace_is_not_recorded_work(self):
        """Пробел в поле — это не запись о работе."""
        self.roe.work_performed = '   \n '
        self.roe.save()

        self.assertIn('work', self._codes())

    def test_the_order_is_ready_only_when_every_unit_is(self):
        """Отгружают коробку, а не статус."""
        self._finish()
        second = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-READY-2'
            ),
        )

        readiness = self.order.readiness()
        self.assertEqual(readiness['total'], 2)
        self.assertEqual(readiness['ready'], 1)
        self.assertEqual(readiness['pending'], [second])
        self.assertFalse(readiness['all_ready'])

    def test_an_order_without_equipment_is_not_ready(self):
        """Отгружать нечего — это не «готово»."""
        empty = RepairOrder.objects.create(client=self.order.client)

        self.assertFalse(empty.readiness()['all_ready'])

    def test_every_check_has_a_reason_written_down(self):
        """Пункт без объяснения читается как придирка программы."""
        for check in self.roe.readiness():
            with self.subTest(code=check['code']):
                self.assertTrue(check['label'])
                self.assertTrue(check['hint'])


class ReadinessScreensTests(TestCase):
    """Где готовность видна и где о ней предупреждают."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='readiness_screens', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»'),
            status='ready_for_shipment',
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-SCREEN-1'
            ),
        )

    def test_the_order_card_lists_what_is_left(self):
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

        self.assertIn('Готовность к отгрузке', html)
        self.assertIn('Заполнена дефектация', html)
        self.assertIn('Осталось 3', html)

    def test_a_finished_order_says_so_instead_of_an_empty_card(self):
        self.roe.diagnosis = 'Высохли конденсаторы.'
        self.roe.work_performed = 'Заменены конденсаторы.'
        self.roe.repair_cost = Decimal('4500.00')
        self.roe.save()

        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

        self.assertIn('заполнено всё', html.lower())
        self.assertIn('Готов', html)

    def test_shipping_warns_but_does_not_forbid(self):
        """Мастер видит прибор, а программа нет. Запрет означал бы, что
        статус проставляют «как-нибудь», лишь бы программа пропустила."""
        response = self.http.post(
            reverse('repair_order_change_status', args=[self.order.pk]),
            {'new_status': 'shipped', 'notes': ''}, follow=True,
        )

        self.order.refresh_from_db()
        self.assertEqual(self.order.status, 'shipped')
        texts = [str(m) for m in response.context['messages']]
        self.assertTrue(
            any('Осталась работа по оборудованию' in text for text in texts),
            texts,
        )
        self.assertTrue(any('SN-SCREEN-1' in text for text in texts), texts)

    def test_no_warning_when_everything_is_filled(self):
        self.roe.diagnosis = 'Высохли конденсаторы.'
        self.roe.work_performed = 'Заменены конденсаторы.'
        self.roe.repair_cost = Decimal('4500.00')
        self.roe.save()

        response = self.http.post(
            reverse('repair_order_change_status', args=[self.order.pk]),
            {'new_status': 'shipped', 'notes': ''}, follow=True,
        )

        texts = [str(m) for m in response.context['messages']]
        self.assertFalse(any('Осталась работа' in text for text in texts), texts)

    def test_no_warning_on_an_ordinary_status_step(self):
        """Недоделки на пути «в ремонт» — это не недоделки, а работа,
        которую ещё не начали."""
        self.order.status = 'accepted'
        self.order.save()

        response = self.http.post(
            reverse('repair_order_change_status', args=[self.order.pk]),
            {'new_status': 'repair', 'notes': ''}, follow=True,
        )

        texts = [str(m) for m in response.context['messages']]
        self.assertFalse(any('Осталась работа' in text for text in texts), texts)

    def test_bulk_shipment_shows_readiness_before_it_happens(self):
        html = self.http.get(
            reverse('repair_order_bulk_status'), {'ids': str(self.order.pk)}
        ).content.decode()

        self.assertIn('осталась работа по 1 из 1', html)
        # и всё-таки отгрузить даёт
        self.assertNotIn('disabled', html.split('Отгрузить')[0][-200:])


# ==================== ЯНДЕКС.ДИСК: ПАПКА ЗАКАЗА (v2.72.0) ====================


class FakeDiskResponse:
    """Ответ Диска для подмены urlopen."""

    def __init__(self, status=201, body=''):
        self.status = status
        self._body = body.encode('utf-8')

    def read(self):
        return self._body

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False


def _disk_conflict(url):
    """409 «папка уже есть» — обычный ход дела, а не ошибка."""
    return urllib_error.HTTPError(
        url, 409, 'Conflict', {},
        io.BytesIO(json.dumps({
            'message': 'Ресурс "/LiftTeam" уже существует.',
            'error': 'DiskPathPointsToExistentDirectoryError',
        }).encode('utf-8')),
    )


@override_settings(
    YANDEX_DISK_TOKEN='test-disk-token',
    YANDEX_DISK_ROOT='LiftTeam',
)
class YandexDiskClientTests(TestCase):
    """Клиент Диска. На живом Диске не проверен ничего — сайт документации
    Яндекса из среды разработки недоступен, как и у Точки; здесь
    проверяется то, что проверить можно: пути, заголовки, разбор ответов
    и поведение при 409.
    """

    def setUp(self):
        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-DISK-1'
            ),
        )

    def _urlopen(self, calls, statuses=None):
        """Подмена urlopen: складывает запросы и отдаёт заданные ответы."""
        queue = list(statuses or [])

        def fake(req, timeout=None):
            calls.append(req)
            status = queue.pop(0) if queue else 201
            if status == 409:
                raise _disk_conflict(req.full_url)
            return FakeDiskResponse(status)

        return fake

    def test_the_path_is_built_in_one_place(self):
        """Разойдись он между кнопкой и проверкой — снимки одного ремонта
        уехали бы в папку другого."""
        self.assertEqual(
            yadisk.unit_path(self.roe),
            'LiftTeam/Заказы/%s/SN-DISK-1' % self.order.order_number,
        )

    def test_a_slash_in_a_serial_number_does_not_become_a_folder(self):
        """Косая черта на Диске означает новый уровень пути: она тихо
        превратила бы одну папку в две."""
        self.roe.equipment.serial_number = 'SN/12:34'
        self.roe.equipment.save()

        path = yadisk.unit_path(self.roe)

        self.assertTrue(path.endswith('SN_12_34'), path)
        self.assertEqual(path.count('/'), 3)

    def test_an_empty_name_still_gives_a_folder(self):
        self.assertEqual(yadisk.safe_name('   '), 'без-номера')
        self.assertEqual(yadisk.safe_name('...'), 'без-номера')

    def test_parents_are_created_top_down(self):
        """Диск родителей сам не заводит."""
        calls = []
        with patch('core.yadisk.request.urlopen', self._urlopen(calls)):
            created = yadisk.ensure_folder('LiftTeam/Заказы/LT-1/SN-1')

        paths = [urllib_parse.parse_qs(urllib_parse.urlsplit(c.full_url).query)['path'][0]
                 for c in calls]
        self.assertEqual(paths, [
            'disk:/LiftTeam',
            'disk:/LiftTeam/Заказы',
            'disk:/LiftTeam/Заказы/LT-1',
            'disk:/LiftTeam/Заказы/LT-1/SN-1',
        ])
        self.assertEqual(len(created), 4)

    def test_an_existing_folder_is_not_an_error(self):
        """Двое мастеров могут нажать кнопку почти одновременно, и вторым
        ответом будет «уже есть»."""
        calls = []
        fake = self._urlopen(calls, statuses=[409, 409, 201, 201])
        with patch('core.yadisk.request.urlopen', fake):
            created = yadisk.ensure_folder('LiftTeam/Заказы/LT-1/SN-1')

        self.assertEqual(len(calls), 4)
        self.assertEqual(created, ['LiftTeam/Заказы/LT-1', 'LiftTeam/Заказы/LT-1/SN-1'])

    def test_the_token_goes_in_the_oauth_header_not_bearer(self):
        """У Диска свой заголовок: с Bearer он отвечает 401."""
        calls = []
        with patch('core.yadisk.request.urlopen', self._urlopen(calls)):
            yadisk.create_folder('LiftTeam')

        self.assertEqual(calls[0].get_header('Authorization'), 'OAuth test-disk-token')
        self.assertEqual(calls[0].get_method(), 'PUT')

    def test_a_bad_token_is_named_by_its_reason(self):
        """401 чаще всего значит просроченный токен, и человек должен
        понять, что чинить, не читая кода ответа."""
        def fake(req, timeout=None):
            raise urllib_error.HTTPError(
                req.full_url, 401, 'Unauthorized', {},
                io.BytesIO(b'{"message": "Unauthorized"}'),
            )

        with patch('core.yadisk.request.urlopen', fake):
            with self.assertRaises(yadisk.YandexDiskError) as caught:
                yadisk.create_folder('LiftTeam')

        self.assertIn('не принял токен', str(caught.exception))
        self.assertIn('YANDEX_DISK_TOKEN', str(caught.exception))

    def test_the_token_never_reaches_the_log_or_the_message(self):
        """Журнал живёт до ротации и уезжает в резервную копию — токен
        в нём перестал бы быть секретом."""
        def fake(req, timeout=None):
            raise urllib_error.HTTPError(
                req.full_url, 500, 'Server Error', {},
                io.BytesIO(b'{"message": "\\u0442\\u043e\\u043a\\u0435\\u043d test-disk-token"}'),
            )

        with self.assertLogs('core.yadisk', level='INFO') as logs:
            with patch('core.yadisk.request.urlopen', fake):
                with self.assertRaises(yadisk.YandexDiskError) as caught:
                    yadisk.create_folder('LiftTeam')

        self.assertNotIn('test-disk-token', '\n'.join(logs.output))
        self.assertNotIn('test-disk-token', str(caught.exception))

    def test_there_is_no_way_to_delete_anything(self):
        """Ни одной функции, стирающей на Диске: та же причина, по которой
        в банковских модулях нет платёжных поручений."""
        source = (settings.BASE_DIR / 'core/yadisk.py').read_text(encoding='utf-8')

        self.assertNotIn("'DELETE'", source)
        self.assertFalse([name for name in dir(yadisk)
                          if 'delete' in name or 'remove' in name])


class YandexDiskFolderButtonTests(TestCase):
    """Кнопка «завести папку» на карточке заказа."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='disk_button', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-DISK-1'
            ),
        )

    def _url(self):
        return reverse('repair_order_unit_disk_folder',
                       args=[self.order.pk, self.roe.pk])

    @override_settings(YANDEX_DISK_TOKEN='')
    def test_without_a_token_it_says_what_is_missing(self):
        """Молчаливое бездействие человек примет за поломку кнопки."""
        response = self.http.post(self._url(), follow=True)

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.yandex_disk_folder, '')
        texts = [str(m) for m in response.context['messages']]
        self.assertTrue(any('YANDEX_DISK_TOKEN' in t for t in texts), texts)

    @override_settings(YANDEX_DISK_TOKEN='test-disk-token')
    def test_the_link_is_written_into_the_unit(self):
        with patch('core.yadisk.request.urlopen',
                   lambda req, timeout=None: FakeDiskResponse(201)):
            self.http.post(self._url(), follow=True)

        self.roe.refresh_from_db()
        self.assertIn('disk.yandex.ru/client/disk/LiftTeam',
                      self.roe.yandex_disk_folder)
        self.assertIn('SN-DISK-1', self.roe.yandex_disk_folder)

    @override_settings(YANDEX_DISK_TOKEN='test-disk-token')
    def test_a_failure_is_reported_and_nothing_is_written(self):
        def fake(req, timeout=None):
            raise urllib_error.URLError('нет связи')

        with patch('core.yadisk.request.urlopen', fake):
            response = self.http.post(self._url(), follow=True)

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.yandex_disk_folder, '')
        texts = [str(m) for m in response.context['messages']]
        self.assertTrue(any('Папка не создана' in t for t in texts), texts)

    def test_the_button_only_shows_while_there_is_no_folder(self):
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()
        self.assertIn('diskFolder%d' % self.roe.pk, html)

        self.roe.yandex_disk_folder = 'https://disk.yandex.ru/client/disk/LiftTeam'
        self.roe.save()

        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()
        self.assertNotIn('diskFolder%d' % self.roe.pk, html)
        self.assertIn('https://disk.yandex.ru/client/disk/LiftTeam', html)

    def test_only_post_creates_a_folder(self):
        """Ссылку с этим адресом мог бы открыть кто угодно, включая
        обходчик поисковика."""
        self.assertEqual(self.http.get(self._url()).status_code, 405)


# ==================== КАМЕРА КАК ВТОРОЙ СКАНЕР (v2.73.0) ====================


class CameraScanTests(SimpleTestCase):
    """Камера читает код и отдаёт его в тот же слой, что и USB-сканер.

    Ни один экран из-за камеры не менялся: у одного и того же скана
    не должно быть двух разных поведений.
    """

    STATIC = Path(__file__).resolve().parent / 'static'
    TEMPLATES = Path(__file__).resolve().parent / 'templates' / 'core'

    def _code(self, path):
        """Сам код, без комментариев: в них как раз и написано, чего
        здесь быть не должно."""
        source = path.read_text(encoding='utf-8')
        source = re.sub(r'/\*.*?\*/', '', source, flags=re.S)
        return re.sub(r'^\s*//.*$', '', source, flags=re.M)

    @property
    def script(self):
        return self._code(self.STATIC / 'js' / 'camera-scan.js')

    def test_loaded_on_every_page(self):
        base = (self.TEMPLATES / 'base.html').read_text(encoding='utf-8')

        self.assertIn('js/camera-scan.js', base)
        self.assertIn('css/camera-scan.css', base)

    def test_the_result_goes_into_the_common_scanner_layer(self):
        """Иначе у камеры завелись бы свои виды кодов, свои сообщения
        и своё поведение — вторая правда о том же самом."""
        self.assertIn('LiftTeamScanner.submit(value)', self.script)

    def test_no_third_party_library(self):
        """Всё стороннее в этом проекте приезжает из интернета, а интернет
        пропадает ровно тогда, когда сканируют у стеллажа."""
        self.assertIn('BarcodeDetector', self.script)
        for forbidden in ('cdn.', 'jsQR', 'zxing', 'import(', 'src ='):
            with self.subTest(forbidden=forbidden):
                self.assertNotIn(forbidden, self.script)

    def test_no_bootstrap(self):
        css = self._code(self.STATIC / 'css' / 'camera-scan.css')
        button = (self.TEMPLATES / '_camera_button.html').read_text(encoding='utf-8')

        self.assertNotIn('data-bs-', self.script + button)
        self.assertNotIn('bootstrap', self.script)
        # Панель прячется атрибутом, а не классом Bootstrap
        self.assertIn('panel.hidden = true', self.script)
        self.assertIn('.camera-scan[hidden] { display: none; }', css)

    def test_every_refusal_is_spoken(self):
        """Молчаливое «не получилось» человек примет за поломку кнопки
        и будет жать её снова."""
        for reason in ('HTTPS', 'BarcodeDetector', 'Доступ к камере запрещён',
                       'Камера на этом устройстве не найдена'):
            with self.subTest(reason=reason):
                self.assertIn(reason, self.script)

    def test_the_secure_context_is_checked_before_asking(self):
        """Браузер отдаёт камеру только в защищённом окружении, и сказать
        об этом надо до запроса, а не после отказа."""
        self.assertIn('window.isSecureContext', self.script)

    def test_the_camera_is_released_when_the_panel_closes(self):
        """Иначе индикатор на телефоне горит и после закрытия, и человек
        справедливо решает, что программа за ним подглядывает."""
        self.assertIn('track.stop()', self.script)
        self.assertIn("window.addEventListener('pagehide', stop)", self.script)

    def test_a_code_left_in_frame_is_one_scan(self):
        """Камера видит наклейку непрерывно: без своего окна повтора
        она пищала бы, пока код не уберут."""
        self.assertIn('SAME_CODE_MS = 2500', self.script)

    def test_the_button_is_an_include_used_by_the_working_screens(self):
        """Новое место — одна вставка, без правки скриптов."""
        button = (self.TEMPLATES / '_camera_button.html').read_text(encoding='utf-8')
        self.assertIn('data-camera-scan', button)

        for page in ('scan.html', 'inventory/count.html', 'storage_cells/grid.html'):
            with self.subTest(page=page):
                html = (self.TEMPLATES / page).read_text(encoding='utf-8')
                self.assertIn("include 'core/_camera_button.html'", html)


class HttpsDeploymentTests(SimpleTestCase):
    """HTTPS нужен ровно ради камеры, но включается на всю программу —
    значит и ломать он не должен ничего."""

    DEPLOY = Path(__file__).resolve().parent.parent / 'deploy'

    @property
    def https_conf(self):
        return (self.DEPLOY / 'nginx-lifteam-https.conf').read_text(encoding='utf-8')

    def test_https_config_keeps_the_address_restriction(self):
        """HTTPS шифрует канал, но никого не отсекает: без allow/deny
        программа была бы открыта миру, просто по защищённому соединению."""
        conf = self.https_conf

        self.assertEqual(conf.count('deny all;'), 2)      # и на 80, и на 443
        self.assertEqual(conf.count('allow 100.64.0.0/10;'), 2)

    def test_http_only_redirects(self):
        """Отдавать то же самое по http хуже, чем не отдавать: страница
        открылась бы, камера бы не заработала, и причина осталась неясной."""
        self.assertIn('return 301 https://$host$request_uri;', self.https_conf)

    def test_websocket_and_media_survive_the_switch(self):
        conf = self.https_conf

        self.assertIn('location /ws/', conf)
        self.assertIn("proxy_set_header Upgrade $http_upgrade;", conf)
        self.assertIn('location /media/', conf)
        self.assertIn('location /static/', conf)

    def test_the_proto_header_is_passed_on(self):
        """По нему Django понимает, что снаружи HTTPS, и перестаёт уводить
        обратно на http."""
        self.assertIn('proxy_set_header X-Forwarded-Proto $scheme;', self.https_conf)

    def test_the_certificate_renewal_is_scheduled(self):
        """Сертификат живёт около трёх месяцев, и просроченный закрывает
        программу целиком, а не только камеру."""
        service = (self.DEPLOY / 'lifteam-cert.service').read_text(encoding='utf-8')
        timer = (self.DEPLOY / 'lifteam-cert.timer').read_text(encoding='utf-8')

        self.assertIn('tailscale cert', service)
        self.assertIn('systemctl reload nginx', service)
        # Pi выключали — обновление должно пройти при следующем включении
        self.assertIn('Persistent=true', timer)

    def test_the_app_already_knows_how_to_be_behind_https(self):
        """Настройки для этого были заведены раньше: включение HTTPS —
        это правка nginx и одна переменная, а не правка кода."""
        pi = (Path(__file__).resolve().parent.parent
              / 'lifteam/settings_pi.py').read_text(encoding='utf-8')

        self.assertIn("SECURE_PROXY_SSL_HEADER = ('HTTP_X_FORWARDED_PROTO', 'https')", pi)
        self.assertIn('SECURE_SSL_REDIRECT', pi)
        self.assertIn('CSRF_TRUSTED_ORIGINS', pi)


# ============ ИЗДЕЛИЕ В КОММЕРЧЕСКОМ ПРЕДЛОЖЕНИИ (v2.74.0) ============


class QuoteEquipmentNameTests(TestCase):
    """Наименование изделия в предложении: своё, но правимое.

    Решение владельца. Исполнение значит разное у разных изделий:
    у преобразователей частоты версии отличаются мощностью, то есть
    и ценой ремонта, а у приводов дверей на цену почти не влияют. Что
    из этого важно заказчику в этом предложении, решает мастер.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='quote_name', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(
            name='Emotron-2.0', kind='Преобразователь частоты'
        )
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='-1.1'
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, version=self.version, serial_number='SN-Q-1'
            ),
            proposed_work='Замена конденсаторов звена постоянного тока',
            estimated_cost=Decimal('12000.00'),
        )

    def test_the_designation_with_the_version_is_used_by_default(self):
        """Одна и та же модель в разных исполнениях — разное изделие,
        и в предложении это должно быть видно."""
        self.assertEqual(
            self.roe.quote_equipment_name,
            'Преобразователь частоты Emotron-2.0-1.1',
        )

    def test_a_hand_written_name_wins(self):
        self.roe.quote_designation = 'ПЧ Emotron 2.0 (7,5 кВт)'
        self.roe.save()

        self.assertEqual(self.roe.quote_equipment_name, 'ПЧ Emotron 2.0 (7,5 кВт)')

    def test_whitespace_is_not_a_name(self):
        self.roe.quote_designation = '   '
        self.roe.save()

        self.assertEqual(
            self.roe.quote_equipment_name,
            'Преобразователь частоты Emotron-2.0-1.1',
        )

    def test_a_unit_without_a_version_prints_without_one(self):
        """Исполнения нет — «исп. 1.0» вместо пустого поля
        не выдумывается."""
        self.roe.equipment.version = None
        self.roe.equipment.save()

        self.assertEqual(
            self.roe.quote_equipment_name, 'Преобразователь частоты Emotron-2.0'
        )

    def test_the_printed_quote_shows_the_equipment_and_the_work(self):
        """Изделие — первой строкой, работы под ним: заказчик должен
        видеть, какое именно ему считают."""
        html = self.http.get(
            reverse('repair_order_quote', args=[self.order.pk])
        ).content.decode()

        self.assertIn('Преобразователь частоты Emotron-2.0-1.1', html)
        self.assertIn('Замена конденсаторов звена постоянного тока', html)

    def test_the_form_offers_the_default_as_a_placeholder(self):
        """«Введите наименование» тут ничего не объясняет: наименование
        уже есть, вопрос только в том, устраивает ли оно."""
        html = self.http.get(
            reverse('repair_order_quote_edit', args=[self.order.pk])
        ).content.decode()

        self.assertIn('placeholder="Преобразователь частоты Emotron-2.0-1.1"', html)

    def test_the_name_is_saved_from_the_form(self):
        response = self.http.post(
            reverse('repair_order_quote_edit', args=[self.order.pk]),
            {
                'quote_subject': '', 'quote_date': '2026-08-25',
                'quote_valid_until': '2026-09-25', 'quote_lead_time': '',
                'quote_payment_terms': '', 'quote_delivery_terms': '',
                'order_equipments-TOTAL_FORMS': '1',
                'order_equipments-INITIAL_FORMS': '1',
                'order_equipments-MIN_NUM_FORMS': '0',
                'order_equipments-MAX_NUM_FORMS': '1000',
                'order_equipments-0-id': str(self.roe.pk),
                'order_equipments-0-quote_designation': 'ПЧ Emotron 2.0 (7,5 кВт)',
                'order_equipments-0-proposed_work': self.roe.proposed_work,
                'order_equipments-0-repair_complexity': '',
                'order_equipments-0-estimated_cost': '12000.00',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.roe.refresh_from_db()
        self.assertEqual(self.roe.quote_designation, 'ПЧ Emotron 2.0 (7,5 кВт)')


class EquipmentOwnerOnRepeatIntakeTests(TestCase):
    """Владельца при повторном приёме не меняем — решение владельца.

    Прибор нередко приезжает от другого обслуживающего предприятия,
    и это не значит, что он сменил хозяина. У кого он был в каждом
    ремонте, и так видно по заказам.
    """

    def setUp(self):
        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.first = ClientModel.objects.create(name='МУП «Лифты»')
        self.second = ClientModel.objects.create(name='ООО «Подъём»')
        self.equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-OWNER-1'
        )

    def test_an_empty_owner_is_filled_in(self):
        order = RepairOrder.objects.create(client=self.first)
        RepairOrderEquipment.objects.create(
            repair_order=order, equipment=self.equipment
        )

        order.assign_equipment_owners()

        self.equipment.refresh_from_db()
        self.assertEqual(self.equipment.current_client, self.first)

    def test_an_existing_owner_survives_a_repeat_intake(self):
        self.equipment.current_client = self.first
        self.equipment.save()
        second_order = RepairOrder.objects.create(client=self.second)
        RepairOrderEquipment.objects.create(
            repair_order=second_order, equipment=self.equipment
        )

        changed = second_order.assign_equipment_owners()

        self.equipment.refresh_from_db()
        self.assertEqual(changed, 0)
        self.assertEqual(self.equipment.current_client, self.first)


# ============ СТРАНИЦА ЕДИНИЦЫ ОБОРУДОВАНИЯ (v2.76.0) ============


class UnitPageTests(TestCase):
    """Всё про один прибор в одной работе — на одной странице.

    Раньше это лежало в трёх местах: часть в строке карточки заказа,
    часть в форме правки заказа, часть на дефектации, — и мастер
    с платой в руках ходил между ними.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='unit_page', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(
            name='БУАД-7-31', kind='Привод дверей'
        )
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        self.fault = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            description='Электролитические конденсаторы потеряли ёмкость.',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, version=self.version, serial_number='SN-UNIT-1'
            ),
            fault_description='Не открывает двери на первом этаже',
            initial_condition='Корпус цел, пломбы на месте',
            work_performed='Заменены конденсаторы C12, C13',
            seal_numbers='7788, 7789',
            diagnosis='Вздулись конденсаторы в цепи питания',
            estimated_cost=Decimal('4500.00'),
        )
        self.roe.faults.add(self.fault)

    def _url(self):
        return reverse('repair_order_unit_detail',
                       args=[self.order.pk, self.roe.pk])

    def test_the_page_gathers_what_used_to_lie_in_three_places(self):
        html = self.http.get(self._url()).content.decode()

        for expected in (
            'Привод дверей БУАД-7-31.4',          # обозначение с исполнением
            'SN-UNIT-1',
            'Не открывает двери на первом этаже',  # приём
            'Корпус цел, пломбы на месте',
            'Вздулись конденсаторы в цепи питания',  # диагностика
            'Электролитические конденсаторы потеряли ёмкость.',
            'Заменены конденсаторы C12, C13',      # ремонт
            '7788, 7789',                          # пломбы
        ):
            with self.subTest(expected=expected):
                self.assertIn(expected, html)

    def test_the_position_in_the_order_is_shown(self):
        """Позиция напечатана на наклейке, и человек с коробкой в руках
        ищет по ней."""
        second = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-UNIT-2'
            ),
        )

        html = self.http.get(
            reverse('repair_order_unit_detail', args=[self.order.pk, second.pk])
        ).content.decode()

        self.assertIn('позиция 2 в заказе', html)

    def test_readiness_is_shown_first(self):
        """С этим вопросом сюда чаще всего и приходят."""
        html = self.http.get(self._url()).content.decode()

        self.assertIn('Записаны выполненные работы', html)
        self.assertIn('Готовность', html)

    def test_only_the_parts_of_this_unit_are_listed(self):
        """Списанные «на заказ целиком» сюда не попадают: программа
        не знает, в какую железку они ушли, и приписывать наугад нельзя."""
        part = SparePart.objects.create(
            part_number='C-100u', name='Конденсатор 100 мкФ', current_stock=10
        )
        other = SparePart.objects.create(
            part_number='R-10K', name='Резистор 10 кОм', current_stock=10
        )
        RepairOrderDetail.objects.create(
            repair_order=self.order, order_equipment=self.roe,
            part=part, quantity_used=2,
        )
        RepairOrderDetail.objects.create(
            repair_order=self.order, order_equipment=None,
            part=other, quantity_used=1,
        )

        html = self.http.get(self._url()).content.decode()

        self.assertIn('C-100u', html)
        self.assertNotIn('R-10K', html)

    def test_a_unit_of_another_order_is_not_opened_by_a_guessed_address(self):
        """Иначе по подобранному адресу открылась бы чужая позиция."""
        other_order = RepairOrder.objects.create(client=self.order.client)

        response = self.http.get(
            reverse('repair_order_unit_detail', args=[other_order.pk, self.roe.pk])
        )

        self.assertEqual(response.status_code, 404)

    def test_the_page_requires_login(self):
        self.http.logout()

        response = self.http.get(self._url())

        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])

    def test_tech_cards_are_linked_with_the_units_version(self):
        """Без исполнения в ссылке мастер увидел бы половину карты
        и не понял бы, что половины нет."""
        card = TechCard.objects.create(
            equipment_model=self.model, title='Разборка корпуса'
        )

        html = self.http.get(self._url()).content.decode()

        self.assertIn(
            '%s?version=%d' % (
                reverse('tech_card_detail', args=[card.pk]), self.version.pk
            ),
            html,
        )


class UnitRowOnOrderCardTests(TestCase):
    """Строка единицы на карточке заказа: клик ведёт на её страницу."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='unit_row', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-ROW-1'
            ),
            fault_description='А' * 200,
        )

    @property
    def html(self):
        return self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

    def test_the_serial_number_is_a_real_link(self):
        """Кликабельной строки мало: ссылка работает и без скриптов,
        и средней кнопкой мыши в новой вкладке."""
        html = self.html

        self.assertIn(
            'href="%s"' % reverse('repair_order_unit_detail',
                                  args=[self.order.pk, self.roe.pk]),
            html,
        )
        self.assertIn('data-href="%s"' % reverse(
            'repair_order_unit_detail', args=[self.order.pk, self.roe.pk]), html)

    def test_long_texts_are_cut_in_the_row(self):
        """Раньше в ячейку шёл весь абзац, и строка вырастала на полэкрана —
        из-за этого таблица и выглядела тяжёлой."""
        html = self.html

        self.assertNotIn('А' * 200, html)
        self.assertIn('А' * 50, html)

    def test_the_anchor_survives(self):
        """С чек-листа готовности и из списка деталей ссылки ведут
        на строку — переход по ним не должен сломаться."""
        self.assertIn('id="unit-%d"' % self.roe.pk, self.html)

    def test_clicks_on_controls_do_not_navigate(self):
        """Галочка отмечает единицу для печати этикеток пачкой, кнопки
        ведут каждая в своё место — увести их на страницу единицы
        значило бы сломать и то, и другое."""
        detail = (settings.BASE_DIR
                  / 'core/templates/core/repair_orders/detail.html'
                  ).read_text(encoding='utf-8')

        self.assertIn("closest('a, button, input, label, select, textarea')", detail)
        # И выделение текста мышью тоже не переход: серийники отсюда копируют
        self.assertIn('window.getSelection()', detail)


# ============ ТРИ МЕЛОЧИ ПО ВИДУ СТРАНИЦ (v2.77.0) ============


class ReadinessEverywhereTests(TestCase):
    """Готовность видна там, где по ней принимают решения: на карточке
    заказа, в списке заказов и на дашборде.

    Список проверок при этом один на всю программу — иначе три экрана
    рано или поздно начали бы говорить разное об одном и том же заказе.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='readiness_everywhere', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.client_obj = ClientModel.objects.create(name='МУП «Лифты»')

    def _order(self, status='repair', ready=True, serial='SN-RE-1'):
        order = RepairOrder.objects.create(client=self.client_obj, status=status)
        roe = RepairOrderEquipment.objects.create(
            repair_order=order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number=serial
            ),
        )
        if ready:
            roe.diagnosis = 'Высохли конденсаторы.'
            roe.work_performed = 'Заменены конденсаторы.'
            roe.repair_cost = Decimal('4500.00')
            roe.save()
        return order

    # --- список заказов ---

    def test_the_list_shows_how_many_units_are_done(self):
        """Чтобы понять, что реально можно отгружать, не открывая
        каждый заказ."""
        order = self._order(ready=False)
        RepairOrderEquipment.objects.create(
            repair_order=order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-RE-2'
            ),
        )

        html = self.http.get(reverse('repair_order_list')).content.decode()

        self.assertIn('Готовность', html)
        self.assertIn('0 из 2', html)

    def test_a_finished_order_is_marked_as_whole(self):
        self._order(ready=True)

        html = self.http.get(reverse('repair_order_list')).content.decode()

        self.assertIn('все 1', html)

    def test_a_shipped_order_is_not_alarmed_about(self):
        """По отгруженному доделывать нечего, и жёлтая отметка была бы
        ложной тревогой."""
        self._order(status='shipped', ready=False)

        html = self.http.get(reverse('repair_order_list')).content.decode()

        self.assertIn('0 из 1', html)
        self.assertNotIn('bg-warning text-dark">0 из 1', html)

    def test_an_order_without_equipment_shows_a_dash(self):
        """Отгружать нечего — это не «не готово», и тревожить нечем."""
        RepairOrder.objects.create(client=self.client_obj, status='accepted')

        html = self.http.get(reverse('repair_order_list')).content.decode()
        cell = html.split('data-label="Готовность">')[1].split('</td>')[0]

        self.assertIn('—', cell)
        self.assertNotIn('bg-warning', cell)

    def test_readiness_does_not_fan_out_into_queries(self):
        """Готовность спрашивает и неисправности, и детали по каждой
        единице: без предзагрузки страница списка уходила бы в сотню
        запросов.

        Проверяем прирост, а не общее число: на странице есть и другие
        запросы, в том числе один на заказ ради суммы (он был здесь
        и раньше). Важно, что готовность к этому не добавляет ничего.
        """
        def queries_for(count, tag):
            RepairOrder.objects.all().delete()
            Equipment.objects.all().delete()
            for number in range(count):
                self._order(serial='SN-%s-%d' % (tag, number))
            with CaptureQueriesContext(connection) as captured:
                self.http.get(reverse('repair_order_list'))
            return len(captured)

        three, six = queries_for(3, 'A'), queries_for(6, 'B')

        # Не больше одного запроса на добавленный заказ — того самого,
        # что считает сумму. Разошлась бы готовность — прирост был бы
        # втрое больше
        self.assertLessEqual(six - three, 3)

    # --- дашборд ---

    def test_the_dashboard_names_what_is_ready_to_ship(self):
        """Первый вопрос дня, и раньше ответа на него здесь не было
        вовсе."""
        ready = self._order(status='ready_for_shipment', ready=True)
        self._order(status='repair', ready=False, serial='SN-RE-3')

        html = self.http.get(reverse('dashboard')).content.decode()

        self.assertIn('Готово к отгрузке', html)
        self.assertIn(ready.order_number, html.split('Готово к отгрузке')[1])

    def test_an_unfinished_order_does_not_reach_the_dashboard_block(self):
        self._order(status='repair', ready=False)

        html = self.http.get(reverse('dashboard')).content.decode()

        self.assertNotIn('Готово к отгрузке', html)

    def test_a_shipped_order_is_no_longer_offered(self):
        """Его уже отгрузили — предлагать отгрузить снова незачем."""
        self._order(status='shipped', ready=True)

        html = self.http.get(reverse('dashboard')).content.decode()

        self.assertNotIn('Готово к отгрузке', html)


class StatusControlPlacementTests(TestCase):
    """Смена статуса стоит рядом со значком статуса, а не отдельной
    карточкой внизу страницы.

    Значок был вверху, а форма под деталями и платежами: на телефоне
    это долгая прокрутка туда и обратно по нескольку раз в день.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='status_place', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )

    def test_the_status_form_sits_in_the_order_info_card(self):
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

        head, _, tail = html.partition('Оборудование в заказе')
        self.assertIn(
            reverse('repair_order_change_status', args=[self.order.pk]), head
        )
        self.assertNotIn(
            reverse('repair_order_change_status', args=[self.order.pk]), tail
        )

    def test_changing_the_status_still_works(self):
        response = self.http.post(
            reverse('repair_order_change_status', args=[self.order.pk]),
            {'new_status': 'repair', 'notes': 'взяли в работу'}, follow=True,
        )

        self.order.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.order.status, 'repair')


# ============ БЫСТРЫЕ ДЕЙСТВИЯ ПО ЕДИНИЦЕ (v2.78.0) ============


class UnitEditTests(TestCase):
    """Правка единицы на её странице.

    Раньше эти поля правились только через редактирование всего заказа —
    формой на все единицы разом: записать работы по одному прибору стоило
    открыть заказ целиком и сохранить всё вместе.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='unit_edit', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.fault = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            description='Конденсаторы потеряли ёмкость.',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-EDIT-1'
            ),
        )

    def _post(self, **extra):
        data = {
            'fault_description': 'Не открывает двери',
            'initial_condition': 'Корпус цел',
            'work_performed': 'Заменены конденсаторы C12, C13',
            'seal_numbers': '7788',
            'repair_cost': '5200.00',
        }
        data.update(extra)
        return self.http.post(
            reverse('repair_order_unit_edit', args=[self.order.pk, self.roe.pk]),
            data,
        )

    def test_the_fields_are_saved_from_the_unit_page(self):
        response = self._post()

        self.roe.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.roe.work_performed, 'Заменены конденсаторы C12, C13')
        self.assertEqual(self.roe.seal_numbers, '7788')
        self.assertEqual(self.roe.repair_cost, Decimal('5200.00'))
        self.assertEqual(self.roe.fault_description, 'Не открывает двери')

    def test_saving_returns_to_the_place_that_was_edited(self):
        """Мастер записывает работы по одному прибору и берётся
        за следующий, а не уходит в заказ целиком."""
        response = self._post(anchor='repair')

        self.assertEqual(
            response['Location'],
            reverse('repair_order_unit_detail',
                    args=[self.order.pk, self.roe.pk]) + '#repair',
        )

    def test_faults_of_another_model_are_not_offered(self):
        """Чужая неисправность означала бы рецепт деталей не от этого
        прибора."""
        other = EquipmentModel.objects.create(name='EkoDrive 2.0')
        alien = FaultType.objects.create(equipment_model=other, name='иная')

        html = self.http.get(
            reverse('repair_order_unit_detail', args=[self.order.pk, self.roe.pk])
        ).content.decode()

        self.assertIn('высохли конденсаторы', html)
        self.assertNotIn('>иная<', html)
        self.assertNotIn('value="%d"' % alien.pk, html)

    def test_choosing_a_fault_changes_readiness_and_documents(self):
        """Описание выбранной неисправности идёт в акт дефектации,
        и из неё же выводится сложность ремонта."""
        self._post(faults=[str(self.fault.pk)])

        self.roe.refresh_from_db()
        self.assertEqual(list(self.roe.faults.all()), [self.fault])
        self.assertIn('Конденсаторы потеряли ёмкость.',
                      self.roe.diagnosis_document_text)

    def test_complexity_is_set_by_hand_on_the_unit_page(self):
        """До v2.84.0 сложность правилась только в форме предложения —
        то есть нигде, пока предложение не понадобится, и цвет сложности
        в списке единиц заказа не появлялся ни у одной строки."""
        self._post(repair_complexity='complex')

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.repair_complexity, 'complex')
        self.assertEqual(self.roe.effective_complexity_css, 'bg-danger')

    def test_empty_complexity_means_derive_it_from_the_faults(self):
        """Пусто — не «простой», а «не задавали»: тогда сложность
        считается по выбранным неисправностям."""
        self.fault.complexity = 'complex'
        self.fault.save()

        self._post(repair_complexity='', faults=[str(self.fault.pk)])

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.repair_complexity, '')
        self.assertEqual(self.roe.effective_complexity, 'complex')
        self.assertTrue(self.roe.complexity_is_derived)

    def test_the_empty_choice_says_what_it_means(self):
        """«---------» не говорит ничего, а разница между «простой»
        и «не задавали» — это разница между записанным решением
        и подсчётом по неисправностям."""
        html = self.http.get(
            reverse('repair_order_unit_detail', args=[self.order.pk, self.roe.pk])
        ).content.decode()

        self.assertIn('По неисправностям', html)
        self.assertNotIn('---------', html)

    def test_a_bad_cost_saves_nothing(self):
        response = self._post(repair_cost='дорого')

        self.roe.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertIsNone(self.roe.repair_cost)
        self.assertEqual(self.roe.work_performed, '')

    def test_only_post_edits(self):
        """Ссылку с этим адресом мог бы открыть кто угодно."""
        response = self.http.get(
            reverse('repair_order_unit_edit', args=[self.order.pk, self.roe.pk])
        )

        self.assertEqual(response.status_code, 405)

    def test_a_unit_of_another_order_is_not_edited(self):
        other = RepairOrder.objects.create(client=self.order.client)

        response = self.http.post(
            reverse('repair_order_unit_edit', args=[other.pk, self.roe.pk]),
            {'work_performed': 'чужое'},
        )

        self.roe.refresh_from_db()
        self.assertEqual(response.status_code, 404)
        self.assertEqual(self.roe.work_performed, '')

    def test_the_recipe_button_uses_the_same_endpoint_as_the_order_form(self):
        """Второй такой обработчик однажды разошёлся бы с первым,
        а списание со склада — не то место, где это можно позволить."""
        html = self.http.get(
            reverse('repair_order_unit_detail', args=[self.order.pk, self.roe.pk])
        ).content.decode()

        self.assertIn(
            reverse('repair_order_apply_fault_template', args=[self.order.pk]),
            html,
        )
        self.assertIn('LiftTeamWS.fetch', html)


class UnitQuickActionsTests(TestCase):
    """Быстрые действия в строке единицы на карточке заказа."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='quick_actions', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-QA-1'
            ),
        )

    @property
    def html(self):
        return self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

    def test_writing_work_is_one_click_from_the_row(self):
        """Самое узкое место: поле входит в чек-лист готовности,
        а добраться до него было дороже всего."""
        expected = reverse(
            'repair_order_unit_detail', args=[self.order.pk, self.roe.pk]
        ) + '#repair'

        self.assertIn(expected, self.html)

    def test_using_a_part_leads_to_the_parts_of_that_unit(self):
        """На странице единицы её не надо выбирать — она известна.
        Форма на карточке заказа осталась для «на заказ целиком»."""
        html = self.html

        self.assertIn(
            reverse('repair_order_unit_detail',
                    args=[self.order.pk, self.roe.pk]) + '#parts',
            html,
        )
        self.assertIn('id="usePartForm"', html)

    def test_the_row_does_not_carry_a_page_button_any_more(self):
        """На страницу ведёт вся строка и серийный номер в ней; место
        отдано тому, что жмут по многу раз в день."""
        html = self.html

        self.assertIn('bi-pencil-square', html)   # записать работы
        self.assertIn('bi-cpu', html)             # списать деталь
        self.assertNotIn('bi-box-arrow-in-right', html)


# ============ РАЗБОР СТРАНИЦЫ ПРАВКИ ЗАКАЗА (v2.79.0) ============


class OrderEditPageIsGoneTests(TestCase):
    """Страницы правки заказа больше нет, и её работу забрали три места.

    Оставленная «облегчённой», она означала бы два места для одних и тех
    же полей — ровно ту беду, которую чинили: статус оплаты правился
    и там, и своей формой на карточке, и вдобавок пересчитывался сам
    после внесения платежа.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='order_edit_gone', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.first = ClientModel.objects.create(name='МУП «Лифты»')
        self.second = ClientModel.objects.create(name='ООО «Подъём»')
        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(client=self.first)
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-GONE-1'
            ),
        )

    def test_the_address_is_gone(self):
        self.assertEqual(
            self.http.get('/repair-orders/%d/edit/' % self.order.pk).status_code,
            404,
        )

    def test_nothing_links_to_it_any_more(self):
        """Ссылка на несуществующую страницу — это «страница не найдена»
        в руках у человека."""
        for url in (reverse('repair_order_list'),
                    reverse('repair_order_detail', args=[self.order.pk])):
            with self.subTest(page=url):
                html = self.http.get(url).content.decode()
                self.assertNotIn('/repair-orders/%d/edit/' % self.order.pk, html)

    def test_the_client_is_changed_from_the_order_card(self):
        response = self.http.post(
            reverse('repair_order_edit_info', args=[self.order.pk]),
            {'client': str(self.second.pk)},
        )

        self.order.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.order.client, self.second)

    def test_the_card_does_not_offer_the_invoice_fields_twice(self):
        """Счёт заполняется при выставлении, статус оплаты — своей формой
        и пересчитывается сам после платежа."""
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()
        form = html.split('edit-info/')[1].split('</form>')[0]

        for name in ('invoice_number', 'invoice_date', 'payment_status'):
            with self.subTest(field=name):
                self.assertNotIn('name="%s"' % name, form)

    def test_only_post_changes_the_order(self):
        response = self.http.get(
            reverse('repair_order_edit_info', args=[self.order.pk])
        )

        self.assertEqual(response.status_code, 405)


class UnitRemovalTests(TestCase):
    """Снятие единицы с заказа — со страницы этой единицы."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='unit_removal', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.equipment = Equipment.objects.create(
            model=self.model, serial_number='SN-RM-1'
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order, equipment=self.equipment
        )

    def _url(self):
        return reverse('repair_order_unit_remove',
                       args=[self.order.pk, self.roe.pk])

    def test_it_asks_before_removing(self):
        """Убрать прибор — значит потерять записанное по нему, и
        переспросить дешевле, чем восстанавливать."""
        response = self.http.get(self._url())

        self.assertEqual(response.status_code, 200)
        self.assertTrue(RepairOrderEquipment.objects.filter(pk=self.roe.pk).exists())
        self.assertContains(response, 'Убрать из заказа')

    def test_it_removes_on_post(self):
        response = self.http.post(self._url())

        self.assertRedirects(
            response, reverse('repair_order_detail', args=[self.order.pk]))
        self.assertFalse(RepairOrderEquipment.objects.filter(pk=self.roe.pk).exists())

    def test_the_equipment_itself_survives(self):
        """Прибор из справочника никуда не девается — его история
        ремонтов остаётся, и принять его снова можно в любой момент."""
        self.http.post(self._url())

        self.equipment.refresh_from_db()
        self.assertTrue(Equipment.objects.filter(pk=self.equipment.pk).exists())

    def test_parts_stay_written_off_and_say_so(self):
        """Детали и правда взяли и потратили: возврат на склад — своё
        действие, и делать его молча за человека нельзя."""
        part = SparePart.objects.create(
            part_number='C-100u', name='Конденсатор', current_stock=10
        )
        detail = RepairOrderDetail.objects.create(
            repair_order=self.order, order_equipment=self.roe,
            part=part, quantity_used=2,
        )

        page = self.http.get(self._url())
        self.assertContains(page, 'На склад они не вернутся')

        self.http.post(self._url())

        detail.refresh_from_db()
        self.assertIsNone(detail.order_equipment)
        self.assertEqual(detail.quantity_used, 2)

    def test_a_unit_of_another_order_is_not_removed(self):
        other = RepairOrder.objects.create(client=self.order.client)

        response = self.http.post(
            reverse('repair_order_unit_remove', args=[other.pk, self.roe.pk])
        )

        self.assertEqual(response.status_code, 404)
        self.assertTrue(RepairOrderEquipment.objects.filter(pk=self.roe.pk).exists())


class NewEquipmentFromOrderTests(TestCase):
    """Прибор, забытый при приёме, заводится с карточки заказа.

    Раньше его заводили окном на странице правки заказа. Страницы больше
    нет, а окно рисует Bootstrap, который приходит из интернета, — поэтому
    обычная страница с возвратом обратно.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='new_equipment', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.client_obj = ClientModel.objects.create(name='МУП «Лифты»')
        self.order = RepairOrder.objects.create(client=self.client_obj)

    def test_the_order_card_offers_it(self):
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

        self.assertIn('Завести новое оборудование', html)
        self.assertIn('client=%d' % self.client_obj.pk, html)

    def test_the_client_is_offered_as_the_owner(self):
        """У прибора заказчик и станет владельцем — перевыбирать его
        руками незачем."""
        page = self.http.get(
            reverse('equipment_create'), {'client': self.client_obj.pk}
        )

        self.assertEqual(
            page.context['form'].initial.get('current_client'),
            str(self.client_obj.pk),
        )

    def test_saving_returns_to_the_order(self):
        back = reverse('repair_order_detail', args=[self.order.pk])

        response = self.http.post(reverse('equipment_create'), {
            'model': str(self.model.pk), 'serial_number': 'SN-NEW-1',
            'current_client': str(self.client_obj.pk), 'next': back,
        })

        self.assertRedirects(response, back)
        self.assertTrue(Equipment.objects.filter(serial_number='SN-NEW-1').exists())

    def test_a_foreign_address_is_not_accepted(self):
        """Иначе кнопка стала бы способом увести человека на постороннюю
        страницу."""
        response = self.http.post(reverse('equipment_create'), {
            'model': str(self.model.pk), 'serial_number': 'SN-NEW-2',
            'next': 'https://example.com/',
        })

        self.assertRedirects(response, reverse('equipment_list'))


# ============ ПРАВКИ ПО ЗАМЕЧАНИЯМ ВЛАДЕЛЬЦА (v2.80.0) ============


class SpaceSeparatorTests(TestCase):
    """У части изделий разделитель между моделью и исполнением — пробел.

    Django у текстовых полей по умолчанию обрезает пробелы по краям,
    поэтому введённое « 21.01» ложилось в базу как «21.01», и печаталось
    «МАГНУС21.01». Дело было не в печати, а в сохранении.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='space_sep', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)
        self.model = EquipmentModel.objects.create(name='МАГНУС')

    def _post(self, name):
        return self.http.post(reverse('equipment_version_create'), {
            'equipment_model': str(self.model.pk), 'name': name, 'note': '',
        })

    def test_a_leading_space_survives_saving(self):
        self._post(' 21.01')

        version = EquipmentVersion.objects.get()
        self.assertEqual(version.name, ' 21.01')

    def test_the_designation_is_printed_with_the_space(self):
        self._post(' 21.01')
        equipment = Equipment.objects.create(
            model=self.model, version=EquipmentVersion.objects.get(),
            serial_number='SN-SPACE-1',
        )

        self.assertEqual(equipment.designation, 'МАГНУС 21.01')
        self.assertEqual(str(EquipmentVersion.objects.get()), 'МАГНУС 21.01')

    def test_other_separators_still_work(self):
        for name in ('.4', '-1.1'):
            with self.subTest(name=name):
                EquipmentVersion.objects.all().delete()
                self._post(name)
                self.assertEqual(EquipmentVersion.objects.get().name, name)

    def test_a_trailing_space_is_refused_as_a_typo(self):
        """Слева пробел — разделитель, справа — задетый случайно."""
        response = self._post('21.01 ')

        self.assertEqual(response.status_code, 200)
        self.assertFalse(EquipmentVersion.objects.exists())
        self.assertContains(response, 'Пробел в конце обозначения')

    def test_spaces_only_is_not_a_designation(self):
        response = self._post('   ')

        self.assertFalse(EquipmentVersion.objects.exists())
        self.assertContains(response, 'не может быть пустым')

    def test_the_form_shows_what_will_be_printed(self):
        """Пробела не видно — значит его надо показать: иначе разница
        всплывает только на бумаге, у заказчика."""
        html = self.http.get(reverse('equipment_version_create')).content.decode()

        self.assertIn('id="versionPreview"', html)
        self.assertIn('Напечатается', html)
        self.assertIn('"%d": "МАГНУС"' % self.model.pk, html)


class VersionChoiceScopedToModelTests(TestCase):
    """В выборе исполнения — только исполнения выбранной модели."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='version_scope', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.version = EquipmentVersion.objects.create(
            equipment_model=self.model, name='.4'
        )
        self.other = EquipmentModel.objects.create(name='EkoDrive 2.0')
        self.alien = EquipmentVersion.objects.create(
            equipment_model=self.other, name='-1.1'
        )

    def test_the_page_filters_the_list_by_the_chosen_model(self):
        """Раньше список приходил целиком — сотни строк, из которых
        к прибору относятся одна-две, и отказ приходил после сохранения."""
        html = self.http.get(reverse('equipment_create')).content.decode()

        self.assertIn(reverse('ajax_equipment_model_list'), html)
        self.assertIn('Исполнение снято', html)

    def test_the_data_source_is_the_one_that_already_exists(self):
        """Один источник, одно поведение: то же, что у окна заведения
        оборудования при приёме заказа."""
        data = self.http.get(reverse('ajax_equipment_model_list')).json()

        by_name = {m['name']: m for m in data['models']}
        self.assertEqual(
            [v['name'] for v in by_name['БУАД-7-31']['versions']], ['.4']
        )
        self.assertEqual(
            [v['name'] for v in by_name['EkoDrive 2.0']['versions']], ['-1.1']
        )

    def test_the_server_still_refuses_a_foreign_version(self):
        """Отбор в браузере — удобство; последнее слово за сервером."""
        response = self.http.post(reverse('equipment_create'), {
            'model': str(self.model.pk), 'version': str(self.alien.pk),
            'serial_number': 'SN-SCOPE-1',
        })

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Equipment.objects.exists())
        self.assertContains(response, 'относится к другой модели')


class DefectActDateTests(TestCase):
    """Дата акта — день первой записи, а не день печати."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='defect_date', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-DATE-1'
            ),
        )

    def _fill(self, **extra):
        data = {
            'defect_act_date': '', 'diagnosis': 'Высохли конденсаторы',
            'error_codes': '', 'warranty_case': '',
            'non_warranty_reason': '', 'estimated_cost': '',
        }
        data.update(extra)
        return self.http.post(
            reverse('repair_order_defect_act_edit',
                    args=[self.order.pk, self.roe.pk]),
            data,
        )

    def test_the_date_is_stamped_on_the_first_entry(self):
        self._fill()

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.defect_act_date, timezone.localdate())

    def test_it_does_not_drift_on_later_saves(self):
        """Заполнили в понедельник, вернулись в среду — в акте должен
        остаться понедельник."""
        monday = timezone.localdate() - datetime.timedelta(days=2)
        self.roe.diagnosis = 'Высохли конденсаторы'
        self.roe.defect_act_date = monday
        self.roe.save()

        self._fill(defect_act_date=monday.strftime('%Y-%m-%d'),
                   diagnosis='Высохли конденсаторы и резистор')

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.defect_act_date, monday)

    def test_a_hand_written_date_wins(self):
        """Диагностировать могли вчера, а записать сегодня."""
        yesterday = timezone.localdate() - datetime.timedelta(days=1)

        self._fill(defect_act_date=yesterday.strftime('%Y-%m-%d'))

        self.roe.refresh_from_db()
        self.assertEqual(self.roe.defect_act_date, yesterday)

    def test_an_empty_act_is_not_stamped(self):
        """Дата на пустом акте означала бы, что диагностика была."""
        self._fill(diagnosis='')

        self.roe.refresh_from_db()
        self.assertIsNone(self.roe.defect_act_date)

    def test_the_printed_act_carries_the_stamped_date(self):
        self._fill()
        self.roe.refresh_from_db()

        html = self.http.get(
            reverse('repair_order_act_defect',
                    args=[self.order.pk, self.roe.pk])
        ).content.decode()

        self.assertIn(self.roe.defect_act_date.strftime('%d.%m.%Y'), html)


class ReadinessButtonsTests(TestCase):
    """Чек-лист: кнопки, порядок по ходу работы, стоимость по факту."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='readiness_buttons', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-BTN-1'
            ),
        )

    def test_the_order_follows_the_work(self):
        """Вскрыли и записали, что нашли; починили и записали, что сделали;
        взяли со склада намеченное; проставили, во что обошлось."""
        self.assertEqual(
            [check['code'] for check in self.roe.readiness()],
            ['defect_act', 'work', 'planned_parts', 'repair_cost'],
        )

    def test_the_cost_asked_for_is_the_one_the_invoice_uses(self):
        """Оценка из дефектации согласуется до ремонта; сумма заказа
        и счёт считаются по фактической стоимости."""
        self.roe.estimated_cost = Decimal('9000.00')
        self.roe.save()
        codes = [c['code'] for c in self.roe.readiness_pending]
        self.assertIn('repair_cost', codes)

        self.roe.repair_cost = Decimal('9000.00')
        self.roe.save()

        self.assertNotIn(
            'repair_cost', [c['code'] for c in self.roe.readiness_pending]
        )

    def test_every_check_leads_to_the_field(self):
        expected = {
            'defect_act': reverse(
                'repair_order_defect_act_edit',
                args=[self.order.pk, self.roe.pk]) + '#id_diagnosis',
            'work': reverse(
                'repair_order_unit_detail',
                args=[self.order.pk, self.roe.pk]) + '#id_work_performed',
            'planned_parts': reverse(
                'repair_order_unit_detail',
                args=[self.order.pk, self.roe.pk]) + '#parts',
            'repair_cost': reverse(
                'repair_order_unit_detail',
                args=[self.order.pk, self.roe.pk]) + '#id_repair_cost',
        }

        for check in self.roe.readiness():
            with self.subTest(code=check['code']):
                self.assertEqual(check['url'], expected[check['code']])

    def test_a_done_check_keeps_its_button(self):
        """Перепроверить написанное надо уметь в один щелчок."""
        self.roe.work_performed = 'Заменены конденсаторы'
        self.roe.save()

        done = [c for c in self.roe.readiness() if c['code'] == 'work'][0]

        self.assertTrue(done['done'])
        self.assertTrue(done['url'])

    def test_the_unit_page_draws_them_as_buttons(self):
        html = self.http.get(
            reverse('repair_order_unit_detail',
                    args=[self.order.pk, self.roe.pk])
        ).content.decode()
        block = html.split('Готовность')[1].split('</div></div>')[0]

        self.assertIn('id_work_performed', block)
        self.assertIn('btn', block)

    def test_the_anchor_puts_the_cursor_in_the_field(self):
        """Браузер к якорю прокрутит и сам, но на длинной странице этого
        мало: видно, что страница съехала, а куда вписывать — нет."""
        base = (settings.BASE_DIR / 'core/templates/core/base.html'
                ).read_text(encoding='utf-8')

        self.assertIn('function focusFromHash', base)
        self.assertIn("field.focus({preventScroll: true})", base)

    def test_the_parts_list_has_the_anchor_it_is_sent_to(self):
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

        self.assertIn('id="parts"', html)


# ============ ТИПОВЫЕ ВЫПОЛНЕННЫЕ РАБОТЫ (v2.81.0) ============


class TypicalWorkTests(TestCase):
    """У типовой неисправности есть не только «что нашли», но и «что делают».

    Два разных текста, и путать их нельзя: акт выполненных работ с фразой
    «конденсаторы потеряли ёмкость» читается как отчёт о том, что ничего
    не делали.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='typical_work', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.dried = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            description='Электролитические конденсаторы потеряли ёмкость.',
            work_description='Заменены электролитические конденсаторы в цепи питания.',
        )
        self.driver = FaultType.objects.create(
            equipment_model=self.model, name='отказал драйвер',
            description='Драйвер силового ключа не открывает транзистор.',
            work_description='Заменён драйвер силового ключа.',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-TW-1'
            ),
        )

    def _unit_url(self):
        return reverse('repair_order_unit_detail',
                       args=[self.order.pk, self.roe.pk])

    def test_the_two_texts_stay_separate(self):
        """Описание отвечает «что нашли», работы — «что сделали»."""
        self.roe.faults.add(self.dried)

        self.assertIn('потеряли ёмкость', self.roe.diagnosis_document_text)
        self.assertNotIn('Заменены', self.roe.diagnosis_document_text)
        self.assertEqual(
            self.roe.typical_work_lines,
            ['Заменены электролитические конденсаторы в цепи питания.'],
        )

    def test_lines_follow_the_order_of_the_faults(self):
        self.roe.faults.add(self.dried, self.driver)

        self.assertEqual(len(self.roe.typical_work_lines), 2)
        self.assertIn('конденсаторы', self.roe.typical_work_lines[0])

    def test_a_fault_without_typical_work_is_skipped(self):
        """Подставлять нечего, а короткое название вместо работ не идёт:
        оно не для документов."""
        silent = FaultType.objects.create(
            equipment_model=self.model, name='непонятный шум',
        )
        self.roe.faults.add(silent)

        self.assertEqual(self.roe.typical_work_lines, [])

    def test_nothing_is_written_into_the_act_by_itself(self):
        """Неисправности выбирают при дефектации, до ремонта: собери
        программа акт сама, он оказался бы готов по неразобранному
        прибору, и чек-лист показал бы «работы записаны»."""
        self.roe.faults.add(self.dried)
        self.roe.refresh_from_db()

        self.assertEqual(self.roe.work_performed, '')
        self.assertIn(
            'work', [check['code'] for check in self.roe.readiness_pending]
        )

    def test_the_button_carries_the_lines_to_the_page(self):
        self.roe.faults.add(self.dried)

        html = self.http.get(self._unit_url()).content.decode()

        self.assertIn('data-typical-work', html)
        self.assertIn('data-target="id_work_performed"', html)
        self.assertIn('Заменены электролитические конденсаторы', html)

    def test_the_lines_travel_as_json(self):
        """В описании работ бывают и запятые, и кавычки, и переводы
        строк — перечислением через запятую их не передать."""
        self.dried.work_description = 'Заменены C12, C13; проверен «стенд»'
        self.dried.save()
        self.roe.faults.add(self.dried)

        self.assertEqual(
            json.loads(self.roe.typical_work_json),
            ['Заменены C12, C13; проверен «стенд»'],
        )

    def test_the_fault_card_asks_for_both_texts(self):
        html = self.http.get(
            reverse('fault_type_edit', args=[self.dried.pk])
        ).content.decode()

        self.assertIn('Описание для документов', html)
        self.assertIn('Типовые выполненные работы', html)

    def test_the_field_is_saved_from_the_fault_card(self):
        response = self.http.post(reverse('fault_type_edit', args=[self.dried.pk]), {
            'equipment_model': str(self.model.pk),
            'name': self.dried.name,
            'description': self.dried.description,
            'work_description': 'Заменены конденсаторы C12, C13 и проверен стенд.',
            'complexity': 'simple',
            'parts-TOTAL_FORMS': '0', 'parts-INITIAL_FORMS': '0',
            'parts-MIN_NUM_FORMS': '0', 'parts-MAX_NUM_FORMS': '1000',
        })

        self.dried.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertIn('C12, C13', self.dried.work_description)

    def test_a_copy_carries_the_typical_work(self):
        """Копирование карточки копирует всё — новое поле не исключение."""
        html = self.http.get(
            reverse('fault_type_create'), {'copy_from': self.dried.pk}
        ).content.decode()

        self.assertIn('Заменены электролитические конденсаторы', html)

    def test_the_quote_offers_the_same_button(self):
        """Предложение делают до ремонта, и текст тот же — только речь
        о будущем."""
        self.roe.faults.add(self.dried)

        html = self.http.get(
            reverse('repair_order_quote_edit', args=[self.order.pk])
        ).content.decode()

        self.assertIn('data-typical-work', html)
        self.assertIn('Подставить типовые работы', html)


class TypicalWorkScriptTests(SimpleTestCase):
    """Правила подстановки записаны в одном скрипте на обе страницы."""

    SCRIPT = Path(__file__).resolve().parent / 'static' / 'js' / 'typical-work.js'

    @property
    def code(self):
        source = self.SCRIPT.read_text(encoding='utf-8')
        source = re.sub(r'/\*.*?\*/', '', source, flags=re.S)
        return re.sub(r'^\s*//.*$', '', source, flags=re.M)

    def test_loaded_once_for_every_page(self):
        base = (settings.BASE_DIR / 'core/templates/core/base.html'
                ).read_text(encoding='utf-8')

        self.assertIn('js/typical-work.js', base)

    def test_typing_is_not_overwritten(self):
        """Мастер уже что-то вписал — строки дописываются, а не заменяют."""
        self.assertIn("var prefix = current.trim() ?", self.code)

    def test_a_second_click_does_not_duplicate(self):
        self.assertIn('added = lines.filter', self.code)
        self.assertIn('flat.indexOf(flatten(line)) === -1', self.code)

    def test_nothing_to_add_is_said_out_loud(self):
        """Молчаливое бездействие человек примет за поломку кнопки."""
        for text in ('типовых работ не заведено', 'уже вписаны', 'Подставлено строк'):
            with self.subTest(text=text):
                self.assertIn(text, self.SCRIPT.read_text(encoding='utf-8'))

    def test_no_bootstrap_and_no_extra_request(self):
        code = self.code

        self.assertNotIn('data-bs-', code)
        self.assertNotIn('fetch(', code)


# ============ ВЫБОР НЕИСПРАВНОСТЕЙ, ДЕТАЛИ, РАСТУЩИЕ ПОЛЯ (v2.82.0) ============


class FaultPickerTests(TestCase):
    """Неисправности выбираются из списка и добавляются кнопкой.

    Раньше это был список с множественным выбором — тот, где надо держать
    Ctrl. На планшете он почти неработоспособен, а при десятке
    неисправностей не видно, что вообще выбрано.
    """

    STATIC = Path(__file__).resolve().parent / 'static'
    TEMPLATES = Path(__file__).resolve().parent / 'templates' / 'core'

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='fault_picker', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.simple = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            complexity='simple',
        )
        self.hard = FaultType.objects.create(
            equipment_model=self.model, name='прошивка процессора',
            complexity='complex',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-FP-1'
            ),
        )

    def _url(self):
        return reverse('repair_order_unit_detail',
                       args=[self.order.pk, self.roe.pk])

    def test_the_multiple_select_is_gone(self):
        html = self.http.get(self._url()).content.decode()

        self.assertNotIn('multiple', html.split('fault-picker')[0][-400:])
        self.assertIn('fault-picker-add', html)
        self.assertIn('Добавить', html)

    def test_the_name_stays_the_same_so_the_view_did_not_change(self):
        """Наружу уходят те же поля под тем же именем — тот же приём,
        что у выбора детали."""
        self.roe.faults.add(self.simple)

        html = self.http.get(self._url()).content.decode()

        self.assertIn('name="faults" value="%d"' % self.simple.pk, html)

    def test_chosen_faults_are_drawn_by_the_server(self):
        """Страница показывает выбранное и до того, как отработал скрипт."""
        self.roe.faults.add(self.simple, self.hard)

        html = self.http.get(self._url()).content.decode()
        chosen = html.split('fault-picker-chosen')[1].split('fault-picker-empty')[0]

        self.assertIn('высохли конденсаторы', chosen)
        self.assertIn('прошивка процессора', chosen)

    def test_complexity_is_coloured(self):
        """Сложный — красный, простой — зелёный. Значений всего два."""
        self.roe.faults.add(self.simple, self.hard)

        html = self.http.get(self._url()).content.decode()
        chosen = html.split('fault-picker-chosen')[1].split('fault-picker-empty')[0]

        self.assertIn('bg-success', chosen)
        self.assertIn('bg-danger', chosen)

    def test_saving_the_choice_still_works(self):
        self.http.post(
            reverse('repair_order_unit_edit', args=[self.order.pk, self.roe.pk]),
            {'fault_description': '', 'initial_condition': '',
             'work_performed': '', 'seal_numbers': '', 'repair_cost': '',
             'yandex_disk_folder': '', 'faults': [str(self.hard.pk)]},
        )

        self.assertEqual(list(self.roe.faults.all()), [self.hard])

    def test_the_script_is_shared_and_bootstrap_free(self):
        base = (self.TEMPLATES / 'base.html').read_text(encoding='utf-8')
        code = (self.STATIC / 'js' / 'fault-picker.js').read_text(encoding='utf-8')

        self.assertIn('js/fault-picker.js', base)
        self.assertNotIn('data-bs-', code)
        self.assertNotIn('bootstrap', code)

    def test_an_added_fault_leaves_the_dropdown(self):
        """Второй раз ту же неисправность не добавить."""
        code = (self.STATIC / 'js' / 'fault-picker.js').read_text(encoding='utf-8')

        self.assertIn('option.hidden = taken.indexOf(option.value) >= 0', code)


class PartsOnUnitPageTests(TestCase):
    """Детали правятся со страницы единицы: там её не надо выбирать."""

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='parts_on_unit', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-PU-1'
            ),
        )
        self.other = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-PU-2'
            ),
        )
        self.part = SparePart.objects.create(
            part_number='C-100u', name='Конденсатор 100 мкФ', current_stock=10
        )

    def _unit_url(self):
        return reverse('repair_order_unit_detail',
                       args=[self.order.pk, self.roe.pk])

    def test_the_form_is_on_the_unit_page(self):
        html = self.http.get(self._unit_url()).content.decode()

        self.assertIn('name="order_equipment" value="%d"' % self.roe.pk, html)
        self.assertIn('Списать', html)
        self.assertIn('В план', html)

    def test_a_part_written_off_here_lands_on_this_unit(self):
        response = self.http.post(
            reverse('repair_order_add_detail', args=[self.order.pk]),
            {'part': str(self.part.pk), 'quantity_used': '2',
             'order_equipment': str(self.roe.pk),
             'next': self._unit_url() + '#parts'},
        )

        detail = RepairOrderDetail.objects.get()
        self.assertEqual(detail.order_equipment, self.roe)
        self.assertEqual(detail.quantity_used, 2)
        # И возвращает туда, откуда пришли: три детали подряд — три
        # раза уезжать в заказ незачем
        self.assertEqual(response['Location'], self._unit_url() + '#parts')

    def test_a_foreign_return_address_is_refused(self):
        response = self.http.post(
            reverse('repair_order_add_detail', args=[self.order.pk]),
            {'part': str(self.part.pk), 'quantity_used': '1',
             'order_equipment': str(self.roe.pk), 'next': 'https://example.com/'},
        )

        self.assertEqual(
            response['Location'],
            reverse('repair_order_detail', args=[self.order.pk]),
        )

    def test_the_order_card_still_writes_off_to_the_order_as_a_whole(self):
        """Там списывают припой, стяжки, промывку — то, что к прибору
        не привязано. Убери форму — списывать общее станет негде."""
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

        self.assertIn('id="usePartForm"', html)
        self.assertIn('На заказ целиком', html)

    def test_planned_parts_are_written_off_from_here(self):
        planned = RepairOrderDetail.objects.create(
            repair_order=self.order, order_equipment=self.roe,
            part=self.part, quantity_used=3, is_planned=True,
        )

        html = self.http.get(self._unit_url()).content.decode()
        self.assertIn(
            reverse('repair_order_write_off_detail',
                    args=[self.order.pk, planned.pk]), html)

        self.http.post(
            reverse('repair_order_write_off_detail',
                    args=[self.order.pk, planned.pk]),
            {'next': self._unit_url() + '#parts'},
        )

        detail = RepairOrderDetail.objects.get()
        self.assertFalse(detail.is_planned)
        self.assertEqual(detail.order_equipment, self.roe)

    def test_parts_of_another_unit_are_not_shown(self):
        RepairOrderDetail.objects.create(
            repair_order=self.order, order_equipment=self.other,
            part=self.part, quantity_used=1,
        )

        html = self.http.get(self._unit_url()).content.decode()
        block = html.split('Детали на эту единицу')[1]

        self.assertIn('деталей не списано', block)


class AutogrowTests(SimpleTestCase):
    """Многострочные поля ростом по содержимому.

    Раньше высота стояла в разметке: под однострочный ответ отводилось
    три строки пустоты, а под настоящий абзац всё равно не хватало.
    """

    STATIC = Path(__file__).resolve().parent / 'static'

    @property
    def code(self):
        return (self.STATIC / 'js' / 'autogrow.js').read_text(encoding='utf-8')

    def test_loaded_for_every_page(self):
        base = (settings.BASE_DIR / 'core/templates/core/base.html'
                ).read_text(encoding='utf-8')

        self.assertIn('js/autogrow.js', base)

    def test_no_form_needs_editing(self):
        """Скрипт находит поля сам — иначе про него забыли бы
        в первой же новой форме."""
        self.assertIn("querySelectorAll('textarea')", self.code)
        self.assertIn("event.target.tagName === 'TEXTAREA'", self.code)

    def test_it_stops_growing_somewhere(self):
        """Страница не должна уезжать из-за одного поля."""
        self.assertIn('MAX_ROWS = 12', self.code)
        self.assertIn("field.style.overflowY = 'auto'", self.code)

    def test_a_hidden_field_is_left_alone(self):
        """У скрытого поля нулевая высота содержимого, и подгонка
        схлопнула бы его в ничто."""
        self.assertIn('if (!field.offsetParent && field.offsetHeight === 0) return;', self.code)


# ============ ЦВЕТ СЛОЖНОСТИ, ОБЩЕЕ ОПИСАНИЕ УБРАНО (v2.83.0) ============


class ComplexityColourTests(TestCase):
    """Сложность помечена цветом, и цвет считается одним местом.

    Раскраска разъезжается быстрее всего: цепочка условий в шаблоне
    копируется в соседний, там её правят, и одно и то же начинает
    выглядеть по-разному. Так уже вышло со статусами заказа.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='complexity_colour', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.simple = FaultType.objects.create(
            equipment_model=self.model, name='высохли конденсаторы',
            complexity='simple',
        )
        self.hard = FaultType.objects.create(
            equipment_model=self.model, name='прошивка процессора',
            complexity='complex',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-CC-1'
            ),
        )

    def test_the_colour_is_decided_in_one_place(self):
        self.assertEqual(complexity_css('simple'), 'bg-success')
        self.assertEqual(complexity_css('complex'), 'bg-danger')
        self.assertEqual(complexity_css(''), '')

    def test_a_fault_and_a_unit_agree_on_the_colour(self):
        self.roe.faults.add(self.hard)

        self.assertEqual(self.hard.complexity_css, 'bg-danger')
        self.assertEqual(self.roe.effective_complexity_css, 'bg-danger')

    def test_not_set_is_not_painted(self):
        """«Не задавали» — это не сложность, и красить его незачем."""
        self.assertEqual(self.roe.effective_complexity, '')
        self.assertEqual(self.roe.effective_complexity_css, '')

    def test_the_units_list_shows_it(self):
        """Владелец попросил цвет и здесь. Столбец называется
        «Стоимость»: пока цифры нет, в нём стоит слово, а цвет
        одинаков в обоих случаях."""
        self.roe.faults.add(self.hard)

        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()
        cell = html.split('data-label="Стоимость"')[1].split('</td>')[0]

        self.assertIn('bg-danger', cell)
        self.assertIn('Сложный', cell)

    def test_the_fault_list_shows_it(self):
        html = self.http.get(reverse('fault_type_list')).content.decode()

        self.assertIn('bg-success', html)
        self.assertIn('bg-danger', html)


class UnitCostCellTests(TestCase):
    """Столбец «Стоимость» в списке единиц заказа.

    Стоимость и сложность слиты в один столбец: цифра — то, ради чего
    в него смотрят, цвет несёт сложность. Слово показывается, только
    пока цифры нет.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='cost_cell', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        self.model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.hard = FaultType.objects.create(
            equipment_model=self.model, name='пробит модуль', complexity='complex',
        )
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=self.model, serial_number='SN-COST-1'
            ),
        )

    def _cell(self):
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()
        return html.split('data-label="Стоимость"')[1].split('</td>')[0]

    def test_the_amount_carries_the_complexity_colour(self):
        """Есть цифра — показывается она, а сложность остаётся цветом."""
        self.roe.repair_cost = Decimal('18500.00')
        self.roe.repair_complexity = 'complex'
        self.roe.save()

        cell = self.roe.cost_cell

        # Пробел в сумме неразрывный: перенос строки посреди числа
        # превращает 18 500 в «18» и «500» на разных строках
        self.assertEqual(cell['text'], '18\u00a0500 \u20bd')
        self.assertEqual(cell['css'], 'bg-danger')
        self.assertIn('18\u00a0500', self._cell())
        self.assertNotIn('Сложный', self._cell())

    def test_an_amount_without_complexity_is_not_painted(self):
        """«Не задавали» — это не сложность, и красить её незачем."""
        self.roe.repair_cost = Decimal('4500.00')
        self.roe.save()

        self.assertEqual(self.roe.cost_cell['css'], '')

    def test_the_word_stands_in_until_the_amount_appears(self):
        self.roe.faults.add(self.hard)

        cell = self.roe.cost_cell

        self.assertEqual(cell['text'], 'Сложный')
        self.assertEqual(cell['css'], 'bg-danger')

    def test_a_warranty_repair_says_so_instead_of_a_dash(self):
        """Прочерк означает «не проставили», а на гарантии проставлять
        нечего — по той же причине пункт стоимости выпадает
        из готовности."""
        self.roe.warranty_case = 'warranty'
        self.roe.save()

        self.assertEqual(self.roe.cost_cell['text'], 'по гарантии')
        self.assertIn('по гарантии', self._cell())

    def test_nothing_filled_is_a_dash(self):
        self.assertEqual(self.roe.cost_cell['text'], '—')
        self.assertTrue(self.roe.cost_cell['muted'])

    def test_the_estimate_never_shows_up_here(self):
        """Решение владельца: в одном столбце две разные цифры путали бы,
        а какая из них перед глазами, из списка не видно."""
        self.roe.estimated_cost = Decimal('12000.00')
        self.roe.save()

        self.assertEqual(self.roe.cost_cell['text'], '—')
        self.assertNotIn('12\u00a0000', self._cell())

    def test_there_is_no_separate_complexity_column_any_more(self):
        """Строка и так широкая, а два значка рядом читаются хуже
        одного."""
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()

        self.assertNotIn('data-label="Сложность"', html)


class UnitReadinessLabelTests(TestCase):
    """«Готов», а не «Готова»: слово согласовано с прибором, а не
    со строкой заказа, и берётся из одного места."""

    def test_the_label_agrees_with_the_device(self):
        model = EquipmentModel.objects.create(name='БУАД-7-31')
        order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        roe = RepairOrderEquipment.objects.create(
            repair_order=order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-LABEL-1'
            ),
            diagnosis='Пробит модуль', work_performed='Заменён модуль',
            repair_cost=Decimal('5000.00'),
        )

        self.assertTrue(roe.is_ready)
        self.assertEqual(roe.readiness_label, 'Готов')


class MediumComplexityIsGoneTests(TestCase):
    """«Среднего» ремонта не бывает — решение владельца."""

    def test_only_two_values_are_offered(self):
        field = RepairOrderEquipment._meta.get_field('repair_complexity')

        self.assertEqual(
            [code for code, _ in field.choices], ['simple', 'complex']
        )

    def test_the_unit_and_the_fault_speak_the_same_language(self):
        """Сложность единицы выводится из неисправностей: разойдись
        наборы значений, вывод дал бы код, которого нет в списке."""
        unit = {code for code, _ in
                RepairOrderEquipment._meta.get_field('repair_complexity').choices}
        fault = {code for code, _ in
                 FaultType._meta.get_field('complexity').choices}

        self.assertEqual(unit, fault)


class OrderFaultDescriptionIsGoneTests(TestCase):
    """Общее описание неисправности у заказа убрано.

    Его заполняли вместо описания по прибору — то есть про одно и то же
    спрашивали дважды. По базе владельца оно не было заполнено ни в одном
    заказе из восьми.
    """

    def setUp(self):
        self.employee = Employee.objects.create_superuser(
            username='no_common_fault', full_name='Мастер', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.employee)

        model = EquipmentModel.objects.create(name='БУАД-7-31')
        self.order = RepairOrder.objects.create(
            client=ClientModel.objects.create(name='МУП «Лифты»')
        )
        self.roe = RepairOrderEquipment.objects.create(
            repair_order=self.order,
            equipment=Equipment.objects.create(
                model=model, serial_number='SN-NC-1'
            ),
            fault_description='Не открывает двери',
        )

    def test_the_field_is_gone_from_the_model(self):
        self.assertFalse(
            [f.name for f in RepairOrder._meta.get_fields()
             if f.name == 'fault_description']
        )

    def test_intake_asks_only_for_the_client(self):
        html = self.http.get(reverse('repair_order_create')).content.decode()
        head = html.split('equipments-0')[0]

        self.assertNotIn('name="fault_description"', head)

    def test_the_order_card_asks_only_for_the_client(self):
        html = self.http.get(
            reverse('repair_order_detail', args=[self.order.pk])
        ).content.decode()
        form = html.split('edit-info/')[1].split('</form>')[0]

        self.assertIn('name="client"', form)
        self.assertNotIn('name="fault_description"', form)

    def test_the_intake_act_prints_the_units_own_text(self):
        """Заявленная неисправность стоит в таблице у каждой единицы —
        там, где её и читают."""
        html = self.http.get(
            reverse('repair_order_act_receive', args=[self.order.pk])
        ).content.decode()

        self.assertIn('Не открывает двери', html)
        self.assertNotIn('Со слов заказчика:', html)

    def test_search_still_finds_the_text(self):
        found = self.http.get(
            reverse('repair_order_list'), {'q': 'двери'}
        ).context['orders']

        self.assertEqual([o.pk for o in found], [self.order.pk])


class OrderFaultDescriptionMigrationTests(TransactionTestCase):
    """Перед удалением столбца текст перекладывается в единицы.

    Ровно так его и печатал акт приёма: описание единицы, а если его нет —
    общее по заказу. Так что ничей текст не пропадает, а оказывается там,
    где его и читали.
    """

    BEFORE = '0046_faulttype_work_description'
    AFTER = '0047_drop_order_fault_description'

    available_apps = None

    def _migrate(self, target):
        from django.db import connection as db_connection
        from django.db.migrations.executor import MigrationExecutor

        executor = MigrationExecutor(db_connection)
        executor.loader.build_graph()
        return executor.migrate([('core', target)])

    def tearDown(self):
        from django.db.migrations.loader import MigrationLoader
        names = sorted(
            name for app, name in
            MigrationLoader(None, ignore_no_migrations=True).disk_migrations
            if app == 'core'
        )
        self._migrate(names[-1])

    def test_the_text_moves_to_units_without_their_own(self):
        old_state = self._migrate(self.BEFORE)

        OldClient = old_state.apps.get_model('core', 'Client')
        OldModel = old_state.apps.get_model('core', 'EquipmentModel')
        OldEquipment = old_state.apps.get_model('core', 'Equipment')
        OldOrder = old_state.apps.get_model('core', 'RepairOrder')
        OldUnit = old_state.apps.get_model('core', 'RepairOrderEquipment')

        model = OldModel.objects.create(name='БУАД-МИГР')
        order = OldOrder.objects.create(
            client=OldClient.objects.create(name='ООО Миграция', inn='7700000906'),
            order_number='FD-1',
            fault_description='Общее: не работает после грозы',
        )
        silent = OldUnit.objects.create(
            repair_order=order,
            equipment=OldEquipment.objects.create(model=model, serial_number='FD-SILENT'),
        )
        speaking = OldUnit.objects.create(
            repair_order=order,
            equipment=OldEquipment.objects.create(model=model, serial_number='FD-OWN'),
            fault_description='Своё: гудит',
        )
        # И единица «среднего» ремонта — такого больше не бывает
        speaking.repair_complexity = 'medium'
        speaking.save()

        self._migrate(self.AFTER)

        silent = RepairOrderEquipment.objects.get(pk=silent.pk)
        speaking = RepairOrderEquipment.objects.get(pk=speaking.pk)

        self.assertEqual(silent.fault_description, 'Общее: не работает после грозы')
        # Своё не перебито общим
        self.assertEqual(speaking.fault_description, 'Своё: гудит')
        # «Средний» переведён в «сложный»: занизить сложность значит
        # занизить цену по прайсу, а завысить — попросить перечитать
        self.assertEqual(speaking.repair_complexity, 'complex')


# ============ ЖИВЫЕ НАСТРОЙКИ: ФАЙЛ .env КАК ХРАНИЛИЩЕ (v2.86.0) ============


class EnvFileMixin:
    """Общая оснастка: свой файл настроек во временном каталоге.

    Своё окружение обязательно: настоящий `.env` разработчика не должен
    ни читаться, ни тем более правиться тестами.

    Примешивается, а не наследуется: части проверок база не нужна вовсе
    (SimpleTestCase), а странице настроек нужна — и её `TestCase` обязан
    откатывать записи между проверками сам.
    """

    def setUp(self):
        super().setUp()
        self.dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.dir.cleanup)
        self.path = Path(self.dir.name) / '.env'
        self.override = override_settings(ENV_FILE_PATH=str(self.path))
        self.override.enable()
        self.addCleanup(self.override.disable)
        envfile.forget()
        self.addCleanup(envfile.forget)

    def write(self, text, mode=0o600):
        self.path.write_text(text, encoding='utf-8')
        self.path.chmod(mode)
        envfile.forget()


class EnvFileTestCase(EnvFileMixin, SimpleTestCase):
    """Проверки самого файла настроек — базы им не нужно."""


class EnvFileReadingTests(EnvFileTestCase):
    """Файл главнее снимка, снятого при запуске, — но только там,
    где его после запуска правили."""

    def test_a_value_from_the_file_wins(self):
        """Ради этого всё и затевалось: правка .env действует без
        перезапуска службы."""
        self.write('TBANK_TOKEN=из-файла\n')

        with override_settings(TBANK_TOKEN='из-настроек'):
            self.assertEqual(envfile.setting('TBANK_TOKEN', ''), 'из-файла')

    def test_an_untouched_file_does_not_win(self):
        """Значение совпало со снимком в окружении — файл не правили,
        и разбирать строку заново незачем. На этом же держится
        override_settings у того, у кого рядом лежит свой .env."""
        self.write('TBANK_TOKEN=одинаково\n')

        with patch.dict(os.environ, {'TBANK_TOKEN': 'одинаково'}):
            with override_settings(TBANK_TOKEN='из-настроек'):
                self.assertEqual(
                    envfile.setting('TBANK_TOKEN', ''), 'из-настроек'
                )

    def test_a_name_absent_from_the_file_falls_through(self):
        self.write('YANDEX_DISK_ROOT=Другое\n')

        with override_settings(TBANK_TOKEN='из-настроек'):
            self.assertEqual(envfile.setting('TBANK_TOKEN', ''), 'из-настроек')

    def test_no_file_at_all_breaks_nothing(self):
        self.assertFalse(envfile.exists())
        self.assertEqual(envfile.values(), {})

        with override_settings(TBANK_TOKEN='из-настроек'):
            self.assertEqual(envfile.setting('TBANK_TOKEN', ''), 'из-настроек')

    def test_the_type_comes_from_the_settings_value(self):
        """Тип настройки описан в settings.py и там же разобран — здесь
        он не дублируется, а берётся у образца."""
        self.write(
            'TBANK_INVOICE_ENABLED=True\n'
            'TBANK_INVOICE_DUE_DAYS=21\n'
            'WEBHOOKS_TRUSTED_PROXIES=127.0.0.1, 10.0.0.1\n'
        )

        with override_settings(TBANK_INVOICE_ENABLED=False,
                               TBANK_INVOICE_DUE_DAYS=14,
                               WEBHOOKS_TRUSTED_PROXIES=[]):
            self.assertIs(envfile.setting('TBANK_INVOICE_ENABLED', False), True)
            self.assertEqual(envfile.setting('TBANK_INVOICE_DUE_DAYS', 14), 21)
            self.assertEqual(
                envfile.setting('WEBHOOKS_TRUSTED_PROXIES', []),
                ['127.0.0.1', '10.0.0.1'],
            )

    def test_a_broken_number_keeps_the_old_one(self):
        """Опечатка в файле не должна ронять выставление счетов."""
        self.write('TBANK_INVOICE_DUE_DAYS=через неделю\n')

        with override_settings(TBANK_INVOICE_DUE_DAYS=14):
            self.assertEqual(envfile.setting('TBANK_INVOICE_DUE_DAYS', 14), 14)

    def test_a_composite_setting_is_left_alone(self):
        """ORDER_OVERDUE_DAYS собирается в settings.py из четырёх
        переменных, и одной строкой её не задать."""
        self.write('ORDER_OVERDUE_DAYS=7\n')

        with override_settings(ORDER_OVERDUE_DAYS={'repair': 7}):
            self.assertEqual(
                envfile.setting('ORDER_OVERDUE_DAYS', {}), {'repair': 7}
            )

    def test_a_rewritten_file_is_re_read(self):
        """Разобранное держится до следующей правки, а не навсегда."""
        self.write('TBANK_TOKEN=первый\n')
        self.assertEqual(envfile.setting('TBANK_TOKEN', ''), 'первый')

        self.write('TBANK_TOKEN=второй\n')

        self.assertEqual(envfile.setting('TBANK_TOKEN', ''), 'второй')


class EnvFileModulesUseItTests(EnvFileTestCase):
    """Модули банков, Диска и ботов читают настройки через envfile,
    а не напрямую: иначе правка .env не значила бы ничего до перезапуска."""

    def test_the_bank_the_disk_and_the_bots_pick_it_up(self):
        self.write(
            'TBANK_TOKEN=банк\n'
            'TBANK_ACCOUNT=40802810000000000001\n'
            'YANDEX_DISK_TOKEN=диск\n'
            'YANDEX_DISK_ROOT=Другой\n'
            'TELEGRAM_BOT_TOKEN=телеграм\n'
            'MAX_BOT_TOKEN=макс\n'
            'TOCHKA_TOKEN=точка\n'
        )

        self.assertEqual(tbank.token(), 'банк')
        self.assertEqual(tbank.account_number(), '40802810000000000001')
        self.assertEqual(yadisk.token(), 'диск')
        self.assertEqual(yadisk.root(), 'Другой')
        self.assertEqual(tochka.token(), 'точка')
        self.assertTrue(messengers.telegram_is_configured())
        self.assertTrue(messengers.max_is_configured())

    def test_the_webhook_secret_picks_it_up_too(self):
        """Секрет уведомлений меняют, когда меняют его у банка, —
        и ждать перезапуска в этот момент неоткуда."""
        self.write('WEBHOOKS_TBANK_SECRET=новый-секрет\n')

        self.assertEqual(
            webhooks._setting('WEBHOOKS_TBANK_SECRET', ''), 'новый-секрет'
        )


class EnvFileWritingTests(EnvFileTestCase):
    """Запись: атомарная, с сохранением комментариев и прав."""

    def test_comments_and_other_lines_survive(self):
        self.write(
            '# Настройки LiftTeam\n'
            'TBANK_TOKEN=старый\n'
            '\n'
            '# Комментарий про Диск\n'
            'YANDEX_DISK_ROOT=LiftTeam\n'
        )

        envfile.set_value('YANDEX_DISK_ROOT', 'Новый')

        text = self.path.read_text(encoding='utf-8')
        self.assertIn('# Настройки LiftTeam', text)
        self.assertIn('# Комментарий про Диск', text)
        self.assertIn('TBANK_TOKEN', text)
        self.assertEqual(envfile.values()['YANDEX_DISK_ROOT'], 'Новый')

    def test_the_file_stays_closed_from_others(self):
        """В файле токены банков. Промежутка, в котором его читают все,
        быть не должно — ни при правке, ни при создании."""
        self.write('YANDEX_DISK_ROOT=LiftTeam\n')

        envfile.set_value('YANDEX_DISK_ROOT', 'Новый')

        self.assertEqual(stat.S_IMODE(self.path.stat().st_mode), 0o600)

    def test_a_new_file_is_created_closed(self):
        envfile.set_value('YANDEX_DISK_ROOT', 'Новый')

        self.assertTrue(self.path.is_file())
        self.assertEqual(stat.S_IMODE(self.path.stat().st_mode), 0o600)

    def test_a_copy_is_left_before_the_change(self):
        """Человеку, сидящему рядом с Pi, должно быть куда откатиться."""
        self.write('YANDEX_DISK_ROOT=LiftTeam\n')

        envfile.set_value('YANDEX_DISK_ROOT', 'Новый')

        backup = self.path.with_name('.env.bak')
        self.assertIn('LiftTeam', backup.read_text(encoding='utf-8'))
        self.assertEqual(stat.S_IMODE(backup.stat().st_mode), 0o600)

    def test_a_value_with_spaces_and_a_hash_survives_a_round_trip(self):
        """Значение записывается и читается одним и тем же разбором —
        своего здесь нет намеренно."""
        envfile.set_value('YANDEX_DISK_ROOT', 'Папка # два слова')

        self.assertEqual(
            envfile.values()['YANDEX_DISK_ROOT'], 'Папка # два слова'
        )

    def test_a_multiline_value_is_refused(self):
        with self.assertRaises(envfile.EnvFileError):
            envfile.set_value('YANDEX_DISK_ROOT', 'первая\nвторая')

    def test_a_bad_name_is_refused(self):
        for name in ('', 'не имя', '1TOKEN', 'TOKEN;rm'):
            with self.assertRaises(envfile.EnvFileError):
                envfile.set_value(name, 'что-нибудь')


class EnvFileSecretsTests(EnvFileTestCase):
    """Секреты: из браузера не пишутся, наружу не показываются."""

    def test_the_web_may_not_write_a_secret(self):
        """Защита от промаха, а не от злоумышленника: файл принадлежит
        пользователю приложения. Но промах здесь и вероятнее."""
        for name in envfile.SECRET_NAMES:
            with self.assertRaises(envfile.EnvFileError):
                envfile.set_value(name, 'значение')

        self.assertFalse(self.path.exists())

    def test_the_command_may(self):
        envfile.set_value('TBANK_TOKEN', 'значение', allow_secrets=True)

        self.assertEqual(envfile.values()['TBANK_TOKEN'], 'значение')

    def test_the_description_gives_the_length_and_not_the_value(self):
        """Показанный один раз токен оседает в истории браузера,
        в кэше и на любом незапертом экране."""
        self.write('TBANK_TOKEN=t1.abcdefgh\n')

        state = envfile.describe_secret('TBANK_TOKEN')

        self.assertTrue(state['filled'])
        self.assertEqual(state['length'], len('t1.abcdefgh'))
        self.assertNotIn('abcdefgh', repr(state))

    def test_an_unset_secret_says_so(self):
        with override_settings(TBANK_TOKEN=''):
            state = envfile.describe_secret('TBANK_TOKEN')

        self.assertFalse(state['filled'])
        self.assertEqual(state['length'], 0)

    def test_the_email_password_is_marked_as_needing_a_restart(self):
        """Его читает почтовый слой Django, а он смотрит прямо в настройки,
        мимо этого модуля."""
        self.assertIn('EMAIL_HOST_PASSWORD', envfile.SECRETS_NEEDING_RESTART)
        self.assertNotIn('TBANK_TOKEN', envfile.SECRETS_NEEDING_RESTART)

    def test_the_secret_key_is_not_in_the_list(self):
        """Его заводят один раз при установке: смена разлогинивает всех
        и обесценивает подписанные ссылки."""
        self.assertNotIn('SECRET_KEY', envfile.SECRET_NAMES)


class SetSecretCommandTests(EnvFileTestCase):
    """Команда у самого Pi: значение спрашивается скрытым вводом
    и нигде не печатается."""

    def test_the_list_shows_state_and_never_a_value(self):
        self.write('TBANK_TOKEN=t1.секретное-значение\n')
        out = io.StringIO()

        call_command('setsecret', '--list', stdout=out)
        text = out.getvalue()

        self.assertIn('TBANK_TOKEN', text)
        self.assertIn('задан', text)
        self.assertNotIn('секретное-значение', text)

    def test_it_writes_what_was_typed(self):
        out = io.StringIO()

        with patch('getpass.getpass', side_effect=['t1.новый', 't1.новый']):
            call_command('setsecret', 'TBANK_TOKEN', '--no-check', stdout=out)

        envfile.forget()
        self.assertEqual(envfile.values()['TBANK_TOKEN'], 't1.новый')
        self.assertNotIn('t1.новый', out.getvalue())

    def test_a_mistyped_repeat_writes_nothing(self):
        """Токен вставляют из письма, ввод не отображается, и потерянный
        при вставке знак не видно."""
        with patch('getpass.getpass', side_effect=['первый', 'второй']):
            with self.assertRaises(CommandError):
                call_command('setsecret', 'TBANK_TOKEN', '--no-check',
                             stdout=io.StringIO())

        self.assertFalse(self.path.exists())

    def test_an_empty_value_asks_for_clear_instead(self):
        with patch('getpass.getpass', side_effect=['  ', '  ']):
            with self.assertRaises(CommandError):
                call_command('setsecret', 'TBANK_TOKEN', '--no-check',
                             stdout=io.StringIO())

    def test_clear_writes_an_empty_value(self):
        self.write('TBANK_TOKEN=старый\n')

        call_command('setsecret', 'TBANK_TOKEN', '--clear',
                     stdout=io.StringIO())

        envfile.forget()
        self.assertEqual(envfile.values()['TBANK_TOKEN'], '')

    def test_an_unknown_name_is_refused(self):
        with self.assertRaises(CommandError):
            call_command('setsecret', 'ALLOWED_HOSTS', '--no-check')

    def test_a_non_secret_setting_is_not_taken_here_either(self):
        """Команда — для секретов. Остальное правится страницей настроек."""
        with self.assertRaises(CommandError):
            call_command('setsecret', 'YANDEX_DISK_ROOT', '--no-check')


class SelfCheckTests(EnvFileTestCase):
    """Проверка связи одна на команду и на будущую страницу настроек:
    об одной и той же службе они обязаны говорить одно и то же."""

    def test_an_unset_service_is_skipped_not_failed(self):
        with override_settings(TBANK_TOKEN=''):
            result = selfcheck.check('TBANK_TOKEN')

        self.assertEqual(result.state, selfcheck.CheckResult.SKIPPED)
        self.assertFalse(result.ok)

    def test_a_reachable_bank_is_reported_ok(self):
        self.write('TBANK_TOKEN=t1.живой\n')

        with patch.object(tbank, 'get_accounts', return_value=[{'id': 1}]):
            result = selfcheck.check('TBANK_TOKEN')

        self.assertTrue(result.ok)

    def test_a_refusing_bank_is_reported_with_its_own_words(self):
        self.write('TBANK_TOKEN=t1.протухший\n')

        with patch.object(tbank, 'get_accounts',
                          side_effect=tbank.TBankError('401 Unauthorized')):
            result = selfcheck.check('TBANK_TOKEN')

        self.assertEqual(result.state, selfcheck.CheckResult.FAIL)
        self.assertIn('401', result.message)

    def test_tochka_says_that_the_check_is_not_written(self):
        """Угаданный адрес отвечал бы «связи нет» на исправном токене —
        это хуже честного «проверка не написана». Тот же порядок,
        что у её проверяющего уведомления."""
        self.write(
            'TOCHKA_TOKEN=точка\n'
            'TOCHKA_CUSTOMER_CODE=300000000\n'
            'TOCHKA_ACCOUNT_ID=40802810000000000001\n'
        )

        result = selfcheck.check('TOCHKA_TOKEN')

        self.assertEqual(result.state, selfcheck.CheckResult.SKIPPED)
        self.assertIn('не написано', result.message)

    def test_the_disk_check_only_reads(self):
        """Проверка связи не должна оставлять следов на Диске.

        Подменяется `urlopen`, а не `_call`: подменяя `_call`, я уже
        записал в заглушку своё представление о том, что он возвращает,
        — и оно разошлось с настоящим. Живой Диск отвечал
        «'int' object has no attribute 'get'», а тест был зелёным.
        """
        self.write('YANDEX_DISK_TOKEN=disk-token\n')
        seen = []

        def fake(req, timeout=None):
            seen.append(req)
            return FakeDiskResponse(200, json.dumps({
                'total_space': 10 * 2 ** 30, 'used_space': 2 ** 30,
            }))

        with patch('core.yadisk.request.urlopen', fake):
            result = selfcheck.check('YANDEX_DISK_TOKEN')

        self.assertEqual(seen[0].method, 'GET')
        self.assertTrue(result.ok, result.message)
        self.assertIn('занято 1 ГБ из 10', result.message)

    def test_an_unknown_name_says_so_plainly(self):
        """Молчаливого «всё хорошо» о непроверенном быть не должно."""
        result = selfcheck.check('НЕИЗВЕСТНАЯ_НАСТРОЙКА')

        self.assertEqual(result.state, selfcheck.CheckResult.SKIPPED)
        self.assertIn('не написано', result.message)


class EnvSettingsGoThroughEnvFileTests(SimpleTestCase):
    """Модули, работающие со сторонними службами, читают настройки только
    через `envfile`.

    Прочитанная напрямую из `settings` настройка застывает на запуске:
    правка `.env` перестаёт действовать до перезапуска, а перезапустить
    себя приложение не может. Заметить это по одному новому вызову
    невозможно — он ничего не ломает, просто настройка молча не меняется.
    """

    MODULES = (
        'tbank.py', 'tochka.py', 'yadisk.py', 'messengers.py',
        'invoicing.py', 'webhooks.py', 'notifications.py',
    )

    def test_no_module_reads_settings_directly(self):
        base = Path(__file__).resolve().parent
        offenders = []
        for name in self.MODULES:
            source = (base / name).read_text(encoding='utf-8')
            for number, line in enumerate(source.splitlines(), start=1):
                if 'getattr(settings' in line:
                    offenders.append('%s:%d' % (name, number))

        self.assertEqual(offenders, [], (
            'Настройка читается мимо envfile — правка .env перестанет '
            'действовать без перезапуска: %s' % ', '.join(offenders)
        ))

    def test_every_secret_has_a_title(self):
        """Список секретов читает человек у Pi, и «TOCHKA_TOKEN» без
        пояснения ему ничего не говорит."""
        for name in envfile.SECRET_NAMES:
            self.assertIn(name, envfile.SECRET_TITLES)


class SettingsPageTests(EnvFileMixin, TestCase):
    """Страница настроек: правимое правится, секреты только видны."""

    def setUp(self):
        super().setUp()
        self.admin = Employee.objects.create_superuser(
            username='settings_admin', full_name='Администратор', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.admin)
        self.url = reverse('admin_settings')

    def test_only_an_admin_gets_in(self):
        """Здесь токены и адреса банков — мастеру тут делать нечего."""
        master = Employee.objects.create_user(
            username='settings_master', full_name='Мастер', password='pass',
            role='master',
        )
        other = TestClient()
        other.force_login(master)

        response = other.get(self.url)

        self.assertEqual(response.status_code, 302)

    def test_a_secret_is_shown_as_a_length_and_never_as_a_value(self):
        self.write('TBANK_TOKEN=t1.очень-секретное\n')

        html = self.http.get(self.url).content.decode()

        self.assertIn('задан', html)
        self.assertIn('18 знаков', html)
        self.assertNotIn('очень-секретное', html)

    def test_there_is_no_input_for_a_secret(self):
        """Поле ввода означало бы, что токен идёт по сети из браузера —
        а по локальному адресу браузер идёт без сертификата."""
        html = self.http.get(self.url).content.decode()

        for name in envfile.SECRET_NAMES:
            self.assertNotIn('name="%s"' % name, html)

    def test_saving_writes_only_what_changed(self):
        self.write('QUOTE_VALID_DAYS=14\nWARRANTY_MONTHS=12\n')

        self.http.post(reverse('admin_settings_save'), {
            'QUOTE_VALID_DAYS': '21',
            'WARRANTY_MONTHS': '12',
        })

        envfile.forget()
        self.assertEqual(envfile.values()['QUOTE_VALID_DAYS'], '21')
        self.assertEqual(
            list(SettingChange.objects.values_list('name', flat=True)),
            ['QUOTE_VALID_DAYS'],
        )

    def test_an_unchecked_flag_means_no_and_not_silence(self):
        """Невыбранный флажок в запрос не попадает вовсе — принять это
        за «поле не показывали» значило бы, что выключить нельзя ничего."""
        self.write('NOTIFY_TELEGRAM=True\n')

        self.http.post(reverse('admin_settings_save'),
                       {'flag': 'NOTIFY_TELEGRAM'})

        envfile.forget()
        self.assertEqual(envfile.values()['NOTIFY_TELEGRAM'], 'False')

    def test_a_flag_missing_from_the_form_is_left_alone(self):
        """Отличить «снял галочку» от «поля на форме не было» по одному
        отсутствию имени нельзя, а молча выключить оповещения нельзя
        тем более."""
        self.write('NOTIFY_TELEGRAM=True\n')

        self.http.post(reverse('admin_settings_save'),
                       {'QUOTE_VALID_DAYS': '21'})

        envfile.forget()
        self.assertEqual(envfile.values()['NOTIFY_TELEGRAM'], 'True')

    def test_the_change_takes_effect_without_a_restart(self):
        self.http.post(reverse('admin_settings_save'),
                       {'YANDEX_DISK_ROOT': 'Другая-папка'})

        self.assertEqual(yadisk.root(), 'Другая-папка')

    def test_the_journal_keeps_who_and_when_but_not_the_value(self):
        self.http.post(reverse('admin_settings_save'),
                       {'MAX_GROUP_CHAT_ID': '-100500'})

        change = SettingChange.objects.get()
        self.assertEqual(change.name, 'MAX_GROUP_CHAT_ID')
        self.assertEqual(change.changed_by, self.admin)
        self.assertNotIn('100500', repr(change.__dict__))

    def test_a_setting_outside_the_list_is_not_written(self):
        """Подсунуть в запрос можно что угодно, а список правимого явный.
        `DEBUG` в нём нет намеренно: включённый, он показывает в браузере
        устройство программы вместе с настройками."""
        self.http.post(reverse('admin_settings_save'),
                       {'DEBUG': 'True', 'SECRET_KEY': 'подменённый'})

        envfile.forget()
        self.assertNotIn('DEBUG', envfile.values())
        self.assertNotIn('SECRET_KEY', envfile.values())
        self.assertFalse(SettingChange.objects.exists())

    def test_a_secret_slipped_into_the_request_is_not_written(self):
        self.http.post(reverse('admin_settings_save'),
                       {'TBANK_TOKEN': 'через-браузер'})

        envfile.forget()
        self.assertNotIn('TBANK_TOKEN', envfile.values())

    def test_settings_needing_a_restart_say_so(self):
        response = self.http.post(
            reverse('admin_settings_save'),
            {'LABEL_BASE_URL': 'http://новое-имя.ts.net'},
            follow=True,
        )
        html = response.content.decode()

        self.assertIn('перезапуск', html)

    def test_the_check_button_answers_with_the_shared_verdict(self):
        """Кнопка и команда у Pi обязаны говорить об одной службе одно
        и то же — проверку держит одно место."""
        with patch.object(selfcheck, 'check',
                          return_value=selfcheck.CheckResult('ok', 'принял')):
            response = self.http.post(reverse('admin_settings_check'),
                                      {'name': 'TBANK_TOKEN'})

        self.assertEqual(response.json(), {'state': 'ok', 'message': 'принял'})

    def test_the_check_refuses_an_unknown_name(self):
        response = self.http.post(reverse('admin_settings_check'),
                                  {'name': 'ALLOWED_HOSTS'})

        self.assertEqual(response.status_code, 400)

    def test_saving_is_post_only(self):
        response = self.http.get(reverse('admin_settings_save'))

        self.assertEqual(response.status_code, 405)

    def test_a_cyrillic_letter_in_a_token_is_named_out_loud(self):
        """«с» вместо «c» видно только так: сообщение самого исключения
        («latin-1 codec can't encode») не говорит об этом ничего."""
        self.write('TBANK_TOKEN=t1.токен\n')
        error = UnicodeEncodeError('latin-1', 'т', 0, 1, 'ordinal not in range')

        with patch.object(tbank, 'get_accounts', side_effect=error):
            result = selfcheck.check('TBANK_TOKEN')

        self.assertIn('кириллическая', result.message)


class SettingsPageIsListedTests(TestCase):
    """Пункт меню: страница, до которой нельзя дойти, всё равно что
    её нет."""

    def test_an_admin_sees_the_link(self):
        admin = Employee.objects.create_superuser(
            username='menu_admin', full_name='Администратор', password='pass',
        )
        http = TestClient()
        http.force_login(admin)

        html = http.get(reverse('dashboard')).content.decode()

        self.assertIn(reverse('admin_settings'), html)


class RestartFromTheInterfaceTests(EnvFileMixin, TestCase):
    """Перезапуск службы по заявке: приложение root не имеет и иметь
    не должно."""

    def setUp(self):
        super().setUp()
        self.admin = Employee.objects.create_superuser(
            username='restart_admin', full_name='Администратор', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.admin)
        self.request_file = Path(self.dir.name) / '.restart-request'
        patcher = patch.object(restarter, 'REQUEST_FILE', self.request_file)
        patcher.start()
        self.addCleanup(patcher.stop)

    def _available(self, yes=True):
        return patch.object(restarter, 'is_available', return_value=yes)

    def test_the_request_is_a_file_and_not_a_command(self):
        """Приложение работает от своего пользователя: дай ему sudo —
        и любая уязвимость в нём означала бы root на устройстве."""
        with self._available():
            self.http.post(reverse('admin_settings_restart'))

        self.assertTrue(self.request_file.exists())
        self.assertIn('restart_admin',
                      self.request_file.read_text(encoding='utf-8'))

    def test_without_the_privileged_script_nothing_is_promised(self):
        """Заявка легла бы и осталась лежать, а страница обещала бы
        перезапуск, которого не будет."""
        with self._available(False):
            response = self.http.post(reverse('admin_settings_restart'),
                                      follow=True)

        self.assertFalse(self.request_file.exists())
        self.assertIn('не настроен', response.content.decode())

    def test_only_an_admin_may_ask(self):
        master = Employee.objects.create_user(
            username='restart_master', full_name='Мастер', password='pass',
            role='master',
        )
        other = TestClient()
        other.force_login(master)

        with self._available():
            other.post(reverse('admin_settings_restart'))

        self.assertFalse(self.request_file.exists())

    def test_it_is_post_only(self):
        response = self.http.get(reverse('admin_settings_restart'))

        self.assertEqual(response.status_code, 405)

    def test_the_need_for_a_restart_is_derived_and_not_stored(self):
        """Флаг «нужен перезапуск» пришлось бы гасить руками, и однажды
        он начал бы врать — то же правило, что у готовности единицы."""
        self.assertEqual(envfile.restart_needed(), [])

        self.write('TIME_ZONE=Asia/Novosibirsk\n')

        self.assertIn('TIME_ZONE', envfile.restart_needed())

    def test_a_live_setting_does_not_ask_for_a_restart(self):
        self.write('QUOTE_VALID_DAYS=21\n')

        self.assertEqual(envfile.restart_needed(), [])

    def test_the_list_empties_itself_after_a_restart(self):
        """После перезапуска снимок в окружении совпадёт с файлом —
        и выводить будет нечего, гасить ничего не надо."""
        self.write('TIME_ZONE=Asia/Novosibirsk\n')

        with patch.dict(os.environ, {'TIME_ZONE': 'Asia/Novosibirsk'}):
            self.assertEqual(envfile.restart_needed(), [])

    def test_the_page_says_what_is_waiting_for_a_restart(self):
        self.write('TIME_ZONE=Asia/Novosibirsk\n')

        html = self.http.get(reverse('admin_settings')).content.decode()

        self.assertIn('Часовой пояс', html)
        self.assertIn('ещё не действуют', html)


class AllowedHostsSafetyTests(EnvFileMixin, TestCase):
    """`ALLOWED_HOSTS` — единственная настройка, ошибка в которой закрывает
    программу целиком: Django ответит «400 Bad Request» на всё, и починить
    это можно будет только по SSH."""

    def setUp(self):
        super().setUp()
        self.admin = Employee.objects.create_superuser(
            username='hosts_admin', full_name='Администратор', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

    def test_the_current_address_is_put_back(self):
        response = self.http.post(
            reverse('admin_settings_save'),
            {'ALLOWED_HOSTS': 'lifteam.example.ts.net'},
            follow=True,
        )

        envfile.forget()
        self.assertIn('testserver', envfile.values()['ALLOWED_HOSTS'])
        self.assertIn('lifteam.example.ts.net', envfile.values()['ALLOWED_HOSTS'])
        self.assertIn('возвращён в список', response.content.decode())

    def test_an_address_that_is_already_there_is_not_doubled(self):
        self.http.post(reverse('admin_settings_save'),
                       {'ALLOWED_HOSTS': 'testserver,lifteam.example.ts.net'})

        envfile.forget()
        self.assertEqual(
            envfile.values()['ALLOWED_HOSTS'],
            'testserver,lifteam.example.ts.net',
        )

    def test_an_empty_list_is_left_empty(self):
        """Пусто — это не «закрылись от себя», а «вернуть встроенное»."""
        self.write('ALLOWED_HOSTS=что-то\n')

        self.http.post(reverse('admin_settings_save'), {'ALLOWED_HOSTS': ''})

        envfile.forget()
        self.assertEqual(envfile.values()['ALLOWED_HOSTS'], '')


class RestartRunnerScriptTests(SimpleTestCase):
    """Привилегированный скрипт: он выполняется от root, и цена ошибки
    в нём — всё устройство."""

    def setUp(self):
        self.script = (Path(__file__).resolve().parent.parent / 'deploy'
                       / 'lifteam-restart-runner.sh').read_text(encoding='utf-8')

    def test_it_reads_nothing_from_the_request(self):
        """Файл пишет веб-приложение. У обновления из него читается версия,
        и её приходится проверять; здесь читать нечего — значит, и подставить
        нечего."""
        self.assertNotIn('$(python3', self.script.replace('\n', ''))
        self.assertNotIn('cat "$REQUEST_FILE"', self.script)
        self.assertNotIn('eval', self.script)

    def test_the_request_is_removed_before_the_restart(self):
        """Перезапуск обрывает и сам скрипт: оставленный файл заставил бы
        службу сработать снова, и приложение ушло бы в бесконечный
        перезапуск."""
        body = self.script
        self.assertLess(body.index('rm -f "$REQUEST_FILE"'),
                        body.index('systemctl restart'))

    def test_the_unit_watches_the_same_file_the_program_writes(self):
        unit = (Path(__file__).resolve().parent.parent / 'deploy'
                / 'lifteam-restart.path').read_text(encoding='utf-8')

        self.assertIn('PathExists=/opt/lifteam/.restart-request', unit)
        self.assertIn('.restart-request', self.script)
        self.assertEqual(restarter.REQUEST_FILE.name, '.restart-request')

    def test_the_script_lives_outside_the_application_directory(self):
        """Каталог приложения принадлежит его пользователю: скрипт оттуда
        взломщик подменил бы и получил выполнение от root."""
        self.assertTrue(str(restarter.RUNNER).startswith('/usr/local/sbin/'))


class ListSettingsRoundTripTests(EnvFileMixin, TestCase):
    """Списочные настройки: в поле ввода — строка через запятую, а не
    список Python.

    Ошибка стоила бы дорого: `['имя.ts.net']` уходит из поля обратно
    в файл, и при следующем запуске скобки с кавычками становятся частью
    адреса — программа перестаёт открываться по своему же имени.
    """

    def setUp(self):
        super().setUp()
        self.admin = Employee.objects.create_superuser(
            username='list_admin', full_name='Администратор', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

    def test_a_list_is_shown_as_a_comma_separated_line(self):
        with override_settings(ALLOWED_HOSTS=['127.0.0.1', 'lifteam.ts.net']):
            row = envfile.describe_editable('ALLOWED_HOSTS')

        self.assertEqual(row['value'], '127.0.0.1,lifteam.ts.net')

    def test_an_empty_list_is_shown_as_an_empty_line(self):
        with override_settings(CSRF_TRUSTED_ORIGINS=[]):
            row = envfile.describe_editable('CSRF_TRUSTED_ORIGINS')

        self.assertEqual(row['value'], '')

    def test_an_untouched_empty_list_is_not_written_at_all(self):
        """Показали пустым, вернули пустым — правки не было, и в файле
        появляться нечему."""
        self.http.post(reverse('admin_settings_save'),
                       {'CSRF_TRUSTED_ORIGINS': ''})

        envfile.forget()
        self.assertNotIn('CSRF_TRUSTED_ORIGINS', envfile.values())
        self.assertFalse(SettingChange.objects.exists())

    def test_what_the_page_shows_is_what_comes_back(self):
        """Полный оборот: страница → форма → файл → страница."""
        with override_settings(CSRF_TRUSTED_ORIGINS=[]):
            self.http.post(
                reverse('admin_settings_save'),
                {'CSRF_TRUSTED_ORIGINS': 'https://lifteam.ts.net,https://192.168.1.50'},
            )
            envfile.forget()
            row = envfile.describe_editable('CSRF_TRUSTED_ORIGINS')

        self.assertEqual(
            row['value'], 'https://lifteam.ts.net,https://192.168.1.50'
        )
        self.assertEqual(
            envfile.setting('CSRF_TRUSTED_ORIGINS', []),
            ['https://lifteam.ts.net', 'https://192.168.1.50'],
        )

    def test_the_field_carries_no_brackets(self):
        # 'testserver' в списке обязателен: без него Django ответит
        # «400 Bad Request» самому тесту — ровно то, от чего бережёт
        # возврат своего адреса в AllowedHostsSafetyTests
        with override_settings(ALLOWED_HOSTS=['lifteam.ts.net', 'testserver']):
            html = self.http.get(reverse('admin_settings')).content.decode()
        field = html.split('id="set_ALLOWED_HOSTS"')[1].split('>')[0]

        self.assertIn('lifteam.ts.net,testserver', field)
        self.assertNotIn('[', field)


class TBankAccountListTests(SimpleTestCase):
    """Разбор списка счетов — одно место на команду и на проверку связи.

    Какой из вариантов ответа приходит на живом счёте, не подтверждено:
    сайт документации банка из среды разработки недоступен. Поэтому
    принимаются оба, и принимаются одинаково — второй разбор сказал бы
    «счетов доступно: 0» на исправном токене, пока команда рядом
    печатала бы их все.
    """

    ACCOUNT = {'accountNumber': '40802810700006096146', 'name': 'Расчётный'}

    def test_a_bare_list_is_understood(self):
        self.assertEqual(tbank.account_list([self.ACCOUNT]), [self.ACCOUNT])

    def test_a_wrapped_list_is_understood(self):
        for key in ('accounts', 'bankAccounts', 'items', 'data', 'result'):
            with self.subTest(key=key):
                self.assertEqual(
                    tbank.account_list({key: [self.ACCOUNT]}), [self.ACCOUNT]
                )

    def test_anything_else_gives_nothing_instead_of_an_error(self):
        for payload in (None, 42, 'счета', {}, {'accounts': 'нет'}):
            with self.subTest(payload=payload):
                self.assertEqual(tbank.account_list(payload), [])

    def test_the_numbers_are_pulled_out_for_a_human(self):
        self.assertEqual(
            tbank.account_numbers({'accounts': [
                self.ACCOUNT, {'number': '40802810700006096147'}, {'name': 'без номера'},
            ]}),
            ['40802810700006096146', '40802810700006096147'],
        )


class TBankCheckTests(EnvFileTestCase):
    """Что видит владелец, нажав «Проверить» у токена Т-Банка."""

    def test_a_wrapped_answer_is_not_reported_as_zero_accounts(self):
        """«Счетов доступно: 0» на исправном токене читается как отказ."""
        self.write('TBANK_TOKEN=t1.demo\n')

        with patch.object(tbank, 'get_accounts',
                          return_value={'accounts': [
                              {'accountNumber': '40802810700006096146'}]}):
            result = selfcheck.check('TBANK_TOKEN')

        self.assertTrue(result.ok)
        self.assertIn('40802810700006096146', result.message)

    def test_an_empty_answer_says_what_to_do_next(self):
        self.write('TBANK_TOKEN=t1.demo\n')

        with patch.object(tbank, 'get_accounts', return_value={}):
            result = selfcheck.check('TBANK_TOKEN')

        self.assertTrue(result.ok)
        self.assertIn('tbank_statement --accounts', result.message)


class StatementIntervalTests(SimpleTestCase):
    """Частоту загрузки выписки решает программа, а не расписание.

    Юнит systemd правится от root, а владелец меняет частоту со страницы
    настроек. Поэтому таймер только тикает, а «пора или нет» отвечает
    одно место.
    """

    def setUp(self):
        self.dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.dir.cleanup)
        self.marker = Path(self.dir.name) / '.tbank-last-run'
        patcher = patch.object(tbank, 'LAST_FETCH_FILE', self.marker)
        patcher.start()
        self.addCleanup(patcher.stop)

    def test_the_first_run_is_always_due(self):
        due, why = tbank.fetch_due()

        self.assertTrue(due)
        self.assertIn('ни разу', why)

    def test_too_soon_is_skipped_and_says_why(self):
        tbank.mark_fetched(timezone.now() - datetime.timedelta(minutes=10))

        with override_settings(TBANK_STATEMENT_INTERVAL_MINUTES=60):
            due, why = tbank.fetch_due()

        self.assertFalse(due)
        self.assertIn('Пропускаю', why)

    def test_after_the_interval_it_is_due_again(self):
        tbank.mark_fetched(timezone.now() - datetime.timedelta(minutes=61))

        with override_settings(TBANK_STATEMENT_INTERVAL_MINUTES=60):
            due, _why = tbank.fetch_due()

        self.assertTrue(due)

    def test_zero_means_every_tick(self):
        tbank.mark_fetched()

        with override_settings(TBANK_STATEMENT_INTERVAL_MINUTES=0):
            due, why = tbank.fetch_due()

        self.assertTrue(due)
        self.assertIn('на каждом тике', why)

    def test_a_broken_value_falls_back_to_an_hour(self):
        """Настройку правит человек, и в поле может оказаться что угодно.
        Загрузка выписки — не то место, где стоит падать."""
        with override_settings(TBANK_STATEMENT_INTERVAL_MINUTES='часто'):
            self.assertEqual(tbank.statement_interval(), 60)

    def test_a_missing_marker_file_is_not_an_error(self):
        self.assertIsNone(tbank.last_fetch_at())

    def test_a_damaged_marker_file_is_not_an_error(self):
        """Файл могли обрезать на середине записи при выключении питания —
        тогда просто тянем заново."""
        self.marker.write_text('позавчера', encoding='utf-8')

        self.assertIsNone(tbank.last_fetch_at())

    def test_the_marker_survives_a_round_trip(self):
        moment = timezone.now().replace(microsecond=0)

        tbank.mark_fetched(moment)

        self.assertEqual(tbank.last_fetch_at(), moment)

    def test_the_marker_is_a_file_and_not_a_row_in_the_database(self):
        """Это состояние установки, а не данные. В базе оно уехало бы
        в облачную копию и после восстановления соврало бы, что выписку
        только что тянули."""
        self.assertTrue(str(tbank.LAST_FETCH_FILE).endswith('.tbank-last-run'))


class StatementScheduleCommandTests(TestCase):
    """Команда: по расписанию промежуток спрашивается, руками — нет."""

    def setUp(self):
        self.dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.dir.cleanup)
        patcher = patch.object(
            tbank, 'LAST_FETCH_FILE', Path(self.dir.name) / '.tbank-last-run'
        )
        patcher.start()
        self.addCleanup(patcher.stop)

    PAYLOAD = {'operations': []}

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='40802810700006096146',
                       TBANK_STATEMENT_INTERVAL_MINUTES=60)
    def test_a_scheduled_run_too_soon_does_not_touch_the_bank(self):
        tbank.mark_fetched(timezone.now() - datetime.timedelta(minutes=5))
        out = io.StringIO()

        with patch('core.tbank.get_statement') as fetch:
            call_command('tbank_statement', '--scheduled', stdout=out)

        fetch.assert_not_called()
        self.assertIn('Пропускаю', out.getvalue())

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='40802810700006096146',
                       TBANK_STATEMENT_INTERVAL_MINUTES=60)
    def test_a_hand_run_does_not_ask_about_the_interval(self):
        """Человек, набравший команду, хочет выписку сейчас."""
        tbank.mark_fetched(timezone.now() - datetime.timedelta(minutes=5))
        out = io.StringIO()

        with patch('core.tbank.get_statement', return_value=self.PAYLOAD) as fetch:
            call_command('tbank_statement', stdout=out)

        fetch.assert_called_once()

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='40802810700006096146',
                       TBANK_STATEMENT_INTERVAL_MINUTES=60)
    def test_a_successful_run_moves_the_marker(self):
        with patch('core.tbank.get_statement', return_value=self.PAYLOAD):
            call_command('tbank_statement', '--scheduled', stdout=io.StringIO())

        self.assertIsNotNone(tbank.last_fetch_at())

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='40802810700006096146')
    def test_a_failed_run_leaves_the_marker_alone(self):
        """Сорвался запрос к банку — следующий тик обязан попробовать
        снова, а не ждать час."""
        with patch('core.tbank.get_statement',
                   side_effect=tbank.TBankError('банк не ответил')):
            call_command('tbank_statement', '--scheduled',
                         stdout=io.StringIO(), stderr=io.StringIO())

        self.assertIsNone(tbank.last_fetch_at())

    @override_settings(TBANK_TOKEN='secret', TBANK_ACCOUNT='40802810700006096146')
    def test_a_dry_run_leaves_the_marker_alone(self):
        """Проверка — не загрузка: после неё расписание не должно считать,
        что выписка уже дома."""
        with patch('core.tbank.get_statement', return_value=self.PAYLOAD):
            call_command('tbank_statement', '--dry-run', stdout=io.StringIO())

        self.assertIsNone(tbank.last_fetch_at())


class StatementScheduleUnitsTests(SimpleTestCase):
    """Юниты systemd и настройка обязаны сходиться: таймер тикает,
    решает программа."""

    def _read(self, name):
        return (Path(__file__).resolve().parent.parent / 'deploy'
                / name).read_text(encoding='utf-8')

    def test_the_service_asks_the_program_whether_it_is_time(self):
        """Без --scheduled команда тянула бы выписку на каждом тике,
        и настройка частоты не значила бы ничего."""
        self.assertIn('tbank_statement --scheduled',
                      self._read('lifteam-tbank.service'))

    def test_the_timer_ticks_more_often_than_the_default_interval(self):
        """Тик реже настройки означал бы, что настройку ниже тика
        выставить нельзя."""
        self.assertIn('/15', self._read('lifteam-tbank.timer'))

    def test_the_interval_is_editable_from_the_page(self):
        self.assertIn('TBANK_STATEMENT_INTERVAL_MINUTES', envfile.EDITABLE_BY_NAME)
        self.assertNotIn('TBANK_STATEMENT_INTERVAL_MINUTES', envfile.SECRET_NAMES)


class WebhookReadinessCheckTests(EnvFileMixin, TestCase):
    """Кнопка «Проверить» у секрета уведомлений.

    Связи тут не проверить: уведомление присылает банк, а секрет — то,
    чего мы ждём от него. Зато проверяется всё, из-за чего уведомления
    молча отбиваются: неверная настройка ничего не роняет, банк просто
    получает отказ, а оплата не отмечается — до первого спорного счёта.
    """

    def test_a_disabled_reception_is_named(self):
        with override_settings(WEBHOOKS_TBANK_ENABLED=False,
                               WEBHOOKS_TBANK_SECRET='Bearer строка'):
            result = selfcheck.check('WEBHOOKS_TBANK_SECRET')

        self.assertEqual(result.state, selfcheck.CheckResult.FAIL)
        self.assertIn('приём выключен', result.message)

    def test_an_empty_secret_is_named(self):
        with override_settings(WEBHOOKS_TBANK_ENABLED=True,
                               WEBHOOKS_TBANK_SECRET=''):
            result = selfcheck.check('WEBHOOKS_TBANK_SECRET')

        self.assertEqual(result.state, selfcheck.CheckResult.FAIL)
        self.assertIn('секрет не задан', result.message)

    def test_a_secret_without_its_scheme_is_named(self):
        """Самая обидная из здешних ошибок: поле выглядит заполненным,
        а не совпадёт ни с одним уведомлением — банк присылает значение
        заголовка целиком, вместе со схемой."""
        with override_settings(WEBHOOKS_TBANK_ENABLED=True,
                               WEBHOOKS_TBANK_SECRET='простострокабезсхемы'):
            result = selfcheck.check('WEBHOOKS_TBANK_SECRET')

        self.assertEqual(result.state, selfcheck.CheckResult.FAIL)
        self.assertIn('без схемы', result.message)

    def test_an_empty_address_list_is_named(self):
        with override_settings(WEBHOOKS_TBANK_ENABLED=True,
                               WEBHOOKS_TBANK_SECRET='Bearer строка',
                               WEBHOOKS_TBANK_IPS=[]):
            result = selfcheck.check('WEBHOOKS_TBANK_SECRET')

        self.assertEqual(result.state, selfcheck.CheckResult.FAIL)
        self.assertIn('список адресов', result.message)

    def test_everything_set_but_nothing_delivered_yet_says_so(self):
        """«Настроено верно» — не то же, что «работает»: совпал секрет
        или нет, знает только банк."""
        with override_settings(WEBHOOKS_TBANK_ENABLED=True,
                               WEBHOOKS_TBANK_SECRET='Bearer строка'):
            result = selfcheck.check('WEBHOOKS_TBANK_SECRET')

        self.assertEqual(result.state, selfcheck.CheckResult.SKIPPED)
        self.assertIn('не приходило ни одного', result.message)

    def test_a_real_delivery_is_the_proof(self):
        """Подтвердить секрет может только сам банк — доставкой."""
        WebhookDelivery.objects.create(
            provider='tbank', dedup_key='inv-1', body_hash='x',
            status=WebhookDelivery.STATUS_PROCESSED,
        )

        with override_settings(WEBHOOKS_TBANK_ENABLED=True,
                               WEBHOOKS_TBANK_SECRET='Bearer строка'):
            result = selfcheck.check('WEBHOOKS_TBANK_SECRET')

        self.assertTrue(result.ok)
        self.assertIn('принято: 1', result.message)

    def test_tochka_says_that_reception_will_refuse_anyway(self):
        """Рапортовать «настроено верно» о том, что работать не будет, —
        прямая неправда."""
        with override_settings(WEBHOOKS_TOCHKA_ENABLED=True,
                               WEBHOOKS_TOCHKA_SECRET='Bearer строка'):
            result = selfcheck.check('WEBHOOKS_TOCHKA_SECRET')

        self.assertEqual(result.state, selfcheck.CheckResult.SKIPPED)
        self.assertIn('откажет при любых настройках', result.message)

    def test_every_secret_now_has_a_check(self):
        """До v2.90.0 у секретов вебхуков ответом было «проверки нет» —
        верно по сути и бесполезно на деле."""
        for name in envfile.SECRET_NAMES:
            with self.subTest(name=name):
                self.assertIn(name, selfcheck.CHECKS)


class LabelReadsBottomUpTests(TestCase):
    """Этикетка читается снизу вверх: артикул у нижнего края.

    На передней стенке кассетницы есть выступ, за который ящик выдвигают,
    и он закрывает верхнюю полосу наклейки — то есть ровно ту, ради которой
    её и читают. Всё важное убрано вниз.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='label_order', full_name='Админ', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.cabinet = Cabinet.objects.create(number=9)
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.first()
        self.part = SparePart.objects.create(
            part_number='BU-1', name='Резистор 10к', component_type='Резистор',
            package='0805', description='Ставится в блок питания',
        )
        self.cell.parts.add(self.part)

    def _pages(self):
        return (
            '/parts/%d/label/' % self.part.pk,
            '/storage-cells/%d/label/' % self.cell.pk,
        )

    def test_the_service_line_comes_before_the_part_number(self):
        """Наверху то, чем жертвуют: корпус, применимость, адрес."""
        for page in self._pages():
            with self.subTest(page=page):
                html = self.http.get(page).content.decode()
                sheet = html.split('label-sheet')[1].split('</div>\n</div>')[0]

                self.assertLess(sheet.index('label-foot'), sheet.index('label-number'))

    def test_the_part_number_is_the_last_thing_on_the_label(self):
        """Нижний край — единственное место, до которого выступ
        не дотянется ни при какой глубине."""
        for page in self._pages():
            with self.subTest(page=page):
                html = self.http.get(page).content.decode()
                sheet = html.split('label-sheet')[1]

                self.assertGreater(sheet.index('label-number'), sheet.index('label-body'))

    def test_the_text_and_the_code_hug_the_bottom(self):
        """Пустое место на этикетке всё равно где-то будет, и лучше ему
        оказаться под выступом, чем между описанием и артикулом."""
        styles = render_to_string('core/_label_part_styles.html')
        block = styles.split('.label-text-block')[1].split('}')[0]
        body = styles.split('.label-body {')[1].split('}')[0]

        self.assertIn('justify-content: flex-end', block)
        self.assertIn('align-items: flex-end', body)


class CellLabelHiddenTopTests(TestCase):
    """Сколько сверху закрыто выступом — настройкой, потому что кассетницы
    у разных серий разные, а число нужно измерить у настоящей."""

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='label_hidden', full_name='Админ', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.cabinet = Cabinet.objects.create(number=11)
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.first()
        self.part = SparePart.objects.create(
            part_number='HD-1', name='Резистор', component_type='Резистор',
        )
        self.cell.parts.add(self.part)

    @override_settings(LABEL_CELL_HIDDEN_TOP_MM=6)
    def test_the_cell_label_carries_it(self):
        html = self.http.get('/storage-cells/%d/label/' % self.cell.pk).content.decode()

        self.assertIn('--hidden: 6mm', html)

    @override_settings(LABEL_CELL_HIDDEN_TOP_MM=6)
    def test_the_part_label_does_not(self):
        """На пакете с деталью выступа нет, и пустая полоса пропала бы зря."""
        html = self.http.get('/parts/%d/label/' % self.part.pk).content.decode()

        self.assertNotIn('--hidden', html.split('label-sheet')[1].split('>')[0])

    @override_settings(LABEL_CELL_HIDDEN_TOP_MM=6)
    def test_the_batch_page_carries_it_too(self):
        """Пачкой печатают чаще, чем по одной: разойдись эти две страницы,
        половина наклеек оказалась бы старой раскладки."""
        html = self.http.get(
            '/storage-cells/labels/?cabinet=%d' % self.cabinet.number
        ).content.decode()

        self.assertIn('--hidden: 6mm', html)

    @override_settings(LABEL_CELL_HIDDEN_TOP_MM=0)
    def test_zero_leaves_the_label_as_it_was(self):
        html = self.http.get('/storage-cells/%d/label/' % self.cell.pk).content.decode()

        self.assertNotIn('--hidden', html.split('label-sheet')[1].split('>')[0])

    def test_nonsense_is_brought_back_to_reason(self):
        """Значение правит человек, а этикетка с закрытой половиной —
        это уже не этикетка."""
        for given, expected in ((-5, 0), (0, 0), (6, 6), (99, 12), ('много', 0)):
            with self.subTest(given=given):
                with override_settings(LABEL_CELL_HIDDEN_TOP_MM=given):
                    self.assertEqual(views._cell_label_hidden_top(), expected)

    def test_it_is_editable_from_the_settings_page(self):
        """Померить выступ можно только у настоящей кассетницы, значит
        число вводит владелец, а не программист."""
        self.assertIn('LABEL_CELL_HIDDEN_TOP_MM', envfile.EDITABLE_BY_NAME)


class QrErrorCorrectionTests(SimpleTestCase):
    """Уровень восстановления кода — самый высокий, и он ничего не стоит.

    Наклейка на оборудовании живёт в машинном отделении: прибор задевают,
    трут, ставят на него что попало. Уровень H держит около 30 %
    повреждённой площади против 15 % у прежнего M.
    """

    # То, что кладётся в код на самом деле: вид объекта и его номер
    # в базе (views.qr_payload). Адреса сервера здесь нет с v2.61.0
    PAYLOADS = ('u/123', 'p/42', 'c/7', 'e/1', 'o/2')
    # Каким код станет, когда номера дорастут до шести знаков
    LONG_ID = 'p/123456'

    def _modules(self, uri):
        from PIL import Image
        raw = base64.b64decode(uri.split(',', 1)[1])
        image = Image.open(io.BytesIO(raw))
        # box_size=6, border=1 с каждой стороны
        return image.size[0] // 6 - 2

    def test_the_level_is_the_highest_one(self):
        import qrcode
        source = (Path(__file__).resolve().parent / 'utils.py').read_text(encoding='utf-8')

        self.assertIn('ERROR_CORRECT_H', source)
        self.assertNotIn('ERROR_CORRECT_M', source)
        self.assertTrue(hasattr(qrcode.constants, 'ERROR_CORRECT_H'))

    def test_it_costs_nothing_on_our_short_payloads(self):
        """Ради этого с v2.61.0 в коде и нет адреса сервера: содержимое
        короткое, и самый высокий уровень умещается в тот же 21 модуль."""
        for payload in self.PAYLOADS:
            with self.subTest(payload=payload):
                self.assertEqual(
                    self._modules(generate_qr_image(payload)), 21
                )

    def test_six_digit_numbers_still_print_comfortably(self):
        """Самый высокий уровень держит семь знаков в 21 модуле; на восьми
        код вырастает до 25. Это не беда: на 12,3 мм при 203 точках на дюйм
        выходит 98 точек, то есть 3,6 точки на модуль вместе с полем —
        столько же, сколько на этикетке заказа, которая печатается годами.
        Проверка стоит здесь, чтобы рост не оказался неожиданностью,
        когда номера дорастут."""
        self.assertLessEqual(self._modules(generate_qr_image(self.LONG_ID)), 25)

    def test_a_full_url_would_no_longer_fit(self):
        """Проверка на будущее: верни кто-нибудь адрес сервера в код —
        и он вырастет с 21 модуля до 37, то есть перестанет читаться
        на 12,3 мм. Это не запрет, а предупреждение в виде теста."""
        long_payload = 'http://lifteam.taile9b605.ts.net/u/123'

        self.assertGreater(self._modules(generate_qr_image(long_payload)), 21)


class EquipmentLabelSerialTests(TestCase):
    """Серийный номер на этикетке оборудования — запасной путь к прибору.

    Если код повреждён, единицу находят в программе поиском по этому
    номеру, а прочитать его надо глазами.
    """

    def test_the_serial_is_set_as_large_as_the_model(self):
        styles = render_to_string(
            'core/repair_orders/_label_order_equipment_styles.html'
        )
        serial = styles.split('.label-serial {')[1].split('}')[0]
        model = styles.split('.label-model {')[1].split('}')[0]

        self.assertIn('font-weight: bold', serial)
        self.assertEqual(
            serial.split('font-size:')[1].split(';')[0].strip(),
            model.split('font-size:')[1].split(';')[0].strip(),
        )


class BatchLabelContextTests(TestCase):
    """Пачка этикеток передаёт поля в шаблон поимённо, и каждое имя
    обязано быть ключом словаря этикетки.

    Пропущенный ключ здесь не оставляет поле пустым: у словаря Python
    есть собственные методы, и `label.items` при отсутствии ключа
    находит метод `items()`, вызывает его и печатает на наклейке весь
    словарь целиком — вместе с `<SparePart: …>` и содержимым QR-кода.
    Ровно это и печаталось на каждой этикетке детали, отправленной
    пачкой.
    """

    def setUp(self):
        self.admin = Employee.objects.create_superuser(
            username='batch_ctx', full_name='Админ', password='pass',
        )
        self.http = TestClient()
        self.http.force_login(self.admin)

        self.cabinet = Cabinet.objects.create(number=13)
        self.cabinet.apply_layout([4])
        self.cell = self.cabinet.cells.first()
        self.part = SparePart.objects.create(
            part_number='BC-1', name='Варистор 100 В', component_type='Варистор',
        )
        self.cell.parts.add(self.part)

    def _names_used(self, template):
        source = (Path(__file__).resolve().parent / 'templates' / 'core'
                  / template).read_text(encoding='utf-8')
        return set(re.findall(r'label\.(\w+)', source))

    def test_the_part_label_carries_every_name_the_batch_asks_for(self):
        keys = set(views._part_label(self.part, 'http://x/').keys())

        missing = self._names_used('parts/labels_batch.html') - keys

        self.assertEqual(missing, set())

    def test_the_cell_label_carries_every_name_the_batch_asks_for(self):
        keys = set(views._cell_label(self.cell, 'http://x/').keys())

        missing = self._names_used('storage_cells/labels_batch.html') - keys

        self.assertEqual(missing, set())

    def test_a_batch_of_part_labels_prints_no_innards(self):
        """Проверка на сам след беды, а не только на её причину."""
        html = self.http.get('/parts/labels/?ids=%d' % self.part.pk).content.decode()
        area = html.split('labels-area')[1]

        for leak in ('SparePart', 'qr_payload', "('part'", 'data:image/png;base64'):
            with self.subTest(leak=leak):
                sheet = area.split('label-text-block')[1].split('label-qr-wrap')[0]
                self.assertNotIn(leak, sheet)

    def test_a_batch_of_cell_labels_prints_no_innards(self):
        html = self.http.get(
            '/storage-cells/labels/?cabinet=%d' % self.cabinet.number
        ).content.decode()
        area = html.split('labels-area')[1]
        sheet = area.split('label-text-block')[1].split('label-qr-wrap')[0]

        for leak in ('SparePart', 'qr_payload', "('title'"):
            with self.subTest(leak=leak):
                self.assertNotIn(leak, sheet)

    def test_a_single_part_label_is_unaffected(self):
        """Одиночная страница берёт поля из своего контекста, а не
        поимённо, и этой беды у неё не было — но проверить надо обе."""
        html = self.http.get('/parts/%d/label/' % self.part.pk).content.decode()
        sheet = html.split('label-text-block')[1].split('label-qr-wrap')[0]

        self.assertNotIn('SparePart', sheet)
